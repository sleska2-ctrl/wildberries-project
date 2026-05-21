from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from .analytics import build_analytics_tables
from .config import load_settings
from .google_sheets import GoogleSheetsClient
from .sqlite_store import SQLiteStore
from .transform import (
    ads_to_sheet_rows,
    aggregate_daily_pnl,
    build_buyout_order_day_rows,
    build_cogs_map,
    build_nm_mapping,
    extract_our_nm_ids,
    extract_orders_filters,
    extract_filter_values,
    filter_orders_rows,
    filter_sales_rows,
    flatten_ads_rows,
    funnel_to_sheet_rows,
    orders_to_sheet_rows,
    pnl_to_sheet_rows,
    sales_to_sheet_rows,
    sheet_values_to_dicts,
    stocks_to_sheet_rows,
)
from .wb_client import WildberriesClient


EXTERNAL_SKU_XLSX_PATH = Path(__file__).resolve().parents[2] / "data" / "sku_iitech.xlsx"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Sync Wildberries data to SQLite")
    parser.add_argument("--date-from", dest="date_from")
    parser.add_argument("--date-to", dest="date_to")
    parser.add_argument(
        "--only",
        choices=("all", "sales", "orders", "stocks", "ads", "funnel"),
        default="all",
        help="Load only one data type and rebuild derived tables from SQLite",
    )
    parser.add_argument("--skip-ads", action="store_true")
    parser.add_argument("--skip-funnel", action="store_true")
    return parser


def _clean_text(value: object) -> str:
    return str(value or "").strip()


def _load_external_sku_map() -> dict[str, dict[str, str]]:
    if not EXTERNAL_SKU_XLSX_PATH.exists():
        return {}

    workbook = load_workbook(EXTERNAL_SKU_XLSX_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    header = {_clean_text(name): idx for idx, name in enumerate(rows[0])}
    first_idx = 0
    nm_idx = header.get("Артикул WB")
    subject_idx = header.get("Предмет")
    supplier_idx = header.get("Артикул поставщика")
    our_article_idx = header.get("НАШ")
    name_idx = header.get("Название")
    cogs_idx = header.get("Себестоимость")
    commission_idx = header.get("Комиссия", header.get("Комиссии"))
    if nm_idx is None:
        return {}

    result: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        if len(row) <= nm_idx:
            continue
        # В новом полном файле пустая первая колонка означает служебную/пустую
        # строку. Ее не нужно добавлять в SKU.
        if len(row) > first_idx and not _clean_text(row[first_idx]):
            continue
        nm_id = _clean_text(row[nm_idx])
        if not nm_id:
            continue
        result[nm_id] = {
            "subject": _clean_text(row[subject_idx]) if subject_idx is not None and len(row) > subject_idx else "",
            "supplier_article": _clean_text(row[supplier_idx]) if supplier_idx is not None and len(row) > supplier_idx else "",
            "our_article": _clean_text(row[our_article_idx]) if our_article_idx is not None and len(row) > our_article_idx else "",
            "name": _clean_text(row[name_idx]) if name_idx is not None and len(row) > name_idx else "",
            "cogs": _clean_text(row[cogs_idx]) if cogs_idx is not None and len(row) > cogs_idx else "",
            "commission": _clean_text(row[commission_idx]) if commission_idx is not None and len(row) > commission_idx else "",
        }
    return result


def _enrich_sku_values(sheet_values: list[list[str]]) -> tuple[list[list[str]], int]:
    if not sheet_values:
        return sheet_values, 0

    external_by_nm = _load_external_sku_map()
    if not external_by_nm:
        return sheet_values, 0

    header = [str(value).strip() for value in sheet_values[0]]
    header_index = {name: idx for idx, name in enumerate(header)}

    def ensure_column(name: str) -> int:
        idx = header_index.get(name)
        if idx is not None:
            return idx
        idx = len(header)
        header.append(name)
        header_index[name] = idx
        return idx

    nm_idx = ensure_column("Артикул WB")
    subject_idx = ensure_column("Предмет")
    supplier_idx = ensure_column("Артикул поставщика")
    our_article_idx = ensure_column("НАШ")
    name_idx = ensure_column("Название")
    cogs_idx = ensure_column("себестоимость")
    commission_idx = ensure_column("% комиссии на вб")
    ensure_column("ИИТех")

    rows = [list(row) + [""] * max(0, len(header) - len(row)) for row in sheet_values[1:]]
    row_by_nm: dict[str, list[str]] = {}
    for row in rows:
        nm_id = _clean_text(row[nm_idx]) if len(row) > nm_idx else ""
        if nm_id:
            row_by_nm[nm_id] = row

    merged = 0
    for nm_id, payload in external_by_nm.items():
        row = row_by_nm.get(nm_id)
        if row is None:
            row = [""] * len(header)
            row[nm_idx] = nm_id
            rows.append(row)
            row_by_nm[nm_id] = row
        if payload["subject"]:
            row[subject_idx] = payload["subject"]
        if payload["supplier_article"]:
            row[supplier_idx] = payload["supplier_article"]
        if payload["our_article"]:
            row[our_article_idx] = payload["our_article"]
        if payload["name"]:
            row[name_idx] = payload["name"]
        if payload["cogs"]:
            row[cogs_idx] = payload["cogs"]
        if payload["commission"]:
            try:
                row[commission_idx] = f"{float(payload['commission']) * 100:.2f}%"
            except ValueError:
                row[commission_idx] = payload["commission"]
        merged += 1

    return [header, *rows], merged


def _stored_rows(store: SQLiteStore, table_name: str) -> list[dict]:
    return sheet_values_to_dicts(store.get_values(table_name))


def _collect_nm_ids(*row_sets: list[dict], sku_values: list[list[str]] | None = None) -> list[int]:
    result: set[int] = set()
    for rows in row_sets:
        for row in rows:
            value = str(row.get("nmId") or row.get("nm_id") or "").strip()
            if value.isdigit():
                result.add(int(value))
    if sku_values:
        header = [str(value).strip() for value in sku_values[0]] if sku_values else []
        if "Артикул WB" in header:
            nm_idx = header.index("Артикул WB")
            for row in sku_values[1:]:
                value = str(row[nm_idx] if len(row) > nm_idx else "").strip()
                if value.isdigit():
                    result.add(int(value))
    return sorted(result)


def _stored_funnel_data(store: SQLiteStore, funnel_table: str) -> list[dict]:
    by_nm: dict[str, dict] = {}
    for row in _stored_rows(store, funnel_table):
        nm_id = str(row.get("nmId", "") or "").strip()
        row_date = str(row.get("date", "") or "").strip()
        if not nm_id or not row_date:
            continue
        entry = by_nm.setdefault(nm_id, {
            "currency": str(row.get("currency", "") or "RUB"),
            "product": {
                "nmId": nm_id,
                "vendorCode": str(row.get("supplierArticle", "") or ""),
                "title": str(row.get("productName", "") or ""),
                "brandName": str(row.get("brand", "") or ""),
            },
            "history": [],
        })
        entry["history"].append({
            "date": row_date,
            "orderSum": row.get("orderSum") or 0,
            "orderCount": row.get("orderCount") or 0,
            "buyoutCount": row.get("buyoutCount") or 0,
            "buyoutSum": row.get("buyoutSum") or 0,
            "cancelCount": row.get("cancelCount") or 0,
        })
    return list(by_nm.values())


def _delete_date_range(store: SQLiteStore, table_name: str, date_column: str, date_from: str, date_to: str) -> None:
    import sqlite3

    resolved = store._table_name(table_name)  # noqa: SLF001 - internal helper keeps table names consistent.
    db_path = str(store._db_path)  # noqa: SLF001
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            f'DELETE FROM "{resolved}" WHERE substr("{date_column}", 1, 10) >= ? AND substr("{date_column}", 1, 10) <= ?',
            (date_from, date_to),
        )
        conn.commit()


def _rebuild_derived_tables(
    *,
    store: SQLiteStore,
    settings,
    cogs_values: list[list[str]],
    nm_mapping: dict[str, str],
    date_from: str,
    date_to: str,
) -> None:
    print("[INFO] Пересобираем производные таблицы из SQLite")
    stored_sales_rows = _stored_rows(store, settings.raw_sales_sheet)
    stored_orders_rows = _stored_rows(store, settings.raw_orders_sheet)
    stored_ads_rows = _stored_rows(store, settings.raw_ads_sheet)
    funnel_data = _stored_funnel_data(store, settings.funnel_analytics_sheet)
    cogs_map = build_cogs_map(cogs_values, article_filter_type=settings.article_filter_type)

    daily_pnl = aggregate_daily_pnl(
        sales_rows=stored_sales_rows,
        ads_rows=stored_ads_rows,
        cogs_map=cogs_map,
        article_filter_type=settings.article_filter_type,
        nm_mapping=nm_mapping,
    )
    store.upsert_table(
        settings.daily_pnl_sheet,
        pnl_to_sheet_rows(daily_pnl),
        key_columns=("date", "article_type", "article"),
        overwrite_existing=True,
    )

    buyout_order_day_rows = build_buyout_order_day_rows(
        orders_rows=stored_orders_rows,
        sales_rows=stored_sales_rows,
        ads_rows=stored_ads_rows,
        nm_mapping=nm_mapping,
        funnel_data=funnel_data,
    )
    store.replace_table("buyout_order_day", buyout_order_day_rows)
    print(f"[INFO] Таблица buyout_order_day: строк {max(0, len(buyout_order_day_rows) - 1)}")

    analytics_tables = build_analytics_tables(
        sales_rows=stored_sales_rows,
        ads_rows=stored_ads_rows,
        cogs_map=cogs_map,
        article_filter_type=settings.article_filter_type,
        nm_mapping=nm_mapping,
        date_from=date_from,
        date_to=date_to,
    )
    store.upsert_table("finance_article_day_detail", analytics_tables["finance_article_day_detail"], key_columns=("Дата", "Артикул"), overwrite_existing=True, allow_new_columns=True)
    store.upsert_table("analytics_article_day", analytics_tables["analytics_article_day"], key_columns=("Артикул", "Дата"), overwrite_existing=True, allow_new_columns=True)
    store.upsert_table("analytics_day", analytics_tables["analytics_day"], key_columns=("Дата",), overwrite_existing=True, allow_new_columns=True)
    store.upsert_table("analytics_article_period", analytics_tables["analytics_article_period"], key_columns=("Период с", "Период по", "Артикул"), overwrite_existing=True, allow_new_columns=True)
    print("[INFO] Производные таблицы обновлены")


def _run_single_load(
    *,
    only: str,
    wb_client: WildberriesClient,
    store: SQLiteStore,
    settings,
    cogs_values: list[list[str]],
    nm_mapping: dict[str, str],
    date_from: str,
    date_to: str,
) -> None:
    if only == "sales":
        print("[INFO] Загружаем только продажи")
        sales_rows = wb_client.fetch_sales_details(date_from=date_from, date_to=date_to, period="daily")
        print(f"[INFO] Продажи: получено строк {len(sales_rows)}")
        store.upsert_table(settings.raw_sales_sheet, sales_to_sheet_rows(sales_rows), key_columns=("rrdId",))
    elif only == "orders":
        print("[INFO] Загружаем только заказы")
        orders_rows = wb_client.fetch_orders(date_from=date_from)
        print(f"[INFO] Заказы: получено строк {len(orders_rows)}")
        store.upsert_table(settings.raw_orders_sheet, orders_to_sheet_rows(orders_rows), key_columns=("srid",))
    elif only == "stocks":
        print("[INFO] Загружаем только остатки")
        stock_rows = wb_client.fetch_stocks()
        store.replace_table("raw_stocks", stocks_to_sheet_rows(stock_rows))
        print(f"[INFO] Остатки: загружено строк {len(stock_rows)}")
    elif only == "ads":
        print("[INFO] Загружаем только рекламу")
        nm_ids = _collect_nm_ids(
            _stored_rows(store, settings.raw_sales_sheet),
            _stored_rows(store, settings.raw_orders_sheet),
            _stored_rows(store, "raw_stocks"),
            sku_values=cogs_values,
        )
        print(f"[INFO] Реклама: ищем кампании по {len(nm_ids)} nmId")
        advert_ids = wb_client.fetch_relevant_campaign_ids(nm_ids)
        print(f"[INFO] Реклама: найдено кампаний {len(advert_ids)}")
        ads_rows = flatten_ads_rows(wb_client.fetch_campaign_stats(advert_ids, date_from, date_to))
        print(f"[INFO] Реклама: получено строк {len(ads_rows)}")
        _delete_date_range(store, settings.raw_ads_sheet, "date", date_from, date_to)
        store.upsert_table(settings.raw_ads_sheet, ads_to_sheet_rows(ads_rows), key_columns=("date", "advertId", "appType", "nmId"), update_existing=True)
    elif only == "funnel":
        print("[INFO] Загружаем только воронку")
        nm_ids = _collect_nm_ids(
            _stored_rows(store, settings.raw_sales_sheet),
            _stored_rows(store, settings.raw_orders_sheet),
            _stored_rows(store, "raw_stocks"),
            sku_values=cogs_values,
        )
        print(f"[INFO] Воронка: загружаем по {len(nm_ids)} nmId")
        funnel_data = wb_client.fetch_funnel_history(date_from, date_to, nm_ids=nm_ids)
        funnel_rows = funnel_to_sheet_rows(funnel_data)
        _delete_date_range(store, settings.funnel_analytics_sheet, "date", date_from, date_to)
        store.upsert_table(settings.funnel_analytics_sheet, funnel_rows, key_columns=("date", "nmId"), update_existing=True, allow_new_columns=True)
        print(f"[INFO] Воронка: загружено строк {max(0, len(funnel_rows) - 1)}")
    else:
        raise ValueError(f"Unknown --only value: {only}")

    _rebuild_derived_tables(
        store=store,
        settings=settings,
        cogs_values=cogs_values,
        nm_mapping=nm_mapping,
        date_from=date_from,
        date_to=date_to,
    )
    print("[INFO] Частичная загрузка завершена")


def main() -> None:
    args = build_parser().parse_args()
    settings = load_settings()

    date_from = args.date_from or settings.default_date_from
    date_to = args.date_to or settings.default_date_to
    print(f"[INFO] Синхронизация: период {date_from}..{date_to}, реклама={'нет' if args.skip_ads else 'да'}, воронка={'нет' if args.skip_funnel else 'да'}")

    wb_client = WildberriesClient(
        finance_token=settings.wb_finance_token,
        adv_token=settings.wb_adv_token,
    )
    store = SQLiteStore(settings.sqlite_db_path)

    cogs_values = store.get_values(settings.cogs_sheet)
    if not cogs_values and settings.google_service_account_file and settings.google_spreadsheet_id:
        sheets_client = GoogleSheetsClient(
            service_account_file=settings.google_service_account_file,
            spreadsheet_id=settings.google_spreadsheet_id,
        )
        cogs_values = sheets_client.get_values(settings.cogs_sheet)
    if cogs_values:
        cogs_values, merged_sku_rows = _enrich_sku_values(cogs_values)
        store.replace_table(settings.cogs_sheet, cogs_values)
        print(f"[INFO] Кэширован лист {settings.cogs_sheet} из Google Sheets в SQLite")
        if merged_sku_rows:
            print(f"[INFO] SKU: обновлено/добавлено {merged_sku_rows} строк из {EXTERNAL_SKU_XLSX_PATH.name}")

    nm_mapping = build_nm_mapping(cogs_values, article_filter_type=settings.article_filter_type)
    our_nm_ids = extract_our_nm_ids(cogs_values)
    article_filter_values = settings.article_filter_values or extract_filter_values(cogs_values, article_filter_type=settings.article_filter_type)
    orders_supplier_articles, orders_nm_ids = extract_orders_filters(cogs_values)

    if args.only != "all":
        _run_single_load(
            only=args.only,
            wb_client=wb_client,
            store=store,
            settings=settings,
            cogs_values=cogs_values,
            nm_mapping=nm_mapping,
            date_from=date_from,
            date_to=date_to,
        )
        return

    print("[INFO] Загружаем продажи")
    sales_rows = wb_client.fetch_sales_details(date_from=date_from, date_to=date_to, period="daily")
    print(f"[INFO] Продажи: получено строк {len(sales_rows)}")
    filtered_sales_rows = sales_rows
    print(f"[INFO] Продажи: сохраняем полный кабинет, строк {len(filtered_sales_rows)}")

    print("[INFO] Загружаем заказы")
    orders_rows = wb_client.fetch_orders(date_from=date_from)
    print(f"[INFO] Заказы: получено строк {len(orders_rows)}")
    filtered_orders_rows = orders_rows
    print(f"[INFO] Заказы: сохраняем полный кабинет, строк {len(filtered_orders_rows)}")

    try:
        print("[INFO] Загружаем остатки")
        stock_rows = wb_client.fetch_stocks()
        filtered_stock_rows = stock_rows
    except Exception as exc:
        print(f"[WARN] Не удалось загрузить остатки: {exc}")
        filtered_stock_rows = []

    ads_rows = []
    if not args.skip_ads:
        try:
            relevant_nm_ids = sorted({
                int(str(row.get("nmId") or "").strip())
                for row in [*sales_rows, *orders_rows, *filtered_stock_rows]
                if str(row.get("nmId") or "").strip().isdigit()
            })
            print(f"[INFO] Реклама: ищем кампании по {len(relevant_nm_ids)} nmId")
            advert_ids = wb_client.fetch_relevant_campaign_ids(relevant_nm_ids)
            print(f"[INFO] Реклама: найдено кампаний {len(advert_ids)}")
            ads_payload = wb_client.fetch_campaign_stats(
                advert_ids=advert_ids,
                date_from=date_from,
                date_to=date_to,
            )
            ads_rows = flatten_ads_rows(ads_payload)
            print(f"[INFO] Реклама: получено строк {len(ads_rows)}")
        except Exception as exc:
            print(f"[WARN] Не удалось загрузить рекламу: {exc}")
            ads_rows = []

    print("[INFO] Собираем производные таблицы")
    cogs_map = build_cogs_map(
        sheet_values=cogs_values,
        article_filter_type=settings.article_filter_type,
    )
    daily_pnl = aggregate_daily_pnl(
        sales_rows=filtered_sales_rows,
        ads_rows=ads_rows,
        cogs_map=cogs_map,
        article_filter_type=settings.article_filter_type,
        nm_mapping=nm_mapping,
    )

    analytics_tables = build_analytics_tables(
        sales_rows=filtered_sales_rows,
        ads_rows=ads_rows,
        cogs_map=cogs_map,
        article_filter_type=settings.article_filter_type,
        nm_mapping=nm_mapping,
        date_from=date_from,
        date_to=date_to,
    )

    store.upsert_table(
        settings.raw_sales_sheet,
        sales_to_sheet_rows(filtered_sales_rows),
        key_columns=("rrdId",),
    )
    store.upsert_table(
        settings.raw_orders_sheet,
        orders_to_sheet_rows(filtered_orders_rows),
        key_columns=("srid",),
    )
    store.upsert_table(
        settings.raw_ads_sheet,
        ads_to_sheet_rows(ads_rows),
        key_columns=("date", "advertId", "appType", "nmId"),
    )
    print("[INFO] Сохраняем исходные таблицы в SQLite")
    store.replace_table("raw_stocks", stocks_to_sheet_rows(filtered_stock_rows))
    print(f"[INFO] Остатки: загружено строк {len(filtered_stock_rows)}")
    store.upsert_table(
        settings.daily_pnl_sheet,
        pnl_to_sheet_rows(daily_pnl),
        key_columns=("date", "article_type", "article"),
        update_existing=True,
    )

    stored_sales_rows = sheet_values_to_dicts(store.get_values(settings.raw_sales_sheet))
    stored_orders_rows = sheet_values_to_dicts(store.get_values(settings.raw_orders_sheet))

    # Fetch funnel before building buyout_order_day so funnel prices populate orders_sum/orders_qty.
    funnel_data: list[dict] = []
    if not args.skip_funnel:
        try:
            target_nm_ids = sorted({
                int(str(row.get("nmId") or "").strip())
                for row in [*stored_sales_rows, *stored_orders_rows, *filtered_stock_rows]
                if str(row.get("nmId") or "").strip().isdigit()
            })
            if target_nm_ids:
                print(f"[INFO] Воронка: загружаем по {len(target_nm_ids)} nmId")
                funnel_data = wb_client.fetch_funnel_history(date_from, date_to, nm_ids=target_nm_ids)
                funnel_rows = funnel_to_sheet_rows(funnel_data)
                store.upsert_table(
                    settings.funnel_analytics_sheet,
                    funnel_rows,
                    key_columns=("date", "nmId"),
                    update_existing=True,
                    allow_new_columns=True,
                )
                print(f"[INFO] Воронка: загружено строк {max(0, len(funnel_rows) - 1)}")
            else:
                print("[WARN] Воронка: нет nmId для запроса (проверь лист SKU / маппинг nmId)")
        except Exception as exc:
            print(f"[WARN] Не удалось загрузить воронку: {exc}")

    # Supplement funnel_data with stored funnel_analytics for any (date, nmId) the API didn't return.
    # buyout_order_day is rebuilt as a full table below, so stored funnel rows from every date
    # must be included; otherwise order sums/counts disappear outside the current sync window.
    api_funnel_keys: set[tuple[str, str]] = {
        (str(h.get("date", "")), str((item.get("product") or {}).get("nmId", "")))
        for item in funnel_data
        for h in (item.get("history") or [])
    }
    stored_funnel_rows = sheet_values_to_dicts(store.get_values(settings.funnel_analytics_sheet))
    allowed_funnel_nm_ids = set()
    extra_by_nm: dict[str, dict] = {}
    for row in stored_funnel_rows:
        d = str(row.get("date", "") or "").strip()
        nm = str(row.get("nmId", "") or "").strip()
        if not d or not nm:
            continue
        if (d, nm) in api_funnel_keys:
            continue
        entry = extra_by_nm.setdefault(nm, {
            "product": {"nmId": nm, "vendorCode": str(row.get("supplierArticle", "") or "")},
            "history": [],
        })
        entry["history"].append({
            "date": d,
            "orderSum": row.get("orderSum") or 0,
            "orderCount": row.get("orderCount") or 0,
            "buyoutCount": row.get("buyoutCount") or 0,
            "buyoutSum": row.get("buyoutSum") or 0,
        })
    if extra_by_nm:
        funnel_data = funnel_data + list(extra_by_nm.values())
        print(f"[INFO] Воронка (из хранилища): добавлено {sum(len(e['history']) for e in extra_by_nm.values())} строк за даты вне API-окна")

    print("[INFO] Пересобираем buyout_order_day")
    buyout_order_day_rows = build_buyout_order_day_rows(
        orders_rows=stored_orders_rows,
        sales_rows=stored_sales_rows,
        ads_rows=sheet_values_to_dicts(store.get_values(settings.raw_ads_sheet)),
        nm_mapping=nm_mapping,
        funnel_data=funnel_data,
    )
    store.replace_table("buyout_order_day", buyout_order_day_rows)
    print(f"[INFO] Таблица buyout_order_day: строк {max(0, len(buyout_order_day_rows) - 1)}")

    # Derived analytics should preserve old dates and only refresh the keys
    # that were recalculated in the current sync run.
    print("[INFO] Обновляем аналитические таблицы")
    store.upsert_table(
        "finance_article_day_detail",
        analytics_tables["finance_article_day_detail"],
        key_columns=("Дата", "Артикул"),
        overwrite_existing=True,
        allow_new_columns=True,
    )
    store.upsert_table(
        "analytics_article_day",
        analytics_tables["analytics_article_day"],
        key_columns=("Артикул", "Дата"),
        overwrite_existing=True,
        allow_new_columns=True,
    )
    store.upsert_table(
        "analytics_day",
        analytics_tables["analytics_day"],
        key_columns=("Дата",),
        overwrite_existing=True,
        allow_new_columns=True,
    )
    store.upsert_table(
        "analytics_article_period",
        analytics_tables["analytics_article_period"],
        key_columns=("Период с", "Период по", "Артикул"),
        overwrite_existing=True,
        allow_new_columns=True,
    )
    print("[INFO] Синхронизация завершена")


def calculate_preliminary_economics(sku, date, orders_count, orders_sum, commission_rate, advertising_cost, additional_expense_rate):
    commission = orders_sum * commission_rate
    acquiring = orders_sum * 0.02
    additional_expenses = orders_sum * additional_expense_rate
    preliminary_profit = orders_sum - commission - acquiring - advertising_cost - additional_expenses

    return {
        "sku": sku,
        "date": date,
        "orders_count": orders_count,
        "orders_sum": orders_sum,
        "commission": commission,
        "acquiring": acquiring,
        "advertising": advertising_cost,
        "additional_expenses": additional_expenses,
        "preliminary_profit": preliminary_profit
    }


def link_orders_with_advertising(sku, date):
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()

        # Fetch orders data
        cursor.execute("SELECT orders_count, orders_sum FROM orders WHERE sku = ? AND date = ?", (sku, date))
        orders_data = cursor.fetchone()

        # Fetch advertising data
        cursor.execute("SELECT advertising_cost FROM advertising WHERE sku = ? AND date = ?", (sku, date))
        advertising_data = cursor.fetchone()

        if orders_data and advertising_data:
            orders_count, orders_sum = orders_data
            advertising_cost = advertising_data[0]
            return {
                "sku": sku,
                "date": date,
                "orders_count": orders_count,
                "orders_sum": orders_sum,
                "advertising_cost": advertising_cost
            }
        else:
            return None


def calculate_financial_metrics(row):
    commission = row['orders_sum'] * 0.1  # Example commission rate
    acquiring = row['orders_sum'] * 0.02
    additional_expenses = row['orders_sum'] * 0.05  # Example additional expense rate
    preliminary_profit = row['orders_sum'] - commission - acquiring - row['advertising_cost'] - additional_expenses

    return {
        "sku": row['sku'],
        "date": row['date'],
        "orders_count": row['orders_count'],
        "orders_sum": row['orders_sum'],
        "commission": commission,
        "acquiring": acquiring,
        "advertising": row['advertising_cost'],
        "additional_expenses": additional_expenses,
        "preliminary_profit": preliminary_profit
    }


def calculate_average_additional_expenses():
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT AVG(additional_expenses / orders_sum) FROM preliminary_economics WHERE orders_sum > 0")
        result = cursor.fetchone()
        return result[0] if result and result[0] else 0.05  # Default to 5% if no data


if __name__ == "__main__":
    main()
