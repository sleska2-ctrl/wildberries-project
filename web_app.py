"""Simple web UI for triggering Wildberries data sync."""
from __future__ import annotations

import base64
import binascii
import queue
import json
import sqlite3
import subprocess
import sys
import threading
import time
import tempfile
from decimal import Decimal, InvalidOperation
from html import escape
from datetime import date, datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import io
import os
import requests
from math import ceil
from pathlib import Path
from urllib.parse import parse_qs, quote_plus, urlparse
from zoneinfo import ZoneInfo

ROOT = Path(__file__).parent
PYTHON = sys.executable
DEFAULT_CABINET_ID = os.getenv("DEFAULT_CABINET_ID", "ewb").strip() or "ewb"
DEFAULT_DB_PATH = str(ROOT / "data" / "cabs" / f"{DEFAULT_CABINET_ID}.db")
DB_PATH = os.getenv("SQLITE_DB_PATH", DEFAULT_DB_PATH).strip() or DEFAULT_DB_PATH
SITE_PASSWORD = os.getenv("WB_SITE_PASSWORD", "").strip()
SYNC_LOG_DIR = ROOT / "data" / "sync_logs"
LATEST_SYNC_LOG_PATH = SYNC_LOG_DIR / "latest.log"

# ── Multi-cabinet platform ──────────────────────────────────────────────────
PLATFORM_DB_PATH = str(ROOT / "data" / "platform.db")
PRICES_DB_PATH   = str(ROOT / "data" / "wb_prices.db")
ADMIN_PIN = os.getenv("ADMIN_PIN", "").strip()

_request_local = threading.local()
_platform_instance = None
_platform_lock = threading.Lock()

def get_platform():
    global _platform_instance
    if _platform_instance is not None:
        return _platform_instance
    with _platform_lock:
        if _platform_instance is None:
            src_path = str(ROOT / "src")
            if src_path not in sys.path:
                sys.path.insert(0, src_path)
            from wb_platform import PlatformStore
            _platform_instance = PlatformStore(PLATFORM_DB_PATH)
    return _platform_instance

def _set_request_cabinet(cabinet: dict | None) -> None:
    _request_local.cabinet = cabinet

def _get_request_cabinet() -> dict | None:
    return getattr(_request_local, "cabinet", None)

def get_current_db_path() -> str:
    cabinet = _get_request_cabinet()
    if cabinet and cabinet.get("cabinet_id"):
        cabs_dir = ROOT / "data" / "cabs"
        cabs_dir.mkdir(parents=True, exist_ok=True)
        return str(cabs_dir / f"{cabinet['cabinet_id']}.db")
    return DB_PATH

def _cabinet_db_path(cabinet: dict | None) -> str:
    cabinet_id = str((cabinet or {}).get("cabinet_id") or "").strip()
    if cabinet_id:
        cabs_dir = ROOT / "data" / "cabs"
        cabs_dir.mkdir(parents=True, exist_ok=True)
        return str(cabs_dir / f"{cabinet_id}.db")
    return DB_PATH
CORE_TABLES = [
  "buyout_order_day",
  "funnel_analytics",
  "raw_sales",
  "raw_orders",
  "raw_ads",
  "raw_stocks",
  "SKU",
  "manager_comments",
]

def _sync_log_paths(date_from: str, date_to: str, only: str) -> list[Path]:
  started_at = datetime.now().strftime("%Y%m%d_%H%M%S")
  mode = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in (only or "all")) or "all"
  run_log_path = SYNC_LOG_DIR / f"sync_{started_at}_{mode}_{date_from}_{date_to}.log"
  return [run_log_path, LATEST_SYNC_LOG_PATH]

def _write_sync_log(paths: list[Path], text: str, reset: bool = False) -> None:
  SYNC_LOG_DIR.mkdir(parents=True, exist_ok=True)
  timestamped = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {text}"
  for path in paths:
    mode = "w" if reset else "a"
    with path.open(mode, encoding="utf-8") as fh:
      fh.write(timestamped + "\n")

def _emit_sync_message(log_q: queue.Queue, log_paths: list[Path], kind: str, text: str) -> None:
  log_q.put((kind, text))
  _write_sync_log(log_paths, f"{kind.upper()}: {text}")

def db_store():
  src_path = str(ROOT / "src")
  if src_path not in sys.path:
    sys.path.insert(0, src_path)
  from wb_gsheets.sqlite_store import SQLiteStore

  return SQLiteStore(get_current_db_path())

REPORT_LINK_ITEMS = [
  ("/", "Загрузка данных"),
  ("/analytics/buyout-order-day", "WB по дням"),
  ("/analytics/buyout-order-week", "WB по неделям"),
  ("/analytics/planning", "WB Планирование"),
  ("/analytics/comments", "WB Комментарии"),
  ("/analytics/competitor-prices", "Цены конкурентов"),
  ("/ads/bidder", "WB Биддер"),
  ("/tasks", "Задачи"),
]

OZON_LINK_ITEMS = [
  ("/ozon/analytics/buyout-order-day", "OZON по дням"),
  ("/ozon/analytics/buyout-order-week", "OZON по неделям"),
  ("/ozon/planning", "OZON Планирование"),
  ("/ozon/comments", "OZON Комментарии"),
]

DB_LINK_ITEM = ("/db", "SQLite")

DISABLED_PAGE_PATHS = {
  "/analytics",
  "/analytics/period",
  "/analytics/day",
  "/analytics/article-day",
  "/analytics/preliminary-economics",
  "/analytics/preliminary-economics-summary",
  "/ozon/analytics/day",
}

DISABLED_API_PATHS = {
  "/api/analytics/period",
  "/api/analytics/day",
  "/api/analytics/article-day",
  "/api/analytics/preliminary-economics",
  "/api/ozon/analytics/day",
}

def _has_wb_cabinet(cabinet: dict | None) -> bool:
    cab = cabinet or {}
    wb_token = str(cab.get("wb_api_token") or "").strip()
    return (
        cab.get("marketplace") in ("wb", "both")
        and bool(wb_token)
    )

def _has_ozon_cabinet(cabinet: dict | None) -> bool:
    cab = cabinet or {}
    ozon_client_id = str(cab.get("ozon_client_id") or "").strip()
    ozon_api_key = str(cab.get("ozon_api_key") or "").strip()
    return (
        cab.get("marketplace") in ("ozon", "both")
        and bool(ozon_client_id)
        and bool(ozon_api_key)
    )

def _send_sse_once(handler: BaseHTTPRequestHandler, kind: str, text: str) -> None:
    handler.send_response(200)
    handler.send_header("Content-Type", "text/event-stream; charset=utf-8")
    handler.send_header("Cache-Control", "no-cache")
    handler.send_header("X-Accel-Buffering", "no")
    handler.end_headers()
    payload = json.dumps({"type": kind, "text": text})
    handler.wfile.write(f"data: {payload}\n\n".encode())
    handler.wfile.flush()

def _report_nav_html_for_cabinet(cabinet: dict | None = None) -> str:
    mp = (cabinet or {}).get("marketplace", "wb")
    load_item = REPORT_LINK_ITEMS[0]
    wb_items = REPORT_LINK_ITEMS[1:]
    if mp in ("ozon", "both"):
        ozon_items = OZON_LINK_ITEMS
    else:
        ozon_items = []
    common_link = (
        "text-decoration:none;font-size:.72rem;font-weight:800;white-space:nowrap;"
        "padding:4px 9px;border-radius:999px;transition:background .15s;"
    )

    def _links(items: list[tuple[str, str]], style: str) -> str:
        return "".join(
            f'<a href="{escape(path)}" style="{common_link}{style}">{escape(label)}</a>'
            for path, label in items
        )

    load_style = "color:#0f766e;background:#ecfdf5;border:1px solid #99f6e4;"
    wb_style = "color:#cb11ab;background:rgba(203,17,171,.09);border:1px solid rgba(203,17,171,.18);"
    ozon_style = "color:#005bff;background:rgba(0,91,255,.09);border:1px solid rgba(0,91,255,.18);"
    db_style = "color:#475569;background:#f1f5f9;border:1px solid #cbd5e1;"
    group_base = "display:inline-flex;align-items:center;gap:3px;padding:2px 4px;border-radius:999px;flex:0 0 auto;"
    load_html = _links([load_item], load_style)
    wb_html = (
        f'<span style="{group_base}background:rgba(203,17,171,.055);">'
        f'{_links(wb_items, wb_style)}</span>'
    )
    ozon_html = (
        f'<span style="{group_base}background:rgba(0,91,255,.055);">'
        f'{_links(ozon_items, ozon_style)}</span>'
        if ozon_items else ""
    )
    db_path, db_label = DB_LINK_ITEM
    db_html = (
        '<span style="flex:1 1 auto;min-width:16px;"></span>'
        f'<a href="{escape(db_path)}" style="{common_link}{db_style}flex:0 0 auto;">{escape(db_label)}</a>'
    )
    return load_html + wb_html + ozon_html + db_html

def _db_connect() -> sqlite3.Connection:
  conn = sqlite3.connect(get_current_db_path())
  conn.execute("PRAGMA journal_mode=WAL")
  conn.row_factory = sqlite3.Row
  return conn

def _to_float(value: object) -> float:
  if value is None:
    return 0.0
  text = str(value).strip().replace(" ", "").replace(",", ".")
  if not text:
    return 0.0
  try:
    return float(text)
  except ValueError:
    return 0.0

def _order_speed_for_day(
  conn: sqlite3.Connection,
  nm_id: str,
  day: str,
) -> float | None:
  row = conn.execute(
    'SELECT SUM(CAST("Заказы, шт" AS REAL)) AS orders_qty '
    'FROM buyout_order_day WHERE "nmId" = ? AND "Дата" = ?',
    [nm_id, day],
  ).fetchone()
  if row is None or row["orders_qty"] is None:
    return None
  return _to_float(row["orders_qty"])

def _format_metric(column: str, value: object) -> str:
  if value is None:
    return ""
  text = str(value)
  if not text.strip():
    return ""

  if column in {"Артикул", "Дата", "Период с", "Период по"}:
    return text

  numeric = _to_float(value)
  if column in {"ДРР", "% маржи"}:
    return f"{int(round(numeric))}%"

  return f"{numeric:,.0f}".replace(",", " ")

def _format_percent(value: float) -> str:
  return f"{int(round(value))}%"

def _safe_percent(numerator: float, denominator: float, min_abs_den: float = 100.0) -> str:
  """Return percent string, or '—' when denominator is too close to zero to be meaningful."""
  if abs(denominator) < min_abs_den:
    return "—"
  return _format_percent(numerator / denominator * 100.0)

def _scope_context(conn: sqlite3.Connection | None = None) -> dict[str, object]:
  own_conn = conn is None
  conn = conn or _db_connect()
  try:
    columns = _table_columns(conn, "sku")
    nmid_col = _first_existing(columns, ["Артикул WB", "nmId", "nm_id"])
    article_col = _first_existing(columns, ["Артикул поставщика", "supplierArticle", "SKU", "sku"])
    cogs_col = _first_existing(columns, ["себестоимость", "cost_price", "cogs"])
    if not nmid_col:
      return {"all_nm_ids": set(), "nm_to_article": {}, "nm_with_cogs": set()}
    select_cols = [nmid_col]
    if article_col:
      select_cols.append(article_col)
    if cogs_col:
      select_cols.append(cogs_col)
    sql = "SELECT " + ", ".join(_sql_ident(col) for col in select_cols) + " FROM sku"
    rows = conn.execute(sql).fetchall()
    all_nm_ids: set[str] = set()
    nm_to_article: dict[str, str] = {}
    nm_with_cogs: set[str] = set()
    for row in rows:
      nm_id = str(row[nmid_col] or "").strip()
      article = str(row[article_col] or "").strip() if article_col else ""
      cogs = str(row[cogs_col] or "").strip() if cogs_col else ""
      if not nm_id:
        continue
      all_nm_ids.add(nm_id)
      if article:
        nm_to_article[nm_id] = article
      if cogs:
        nm_with_cogs.add(nm_id)
    return {"all_nm_ids": all_nm_ids, "nm_to_article": nm_to_article, "nm_with_cogs": nm_with_cogs}
  finally:
    if own_conn:
      conn.close()

def _fetch_period_analytics(date_from: str, date_to: str, article_query: str = "") -> list[dict[str, str]]:
  with _db_connect() as conn:
    where = ['"Дата" >= ?', '"Дата" <= ?']
    params: list[str] = [date_from, date_to]
    article_query = article_query.strip()
    if article_query:
      where.append('"Артикул" = ?')
      params.append(article_query)
    where_sql = f"WHERE {' AND '.join(where)}"
    rows = conn.execute(
      (
        "SELECT \"Артикул\", "
        "SUM(\"Продажи по нашей цене\") AS \"Продажи по нашей цене\", "
        "SUM(\"Реклама\") AS \"Реклама\", "
        "SUM(\"Чистая прибыль\") AS \"Чистая прибыль\", "
        "CASE WHEN SUM(\"Продажи по нашей цене\") != 0 "
        "THEN SUM(\"Реклама\") / SUM(\"Продажи по нашей цене\") * 100 ELSE 0 END AS \"ДРР\", "
        "CASE WHEN SUM(\"Продажи по нашей цене\") != 0 "
        "THEN SUM(\"Чистая прибыль\") / SUM(\"Продажи по нашей цене\") * 100 ELSE 0 END AS \"% маржи\" "
        f"FROM analytics_article_day {where_sql} GROUP BY \"Артикул\" ORDER BY \"Артикул\" ASC LIMIT 5000"
      ),
      params,
    ).fetchall()
    return [
      {column: _format_metric(column, row[idx]) for idx, column in enumerate(rows[0].keys())}
      for row in rows
    ] if rows else []

def _fetch_article_day_analytics(article_query: str, date_from: str | None, date_to: str | None) -> list[dict[str, str]]:
  with _db_connect() as conn:
    where = []
    params: list[str] = []
    article_query = article_query.strip()
    if article_query:
      where.append('"Артикул" = ?')
      params.append(article_query)
    if date_from:
      where.append('"Дата" >= ?')
      params.append(date_from)
    if date_to:
      where.append('"Дата" <= ?')
      params.append(date_to)
    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    rows = conn.execute(
      "SELECT \"Артикул\", \"Дата\", \"Продажи по нашей цене\", \"Реклама\", \"Чистая прибыль\", \"ДРР\", \"% маржи\" "
      f"FROM analytics_article_day {where_sql} ORDER BY \"Дата\" DESC, \"Артикул\" ASC LIMIT 500",
      params,
    ).fetchall()
    return [
      {
        "Артикул": _format_metric("Артикул", row["Артикул"]),
        "Дата": _format_metric("Дата", row["Дата"]),
        "Продажи по нашей цене": _format_metric("Продажи по нашей цене", row["Продажи по нашей цене"]),
        "Реклама": _format_metric("Реклама", row["Реклама"]),
        "Чистая прибыль": _format_metric("Чистая прибыль", row["Чистая прибыль"]),
        "ДРР": _format_metric("ДРР", row["ДРР"]),
        "% маржи": _format_metric("% маржи", row["% маржи"]),
      }
      for row in rows
    ]

def _fetch_day_analytics(date_from: str, date_to: str) -> dict[str, object]:
  with _db_connect() as conn:
    rows = conn.execute(
      """
      SELECT
        "Дата",
        "Продажи по нашей цене, р",
        "Реклама, р",
        "Чистая прибыль",
        "ДРР",
        "% маржи"
      FROM analytics_day
      WHERE "Дата" >= ? AND "Дата" <= ?
      ORDER BY "Дата" DESC
      """,
      (date_from, date_to),
    ).fetchall()
    payload_rows: list[dict[str, str]] = []
    total_sales = 0.0
    total_ads = 0.0
    total_profit = 0.0
    for row in rows:
      sales = _to_float(row["Продажи по нашей цене, р"])
      ads = _to_float(row["Реклама, р"])
      profit = _to_float(row["Чистая прибыль"])
      total_sales += sales
      total_ads += ads
      total_profit += profit
      payload_rows.append(
        {
          "Дата": _format_metric("Дата", row["Дата"]),
          "Продажи": _format_metric("Продажи", sales),
          "Реклама": _format_metric("Реклама", ads),
          "Чистая прибыль": _format_metric("Чистая прибыль", profit),
          "ДРР": _format_metric("ДРР", row["ДРР"]),
          "% маржи": _format_metric("% маржи", row["% маржи"]),
        }
      )
    return {
      "rows": payload_rows,
      "summary": {
        "days": str(len(payload_rows)),
        "sales": _format_metric("Продажи", total_sales),
        "ads": _format_metric("Реклама", total_ads),
        "profit": _format_metric("Чистая прибыль", total_profit),
        "drr": _format_metric("ДРР", (total_ads / total_sales * 100.0) if total_sales else 0.0),
        "margin": _format_metric("% маржи", (total_profit / total_sales * 100.0) if total_sales else 0.0),
      },
    }

def _date_range_limited(date_from: str, date_to: str, max_days: int = 30) -> tuple[list[str], str, str]:
  start = date.fromisoformat(date_from)
  end = date.fromisoformat(date_to)
  if start > end:
    start, end = end, start
  if (end - start).days + 1 > max_days:
    start = end - timedelta(days=max_days - 1)
  dates = [(start + timedelta(days=idx)).isoformat() for idx in range((end - start).days + 1)]
  return dates, start.isoformat(), end.isoformat()

def _week_label(day: str, effective_from: str, effective_to: str) -> str:
  current = date.fromisoformat(day)
  week_start = current - timedelta(days=current.weekday())
  week_end = week_start + timedelta(days=6)
  start = max(week_start, date.fromisoformat(effective_from))
  end = min(week_end, date.fromisoformat(effective_to))
  return f"{start.isoformat()}..{end.isoformat()}"

def _aggregate_daily_dicts_by_week(rows: list[dict[str, object]], effective_from: str, effective_to: str) -> list[dict[str, object]]:
  buckets: dict[str, dict[str, object]] = {}
  for row in rows:
    label = _week_label(str(row.get("Дата", "")), effective_from, effective_to)
    bucket = buckets.setdefault(label, {"Дата": label})
    for key, value in row.items():
      if key == "Дата":
        continue
      bucket[key] = _to_float(bucket.get(key, 0.0)) + _to_float(value)
  return [buckets[key] for key in sorted(buckets)]

def _aggregate_daily_values_by_week(
  values_by_date: dict[str, float],
  daily_dates: list[str],
  effective_from: str,
  effective_to: str,
) -> dict[str, float]:
  result: dict[str, float] = {}
  for day in daily_dates:
    label = _week_label(day, effective_from, effective_to)
    result[label] = result.get(label, 0.0) + float(values_by_date.get(day, 0.0))
  return result

def _aggregate_stock_values_by_week(
  values_by_date: dict[str, float | None],
  daily_dates: list[str],
  effective_from: str,
  effective_to: str,
) -> dict[str, float | None]:
  result: dict[str, float | None] = {}
  for day in daily_dates:
    value = values_by_date.get(day)
    if value is None:
      continue
    result[_week_label(day, effective_from, effective_to)] = float(value)
  return result

def _aggregate_nested_daily_values_by_week(
  values_by_date: dict[str, dict[str, float]],
  daily_dates: list[str],
  effective_from: str,
  effective_to: str,
) -> dict[str, dict[str, float]]:
  result: dict[str, dict[str, float]] = {}
  for day in daily_dates:
    label = _week_label(day, effective_from, effective_to)
    bucket = result.setdefault(label, {})
    for key, value in values_by_date.get(day, {}).items():
      bucket[key] = bucket.get(key, 0.0) + float(value)
  return result

def _buyout_subject_join(conn: sqlite3.Connection) -> str:
  columns = _table_columns(conn, "sku")
  nmid_col = _first_existing(columns, ["Артикул WB", "nmId", "nm_id"])
  subject_expr = _sql_ident("Предмет") if "Предмет" in columns else "''"
  yarlyky_expr = _sql_ident("Ярлыки") if "Ярлыки" in columns else "''"
  revenue_category_expr = _sql_ident("Категория по выручке") if "Категория по выручке" in columns else "''"
  if not nmid_col:
    return "LEFT JOIN (SELECT NULL AS nmid, '' AS subject, '' AS yarlyky, '' AS revenue_category WHERE 0) sku_subject ON 1 = 0"
  return (
    "LEFT JOIN ("
    "SELECT "
    f"{_sql_ident(nmid_col)} AS nmid, "
    f"MIN({subject_expr}) AS subject, "
    f"MIN({yarlyky_expr}) AS yarlyky, "
    f"MIN({revenue_category_expr}) AS revenue_category "
    "FROM sku "
    f"WHERE TRIM(COALESCE({_sql_ident(nmid_col)}, '')) != '' GROUP BY {_sql_ident(nmid_col)}"
    ") sku_subject ON sku_subject.nmid = buyout_order_day.\"nmId\""
  )

def _append_sku_attribute_filter(
  where: list[str],
  params: list[str],
  nmid_expr: str,
  *,
  subject: str = "",
  labels: list[str] | None = None,
  revenue_category: str = "",
) -> None:
  filters: list[str] = ["TRIM(COALESCE(\"Артикул WB\", '')) != ''"]
  if subject.strip():
    filters.append(f"{_sql_ident('Предмет')} = ?")
    params.append(subject.strip())
  if labels:
    lc = [
      "(INSTR(',' || TRIM(COALESCE(\"Ярлыки\",'')) || ',', ',' || ? || ',') > 0)"
      for _ in labels
    ]
    filters.append('(' + ' OR '.join(lc) + ')')
    params.extend(labels)
  if revenue_category.strip():
    filters.append(f"{_sql_ident('Категория по выручке')} = ?")
    params.append(revenue_category.strip())
  if len(filters) > 1:
    where.append(f"{nmid_expr} IN (SELECT TRIM(\"Артикул WB\") FROM SKU WHERE {' AND '.join(filters)})")

def _buyout_cogs_join(conn: sqlite3.Connection) -> str:
  columns = _table_columns(conn, "sku")
  nmid_col = _first_existing(columns, ["Артикул WB", "nmId", "nm_id"])
  cogs_col = _first_existing(columns, ["себестоимость", "cost_price", "cogs"])
  if nmid_col and cogs_col:
    return (
      "LEFT JOIN ("
      f"SELECT {_sql_ident(nmid_col)} AS nmid, "
      f"MAX(CAST(REPLACE(REPLACE({_sql_ident(cogs_col)}, ' ', ''), ',', '.') AS REAL)) AS cogs "
      "FROM sku "
      f"WHERE TRIM(COALESCE({_sql_ident(nmid_col)}, '')) != '' "
      f"GROUP BY {_sql_ident(nmid_col)}"
      ") cogs_data ON cogs_data.nmid = buyout_order_day.\"nmId\""
    )
  return "LEFT JOIN (SELECT NULL AS nmid, 0 AS cogs WHERE 0) cogs_data ON 1 = 0"

def _ozon_sku_cogs_join(conn: sqlite3.Connection, source_alias: str = "s") -> str:
  columns = _table_columns(conn, "sku")
  ozon_sku_col = _first_existing(columns, ["SKU Ozon", "ozon_sku", "sku", "SKU"])
  cogs_col = _first_existing(columns, ["себестоимость", "cost_price", "cogs"])
  if ozon_sku_col and cogs_col:
    ozon_sku_expr = _sql_ident(ozon_sku_col)
    cogs_expr = (
      f"CAST(REPLACE(REPLACE(COALESCE(NULLIF(TRIM({_sql_ident(cogs_col)}), ''), '0'), ' ', ''), ',', '.') AS REAL)"
    )
    return (
      "LEFT JOIN ("
      f"SELECT CAST({ozon_sku_expr} AS TEXT) AS sku, "
      f"MAX({cogs_expr}) AS cogs "
      "FROM sku "
      f"WHERE TRIM(COALESCE({ozon_sku_expr}, '')) != '' "
      f"GROUP BY CAST({ozon_sku_expr} AS TEXT)"
      f") cogs_data ON cogs_data.sku = CAST({source_alias}.sku AS TEXT)"
    )
  return "LEFT JOIN (SELECT NULL AS sku, 0 AS cogs WHERE 0) cogs_data ON 1 = 0"

def _stocks_join(conn: sqlite3.Connection) -> str:
  columns = _table_columns(conn, "raw_stocks")
  if {"nmId", "quantity"}.issubset(columns):
    return (
      "LEFT JOIN ("
      "SELECT \"nmId\" AS nmid, "
      "SUM(CAST(REPLACE(REPLACE(\"quantity\", ' ', ''), ',', '.') AS REAL)) AS stock "
      "FROM raw_stocks WHERE TRIM(COALESCE(\"nmId\", '')) != '' GROUP BY \"nmId\""
      ") stock_data ON stock_data.nmid = buyout_order_day.\"nmId\""
    )
  return "LEFT JOIN (SELECT NULL AS nmid, NULL AS stock WHERE 0) stock_data ON 1 = 0"

def _stock_row_total(raw_values: list[float | None], formatter) -> str:
  present = [float(v) for v in raw_values if v is not None]
  if not present:
    return "—"
  return f"{formatter(present[-1])} / {formatter(sum(present) / len(present))}"

def _append_stock_pivot_row(
  pivot_rows: list[dict[str, object]],
  *,
  group: str,
  label: str,
  dates: list[str],
  values_by_date: dict[str, float | None],
  formatter,
  replenishment_threshold: float | None = None,
) -> None:
  raw_values: list[float | None] = []
  values: list[str] = []
  for day in dates:
    raw = values_by_date.get(day)
    raw_values.append(raw if raw is None else float(raw))
    values.append("—" if raw is None else formatter(float(raw)))
  pivot_rows.append({
    "group": group,
    "metric": label,
    "total": _stock_row_total(raw_values, formatter),
    "values": values,
    "raw_values": raw_values,
    "kind": "stock",
    "color_mode": "absolute",
    "color_threshold": max(1.0, float(replenishment_threshold if replenishment_threshold is not None else 20.0)),
  })

def _stock_values_for_output_dates(
  values_by_date: dict[str, float],
  *,
  out_dates: list[str],
  daily_dates: list[str],
  effective_from: str,
  effective_to: str,
  granularity: str,
) -> dict[str, float | None]:
  if granularity == "week":
    weekly = _aggregate_stock_values_by_week(values_by_date, daily_dates, effective_from, effective_to)
    return {day: weekly.get(day) for day in out_dates}
  return {day: values_by_date.get(day) for day in out_dates}

def _fetch_wb_stock_values_by_date(
  conn: sqlite3.Connection,
  *,
  effective_from: str,
  effective_to: str,
  article_query: str,
  selected_articles: list[str],
  subject: str,
  labels: list[str],
  revenue_category: str,
) -> dict[str, float]:
  def _build_common_filters(alias: str, nmid_col: str, article_col: str | None = None) -> tuple[list[str], list[object]]:
    where: list[str] = []
    params: list[object] = []
    if selected_articles:
      placeholders = ", ".join("?" for _ in selected_articles)
      if article_col:
        where.append(
          f"(TRIM(CAST({alias}.{nmid_col} AS TEXT)) IN ({placeholders}) "
          f"OR TRIM(CAST(COALESCE({alias}.{article_col}, '') AS TEXT)) IN ({placeholders}))"
        )
        params.extend(selected_articles)
        params.extend(selected_articles)
      else:
        where.append(f"TRIM(CAST({alias}.{nmid_col} AS TEXT)) IN ({placeholders})")
        params.extend(selected_articles)
    elif article_query:
      if article_query.isdigit():
        where.append(f"TRIM(CAST({alias}.{nmid_col} AS TEXT)) = ?")
        params.append(article_query)
      elif article_col:
        where.append(f"TRIM(CAST(COALESCE({alias}.{article_col}, '') AS TEXT)) = ?")
        params.append(article_query)
    else:
      where.append(
        f"TRIM(CAST({alias}.{nmid_col} AS TEXT)) IN ("
        "SELECT TRIM(\"Артикул WB\") FROM SKU WHERE TRIM(COALESCE(\"Артикул WB\", '')) != ''"
        ")"
      )
    _append_sku_attribute_filter(
      where,
      params,
      f"{alias}.{nmid_col}",
      subject=subject,
      labels=labels,
      revenue_category=revenue_category,
    )
    return where, params

  if _table_exists(conn, "wb_stock_daily_snapshot"):
    columns = _table_columns(conn, "wb_stock_daily_snapshot")
    if {"snapshot_date", "nmId", "quantity"}.issubset(columns):
      where, params = _build_common_filters("s", "nmId", "supplierArticle" if "supplierArticle" in columns else None)
      where.insert(0, "s.snapshot_date >= ?")
      where.insert(1, "s.snapshot_date <= ?")
      params = [effective_from, effective_to, *params]
      rows = conn.execute(
        (
          "SELECT s.snapshot_date AS day, SUM(CAST(s.quantity AS REAL)) AS stock "
          "FROM wb_stock_daily_snapshot s "
          f"WHERE {' AND '.join(where)} GROUP BY s.snapshot_date ORDER BY s.snapshot_date ASC"
        ),
        params,
      ).fetchall()
      result = {str(row["day"]): _to_float(row["stock"]) for row in rows}
      if result:
        return result

  columns = _table_columns(conn, "raw_stocks")
  if not {"nmId", "quantity"}.issubset(columns):
    return {}
  where, params = _build_common_filters("raw_stocks", "nmId", "supplierArticle" if "supplierArticle" in columns else None)
  today = date.today().isoformat()
  if effective_to != today:
    return {}
  rows = conn.execute(
    (
      "SELECT SUM(CAST(REPLACE(REPLACE(raw_stocks.quantity, ' ', ''), ',', '.') AS REAL)) AS stock "
      "FROM raw_stocks "
      f"WHERE {' AND '.join(where)}"
    ),
    params,
  ).fetchall()
  stock = _to_float(rows[0]["stock"]) if rows else 0.0
  return {effective_to: stock} if stock else {}

def _buyout_nm_join(conn: sqlite3.Connection) -> str:
  columns = _table_columns(conn, "sku")
  article_col = _first_existing(columns, ["Артикул поставщика", "supplierArticle", "SKU", "sku"])
  nmid_col = _first_existing(columns, ["Артикул WB", "nmId", "nm_id"])
  if article_col and nmid_col:
    # Берем nmId из SKU-таблицы, чтобы построить прямую ссылку на карточку WB.
    return (
      "LEFT JOIN ("
      f"SELECT {_sql_ident(nmid_col)} AS nmid, MIN({_sql_ident(article_col)}) AS article "
      "FROM sku "
      f"WHERE TRIM(COALESCE({_sql_ident(nmid_col)}, '')) != '' "
      f"GROUP BY {_sql_ident(nmid_col)}"
      ") nm_data ON nm_data.nmid = buyout_order_day.\"nmId\""
    )
  return "LEFT JOIN (SELECT NULL AS article, NULL AS nmid WHERE 0) nm_data ON 1 = 0"

def _marketplace_article_by_nm(conn: sqlite3.Connection, nm_ids: list[str]) -> dict[str, str]:
  nm_ids = [str(nm_id).strip() for nm_id in nm_ids if str(nm_id).strip()]
  if not nm_ids or not _table_exists(conn, "buyout_order_day"):
    return {}
  columns = _table_columns(conn, "buyout_order_day")
  if not {"nmId", "Артикул"}.issubset(columns):
    return {}
  result: dict[str, str] = {}
  for idx in range(0, len(nm_ids), 900):
    chunk = nm_ids[idx:idx + 900]
    ph = ", ".join("?" for _ in chunk)
    for row in conn.execute(
      f'SELECT CAST("nmId" AS TEXT) AS nmid, MAX(NULLIF(TRIM("Артикул"), "")) AS article '
      f'FROM buyout_order_day WHERE CAST("nmId" AS TEXT) IN ({ph}) GROUP BY CAST("nmId" AS TEXT)',
      chunk,
    ).fetchall():
      article = str(row["article"] or "").strip()
      if article:
        result[str(row["nmid"])] = article
  return result

def _parse_labels(val: str) -> list[str]:
  """Parse comma-separated labels from Ярлыки column value."""
  return [l.strip() for l in (val or '').split(',') if l.strip()]


def _labels_match(article_labels: list[str], selected: list[str]) -> bool:
  """True if article has at least one of the selected labels (OR logic)."""
  if not selected:
    return True
  a = {l.lower() for l in article_labels}
  s = {l.lower() for l in selected}
  return bool(a & s)


def _fetch_buyout_filter_options(date_from: str, date_to: str) -> dict[str, list[str]]:
  dates, effective_from, effective_to = _date_range_limited(date_from, date_to, max_days=30)
  with _db_connect() as conn:
    ctx = _scope_context(conn)
    rows = conn.execute(
      (
        "SELECT sku_subject.subject AS subject, sku_subject.yarlyky AS yarlyky, "
        "sku_subject.revenue_category AS revenue_category, "
        "SUM(CAST(buyout_order_day.\"Сумма заказов\" AS REAL)) AS orders_sum "
        "FROM buyout_order_day "
        f"{_buyout_subject_join(conn)} "
        "WHERE buyout_order_day.\"Дата\" >= ? AND buyout_order_day.\"Дата\" <= ? "
        "AND (TRIM(COALESCE(sku_subject.subject, '')) != '' "
        "OR TRIM(COALESCE(sku_subject.yarlyky, '')) != '' "
        "OR TRIM(COALESCE(sku_subject.revenue_category, '')) != '') "
        "GROUP BY sku_subject.subject, sku_subject.yarlyky, sku_subject.revenue_category "
        "ORDER BY orders_sum DESC "
      ),
      [effective_from, effective_to],
    ).fetchall()
  subjects: set[str] = set()
  labels: set[str] = set()
  revenue_categories: set[str] = set()
  for row in rows:
    if str(row["subject"] or "").strip():
      subjects.add(str(row["subject"]).strip())
    for lbl in _parse_labels(str(row["yarlyky"] or "")):
      labels.add(lbl)
    if str(row["revenue_category"] or "").strip():
      revenue_categories.add(str(row["revenue_category"]).strip())
  return {
    "subjects": sorted(subjects, key=str.lower),
    "labels": sorted(labels, key=str.lower),
    "revenue_categories": sorted(revenue_categories, key=str.lower),
  }

def _fetch_buyout_subjects(date_from: str, date_to: str) -> list[str]:
  return _fetch_buyout_filter_options(date_from, date_to)["subjects"]

def _fetch_buyout_articles(
  date_from: str,
  date_to: str,
  subject: str = "",
  labels: list[str] | None = None,
  revenue_category: str = "",
) -> list[dict[str, object]]:
  dates, effective_from, effective_to = _date_range_limited(date_from, date_to, max_days=30)
  day_count = max(1, len(dates))
  where = ['buyout_order_day."Дата" >= ?', 'buyout_order_day."Дата" <= ?']
  params: list[str] = [effective_from, effective_to]
  subject = subject.strip()
  labels = [l.strip() for l in (labels or []) if l.strip()]
  revenue_category = revenue_category.strip()
  if subject:
    where.append("sku_subject.subject = ?")
    params.append(subject)
  if labels:
      lc=[]; lp:list[str]=[]
      for lbl in labels:
        lc.append("(INSTR(',' || TRIM(COALESCE(sku_subject.yarlyky,'')) || ',', ',' || ? || ',') > 0)")
        lp.append(lbl)
      where.append('('+' OR '.join(lc)+')')
      params.extend(lp)
  if revenue_category:
    where.append("sku_subject.revenue_category = ?")
    params.append(revenue_category)
  with _db_connect() as conn:
    ctx = _scope_context(conn)
    tables = {str(r["name"]) for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    has_raw_ads = "raw_ads" in tables
    has_funnel = "funnel_analytics" in tables
    ads_join = (
      "LEFT JOIN (SELECT \"nmId\", SUM(CAST(\"sum\" AS REAL)) AS ads_sum FROM raw_ads "
      f"WHERE date >= '{effective_from}' AND date <= '{effective_to}' GROUP BY \"nmId\") "
      "ads_data ON ads_data.\"nmId\" = buyout_order_day.\"nmId\" "
    ) if has_raw_ads else ""
    ads_select = "COALESCE(ads_data.ads_sum, 0) AS ads_sum, " if has_raw_ads else "0 AS ads_sum, "
    fa_join = (
      "LEFT JOIN ("
      "  SELECT nmId, SUM(CAST(buyoutCount AS REAL)) AS total_buyout, "
      + ("SUM(CAST(cancelCount AS REAL))" if "cancelCount" in _table_columns(conn, "funnel_analytics") else "0") +
      " AS total_cancel"
      "  FROM funnel_analytics WHERE date >= ? AND date <= ? GROUP BY nmId"
      ") fa ON fa.nmId = buyout_order_day.\"nmId\" "
    ) if has_funnel else ""
    fa_select = (
      "COALESCE(fa.total_buyout, 0) AS fa_buyout, COALESCE(fa.total_cancel, 0) AS fa_cancel, "
      if has_funnel
      else "0 AS fa_buyout, 0 AS fa_cancel, "
    )
    query_params: list[str] = []
    if has_funnel:
      query_params.extend([effective_from, effective_to])
    query_params.extend(params)
    rows = conn.execute(
      (
        "SELECT "
        "buyout_order_day.\"nmId\" AS nmid, "
        "MAX(buyout_order_day.\"Артикул\") AS \"Артикул\", "
        "MAX(stock_data.stock) AS stock, "
        "SUM(CAST(buyout_order_day.\"Выкупы товаров, заказанных в эту дату, шт\" AS REAL)) AS ordered_buyouts, "
        "SUM(CAST(buyout_order_day.\"Сумма выкупов в эту дату\" AS REAL)) AS revenue, "
        "SUM(CAST(buyout_order_day.\"Сумма заказов\" AS REAL)) AS orders_sum, "
        + ads_select +
        fa_select +
        "COALESCE(ord3d.avg_orders_3d, 0) AS avg_orders_3d "
        "FROM buyout_order_day "
        f"{_buyout_subject_join(conn)} "
        f"{_stocks_join(conn)} "
        + ads_join +
        fa_join +
        f"LEFT JOIN ("
        f"  SELECT \"nmId\", SUM(CAST(\"Заказы, шт\" AS REAL)) / 3.0 AS avg_orders_3d"
        f"  FROM buyout_order_day"
        f"  WHERE \"Дата\" >= date('{effective_to}', '-2 days') AND \"Дата\" <= '{effective_to}'"
        f"  GROUP BY \"nmId\""
        f") ord3d ON ord3d.\"nmId\" = buyout_order_day.\"nmId\" "
        f"WHERE {' AND '.join(where)} "
        "GROUP BY buyout_order_day.\"nmId\" "
        "ORDER BY revenue DESC, orders_sum DESC, buyout_order_day.\"Артикул\" ASC "
        "LIMIT 1000"
      ),
      query_params,
    ).fetchall()
  result = []
  for row in rows:
    article = str(row["Артикул"]).strip()
    if not article:
      continue
    revenue = _to_float(row["revenue"])
    ads = _to_float(row["ads_sum"])
    drr = ads / revenue * 100.0 if revenue > 0 else None
    stock = _to_float(row["stock"]) if row["stock"] is not None else None
    avg_orders_3d = _to_float(row["avg_orders_3d"])
    turnover = round(stock / avg_orders_3d) if (stock is not None and avg_orders_3d and avg_orders_3d > 0) else 999
    result.append({
      "article": article,
      "nmid": str(row["nmid"]).strip() if row["nmid"] is not None else "",
      "stock": stock,
      "buyouts": _to_float(row["ordered_buyouts"]) / day_count,
      "revenue": revenue,
      "drr": drr,
      "turnover": turnover,
    })
  return result

def _fetch_buyout_order_day_pivot(
  date_from: str,
  date_to: str,
  article_query: str = "",
  articles: list[str] | None = None,
  subject: str = "",
  labels: list[str] | None = None,
  revenue_category: str = "",
  granularity: str = "day",
) -> dict[str, object]:
  dates, effective_from, effective_to = _date_range_limited(date_from, date_to, max_days=93 if granularity == "week" else 45)
  daily_dates = list(dates)
  where = ['buyout_order_day."Дата" >= ?', 'buyout_order_day."Дата" <= ?']
  params: list[str] = [effective_from, effective_to]
  article_query = article_query.strip()
  subject = subject.strip()
  labels = [l.strip() for l in (labels or []) if l.strip()]
  revenue_category = revenue_category.strip()
  selected_articles = [article.strip() for article in (articles or []) if article.strip()]
  has_article_filter = bool(selected_articles or article_query or subject or labels or revenue_category)
  stock_by_date: dict[str, float] = {}
  if selected_articles:
    placeholders = ", ".join("?" for _ in selected_articles)
    where.append(
      "(TRIM(CAST(buyout_order_day.\"nmId\" AS TEXT)) IN ("
      + placeholders
      + ") OR TRIM(CAST(buyout_order_day.\"Артикул\" AS TEXT)) IN ("
      + placeholders
      + "))"
    )
    params.extend(selected_articles)
    params.extend(selected_articles)
  elif article_query:
    if article_query.isdigit():
      where.append('buyout_order_day."nmId" = ?')
      params.append(article_query)
    else:
      where.append('buyout_order_day."Артикул" = ?')
      params.append(article_query)
  if subject:
    where.append("sku_subject.subject = ?")
    params.append(subject)
  if labels:
      lc=[]; lp:list[str]=[]
      for lbl in labels:
        lc.append("(INSTR(',' || TRIM(COALESCE(sku_subject.yarlyky,'')) || ',', ',' || ? || ',') > 0)")
        lp.append(lbl)
      where.append('('+' OR '.join(lc)+')')
      params.extend(lp)
  if revenue_category:
    where.append("sku_subject.revenue_category = ?")
    params.append(revenue_category)

  with _db_connect() as conn:
    ctx = _scope_context(conn)
    where_sql = f"WHERE {' AND '.join(where)}"
    spp_without_by_date: dict[str, float] = {}
    spp_with_by_date: dict[str, float] = {}
    spp_qty_by_date: dict[str, float] = {}
    cogs_join = _buyout_cogs_join(conn)

    raw_order_columns = _table_columns(conn, "raw_orders")
    if {
      "date",
      "priceWithDisc",
      "finishedPrice",
      "nmId",
    }.issubset(raw_order_columns):
      raw_where = [
        "substr(raw_orders.date, 1, 10) >= ?",
        "substr(raw_orders.date, 1, 10) <= ?",
      ]
      raw_params: list[str] = [effective_from, effective_to]
      raw_has_supplier_article = "supplierArticle" in raw_order_columns
      if selected_articles:
        placeholders = ", ".join("?" for _ in selected_articles)
        if raw_has_supplier_article:
          raw_where.append(
            "(TRIM(CAST(raw_orders.nmId AS TEXT)) IN ("
            + placeholders
            + ") OR TRIM(CAST(COALESCE(raw_orders.supplierArticle, '') AS TEXT)) IN ("
            + placeholders
            + "))"
          )
          raw_params.extend(selected_articles)
        else:
          raw_where.append(
            "TRIM(CAST(raw_orders.nmId AS TEXT)) IN ("
            + placeholders
            + ")"
          )
        raw_params.extend(selected_articles)
      elif article_query:
        if article_query.isdigit():
          raw_where.append("raw_orders.nmId = ?")
          raw_params.append(article_query)
        else:
          raw_where.append("raw_orders.supplierArticle = ?")
          raw_params.append(article_query)
      else:
        raw_where.append(
          "raw_orders.nmId IN ("
          "SELECT TRIM(\"Артикул WB\") FROM SKU WHERE TRIM(COALESCE(\"Артикул WB\", '')) != ''"
          ")"
        )
      _append_sku_attribute_filter(
        raw_where,
        raw_params,
        "raw_orders.nmId",
        subject=subject,
        labels=labels,
        revenue_category=revenue_category,
      )
      spp_rows = conn.execute(
        (
          "SELECT "
          "substr(raw_orders.date, 1, 10) AS order_date, "
          "SUM(CAST(REPLACE(REPLACE(raw_orders.priceWithDisc, ' ', ''), ',', '.') AS REAL)) AS without_spp, "
          "SUM(CAST(REPLACE(REPLACE(COALESCE(NULLIF(raw_orders.finishedPrice, ''), raw_orders.priceWithDisc), ' ', ''), ',', '.') AS REAL)) AS with_spp, "
          "COUNT(*) AS qty "
          "FROM raw_orders "
          f"WHERE {' AND '.join(raw_where)} "
          "GROUP BY order_date ORDER BY order_date ASC"
        ),
        raw_params,
      ).fetchall()
      spp_without_by_date = {
        str(row["order_date"]): _to_float(row["without_spp"])
        for row in spp_rows
      }
      spp_with_by_date = {
        str(row["order_date"]): _to_float(row["with_spp"])
        for row in spp_rows
      }
      spp_qty_by_date = {
        str(row["order_date"]): _to_float(row["qty"])
        for row in spp_rows
      }

    funnel_exists = _table_exists(conn, "funnel_analytics")
    funnel_columns = _table_columns(conn, "funnel_analytics") if funnel_exists else []
    funnel_has_supplier_article = "supplierArticle" in funnel_columns

    funnel_where = ["date >= ?", "date <= ?"]
    funnel_params: list[str] = [effective_from, effective_to]
    if selected_articles:
      placeholders = ", ".join("?" for _ in selected_articles)
      if funnel_has_supplier_article:
        funnel_where.append(
          "(TRIM(CAST(nmId AS TEXT)) IN ("
          + placeholders
          + ") OR TRIM(CAST(supplierArticle AS TEXT)) IN ("
          + placeholders
          + "))"
        )
        funnel_params.extend(selected_articles)
      else:
        funnel_where.append(
          "TRIM(CAST(nmId AS TEXT)) IN ("
          + placeholders
          + ")"
        )
      funnel_params.extend(selected_articles)
    elif article_query:
      if article_query.isdigit():
        funnel_where.append("nmId = ?")
        funnel_params.append(article_query)
      else:
        funnel_where.append("supplierArticle = ?")
        funnel_params.append(article_query)
    else:
      funnel_where.append(
        "nmId IN ("
        "SELECT TRIM(\"Артикул WB\") FROM SKU WHERE TRIM(COALESCE(\"Артикул WB\", '')) != ''"
        ")"
      )
    _append_sku_attribute_filter(
      funnel_where,
      funnel_params,
      "nmId",
      subject=subject,
      labels=labels,
      revenue_category=revenue_category,
    )

    funnel_by_date: dict[str, dict[str, float]] = {}
    if funnel_exists:
      funnel_cancel_expr = (
        "SUM(CAST(cancelCount AS REAL))"
        if "cancelCount" in funnel_columns
        else "0"
      )
      funnel_rows = conn.execute(
        (
          "SELECT date, "
          "SUM(CAST(openCount AS REAL)) AS open_count, "
          "SUM(CAST(cartCount AS REAL)) AS cart_count, "
          "SUM(CAST(orderCount AS REAL)) AS order_count, "
          "SUM(CAST(orderSum AS REAL)) AS order_sum, "
          "SUM(CAST(buyoutCount AS REAL)) AS buyout_count, "
          + funnel_cancel_expr +
          " AS cancel_count "
          "FROM funnel_analytics "
          f"WHERE {' AND '.join(funnel_where)} "
          "GROUP BY date ORDER BY date ASC"
        ),
        funnel_params,
      ).fetchall()
      funnel_by_date = {
        str(row["date"]): {
          "open_count": _to_float(row["open_count"]),
          "cart_count": _to_float(row["cart_count"]),
          "order_count": _to_float(row["order_count"]),
          "order_sum": _to_float(row["order_sum"]),
          "buyout_count": _to_float(row["buyout_count"]),
          "cancel_count": _to_float(row["cancel_count"]),
        }
        for row in funnel_rows
      }

    impression_by_date: dict[str, float] = {}
    uploaded_open_by_date: dict[str, float] = {}
    if "funnel_impressions_upload" in {
      str(row["name"])
      for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
    }:
      impression_columns = _table_columns(conn, "funnel_impressions_upload")
      impression_has_supplier_article = "supplierArticle" in impression_columns
      impression_where = ["date >= ?", "date <= ?"]
      impression_params: list[str] = [effective_from, effective_to]
      if selected_articles:
        placeholders = ", ".join("?" for _ in selected_articles)
        if impression_has_supplier_article:
          impression_where.append(
            "(TRIM(CAST(nmId AS TEXT)) IN ("
            + placeholders
            + ") OR TRIM(CAST(supplierArticle AS TEXT)) IN ("
            + placeholders
            + "))"
          )
          impression_params.extend(selected_articles)
        else:
          impression_where.append(
            "TRIM(CAST(nmId AS TEXT)) IN ("
            + placeholders
            + ")"
          )
        impression_params.extend(selected_articles)
      elif article_query:
        if article_query.isdigit():
          impression_where.append("nmId = ?")
          impression_params.append(article_query)
        else:
          impression_where.append("supplierArticle = ?")
          impression_params.append(article_query)
      else:
        impression_where.append(
          "nmId IN ("
          "SELECT TRIM(\"Артикул WB\") FROM SKU WHERE TRIM(COALESCE(\"Артикул WB\", '')) != ''"
          ")"
        )
      _append_sku_attribute_filter(
        impression_where,
        impression_params,
        "nmId",
        subject=subject,
        labels=labels,
        revenue_category=revenue_category,
      )
      impression_rows = conn.execute(
        (
          "SELECT date, "
          "SUM(CAST(impressions AS REAL)) AS impressions, "
          "SUM(CAST(openCount AS REAL)) AS open_count "
          "FROM funnel_impressions_upload "
          f"WHERE {' AND '.join(impression_where)} "
          "GROUP BY date ORDER BY date ASC"
        ),
        impression_params,
      ).fetchall()
      impression_by_date = {
        str(row["date"]): _to_float(row["impressions"])
        for row in impression_rows
      }
      uploaded_open_by_date = {
        str(row["date"]): _to_float(row["open_count"])
        for row in impression_rows
      }

    rows = conn.execute(
      (
        "SELECT "
        "buyout_order_day.\"Дата\", "
        "SUM(CAST(buyout_order_day.\"Сумма заказов\" AS REAL)) AS orders_sum, "
        "SUM(CAST(buyout_order_day.\"Сумма выкупов в эту дату\" AS REAL)) AS event_buyout_sum, "
        "SUM(CAST(buyout_order_day.\"Сумма выкупов товаров, заказанных в эту дату\" AS REAL)) AS ordered_buyout_sum, "
        "SUM(CAST(buyout_order_day.\"К перечислению от выкупов в эту дату\" AS REAL)) AS event_for_pay_sum, "
        "SUM(CAST(buyout_order_day.\"К перечислению от выкупов товаров, заказанных в эту дату\" AS REAL)) AS ordered_for_pay_sum, "
        "SUM(CAST(buyout_order_day.\"Реклама\" AS REAL)) AS ads, "
        "SUM(CAST(buyout_order_day.\"Выкупы в эту дату, шт\" AS REAL)) AS event_buyout_qty, "
        "SUM(CAST(buyout_order_day.\"Выкупы товаров, заказанных в эту дату, шт\" AS REAL)) AS ordered_buyout_qty, "
        "SUM(COALESCE(cogs_data.cogs, 0) * CAST(buyout_order_day.\"Выкупы в эту дату, шт\" AS REAL)) AS event_cogs_sum, "
        "SUM(COALESCE(cogs_data.cogs, 0) * CAST(buyout_order_day.\"Выкупы товаров, заказанных в эту дату, шт\" AS REAL)) AS ordered_cogs_sum, "
        "SUM(CAST(buyout_order_day.\"Заказы, шт\" AS REAL)) AS orders_qty, "
        "SUM(CAST(buyout_order_day.\"Продажи до возвратов, шт\" AS REAL)) AS gross_sale_qty, "
        "SUM(CAST(buyout_order_day.\"Сумма продаж до возвратов\" AS REAL)) AS gross_sale_sum, "
        "SUM(CAST(buyout_order_day.\"Возвраты, шт\" AS REAL)) AS return_qty, "
        "SUM(CAST(buyout_order_day.\"Сумма возвратов\" AS REAL)) AS return_sum "
        "FROM buyout_order_day "
        f"{_buyout_subject_join(conn)} "
        f"{cogs_join} "
        f"{where_sql} GROUP BY buyout_order_day.\"Дата\" ORDER BY buyout_order_day.\"Дата\" ASC"
      ),
      params,
    ).fetchall()

    if article_query and not selected_articles and not rows:
      fallback_where = ['buyout_order_day."Дата" >= ?', 'buyout_order_day."Дата" <= ?']
      fallback_params = [effective_from, effective_to]
      if article_query.isdigit():
        fallback_where.append('buyout_order_day."nmId" = ?')
        fallback_params.append(article_query)
      else:
        fallback_where.append('buyout_order_day."Артикул" LIKE ?')
        fallback_params.append(f"%{article_query}%")
      rows = conn.execute(
        (
          "SELECT "
          "buyout_order_day.\"Дата\", "
          "SUM(CAST(buyout_order_day.\"Сумма заказов\" AS REAL)) AS orders_sum, "
          "SUM(CAST(buyout_order_day.\"Сумма выкупов в эту дату\" AS REAL)) AS event_buyout_sum, "
          "SUM(CAST(buyout_order_day.\"Сумма выкупов товаров, заказанных в эту дату\" AS REAL)) AS ordered_buyout_sum, "
          "SUM(CAST(buyout_order_day.\"К перечислению от выкупов в эту дату\" AS REAL)) AS event_for_pay_sum, "
          "SUM(CAST(buyout_order_day.\"К перечислению от выкупов товаров, заказанных в эту дату\" AS REAL)) AS ordered_for_pay_sum, "
          "SUM(CAST(buyout_order_day.\"Реклама\" AS REAL)) AS ads, "
          "SUM(CAST(buyout_order_day.\"Выкупы в эту дату, шт\" AS REAL)) AS event_buyout_qty, "
          "SUM(CAST(buyout_order_day.\"Выкупы товаров, заказанных в эту дату, шт\" AS REAL)) AS ordered_buyout_qty, "
          "SUM(COALESCE(cogs_data.cogs, 0) * CAST(buyout_order_day.\"Выкупы в эту дату, шт\" AS REAL)) AS event_cogs_sum, "
          "SUM(COALESCE(cogs_data.cogs, 0) * CAST(buyout_order_day.\"Выкупы товаров, заказанных в эту дату, шт\" AS REAL)) AS ordered_cogs_sum, "
          "SUM(CAST(buyout_order_day.\"Заказы, шт\" AS REAL)) AS orders_qty, "
          "SUM(CAST(buyout_order_day.\"Продажи до возвратов, шт\" AS REAL)) AS gross_sale_qty, "
          "SUM(CAST(buyout_order_day.\"Сумма продаж до возвратов\" AS REAL)) AS gross_sale_sum, "
          "SUM(CAST(buyout_order_day.\"Возвраты, шт\" AS REAL)) AS return_qty, "
          "SUM(CAST(buyout_order_day.\"Сумма возвратов\" AS REAL)) AS return_sum "
          "FROM buyout_order_day "
          f"{cogs_join} "
          f"WHERE {' AND '.join(fallback_where)} GROUP BY buyout_order_day.\"Дата\" ORDER BY buyout_order_day.\"Дата\" ASC"
        ),
        fallback_params,
      ).fetchall()

    stock_by_date = _fetch_wb_stock_values_by_date(
      conn,
      effective_from=effective_from,
      effective_to=effective_to,
      article_query=article_query,
      selected_articles=selected_articles,
      subject=subject,
      labels=labels,
      revenue_category=revenue_category,
    )

  row_dicts = [dict(row) for row in rows]
  if granularity == "week":
    dates = []
    for day in daily_dates:
      label = _week_label(day, effective_from, effective_to)
      if not dates or dates[-1] != label:
        dates.append(label)
    row_dicts = _aggregate_daily_dicts_by_week(row_dicts, effective_from, effective_to)
    spp_without_by_date = _aggregate_daily_values_by_week(spp_without_by_date, daily_dates, effective_from, effective_to)
    spp_with_by_date = _aggregate_daily_values_by_week(spp_with_by_date, daily_dates, effective_from, effective_to)
    spp_qty_by_date = _aggregate_daily_values_by_week(spp_qty_by_date, daily_dates, effective_from, effective_to)
    impression_by_date = _aggregate_daily_values_by_week(impression_by_date, daily_dates, effective_from, effective_to)
    uploaded_open_by_date = _aggregate_daily_values_by_week(uploaded_open_by_date, daily_dates, effective_from, effective_to)
    funnel_by_date = _aggregate_nested_daily_values_by_week(funnel_by_date, daily_dates, effective_from, effective_to)
    stock_values_by_date = _stock_values_for_output_dates(
      stock_by_date,
      out_dates=dates,
      daily_dates=daily_dates,
      effective_from=effective_from,
      effective_to=effective_to,
      granularity=granularity,
    )
  else:
    stock_values_by_date = _stock_values_for_output_dates(
      stock_by_date,
      out_dates=dates,
      daily_dates=daily_dates,
      effective_from=effective_from,
      effective_to=effective_to,
      granularity=granularity,
    )

  by_date = {str(row["Дата"]): row for row in row_dicts}
  def day_value(day: str, key: str) -> float:
    row = by_date.get(day)
    value = _to_float(row[key]) if row else 0.0
    # WB buyout_order_day may have zeroed order columns for some slices; fallback to raw_orders aggregates.
    if key == "orders_sum" and value == 0.0:
      return _to_float(spp_without_by_date.get(day, 0.0))
    if key == "orders_qty" and value == 0.0:
      return _to_float(spp_qty_by_date.get(day, 0.0))
    return value

  def format_number(raw: float) -> str:
    return f"{int(round(raw)):,}".replace(",", " ")

  def format_money(label: str, raw: float) -> str:
    return _format_metric(label, raw)

  def format_percent(raw: float) -> str:
    return _format_percent(raw)

  def format_ratio(numerator: float, denominator: float) -> str:
    return _format_percent(numerator / denominator * 100.0) if denominator else "—"

  def format_unit_cost(cost: float, count: float) -> str:
    return f"{(cost / count):,.2f}".replace(",", " ").replace(".", ",") if count else "—"

  def format_total_with_average(total: float, formatter) -> str:
    # В столбце "Итого / среднее" показываем сумму и среднее за видимые даты.
    period_count = len(dates)
    average = total / period_count if period_count else 0.0
    return f"{formatter(total)} / {formatter(average)}"

  metrics = [
    ("Суммы", "Сумма заказов", "orders_sum", "money", "income"),
    ("Суммы", "Сумма выкупов в эту дату", "event_buyout_sum", "money", "income"),
    ("Суммы", "Сумма выкупов товаров, заказанных в эту дату", "ordered_buyout_sum", "money", "income"),
    ("Штуки", "Заказы, шт", "orders_qty", "number", "income"),
    ("Штуки", "Выкупы в эту дату, шт", "event_buyout_qty", "number", "income"),
    ("Штуки", "Выкупы товаров, заказанных в эту дату, шт", "ordered_buyout_qty", "number", "income"),
  ]
  pivot_rows: list[dict[str, object]] = []
  for group, label, key, kind, color_kind in metrics:
    values = []
    raw_values = []
    total = 0.0
    for day in dates:
      raw = day_value(day, key)
      total += raw
      raw_values.append(raw)
      if kind == "number":
        values.append(format_number(raw))
      else:
        values.append(format_money(label, raw))
    if kind == "number":
      total_value = format_total_with_average(total, format_number)
    else:
      total_value = format_total_with_average(total, lambda value: format_money(label, value))
    pivot_rows.append({"group": group, "metric": label, "total": total_value, "values": values, "raw_values": raw_values, "kind": color_kind})

  avg_sales_for_replenishment = 0.0
  if dates:
    avg_sales_for_replenishment = sum(
      max(
        day_value(day, "event_buyout_qty"),
        day_value(day, "ordered_buyout_qty"),
        day_value(day, "orders_qty"),
      )
      for day in dates
    ) / len(dates)
  stock_replenishment_threshold = max(3.0, avg_sales_for_replenishment * 2.0)

  _append_stock_pivot_row(
    pivot_rows,
    group="Остатки",
    label="Остатки, шт",
    dates=dates,
    values_by_date=stock_values_by_date,
    formatter=format_number,
    replenishment_threshold=stock_replenishment_threshold,
  )

  average_check_label = "Средний чек по заказам"
  average_check_values = []
  total_orders_sum = 0.0
  total_orders_qty_for_average = 0.0
  for day in dates:
    orders_sum = day_value(day, "orders_sum")
    orders_qty = day_value(day, "orders_qty")
    total_orders_sum += orders_sum
    total_orders_qty_for_average += orders_qty
    average_check = orders_sum / orders_qty if orders_qty else 0.0
    average_check_values.append(format_money(average_check_label, average_check))
  average_check_raw = []
  for day in dates:
    orders_sum_v = day_value(day, "orders_sum")
    orders_qty_v = day_value(day, "orders_qty")
    average_check_raw.append(orders_sum_v / orders_qty_v if orders_qty_v else 0.0)
  pivot_rows.append({
    "group": "Средние показатели",
    "metric": average_check_label,
    "total": format_money(
      average_check_label,
      total_orders_sum / total_orders_qty_for_average if total_orders_qty_for_average else 0.0,
    ),
    "values": average_check_values,
    "raw_values": average_check_raw,
    "kind": "income",
    "color_mode": "absolute",
    "color_threshold": 10,
  })

  seller_price_values = []
  seller_price_raw = []
  total_spp_qty = 0.0
  total_without_spp_for_average = 0.0
  for day in dates:
    without_spp = spp_without_by_date.get(day, 0.0)
    qty = spp_qty_by_date.get(day, 0.0)
    total_spp_qty += qty
    total_without_spp_for_average += without_spp
    seller_price = without_spp / qty if qty else 0.0
    seller_price_values.append(format_money("Средняя цена продавца", seller_price))
    seller_price_raw.append(seller_price)
  pivot_rows.append({
    "group": "Средние показатели",
    "metric": "Средняя цена продавца",
    "total": format_money(
      "Средняя цена продавца",
      total_without_spp_for_average / total_spp_qty if total_spp_qty else 0.0,
    ),
    "values": seller_price_values,
    "raw_values": seller_price_raw,
    "kind": "income",
    "color_mode": "absolute",
    "color_threshold": 10,
  })

  spp_values = []
  spp_raw = []
  total_without_spp = 0.0
  total_with_spp = 0.0
  for day in dates:
    without_spp = spp_without_by_date.get(day, 0.0)
    with_spp = spp_with_by_date.get(day, 0.0)
    total_without_spp += without_spp
    total_with_spp += with_spp
    spp_percent = (without_spp - with_spp) / without_spp * 100.0 if without_spp else 0.0
    spp_values.append(format_percent(spp_percent))
    spp_raw.append(spp_percent)
  total_spp_percent = (
    (total_without_spp - total_with_spp) / total_without_spp * 100.0
    if total_without_spp else 0.0
  )
  pivot_rows.append({
    "group": "Средние показатели",
    "metric": "СПП",
    "total": format_percent(total_spp_percent),
    "values": spp_values,
    "raw_values": spp_raw,
    "kind": "percent_income",
    "color_threshold": 3,
  })

  client_price_values = []
  client_price_raw = []
  total_client_price_sum = 0.0
  total_client_price_qty = 0.0
  for day in dates:
    with_spp = spp_with_by_date.get(day, 0.0)
    qty = spp_qty_by_date.get(day, 0.0)
    total_client_price_sum += with_spp
    total_client_price_qty += qty
    client_price = with_spp / qty if qty else 0.0
    client_price_values.append(format_money("Средняя цена клиента", client_price))
    client_price_raw.append(client_price)
  pivot_rows.append({
    "group": "Средние показатели",
    "metric": "Средняя цена клиента",
    "total": format_money(
      "Средняя цена клиента",
      total_client_price_sum / total_client_price_qty if total_client_price_qty else 0.0,
    ),
    "values": client_price_values,
    "raw_values": client_price_raw,
    "kind": "income",
    "color_mode": "absolute",
    "color_threshold": 10,
  })

  def funnel_value(day: str, key: str) -> float:
    value = float(funnel_by_date.get(day, {}).get(key, 0.0))
    if key == "order_count" and value == 0.0 and float(funnel_by_date.get(day, {}).get("order_sum", 0.0)) > 0:
      return day_value(day, "orders_qty")
    return value

  def open_value(day: str) -> float:
    uploaded_value = uploaded_open_by_date.get(day)
    if uploaded_value is not None:
      return uploaded_value
    return funnel_value(day, "open_count")

  funnel_specs = [
    ("Воронка", "Показы", lambda day: impression_by_date.get(day, 0.0), "number", "income"),
    ("Воронка", "Переходы", lambda day: open_value(day), "number", "income"),
    ("Воронка", "CTR", lambda day: (open_value(day), impression_by_date.get(day, 0.0)), "percent_ratio", "percent_income"),
    ("Воронка", "Корзины", lambda day: funnel_value(day, "cart_count"), "number", "income"),
    ("Воронка", "CR1", lambda day: (funnel_value(day, "cart_count"), open_value(day)), "percent_ratio", "percent_income"),
    ("Воронка", "Заказы", lambda day: funnel_value(day, "order_count"), "number", "income"),
    ("Воронка", "CR2", lambda day: (funnel_value(day, "order_count"), funnel_value(day, "cart_count")), "percent_ratio", "percent_income"),
    ("Воронка", "% выкупа WB", lambda day: (funnel_value(day, "buyout_count"), funnel_value(day, "buyout_count") + funnel_value(day, "cancel_count")), "percent_ratio", "percent_income"),
    ("Воронка", "% выкупа от заказов", lambda day: (funnel_value(day, "buyout_count"), funnel_value(day, "order_count")), "percent_ratio", "percent_income"),
    ("Стоимость воронки", "Цена показа", lambda day: (day_value(day, "ads"), impression_by_date.get(day, 0.0)), "money_ratio", "expense"),
    ("Стоимость воронки", "Цена клика", lambda day: (day_value(day, "ads"), open_value(day)), "money_ratio", "expense"),
    ("Стоимость воронки", "Цена корзины", lambda day: (day_value(day, "ads"), funnel_value(day, "cart_count")), "money_ratio", "expense"),
    ("Стоимость воронки", "Цена заказа", lambda day: (day_value(day, "ads"), funnel_value(day, "order_count")), "money_ratio", "expense"),
  ]
  for group, label, value_fn, value_kind, color_kind in funnel_specs:
    values = []
    raw_values = []
    total_numerator = 0.0
    total_denominator = 0.0
    total_value_raw = 0.0
    for day in dates:
      raw = value_fn(day)
      if value_kind in {"percent_ratio", "money_ratio"}:
        numerator, denominator = raw
        if denominator:
          total_numerator += numerator
          total_denominator += denominator
        raw_value = numerator / denominator * (100.0 if value_kind == "percent_ratio" else 1.0) if denominator else None
        raw_values.append(raw_value)
        values.append(format_ratio(numerator, denominator) if value_kind == "percent_ratio" else format_unit_cost(numerator, denominator))
      else:
        total_value_raw += raw
        raw_values.append(raw)
        values.append(format_number(raw))
    if value_kind == "percent_ratio":
      total_value = format_ratio(total_numerator, total_denominator)
    elif value_kind == "money_ratio":
      total_value = format_unit_cost(total_numerator, total_denominator)
    else:
      total_value = format_total_with_average(total_value_raw, format_number)
    pivot_rows.append({"group": group, "metric": label, "total": total_value, "values": values, "raw_values": raw_values, "kind": color_kind})

  expense_metrics = [
    ("Расходы", "Реклама", "ads", "money", "expense"),
  ]
  for group, label, key, kind, color_kind in expense_metrics:
    values = []
    raw_values = []
    total = 0.0
    for day in dates:
      raw = day_value(day, key)
      total += raw
      raw_values.append(raw)
      values.append(format_number(raw) if kind == "number" else format_money(label, raw))
    if kind == "number":
      total_value = format_total_with_average(total, format_number)
    else:
      total_value = format_total_with_average(total, lambda value: format_money(label, value))
    pivot_rows.append({"group": group, "metric": label, "total": total_value, "values": values, "raw_values": raw_values, "kind": color_kind})

  derived_metrics = [
    (
      "Расходы",
      "ДРР от заказов",
      lambda row, ads: ads,
      lambda total_num, total_den: _safe_percent(total_num, total_den),
      lambda num, den: _safe_percent(num, den),
      "orders_sum",
      "percent_expense",
    ),
    (
      "Расходы",
      "ДРР от выкупов в эту дату",
      lambda row, ads: ads,
      lambda total_num, total_den: _safe_percent(total_num, total_den),
      lambda num, den: _safe_percent(num, den),
      "event_buyout_sum",
      "percent_expense",
    ),
    (
      "Доходность",
      "Себестоимость выкупов в эту дату",
      lambda row, ads: _to_float(row["event_cogs_sum"]) if row else 0.0,
      lambda total_num, total_den: format_money("Себестоимость выкупов в эту дату", total_num),
      lambda num, den: format_money("Себестоимость выкупов в эту дату", num),
      "",
      "expense",
    ),
    (
      "Доходность",
      "Прибыль от выкупов в эту дату",
      lambda row, ads: _to_float(row["event_for_pay_sum"]) - ads - _to_float(row["event_cogs_sum"]) if row else -ads,
      lambda total_num, total_den: format_money("Прибыль от выкупов в эту дату", total_num),
      lambda num, den: format_money("Прибыль от выкупов в эту дату", num),
      "",
      "income",
    ),
    (
      "Доходность",
      "Маржинальность от выкупов за дату",
      lambda row, ads: (_to_float(row["event_for_pay_sum"]) - ads - _to_float(row["event_cogs_sum"])) if row else -ads,
      lambda total_num, total_den: _safe_percent(total_num, total_den),
      lambda num, den: _safe_percent(num, den),
      "event_buyout_sum",
      "percent_income",
    ),
  ]

  for metric in derived_metrics:
    group, label, numerator_fn, total_formatter, value_formatter, denominator_name, color_kind = metric
    values = []
    raw_values = []
    total_numerator = 0.0
    total_denominator = 0.0
    for day in dates:
      row = by_date.get(day)
      ads = day_value(day, "ads")
      numerator = numerator_fn(row, ads)
      denominator = day_value(day, denominator_name) if denominator_name else 0.0
      total_numerator += numerator
      total_denominator += denominator
      raw_values.append(numerator if not denominator_name else (numerator / denominator * 100.0 if abs(denominator) >= 100.0 else None))
      values.append(value_formatter(numerator, denominator))
    if denominator_name:
      total_value = total_formatter(total_numerator, total_denominator)
    else:
      total_value = format_total_with_average(
        total_numerator,
        lambda value: total_formatter(value, 0.0),
      )
    pivot_rows.append({"group": group, "metric": label, "total": total_value, "values": values, "raw_values": raw_values, "kind": color_kind})

  return {
    "dates": dates,
    "rows": pivot_rows,
    "effective_from": effective_from,
    "effective_to": effective_to,
    "max_days": 93 if granularity == "week" else 45,
    "granularity": granularity,
  }

def _calc_elasticity(points: list[tuple[float, float]]) -> float | None:
  import math
  prepared = [(price, qty) for price, qty in points if price > 0 and qty > 0]
  if len(prepared) < 7:
    return None
  prices = [price for price, _qty in prepared]
  if max(prices) / min(prices) < 1.05:
    return None
  # require at least 2 distinct price clusters separated by >3%
  sorted_prices = sorted(set(round(p / 10) * 10 for p in prices))
  has_gap = any(sorted_prices[i + 1] / sorted_prices[i] >= 1.03 for i in range(len(sorted_prices) - 1))
  if not has_gap:
    return None
  xs = [math.log(price) for price, _qty in prepared]
  ys = [math.log(qty) for _price, qty in prepared]
  n = len(xs)
  avg_x = sum(xs) / n
  avg_y = sum(ys) / n
  ss_xx = sum((x - avg_x) ** 2 for x in xs)
  if ss_xx <= 0:
    return None
  ss_xy = sum((x - avg_x) * (y - avg_y) for x, y in zip(xs, ys))
  value = ss_xy / ss_xx
  # R² — share of demand variance explained by price
  ss_yy = sum((y - avg_y) ** 2 for y in ys)
  if ss_yy <= 0:
    return None
  r2 = (ss_xy ** 2) / (ss_xx * ss_yy)
  if r2 < 0.3:
    return None
  # clip to economically plausible range for WB: demand drops when price rises
  return max(-3.0, min(0.0, value))

def _calc_demand_model(
  points: list[tuple[float, float, float]],  # (price, ads_per_day, qty)
) -> tuple[float, float] | None:
  """Bivariate log-log OLS: log(qty) = e_price*log(price) + e_ads*log(1+ads) + const.
  Returns (e_price, e_ads). Falls back to one-factor models when only price or ads has signal."""
  import math
  prepared = [(p, a, q) for p, a, q in points if p > 0 and q > 0]
  if len(prepared) < 7:
    return None
  prices = [p for p, _, _ in prepared]
  ads_values = [a for _, a, _ in prepared]
  sorted_prices = sorted(set(round(p / 10) * 10 for p in prices))
  has_price_signal = (
    max(prices) / min(prices) >= 1.05
    and any(sorted_prices[i + 1] / sorted_prices[i] >= 1.03 for i in range(len(sorted_prices) - 1))
  )
  has_ads_signal = max(ads_values) > 0 and (max(ads_values) - min(ads_values)) > max(50.0, max(ads_values) * 0.15)
  xs1 = [math.log(p) for p, _, _ in prepared]
  xs2 = [math.log(1.0 + a) for _, a, _ in prepared]
  ys  = [math.log(q) for _, _, q in prepared]
  n = len(xs1)
  if not has_price_signal and has_ads_signal:
    mx = sum(xs2) / n
    my = sum(ys) / n
    ss_xx = sum((x - mx) ** 2 for x in xs2)
    ss_xy = sum((x - mx) * (y - my) for x, y in zip(xs2, ys))
    ss_yy = sum((y - my) ** 2 for y in ys)
    if ss_xx <= 0 or ss_yy <= 0:
      return None
    r2 = (ss_xy ** 2) / (ss_xx * ss_yy)
    if r2 < 0.12:
      return None
    return 0.0, max(0.0, min(1.5, ss_xy / ss_xx))
  if not has_price_signal:
    return None
  mx1 = sum(xs1) / n; mx2 = sum(xs2) / n; my = sum(ys) / n
  x1 = [v - mx1 for v in xs1]; x2 = [v - mx2 for v in xs2]; y = [v - my for v in ys]
  sa = sum(v * v for v in x1)
  sb = sum(i * j for i, j in zip(x1, x2))
  sd = sum(v * v for v in x2)
  sf = sum(i * j for i, j in zip(x1, y))
  sg = sum(i * j for i, j in zip(x2, y))
  det = sa * sd - sb * sb
  ads_has_variation = sd > 1e-12
  if abs(det) < 1e-12 or not ads_has_variation:
    e_p = _calc_elasticity([(p, q) for p, _, q in prepared])
    return (e_p, 0.0) if e_p is not None else None
  e_price_raw = (sd * sf - sb * sg) / det
  e_ads_raw   = (sa * sg - sb * sf) / det
  ss_tot = sum(v * v for v in y)
  if ss_tot <= 0:
    return None
  pred = [e_price_raw * i + e_ads_raw * j for i, j in zip(x1, x2)]
  ss_res = sum((yi - pi) ** 2 for yi, pi in zip(y, pred))
  r2 = 1.0 - ss_res / ss_tot
  if r2 < 0.3:
    e_p = _calc_elasticity([(p, q) for p, _, q in prepared])
    return (e_p, 0.0) if e_p is not None else None
  return max(-3.0, min(0.0, e_price_raw)), max(0.0, min(1.5, e_ads_raw))

def _calc_ads_confidence(stats: dict[str, object] | None) -> float:
  if not stats:
    return 0.0
  import math
  ad_days = _to_float(stats.get("ad_days") or 0)
  ad_orders = _to_float(stats.get("ad_orders") or 0)
  ad_views = _to_float(stats.get("ad_views") or 0)
  spend_values = [float(v) for v in (stats.get("spend_values") or []) if _to_float(v) > 0]
  variation = 0.0
  if len(spend_values) >= 5 and min(spend_values) > 0:
    variation = min(1.0, math.log(max(spend_values) / min(spend_values)) / math.log(5.0))
  confidence = (
    min(1.0, ad_days / 30.0) * 0.30
    + min(1.0, ad_orders / 50.0) * 0.30
    + min(1.0, ad_views / 50000.0) * 0.25
    + variation * 0.15
  )
  return max(0.0, min(1.0, confidence))

def _effective_ads_elasticity(measured: float | None, stats: dict[str, object] | None, *, group_fallback: bool = False) -> float:
  prior = 0.22
  confidence = _calc_ads_confidence(stats)
  if group_fallback:
    confidence = min(0.35, confidence * 0.5)
  else:
    confidence = min(0.85, confidence)
  measured_v = _to_float(measured)
  if measured_v <= 0:
    confidence = 0.0
  value = confidence * measured_v + (1.0 - confidence) * prior
  return max(0.08, min(0.85, value))

def _simulate_planning_orders(
  stock: float,
  daily_demand: float,
  buyout_rate: float,
  forecast_days: int,
  return_delay_days: int,
  start_day: date,
  warehouse_stock: float = 0.0,
  transit_days: int = 0,
) -> tuple[float, str]:
  available = max(0.0, stock)
  demand = max(0.0, daily_demand)
  buyout_rate = max(0.0, min(1.0, buyout_rate))
  delay = max(1, int(return_delay_days))
  transit = max(0, int(transit_days))
  # warehouse stock arrives at WB on day transit
  arrivals: dict[int, float] = {}
  if warehouse_stock > 0:
    arrivals[transit] = max(0.0, warehouse_stock)
  returns_by_day: dict[int, float] = {}
  total_orders = 0.0
  stockout_date = ""
  for idx in range(max(0, forecast_days)):
    available += returns_by_day.pop(idx, 0.0) + arrivals.pop(idx, 0.0)
    planned = demand
    ordered = min(planned, available)
    if planned > available and not stockout_date:
      stockout_date = (start_day + timedelta(days=idx)).isoformat()
    available -= ordered
    total_orders += ordered
    returning_qty = ordered * (1.0 - buyout_rate)
    due_idx = idx + delay
    if returning_qty > 0 and due_idx < forecast_days:
      returns_by_day[due_idx] = returns_by_day.get(due_idx, 0.0) + returning_qty
  return total_orders, stockout_date

def _fetch_planning(
  forecast_to: str,
  return_delay_days: int = 8,
  transit_days: int = 3,
  subject: str = "",
  labels: list[str] | None = None,
  revenue_category: str = "",
  price_change_pct: float = 0.0,
) -> dict[str, object]:
  target_date = date.fromisoformat(forecast_to)
  subject = subject.strip()
  labels = [l.strip() for l in (labels or []) if l.strip()]
  revenue_category = revenue_category.strip()
  price_change_pct = max(-90.0, min(300.0, float(price_change_pct)))
  return_delay_days = max(1, int(return_delay_days))
  transit_days = max(0, int(transit_days))

  with _db_connect() as conn:
    latest_raw = conn.execute('SELECT MAX("Дата") AS max_date FROM buyout_order_day').fetchone()["max_date"]
    if not latest_raw:
      return {"summary": {}, "rows": [], "options": {"subjects": [], "labels": [], "revenue_categories": []}}
    latest_day = date.fromisoformat(str(latest_raw))
    # exclude the latest day from speed/funnel baseline — it may be partial
    speed_baseline_to = latest_day - timedelta(days=1)
    baseline_from = (speed_baseline_to - timedelta(days=6)).isoformat()
    baseline_to = speed_baseline_to.isoformat()
    # buyout rate uses older data (14–44 days ago) — recent data is incomplete
    # because WB takes 7–14 days to finalize buyout/cancel counts
    buyout_rate_to = (latest_day - timedelta(days=14)).isoformat()
    buyout_rate_from = (latest_day - timedelta(days=44)).isoformat()
    # ads baseline: 14 days ending 2 days before latest (позавчера)
    ads_day_count = 14.0
    ads_baseline_to = (latest_day - timedelta(days=2)).isoformat()
    ads_baseline_from = (latest_day - timedelta(days=15)).isoformat()
    forecast_days = max(0, (target_date - latest_day).days)
    forecast_start = latest_day + timedelta(days=1)

    ctx = _scope_context(conn)
    sku_where = ["TRIM(COALESCE(\"Артикул WB\", '')) != ''"]
    sku_params: list[str] = []
    if subject:
      sku_where.append('"Предмет" = ?')
      sku_params.append(subject)
    if labels:
      lc_sku=[]; lp_sku:list[str]=[]
      for lbl in labels:
        lc_sku.append("(INSTR(',' || TRIM(COALESCE(\"Ярлыки\",'')) || ',', ',' || ? || ',') > 0)")
        lp_sku.append(lbl)
      sku_where.append('('+' OR '.join(lc_sku)+')')
      sku_params.extend(lp_sku)
    if revenue_category:
      sku_where.append('"Категория по выручке" = ?')
      sku_params.append(revenue_category)
    sku_rows = conn.execute(
      (
        'SELECT TRIM("Артикул WB") AS nmid, '
        'COALESCE(NULLIF(TRIM("Артикул поставщика"), \'\'), TRIM("Артикул WB")) AS article, '
        'COALESCE("Название", "Наименование", "") AS name, '
        'COALESCE("Предмет", "") AS subject, COALESCE("Ярлыки", "") AS strategy, '
        'COALESCE("Категория по выручке", "") AS revenue_category, '
        'COALESCE(CAST(REPLACE(REPLACE(COALESCE("себестоимость","0")," ",""),",",".") AS REAL), 0) AS cogs, '
        'COALESCE(CAST("склад" AS REAL), 0) AS warehouse_stock, '
        'COALESCE(CAST(REPLACE(REPLACE(COALESCE("% комиссии на вб","0"),"%","")," ","") AS REAL), 0) AS wb_commission_pct '
        'FROM SKU '
        f'WHERE {" AND ".join(sku_where)}'
      ),
      sku_params,
    ).fetchall()
    sku_by_nm = {str(row["nmid"]): dict(row) for row in sku_rows if str(row["nmid"]).strip()}

    options = _fetch_buyout_filter_options(baseline_from, baseline_to)
    if not sku_by_nm:
      return {
        "summary": {"forecast_days": forecast_days, "baseline_from": baseline_from, "baseline_to": baseline_to},
        "rows": [],
        "options": options,
      }

    nm_ids = sorted(sku_by_nm)
    marketplace_article_by_nm = _marketplace_article_by_nm(conn, nm_ids)
    placeholders = ", ".join("?" for _ in nm_ids)
    stock_rows = conn.execute(
      (
        'SELECT "nmId" AS nmid, SUM(CAST(REPLACE(REPLACE("quantity", \' \', \'\'), \',\', \'.\') AS REAL)) AS stock '
        'FROM raw_stocks '
        f'WHERE "nmId" IN ({placeholders}) GROUP BY "nmId"'
      ),
      nm_ids,
    ).fetchall()
    stock_by_nm = {str(row["nmid"]): _to_float(row["stock"]) for row in stock_rows}

    metric_rows = conn.execute(
      (
        'SELECT "nmId" AS nmid, SUM(CAST("Заказы, шт" AS REAL)) AS orders_qty, '
        'SUM(CAST("Сумма заказов" AS REAL)) AS orders_sum '
        'FROM buyout_order_day '
        f'WHERE "Дата" >= ? AND "Дата" <= ? AND "nmId" IN ({placeholders}) '
        'GROUP BY "nmId"'
      ),
      [baseline_from, baseline_to, *nm_ids],
    ).fetchall()
    metrics_by_nm = {str(row["nmid"]): row for row in metric_rows}

    _cancel_expr = 'SUM(CAST("cancelCount" AS REAL))' if "cancelCount" in _table_columns(conn, "funnel_analytics") else "0"
    buyout_rows = conn.execute(
      (
        'SELECT "nmId" AS nmid, SUM(CAST("buyoutCount" AS REAL)) AS buyouts, '
        f'{_cancel_expr} AS cancels '
        'FROM funnel_analytics '
        f'WHERE "date" >= ? AND "date" <= ? AND "nmId" IN ({placeholders}) '
        'GROUP BY "nmId"'
      ),
      [baseline_from, baseline_to, *nm_ids],
    ).fetchall()
    wb_buyout_by_nm = {str(row["nmid"]): (_to_float(row["buyouts"]), _to_float(row["cancels"])) for row in buyout_rows}

    # ABC-классификация по выкупам (buyoutSum) за 7 дней: A=80%, B=15%, C=5%
    abc_sum_rows = conn.execute(
      (
        'SELECT "nmId" AS nmid, SUM(CAST("buyoutSum" AS REAL)) AS bsum '
        'FROM funnel_analytics '
        f'WHERE "date" >= ? AND "date" <= ? AND "nmId" IN ({placeholders}) '
        'GROUP BY "nmId"'
      ),
      [baseline_from, baseline_to, *nm_ids],
    ).fetchall()
    abc_sum_by_nm: dict[str, float] = {str(r["nmid"]): _to_float(r["bsum"]) for r in abc_sum_rows}
    total_abc = sum(abc_sum_by_nm.values())
    abc_category_by_nm: dict[str, str] = {}
    if total_abc > 0:
      sorted_nms = sorted(nm_ids, key=lambda n: abc_sum_by_nm.get(n, 0.0), reverse=True)
      cum = 0.0
      for n in sorted_nms:
        cum += abc_sum_by_nm.get(n, 0.0)
        pct = cum / total_abc
        if pct <= 0.80:
          abc_category_by_nm[n] = "A"
        elif pct <= 0.95:
          abc_category_by_nm[n] = "B"
        else:
          abc_category_by_nm[n] = "C"
      options["revenue_categories"] = ["A", "B", "C"]

    group_wb_buyout: dict[str, tuple[float, float]] = {}
    group_check: dict[str, tuple[float, float]] = {}
    total_wb_buyouts = 0.0
    total_wb_cancels = 0.0
    for nm_id, sku in sku_by_nm.items():
      row = metrics_by_nm.get(nm_id)
      orders_qty = _to_float(row["orders_qty"]) if row else 0.0
      orders_sum = _to_float(row["orders_sum"]) if row else 0.0
      wb_buyouts, wb_cancels = wb_buyout_by_nm.get(nm_id, (0.0, 0.0))
      total_wb_buyouts += wb_buyouts
      total_wb_cancels += wb_cancels
      computed_cat = abc_category_by_nm.get(nm_id, "") if total_abc > 0 else str(sku.get("revenue_category") or "")
      for key in (str(sku.get("subject") or ""), str(sku.get("strategy") or ""), computed_cat):
        if not key:
          continue
        bought, canceled = group_wb_buyout.get(key, (0.0, 0.0))
        group_wb_buyout[key] = (bought + wb_buyouts, canceled + wb_cancels)
        amount, qty = group_check.get(key, (0.0, 0.0))
        group_check[key] = (amount + orders_sum, qty + orders_qty)

    order_columns = _table_columns(conn, "raw_orders")
    price_col = "finishedPrice" if "finishedPrice" in order_columns else "priceWithDisc"
    elasticity_by_nm: dict[str, float] = {}
    group_elasticities: dict[str, list[float]] = {}
    if {"date", "nmId", price_col}.issubset(order_columns):
      history_from = (latest_day - timedelta(days=59)).isoformat()
      rows = conn.execute(
        (
          f'SELECT substr("date", 1, 10) AS day, "nmId" AS nmid, '
          f'AVG(CAST(REPLACE(REPLACE("{price_col}", \' \', \'\'), \',\', \'.\') AS REAL)) AS price, COUNT(*) AS qty '
          'FROM raw_orders '
          f'WHERE substr("date", 1, 10) >= ? AND "nmId" IN ({placeholders}) '
          'AND COALESCE("isCancel", "") NOT IN ("true", "1", "True") '
          'GROUP BY "nmId", substr("date", 1, 10)'
        ),
        [history_from, *nm_ids],
      ).fetchall()
      # store (day, price, qty) so we can join with ads by day
      day_points_by_nm: dict[str, list[tuple[str, float, float]]] = {}
      points_by_nm: dict[str, list[tuple[float, float]]] = {}
      for row in rows:
        nm = str(row["nmid"])
        p = _to_float(row["price"]); q = _to_float(row["qty"])
        day_points_by_nm.setdefault(nm, []).append((str(row["day"]), p, q))
        points_by_nm.setdefault(nm, []).append((p, q))

      # 60-day daily ads spend per nmId for bivariate model
      ads_tables = {str(r["name"]) for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
      ads_by_day_nm: dict[str, dict[str, float]] = {}
      ads_stats_by_nm: dict[str, dict[str, object]] = {}
      if "raw_ads" in ads_tables:
        ads_rows = conn.execute(
          f'SELECT substr("date",1,10) AS day, "nmId" AS nmid, '
          f'SUM(CAST("sum" AS REAL)) AS ads_sum, '
          f'SUM(CAST(COALESCE("views", "0") AS REAL)) AS ad_views, '
          f'SUM(CAST(COALESCE("clicks", "0") AS REAL)) AS ad_clicks, '
          f'SUM(CAST(COALESCE("orders", "0") AS REAL)) AS ad_orders '
          f'FROM raw_ads WHERE substr("date",1,10) >= ? AND "nmId" IN ({placeholders}) '
          'GROUP BY "nmId", substr("date",1,10)',
          [history_from, *nm_ids],
        ).fetchall()
        for ar in ads_rows:
          nm = str(ar["nmid"])
          ads_sum = _to_float(ar["ads_sum"])
          ads_by_day_nm.setdefault(nm, {})[str(ar["day"])] = ads_sum
          stats = ads_stats_by_nm.setdefault(nm, {
            "ad_days": 0.0,
            "ad_spend": 0.0,
            "ad_views": 0.0,
            "ad_clicks": 0.0,
            "ad_orders": 0.0,
            "spend_values": [],
          })
          if ads_sum > 0 or _to_float(ar["ad_views"]) > 0:
            stats["ad_days"] = _to_float(stats.get("ad_days") or 0) + 1.0
          stats["ad_spend"] = _to_float(stats.get("ad_spend") or 0) + ads_sum
          stats["ad_views"] = _to_float(stats.get("ad_views") or 0) + _to_float(ar["ad_views"])
          stats["ad_clicks"] = _to_float(stats.get("ad_clicks") or 0) + _to_float(ar["ad_clicks"])
          stats["ad_orders"] = _to_float(stats.get("ad_orders") or 0) + _to_float(ar["ad_orders"])
          stats.setdefault("spend_values", []).append(ads_sum)

      ads_elasticity_by_nm: dict[str, float] = {}
      group_ads_elasticities: dict[str, list[float]] = {}
      for nm_id, day_pts in day_points_by_nm.items():
        ads_day = ads_by_day_nm.get(nm_id, {})
        pts3: list[tuple[float, float, float]] = [
          (p, ads_day.get(d, 0.0), q) for d, p, q in day_pts
        ]
        result = _calc_demand_model(pts3) if pts3 else None
        if result is not None:
          e_p, e_a = result
        else:
          price_pts = [(p, q) for _, p, q in day_pts]
          e_p_only = _calc_elasticity(price_pts)
          if e_p_only is None:
            continue
          e_p, e_a = e_p_only, 0.0
        elasticity_by_nm[nm_id] = e_p
        if e_a > 0:
          ads_elasticity_by_nm[nm_id] = e_a
        sku = sku_by_nm.get(nm_id, {})
        _cat = abc_category_by_nm.get(nm_id, "") if total_abc > 0 else str(sku.get("revenue_category") or "")
        for key in (str(sku.get("subject") or ""), str(sku.get("strategy") or ""), _cat):
          if key:
            group_elasticities.setdefault(key, []).append(e_p)
            if e_a > 0:
              group_ads_elasticities.setdefault(key, []).append(e_a)

      # 14-day ads baseline (ending 2 days before latest) for DRR calculation
      base_ads_by_nm: dict[str, float] = {}
      if "raw_ads" in ads_tables:
        base_ads_rows = conn.execute(
          f'SELECT "nmId" AS nmid, SUM(CAST("sum" AS REAL)) AS ads_sum '
          f'FROM raw_ads WHERE "date" >= ? AND "date" <= ? AND "nmId" IN ({placeholders}) '
          'GROUP BY "nmId"',
          [ads_baseline_from, ads_baseline_to, *nm_ids],
        ).fetchall()
        for ar in base_ads_rows:
          base_ads_by_nm[str(ar["nmid"])] = _to_float(ar["ads_sum"])

      # actual buyout revenue over the same 14-day ads window (for DRR denominator)
      ads_revenue_rows = conn.execute(
        f'SELECT "nmId" AS nmid, SUM(CAST("buyoutSum" AS REAL)) AS buyout_sum '
        f'FROM funnel_analytics '
        f'WHERE "date" >= ? AND "date" <= ? AND "nmId" IN ({placeholders}) '
        'GROUP BY "nmId"',
        [ads_baseline_from, ads_baseline_to, *nm_ids],
      ).fetchall()
      ads_revenue_by_nm: dict[str, float] = {str(r["nmid"]): _to_float(r["buyout_sum"]) for r in ads_revenue_rows}

    else:
      ads_elasticity_by_nm = {}
      group_ads_elasticities = {}
      base_ads_by_nm = {}
      ads_revenue_by_nm = {}
      ads_stats_by_nm = {}

    # 7-day funnel baseline: CTR, CR1, CR2 from funnel_impressions_upload
    funnel_by_nm: dict[str, dict[str, float]] = {}
    all_tables = {str(r["name"]) for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    if "funnel_impressions_upload" in all_tables:
      fi_rows = conn.execute(
        f'SELECT "nmId" AS nmid, '
        f'SUM(CAST("impressions" AS REAL)) AS impr, '
        f'SUM(CAST("openCount" AS REAL)) AS opens, '
        f'SUM(CAST("cartCount" AS REAL)) AS carts, '
        f'SUM(CAST("orderCount" AS REAL)) AS orders_f '
        f'FROM funnel_impressions_upload '
        f'WHERE "date" >= ? AND "date" <= ? AND "nmId" IN ({placeholders}) '
        'GROUP BY "nmId"',
        [baseline_from, baseline_to, *nm_ids],
      ).fetchall()
      for fi in fi_rows:
        nm = str(fi["nmid"])
        impr  = _to_float(fi["impr"])
        opens = _to_float(fi["opens"])
        carts = _to_float(fi["carts"])
        ords  = _to_float(fi["orders_f"])
        funnel_by_nm[nm] = {
          "ctr":  opens / impr  * 100.0 if impr  > 0 else 0.0,
          "cr1":  carts / opens * 100.0 if opens > 0 else 0.0,
          "cr2":  ords  / carts * 100.0 if carts > 0 else 0.0,
        }

    current_price_by_nm: dict[str, float] = {}
    current_spp_by_nm: dict[str, float] = {}
    client_price_by_nm: dict[str, float] = {}
    baseline_avg_price_by_nm: dict[str, float] = {}
    baseline_client_check_by_nm: dict[str, float] = {}
    has_finished = "finishedPrice" in order_columns and "priceWithDisc" in order_columns
    if {"date", "nmId", price_col}.issubset(order_columns):
      seller_col = "priceWithDisc"
      client_col = "finishedPrice" if has_finished else "priceWithDisc"
      last_day_rows = conn.execute(
        (
          f'SELECT r."nmId" AS nmid, '
          f'AVG(CAST(REPLACE(REPLACE(r."{seller_col}", \' \', \'\'), \',\', \'.\') AS REAL)) AS seller_price, '
          f'AVG(CAST(REPLACE(REPLACE(r."{client_col}", \' \', \'\'), \',\', \'.\') AS REAL)) AS client_price, '
          f'AVG(CAST(REPLACE(REPLACE(COALESCE(r."spp", "0"), \' \', \'\'), \',\', \'.\') AS REAL)) AS spp '
          'FROM raw_orders r '
          f'JOIN (SELECT "nmId", MAX(substr("date",1,10)) AS max_date FROM raw_orders WHERE "nmId" IN ({placeholders}) GROUP BY "nmId") md '
          'ON r."nmId" = md."nmId" AND substr(r."date",1,10) = md.max_date '
          f'WHERE r."nmId" IN ({placeholders}) '
          'GROUP BY r."nmId"'
        ),
        [*nm_ids, *nm_ids],
      ).fetchall()
      for row in last_day_rows:
        if _to_float(row["seller_price"]) > 0:
          current_price_by_nm[str(row["nmid"])] = _to_float(row["seller_price"])
        if _to_float(row["client_price"]) > 0:
          client_price_by_nm[str(row["nmid"])] = _to_float(row["client_price"])
        current_spp_by_nm[str(row["nmid"])] = _to_float(row["spp"])
      baseline_rows = conn.execute(
        (
          f'SELECT "nmId" AS nmid, '
          f'AVG(CAST(REPLACE(REPLACE("{seller_col}", \' \', \'\'), \',\', \'.\') AS REAL)) AS price, '
          f'AVG(CAST(REPLACE(REPLACE("{client_col}", \' \', \'\'), \',\', \'.\') AS REAL)) AS client_price, '
          f'AVG(CAST(REPLACE(REPLACE(COALESCE("spp","0"), \' \', \'\'), \',\', \'.\') AS REAL)) AS avg_spp '
          'FROM raw_orders '
          f'WHERE substr("date", 1, 10) >= ? AND substr("date", 1, 10) <= ? AND "nmId" IN ({placeholders}) '
          'GROUP BY "nmId"'
        ),
        [baseline_from, baseline_to, *nm_ids],
      ).fetchall()
      baseline_spp_by_nm: dict[str, float] = {}
      baseline_client_check_by_nm: dict[str, float] = {}
      for row in baseline_rows:
        if _to_float(row["price"]) > 0:
          baseline_avg_price_by_nm[str(row["nmid"])] = _to_float(row["price"])
        if _to_float(row["client_price"]) > 0:
          baseline_client_check_by_nm[str(row["nmid"])] = _to_float(row["client_price"])
        baseline_spp_by_nm[str(row["nmid"])] = _to_float(row["avg_spp"])

    def _sku_keys(nm_id: str, sku: dict[str, object]) -> tuple[str, ...]:
      cat = abc_category_by_nm.get(nm_id, "") if total_abc > 0 else str(sku.get("revenue_category") or "")
      return (str(sku.get("subject") or ""), str(sku.get("strategy") or ""), cat)

    def fallback_buyout_rate(nm_id: str, sku: dict[str, object]) -> float:
      for key in _sku_keys(nm_id, sku):
        bought, canceled = group_wb_buyout.get(key, (0.0, 0.0))
        if bought + canceled > 0:
          return bought / (bought + canceled)
      if total_wb_buyouts + total_wb_cancels > 0:
        return total_wb_buyouts / (total_wb_buyouts + total_wb_cancels)
      return 0.3

    def fallback_check(nm_id: str, sku: dict[str, object]) -> float:
      for key in _sku_keys(nm_id, sku):
        amount, qty = group_check.get(key, (0.0, 0.0))
        if qty > 0:
          return amount / qty
      return 0.0

    def fallback_elasticity(nm_id: str, sku: dict[str, object]) -> float:
      if nm_id in elasticity_by_nm:
        v = elasticity_by_nm[nm_id]
        return v if v < 0.0 else -1.0  # 0.0 = positive slope clipped, treat as no data
      for key in _sku_keys(nm_id, sku):
        values = group_elasticities.get(key, [])
        if values:
          avg = sum(values) / len(values)
          return avg if avg < 0.0 else -1.0
      return -1.0

    def fallback_ads_elasticity(nm_id: str, sku: dict[str, object]) -> float:
      stats = ads_stats_by_nm.get(nm_id, {})
      if nm_id in ads_elasticity_by_nm:
        return _effective_ads_elasticity(ads_elasticity_by_nm[nm_id], stats)
      for key in _sku_keys(nm_id, sku):
        values = group_ads_elasticities.get(key, [])
        if values:
          return _effective_ads_elasticity(sum(values) / len(values), stats, group_fallback=True)
      return _effective_ads_elasticity(None, stats, group_fallback=True)

    rows_out: list[dict[str, object]] = []
    summary = {
      "stock": 0.0,
      "stock_cost": 0.0,
      "stock_value": 0.0,
      "forecast_orders": 0.0,
      "forecast_buyouts": 0.0,
      "orders_revenue": 0.0,
      "forecast_revenue": 0.0,
      "lost_orders": 0.0,
      "lost_revenue": 0.0,
    }
    day_count = 7.0
    for nm_id in nm_ids:
      sku = sku_by_nm[nm_id]
      metric = metrics_by_nm.get(nm_id)
      orders_qty = _to_float(metric["orders_qty"]) if metric else 0.0
      orders_sum = _to_float(metric["orders_sum"]) if metric else 0.0
      base_speed = orders_qty / day_count
      wb_buyouts, wb_cancels = wb_buyout_by_nm.get(nm_id, (0.0, 0.0))
      buyout_rate = wb_buyouts / (wb_buyouts + wb_cancels) if (wb_buyouts + wb_cancels) > 0 else fallback_buyout_rate(nm_id, sku)
      buyout_rate = max(0.0, min(1.0, buyout_rate))
      average_check = orders_sum / orders_qty if orders_qty > 0 else fallback_check(nm_id, sku)
      elasticity = fallback_elasticity(nm_id, sku)
      ads_elasticity = fallback_ads_elasticity(nm_id, sku)
      current_price = current_price_by_nm.get(nm_id, 0.0)
      current_spp = current_spp_by_nm.get(nm_id, 0.0)
      current_client_price = client_price_by_nm.get(nm_id, 0.0)
      baseline_client_price = baseline_client_check_by_nm.get(nm_id, current_client_price)
      # speed corrections use client price — that's what drives demand
      if baseline_client_price > 0 and current_client_price > 0:
        current_price_delta = (current_client_price / baseline_client_price - 1.0)
        current_factor = max(0.0, 1.0 + elasticity * current_price_delta)
      else:
        current_factor = 1.0
      model_current_speed = base_speed * current_factor
      current_speed_day = baseline_to
      yesterday_speed = _order_speed_for_day(conn, nm_id, current_speed_day)
      if yesterday_speed is None:
        current_speed_day = (date.fromisoformat(baseline_to) - timedelta(days=1)).isoformat()
        yesterday_speed = _order_speed_for_day(conn, nm_id, current_speed_day)
      current_speed = yesterday_speed if yesterday_speed is not None else model_current_speed
      current_speed_source = "day" if yesterday_speed is not None else "price_model"
      # planned: user sets % change from current seller price → derive planned client price
      planned_seller_price = current_price * (1.0 + price_change_pct / 100.0) if current_price > 0 else 0.0
      planned_client_price = planned_seller_price * (1.0 - current_spp / 100.0) if planned_seller_price > 0 else 0.0
      planned_anchor_speed = current_speed if current_speed > 0 else base_speed
      planned_anchor_client = current_client_price if current_client_price > 0 else baseline_client_price
      if planned_anchor_client > 0 and planned_client_price > 0:
        planned_price_delta = (planned_client_price / planned_anchor_client - 1.0)
        planned_factor = max(0.0, 1.0 + elasticity * planned_price_delta)
      else:
        planned_factor = 1.0
      planned_speed = planned_anchor_speed * planned_factor
      stock = stock_by_nm.get(nm_id, 0.0)
      warehouse_stock = _to_float(sku.get("warehouse_stock") or 0)
      potential_orders = planned_speed * forecast_days
      forecast_orders, stockout_date = _simulate_planning_orders(
        stock,
        planned_speed,
        buyout_rate,
        forecast_days,
        return_delay_days,
        forecast_start,
        warehouse_stock=warehouse_stock,
        transit_days=transit_days,
      )
      effective_check = planned_client_price if planned_client_price > 0 else average_check
      forecast_buyouts = forecast_orders * buyout_rate
      forecast_revenue = forecast_buyouts * effective_check
      lost_orders = max(0.0, potential_orders - forecast_orders)
      cogs = _to_float(sku.get("cogs") or 0)
      # ДРР = реклама / фактические выкупы (buyoutSum из funnel_analytics за 14д)
      base_ads_per_day = base_ads_by_nm.get(nm_id, 0.0) / ads_day_count
      actual_buyout_revenue_14d = ads_revenue_by_nm.get(nm_id, 0.0)
      base_revenue_per_day = (actual_buyout_revenue_14d / ads_day_count) if actual_buyout_revenue_14d > 0 else (effective_check * base_speed * buyout_rate)
      base_drr_pct = base_ads_per_day / base_revenue_per_day * 100.0 if base_revenue_per_day > 0 else 0.0
      drr_anchor_pct = max(6.0, min(12.0, base_drr_pct if base_drr_pct > 0 else 6.0))
      organic_floor = max(0.25, min(0.70, 0.55 - ads_elasticity * 0.35))
      # Воронка baseline
      fi = funnel_by_nm.get(nm_id, {})
      base_ctr = fi.get("ctr", 0.0)
      base_cr1 = fi.get("cr1", 0.0)
      base_cr2 = fi.get("cr2", 0.0)
      summary["stock"] += stock + warehouse_stock
      summary["stock_cost"] += (stock + warehouse_stock) * cogs
      summary["stock_value"] += (stock + warehouse_stock) * (current_client_price or average_check)
      summary["forecast_orders"] += forecast_orders
      summary["forecast_buyouts"] += forecast_buyouts
      summary["orders_revenue"] += forecast_orders * effective_check
      summary["forecast_revenue"] += forecast_revenue
      summary["lost_orders"] += lost_orders
      summary["lost_revenue"] += lost_orders * buyout_rate * effective_check
      rows_out.append({
        "nmid": nm_id,
        "article": str(marketplace_article_by_nm.get(nm_id) or sku.get("article") or nm_id),
        "name": str(sku.get("name") or ""),
        "subject": str(sku.get("subject") or ""),
        "strategy": str(sku.get("strategy") or ""),
        "revenue_category": abc_category_by_nm.get(nm_id, "C") if total_abc > 0 else str(sku.get("revenue_category") or ""),
        "cogs": cogs,
        "wb_commission_pct": _to_float(sku.get("wb_commission_pct") or 0),
        "stock": stock,
        "warehouse_stock": _to_float(sku.get("warehouse_stock") or 0),
        "baseline_seller_price": baseline_avg_price_by_nm.get(nm_id, 0.0),
        "baseline_spp": baseline_spp_by_nm.get(nm_id, 0.0),
        "baseline_client_check": baseline_client_check_by_nm.get(nm_id, 0.0),
        "base_speed": base_speed,
        "buyout_percent": buyout_rate * 100.0,
        "average_check": average_check,
        "current_price": current_price,
        "current_spp": current_spp,
        "client_price": current_client_price,
        "elasticity": elasticity,
        "ads_elasticity": ads_elasticity,
        "base_ads_per_day": base_ads_per_day,
        "base_revenue_per_day": base_revenue_per_day,
        "base_drr_pct": round(base_drr_pct, 2),
        "drr_anchor_pct": round(drr_anchor_pct, 2),
        "organic_floor": round(organic_floor, 3),
        "base_ctr": round(base_ctr, 2),
        "base_cr1": round(base_cr1, 2),
        "base_cr2": round(base_cr2, 2),
        "current_speed": current_speed,
        "model_current_speed": model_current_speed,
        "current_speed_source": current_speed_source,
        "current_speed_day": current_speed_day if yesterday_speed is not None else "",
        "planned_seller_price": planned_seller_price,
        "planned_price": planned_client_price,
        "planned_speed": planned_speed,
        "price_change_pct": price_change_pct,
        "potential_orders": potential_orders,
        "effective_check": effective_check,
        "forecast_orders": forecast_orders,
        "forecast_buyouts": forecast_buyouts,
        "forecast_revenue": forecast_revenue,
        "lost_orders": lost_orders,
        "lost_revenue": lost_orders * buyout_rate * effective_check,
        "stockout_date": stockout_date,
      })

  rows_out.sort(key=lambda row: (-_to_float(row.get("forecast_revenue")), str(row.get("article") or "")))
  return {
    "summary": summary,
    "rows": rows_out,
    "options": options,
    "baseline_from": baseline_from,
    "baseline_to": baseline_to,
    "forecast_from": forecast_start.isoformat(),
    "forecast_to": target_date.isoformat(),
    "forecast_days": forecast_days,
    "return_delay_days": return_delay_days,
    "transit_days": transit_days,
    "price_change_pct": price_change_pct,
  }

def _sql_ident(name: str) -> str:
  return '"' + name.replace('"', '""') + '"'

def _first_existing(columns: set[str], candidates: list[str]) -> str | None:
  for candidate in candidates:
    if candidate in columns:
      return candidate
  return None

def _table_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
  safe_table = table_name.replace('"', '""')
  rows = conn.execute(f'PRAGMA table_info("{safe_table}")').fetchall()
  return {str(row["name"]) for row in rows}

def _table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
  row = conn.execute(
    "SELECT 1 FROM sqlite_master WHERE type IN ('table', 'view') AND lower(name) = lower(?) LIMIT 1",
    (table_name,),
  ).fetchone()
  return row is not None

def _sum_expr(columns: set[str], candidates: list[str]) -> tuple[str, list[str]]:
  used = [candidate for candidate in candidates if candidate in columns]
  if not used:
    return "0", []
  expr = " + ".join(f"COALESCE(SUM({_sql_ident(col)}), 0)" for col in used)
  return expr, used

AD_BIDDER_TZ = ZoneInfo("Europe/Moscow")
AD_BIDDER_DAYS = ("mon", "tue", "wed", "thu", "fri", "sat", "sun")
AD_BIDDER_DAY_LABELS = {
  "mon": "Пн",
  "tue": "Вт",
  "wed": "Ср",
  "thu": "Чт",
  "fri": "Пт",
  "sat": "Сб",
  "sun": "Вс",
}
AD_API_BASE = "https://advert-api.wildberries.ru"
_ad_executor_started = False
_ad_executor_lock = threading.Lock()
_ad_refresh_jobs: set[str] = set()
_ad_refresh_jobs_lock = threading.Lock()

def _has_wb_ad_cabinet(cabinet: dict | None) -> bool:
  cab = cabinet or {}
  token = str(cab.get("wb_adv_token") or cab.get("wb_api_token") or "").strip()
  return cab.get("marketplace") in ("wb", "both") and bool(token)

def _ad_token(cabinet: dict | None) -> str:
  cab = cabinet or {}
  return str(cab.get("wb_adv_token") or cab.get("wb_api_token") or "").strip()

def _ad_headers(cabinet: dict | None) -> dict[str, str]:
  return {"Authorization": _ad_token(cabinet), "Accept": "application/json"}

def _ad_db_connect_for_cabinet(cabinet: dict | None) -> sqlite3.Connection:
  conn = sqlite3.connect(_cabinet_db_path(cabinet))
  conn.execute("PRAGMA journal_mode=WAL")
  conn.row_factory = sqlite3.Row
  return conn

def _ensure_ad_bidding_tables(conn: sqlite3.Connection) -> None:
  conn.executescript("""
    CREATE TABLE IF NOT EXISTS ad_bidding_rules (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      enabled INTEGER NOT NULL DEFAULT 1,
      name TEXT NOT NULL,
      nmid TEXT NOT NULL,
      article TEXT NOT NULL,
      timezone TEXT NOT NULL DEFAULT 'Europe/Moscow',
      schedule_json TEXT NOT NULL,
      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS ad_bidding_rule_campaigns (
      rule_id INTEGER NOT NULL,
      advert_id TEXT NOT NULL,
      PRIMARY KEY (rule_id, advert_id)
    );
    CREATE TABLE IF NOT EXISTS ad_bidding_actions (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      rule_id INTEGER,
      advert_id TEXT NOT NULL,
      desired_state TEXT NOT NULL,
      actual_action TEXT NOT NULL,
      status_code INTEGER,
      message TEXT,
      created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS ad_bidding_log (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      event_type TEXT NOT NULL,
      advert_id TEXT NOT NULL DEFAULT '',
      title TEXT NOT NULL DEFAULT '',
      details TEXT NOT NULL DEFAULT '',
      payload_json TEXT NOT NULL DEFAULT '{}',
      created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS ad_bidding_state (
      advert_id TEXT PRIMARY KEY,
      last_desired_state TEXT,
      last_success_at TEXT,
      last_error_at TEXT,
      last_error TEXT
    );
    CREATE TABLE IF NOT EXISTS ad_bidding_global_settings (
      key TEXT PRIMARY KEY,
      schedule_json TEXT NOT NULL,
      updated_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS ad_bidding_campaign_settings (
      advert_id TEXT PRIMARY KEY,
      enabled INTEGER NOT NULL DEFAULT 0,
      nmid TEXT NOT NULL DEFAULT '',
      article TEXT NOT NULL DEFAULT '',
      name TEXT NOT NULL DEFAULT '',
      daily_budget REAL NOT NULL DEFAULT 0,
      auto_pause INTEGER NOT NULL DEFAULT 0,
      today_spend REAL NOT NULL DEFAULT 0,
      spend_day TEXT NOT NULL DEFAULT '',
      last_spend_sync_at TEXT,
      last_seen_at TEXT,
      updated_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS ad_campaign_cache (
      advert_id TEXT PRIMARY KEY,
      name TEXT NOT NULL DEFAULT '',
      status TEXT NOT NULL DEFAULT '',
      type TEXT NOT NULL DEFAULT '',
      nmids_json TEXT NOT NULL DEFAULT '[]',
      nm_count INTEGER NOT NULL DEFAULT 0,
      updated_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS ad_bidding_executor_lock (
      key TEXT PRIMARY KEY,
      locked_until TEXT NOT NULL
    );
  """)
  existing_cols = set(_table_columns(conn, "ad_bidding_campaign_settings"))
  for col_name, col_sql in [
    ("pause_start", "TEXT NOT NULL DEFAULT '23:00'"),
    ("pause_end", "TEXT NOT NULL DEFAULT '08:00'"),
  ]:
    if col_name not in existing_cols:
      conn.execute(f"ALTER TABLE ad_bidding_campaign_settings ADD COLUMN {col_name} {col_sql}")
  conn.execute("CREATE INDEX IF NOT EXISTS idx_ad_bidding_rules_enabled ON ad_bidding_rules(enabled)")
  conn.execute("CREATE INDEX IF NOT EXISTS idx_ad_bidding_actions_created ON ad_bidding_actions(created_at)")
  conn.execute("CREATE INDEX IF NOT EXISTS idx_ad_bidding_log_created ON ad_bidding_log(created_at)")
  conn.execute("CREATE INDEX IF NOT EXISTS idx_ad_bidding_campaign_settings_enabled ON ad_bidding_campaign_settings(enabled)")
  conn.execute("CREATE INDEX IF NOT EXISTS idx_ad_campaign_cache_updated ON ad_campaign_cache(updated_at)")

def _ad_log(
  conn: sqlite3.Connection,
  event_type: str,
  advert_id: str = "",
  title: str = "",
  details: str = "",
  payload: object | None = None,
  created_at: str | None = None,
) -> None:
  conn.execute(
    "INSERT INTO ad_bidding_log(event_type, advert_id, title, details, payload_json, created_at) VALUES (?, ?, ?, ?, ?, ?)",
    (
      str(event_type or "event")[:40],
      str(advert_id or "")[:80],
      str(title or "")[:300],
      str(details or "")[:2000],
      json.dumps(payload if payload is not None else {}, ensure_ascii=False, default=str),
      created_at or datetime.now(AD_BIDDER_TZ).isoformat(timespec="seconds"),
    ),
  )

def _ad_money_label(value: object) -> str:
  return f"{_to_float(value):,.0f}".replace(",", " ") + " ₽"

def _ad_campaign_log_context(advert_id: str, campaign_name: str, article: str, daily_budget: object, today_spend: object) -> str:
  clean_name = str(campaign_name or f"Кампания {advert_id}").strip()
  clean_article = str(article or "").strip() or "—"
  return (
    f"Кампания: {clean_name} ({advert_id}); "
    f"артикул: {clean_article}; "
    f"бюджет: {_ad_money_label(daily_budget)}/день; "
    f"расход сегодня: {_ad_money_label(today_spend)}"
  )

def _normalize_ad_schedule(schedule: object) -> dict[str, dict[str, list[list[str]]]]:
  raw_weekly = {}
  if isinstance(schedule, dict):
    raw_weekly = schedule.get("weekly") if isinstance(schedule.get("weekly"), dict) else schedule
  weekly: dict[str, list[list[str]]] = {}
  for day in AD_BIDDER_DAYS:
    intervals = []
    for item in (raw_weekly.get(day) if isinstance(raw_weekly, dict) else []) or []:
      if not isinstance(item, (list, tuple)) or len(item) != 2:
        continue
      start = str(item[0]).strip()
      end = str(item[1]).strip()
      if _ad_valid_time(start) and _ad_valid_time(end) and start != end:
        intervals.append([start, end])
    weekly[day] = intervals
  return {"weekly": weekly}

def _ad_valid_time(value: str) -> bool:
  try:
    hh, mm = value.split(":", 1)
    return 0 <= int(hh) <= 23 and 0 <= int(mm) <= 59 and len(hh) == 2 and len(mm) == 2
  except Exception:
    return False

def _ad_minutes(value: str) -> int:
  hh, mm = value.split(":", 1)
  return int(hh) * 60 + int(mm)

def _ad_schedule_active(schedule: object, now: datetime | None = None) -> bool:
  now = now or datetime.now(AD_BIDDER_TZ)
  normalized = _normalize_ad_schedule(schedule)
  weekly = normalized["weekly"]
  day_idx = now.weekday()
  today_key = AD_BIDDER_DAYS[day_idx]
  prev_key = AD_BIDDER_DAYS[(day_idx - 1) % 7]
  minute = now.hour * 60 + now.minute
  for start, end in weekly.get(today_key, []):
    start_m = _ad_minutes(start)
    end_m = _ad_minutes(end)
    if start_m < end_m and start_m <= minute < end_m:
      return True
    if start_m > end_m and minute >= start_m:
      return True
  for start, end in weekly.get(prev_key, []):
    start_m = _ad_minutes(start)
    end_m = _ad_minutes(end)
    if start_m > end_m and minute < end_m:
      return True
  return False

def _ad_normalize_time(value: object, default: str) -> str:
  text = str(value or "").strip()
  if not text:
    return default
  if ":" not in text and text.isdigit():
    text = f"{int(text):02d}:00"
  if ":" in text:
    left, right = text.split(":", 1)
    try:
      text = f"{int(left):02d}:00"
    except ValueError:
      return default
  return text if _ad_valid_time(text) else default

def _ad_hour_label(value: object, default: str) -> str:
  normalized = _ad_normalize_time(value, default)
  return f"{int(normalized.split(':', 1)[0]):02d}"

def _ad_hour_interval_label(start: object, end: object) -> str:
  return f"{_ad_hour_label(start, '23:00')}-{_ad_hour_label(end, '08:00')}"

def _ad_parse_pause_interval(value: object, start_default: str = "23:00", end_default: str = "08:00") -> tuple[str, str]:
  text = str(value or "").strip()
  if "-" not in text:
    return (_ad_normalize_time(start_default, "23:00"), _ad_normalize_time(end_default, "08:00"))
  start, end = text.split("-", 1)
  return (_ad_normalize_time(start, start_default), _ad_normalize_time(end, end_default))

def _ad_interval_active(start: object, end: object, now: datetime | None = None) -> bool:
  now = now or datetime.now(AD_BIDDER_TZ)
  start_s = _ad_normalize_time(start, "23:00")
  end_s = _ad_normalize_time(end, "08:00")
  if start_s == end_s:
    return False
  minute = now.hour * 60 + now.minute
  start_m = _ad_minutes(start_s)
  end_m = _ad_minutes(end_s)
  if start_m < end_m:
    return start_m <= minute < end_m
  return minute >= start_m or minute < end_m

def _ad_empty_schedule() -> dict[str, dict[str, list[list[str]]]]:
  return {"weekly": {day: [] for day in AD_BIDDER_DAYS}}

def _ad_default_schedule() -> dict[str, dict[str, list[list[str]]]]:
  return {"weekly": {day: [["09:00", "22:00"]] for day in AD_BIDDER_DAYS}}

def _fetch_ad_global_settings() -> dict[str, object]:
  with _db_connect() as conn:
    _ensure_ad_bidding_tables(conn)
    row = conn.execute("SELECT schedule_json, updated_at FROM ad_bidding_global_settings WHERE key='main'").fetchone()
  schedule = json.loads(row["schedule_json"]) if row else _ad_default_schedule()
  normalized = _normalize_ad_schedule(schedule)
  return {
    "schedule": normalized,
    "updated_at": row["updated_at"] if row else "",
    "desired_now": "active" if _ad_schedule_active(normalized) else "paused",
  }

def _save_ad_global_schedule(payload: dict[str, object]) -> dict[str, object]:
  schedule = _normalize_ad_schedule(payload.get("schedule") or {})
  now = datetime.now(AD_BIDDER_TZ).isoformat(timespec="seconds")
  with _db_connect() as conn:
    _ensure_ad_bidding_tables(conn)
    conn.execute(
      "INSERT INTO ad_bidding_global_settings(key, schedule_json, updated_at) VALUES ('main', ?, ?) "
      "ON CONFLICT(key) DO UPDATE SET schedule_json=excluded.schedule_json, updated_at=excluded.updated_at",
      (json.dumps(schedule, ensure_ascii=False), now),
    )
    conn.commit()
  return {"ok": True, "schedule": schedule, "desired_now": "active" if _ad_schedule_active(schedule) else "paused"}

def _ad_api_get(cabinet: dict | None, path: str, params: dict[str, str] | None = None, *, retries: int = 3) -> requests.Response:
  last_response = None
  last_exc: Exception | None = None
  for attempt in range(max(1, retries)):
    try:
      response = requests.get(f"{AD_API_BASE}{path}", headers=_ad_headers(cabinet), params=params or {}, timeout=20)
    except requests.RequestException as exc:
      last_exc = exc
      time.sleep(min(2 + attempt * 2, 8))
      continue
    last_response = response
    if response.status_code not in {429, 500, 502, 503, 504}:
      return response
    time.sleep(min(2 + attempt * 2, 8))
  if last_response is None and last_exc is not None:
    raise last_exc
  return last_response  # type: ignore[return-value]

def _ad_fetch_campaign_ids(cabinet: dict | None) -> list[int]:
  response = _ad_api_get(cabinet, "/adv/v1/promotion/count", {})
  response.raise_for_status()
  data = response.json()
  result: list[int] = []
  for group in data.get("adverts", []) or []:
    if group.get("status") not in {4, 9, 11}:
      continue
    for advert in group.get("advert_list", []) or []:
      advert_id = advert.get("advertId")
      if isinstance(advert_id, int):
        result.append(advert_id)
  return sorted(set(result))

def _ad_fetch_campaign_details(cabinet: dict | None, advert_ids: list[int]) -> list[dict[str, object]]:
  if not advert_ids:
    return []
  rows: list[dict[str, object]] = []
  for idx in range(0, len(advert_ids), 50):
    chunk = advert_ids[idx:idx + 50]
    response = _ad_api_get(cabinet, "/api/advert/v2/adverts", {"ids": ",".join(str(x) for x in chunk)})
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, dict):
      rows.extend(payload.get("adverts", []) or [])
  return rows

def _ad_type_label(value: object) -> str:
  text = str(value or "").strip()
  if not text:
    return ""
  if "·" in text:
    return text
  lowered = text.lower()
  if lowered in {"cpm", "cpc"}:
    return lowered.upper()
  labels = {
    "4": "Каталог",
    "5": "Карточка",
    "6": "Поиск",
    "7": "Рекомендации",
    "8": "Авто",
    "9": "Аукцион",
  }
  return labels.get(text, text)

def _ad_extract_nmids(advert: dict[str, object]) -> list[str]:
  nmids: set[str] = set()
  for key in ("nm_settings", "nms", "nmIds"):
    value = advert.get(key)
    if isinstance(value, list):
      for item in value:
        if isinstance(item, dict):
          nm_id = item.get("nm_id") or item.get("nmId") or item.get("nm")
        else:
          nm_id = item
        if nm_id is not None and str(nm_id).strip().isdigit():
          nmids.add(str(nm_id).strip())
  return sorted(nmids, key=lambda item: int(item))

def _ad_upsert_campaign_cache(conn: sqlite3.Connection, adverts: list[dict[str, object]]) -> None:
  now = datetime.now(AD_BIDDER_TZ).isoformat(timespec="seconds")
  for advert in adverts:
    advert_id = str(advert.get("id") or advert.get("advertId") or "").strip()
    if not advert_id.isdigit():
      continue
    settings = advert.get("settings") if isinstance(advert.get("settings"), dict) else {}
    payment_type = str((settings or {}).get("payment_type") or advert.get("payment_type") or advert.get("type") or "").strip()
    bid_type = str(advert.get("bid_type") or "").strip().lower()
    bid_label = {"unified": "Единая", "manual": "Ручная"}.get(bid_type, bid_type.title() if bid_type else "")
    type_parts = [part for part in [payment_type.upper() if payment_type.lower() in {"cpm", "cpc"} else payment_type, bid_label] if part]
    nmids = _ad_extract_nmids(advert)
    conn.execute(
      """
      INSERT INTO ad_campaign_cache(advert_id, name, status, type, nmids_json, nm_count, updated_at)
      VALUES (?, ?, ?, ?, ?, ?, ?)
      ON CONFLICT(advert_id) DO UPDATE SET
        name=excluded.name,
        status=excluded.status,
        type=excluded.type,
        nmids_json=excluded.nmids_json,
        nm_count=excluded.nm_count,
        updated_at=excluded.updated_at
      """,
      (
        advert_id,
        str(advert.get("name") or (settings or {}).get("name") or ""),
        str(advert.get("status") or ""),
        " · ".join(type_parts),
        json.dumps(nmids, ensure_ascii=False),
        len(nmids),
        now,
      ),
    )

def _ad_cached_campaign_details(conn: sqlite3.Connection, advert_ids: list[str]) -> dict[str, dict[str, object]]:
  if not advert_ids:
    return {}
  placeholders = ",".join("?" for _ in advert_ids)
  rows = conn.execute(
    f"SELECT * FROM ad_campaign_cache WHERE advert_id IN ({placeholders})",
    advert_ids,
  ).fetchall()
  return {str(row["advert_id"]): dict(row) for row in rows}

def _ad_cached_campaign_ids_for_nmid(conn: sqlite3.Connection, nmid: str) -> list[str]:
  clean_nmid = str(nmid or "").strip()
  if not clean_nmid:
    return []
  rows = conn.execute("SELECT advert_id, status, nmids_json FROM ad_campaign_cache").fetchall()
  result: list[str] = []
  for row in rows:
    if str(row["status"] or "").strip() not in {"4", "9", "11"}:
      continue
    try:
      nmids = json.loads(row["nmids_json"] or "[]")
    except Exception:
      nmids = []
    if clean_nmid in {str(item).strip() for item in nmids}:
      result.append(str(row["advert_id"]))
  return sorted({item for item in result if item.isdigit()}, key=lambda item: int(item))

def _ad_campaign_action(cabinet: dict | None, advert_id: str, desired_state: str) -> tuple[str, int | None, str]:
  path = "/adv/v0/start" if desired_state == "active" else "/adv/v0/pause"
  action = "start" if desired_state == "active" else "pause"
  try:
    response = _ad_api_get(cabinet, path, {"id": str(advert_id)})
  except requests.RequestException as exc:
    return "error", None, str(exc)[:500]
  status = response.status_code if response is not None else None
  text = response.text[:500] if response is not None else ""
  if status and 200 <= status < 300:
    return action, status, "ok"
  if status == 422:
    return "noop", status, text or "status not changed"
  return "error", status, text or "request failed"

def _ad_campaign_settings_by_ids(conn: sqlite3.Connection, advert_ids: list[str]) -> dict[str, dict[str, object]]:
  if not advert_ids:
    return {}
  placeholders = ",".join("?" for _ in advert_ids)
  rows = conn.execute(
    f"SELECT * FROM ad_bidding_campaign_settings WHERE advert_id IN ({placeholders})",
    advert_ids,
  ).fetchall()
  return {str(row["advert_id"]): dict(row) for row in rows}

def _ad_default_budget_by_nmid(nmid: str, date_to: str | None = None) -> dict[str, float]:
  clean_nmid = str(nmid or "").strip()
  if not clean_nmid:
    return {"daily_budget": 0.0, "avg_orders_7d": 0.0, "orders_sum_7d": 0.0}
  with _db_connect() as conn:
    if not _table_exists(conn, "buyout_order_day"):
      return {"daily_budget": 0.0, "avg_orders_7d": 0.0, "orders_sum_7d": 0.0}
    try:
      end_day = date.fromisoformat(str(date_to or "").strip()[:10]) if date_to else None
    except ValueError:
      end_day = None
    if end_day is None:
      latest = conn.execute(
        'SELECT MAX("Дата") AS day FROM buyout_order_day WHERE CAST("nmId" AS TEXT)=?',
        (clean_nmid,),
      ).fetchone()
      latest_day = str((latest or {})["day"] or "")[:10] if latest else ""
      try:
        end_day = date.fromisoformat(latest_day)
      except ValueError:
        return {"daily_budget": 0.0, "avg_orders_7d": 0.0, "orders_sum_7d": 0.0}
    start_day = end_day - timedelta(days=6)
    row = conn.execute(
      """
      SELECT SUM(CAST("Сумма заказов" AS REAL)) AS orders_sum
      FROM buyout_order_day
      WHERE CAST("nmId" AS TEXT)=? AND "Дата" >= ? AND "Дата" <= ?
      """,
      (clean_nmid, start_day.isoformat(), end_day.isoformat()),
    ).fetchone()
  orders_sum = _to_float(row["orders_sum"] if row else 0)
  avg_orders = orders_sum / 7.0 if orders_sum > 0 else 0.0
  return {
    "daily_budget": round(avg_orders * 0.02),
    "avg_orders_7d": avg_orders,
    "orders_sum_7d": orders_sum,
  }

def _ad_last_actions_by_advert(conn: sqlite3.Connection, advert_ids: list[str]) -> dict[str, dict[str, object]]:
  if not advert_ids:
    return {}
  placeholders = ",".join("?" for _ in advert_ids)
  rows = conn.execute(
    f"""
    SELECT a.*
    FROM ad_bidding_actions a
    JOIN (
      SELECT advert_id, MAX(id) AS max_id
      FROM ad_bidding_actions
      WHERE advert_id IN ({placeholders})
      GROUP BY advert_id
    ) x ON x.max_id = a.id
    """,
    advert_ids,
  ).fetchall()
  return {str(row["advert_id"]): dict(row) for row in rows}

def _save_ad_campaign_setting(payload: dict[str, object]) -> dict[str, object]:
  advert_id = str(payload.get("advert_id") or "").strip()
  if not advert_id.isdigit():
    raise ValueError("advert_id required")
  enabled = 1 if bool(payload.get("enabled") or payload.get("schedule_auto")) else 0
  auto_pause = 1 if bool(payload.get("auto_pause", False)) else 0
  try:
    daily_budget = max(0.0, float(str(payload.get("daily_budget") or "0").replace(",", ".")))
  except ValueError:
    daily_budget = 0.0
  pause_start, pause_end = _ad_parse_pause_interval(payload.get("pause_interval") or payload.get("pause") or "")
  nmid = str(payload.get("nmid") or "").strip()
  article = str(payload.get("article") or "").strip()
  name = str(payload.get("name") or "").strip()[:300]
  now = datetime.now(AD_BIDDER_TZ).isoformat(timespec="seconds")
  with _db_connect() as conn:
    _ensure_ad_bidding_tables(conn)
    existing = conn.execute("SELECT * FROM ad_bidding_campaign_settings WHERE advert_id=?", (advert_id,)).fetchone()
    conn.execute(
      """
      INSERT INTO ad_bidding_campaign_settings(
        advert_id, enabled, nmid, article, name, daily_budget, auto_pause,
        today_spend, spend_day, last_spend_sync_at, last_seen_at, updated_at,
        pause_start, pause_end
      ) VALUES (?, ?, ?, ?, ?, ?, ?, 0, '', NULL, ?, ?, ?, ?)
      ON CONFLICT(advert_id) DO UPDATE SET
        enabled=excluded.enabled,
        nmid=COALESCE(NULLIF(excluded.nmid, ''), ad_bidding_campaign_settings.nmid),
        article=COALESCE(NULLIF(excluded.article, ''), ad_bidding_campaign_settings.article),
        name=COALESCE(NULLIF(excluded.name, ''), ad_bidding_campaign_settings.name),
        daily_budget=excluded.daily_budget,
        auto_pause=excluded.auto_pause,
        pause_start=excluded.pause_start,
        pause_end=excluded.pause_end,
        last_seen_at=excluded.last_seen_at,
        updated_at=excluded.updated_at
      """,
      (advert_id, enabled, nmid, article, name, daily_budget, auto_pause, now, now, pause_start, pause_end),
    )
    row = conn.execute("SELECT * FROM ad_bidding_campaign_settings WHERE advert_id=?", (advert_id,)).fetchone()
    pause_interval = _ad_hour_interval_label(pause_start, pause_end)
    detail_parts = [
      f"Бюджет: {daily_budget:g} ₽/день",
      f"автовыключение: {'вкл' if auto_pause else 'выкл'}",
      f"пауза: {pause_interval}",
      f"авто-пауза по расписанию: {'вкл' if enabled else 'выкл'}",
    ]
    if article or nmid:
      detail_parts.append(f"артикул: {article or nmid}")
    if name:
      detail_parts.append(f"кампания: {name}")
    changed: list[str] = []
    if existing:
      checks = [
        ("бюджет", _to_float(existing["daily_budget"]), daily_budget),
        ("автовыключение", int(existing["auto_pause"] or 0), auto_pause),
        ("авто расписание", int(existing["enabled"] or 0), enabled),
        ("начало паузы", str(existing["pause_start"] or ""), pause_start),
        ("конец паузы", str(existing["pause_end"] or ""), pause_end),
      ]
      for label, before, after in checks:
        if before != after:
          changed.append(label)
    else:
      changed.append("новые настройки")
    _ad_log(
      conn,
      "manual",
      advert_id,
      "Сохранены настройки кампании",
      "; ".join(detail_parts) + (f"; изменено: {', '.join(changed)}" if changed else "; без изменений"),
      {
        "advert_id": advert_id,
        "nmid": nmid,
        "article": article,
        "name": name,
        "daily_budget": daily_budget,
        "auto_pause": bool(auto_pause),
        "schedule_auto": bool(enabled),
        "pause_interval": pause_interval,
        "changed": changed,
      },
      now,
    )
    conn.commit()
  result = dict(row) if row else {}
  result["ok"] = True
  result["was_enabled"] = bool(existing["enabled"]) if existing else False
  return result

def _fetch_ad_logs(limit: int = 200) -> dict[str, object]:
  clean_limit = max(1, min(500, int(limit or 200)))
  with _db_connect() as conn:
    _ensure_ad_bidding_tables(conn)
    rows = conn.execute(
      "SELECT id, event_type, advert_id, title, details, payload_json, created_at FROM ad_bidding_log ORDER BY id DESC LIMIT ?",
      (clean_limit,),
    ).fetchall()
    advert_ids = sorted({str(row["advert_id"] or "").strip() for row in rows if str(row["advert_id"] or "").strip().isdigit()}, key=lambda item: int(item))
    settings_by_id = _ad_campaign_settings_by_ids(conn, advert_ids)
    cache_by_id = _ad_cached_campaign_details(conn, advert_ids)
  logs = []
  today = datetime.now(AD_BIDDER_TZ).date().isoformat()
  for row in rows:
    item = dict(row)
    try:
      item["payload"] = json.loads(item.pop("payload_json") or "{}")
    except Exception:
      item["payload"] = {}
    advert_id = str(item.get("advert_id") or "").strip()
    details = str(item.get("details") or "")
    if advert_id and str(item.get("event_type") or "") == "auto" and "Кампания:" not in details:
      payload = item["payload"] if isinstance(item.get("payload"), dict) else {}
      settings = settings_by_id.get(advert_id) or {}
      cached = cache_by_id.get(advert_id) or {}
      campaign_name = payload.get("campaign_name") or settings.get("name") or cached.get("name") or f"Кампания {advert_id}"
      article = payload.get("article") or settings.get("article") or settings.get("nmid") or "—"
      budget = payload.get("daily_budget") if "daily_budget" in payload else settings.get("daily_budget")
      if "today_spend" in payload:
        spend = payload.get("today_spend")
      else:
        spend = settings.get("today_spend") if settings and str(settings.get("spend_day") or "") == today else 0.0
      item["details"] = _ad_campaign_log_context(advert_id, str(campaign_name or ""), str(article or ""), budget, spend) + "; " + details
    logs.append(item)
  return {"logs": logs}

def _ad_fetch_today_spend(cabinet: dict | None, advert_ids: list[str], day: str) -> dict[str, float]:
  clean_ids = sorted({str(value).strip() for value in advert_ids if str(value).strip().isdigit()})
  if not clean_ids:
    return {}
  spend = {advert_id: 0.0 for advert_id in clean_ids}
  for idx in range(0, len(clean_ids), 50):
    chunk = clean_ids[idx:idx + 50]
    response = _ad_api_get(
      cabinet,
      "/adv/v3/fullstats",
      {"ids": ",".join(chunk), "beginDate": day, "endDate": day},
    )
    response.raise_for_status()
    payload = response.json()
    if not isinstance(payload, list):
      continue
    for campaign in payload:
      advert_id = str(campaign.get("advertId") or "").strip()
      if advert_id not in spend:
        continue
      total = 0.0
      for day_item in campaign.get("days", []) or []:
        for app in day_item.get("apps", []) or []:
          for item in app.get("nms", []) or []:
            total += _to_float(item.get("sum"))
      spend[advert_id] = total
  return spend

def _ad_metrics_from_fullstats(cabinet: dict | None, nmid: str, advert_ids: list[str], date_from: str, date_to: str) -> dict[str, dict[str, float]]:
  clean_nmid = str(nmid or "").strip()
  clean_ids = sorted({str(value).strip() for value in advert_ids if str(value).strip().isdigit()}, key=lambda item: int(item))
  if not clean_nmid or not clean_ids:
    return {}
  metrics: dict[str, dict[str, float]] = {advert_id: _empty_ad_metrics() for advert_id in clean_ids}
  for idx in range(0, len(clean_ids), 50):
    chunk = clean_ids[idx:idx + 50]
    response = _ad_api_get(
      cabinet,
      "/adv/v3/fullstats",
      {"ids": ",".join(chunk), "beginDate": date_from, "endDate": date_to},
    )
    response.raise_for_status()
    payload = response.json()
    if not isinstance(payload, list):
      continue
    for campaign in payload:
      advert_id = str(campaign.get("advertId") or "").strip()
      if advert_id not in metrics:
        continue
      target = metrics[advert_id]
      for day_item in campaign.get("days", []) or []:
        for app in day_item.get("apps", []) or []:
          for item in app.get("nms", []) or []:
            if str(item.get("nmId") or item.get("nm_id") or "").strip() != clean_nmid:
              continue
            target["views"] += _to_float(item.get("views"))
            target["clicks"] += _to_float(item.get("clicks"))
            target["spend"] += _to_float(item.get("sum"))
            target["orders"] += _to_float(item.get("orders"))
            target["revenue"] += _to_float(item.get("sum_price") or item.get("sumPrice"))
  result: dict[str, dict[str, float]] = {}
  for advert_id, row in metrics.items():
    views = _to_float(row.get("views"))
    clicks = _to_float(row.get("clicks"))
    spend = _to_float(row.get("spend"))
    orders_qty = _to_float(row.get("orders"))
    revenue = _to_float(row.get("revenue"))
    if not any([views, clicks, spend, orders_qty, revenue]):
      continue
    result[advert_id] = {
      "views": views,
      "clicks": clicks,
      "spend": spend,
      "orders": orders_qty,
      "revenue": revenue,
      "ctr": clicks / views * 100.0 if views else 0.0,
      "cpc": spend / clicks if clicks else 0.0,
      "cpm": spend / views * 1000.0 if views else 0.0,
      "cpo": spend / orders_qty if orders_qty else 0.0,
      "drr": spend / revenue * 100.0 if revenue else 0.0,
    }
  return result

def _ad_write_today_spend_for_campaigns(
  conn: sqlite3.Connection,
  advert_ids: list[str],
  spend_by_id: dict[str, float],
  day: str,
  *,
  nmid: str = "",
  article: str = "",
  names_by_id: dict[str, str] | None = None,
) -> int:
  clean_ids = sorted({str(value).strip() for value in advert_ids if str(value).strip().isdigit()}, key=lambda item: int(item))
  if not clean_ids:
    return 0
  now_iso = datetime.now(AD_BIDDER_TZ).isoformat(timespec="seconds")
  names_by_id = names_by_id or {}
  for advert_id in clean_ids:
    conn.execute(
      """
      INSERT INTO ad_bidding_campaign_settings(
        advert_id, enabled, nmid, article, name, daily_budget, auto_pause,
        today_spend, spend_day, last_spend_sync_at, last_seen_at, updated_at,
        pause_start, pause_end
      ) VALUES (?, 0, ?, ?, ?, 0, 0, ?, ?, ?, ?, ?, '23:00', '08:00')
      ON CONFLICT(advert_id) DO UPDATE SET
        nmid=COALESCE(NULLIF(excluded.nmid, ''), ad_bidding_campaign_settings.nmid),
        article=COALESCE(NULLIF(excluded.article, ''), ad_bidding_campaign_settings.article),
        name=COALESCE(NULLIF(excluded.name, ''), ad_bidding_campaign_settings.name),
        today_spend=excluded.today_spend,
        spend_day=excluded.spend_day,
        last_spend_sync_at=excluded.last_spend_sync_at,
        last_seen_at=excluded.last_seen_at,
        updated_at=excluded.updated_at
      """,
      (
        advert_id,
        str(nmid or ""),
        str(article or ""),
        str(names_by_id.get(advert_id) or "")[:300],
        _to_float(spend_by_id.get(advert_id)),
        day,
        now_iso,
        now_iso,
        now_iso,
      ),
    )
  return len(clean_ids)

def _ad_refresh_campaigns_job(cabinet: dict | None, nmid: str, date_from: str | None, date_to: str | None, *, refresh_details: bool = False) -> None:
  clean_nmid = str(nmid or "").strip()
  cabinet_id = str((cabinet or {}).get("cabinet_id") or "")
  job_key = f"{cabinet_id}:{clean_nmid}:{date_from or ''}:{date_to or ''}:{int(refresh_details)}"
  try:
    with _ad_db_connect_for_cabinet(cabinet) as conn:
      _ensure_ad_bidding_tables(conn)
      started_at = datetime.now(AD_BIDDER_TZ).isoformat(timespec="seconds")
      _ad_log(
        conn,
        "poll",
        "",
        "Фоновое обновление WB запущено",
        f"Артикул WB: {clean_nmid}; период: {date_from or '—'}..{date_to or '—'}",
        {"nmid": clean_nmid, "date_from": date_from, "date_to": date_to, "refresh_details": refresh_details},
        started_at,
      )
      conn.commit()
      ids = _ad_fetch_campaign_ids(cabinet)
      details = _ad_fetch_campaign_details(cabinet, ids)
      _ad_upsert_campaign_cache(conn, details)
      cached_ids = _ad_cached_campaign_ids_for_nmid(conn, clean_nmid)
      cache_by_id = _ad_cached_campaign_details(conn, cached_ids)
      names_by_id = {advert_id: str((cache_by_id.get(advert_id) or {}).get("name") or "") for advert_id in cached_ids}
      metrics_count = 0
      today_spend_updated = 0
      today_spend_total = 0.0
      period_metrics: dict[str, dict[str, float]] = {}
      if cached_ids and date_from and date_to:
        from_day = str(date_from)[:10]
        to_day = str(date_to)[:10]
        period_metrics = _ad_metrics_from_fullstats(cabinet, clean_nmid, cached_ids, from_day, to_day)
        metrics_count = len(period_metrics)
        today = datetime.now(AD_BIDDER_TZ).date().isoformat()
        if from_day <= today <= to_day:
          today_metrics = period_metrics if from_day == today and to_day == today else _ad_metrics_from_fullstats(cabinet, clean_nmid, cached_ids, today, today)
          today_spend = {advert_id: _to_float((today_metrics.get(advert_id) or {}).get("spend")) for advert_id in cached_ids}
          today_spend_updated = _ad_write_today_spend_for_campaigns(
            conn,
            cached_ids,
            today_spend,
            today,
            nmid=clean_nmid,
            names_by_id=names_by_id,
          )
          today_spend_total = sum(today_spend.values())
      _ad_log(
        conn,
        "poll",
        "",
        "Фоновое обновление WB завершено",
        f"Получено РК из WB: {len(ids)}; РК товара: {len(cached_ids)}; метрики WB по периоду: {metrics_count}"
        + (f"; расход сегодня записан: {today_spend_updated} РК, {today_spend_total:g} ₽" if today_spend_updated else ""),
        {
          "nmid": clean_nmid,
          "date_from": date_from,
          "date_to": date_to,
          "campaign_ids": cached_ids,
          "metrics_count": metrics_count,
          "today_spend_updated": today_spend_updated,
          "today_spend_total": today_spend_total,
        },
      )
      conn.commit()
  except Exception as exc:
    try:
      with _ad_db_connect_for_cabinet(cabinet) as conn:
        _ensure_ad_bidding_tables(conn)
        _ad_log(
          conn,
          "error",
          "",
          "Ошибка фонового обновления WB",
          str(exc),
          {"nmid": clean_nmid, "date_from": date_from, "date_to": date_to, "refresh_details": refresh_details},
        )
        conn.commit()
    except Exception:
      pass
  finally:
    with _ad_refresh_jobs_lock:
      _ad_refresh_jobs.discard(job_key)

def _ad_start_campaign_refresh(cabinet: dict | None, nmid: str, date_from: str | None, date_to: str | None, *, refresh_details: bool = False) -> bool:
  if not _has_wb_ad_cabinet(cabinet):
    return False
  clean_nmid = str(nmid or "").strip()
  if not clean_nmid:
    return False
  cabinet_id = str((cabinet or {}).get("cabinet_id") or "")
  job_key = f"{cabinet_id}:{clean_nmid}:{date_from or ''}:{date_to or ''}:{int(refresh_details)}"
  with _ad_refresh_jobs_lock:
    if job_key in _ad_refresh_jobs:
      return False
    _ad_refresh_jobs.add(job_key)
  thread = threading.Thread(
    target=_ad_refresh_campaigns_job,
    args=(cabinet, clean_nmid, date_from, date_to),
    kwargs={"refresh_details": refresh_details},
    name=f"wb-ad-refresh-{clean_nmid}",
    daemon=True,
  )
  thread.start()
  return True

def _fetch_ad_articles(query: str = "") -> dict[str, object]:
  query = query.strip().lower()
  with _db_connect() as conn:
    if not _table_exists(conn, "SKU"):
      return {"articles": []}
    cols = _table_columns(conn, "SKU")
    nmid_col = _first_existing(cols, ["Артикул WB", "nmId", "nm_id"])
    article_col = _first_existing(cols, ["Артикул поставщика", "supplierArticle", "НАШ", "Артикул"])
    name_col = _first_existing(cols, ["Название", "Наименование", "name"])
    subject_col = _first_existing(cols, ["Предмет", "subject"])
    if not nmid_col:
      return {"articles": []}
    select_parts = [
      f"TRIM({_sql_ident(nmid_col)}) AS nmid",
      f"COALESCE(NULLIF(TRIM({_sql_ident(article_col)}), ''), TRIM({_sql_ident(nmid_col)})) AS article" if article_col else f"TRIM({_sql_ident(nmid_col)}) AS article",
      f"COALESCE({_sql_ident(name_col)}, '') AS name" if name_col else "'' AS name",
      f"COALESCE({_sql_ident(subject_col)}, '') AS subject" if subject_col else "'' AS subject",
    ]
    where = [f"TRIM(COALESCE({_sql_ident(nmid_col)}, '')) != ''"]
    params: list[str] = []
    if query:
      hay = []
      for col in [nmid_col, article_col, name_col]:
        if col:
          hay.append(f"LOWER(COALESCE({_sql_ident(col)}, '')) LIKE ?")
          params.append(f"%{query}%")
      if hay:
        where.append("(" + " OR ".join(hay) + ")")
    rows = conn.execute(
      "SELECT " + ", ".join(select_parts) + " FROM SKU WHERE " + " AND ".join(where) + " LIMIT 50",
      params,
    ).fetchall()
  return {"articles": [dict(row) for row in rows]}

def _ad_metrics_by_campaign(nmid: str, days: int, date_from: str | None = None, date_to: str | None = None) -> dict[str, dict[str, float]]:
  with _db_connect() as conn:
    if not _table_exists(conn, "raw_ads"):
      return {}
    if date_from and date_to:
      try:
        _, start_day, latest_day = _date_range_limited(date_from, date_to, max_days=93)
      except ValueError:
        start_day = ""
        latest_day = ""
    else:
      latest = conn.execute('SELECT MAX("date") AS day FROM raw_ads').fetchone()
      latest_day = str((latest or {})["day"] or "")[:10] if latest else ""
      start_day = (date.fromisoformat(latest_day) - timedelta(days=max(1, days) - 1)).isoformat() if latest_day else ""
    if not latest_day or not start_day:
      return {}
    rows = conn.execute(
      """
      SELECT "advertId" AS advert_id,
             SUM(CAST(COALESCE("views", "0") AS REAL)) AS views,
             SUM(CAST(COALESCE("clicks", "0") AS REAL)) AS clicks,
             SUM(CAST(COALESCE("sum", "0") AS REAL)) AS spend,
             SUM(CAST(COALESCE("orders", "0") AS REAL)) AS orders_qty,
             SUM(CAST(COALESCE("sum_price", "0") AS REAL)) AS revenue
      FROM raw_ads
      WHERE "nmId" = ? AND substr("date", 1, 10) >= ? AND substr("date", 1, 10) <= ?
      GROUP BY "advertId"
      """,
      (nmid, start_day, latest_day),
    ).fetchall()
  metrics: dict[str, dict[str, float]] = {}
  for row in rows:
    views = _to_float(row["views"])
    clicks = _to_float(row["clicks"])
    spend = _to_float(row["spend"])
    orders_qty = _to_float(row["orders_qty"])
    revenue = _to_float(row["revenue"])
    metrics[str(row["advert_id"])] = {
      "views": views,
      "clicks": clicks,
      "spend": spend,
      "orders": orders_qty,
      "revenue": revenue,
      "ctr": clicks / views * 100.0 if views else 0.0,
      "cpc": spend / clicks if clicks else 0.0,
      "cpm": spend / views * 1000.0 if views else 0.0,
      "cpo": spend / orders_qty if orders_qty else 0.0,
      "drr": spend / revenue * 100.0 if revenue else 0.0,
    }
  return metrics

def _fetch_ad_campaigns(nmid: str, days: int, refresh_details: bool = False, date_from: str | None = None, date_to: str | None = None, show_all: bool = False) -> dict[str, object]:
  nmid = str(nmid or "").strip()
  days = days if days in {7, 14, 30} else 7
  if not nmid:
    return {"campaigns": []}
  cabinet = _get_request_cabinet()
  metrics = _ad_metrics_by_campaign(nmid, days, date_from=date_from, date_to=date_to)
  default_budget_info = _ad_default_budget_by_nmid(nmid, date_to=date_to)
  default_daily_budget = _to_float(default_budget_info.get("daily_budget"))
  campaigns: list[dict[str, object]] = []
  api_error = ""
  refresh_started = False
  advert_ids: list[str] = []
  cache_by_id: dict[str, dict[str, object]] = {}
  with _db_connect() as conn:
    _ensure_ad_bidding_tables(conn)
    advert_ids = _ad_cached_campaign_ids_for_nmid(conn, nmid)
    advert_ids = sorted(advert_ids, key=lambda item: -_to_float(metrics.get(item, {}).get("spend")))
    cache_by_id = _ad_cached_campaign_details(conn, advert_ids)
  if refresh_details or not cache_by_id:
    refresh_started = _ad_start_campaign_refresh(cabinet, nmid, date_from, date_to, refresh_details=refresh_details)
  advert_ids = sorted(advert_ids, key=lambda item: -_to_float(metrics.get(item, {}).get("spend")))
  for advert_id in advert_ids:
    metric = metrics.get(advert_id, _empty_ad_metrics())
    cached = cache_by_id.get(advert_id) or {}
    campaigns.append({
      "advert_id": advert_id,
      "name": cached.get("name") or f"Кампания {advert_id}",
      "status": cached.get("status") or "",
      "type": _ad_type_label(cached.get("type")),
      "type_raw": cached.get("type") or "",
      "nm_count": int(cached.get("nm_count") or 0),
      "source": "db",
      "metrics": metric,
    })
  if advert_ids:
    with _db_connect() as conn:
      _ensure_ad_bidding_tables(conn)
      settings_by_id = _ad_campaign_settings_by_ids(conn, advert_ids)
      actions_by_id = _ad_last_actions_by_advert(conn, advert_ids)
    today = datetime.now(AD_BIDDER_TZ).date().isoformat()
    for campaign in campaigns:
      advert_id = str(campaign.get("advert_id") or "")
      settings = settings_by_id.get(advert_id) or {}
      metric = campaign.get("metrics") if isinstance(campaign.get("metrics"), dict) else {}
      saved_budget = _to_float(settings.get("daily_budget")) if settings else 0.0
      has_spend = _to_float(metric.get("spend") if isinstance(metric, dict) else 0) > 0
      schedule_auto = bool(settings.get("enabled")) if settings else False
      auto_pause = bool(settings.get("auto_pause")) if settings else False
      use_default_budget = saved_budget <= 0 and has_spend and not schedule_auto and not auto_pause and default_daily_budget > 0
      daily_budget = default_daily_budget if use_default_budget else saved_budget
      campaign["settings"] = {
        "enabled": schedule_auto,
        "daily_budget": daily_budget,
        "daily_budget_default": default_daily_budget,
        "daily_budget_is_default": use_default_budget,
        "avg_orders_7d": _to_float(default_budget_info.get("avg_orders_7d")),
        "orders_sum_7d": _to_float(default_budget_info.get("orders_sum_7d")),
        "auto_pause": auto_pause,
        "schedule_auto": schedule_auto,
        "pause_start": settings.get("pause_start") or "23:00",
        "pause_end": settings.get("pause_end") or "08:00",
        "pause_interval": _ad_hour_interval_label(settings.get("pause_start") or "23:00", settings.get("pause_end") or "08:00"),
        "today_spend": _to_float(settings.get("today_spend")) if settings and str(settings.get("spend_day") or "") == today else 0.0,
        "spend_day": settings.get("spend_day") or "",
        "last_spend_sync_at": settings.get("last_spend_sync_at") or "",
      }
      last_action = actions_by_id.get(advert_id)
      if last_action and int(last_action.get("status_code") or 0) in range(200, 300):
        if str(last_action.get("actual_action") or "") == "pause":
          campaign["status"] = "11"
        elif str(last_action.get("actual_action") or "") == "start":
          campaign["status"] = "9"
      campaign["last_action"] = last_action
    campaigns.sort(key=lambda row: (
      -_to_float((row.get("metrics") or {}).get("spend") if isinstance(row.get("metrics"), dict) else 0),
      str(row.get("name") or ""),
    ))
  return {"campaigns": campaigns, "api_error": api_error, "refresh_started": refresh_started, "show_all": show_all, "days": days, "date_from": date_from or "", "date_to": date_to or ""}

def _empty_ad_metrics() -> dict[str, float]:
  return {"views": 0.0, "clicks": 0.0, "spend": 0.0, "orders": 0.0, "revenue": 0.0, "ctr": 0.0, "cpc": 0.0, "cpm": 0.0, "cpo": 0.0, "drr": 0.0}

def _fetch_ad_rules() -> dict[str, object]:
  with _db_connect() as conn:
    _ensure_ad_bidding_tables(conn)
    rows = conn.execute(
      """
      SELECT r.*, GROUP_CONCAT(rc.advert_id) AS advert_ids
      FROM ad_bidding_rules r
      LEFT JOIN ad_bidding_rule_campaigns rc ON rc.rule_id = r.id
      GROUP BY r.id
      ORDER BY r.updated_at DESC, r.id DESC
      """
    ).fetchall()
    actions = conn.execute(
      """
      SELECT a.*
      FROM ad_bidding_actions a
      JOIN (
        SELECT advert_id, MAX(id) AS max_id FROM ad_bidding_actions GROUP BY advert_id
      ) x ON x.max_id = a.id
      """
    ).fetchall()
  last_by_advert = {str(row["advert_id"]): dict(row) for row in actions}
  rules = []
  for row in rows:
    advert_ids = [x for x in str(row["advert_ids"] or "").split(",") if x]
    last = None
    for advert_id in advert_ids:
      item = last_by_advert.get(advert_id)
      if item and (last is None or int(item["id"]) > int(last["id"])):
        last = item
    rules.append({
      "id": row["id"],
      "enabled": bool(row["enabled"]),
      "name": row["name"],
      "nmid": row["nmid"],
      "article": row["article"],
      "schedule": json.loads(row["schedule_json"] or "{}"),
      "advert_ids": advert_ids,
      "updated_at": row["updated_at"],
      "last_action": last,
      "desired_now": "active" if _ad_schedule_active(json.loads(row["schedule_json"] or "{}")) else "paused",
    })
  return {"rules": rules}

def _save_ad_rule(payload: dict[str, object]) -> dict[str, object]:
  rule_id = int(payload.get("id") or 0)
  nmid = str(payload.get("nmid") or "").strip()
  article = str(payload.get("article") or nmid).strip()
  name = str(payload.get("name") or article or nmid).strip()
  advert_ids = [str(x).strip() for x in (payload.get("advert_ids") or []) if str(x).strip().isdigit()]
  schedule = _normalize_ad_schedule(payload.get("schedule") or {})
  enabled = 1 if bool(payload.get("enabled", True)) else 0
  if not nmid or not nmid.isdigit():
    raise ValueError("nmid required")
  if not advert_ids:
    raise ValueError("Нужно выбрать хотя бы одну кампанию")
  now = datetime.now(AD_BIDDER_TZ).isoformat(timespec="seconds")
  with _db_connect() as conn:
    _ensure_ad_bidding_tables(conn)
    if rule_id:
      conn.execute(
        "UPDATE ad_bidding_rules SET enabled=?, name=?, nmid=?, article=?, schedule_json=?, updated_at=? WHERE id=?",
        (enabled, name, nmid, article, json.dumps(schedule, ensure_ascii=False), now, rule_id),
      )
      conn.execute("DELETE FROM ad_bidding_rule_campaigns WHERE rule_id=?", (rule_id,))
    else:
      cur = conn.execute(
        "INSERT INTO ad_bidding_rules(enabled, name, nmid, article, timezone, schedule_json, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (enabled, name, nmid, article, "Europe/Moscow", json.dumps(schedule, ensure_ascii=False), now, now),
      )
      rule_id = int(cur.lastrowid)
    for advert_id in sorted(set(advert_ids)):
      conn.execute("INSERT OR IGNORE INTO ad_bidding_rule_campaigns(rule_id, advert_id) VALUES (?, ?)", (rule_id, advert_id))
    conn.commit()
  return {"ok": True, "id": rule_id}

def _toggle_ad_rule(rule_id: int, enabled: bool) -> dict[str, object]:
  with _db_connect() as conn:
    _ensure_ad_bidding_tables(conn)
    conn.execute(
      "UPDATE ad_bidding_rules SET enabled=?, updated_at=? WHERE id=?",
      (1 if enabled else 0, datetime.now(AD_BIDDER_TZ).isoformat(timespec="seconds"), rule_id),
    )
    conn.commit()
  return {"ok": True}

def _ad_try_lock(conn: sqlite3.Connection) -> bool:
  _ensure_ad_bidding_tables(conn)
  now = datetime.now(AD_BIDDER_TZ)
  now_iso = now.isoformat(timespec="seconds")
  locked_until = (now + timedelta(seconds=120)).isoformat(timespec="seconds")
  row = conn.execute("SELECT locked_until FROM ad_bidding_executor_lock WHERE key='main'").fetchone()
  if row and str(row["locked_until"] or "") > now_iso:
    return False
  conn.execute(
    "INSERT INTO ad_bidding_executor_lock(key, locked_until) VALUES ('main', ?) "
    "ON CONFLICT(key) DO UPDATE SET locked_until=excluded.locked_until",
    (locked_until,),
  )
  conn.commit()
  return True

def _ad_release_lock(conn: sqlite3.Connection) -> None:
  conn.execute(
    "UPDATE ad_bidding_executor_lock SET locked_until=? WHERE key='main'",
    ((datetime.now(AD_BIDDER_TZ) - timedelta(seconds=1)).isoformat(timespec="seconds"),),
  )
  conn.commit()

def _ad_sync_spend_if_due(cabinet: dict | None, conn: sqlite3.Connection, rows: list[sqlite3.Row], *, force: bool = False) -> tuple[dict[str, float], str]:
  now = datetime.now(AD_BIDDER_TZ)
  today = now.date().isoformat()
  due_ids = []
  for row in rows:
    advert_id = str(row["advert_id"])
    spend_day = str(row["spend_day"] or "")
    last_sync = str(row["last_spend_sync_at"] or "")
    due = force or spend_day != today or not last_sync
    if not due:
      try:
        last_dt = datetime.fromisoformat(last_sync)
        if last_dt.tzinfo is None:
          last_dt = last_dt.replace(tzinfo=AD_BIDDER_TZ)
        due = now - last_dt >= timedelta(hours=1)
      except ValueError:
        due = True
    if due:
      due_ids.append(advert_id)
  if not due_ids:
    if force:
      now_iso = now.isoformat(timespec="seconds")
      _ad_log(
        conn,
        "poll",
        "",
        f"Опрос WB: расход за {today}",
        "Нет кампаний, которым сейчас нужен опрос расхода.",
        {"day": today, "force": force, "spend": {}},
        now_iso,
      )
      conn.commit()
    return {}, today
  spend = _ad_fetch_today_spend(cabinet, due_ids, today)
  now_iso = now.isoformat(timespec="seconds")
  for advert_id, value in spend.items():
    conn.execute(
      "UPDATE ad_bidding_campaign_settings SET today_spend=?, spend_day=?, last_spend_sync_at=? WHERE advert_id=?",
      (value, today, now_iso, advert_id),
    )
  top_items = sorted(spend.items(), key=lambda item: -item[1])[:20]
  top_text = ", ".join(f"{advert_id}: {value:g} ₽" for advert_id, value in top_items)
  _ad_log(
    conn,
    "poll",
    "",
    f"Опрос WB: расход за {today}",
    f"Кампаний: {len(spend)}; суммарный расход: {sum(spend.values()):g} ₽" + (f"; топ: {top_text}" if top_text else ""),
    {"day": today, "force": force, "spend": spend},
    now_iso,
  )
  conn.commit()
  return spend, today

def _run_ad_bidding_once_for_cabinet(cabinet: dict | None, *, force_spend_sync: bool = False) -> dict[str, object]:
  if not _has_wb_ad_cabinet(cabinet):
    return {"ok": False, "message": "no WB advert token", "actions": []}
  db_path = Path(_cabinet_db_path(cabinet))
  if not db_path.exists():
    return {"ok": False, "message": "cabinet db missing", "actions": []}
  actions_out = []
  with _ad_db_connect_for_cabinet(cabinet) as conn:
    if not _ad_try_lock(conn):
      return {"ok": True, "message": "locked", "actions": []}
    try:
      _ensure_ad_bidding_tables(conn)
      rows = conn.execute("SELECT * FROM ad_bidding_campaign_settings WHERE enabled=1 OR auto_pause=1").fetchall()
      spend_sync_error = False
      try:
        synced_spend, today = _ad_sync_spend_if_due(cabinet, conn, rows, force=force_spend_sync)
      except Exception as exc:
        today = datetime.now(AD_BIDDER_TZ).date().isoformat()
        synced_spend = {}
        spend_sync_error = True
        _ad_log(
          conn,
          "error",
          "",
          "Ошибка опроса WB",
          str(exc),
          {"force_spend_sync": force_spend_sync},
        )
        conn.commit()
      if synced_spend:
        rows = conn.execute("SELECT * FROM ad_bidding_campaign_settings WHERE enabled=1 OR auto_pause=1").fetchall()
      now = datetime.now(AD_BIDDER_TZ).isoformat(timespec="seconds")
      for row in rows:
        advert_id = str(row["advert_id"])
        budget = _to_float(row["daily_budget"])
        today_spend = _to_float(row["today_spend"]) if str(row["spend_day"] or "") == today else 0.0
        cache_row = conn.execute("SELECT name FROM ad_campaign_cache WHERE advert_id=?", (advert_id,)).fetchone()
        campaign_name = str(row["name"] or (cache_row["name"] if cache_row else "") or f"Кампания {advert_id}").strip()
        nmid = str(row["nmid"] or "").strip()
        article = str(row["article"] or "").strip() or nmid or "—"
        over_budget = bool(row["auto_pause"]) and budget > 0 and today_spend >= budget
        pause_now = bool(row["enabled"]) and _ad_interval_active(row["pause_start"], row["pause_end"])
        if over_budget:
          desired_state = "paused"
          reason = "over_budget"
        elif pause_now:
          desired_state = "paused"
          reason = "pause_interval"
        elif bool(row["enabled"]):
          desired_state = "active"
          reason = "pause_interval_inactive"
        else:
          continue
        state_row = conn.execute("SELECT last_desired_state FROM ad_bidding_state WHERE advert_id=?", (advert_id,)).fetchone()
        if state_row and str(state_row["last_desired_state"] or "") == desired_state:
          continue
        actual_action, status_code, message = _ad_campaign_action(cabinet, advert_id, desired_state)
        conn.execute(
          "INSERT INTO ad_bidding_actions(rule_id, advert_id, desired_state, actual_action, status_code, message, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
          (None, advert_id, desired_state, actual_action, status_code, f"{reason}: {message}", now),
        )
        if actual_action in {"start", "pause", "noop"}:
          if actual_action in {"start", "pause"}:
            conn.execute(
              "UPDATE ad_campaign_cache SET status=?, updated_at=? WHERE advert_id=?",
              ("9" if actual_action == "start" else "11", now, advert_id),
            )
          conn.execute(
            "INSERT INTO ad_bidding_state(advert_id, last_desired_state, last_success_at, last_error_at, last_error) VALUES (?, ?, ?, NULL, NULL) "
            "ON CONFLICT(advert_id) DO UPDATE SET last_desired_state=excluded.last_desired_state, last_success_at=excluded.last_success_at, last_error_at=NULL, last_error=NULL",
            (advert_id, desired_state, now),
          )
        else:
          conn.execute(
            "INSERT INTO ad_bidding_state(advert_id, last_desired_state, last_success_at, last_error_at, last_error) VALUES (?, NULL, NULL, ?, ?) "
            "ON CONFLICT(advert_id) DO UPDATE SET last_error_at=excluded.last_error_at, last_error=excluded.last_error",
            (advert_id, now, message),
          )
        action_payload = {
          "advert_id": advert_id,
          "campaign_name": campaign_name,
          "nmid": nmid,
          "article": article,
          "daily_budget": budget,
          "today_spend": today_spend,
          "spend_day": today,
          "desired_state": desired_state,
          "actual_action": actual_action,
          "status_code": status_code,
          "message": message,
          "reason": reason,
        }
        actions_out.append(action_payload)
        log_details = (
          _ad_campaign_log_context(advert_id, campaign_name, article, budget, today_spend)
          + f"; причина: {reason}; целевое состояние: {desired_state}; статус: {status_code}; {message}"
        )
        _ad_log(
          conn,
          "auto",
          advert_id,
          f"Автодействие: {actual_action}",
          log_details,
          action_payload,
          now,
        )
      conn.commit()
      return {"ok": True, "actions": actions_out, "spend_synced": sorted(synced_spend), "spend_sync_error": spend_sync_error}
    finally:
      _ad_release_lock(conn)

def _run_ad_bidding_once_all() -> bool:
  had_spend_sync_error = False
  try:
    cabinets = get_platform().list_cabinets()
  except Exception:
    return False
  for item in cabinets:
    cabinet = get_platform().get_cabinet(str(item.get("cabinet_id") or ""))
    if not _has_wb_ad_cabinet(cabinet):
      continue
    try:
      result = _run_ad_bidding_once_for_cabinet(cabinet)
      had_spend_sync_error = had_spend_sync_error or bool(result.get("spend_sync_error"))
    except Exception as exc:
      print(f"[WARN] ad bidder {item.get('cabinet_id')}: {exc}")
  return had_spend_sync_error

def _ad_executor_loop() -> None:
  retry_after_error = False
  while True:
    if retry_after_error:
      time.sleep(60)
    else:
      now = datetime.now(AD_BIDDER_TZ)
      next_hour = (now.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1))
      time.sleep(max(1, int((next_hour - now).total_seconds())))
    retry_after_error = _run_ad_bidding_once_all()

def _start_ad_executor() -> None:
  global _ad_executor_started
  with _ad_executor_lock:
    if _ad_executor_started:
      return
    thread = threading.Thread(target=_ad_executor_loop, name="wb-ad-bidder", daemon=True)
    thread.start()
    _ad_executor_started = True

def _ensure_preliminary_economics_tables(conn: sqlite3.Connection) -> None:
  conn.execute(
    """
    CREATE TABLE IF NOT EXISTS app_settings (
      key TEXT PRIMARY KEY,
      value TEXT NOT NULL,
      updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    )
    """
  )
  conn.execute(
    """
    CREATE TABLE IF NOT EXISTS preliminary_order_economics (
      sku TEXT NOT NULL,
      date TEXT NOT NULL,
      orders_count REAL NOT NULL,
      orders_sum REAL NOT NULL,
      commission_rub REAL NOT NULL,
      acquiring_rub REAL NOT NULL,
      advertising_rub REAL NOT NULL,
      additional_expenses_rub REAL NOT NULL,
      preliminary_profit_rub REAL NOT NULL,
      additional_rate REAL NOT NULL,
      updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
      PRIMARY KEY (sku, date)
    )
    """
  )

def _ensure_manager_comments_table(conn: sqlite3.Connection) -> None:
  conn.execute(
    """
    CREATE TABLE IF NOT EXISTS manager_comments (
      id           INTEGER PRIMARY KEY AUTOINCREMENT,
      nmId         INTEGER NOT NULL,
      comment_type TEXT NOT NULL,
      comment_text TEXT NOT NULL,
      created_at   DATETIME DEFAULT CURRENT_TIMESTAMP
    )
    """
  )
  conn.execute(
    "CREATE INDEX IF NOT EXISTS idx_mc_nmid ON manager_comments (nmId)"
  )
  now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
  try:
    conn.execute(
      "UPDATE manager_comments SET created_at = ? "
      "WHERE TRIM(COALESCE(created_at, '')) = '' OR LOWER(TRIM(COALESCE(created_at, ''))) = 'none'",
      (now,),
    )
    conn.execute(
      "UPDATE manager_comments SET id = rowid "
      "WHERE TRIM(COALESCE(CAST(id AS TEXT), '')) = '' OR LOWER(TRIM(COALESCE(CAST(id AS TEXT), ''))) = 'none'"
    )
  except sqlite3.DatabaseError:
    pass

def _fetch_comments_data() -> dict:
  with _db_connect() as conn:
    _ensure_manager_comments_table(conn)
    ctx = _scope_context(conn)

    sku_where = ["TRIM(COALESCE(\"Артикул WB\", '')) != ''"]
    sku_params: list = []
    sku_rows = conn.execute(
      'SELECT CAST(TRIM("Артикул WB") AS INTEGER) AS nmid, '
      'COALESCE(NULLIF(TRIM("Артикул поставщика"), \'\'), TRIM("Артикул WB")) AS article, '
      'COALESCE(NULLIF(TRIM("Название"), \'\'), NULLIF(TRIM("Наименование"), \'\'), "") AS name, '
      'COALESCE(TRIM("Предмет"), "") AS subject, '
      'COALESCE(TRIM("Ярлыки"), "") AS strategy, '
      'COALESCE(TRIM("Категория по выручке"), "") AS revenue_category '
      'FROM SKU '
      f'WHERE {" AND ".join(sku_where)}',
      sku_params,
    ).fetchall()

    nm_ids_int = [r["nmid"] for r in sku_rows if r["nmid"]]
    if not nm_ids_int:
      return {"rows": []}
    marketplace_article_by_nm = _marketplace_article_by_nm(conn, [str(nm) for nm in nm_ids_int])

    def _in_chunks(lst, n=900):
      for i in range(0, len(lst), n):
        yield lst[i:i + n]

    from datetime import date, timedelta
    today          = date.today()
    yesterday      = (today - timedelta(days=1)).isoformat()
    yday2          = (today - timedelta(days=2)).isoformat()
    seven_days_ago = (today - timedelta(days=7)).isoformat()
    today_iso      = today.isoformat()

    metrics_by_nm: dict = {}
    for chunk in _in_chunks(nm_ids_int):
      ph = ", ".join("?" for _ in chunk)
      for r in conn.execute(
        f'SELECT "nmId" AS nmid, '
        f'SUM(CASE WHEN "Дата" >= ? AND "Дата" <= ? THEN CAST(COALESCE("Реклама","0") AS REAL) ELSE 0 END) AS ads_7d, '
        f'SUM(CASE WHEN "Дата" = ? THEN CAST(COALESCE("Реклама","0") AS REAL) ELSE 0 END) AS ads_yday, '
        f'SUM(CASE WHEN "Дата" = ? THEN CAST(COALESCE("Реклама","0") AS REAL) ELSE 0 END) AS ads_yday2, '
        f'SUM(CASE WHEN "Дата" >= ? AND "Дата" <= ? AND CAST(COALESCE("Сумма выкупов в эту дату","0") AS REAL) > 0 '
        f'    THEN CAST("Сумма выкупов в эту дату" AS REAL) ELSE 0 END) AS rev_7d, '
        f'SUM(CASE WHEN "Дата" = ? AND CAST(COALESCE("Сумма выкупов в эту дату","0") AS REAL) > 0 '
        f'    THEN CAST("Сумма выкупов в эту дату" AS REAL) ELSE 0 END) AS rev_yday, '
        f'SUM(CASE WHEN "Дата" = ? AND CAST(COALESCE("Сумма выкупов в эту дату","0") AS REAL) > 0 '
        f'    THEN CAST("Сумма выкупов в эту дату" AS REAL) ELSE 0 END) AS rev_yday2, '
        f'SUM(CASE WHEN "Дата" >= ? AND "Дата" <= ? THEN CAST(COALESCE("Заказы, шт","0") AS REAL) ELSE 0 END) AS ord_7d, '
        f'SUM(CASE WHEN "Дата" = ? THEN CAST(COALESCE("Заказы, шт","0") AS REAL) ELSE 0 END) AS ord_yday, '
        f'SUM(CASE WHEN "Дата" = ? THEN CAST(COALESCE("Заказы, шт","0") AS REAL) ELSE 0 END) AS ord_yday2 '
        f'FROM buyout_order_day '
        f'WHERE "Дата" >= ? AND "Дата" <= ? AND "nmId" IN ({ph}) '
        f'GROUP BY "nmId"',
        [seven_days_ago, yesterday, yesterday, yday2,
         seven_days_ago, yesterday, yesterday, yday2,
         seven_days_ago, yesterday, yesterday, yday2,
         seven_days_ago, today_iso, *chunk],
      ).fetchall():
        nm = int(r["nmid"])
        metrics_by_nm[nm] = {
          "ads_7d":    _to_float(r["ads_7d"]),
          "ads_yday":  _to_float(r["ads_yday"]),
          "ads_yday2": _to_float(r["ads_yday2"]),
          "rev_7d":    _to_float(r["rev_7d"]),
          "rev_yday":  _to_float(r["rev_yday"]),
          "rev_yday2": _to_float(r["rev_yday2"]),
          "ord_7d":    _to_float(r["ord_7d"]),
          "ord_yday":  _to_float(r["ord_yday"]),
          "ord_yday2": _to_float(r["ord_yday2"]),
        }

    stock_by_nm: dict = {}
    for chunk in _in_chunks(nm_ids_int):
      ph = ", ".join("?" for _ in chunk)
      for r in conn.execute(
        f'SELECT "nmId" AS nmid, '
        f'SUM(CAST(REPLACE(REPLACE(COALESCE("quantity","0"), \' \', \'\'), \',\', \'.\') AS REAL)) AS stock '
        f'FROM raw_stocks WHERE "nmId" IN ({ph}) GROUP BY "nmId"',
        chunk,
      ).fetchall():
        stock_by_nm[int(r["nmid"])] = _to_float(r["stock"])

    latest: dict = {}
    for chunk in _in_chunks(nm_ids_int):
      ph = ", ".join("?" for _ in chunk)
      for r in conn.execute(
        f'SELECT nmId, comment_type, comment_text, created_at '
        f'FROM manager_comments m '
        f'WHERE nmId IN ({ph}) '
        f'AND NOT EXISTS ('
        f'  SELECT 1 FROM manager_comments m2 '
        f'  WHERE m2.nmId = m.nmId AND m2.comment_type = m.comment_type '
        f'  AND (COALESCE(m2.created_at, "") > COALESCE(m.created_at, "") '
        f'       OR (COALESCE(m2.created_at, "") = COALESCE(m.created_at, "") '
        f'           AND CAST(COALESCE(NULLIF(m2.id, ""), "0") AS INTEGER) > CAST(COALESCE(NULLIF(m.id, ""), "0") AS INTEGER)))'
        f')',
        chunk,
      ).fetchall():
        nm = int(r["nmId"])
        created_at = str(r["created_at"] or "").strip()
        _dt = created_at[:10].split("-")
        _dt_fmt = f"{_dt[2]}.{_dt[1]}" if len(_dt) == 3 else ""
        latest.setdefault(nm, {})[str(r["comment_type"])] = (
          (("(" + _dt_fmt + ")  ") if _dt_fmt else "") + str(r["comment_text"])
        )

    def _drr(ads, rev):
      if ads == 0:
        return 0.0
      return round(ads / rev * 100, 1) if rev > 0 else None

    rows = []
    for sku in sku_rows:
      nm   = int(sku["nmid"])
      m    = metrics_by_nm.get(nm, {})
      a7   = m.get("ads_7d",    0.0)
      ay   = m.get("ads_yday",  0.0)
      ay2  = m.get("ads_yday2", 0.0)
      r7   = m.get("rev_7d",    0.0)
      ry   = m.get("rev_yday",  0.0)
      ry2  = m.get("rev_yday2", 0.0)
      o7   = m.get("ord_7d",    0.0)
      oy   = m.get("ord_yday",  0.0)
      oy2  = m.get("ord_yday2", 0.0)
      stk  = stock_by_nm.get(nm, 0.0)
      s7   = round(o7 / 7, 1) if o7 else 0.0
      trn  = round(stk / oy, 0) if oy > 0 else (round(stk / (o7 / 7), 0) if o7 > 0 else None)
      last = latest.get(nm, {})
      rows.append({
        "nmId":             str(nm),
        "article":          str(marketplace_article_by_nm.get(str(nm)) or sku["article"] or nm),
        "name":             str(sku["name"] or ""),
        "subject":          str(sku["subject"] or ""),
        "labels":         _parse_labels(str(sku["strategy"] or "")),
        "revenue_category": str(sku["revenue_category"] or ""),
        "drr_7d":           _drr(a7, r7),
        "drr_yday2":        _drr(ay2, ry2),
        "drr_yday":         _drr(ay, ry),
        "speed_7d":         s7,
        "speed_yday2":      oy2,
        "speed_yday":       oy,
        "turnover":         trn,
        "stock":            int(round(stk)) if stk else 0,
        "last_general":     last.get("general", ""),
        "last_cpc":         last.get("cpc", ""),
        "last_shelves":     last.get("shelves", ""),
        "last_unified":     last.get("unified", ""),
        "last_search":      last.get("search", ""),
      })
    rows.sort(key=lambda r: r["article"])
    return {"rows": rows}

def _fetch_comment_log(nm_id: str = "", limit: int = 300) -> dict:
  with _db_connect() as conn:
    _ensure_manager_comments_table(conn)
    ctx = _scope_context(conn)

    where: list = []
    params: list = []
    nm_id = (nm_id or "").strip()
    if nm_id:
      where.append("mc.nmId = ?")
      params.append(int(nm_id))
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    rows = conn.execute(
      'SELECT mc.id, mc.nmId, mc.comment_type, mc.comment_text, mc.created_at, '
      'COALESCE(NULLIF(TRIM(bod.article), \'\'), NULLIF(TRIM(s."Артикул поставщика"), \'\'), CAST(mc.nmId AS TEXT)) AS article '
      'FROM manager_comments mc '
      'LEFT JOIN SKU s ON CAST(TRIM(s."Артикул WB") AS INTEGER) = mc.nmId '
      'LEFT JOIN ('
      '  SELECT CAST("nmId" AS INTEGER) AS nmid, MAX(NULLIF(TRIM("Артикул"), "")) AS article '
      '  FROM buyout_order_day GROUP BY CAST("nmId" AS INTEGER)'
      ') bod ON bod.nmid = mc.nmId '
      f'{where_sql} '
      'ORDER BY datetime(COALESCE(NULLIF(mc.created_at, ""), "1970-01-01 00:00:00")) DESC, '
      'CAST(COALESCE(NULLIF(mc.id, ""), "0") AS INTEGER) DESC LIMIT ?',
      [*params, min(limit, 2000)],
    ).fetchall()
    return {
      "rows": [
        {
          "id":           str(r["id"] or ""),
          "nmId":         str(r["nmId"]),
          "article":      str(r["article"]),
          "comment_type": str(r["comment_type"]),
          "comment_text": str(r["comment_text"]),
          "created_at":   str(r["created_at"] or ""),
        }
        for r in rows
      ]
    }

def _save_manager_comments(
  nm_ids: list,
  comment_type: str,
  comment_text: str,
) -> dict:
  VALID_TYPES = {"general", "cpc", "shelves", "unified", "search"}
  comment_type = (comment_type or "").strip()
  comment_text = (comment_text or "").strip()
  if comment_type not in VALID_TYPES:
    return {"error": f"invalid comment_type: {comment_type}"}
  if not comment_text:
    return {"error": "comment_text is required"}
  if not nm_ids:
    return {"error": "nm_ids list is required"}
  with _db_connect() as conn:
    _ensure_manager_comments_table(conn)
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.executemany(
      "INSERT INTO manager_comments (id, nmId, comment_type, comment_text, created_at) VALUES (?, ?, ?, ?, ?)",
      [
        (int(datetime.now().timestamp() * 1000000) + idx, int(nm_id), comment_type, comment_text, created_at)
        for idx, nm_id in enumerate(nm_ids)
      ],
    )
  return {"saved": len(nm_ids)}

# ── Tasks (Kanban) ────────────────────────────────────────────────────────────

TASK_STATUSES = ("backlog", "in_progress", "review", "done")

def _ensure_tasks_table(conn: sqlite3.Connection) -> None:
  conn.execute(
    """
    CREATE TABLE IF NOT EXISTS tasks (
      id         INTEGER PRIMARY KEY AUTOINCREMENT,
      title      TEXT NOT NULL,
      why        TEXT NOT NULL,
      result     TEXT NOT NULL,
      due_date   TEXT NOT NULL,
      status     TEXT NOT NULL DEFAULT 'backlog',
      created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
      updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    )
    """
  )
  conn.commit()

def _list_tasks() -> list:
  with _db_connect() as conn:
    _ensure_tasks_table(conn)
    rows = conn.execute(
      "SELECT * FROM tasks ORDER BY created_at"
    ).fetchall()
    return [dict(r) for r in rows]

def _create_task(title: str, why: str, result: str, due_date: str) -> dict:
  with _db_connect() as conn:
    _ensure_tasks_table(conn)
    cur = conn.execute(
      "INSERT INTO tasks (title, why, result, due_date) VALUES (?,?,?,?)",
      (title, why, result, due_date),
    )
    conn.commit()
    return {"id": cur.lastrowid}

def _update_task(task_id: int, **fields) -> dict:
  allowed = {"title", "why", "result", "due_date", "status"}
  updates = {k: v for k, v in fields.items() if k in allowed}
  if not updates:
    return {"error": "no valid fields"}
  updates["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
  set_clause = ", ".join(f"{k} = ?" for k in updates)
  values = list(updates.values()) + [task_id]
  with _db_connect() as conn:
    _ensure_tasks_table(conn)
    conn.execute(f"UPDATE tasks SET {set_clause} WHERE id = ?", values)
    conn.commit()
  return {"ok": True}

def _delete_task(task_id: int) -> dict:
  with _db_connect() as conn:
    _ensure_tasks_table(conn)
    conn.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
    conn.commit()
  return {"ok": True}

# ── OZON manager comments ────────────────────────────────────────────────────

def _ensure_ozon_manager_comments_table(conn: sqlite3.Connection) -> None:
  conn.execute(
    """
    CREATE TABLE IF NOT EXISTS ozon_manager_comments (
      id           INTEGER PRIMARY KEY AUTOINCREMENT,
      ozon_sku     TEXT NOT NULL,
      comment_type TEXT NOT NULL,
      comment_text TEXT NOT NULL,
      created_at   DATETIME DEFAULT CURRENT_TIMESTAMP
    )
    """
  )
  conn.execute(
    "CREATE INDEX IF NOT EXISTS idx_omc_sku ON ozon_manager_comments (ozon_sku)"
  )

def _fetch_ozon_comments_data() -> dict:
  from datetime import date, timedelta
  today = date.today()
  two_days_ago   = (today - timedelta(days=2)).isoformat()
  seven_days_ago = (today - timedelta(days=7)).isoformat()
  three_days_ago = (today - timedelta(days=3)).isoformat()
  today_iso      = today.isoformat()

  with _db_connect() as conn:
    _ensure_ozon_manager_comments_table(conn)

    sku_rows = conn.execute(
      "SELECT CAST(s.sku AS TEXT) AS ozon_sku, "
      "MAX(COALESCE(st.offer_id, '')) AS offer_id, "
      "MAX(COALESCE(s.item_name, '')) AS item_name "
      "FROM ozon_sku_day_analytics s "
      "LEFT JOIN ("
      "  SELECT CAST(sku AS TEXT) AS sku, MIN(offer_id) AS offer_id "
      "  FROM ozon_stock_on_warehouses GROUP BY sku"
      ") st ON st.sku = CAST(s.sku AS TEXT) "
      "WHERE s.day >= ? "
      "GROUP BY s.sku",
      [seven_days_ago],
    ).fetchall()

    if not sku_rows:
      return {"rows": []}

    sku_ids = [r["ozon_sku"] for r in sku_rows]

    def _in_chunks(lst, n=900):
      for i in range(0, len(lst), n):
        yield lst[i:i + n]

    ads_by_sku: dict = {}
    revenue_by_sku: dict = {}
    orders_7d_by_sku: dict = {}
    orders_3d_by_sku: dict = {}

    for chunk in _in_chunks(sku_ids):
      ph = ", ".join("?" for _ in chunk)
      for r in conn.execute(
        f"SELECT CAST(sku AS TEXT) AS sk, "
        f"SUM(CASE WHEN day >= ? THEN COALESCE(orders_qty,0) ELSE 0 END) AS ord_7d, "
        f"SUM(CASE WHEN day >= ? THEN COALESCE(orders_qty,0) ELSE 0 END) AS ord_3d, "
        f"SUM(CASE WHEN day >= ? THEN COALESCE(orders_revenue,0) ELSE 0 END) AS rev_2d "
        f"FROM ozon_sku_day_analytics "
        f"WHERE day >= ? AND day <= ? AND CAST(sku AS TEXT) IN ({ph}) "
        f"GROUP BY sku",
        [seven_days_ago, three_days_ago, two_days_ago, seven_days_ago, today_iso, *chunk],
      ).fetchall():
        sk = str(r["sk"])
        orders_7d_by_sku[sk] = _to_float(r["ord_7d"])
        orders_3d_by_sku[sk] = _to_float(r["ord_3d"])
        revenue_by_sku[sk]   = _to_float(r["rev_2d"])

    for chunk in _in_chunks(sku_ids):
      ph = ", ".join("?" for _ in chunk)
      for r in conn.execute(
        f"SELECT CAST(sku AS TEXT) AS sk, SUM(COALESCE(ad_spend,0)) AS ads "
        f"FROM ozon_sku_day_ad_spend "
        f"WHERE day >= ? AND day <= ? AND CAST(sku AS TEXT) IN ({ph}) "
        f"GROUP BY sku",
        [two_days_ago, today_iso, *chunk],
      ).fetchall():
        ads_by_sku[str(r["sk"])] = _to_float(r["ads"])

    stock_by_sku: dict = {}
    for chunk in _in_chunks(sku_ids):
      ph = ", ".join("?" for _ in chunk)
      for r in conn.execute(
        f"SELECT CAST(sku AS TEXT) AS sk, "
        f"SUM(COALESCE(free_to_sell_amount,0)+COALESCE(reserved_amount,0)) AS stock "
        f"FROM ozon_stock_on_warehouses WHERE CAST(sku AS TEXT) IN ({ph}) GROUP BY sku",
        chunk,
      ).fetchall():
        stock_by_sku[str(r["sk"])] = _to_float(r["stock"])

    latest: dict = {}
    for chunk in _in_chunks(sku_ids):
      ph = ", ".join("?" for _ in chunk)
      for r in conn.execute(
        f"SELECT ozon_sku, comment_type, comment_text, created_at "
        f"FROM ozon_manager_comments m "
        f"WHERE ozon_sku IN ({ph}) "
        f"AND NOT EXISTS ("
        f"  SELECT 1 FROM ozon_manager_comments m2 "
        f"  WHERE m2.ozon_sku = m.ozon_sku AND m2.comment_type = m.comment_type "
        f"  AND (m2.created_at > m.created_at "
        f"       OR (m2.created_at = m.created_at AND m2.id > m.id))"
        f")",
        chunk,
      ).fetchall():
        sk = str(r["ozon_sku"])
        latest.setdefault(sk, {})[str(r["comment_type"])] = (
          str(r["comment_text"]) + "  (" + str(r["created_at"])[:10] + ")"
        )

    rows = []
    for sku in sku_rows:
      sk   = str(sku["ozon_sku"])
      ads  = ads_by_sku.get(sk, 0.0)
      rev  = revenue_by_sku.get(sk, 0.0)
      ord7 = orders_7d_by_sku.get(sk, 0.0)
      ord3 = orders_3d_by_sku.get(sk, 0.0)
      stk  = stock_by_sku.get(sk, 0.0)
      drr_pct  = 0.0 if ads == 0 else (round(ads / rev * 100, 1) if rev > 0 else None)
      speed    = round(ord7 / 7, 1) if ord7 else 0.0
      avg3     = ord3 / 3 if ord3 else 0.0
      turnover = round(stk / avg3, 0) if avg3 > 0 else None
      last = latest.get(sk, {})
      rows.append({
        "ozon_sku":      sk,
        "offer_id":      str(sku["offer_id"] or ""),
        "item_name":     str(sku["item_name"] or ""),
        "drr_pct":       drr_pct,
        "drr_rub":       round(ads, 0) if ads else 0.0,
        "speed":         speed,
        "turnover":      turnover,
        "last_general":  last.get("general", ""),
        "last_cpc":      last.get("cpc", ""),
        "last_shelves":  last.get("shelves", ""),
        "last_unified":  last.get("unified", ""),
        "last_search":   last.get("search", ""),
      })
    rows.sort(key=lambda r: r["offer_id"] or r["ozon_sku"])
    return {"rows": rows}

def _fetch_ozon_comment_log(ozon_sku: str = "", limit: int = 300) -> dict:
  with _db_connect() as conn:
    _ensure_ozon_manager_comments_table(conn)
    where: list = []
    params: list = []
    ozon_sku = (ozon_sku or "").strip()
    if ozon_sku:
      where.append("mc.ozon_sku = ?")
      params.append(ozon_sku)
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    rows = conn.execute(
      "SELECT mc.id, mc.ozon_sku, mc.comment_type, mc.comment_text, mc.created_at, "
      "COALESCE(NULLIF(TRIM(st.offer_id), ''), mc.ozon_sku) AS offer_id "
      "FROM ozon_manager_comments mc "
      "LEFT JOIN ("
      "  SELECT CAST(sku AS TEXT) AS sku, MIN(offer_id) AS offer_id "
      "  FROM ozon_stock_on_warehouses GROUP BY sku"
      ") st ON st.sku = mc.ozon_sku "
      f"{where_sql} ORDER BY mc.id DESC LIMIT ?",
      [*params, min(limit, 2000)],
    ).fetchall()
    return {
      "rows": [
        {
          "id":           r["id"],
          "ozon_sku":     str(r["ozon_sku"]),
          "offer_id":     str(r["offer_id"]),
          "comment_type": str(r["comment_type"]),
          "comment_text": str(r["comment_text"]),
          "created_at":   str(r["created_at"]),
        }
        for r in rows
      ]
    }

def _save_ozon_manager_comments(
  skus: list,
  comment_type: str,
  comment_text: str,
) -> dict:
  VALID_TYPES = {"general", "cpc", "shelves", "unified", "search"}
  comment_type = (comment_type or "").strip()
  comment_text = (comment_text or "").strip()
  if comment_type not in VALID_TYPES:
    return {"error": f"invalid comment_type: {comment_type}"}
  if not comment_text:
    return {"error": "comment_text is required"}
  if not skus:
    return {"error": "skus list is required"}
  with _db_connect() as conn:
    _ensure_ozon_manager_comments_table(conn)
    conn.executemany(
      "INSERT INTO ozon_manager_comments (ozon_sku, comment_type, comment_text) VALUES (?, ?, ?)",
      [(str(sku), comment_type, comment_text) for sku in skus],
    )
  return {"saved": len(skus)}

def _get_or_set_preliminary_additional_rate(
  conn: sqlite3.Connection,
  columns: set[str],
  sales_column: str,
) -> tuple[float, list[str]]:
  setting_key = "preliminary_additional_rate"
  stored = conn.execute("SELECT value FROM app_settings WHERE key = ?", (setting_key,)).fetchone()
  if stored:
    try:
      rate = max(0.0, float(stored["value"]))
      return rate, []
    except ValueError:
      pass

  additional_candidates = [
    "Логистика",
    "Хранение",
    "Приемка",
    "Штрафы",
    "Удержания",
    "Доплаты",
    "НДС WB",
    "Платная приемка",
    "Платное хранение",
    "Платные услуги",
  ]
  additional_expr, used_columns = _sum_expr(columns, additional_candidates)
  sales_expr = f"COALESCE(SUM({_sql_ident(sales_column)}), 0)"
  row = conn.execute(
    f"SELECT CASE WHEN {sales_expr} > 0 THEN ({additional_expr}) / {sales_expr} ELSE 0 END AS rate "
    "FROM finance_article_day_detail"
  ).fetchone()
  rate = _to_float(row["rate"] if row else 0.0)
  if rate <= 0:
    rate = 0.05

  conn.execute(
    """
    INSERT INTO app_settings(key, value, updated_at)
    VALUES(?, ?, CURRENT_TIMESTAMP)
    ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=CURRENT_TIMESTAMP
    """,
    (setting_key, f"{rate:.8f}"),
  )
  return rate, used_columns

def _fetch_preliminary_economics(
  date_from: str,
  date_to: str,
  article_query: str = "",
  buyout_percent: float = 30.0,
  aggregate_by_period: bool = False,
) -> dict[str, object]:
  article_query = article_query.strip()
  buyout_percent = max(0.0, min(100.0, buyout_percent))
  buyout_factor = buyout_percent / 100.0
  with _db_connect() as conn:
    _ensure_preliminary_economics_tables(conn)
    columns = _table_columns(conn, "finance_article_day_detail")
    ctx = _scope_context(conn)

    sku_col = _first_existing(columns, ["Артикул", "SKU"])
    date_col = _first_existing(columns, ["Дата", "date"])
    orders_count_col = _first_existing(columns, ["Продажи, шт", "Заказы, шт", "Заказано, шт", "Количество заказов"])
    sales_col = _first_existing(columns, ["Наша цена", "Продажи по нашей цене", "Продажи по нашей цене, р"])
    if not sku_col or not date_col or not sales_col:
      return {"rows": [], "additional_rate": "0.00%", "expense_components": []}

    commission_expr, _ = _sum_expr(columns, ["Комиссия WB", "Вознаграждение WB"])
    ads_expr, _ = _sum_expr(columns, ["Реклама", "Рекламные расходы"])
    additional_rate, used_additional_cols = _get_or_set_preliminary_additional_rate(conn, columns, sales_col)

    where = [f"{_sql_ident(date_col)} >= ?", f"{_sql_ident(date_col)} <= ?"]
    params: list[str] = [date_from, date_to]
    if article_query:
      where.append(f"{_sql_ident(sku_col)} = ?")
      params.append(article_query)
    orders_count_expr = f"COALESCE(SUM({_sql_ident(orders_count_col)}), 0)" if orders_count_col else "0"
    select_date = f"{_sql_ident(date_col)} AS date, " if not aggregate_by_period else ""
    group_by = "GROUP BY sku, date" if not aggregate_by_period else "GROUP BY sku"
    order_by = "ORDER BY date DESC, sku ASC" if not aggregate_by_period else "ORDER BY sku ASC"
    sql = (
      "SELECT "
      f"{_sql_ident(sku_col)} AS sku, "
      f"{select_date}"
      f"{orders_count_expr} AS orders_count, "
      f"COALESCE(SUM({_sql_ident(sales_col)}), 0) AS orders_sum, "
      f"{commission_expr} AS commission_rub, "
      f"{ads_expr} AS advertising_rub "
      "FROM finance_article_day_detail "
      f"WHERE {' AND '.join(where)} "
      f"{group_by} "
      f"{order_by} "
      "LIMIT 10000"
    )
    base_rows = conn.execute(sql, params).fetchall()

    funnel_where = ["date >= ?", "date <= ?"]
    funnel_params: list[str] = [date_from, date_to]
    if article_query:
      funnel_where.append("supplierArticle = ?")
      funnel_params.append(article_query)
    funnel_select_date = "date, " if not aggregate_by_period else ""
    funnel_group_by = "GROUP BY date, supplierArticle" if not aggregate_by_period else "GROUP BY supplierArticle"
    funnel_rows = conn.execute(
      (
        f"SELECT {funnel_select_date} supplierArticle AS sku, "
        "SUM(COALESCE(CAST(orderCount AS REAL), 0)) AS funnel_orders_count "
        "FROM funnel_analytics "
        f"WHERE {' AND '.join(funnel_where)} "
        f"{funnel_group_by}"
      ),
      funnel_params,
    ).fetchall()
    if aggregate_by_period:
      funnel_orders_map = {
        str(row["sku"]): _to_float(row["funnel_orders_count"])
        for row in funnel_rows
      }
    else:
      funnel_orders_map = {
        (str(row["sku"]), str(row["date"])): _to_float(row["funnel_orders_count"])
        for row in funnel_rows
      }

    payload: list[dict[str, str]] = []
    to_store: list[tuple[object, ...]] = []
    for row in base_rows:
      sku = str(row["sku"])
      row_date = str(row["date"]) if not aggregate_by_period else f"{date_from}..{date_to}"
      finance_orders_count = _to_float(row["orders_count"])
      if aggregate_by_period:
        orders_count = funnel_orders_map.get(sku, finance_orders_count)
      else:
        orders_count = funnel_orders_map.get((sku, row_date), finance_orders_count)

      raw_orders_sum = _to_float(row["orders_sum"])
      raw_commission = _to_float(row["commission_rub"])
      advertising = _to_float(row["advertising_rub"])
      orders_sum = raw_orders_sum * buyout_factor
      commission = raw_commission * buyout_factor
      acquiring = (raw_orders_sum * 0.02) * buyout_factor
      additional_expenses = (raw_orders_sum * additional_rate) * buyout_factor
      preliminary_profit = orders_sum - commission - acquiring - advertising - additional_expenses
      ad_pct = (advertising / orders_sum * 100.0) if orders_sum else 0.0
      margin_pct = (preliminary_profit / orders_sum * 100.0) if orders_sum else 0.0

      date_key = "Период" if aggregate_by_period else "Дата"
      payload.append(
        {
          "Артикул / SKU": sku,
          date_key: row_date,
          "Количество заказов": str(int(round(orders_count))),
          "Сумма заказов": _format_metric("Сумма заказов", orders_sum),
          "% выкупа": _format_percent(buyout_percent),
          "Комиссия, ₽": _format_metric("Комиссия, ₽", commission),
          "Эквайринг, ₽": _format_metric("Эквайринг, ₽", acquiring),
          "Реклама, ₽": _format_metric("Реклама, ₽", advertising),
          "% рекламы": _format_percent(ad_pct),
          "Дополнительные расходы, ₽": _format_metric("Дополнительные расходы, ₽", additional_expenses),
          "Предварительная прибыль, ₽": _format_metric("Предварительная прибыль, ₽", preliminary_profit),
          "% маржинальности": _format_percent(margin_pct),
        }
      )
      if not aggregate_by_period:
        to_store.append(
          (
            sku,
            row_date,
            orders_count,
            orders_sum,
            commission,
            acquiring,
            advertising,
            additional_expenses,
            preliminary_profit,
            additional_rate,
          )
        )

    if to_store:
      conn.executemany(
        """
        INSERT INTO preliminary_order_economics(
          sku, date, orders_count, orders_sum, commission_rub, acquiring_rub,
          advertising_rub, additional_expenses_rub, preliminary_profit_rub, additional_rate, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(sku, date) DO UPDATE SET
          orders_count=excluded.orders_count,
          orders_sum=excluded.orders_sum,
          commission_rub=excluded.commission_rub,
          acquiring_rub=excluded.acquiring_rub,
          advertising_rub=excluded.advertising_rub,
          additional_expenses_rub=excluded.additional_expenses_rub,
          preliminary_profit_rub=excluded.preliminary_profit_rub,
          additional_rate=excluded.additional_rate,
          updated_at=CURRENT_TIMESTAMP
        """,
        to_store,
      )
    conn.commit()

  return {
    "rows": payload,
    "additional_rate": f"{additional_rate * 100:.2f}%",
    "buyout_percent": _format_percent(buyout_percent),
    "expense_components": used_additional_cols,
  }

def _render_table_links_html() -> str:
  try:
    existing_tables = db_store().list_tables()
  except Exception:
    return '<div class="tables-links empty">Не удалось прочитать список таблиц.</div>'

  # Показываем фиксированный набор ключевых таблиц даже до первого sync.
  ordered_tables = list(dict.fromkeys(CORE_TABLES + existing_tables))

  links = []
  existing_set = set(existing_tables)
  for table in ordered_tables:
    table_q = quote_plus(table)
    cls = "" if table in existing_set else " pending"
    suffix = "" if table in existing_set else " (пусто)"
    links.append(f'<a class="tbl-chip{cls}" href="{escape(f"/db?table={table_q}&page=1")}">{escape(table)}{escape(suffix)}</a>')
  return '<div class="tables-links">' + "".join(links) + "</div>"

def run_sync(date_from: str, date_to: str, skip_ads: bool, skip_funnel: bool, log_q: queue.Queue, only: str = "all", cabinet: dict | None = None) -> None:
    cmd = [PYTHON, "-u", "-m", "wb_gsheets.main", "--date-from", date_from, "--date-to", date_to]
    if only != "all":
        cmd.extend(["--only", only])
    elif skip_ads:
        cmd.append("--skip-ads")
    if only == "all" and skip_funnel:
        cmd.append("--skip-funnel")
    if only == "all":
        cmd.append("--slim")

    cab = cabinet or {}
    wb_token = cab.get("wb_api_token", "")
    db_path = _cabinet_db_path(cab)
    env_patch = {
        "PYTHONPATH": str(ROOT / "src"),
        "PYTHONUNBUFFERED": "1",
        # Фоновый поток не видит request-local кабинет, поэтому путь передаем явно.
        "SQLITE_DB_PATH": db_path,
        "WB_API_TOKEN": wb_token,
        "WB_FINANCE_TOKEN": cab.get("wb_finance_token", "") or wb_token,
        "WB_ADV_TOKEN": cab.get("wb_adv_token", "") or wb_token,
        "ARTICLE_FILTER_TYPE": cab.get("article_filter_type", "") or "nmId",
        "DISABLE_SCOPE_FILTER": "1" if cab.get("disable_scope_filter") else "",
    }
    env = {**os.environ, **env_patch}
    log_paths = _sync_log_paths(date_from, date_to, only)
    _write_sync_log(
        log_paths,
        f"Новый запуск sync: период {date_from}..{date_to}, режим {only}.",
        reset=True,
    )

    try:
        mode_label = {
            "all": "все данные",
            "sales": "только продажи",
            "orders": "только заказы",
            "stocks": "только остатки",
            "ads": "только реклама",
            "funnel": "только воронка",
        }.get(only, only)
        _emit_sync_message(log_q, log_paths, "log", f"Старт: период {date_from}..{date_to}, режим: {mode_label}")
        _emit_sync_message(log_q, log_paths, "log", f"База SQLite: {db_path}")
        _emit_sync_message(log_q, log_paths, "log", f"Команда: {' '.join(cmd)}")
        _emit_sync_message(log_q, log_paths, "log", f"Файл лога: {log_paths[0]}")
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            cwd=str(ROOT),
            env=env,
        )
        assert proc.stdout
        for line in proc.stdout:
            _emit_sync_message(log_q, log_paths, "log", line.rstrip())
        proc.wait()
        if proc.returncode == 0:
            _emit_sync_message(log_q, log_paths, "done", "✅ Готово! Данные загружены.")
        else:
            _emit_sync_message(log_q, log_paths, "error", f"❌ Ошибка (код {proc.returncode})")
    except Exception as exc:
        _emit_sync_message(log_q, log_paths, "error", f"❌ {exc}")

def run_ozon_sync(
    date_from: str,
    date_to: str,
    log_q: queue.Queue,
    skip_ads: bool = True,
    cabinet: dict | None = None,
    ads_only: bool = False,
) -> None:
    cab = cabinet or {}
    cabinet_id = cab.get("cabinet_id", "default")
    db_path = _cabinet_db_path(cab)
    cmd = [PYTHON, "-u", str(ROOT / "scripts" / "ozon_sync.py"),
           "--date-from", date_from, "--date-to", date_to,
           "--db", db_path,
           "--cabinet-id", cabinet_id,
           "--skip-spp"]
    if ads_only:
        cmd.append("--ads-only")
    elif skip_ads:
        cmd.append("--skip-ads")
    env_patch = {
        "PYTHONUNBUFFERED": "1",
        "OZON_CLIENT_ID": cab.get("ozon_client_id", ""),
        "OZON_API_KEY": cab.get("ozon_api_key", ""),
        "OZON_PERFORMANCE_CLIENT_ID": cab.get("ozon_performance_client_id", ""),
        "OZON_PERFORMANCE_CLIENT_SECRET": cab.get("ozon_performance_client_secret", ""),
    }
    env = {**os.environ, **env_patch}
    mode_suffix = "ads" if ads_only else "full" if not skip_ads else "fast"
    log_paths = _sync_log_paths(date_from, date_to, f"ozon_{cabinet_id}_{mode_suffix}")
    _write_sync_log(log_paths, f"OZON синк: {date_from}..{date_to}, кабинет {cabinet_id}, режим {mode_suffix}", reset=True)
    try:
        _emit_sync_message(log_q, log_paths, "log", f"OZON синк: {date_from}..{date_to}")
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, cwd=str(ROOT), env=env)
        assert proc.stdout
        for line in proc.stdout:
            _emit_sync_message(log_q, log_paths, "log", line.rstrip())
        proc.wait()
        if proc.returncode == 0:
            _emit_sync_message(log_q, log_paths, "done", "✅ OZON данные загружены.")
        else:
            _emit_sync_message(log_q, log_paths, "error", f"❌ Ошибка (код {proc.returncode})")
    except Exception as exc:
        _emit_sync_message(log_q, log_paths, "error", f"❌ {exc}")

# ── OZON analytics functions (from mipao) ────────────────────────────────
def _ozon_label_skus(
  conn: sqlite3.Connection,
  labels: list[str],
  date_from: str,
  date_to: str,
) -> set[str]:
  selected = [str(v).strip() for v in labels if str(v).strip()]
  if not selected:
    return set()
  columns = _table_columns(conn, "SKU")
  sku_col = _first_existing(columns, ["SKU Ozon", "ID товара маркетплейса", "ozon_sku", "sku", "SKU"])
  labels_col = _first_existing(columns, ["Ярлыки", "Стратегия"])
  if not sku_col or not labels_col:
    return set()
  ors = " OR ".join(
    (
      f"(INSTR(',' || REPLACE(REPLACE(TRIM(COALESCE({_sql_ident(labels_col)},'')), ', ', ','), ' ,', ',') || ',', "
      "',' || ? || ',') > 0)"
    )
    for _ in selected
  )
  rows = conn.execute(
    (
      f"SELECT DISTINCT CAST({_sql_ident(sku_col)} AS TEXT) AS sku "
      "FROM SKU "
      f"WHERE TRIM(COALESCE({_sql_ident(sku_col)}, '')) != '' "
      f"AND ({ors}) "
      f"AND CAST({_sql_ident(sku_col)} AS TEXT) IN ("
      "  SELECT DISTINCT CAST(sku AS TEXT) FROM ozon_sku_day_analytics "
      "  WHERE day >= ? AND day <= ?"
      ")"
    ),
    [*selected, date_from, date_to],
  ).fetchall()
  return {str(r["sku"]).strip() for r in rows if str(r["sku"] or "").strip()}

def _fetch_ozon_stock_values_by_date(
  conn: sqlite3.Connection,
  *,
  effective_from: str,
  effective_to: str,
  article_query: str,
  selected_articles: list[str],
  subject: str,
  labels: list[str],
  revenue_category: str,
) -> dict[str, float]:
  def _append_ozon_stock_filters(where: list[str], params: list[object], alias: str) -> None:
    if subject:
      where.append(f"{alias}.item_name = ?")
      params.append(subject)
    if labels:
      label_skus = sorted(_ozon_label_skus(conn, labels, effective_from, effective_to))
      if not label_skus:
        where.append("1 = 0")
      else:
        placeholders = ", ".join("?" for _ in label_skus)
        where.append(f"CAST({alias}.sku AS TEXT) IN ({placeholders})")
        params.extend(label_skus)
    if selected_articles:
      placeholders = ", ".join("?" for _ in selected_articles)
      where.append(
        f"(CAST({alias}.sku AS TEXT) IN ({placeholders}) "
        f"OR TRIM(CAST(COALESCE({alias}.offer_id, '') AS TEXT)) IN ({placeholders}))"
      )
      params.extend(selected_articles)
      params.extend(selected_articles)
    elif article_query:
      where.append(
        f"(CAST({alias}.sku AS TEXT) = ? "
        f"OR TRIM(CAST(COALESCE({alias}.offer_id, '') AS TEXT)) = ?)"
      )
      params.extend([article_query, article_query])
    if revenue_category:
      where.append(
        f"CAST({alias}.sku AS TEXT) IN ("
        "SELECT sku FROM ("
        "SELECT sku, SUM(orders_revenue) AS rev, "
        "SUM(SUM(orders_revenue)) OVER (ORDER BY SUM(orders_revenue) DESC ROWS UNBOUNDED PRECEDING) AS cum, "
        "SUM(SUM(orders_revenue)) OVER () AS total "
        "FROM ozon_sku_day_analytics WHERE day >= ? AND day <= ? GROUP BY sku"
        ") WHERE CASE WHEN total > 0 AND cum/total <= 0.80 THEN 'A' "
        "WHEN total > 0 AND cum/total <= 0.95 THEN 'B' ELSE 'C' END = ?"
        ")"
      )
      params.extend([effective_from, effective_to, revenue_category])

  if _table_exists(conn, "ozon_stock_daily_snapshot"):
    columns = _table_columns(conn, "ozon_stock_daily_snapshot")
    if {"snapshot_date", "sku", "stock"}.issubset(columns):
      where = ["s.snapshot_date >= ?", "s.snapshot_date <= ?"]
      params: list[object] = [effective_from, effective_to]
      _append_ozon_stock_filters(where, params, "s")
      rows = conn.execute(
        (
          "SELECT s.snapshot_date AS day, SUM(CAST(s.stock AS REAL)) AS stock "
          "FROM ozon_stock_daily_snapshot s "
          f"WHERE {' AND '.join(where)} GROUP BY s.snapshot_date ORDER BY s.snapshot_date ASC"
        ),
        params,
      ).fetchall()
      result = {str(row["day"]): _to_float(row["stock"]) for row in rows}
      if result:
        # Fill-forward missing dates so sync gaps don't show as empty cells
        d = date.fromisoformat(effective_from)
        end = date.fromisoformat(effective_to)
        last_val = 0.0
        while d <= end:
          ds = d.isoformat()
          if ds in result:
            last_val = result[ds]
          elif last_val > 0:
            result[ds] = last_val
          d += timedelta(days=1)
        return result

  columns = _table_columns(conn, "ozon_stock_on_warehouses")
  if not {"sku", "free_to_sell_amount", "reserved_amount", "synced_at"}.issubset(columns):
    return {}
  where = ["substr(s.synced_at, 1, 10) >= ?", "substr(s.synced_at, 1, 10) <= ?"]
  params: list[object] = [effective_from, effective_to]
  _append_ozon_stock_filters(where, params, "s")
  rows = conn.execute(
    (
      "SELECT substr(s.synced_at, 1, 10) AS day, "
      "SUM(CAST(s.free_to_sell_amount AS REAL) + CAST(s.reserved_amount AS REAL)) AS stock "
      "FROM ozon_stock_on_warehouses s "
      f"WHERE {' AND '.join(where)} GROUP BY day ORDER BY day ASC"
    ),
    params,
  ).fetchall()
  return {str(row["day"]): _to_float(row["stock"]) for row in rows}


def _fetch_ozon_buyout_filter_options(
  date_from: str,
  date_to: str,
) -> dict[str, list[str]]:
  with _db_connect() as conn:
    # Subjects = unique item names from per-SKU data for the period
    subject_rows = conn.execute(
      """
      SELECT DISTINCT item_name FROM ozon_sku_day_analytics
      WHERE day >= ? AND day <= ? AND COALESCE(item_name, '') != ''
      ORDER BY item_name
      """,
      [date_from, date_to],
    ).fetchall()
    subjects = [str(r["item_name"]) for r in subject_rows]

    # ABC revenue categories — computed from per-SKU revenue totals
    abc_rows = conn.execute(
      """
      SELECT sku, SUM(orders_revenue) AS rev
      FROM ozon_sku_day_analytics
      WHERE day >= ? AND day <= ?
      GROUP BY sku ORDER BY rev DESC
      """,
      [date_from, date_to],
    ).fetchall()
    revenue_categories: list[str] = []
    if abc_rows:
      total = sum(_to_float(r["rev"]) for r in abc_rows)
      if total > 0:
        revenue_categories = ["A", "B", "C"]

    labels: set[str] = set()
    sku_columns = _table_columns(conn, "SKU")
    if "SKU Ozon" in sku_columns and "Ярлыки" in sku_columns:
      label_rows = conn.execute(
        """
        SELECT COALESCE("Ярлыки", '') AS yarlyky
        FROM SKU
        WHERE TRIM(COALESCE("Ярлыки", '')) != ''
          AND CAST("SKU Ozon" AS TEXT) IN (
            SELECT DISTINCT CAST(sku AS TEXT)
            FROM ozon_sku_day_analytics
            WHERE day >= ? AND day <= ?
          )
        """,
        [date_from, date_to],
      ).fetchall()
      for row in label_rows:
        for lbl in _parse_labels(str(row["yarlyky"] or "")):
          labels.add(lbl)

  return {
    "subjects": subjects,
    "labels": sorted(labels, key=str.lower),
    "revenue_categories": revenue_categories,
  }

def _fetch_ozon_buyout_articles(
  date_from: str,
  date_to: str,
  subject: str = "",
  labels: list[str] | None = None,
  revenue_category: str = "",
) -> list[dict[str, object]]:
  subject = subject.strip()
  labels = [str(v).strip() for v in (labels or []) if str(v).strip()]
  revenue_category = revenue_category.strip()

  with _db_connect() as conn:
    # Aggregate per-SKU metrics for the period
    where = ["s.day >= ?", "s.day <= ?"]
    params: list[object] = [date_from, date_to]
    if subject:
      where.append("s.item_name = ?")
      params.append(subject)
    if labels:
      label_skus = sorted(_ozon_label_skus(conn, labels, date_from, date_to))
      if not label_skus:
        return []
      placeholders = ", ".join("?" for _ in label_skus)
      where.append(f"CAST(s.sku AS TEXT) IN ({placeholders})")
      params.extend(label_skus)

    sku_rows = conn.execute(
      f"""
      SELECT s.sku,
        MAX(COALESCE(st.offer_id, '')) AS offer_id,
        MAX(COALESCE(s.item_name, '')) AS item_name,
        SUM(s.orders_revenue) AS revenue,
        SUM(s.orders_qty) AS orders_qty,
        SUM(s.delivered_qty) AS delivered_qty
      FROM ozon_sku_day_analytics s
      LEFT JOIN (
        SELECT sku, MIN(offer_id) AS offer_id FROM ozon_stock_on_warehouses GROUP BY sku
      ) st ON st.sku = s.sku
      WHERE {' AND '.join(where)}
      GROUP BY s.sku
      """,
      params,
    ).fetchall()

    # Stock per SKU
    stock_rows = conn.execute(
      """
      SELECT sku, SUM(COALESCE(free_to_sell_amount, 0) + COALESCE(reserved_amount, 0)) AS stock
      FROM ozon_stock_on_warehouses GROUP BY sku
      """
    ).fetchall()
    stock_by_sku = {str(r["sku"]): _to_float(r["stock"]) for r in stock_rows}

    ad_rows = conn.execute(
      """
      SELECT sku, SUM(COALESCE(ad_spend, 0)) AS ad_spend
      FROM ozon_sku_day_ad_spend
      WHERE day >= ? AND day <= ?
      GROUP BY sku
      """,
      [date_from, date_to],
    ).fetchall()
    ad_spend_by_sku = {str(r["sku"]): _to_float(r["ad_spend"]) for r in ad_rows}

  if not sku_rows:
    return []

  # Compute ABC
  total_rev = sum(_to_float(r["revenue"]) for r in sku_rows)
  abc_by_sku: dict[str, str] = {}
  if total_rev > 0:
    sorted_rows = sorted(sku_rows, key=lambda r: _to_float(r["revenue"]), reverse=True)
    cum = 0.0
    for r in sorted_rows:
      cum += _to_float(r["revenue"])
      pct = cum / total_rev
      sku = str(r["sku"])
      if pct <= 0.80:
        abc_by_sku[sku] = "A"
      elif pct <= 0.95:
        abc_by_sku[sku] = "B"
      else:
        abc_by_sku[sku] = "C"

  days_in_period = max(1, (date.fromisoformat(date_to) - date.fromisoformat(date_from)).days + 1)

  result = []
  for r in sku_rows:
    sku = str(r["sku"])
    abc = abc_by_sku.get(sku, "")
    if revenue_category and abc != revenue_category:
      continue
    revenue = _to_float(r["revenue"])
    orders_qty = _to_float(r["orders_qty"])
    stock = stock_by_sku.get(sku, 0.0)
    daily_demand = orders_qty / days_in_period
    turnover = int(stock / daily_demand) if daily_demand > 0 else 999
    offer_id = str(r["offer_id"] or sku)
    result.append({
      "article": offer_id,
      "nmid": sku,
      "stock": int(round(stock)),
      "revenue": round(revenue),
      "drr": round(ad_spend_by_sku.get(sku, 0.0) / revenue * 100.0, 1) if revenue > 0 and ad_spend_by_sku.get(sku, 0.0) > 0 else None,
      "turnover": min(999, turnover),
      "abc": abc,
    })

  result.sort(key=lambda r: r["revenue"], reverse=True)
  return result

def _fetch_ozon_buyout_order_day_pivot(
  date_from: str,
  date_to: str,
  article_query: str = "",
  articles: list[str] | None = None,
  subject: str = "",
  labels: list[str] | None = None,
  revenue_category: str = "",
  granularity: str = "day",
) -> dict[str, object]:
  dates, effective_from, effective_to = _date_range_limited(date_from, date_to, max_days=93 if granularity == "week" else 45)
  daily_dates = list(dates)
  article_query = (article_query or "").strip()
  subject = (subject or "").strip()
  labels = [str(v).strip() for v in (labels or []) if str(v).strip()]
  revenue_category = (revenue_category or "").strip()
  selected_articles = [a.strip() for a in (articles or []) if a.strip()]
  has_article_filter = bool(selected_articles or article_query or subject or labels or revenue_category)
  stock_by_date: dict[str, float] = {}

  with _db_connect() as conn:
    cogs_join = _ozon_sku_cogs_join(conn, "s")
    finance_columns = _table_columns(conn, "ozon_sku_day_finance") if _table_exists(conn, "ozon_sku_day_finance") else []
    if {"day", "sku"}.issubset(finance_columns):
      finance_select_parts = [
        "day",
        "sku",
        "COALESCE(sale_commission, 0) AS sale_commission" if "sale_commission" in finance_columns else "0 AS sale_commission",
        "COALESCE(delivery_charge, 0) AS delivery_charge" if "delivery_charge" in finance_columns else "0 AS delivery_charge",
        "COALESCE(delivered_qty, 0) AS delivered_qty" if "delivered_qty" in finance_columns else "0 AS delivered_qty",
        "COALESCE(accruals_for_sale, 0) AS accruals_for_sale" if "accruals_for_sale" in finance_columns else "0 AS accruals_for_sale",
      ]
      finance_join = (
        "LEFT JOIN (SELECT "
        + ", ".join(finance_select_parts)
        + " FROM ozon_sku_day_finance) f ON f.day = s.day AND f.sku = s.sku"
      )
    else:
      finance_join = (
        "LEFT JOIN (SELECT NULL AS day, NULL AS sku, 0 AS sale_commission, "
        "0 AS delivery_charge, 0 AS delivered_qty, 0 AS accruals_for_sale WHERE 0) f "
        "ON f.day = s.day AND f.sku = s.sku"
      )
    if has_article_filter:
      # Per-SKU daily pivot from ozon_sku_day_analytics
      sku_where = ["s.day >= ?", "s.day <= ?"]
      sku_params: list[object] = [effective_from, effective_to]
      if subject:
        sku_where.append("s.item_name = ?")
        sku_params.append(subject)
      if labels:
        label_skus = sorted(_ozon_label_skus(conn, labels, effective_from, effective_to))
        if not label_skus:
          if granularity == "week":
            out_dates: list[str] = []
            for day in daily_dates:
              label = _week_label(day, effective_from, effective_to)
              if not out_dates or out_dates[-1] != label:
                out_dates.append(label)
          else:
            out_dates = daily_dates
          return {
            "dates": out_dates,
            "rows": [],
            "effective_from": effective_from,
            "effective_to": effective_to,
            "max_days": 93 if granularity == "week" else 45,
            "granularity": granularity,
          }
        placeholders = ", ".join("?" for _ in label_skus)
        sku_where.append(f"CAST(s.sku AS TEXT) IN ({placeholders})")
        sku_params.extend(label_skus)
      if selected_articles:
        # articles are offer_ids; map to skus via ozon_stock_on_warehouses
        placeholders = ", ".join("?" for _ in selected_articles)
        sku_where.append(
          f"(s.sku IN (SELECT sku FROM ozon_stock_on_warehouses WHERE offer_id IN ({placeholders})) "
          f"OR s.sku IN ({placeholders}))"
        )
        sku_params.extend(selected_articles)
        sku_params.extend(selected_articles)
      elif article_query:
        sku_where.append("(st.offer_id = ? OR s.sku = ?)")
        sku_params.extend([article_query, article_query])
      if revenue_category:
        # Filter by ABC — compute inline
        sku_where.append(
          "s.sku IN ("
          "SELECT sku FROM ("
          "SELECT sku, SUM(orders_revenue) AS rev, "
          "SUM(SUM(orders_revenue)) OVER (ORDER BY SUM(orders_revenue) DESC ROWS UNBOUNDED PRECEDING) AS cum, "
          "SUM(SUM(orders_revenue)) OVER () AS total "
          "FROM ozon_sku_day_analytics WHERE day >= ? AND day <= ? GROUP BY sku"
          ") WHERE CASE WHEN cum/total <= 0.80 THEN 'A' WHEN cum/total <= 0.95 THEN 'B' ELSE 'C' END = ?"
          ")"
        )
        sku_params.extend([effective_from, effective_to, revenue_category])
      summary_rows = conn.execute(
        f"""
        SELECT s.day,
          SUM(s.orders_revenue) AS orders_revenue,
          SUM(s.orders_qty) AS orders_qty,
          SUM(s.delivered_qty) AS delivered_qty,
          SUM(s.cancellations_qty) AS cancellations_qty,
          CASE
            WHEN COALESCE(SUM(f.accruals_for_sale), 0) > 0 THEN SUM(f.accruals_for_sale)
            ELSE SUM(s.delivered_qty * CASE WHEN s.orders_qty > 0 THEN s.orders_revenue / s.orders_qty ELSE 0 END)
          END AS accruals_for_sale,
          SUM(COALESCE(f.sale_commission, 0)) AS sale_commission,
          SUM(COALESCE(f.delivery_charge, 0)) AS delivery_charge,
          0 AS for_pay,
          SUM(COALESCE(CASE WHEN f.delivered_qty > 0 THEN f.delivered_qty ELSE s.delivered_qty END, 0) * COALESCE(cogs_data.cogs, 0)) AS cogs_total,
          SUM(COALESCE(a.ad_spend, 0)) AS ad_spend,
          SUM(COALESCE(a.ad_views, 0)) AS ad_impressions,
          SUM(COALESCE(a.ad_clicks, 0)) AS ad_clicks,
          SUM(COALESCE(a.ad_orders, 0)) AS ad_orders,
          SUM(COALESCE(a.ad_revenue, 0)) AS ad_revenue,
          SUM(COALESCE(s.hits_view, 0)) AS hits_view,
          SUM(COALESCE(s.hits_view_pdp, 0)) AS hits_view_pdp,
          SUM(COALESCE(s.hits_tocart_pdp, 0)) AS hits_tocart_pdp
        FROM ozon_sku_day_analytics s
        LEFT JOIN (SELECT sku, MIN(offer_id) AS offer_id FROM ozon_stock_on_warehouses GROUP BY sku) st
          ON st.sku = s.sku
        LEFT JOIN ozon_sku_day_ad_spend a
          ON a.day = s.day AND a.sku = s.sku
        {finance_join}
        {cogs_join}
        WHERE {' AND '.join(sku_where)}
        GROUP BY s.day ORDER BY s.day ASC
        """,
        sku_params,
      ).fetchall()
    else:
      summary_rows = conn.execute(
        f"""
        SELECT s.day,
          SUM(s.orders_revenue) AS orders_revenue,
          SUM(s.orders_qty) AS orders_qty,
          SUM(s.delivered_qty) AS delivered_qty,
          SUM(s.cancellations_qty) AS cancellations_qty,
          COALESCE(ds.accruals_for_sale, 0) AS accruals_for_sale,
          COALESCE(ds.accruals_for_sale, 0) AS accruals_finance,
          COALESCE(ds.sale_commission, 0) AS sale_commission,
          COALESCE(ds.delivery_charge, 0) AS delivery_charge,
          COALESCE(ds.for_pay, 0) AS for_pay,
          SUM(COALESCE(s.delivered_qty, 0) * COALESCE(cogs_data.cogs, 0)) AS cogs_total,
          MAX(COALESCE(ds.ad_spend, 0)) AS ad_spend,
          MAX(COALESCE(ds.ad_impressions, 0)) AS ad_impressions,
          MAX(COALESCE(ds.ad_clicks, 0)) AS ad_clicks,
          MAX(COALESCE(ds.ad_orders, 0)) AS ad_orders,
          MAX(COALESCE(ds.ad_revenue, 0)) AS ad_revenue,
          SUM(COALESCE(s.hits_view, 0)) AS hits_view,
          SUM(COALESCE(s.hits_view_pdp, 0)) AS hits_view_pdp,
          SUM(COALESCE(s.hits_tocart_pdp, 0)) AS hits_tocart_pdp
        FROM ozon_sku_day_analytics s
        LEFT JOIN ozon_sku_day_ad_spend a
          ON a.day = s.day AND a.sku = s.sku
        LEFT JOIN ozon_daily_summary ds
          ON ds.day = s.day
        {cogs_join}
        WHERE s.day >= ? AND s.day <= ?
        GROUP BY s.day ORDER BY s.day ASC
        """,
        [effective_from, effective_to],
      ).fetchall()

    stock_by_date = _fetch_ozon_stock_values_by_date(
      conn,
      effective_from=effective_from,
      effective_to=effective_to,
      article_query=article_query,
      selected_articles=selected_articles,
      subject=subject,
      labels=labels,
      revenue_category=revenue_category,
    )

  sd: dict[str, dict] = {str(r["day"]): dict(r) for r in summary_rows}
  has_plugin = any(_to_float(r["hits_view"]) > 0 for r in summary_rows)

  if granularity == "week":
    out_dates: list[str] = []
    for day in daily_dates:
      label = _week_label(day, effective_from, effective_to)
      if not out_dates or out_dates[-1] != label:
        out_dates.append(label)
    w_sd: dict[str, dict] = {}
    for day in daily_dates:
      label = _week_label(day, effective_from, effective_to)
      sb = w_sd.setdefault(label, {})
      row = sd.get(day, {})
      for k in ["orders_revenue", "orders_qty", "delivered_qty", "cancellations_qty",
                "accruals_for_sale", "accruals_finance", "sale_commission", "delivery_charge", "for_pay",
                "cogs_total", "ad_spend", "ad_impressions", "ad_clicks", "ad_orders", "ad_revenue",
                "hits_view", "hits_view_pdp", "hits_tocart_pdp"]:
        sb[k] = sb.get(k, 0.0) + _to_float(row.get(k, 0))
    sd = w_sd
    stock_values_by_date = _stock_values_for_output_dates(
      stock_by_date,
      out_dates=out_dates,
      daily_dates=daily_dates,
      effective_from=effective_from,
      effective_to=effective_to,
      granularity=granularity,
    )
  else:
    out_dates = daily_dates
    stock_values_by_date = _stock_values_for_output_dates(
      stock_by_date,
      out_dates=out_dates,
      daily_dates=daily_dates,
      effective_from=effective_from,
      effective_to=effective_to,
      granularity=granularity,
    )

  def sv(day: str, key: str) -> float:
    r = sd.get(day)
    return _to_float(r[key]) if r and r.get(key) is not None else 0.0

  def fmt_m(v: float) -> str:
    return f"{v:,.0f}".replace(",", " ")

  def fmt_n(v: float) -> str:
    return f"{int(round(v)):,}".replace(",", " ")

  def fmt_p(v: float) -> str:
    return f"{v:.1f}%"

  n_dates = len(out_dates)

  def total_avg(total: float, fmt) -> str:
    avg = total / n_dates if n_dates else 0.0
    return f"{fmt(total)} / {fmt(avg)}"

  pivot_rows: list[dict] = []

  def add_simple(group: str, label: str, key_fn, kind: str, color: str) -> None:
    vals, raws, total = [], [], 0.0
    for day in out_dates:
      v = key_fn(day)
      total += v
      raws.append(v)
      vals.append(fmt_m(v) if kind == "money" else fmt_n(v))
    tot_str = total_avg(total, fmt_m if kind == "money" else fmt_n)
    pivot_rows.append({"group": group, "metric": label, "total": tot_str, "values": vals, "raw_values": raws, "kind": color})

  def add_ratio(group: str, label: str, num_fn, den_fn, mult: float, fmt, color: str) -> None:
    vals, raws, tn, td = [], [], 0.0, 0.0
    for day in out_dates:
      num, den = num_fn(day), den_fn(day)
      tn += num; td += den
      if den:
        v = num / den * mult
        vals.append(fmt(v)); raws.append(v)
      else:
        vals.append("—"); raws.append(None)
    tot = fmt(tn / td * mult) if td else "—"
    pivot_rows.append({"group": group, "metric": label, "total": tot, "values": vals, "raw_values": raws, "kind": color})

  add_simple("Суммы", "Сумма заказов, ₽", lambda d: sv(d, "orders_revenue"), "money", "income")
  add_simple("Суммы", "Сумма выкупов (факт), ₽", lambda d: sv(d, "accruals_for_sale"), "money", "income")
  add_simple("Штуки", "Заказы, шт", lambda d: sv(d, "orders_qty"), "number", "income")
  add_simple("Штуки", "Выкупы в эту дату, шт", lambda d: sv(d, "delivered_qty"), "number", "income")
  add_simple("Штуки", "Отмены, шт", lambda d: sv(d, "cancellations_qty"), "number", "expense")
  avg_sales_for_replenishment = (
    sum(max(sv(day, "delivered_qty"), sv(day, "orders_qty")) for day in out_dates) / len(out_dates)
    if out_dates else 0.0
  )
  _append_stock_pivot_row(
    pivot_rows,
    group="Остатки",
    label="Остатки, шт",
    dates=out_dates,
    values_by_date=stock_values_by_date,
    formatter=fmt_n,
    replenishment_threshold=max(3.0, avg_sales_for_replenishment * 2.0),
  )

  avck_vals, avck_raw, t_rev, t_qty = [], [], 0.0, 0.0
  for day in out_dates:
    rev, qty = sv(day, "orders_revenue"), sv(day, "orders_qty")
    t_rev += rev; t_qty += qty
    v = rev / qty if qty else 0.0
    avck_vals.append(fmt_m(v)); avck_raw.append(v)
  pivot_rows.append({"group": "Средние показатели", "metric": "Средний чек, ₽",
    "total": fmt_m(t_rev / t_qty if t_qty else 0.0),
    "values": avck_vals, "raw_values": avck_raw, "kind": "income"})

  if has_plugin:
    add_simple("Воронка", "Показы", lambda d: sv(d, "hits_view"), "number", "income")
    add_simple("Воронка", "Переходы в карточку", lambda d: sv(d, "hits_view_pdp"), "number", "income")
    add_ratio("Воронка", "CTR органический, %",
      lambda d: sv(d, "hits_view_pdp"), lambda d: sv(d, "hits_view"), 100.0, fmt_p, "percent_income")
    add_simple("Воронка", "В корзину", lambda d: sv(d, "hits_tocart_pdp"), "number", "income")
    add_ratio("Воронка", "CR в корзину, %",
      lambda d: sv(d, "hits_tocart_pdp"), lambda d: sv(d, "hits_view_pdp"), 100.0, fmt_p, "percent_income")
    add_ratio("Воронка", "% выкупа",
      lambda d: min(sv(d, "delivered_qty"), sv(d, "orders_qty")),
      lambda d: sv(d, "orders_qty"), 100.0, fmt_p, "percent_income")

  add_simple("Реклама", "Показы рекламы", lambda d: sv(d, "ad_impressions"), "number", "income")
  add_simple("Реклама", "Клики", lambda d: sv(d, "ad_clicks"), "number", "income")
  add_ratio("Реклама", "CTR рекламы, %",
    lambda d: sv(d, "ad_clicks"), lambda d: sv(d, "ad_impressions"), 100.0, fmt_p, "percent_income")
  add_ratio("Реклама", "CPM (1000 показов), ₽",
    lambda d: sv(d, "ad_spend"), lambda d: sv(d, "ad_impressions") / 1000.0, 1.0, fmt_m, "expense")
  add_ratio("Реклама", "CPC (цена клика), ₽",
    lambda d: sv(d, "ad_spend"), lambda d: sv(d, "ad_clicks"), 1.0, fmt_m, "expense")
  add_ratio("Реклама", "CPO (цена заказа), ₽",
    lambda d: sv(d, "ad_spend"), lambda d: sv(d, "ad_orders"), 1.0, fmt_m, "expense")
  add_simple("ДРР", "Расход, ₽", lambda d: sv(d, "ad_spend"), "money", "expense")
  add_ratio("ДРР", "ACOS (расход/выручка), %",
    lambda d: sv(d, "ad_spend"), lambda d: sv(d, "orders_revenue"), 100.0, fmt_p, "percent_expense")
  add_ratio("ДРР", "TACOS (от общих выкупов), %",
    lambda d: sv(d, "ad_spend"), lambda d: sv(d, "accruals_for_sale"), 100.0, fmt_p, "percent_expense")
  def _base_accruals(d: str) -> float:
    af = sv(d, "accruals_finance")
    return af if af > 0 else sv(d, "accruals_for_sale")

  def _ozon_commissions(d: str) -> float:
    return abs(sv(d, "sale_commission")) + abs(sv(d, "delivery_charge"))

  def _marginal_profit(d: str) -> float:
    return _base_accruals(d) - _ozon_commissions(d) - sv(d, "ad_spend") - sv(d, "cogs_total")

  add_simple("Финансы", "Комиссии Ozon, ₽", _ozon_commissions, "money", "expense")
  add_simple("Финансы", "Себестоимость, ₽", lambda d: sv(d, "cogs_total"), "money", "expense")
  add_simple("Финансы", "Маржинальная прибыль, ₽", _marginal_profit, "money", "income")
  add_ratio("Финансы", "Маржинальность, %",
    _marginal_profit,
    lambda d: _base_accruals(d), 100.0, fmt_p, "percent_income")

  return {
    "dates": out_dates,
    "rows": pivot_rows,
    "effective_from": effective_from,
    "effective_to": effective_to,
    "max_days": 93 if granularity == "week" else 45,
    "granularity": granularity,
  }

def _calc_elasticity(points: list[tuple[float, float]]) -> float | None:
  import math
  prepared = [(price, qty) for price, qty in points if price > 0 and qty > 0]
  if len(prepared) < 7:
    return None
  prices = [price for price, _qty in prepared]
  if max(prices) / min(prices) < 1.05:
    return None
  # require at least 2 distinct price clusters separated by >3%
  sorted_prices = sorted(set(round(p / 10) * 10 for p in prices))
  has_gap = any(sorted_prices[i + 1] / sorted_prices[i] >= 1.03 for i in range(len(sorted_prices) - 1))
  if not has_gap:
    return None
  xs = [math.log(price) for price, _qty in prepared]
  ys = [math.log(qty) for _price, qty in prepared]
  n = len(xs)
  avg_x = sum(xs) / n
  avg_y = sum(ys) / n
  ss_xx = sum((x - avg_x) ** 2 for x in xs)
  if ss_xx <= 0:
    return None
  ss_xy = sum((x - avg_x) * (y - avg_y) for x, y in zip(xs, ys))
  value = ss_xy / ss_xx
  # R² — share of demand variance explained by price
  ss_yy = sum((y - avg_y) ** 2 for y in ys)
  if ss_yy <= 0:
    return None
  r2 = (ss_xy ** 2) / (ss_xx * ss_yy)
  if r2 < 0.3:
    return None
  # clip to economically plausible range for WB: demand drops when price rises
  return max(-3.0, min(0.0, value))

def _calc_demand_model(
  points: list[tuple[float, float, float]],  # (price, ads_per_day, qty)
) -> tuple[float, float] | None:
  """Bivariate log-log OLS: log(qty) = e_price*log(price) + e_ads*log(1+ads) + const.
  Returns (e_price, e_ads). Falls back to one-factor models when only price or ads has signal."""
  import math
  prepared = [(p, a, q) for p, a, q in points if p > 0 and q > 0]
  if len(prepared) < 7:
    return None
  prices = [p for p, _, _ in prepared]
  ads_values = [a for _, a, _ in prepared]
  sorted_prices = sorted(set(round(p / 10) * 10 for p in prices))
  has_price_signal = (
    max(prices) / min(prices) >= 1.05
    and any(sorted_prices[i + 1] / sorted_prices[i] >= 1.03 for i in range(len(sorted_prices) - 1))
  )
  has_ads_signal = max(ads_values) > 0 and (max(ads_values) - min(ads_values)) > max(50.0, max(ads_values) * 0.15)
  xs1 = [math.log(p) for p, _, _ in prepared]
  xs2 = [math.log(1.0 + a) for _, a, _ in prepared]
  ys  = [math.log(q) for _, _, q in prepared]
  n = len(xs1)
  if not has_price_signal and has_ads_signal:
    mx = sum(xs2) / n
    my = sum(ys) / n
    ss_xx = sum((x - mx) ** 2 for x in xs2)
    ss_xy = sum((x - mx) * (y - my) for x, y in zip(xs2, ys))
    ss_yy = sum((y - my) ** 2 for y in ys)
    if ss_xx <= 0 or ss_yy <= 0:
      return None
    r2 = (ss_xy ** 2) / (ss_xx * ss_yy)
    if r2 < 0.12:
      return None
    return 0.0, max(0.0, min(1.5, ss_xy / ss_xx))
  if not has_price_signal:
    return None
  mx1 = sum(xs1) / n; mx2 = sum(xs2) / n; my = sum(ys) / n
  x1 = [v - mx1 for v in xs1]; x2 = [v - mx2 for v in xs2]; y = [v - my for v in ys]
  sa = sum(v * v for v in x1)
  sb = sum(i * j for i, j in zip(x1, x2))
  sd = sum(v * v for v in x2)
  sf = sum(i * j for i, j in zip(x1, y))
  sg = sum(i * j for i, j in zip(x2, y))
  det = sa * sd - sb * sb
  ads_has_variation = sd > 1e-12
  if abs(det) < 1e-12 or not ads_has_variation:
    e_p = _calc_elasticity([(p, q) for p, _, q in prepared])
    return (e_p, 0.0) if e_p is not None else None
  e_price_raw = (sd * sf - sb * sg) / det
  e_ads_raw   = (sa * sg - sb * sf) / det
  ss_tot = sum(v * v for v in y)
  if ss_tot <= 0:
    return None
  pred = [e_price_raw * i + e_ads_raw * j for i, j in zip(x1, x2)]
  ss_res = sum((yi - pi) ** 2 for yi, pi in zip(y, pred))
  r2 = 1.0 - ss_res / ss_tot
  if r2 < 0.3:
    e_p = _calc_elasticity([(p, q) for p, _, q in prepared])
    return (e_p, 0.0) if e_p is not None else None
  return max(-3.0, min(0.0, e_price_raw)), max(0.0, min(1.5, e_ads_raw))

def _calc_ads_confidence(stats: dict[str, object] | None) -> float:
  if not stats:
    return 0.0
  import math
  ad_days = _to_float(stats.get("ad_days") or 0)
  ad_orders = _to_float(stats.get("ad_orders") or 0)
  ad_views = _to_float(stats.get("ad_views") or 0)
  spend_values = [float(v) for v in (stats.get("spend_values") or []) if _to_float(v) > 0]
  variation = 0.0
  if len(spend_values) >= 5 and min(spend_values) > 0:
    variation = min(1.0, math.log(max(spend_values) / min(spend_values)) / math.log(5.0))
  confidence = (
    min(1.0, ad_days / 30.0) * 0.30
    + min(1.0, ad_orders / 50.0) * 0.30
    + min(1.0, ad_views / 50000.0) * 0.25
    + variation * 0.15
  )
  return max(0.0, min(1.0, confidence))

def _effective_ads_elasticity(measured: float | None, stats: dict[str, object] | None, *, group_fallback: bool = False) -> float:
  prior = 0.22
  confidence = _calc_ads_confidence(stats)
  if group_fallback:
    confidence = min(0.35, confidence * 0.5)
  else:
    confidence = min(0.85, confidence)
  measured_v = _to_float(measured)
  if measured_v <= 0:
    confidence = 0.0
  value = confidence * measured_v + (1.0 - confidence) * prior
  return max(0.08, min(0.85, value))

def _simulate_planning_orders(
  stock: float,
  daily_demand: float,
  buyout_rate: float,
  forecast_days: int,
  return_delay_days: int,
  start_day: date,
  warehouse_stock: float = 0.0,
  transit_days: int = 0,
) -> tuple[float, str]:
  available = max(0.0, stock)
  demand = max(0.0, daily_demand)
  buyout_rate = max(0.0, min(1.0, buyout_rate))
  delay = max(1, int(return_delay_days))
  transit = max(0, int(transit_days))
  # warehouse stock arrives at WB on day transit
  arrivals: dict[int, float] = {}
  if warehouse_stock > 0:
    arrivals[transit] = max(0.0, warehouse_stock)
  returns_by_day: dict[int, float] = {}
  total_orders = 0.0
  stockout_date = ""
  for idx in range(max(0, forecast_days)):
    available += returns_by_day.pop(idx, 0.0) + arrivals.pop(idx, 0.0)
    planned = demand
    ordered = min(planned, available)
    if planned > available and not stockout_date:
      stockout_date = (start_day + timedelta(days=idx)).isoformat()
    available -= ordered
    total_orders += ordered
    returning_qty = ordered * (1.0 - buyout_rate)
    due_idx = idx + delay
    if returning_qty > 0 and due_idx < forecast_days:
      returns_by_day[due_idx] = returns_by_day.get(due_idx, 0.0) + returning_qty
  return total_orders, stockout_date

def _fetch_planning(
  forecast_to: str,
  return_delay_days: int = 8,
  transit_days: int = 3,
  subject: str = "",
  labels: list[str] | None = None,
  revenue_category: str = "",
  price_change_pct: float = 0.0,
) -> dict[str, object]:
  target_date = date.fromisoformat(forecast_to)
  subject = subject.strip()
  labels = [l.strip() for l in (labels or []) if l.strip()]
  revenue_category = revenue_category.strip()
  price_change_pct = max(-90.0, min(300.0, float(price_change_pct)))
  return_delay_days = max(1, int(return_delay_days))
  transit_days = max(0, int(transit_days))

  with _db_connect() as conn:
    latest_raw = conn.execute('SELECT MAX("Дата") AS max_date FROM buyout_order_day').fetchone()["max_date"]
    if not latest_raw:
      return {"summary": {}, "rows": [], "options": {"subjects": [], "labels": [], "revenue_categories": []}}
    latest_day = date.fromisoformat(str(latest_raw))
    # exclude the latest day from speed/funnel baseline — it may be partial
    speed_baseline_to = latest_day - timedelta(days=1)
    baseline_from = (speed_baseline_to - timedelta(days=6)).isoformat()
    baseline_to = speed_baseline_to.isoformat()
    # buyout rate uses older data (14–44 days ago) — recent data is incomplete
    # because WB takes 7–14 days to finalize buyout/cancel counts
    buyout_rate_to = (latest_day - timedelta(days=14)).isoformat()
    buyout_rate_from = (latest_day - timedelta(days=44)).isoformat()
    # ads baseline: 14 days ending 2 days before latest (позавчера)
    ads_day_count = 14.0
    ads_baseline_to = (latest_day - timedelta(days=2)).isoformat()
    ads_baseline_from = (latest_day - timedelta(days=15)).isoformat()
    forecast_days = max(0, (target_date - latest_day).days)
    forecast_start = latest_day + timedelta(days=1)

    ctx = _scope_context(conn)
    sku_where = ["TRIM(COALESCE(\"Артикул WB\", '')) != ''"]
    sku_params: list[str] = []
    if subject:
      sku_where.append('"Предмет" = ?')
      sku_params.append(subject)
    if labels:
      lc_sku=[]; lp_sku:list[str]=[]
      for lbl in labels:
        lc_sku.append("(INSTR(',' || TRIM(COALESCE(\"Ярлыки\",'')) || ',', ',' || ? || ',') > 0)")
        lp_sku.append(lbl)
      sku_where.append('('+' OR '.join(lc_sku)+')')
      sku_params.extend(lp_sku)
    if revenue_category:
      sku_where.append('"Категория по выручке" = ?')
      sku_params.append(revenue_category)
    sku_rows = conn.execute(
      (
        'SELECT TRIM("Артикул WB") AS nmid, '
        'COALESCE(NULLIF(TRIM("Артикул поставщика"), \'\'), TRIM("Артикул WB")) AS article, '
        'COALESCE("Название", "Наименование", "") AS name, '
        'COALESCE("Предмет", "") AS subject, COALESCE("Ярлыки", "") AS strategy, '
        'COALESCE("Категория по выручке", "") AS revenue_category, '
        'COALESCE(CAST(REPLACE(REPLACE(COALESCE("себестоимость","0")," ",""),",",".") AS REAL), 0) AS cogs, '
        'COALESCE(CAST("склад" AS REAL), 0) AS warehouse_stock, '
        'COALESCE(CAST(REPLACE(REPLACE(COALESCE("% комиссии на вб","0"),"%","")," ","") AS REAL), 0) AS wb_commission_pct '
        'FROM SKU '
        f'WHERE {" AND ".join(sku_where)}'
      ),
      sku_params,
    ).fetchall()
    sku_by_nm = {str(row["nmid"]): dict(row) for row in sku_rows if str(row["nmid"]).strip()}

    options = _fetch_buyout_filter_options(baseline_from, baseline_to)
    if not sku_by_nm:
      return {
        "summary": {"forecast_days": forecast_days, "baseline_from": baseline_from, "baseline_to": baseline_to},
        "rows": [],
        "options": options,
      }

    nm_ids = sorted(sku_by_nm)
    marketplace_article_by_nm = _marketplace_article_by_nm(conn, nm_ids)
    placeholders = ", ".join("?" for _ in nm_ids)
    stock_rows = conn.execute(
      (
        'SELECT "nmId" AS nmid, SUM(CAST(REPLACE(REPLACE("quantity", \' \', \'\'), \',\', \'.\') AS REAL)) AS stock '
        'FROM raw_stocks '
        f'WHERE "nmId" IN ({placeholders}) GROUP BY "nmId"'
      ),
      nm_ids,
    ).fetchall()
    stock_by_nm = {str(row["nmid"]): _to_float(row["stock"]) for row in stock_rows}

    metric_rows = conn.execute(
      (
        'SELECT "nmId" AS nmid, SUM(CAST("Заказы, шт" AS REAL)) AS orders_qty, '
        'SUM(CAST("Сумма заказов" AS REAL)) AS orders_sum '
        'FROM buyout_order_day '
        f'WHERE "Дата" >= ? AND "Дата" <= ? AND "nmId" IN ({placeholders}) '
        'GROUP BY "nmId"'
      ),
      [baseline_from, baseline_to, *nm_ids],
    ).fetchall()
    metrics_by_nm = {str(row["nmid"]): row for row in metric_rows}

    funnel_columns = _table_columns(conn, "funnel_analytics")
    cancel_count_expr = 'SUM(CAST("cancelCount" AS REAL))' if "cancelCount" in funnel_columns else "0"
    buyout_rows = conn.execute(
      (
        'SELECT "nmId" AS nmid, SUM(CAST("buyoutCount" AS REAL)) AS buyouts, '
        f'{cancel_count_expr} AS cancels '
        'FROM funnel_analytics '
        f'WHERE "date" >= ? AND "date" <= ? AND "nmId" IN ({placeholders}) '
        'GROUP BY "nmId"'
      ),
      [baseline_from, baseline_to, *nm_ids],
    ).fetchall()
    wb_buyout_by_nm = {str(row["nmid"]): (_to_float(row["buyouts"]), _to_float(row["cancels"])) for row in buyout_rows}

    # ABC-классификация по выкупам (buyoutSum) за 7 дней: A=80%, B=15%, C=5%
    abc_sum_rows = conn.execute(
      (
        'SELECT "nmId" AS nmid, SUM(CAST("buyoutSum" AS REAL)) AS bsum '
        'FROM funnel_analytics '
        f'WHERE "date" >= ? AND "date" <= ? AND "nmId" IN ({placeholders}) '
        'GROUP BY "nmId"'
      ),
      [baseline_from, baseline_to, *nm_ids],
    ).fetchall()
    abc_sum_by_nm: dict[str, float] = {str(r["nmid"]): _to_float(r["bsum"]) for r in abc_sum_rows}
    total_abc = sum(abc_sum_by_nm.values())
    abc_category_by_nm: dict[str, str] = {}
    if total_abc > 0:
      sorted_nms = sorted(nm_ids, key=lambda n: abc_sum_by_nm.get(n, 0.0), reverse=True)
      cum = 0.0
      for n in sorted_nms:
        cum += abc_sum_by_nm.get(n, 0.0)
        pct = cum / total_abc
        if pct <= 0.80:
          abc_category_by_nm[n] = "A"
        elif pct <= 0.95:
          abc_category_by_nm[n] = "B"
        else:
          abc_category_by_nm[n] = "C"
      options["revenue_categories"] = ["A", "B", "C"]

    group_wb_buyout: dict[str, tuple[float, float]] = {}
    group_check: dict[str, tuple[float, float]] = {}
    total_wb_buyouts = 0.0
    total_wb_cancels = 0.0
    for nm_id, sku in sku_by_nm.items():
      row = metrics_by_nm.get(nm_id)
      orders_qty = _to_float(row["orders_qty"]) if row else 0.0
      orders_sum = _to_float(row["orders_sum"]) if row else 0.0
      wb_buyouts, wb_cancels = wb_buyout_by_nm.get(nm_id, (0.0, 0.0))
      total_wb_buyouts += wb_buyouts
      total_wb_cancels += wb_cancels
      computed_cat = abc_category_by_nm.get(nm_id, "") if total_abc > 0 else str(sku.get("revenue_category") or "")
      for key in (str(sku.get("subject") or ""), str(sku.get("strategy") or ""), computed_cat):
        if not key:
          continue
        bought, canceled = group_wb_buyout.get(key, (0.0, 0.0))
        group_wb_buyout[key] = (bought + wb_buyouts, canceled + wb_cancels)
        amount, qty = group_check.get(key, (0.0, 0.0))
        group_check[key] = (amount + orders_sum, qty + orders_qty)

    order_columns = _table_columns(conn, "raw_orders")
    price_col = "finishedPrice" if "finishedPrice" in order_columns else "priceWithDisc"
    elasticity_by_nm: dict[str, float] = {}
    group_elasticities: dict[str, list[float]] = {}
    if {"date", "nmId", price_col}.issubset(order_columns):
      history_from = (latest_day - timedelta(days=59)).isoformat()
      rows = conn.execute(
        (
          f'SELECT substr("date", 1, 10) AS day, "nmId" AS nmid, '
          f'AVG(CAST(REPLACE(REPLACE("{price_col}", \' \', \'\'), \',\', \'.\') AS REAL)) AS price, COUNT(*) AS qty '
          'FROM raw_orders '
          f'WHERE substr("date", 1, 10) >= ? AND "nmId" IN ({placeholders}) '
          'AND COALESCE("isCancel", "") NOT IN ("true", "1", "True") '
          'GROUP BY "nmId", substr("date", 1, 10)'
        ),
        [history_from, *nm_ids],
      ).fetchall()
      # store (day, price, qty) so we can join with ads by day
      day_points_by_nm: dict[str, list[tuple[str, float, float]]] = {}
      points_by_nm: dict[str, list[tuple[float, float]]] = {}
      for row in rows:
        nm = str(row["nmid"])
        p = _to_float(row["price"]); q = _to_float(row["qty"])
        day_points_by_nm.setdefault(nm, []).append((str(row["day"]), p, q))
        points_by_nm.setdefault(nm, []).append((p, q))

      # 60-day daily ads spend per nmId for bivariate model
      ads_tables = {str(r["name"]) for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
      ads_by_day_nm: dict[str, dict[str, float]] = {}
      ads_stats_by_nm: dict[str, dict[str, object]] = {}
      if "raw_ads" in ads_tables:
        ads_rows = conn.execute(
          f'SELECT substr("date",1,10) AS day, "nmId" AS nmid, '
          f'SUM(CAST("sum" AS REAL)) AS ads_sum, '
          f'SUM(CAST(COALESCE("views", "0") AS REAL)) AS ad_views, '
          f'SUM(CAST(COALESCE("clicks", "0") AS REAL)) AS ad_clicks, '
          f'SUM(CAST(COALESCE("orders", "0") AS REAL)) AS ad_orders '
          f'FROM raw_ads WHERE substr("date",1,10) >= ? AND "nmId" IN ({placeholders}) '
          'GROUP BY "nmId", substr("date",1,10)',
          [history_from, *nm_ids],
        ).fetchall()
        for ar in ads_rows:
          nm = str(ar["nmid"])
          ads_sum = _to_float(ar["ads_sum"])
          ads_by_day_nm.setdefault(nm, {})[str(ar["day"])] = ads_sum
          stats = ads_stats_by_nm.setdefault(nm, {
            "ad_days": 0.0,
            "ad_spend": 0.0,
            "ad_views": 0.0,
            "ad_clicks": 0.0,
            "ad_orders": 0.0,
            "spend_values": [],
          })
          if ads_sum > 0 or _to_float(ar["ad_views"]) > 0:
            stats["ad_days"] = _to_float(stats.get("ad_days") or 0) + 1.0
          stats["ad_spend"] = _to_float(stats.get("ad_spend") or 0) + ads_sum
          stats["ad_views"] = _to_float(stats.get("ad_views") or 0) + _to_float(ar["ad_views"])
          stats["ad_clicks"] = _to_float(stats.get("ad_clicks") or 0) + _to_float(ar["ad_clicks"])
          stats["ad_orders"] = _to_float(stats.get("ad_orders") or 0) + _to_float(ar["ad_orders"])
          stats.setdefault("spend_values", []).append(ads_sum)

      ads_elasticity_by_nm: dict[str, float] = {}
      group_ads_elasticities: dict[str, list[float]] = {}
      for nm_id, day_pts in day_points_by_nm.items():
        ads_day = ads_by_day_nm.get(nm_id, {})
        pts3: list[tuple[float, float, float]] = [
          (p, ads_day.get(d, 0.0), q) for d, p, q in day_pts
        ]
        result = _calc_demand_model(pts3) if pts3 else None
        if result is not None:
          e_p, e_a = result
        else:
          price_pts = [(p, q) for _, p, q in day_pts]
          e_p_only = _calc_elasticity(price_pts)
          if e_p_only is None:
            continue
          e_p, e_a = e_p_only, 0.0
        elasticity_by_nm[nm_id] = e_p
        if e_a > 0:
          ads_elasticity_by_nm[nm_id] = e_a
        sku = sku_by_nm.get(nm_id, {})
        _cat = abc_category_by_nm.get(nm_id, "") if total_abc > 0 else str(sku.get("revenue_category") or "")
        for key in (str(sku.get("subject") or ""), str(sku.get("strategy") or ""), _cat):
          if key:
            group_elasticities.setdefault(key, []).append(e_p)
            if e_a > 0:
              group_ads_elasticities.setdefault(key, []).append(e_a)

      # 14-day ads baseline (ending 2 days before latest) for DRR calculation
      base_ads_by_nm: dict[str, float] = {}
      if "raw_ads" in ads_tables:
        base_ads_rows = conn.execute(
          f'SELECT "nmId" AS nmid, SUM(CAST("sum" AS REAL)) AS ads_sum '
          f'FROM raw_ads WHERE "date" >= ? AND "date" <= ? AND "nmId" IN ({placeholders}) '
          'GROUP BY "nmId"',
          [ads_baseline_from, ads_baseline_to, *nm_ids],
        ).fetchall()
        for ar in base_ads_rows:
          base_ads_by_nm[str(ar["nmid"])] = _to_float(ar["ads_sum"])

      # actual buyout revenue over the same 14-day ads window (for DRR denominator)
      ads_revenue_rows = conn.execute(
        f'SELECT "nmId" AS nmid, SUM(CAST("buyoutSum" AS REAL)) AS buyout_sum '
        f'FROM funnel_analytics '
        f'WHERE "date" >= ? AND "date" <= ? AND "nmId" IN ({placeholders}) '
        'GROUP BY "nmId"',
        [ads_baseline_from, ads_baseline_to, *nm_ids],
      ).fetchall()
      ads_revenue_by_nm: dict[str, float] = {str(r["nmid"]): _to_float(r["buyout_sum"]) for r in ads_revenue_rows}

    else:
      ads_elasticity_by_nm = {}
      group_ads_elasticities = {}
      base_ads_by_nm = {}
      ads_revenue_by_nm = {}
      ads_stats_by_nm = {}

    # 7-day funnel baseline: CTR, CR1, CR2 from funnel_impressions_upload
    funnel_by_nm: dict[str, dict[str, float]] = {}
    all_tables = {str(r["name"]) for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    if "funnel_impressions_upload" in all_tables:
      fi_rows = conn.execute(
        f'SELECT "nmId" AS nmid, '
        f'SUM(CAST("impressions" AS REAL)) AS impr, '
        f'SUM(CAST("openCount" AS REAL)) AS opens, '
        f'SUM(CAST("cartCount" AS REAL)) AS carts, '
        f'SUM(CAST("orderCount" AS REAL)) AS orders_f '
        f'FROM funnel_impressions_upload '
        f'WHERE "date" >= ? AND "date" <= ? AND "nmId" IN ({placeholders}) '
        'GROUP BY "nmId"',
        [baseline_from, baseline_to, *nm_ids],
      ).fetchall()
      for fi in fi_rows:
        nm = str(fi["nmid"])
        impr  = _to_float(fi["impr"])
        opens = _to_float(fi["opens"])
        carts = _to_float(fi["carts"])
        ords  = _to_float(fi["orders_f"])
        funnel_by_nm[nm] = {
          "ctr":  opens / impr  * 100.0 if impr  > 0 else 0.0,
          "cr1":  carts / opens * 100.0 if opens > 0 else 0.0,
          "cr2":  ords  / carts * 100.0 if carts > 0 else 0.0,
        }

    current_price_by_nm: dict[str, float] = {}
    current_spp_by_nm: dict[str, float] = {}
    client_price_by_nm: dict[str, float] = {}
    baseline_avg_price_by_nm: dict[str, float] = {}
    baseline_client_check_by_nm: dict[str, float] = {}
    has_finished = "finishedPrice" in order_columns and "priceWithDisc" in order_columns
    if {"date", "nmId", price_col}.issubset(order_columns):
      seller_col = "priceWithDisc"
      client_col = "finishedPrice" if has_finished else "priceWithDisc"
      last_day_rows = conn.execute(
        (
          f'SELECT r."nmId" AS nmid, '
          f'AVG(CAST(REPLACE(REPLACE(r."{seller_col}", \' \', \'\'), \',\', \'.\') AS REAL)) AS seller_price, '
          f'AVG(CAST(REPLACE(REPLACE(r."{client_col}", \' \', \'\'), \',\', \'.\') AS REAL)) AS client_price, '
          f'AVG(CAST(REPLACE(REPLACE(COALESCE(r."spp", "0"), \' \', \'\'), \',\', \'.\') AS REAL)) AS spp '
          'FROM raw_orders r '
          f'JOIN (SELECT "nmId", MAX(substr("date",1,10)) AS max_date FROM raw_orders WHERE "nmId" IN ({placeholders}) GROUP BY "nmId") md '
          'ON r."nmId" = md."nmId" AND substr(r."date",1,10) = md.max_date '
          f'WHERE r."nmId" IN ({placeholders}) '
          'GROUP BY r."nmId"'
        ),
        [*nm_ids, *nm_ids],
      ).fetchall()
      for row in last_day_rows:
        if _to_float(row["seller_price"]) > 0:
          current_price_by_nm[str(row["nmid"])] = _to_float(row["seller_price"])
        if _to_float(row["client_price"]) > 0:
          client_price_by_nm[str(row["nmid"])] = _to_float(row["client_price"])
        current_spp_by_nm[str(row["nmid"])] = _to_float(row["spp"])
      baseline_rows = conn.execute(
        (
          f'SELECT "nmId" AS nmid, '
          f'AVG(CAST(REPLACE(REPLACE("{seller_col}", \' \', \'\'), \',\', \'.\') AS REAL)) AS price, '
          f'AVG(CAST(REPLACE(REPLACE("{client_col}", \' \', \'\'), \',\', \'.\') AS REAL)) AS client_price, '
          f'AVG(CAST(REPLACE(REPLACE(COALESCE("spp","0"), \' \', \'\'), \',\', \'.\') AS REAL)) AS avg_spp '
          'FROM raw_orders '
          f'WHERE substr("date", 1, 10) >= ? AND substr("date", 1, 10) <= ? AND "nmId" IN ({placeholders}) '
          'GROUP BY "nmId"'
        ),
        [baseline_from, baseline_to, *nm_ids],
      ).fetchall()
      baseline_spp_by_nm: dict[str, float] = {}
      baseline_client_check_by_nm: dict[str, float] = {}
      for row in baseline_rows:
        if _to_float(row["price"]) > 0:
          baseline_avg_price_by_nm[str(row["nmid"])] = _to_float(row["price"])
        if _to_float(row["client_price"]) > 0:
          baseline_client_check_by_nm[str(row["nmid"])] = _to_float(row["client_price"])
        baseline_spp_by_nm[str(row["nmid"])] = _to_float(row["avg_spp"])

    def _sku_keys(nm_id: str, sku: dict[str, object]) -> tuple[str, ...]:
      cat = abc_category_by_nm.get(nm_id, "") if total_abc > 0 else str(sku.get("revenue_category") or "")
      return (str(sku.get("subject") or ""), str(sku.get("strategy") or ""), cat)

    def fallback_buyout_rate(nm_id: str, sku: dict[str, object]) -> float:
      for key in _sku_keys(nm_id, sku):
        bought, canceled = group_wb_buyout.get(key, (0.0, 0.0))
        if bought + canceled > 0:
          return bought / (bought + canceled)
      if total_wb_buyouts + total_wb_cancels > 0:
        return total_wb_buyouts / (total_wb_buyouts + total_wb_cancels)
      return 0.3

    def fallback_check(nm_id: str, sku: dict[str, object]) -> float:
      for key in _sku_keys(nm_id, sku):
        amount, qty = group_check.get(key, (0.0, 0.0))
        if qty > 0:
          return amount / qty
      return 0.0

    def fallback_elasticity(nm_id: str, sku: dict[str, object]) -> float:
      if nm_id in elasticity_by_nm:
        v = elasticity_by_nm[nm_id]
        return v if v < 0.0 else -1.0  # 0.0 = positive slope clipped, treat as no data
      for key in _sku_keys(nm_id, sku):
        values = group_elasticities.get(key, [])
        if values:
          avg = sum(values) / len(values)
          return avg if avg < 0.0 else -1.0
      return -1.0

    def fallback_ads_elasticity(nm_id: str, sku: dict[str, object]) -> float:
      stats = ads_stats_by_nm.get(nm_id, {})
      if nm_id in ads_elasticity_by_nm:
        return _effective_ads_elasticity(ads_elasticity_by_nm[nm_id], stats)
      for key in _sku_keys(nm_id, sku):
        values = group_ads_elasticities.get(key, [])
        if values:
          return _effective_ads_elasticity(sum(values) / len(values), stats, group_fallback=True)
      return _effective_ads_elasticity(None, stats, group_fallback=True)

    rows_out: list[dict[str, object]] = []
    summary = {
      "stock": 0.0,
      "stock_cost": 0.0,
      "stock_value": 0.0,
      "forecast_orders": 0.0,
      "forecast_buyouts": 0.0,
      "orders_revenue": 0.0,
      "forecast_revenue": 0.0,
      "lost_orders": 0.0,
      "lost_revenue": 0.0,
    }
    day_count = 7.0
    for nm_id in nm_ids:
      sku = sku_by_nm[nm_id]
      metric = metrics_by_nm.get(nm_id)
      orders_qty = _to_float(metric["orders_qty"]) if metric else 0.0
      orders_sum = _to_float(metric["orders_sum"]) if metric else 0.0
      base_speed = orders_qty / day_count
      wb_buyouts, wb_cancels = wb_buyout_by_nm.get(nm_id, (0.0, 0.0))
      buyout_rate = wb_buyouts / (wb_buyouts + wb_cancels) if (wb_buyouts + wb_cancels) > 0 else fallback_buyout_rate(nm_id, sku)
      buyout_rate = max(0.0, min(1.0, buyout_rate))
      average_check = orders_sum / orders_qty if orders_qty > 0 else fallback_check(nm_id, sku)
      elasticity = fallback_elasticity(nm_id, sku)
      ads_elasticity = fallback_ads_elasticity(nm_id, sku)
      current_price = current_price_by_nm.get(nm_id, 0.0)
      current_spp = current_spp_by_nm.get(nm_id, 0.0)
      current_client_price = client_price_by_nm.get(nm_id, 0.0)
      baseline_client_price = baseline_client_check_by_nm.get(nm_id, current_client_price)
      # speed corrections use client price — that's what drives demand
      if baseline_client_price > 0 and current_client_price > 0:
        current_price_delta = (current_client_price / baseline_client_price - 1.0)
        current_factor = max(0.0, 1.0 + elasticity * current_price_delta)
      else:
        current_factor = 1.0
      model_current_speed = base_speed * current_factor
      current_speed_day = baseline_to
      yesterday_speed = _order_speed_for_day(conn, nm_id, current_speed_day)
      if yesterday_speed is None:
        current_speed_day = (date.fromisoformat(baseline_to) - timedelta(days=1)).isoformat()
        yesterday_speed = _order_speed_for_day(conn, nm_id, current_speed_day)
      current_speed = yesterday_speed if yesterday_speed is not None else model_current_speed
      current_speed_source = "day" if yesterday_speed is not None else "price_model"
      # planned: user sets % change from current seller price → derive planned client price
      planned_seller_price = current_price * (1.0 + price_change_pct / 100.0) if current_price > 0 else 0.0
      planned_client_price = planned_seller_price * (1.0 - current_spp / 100.0) if planned_seller_price > 0 else 0.0
      planned_anchor_speed = current_speed if current_speed > 0 else base_speed
      planned_anchor_client = current_client_price if current_client_price > 0 else baseline_client_price
      if planned_anchor_client > 0 and planned_client_price > 0:
        planned_price_delta = (planned_client_price / planned_anchor_client - 1.0)
        planned_factor = max(0.0, 1.0 + elasticity * planned_price_delta)
      else:
        planned_factor = 1.0
      planned_speed = planned_anchor_speed * planned_factor
      stock = stock_by_nm.get(nm_id, 0.0)
      warehouse_stock = _to_float(sku.get("warehouse_stock") or 0)
      potential_orders = planned_speed * forecast_days
      forecast_orders, stockout_date = _simulate_planning_orders(
        stock,
        planned_speed,
        buyout_rate,
        forecast_days,
        return_delay_days,
        forecast_start,
        warehouse_stock=warehouse_stock,
        transit_days=transit_days,
      )
      effective_check = planned_client_price if planned_client_price > 0 else average_check
      forecast_buyouts = forecast_orders * buyout_rate
      forecast_revenue = forecast_buyouts * effective_check
      lost_orders = max(0.0, potential_orders - forecast_orders)
      cogs = _to_float(sku.get("cogs") or 0)
      # ДРР = реклама / фактические выкупы (buyoutSum из funnel_analytics за 14д)
      base_ads_per_day = base_ads_by_nm.get(nm_id, 0.0) / ads_day_count
      actual_buyout_revenue_14d = ads_revenue_by_nm.get(nm_id, 0.0)
      base_revenue_per_day = (actual_buyout_revenue_14d / ads_day_count) if actual_buyout_revenue_14d > 0 else (effective_check * base_speed * buyout_rate)
      base_drr_pct = base_ads_per_day / base_revenue_per_day * 100.0 if base_revenue_per_day > 0 else 0.0
      drr_anchor_pct = max(6.0, min(12.0, base_drr_pct if base_drr_pct > 0 else 6.0))
      organic_floor = max(0.25, min(0.70, 0.55 - ads_elasticity * 0.35))
      # Воронка baseline
      fi = funnel_by_nm.get(nm_id, {})
      base_ctr = fi.get("ctr", 0.0)
      base_cr1 = fi.get("cr1", 0.0)
      base_cr2 = fi.get("cr2", 0.0)
      summary["stock"] += stock
      summary["stock_cost"] += stock * cogs
      summary["stock_value"] += stock * (current_client_price or average_check)
      summary["forecast_orders"] += forecast_orders
      summary["forecast_buyouts"] += forecast_buyouts
      summary["orders_revenue"] += forecast_orders * effective_check
      summary["forecast_revenue"] += forecast_revenue
      summary["lost_orders"] += lost_orders
      summary["lost_revenue"] += lost_orders * buyout_rate * effective_check
      rows_out.append({
        "nmid": nm_id,
        "article": str(marketplace_article_by_nm.get(nm_id) or sku.get("article") or nm_id),
        "name": str(sku.get("name") or ""),
        "subject": str(sku.get("subject") or ""),
        "strategy": str(sku.get("strategy") or ""),
        "revenue_category": abc_category_by_nm.get(nm_id, "C") if total_abc > 0 else str(sku.get("revenue_category") or ""),
        "cogs": cogs,
        "wb_commission_pct": _to_float(sku.get("wb_commission_pct") or 0),
        "stock": stock,
        "warehouse_stock": _to_float(sku.get("warehouse_stock") or 0),
        "baseline_seller_price": baseline_avg_price_by_nm.get(nm_id, 0.0),
        "baseline_spp": baseline_spp_by_nm.get(nm_id, 0.0),
        "baseline_client_check": baseline_client_check_by_nm.get(nm_id, 0.0),
        "base_speed": base_speed,
        "buyout_percent": buyout_rate * 100.0,
        "average_check": average_check,
        "current_price": current_price,
        "current_spp": current_spp,
        "client_price": current_client_price,
        "elasticity": elasticity,
        "ads_elasticity": ads_elasticity,
        "base_ads_per_day": base_ads_per_day,
        "base_revenue_per_day": base_revenue_per_day,
        "base_drr_pct": round(base_drr_pct, 2),
        "drr_anchor_pct": round(drr_anchor_pct, 2),
        "organic_floor": round(organic_floor, 3),
        "base_ctr": round(base_ctr, 2),
        "base_cr1": round(base_cr1, 2),
        "base_cr2": round(base_cr2, 2),
        "current_speed": current_speed,
        "model_current_speed": model_current_speed,
        "current_speed_source": current_speed_source,
        "current_speed_day": current_speed_day if yesterday_speed is not None else "",
        "planned_seller_price": planned_seller_price,
        "planned_price": planned_client_price,
        "planned_speed": planned_speed,
        "price_change_pct": price_change_pct,
        "potential_orders": potential_orders,
        "effective_check": effective_check,
        "forecast_orders": forecast_orders,
        "forecast_buyouts": forecast_buyouts,
        "forecast_revenue": forecast_revenue,
        "lost_orders": lost_orders,
        "lost_revenue": lost_orders * buyout_rate * effective_check,
        "stockout_date": stockout_date,
      })

  rows_out.sort(key=lambda row: (-_to_float(row.get("forecast_revenue")), str(row.get("article") or "")))
  return {
    "summary": summary,
    "rows": rows_out,
    "options": options,
    "baseline_from": baseline_from,
    "baseline_to": baseline_to,
    "forecast_from": forecast_start.isoformat(),
    "forecast_to": target_date.isoformat(),
    "forecast_days": forecast_days,
    "return_delay_days": return_delay_days,
    "transit_days": transit_days,
    "price_change_pct": price_change_pct,
  }

def _compute_ozon_elasticity(sku_day_rows: list) -> dict[str, float | None]:
  """Compute price elasticity per SKU from daily revenue/qty data."""
  import math, statistics as stats_mod
  from collections import defaultdict
  daily: dict[str, list[tuple[float, float]]] = defaultdict(list)
  for r in sku_day_rows:
    qty = _to_float(r["orders_qty"])
    rev = _to_float(r["orders_revenue"])
    if qty > 0 and rev > 0:
      daily[str(r["sku"])].append((rev / qty, qty))
  result: dict[str, float | None] = {}
  for sku, pts in daily.items():
    if len(pts) < 7:
      result[sku] = None
      continue
    prices = [p for p, _ in pts]
    qtys   = [q for _, q in pts]
    avg_p = sum(prices) / len(prices)
    if avg_p <= 0:
      result[sku] = None
      continue
    try:
      pvar = stats_mod.stdev(prices) / avg_p
    except Exception:
      result[sku] = None
      continue
    if pvar < 0.05:
      result[sku] = None
      continue
    try:
      lp = [math.log(p) for p in prices]
      lq = [math.log(q) for q in qtys]
      n = len(lp)
      mean_lp = sum(lp) / n
      mean_lq = sum(lq) / n
      cov = sum((lp[i] - mean_lp) * (lq[i] - mean_lq) for i in range(n)) / n
      var_p = sum((x - mean_lp) ** 2 for x in lp) / n
      corr = cov / max(var_p ** 0.5 * (sum((x - mean_lq) ** 2 for x in lq) / n) ** 0.5, 1e-9)
      result[sku] = round(-abs(corr) * 2.0, 2)
    except Exception:
      result[sku] = None
  return result

def _fetch_ozon_planning(
  forecast_to: str,
  return_delay_days: int = 8,
  transit_days: int = 3,
  subject: str = "",
  revenue_category: str = "",
) -> dict[str, object]:
  target_date = date.fromisoformat(forecast_to)
  return_delay_days = max(1, int(return_delay_days))
  transit_days = max(0, int(transit_days))
  subject = (subject or "").strip()
  revenue_category = (revenue_category or "").strip().upper()

  with _db_connect() as conn:
    # Prefer ozon_sku_day_analytics (per-day data) over period totals
    has_sku_day = conn.execute(
      "SELECT COUNT(*) AS n FROM ozon_sku_day_analytics"
    ).fetchone()["n"] > 0

    if has_sku_day:
      cutoff = (date.today() - timedelta(days=30)).isoformat()
      period_to = (date.today() - timedelta(days=1)).isoformat()
      sku_rows = conn.execute(
        """
        SELECT s.sku, MAX(s.item_name) AS item_name,
          SUM(COALESCE(s.orders_qty, 0)) AS ordered_units,
          SUM(COALESCE(s.delivered_qty, 0)) AS delivered_units,
          SUM(COALESCE(s.orders_revenue, 0)) AS revenue,
          MIN(s.day) AS period_from, MAX(s.day) AS period_to
        FROM ozon_sku_day_analytics s
        WHERE s.day >= ?
        GROUP BY s.sku
        """,
        [cutoff],
      ).fetchall()
      period_from = cutoff
    else:
      latest = conn.execute(
        "SELECT MAX(period_to) AS max_to, MIN(period_from) AS min_from FROM ozon_product_day_analytics"
      ).fetchone()
      if not latest or not latest["max_to"]:
        return {"summary": {}, "rows": [], "forecast_days": 0, "baseline_from": "", "baseline_to": "",
                "options": {"subjects": [], "labels": [], "revenue_categories": []}}
      period_to = str(latest["max_to"])
      period_from = str(latest["min_from"])
      sku_rows = conn.execute(
        """
        SELECT sku, MAX(item_name) AS item_name,
          SUM(COALESCE(ordered_units, 0)) AS ordered_units,
          SUM(COALESCE(delivered_units, 0)) AS delivered_units,
          SUM(COALESCE(revenue, 0)) AS revenue,
          MIN(period_from) AS period_from, MAX(period_to) AS period_to
        FROM ozon_product_day_analytics
        WHERE period_from = ? AND period_to = ?
        GROUP BY sku
        """,
        [period_from, period_to],
      ).fetchall()

    # offer_id lookup from warehouse stocks
    offer_map = {}
    for r in conn.execute(
      "SELECT sku, offer_id FROM ozon_stock_on_warehouses WHERE offer_id IS NOT NULL GROUP BY sku"
    ).fetchall():
      offer_map[str(r["sku"])] = str(r["offer_id"])

    # Marketplace stock (Склад маркетплейса) — from Ozon API
    stock_rows = conn.execute(
      """
      SELECT sku, SUM(COALESCE(free_to_sell_amount, 0) + COALESCE(reserved_amount, 0)) AS stock
      FROM ozon_stock_on_warehouses
      GROUP BY sku
      """
    ).fetchall()
    stock_by_sku = {str(r["sku"]): _to_float(r["stock"]) for r in stock_rows}

    # Our warehouse stock + cogs from SKU file table (Наш склад + Себестоимость).
    sku_columns = _table_columns(conn, "SKU")
    sku_ozon_col = _first_existing(sku_columns, ["SKU Ozon", "ID товара маркетплейса", "ozon_sku", "sku", "SKU"])
    cogs_col = _first_existing(sku_columns, ["себестоимость", "Себестоимость единицы", "cost_price", "cogs"])
    warehouse_col = _first_existing(sku_columns, ["склад", "Склад", "Наш склад", "warehouse_stock"])
    if sku_ozon_col:
      cogs_expr = (
        f"CAST(COALESCE(NULLIF(TRIM({_sql_ident(cogs_col)}),''), '0') AS REAL)"
        if cogs_col else "0"
      )
      warehouse_expr = (
        f"CAST(COALESCE(NULLIF(TRIM({_sql_ident(warehouse_col)}),''), '0') AS REAL)"
        if warehouse_col else "0"
      )
      sku_file_rows = conn.execute(
        (
          f"SELECT CAST({_sql_ident(sku_ozon_col)} AS TEXT) AS ozon_sku, "
          f"{cogs_expr} AS cogs, "
          f"{warehouse_expr} AS warehouse_stock "
          "FROM SKU "
          f"WHERE TRIM(COALESCE({_sql_ident(sku_ozon_col)}, '')) != ''"
        )
      ).fetchall()
    else:
      sku_file_rows = []
    cogs_by_sku: dict[str, float] = {}
    wh_stock_by_sku: dict[str, float] = {}
    for r in sku_file_rows:
      k = str(r["ozon_sku"]).strip()
      cogs_by_sku[k] = _to_float(r["cogs"])
      wh_stock_by_sku[k] = _to_float(r["warehouse_stock"])

    # Per-day rows for elasticity calculation (need daily price variation per SKU)
    sku_day_rows_for_elasticity = conn.execute(
      "SELECT sku, orders_qty, orders_revenue FROM ozon_sku_day_analytics "
      "WHERE day BETWEEN ? AND ? AND orders_qty > 0",
      [period_from, period_to],
    ).fetchall()

    # Per-SKU SPP: avg_spp from ozon_sku_day_analytics
    sku_spp: dict[str, float] = {}
    for r in conn.execute(
      "SELECT sku, AVG(NULLIF(avg_spp, 0)) AS spp FROM ozon_sku_day_analytics "
      "WHERE day BETWEEN ? AND ? AND avg_spp > 0 GROUP BY sku",
      [period_from, period_to],
    ).fetchall():
      if r["spp"]:
        sku_spp[str(r["sku"])] = round(_to_float(r["spp"]), 1)

    # Per-SKU funnel: hits_view/pdp/tocart from ozon_sku_day_analytics
    sku_funnel: dict[str, dict[str, float]] = {}
    for r in conn.execute(
      "SELECT sku, SUM(COALESCE(hits_view,0)) v, SUM(COALESCE(hits_view_pdp,0)) pdp, "
      "SUM(COALESCE(hits_tocart_pdp,0)) cart, SUM(COALESCE(orders_qty,0)) orders "
      "FROM ozon_sku_day_analytics WHERE day BETWEEN ? AND ? GROUP BY sku",
      [period_from, period_to],
    ).fetchall():
      v, pdp, cart, orders_q = (r["v"] or 0, r["pdp"] or 0, r["cart"] or 0, r["orders"] or 0)
      sku_funnel[str(r["sku"])] = {
        "ctr": round(pdp  / max(v,     1) * 100, 2),
        "cr1": round(cart / max(pdp,   1) * 100, 2),
        "cr2": round(orders_q / max(cart, 1) * 100, 2),
      }

    # Per-SKU ad spend from CPO/CPC orders attribution
    sku_ad_spend: dict[str, float] = {}
    for r in conn.execute(
      "SELECT sku, SUM(ad_spend) AS spend FROM ozon_sku_period_ad_spend "
      "WHERE period_from >= ? AND period_to <= ? GROUP BY sku",
      [period_from, period_to],
    ).fetchall():
      sku_ad_spend[str(r["sku"])] = _to_float(r["spend"])

    # Cabinet-level commission and logistics rates from Finance API (no per-SKU data from Finance API)
    fin_row = conn.execute(
      "SELECT SUM(ABS(sale_commission)) comm, SUM(ABS(delivery_charge)) deliv, "
      "SUM(accruals_for_sale) accruals FROM ozon_daily_summary WHERE day BETWEEN ? AND ?",
      [period_from, period_to],
    ).fetchone()
    _accruals = max(_to_float(fin_row["accruals"] if fin_row else 0), 1)
    commission_pct = round(_to_float(fin_row["comm"] if fin_row else 0) / _accruals * 100, 2)
    delivery_pct   = round(_to_float(fin_row["deliv"] if fin_row else 0) / _accruals * 100, 2)

  if not sku_rows:
    return {"summary": {}, "rows": [], "forecast_days": 0,
            "baseline_from": period_from, "baseline_to": period_to,
            "options": {"subjects": [], "labels": [], "revenue_categories": []}}

  period_days = max(1, (date.fromisoformat(period_to) - date.fromisoformat(period_from)).days + 1)
  latest_day = date.fromisoformat(period_to)
  forecast_days = max(0, (target_date - latest_day).days)
  forecast_start = latest_day + timedelta(days=1)

  # ABC classification by revenue
  rev_list = sorted(
    [(str(r["sku"]), _to_float(r["revenue"])) for r in sku_rows],
    key=lambda x: x[1], reverse=True
  )
  total_rev = sum(v for _, v in rev_list) or 1.0
  cumul = 0.0
  abc_by_sku: dict[str, str] = {}
  for sku, rev in rev_list:
    cumul += rev
    if cumul / total_rev <= 0.80:
      abc_by_sku[sku] = "A"
    elif cumul / total_rev <= 0.95:
      abc_by_sku[sku] = "B"
    else:
      abc_by_sku[sku] = "C"

  all_subjects = sorted({str(r["item_name"] or "") for r in sku_rows if r["item_name"]})

  # Compute elasticity per SKU using per-day rows in the baseline period
  elasticity_by_sku = _compute_ozon_elasticity(sku_day_rows_for_elasticity)

  rows_out = []
  total_forecast = 0.0; total_lost = 0.0
  for row in sku_rows:
    sku = str(row["sku"])
    item_name = str(row["item_name"] or "")
    abc = abc_by_sku.get(sku, "C")
    if subject and item_name != subject:
      continue
    if revenue_category and abc != revenue_category:
      continue

    ordered = _to_float(row["ordered_units"])
    delivered = _to_float(row["delivered_units"])
    revenue = _to_float(row["revenue"])
    marketplace_stock = stock_by_sku.get(sku, 0.0)
    our_stock = wh_stock_by_sku.get(sku, 0.0)
    cogs = cogs_by_sku.get(sku, 0.0)
    daily_demand = ordered / period_days
    buyout_rate = min(1.0, max(0.0, delivered / ordered if ordered else 0.8))
    avg_rev_per_order = revenue / ordered if ordered else 0.0
    offer_id = offer_map.get(sku, sku)

    # Use effective_stock = marketplace + our warehouse for lost-orders simulation
    effective_stock = marketplace_stock + our_stock
    forecast_orders, stockout_date = _simulate_planning_orders(
      stock=effective_stock,
      daily_demand=daily_demand,
      buyout_rate=buyout_rate,
      forecast_days=forecast_days,
      return_delay_days=return_delay_days,
      start_day=forecast_start,
    )
    max_orders = daily_demand * forecast_days
    lost_orders = max(0.0, max_orders - forecast_orders)
    forecast_revenue = forecast_orders * avg_rev_per_order
    lost_revenue = lost_orders * avg_rev_per_order
    total_forecast += forecast_orders; total_lost += lost_orders

    # Per-SKU metrics
    spp = sku_spp.get(sku, 0.0)
    funnel = sku_funnel.get(sku, {"ctr": 0.0, "cr1": 0.0, "cr2": 0.0})
    ad_spend_sku = sku_ad_spend.get(sku, 0.0)
    drr_pct = round(ad_spend_sku / max(revenue, 1) * 100, 2) if ad_spend_sku else 0.0

    rows_out.append({
      "nmid": sku,
      "article": offer_id,
      "subject": item_name,
      "strategy": "",
      "revenue_category": abc,
      "stock": int(round(marketplace_stock)),
      "warehouse_stock": int(round(our_stock)),
      "ordered_units": int(round(ordered)),
      "base_speed": round(daily_demand, 2),
      "current_speed": round(daily_demand, 2),
      "buyout_percent": round(buyout_rate * 100, 1),
      "forecast_orders": int(round(forecast_orders)),
      "forecast_revenue": int(round(forecast_revenue)),
      "lost_orders": int(round(lost_orders)),
      "lost_revenue": int(round(lost_revenue)),
      "stockout_date": stockout_date,
      "base_drr_pct": drr_pct,
      "base_ctr": funnel["ctr"],
      "base_cr1": funnel["cr1"],
      "base_cr2": funnel["cr2"],
      "elasticity": elasticity_by_sku.get(sku),
      "ads_elasticity": None,
      "cogs": round(cogs, 2),
      "wb_commission_pct": commission_pct,
      "delivery_charge_pct": delivery_pct,
      "baseline_seller_price": round(avg_rev_per_order, 2),
      "baseline_spp": spp,
      "baseline_client_check": round(avg_rev_per_order, 2),
      "current_price": round(avg_rev_per_order, 2),
      "current_spp": spp,
      "client_price": round(avg_rev_per_order, 2),
      "average_check": round(avg_rev_per_order, 2),
    })

  rows_out.sort(key=lambda r: r["lost_revenue"], reverse=True)

  return {
    "rows": rows_out,
    "options": {
      "subjects": all_subjects,
      "labels": [],
      "revenue_categories": ["A", "B", "C"],
    },
    "summary": {
      "total_forecast_orders": int(round(total_forecast)),
      "total_lost_orders": int(round(total_lost)),
    },
    "forecast_days": forecast_days,
    "baseline_from": period_from,
    "baseline_to": period_to,
    "forecast_from": forecast_start.isoformat(),
    "forecast_to": target_date.isoformat(),
    "return_delay_days": return_delay_days,
  }

def _sql_ident(name: str) -> str:
  return '"' + name.replace('"', '""') + '"'

def _first_existing(columns: set[str], candidates: list[str]) -> str | None:
  for candidate in candidates:
    if candidate in columns:
      return candidate
  return None

def _table_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
  safe_table = table_name.replace('"', '""')
  rows = conn.execute(f'PRAGMA table_info("{safe_table}")').fetchall()
  return {str(row["name"]) for row in rows}

def _sum_expr(columns: set[str], candidates: list[str]) -> tuple[str, list[str]]:
  used = [candidate for candidate in candidates if candidate in columns]
  if not used:
    return "0", []
  expr = " + ".join(f"COALESCE(SUM({_sql_ident(col)}), 0)" for col in used)
  return expr, used

def _db_scope_filter(conn: sqlite3.Connection, table_name: str) -> tuple[str, list[str]]:
  return "", []

def _ensure_preliminary_economics_tables(conn: sqlite3.Connection) -> None:
  conn.execute(
    """
    CREATE TABLE IF NOT EXISTS app_settings (
      key TEXT PRIMARY KEY,
      value TEXT NOT NULL,
      updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    )
    """
  )
  conn.execute(
    """
    CREATE TABLE IF NOT EXISTS preliminary_order_economics (
      sku TEXT NOT NULL,
      date TEXT NOT NULL,
      orders_count REAL NOT NULL,
      orders_sum REAL NOT NULL,
      commission_rub REAL NOT NULL,
      acquiring_rub REAL NOT NULL,
      advertising_rub REAL NOT NULL,
      additional_expenses_rub REAL NOT NULL,
      preliminary_profit_rub REAL NOT NULL,
      additional_rate REAL NOT NULL,
      updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
      PRIMARY KEY (sku, date)
    )
    """
  )

def _get_or_set_preliminary_additional_rate(
  conn: sqlite3.Connection,
  columns: set[str],
  sales_column: str,
) -> tuple[float, list[str]]:
  setting_key = "preliminary_additional_rate"
  stored = conn.execute("SELECT value FROM app_settings WHERE key = ?", (setting_key,)).fetchone()
  if stored:
    try:
      rate = max(0.0, float(stored["value"]))
      return rate, []
    except ValueError:
      pass

  additional_candidates = [
    "Логистика",
    "Хранение",
    "Приемка",
    "Штрафы",
    "Удержания",
    "Доплаты",
    "НДС WB",
    "Платная приемка",
    "Платное хранение",
    "Платные услуги",
  ]
  additional_expr, used_columns = _sum_expr(columns, additional_candidates)
  sales_expr = f"COALESCE(SUM({_sql_ident(sales_column)}), 0)"
  row = conn.execute(
    f"SELECT CASE WHEN {sales_expr} > 0 THEN ({additional_expr}) / {sales_expr} ELSE 0 END AS rate "
    "FROM finance_article_day_detail"
  ).fetchone()
  rate = _to_float(row["rate"] if row else 0.0)
  if rate <= 0:
    rate = 0.05

  conn.execute(
    """
    INSERT INTO app_settings(key, value, updated_at)
    VALUES(?, ?, CURRENT_TIMESTAMP)
    ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=CURRENT_TIMESTAMP
    """,
    (setting_key, f"{rate:.8f}"),
  )
  return rate, used_columns

def _fetch_preliminary_economics(
  date_from: str,
  date_to: str,
  article_query: str = "",
  buyout_percent: float = 30.0,
  aggregate_by_period: bool = False,
) -> dict[str, object]:
  article_query = article_query.strip()
  buyout_percent = max(0.0, min(100.0, buyout_percent))
  buyout_factor = buyout_percent / 100.0
  with _db_connect() as conn:
    _ensure_preliminary_economics_tables(conn)
    columns = _table_columns(conn, "finance_article_day_detail")
    ctx = _scope_context(conn)

    sku_col = _first_existing(columns, ["Артикул", "SKU"])
    date_col = _first_existing(columns, ["Дата", "date"])
    orders_count_col = _first_existing(columns, ["Продажи, шт", "Заказы, шт", "Заказано, шт", "Количество заказов"])
    sales_col = _first_existing(columns, ["Наша цена", "Продажи по нашей цене", "Продажи по нашей цене, р"])
    if not sku_col or not date_col or not sales_col:
      return {"rows": [], "additional_rate": "0.00%", "expense_components": []}

    commission_expr, _ = _sum_expr(columns, ["Комиссия WB", "Вознаграждение WB"])
    ads_expr, _ = _sum_expr(columns, ["Реклама", "Рекламные расходы"])
    additional_rate, used_additional_cols = _get_or_set_preliminary_additional_rate(conn, columns, sales_col)

    where = [f"{_sql_ident(date_col)} >= ?", f"{_sql_ident(date_col)} <= ?"]
    params: list[str] = [date_from, date_to]
    if article_query:
      where.append(f"{_sql_ident(sku_col)} = ?")
      params.append(article_query)
    orders_count_expr = f"COALESCE(SUM({_sql_ident(orders_count_col)}), 0)" if orders_count_col else "0"
    select_date = f"{_sql_ident(date_col)} AS date, " if not aggregate_by_period else ""
    group_by = "GROUP BY sku, date" if not aggregate_by_period else "GROUP BY sku"
    order_by = "ORDER BY date DESC, sku ASC" if not aggregate_by_period else "ORDER BY sku ASC"
    sql = (
      "SELECT "
      f"{_sql_ident(sku_col)} AS sku, "
      f"{select_date}"
      f"{orders_count_expr} AS orders_count, "
      f"COALESCE(SUM({_sql_ident(sales_col)}), 0) AS orders_sum, "
      f"{commission_expr} AS commission_rub, "
      f"{ads_expr} AS advertising_rub "
      "FROM finance_article_day_detail "
      f"WHERE {' AND '.join(where)} "
      f"{group_by} "
      f"{order_by} "
      "LIMIT 10000"
    )
    base_rows = conn.execute(sql, params).fetchall()

    funnel_where = ["date >= ?", "date <= ?"]
    funnel_params: list[str] = [date_from, date_to]
    if article_query:
      funnel_where.append("supplierArticle = ?")
      funnel_params.append(article_query)
    funnel_select_date = "date, " if not aggregate_by_period else ""
    funnel_group_by = "GROUP BY date, supplierArticle" if not aggregate_by_period else "GROUP BY supplierArticle"
    funnel_rows = conn.execute(
      (
        f"SELECT {funnel_select_date} supplierArticle AS sku, "
        "SUM(COALESCE(CAST(orderCount AS REAL), 0)) AS funnel_orders_count "
        "FROM funnel_analytics "
        f"WHERE {' AND '.join(funnel_where)} "
        f"{funnel_group_by}"
      ),
      funnel_params,
    ).fetchall()
    if aggregate_by_period:
      funnel_orders_map = {
        str(row["sku"]): _to_float(row["funnel_orders_count"])
        for row in funnel_rows
      }
    else:
      funnel_orders_map = {
        (str(row["sku"]), str(row["date"])): _to_float(row["funnel_orders_count"])
        for row in funnel_rows
      }

    payload: list[dict[str, str]] = []
    to_store: list[tuple[object, ...]] = []
    for row in base_rows:
      sku = str(row["sku"])
      row_date = str(row["date"]) if not aggregate_by_period else f"{date_from}..{date_to}"
      finance_orders_count = _to_float(row["orders_count"])
      if aggregate_by_period:
        orders_count = funnel_orders_map.get(sku, finance_orders_count)
      else:
        orders_count = funnel_orders_map.get((sku, row_date), finance_orders_count)

      raw_orders_sum = _to_float(row["orders_sum"])
      raw_commission = _to_float(row["commission_rub"])
      advertising = _to_float(row["advertising_rub"])
      orders_sum = raw_orders_sum * buyout_factor
      commission = raw_commission * buyout_factor
      acquiring = (raw_orders_sum * 0.02) * buyout_factor
      additional_expenses = (raw_orders_sum * additional_rate) * buyout_factor
      preliminary_profit = orders_sum - commission - acquiring - advertising - additional_expenses
      ad_pct = (advertising / orders_sum * 100.0) if orders_sum else 0.0
      margin_pct = (preliminary_profit / orders_sum * 100.0) if orders_sum else 0.0

      date_key = "Период" if aggregate_by_period else "Дата"
      payload.append(
        {
          "Артикул / SKU": sku,
          date_key: row_date,
          "Количество заказов": str(int(round(orders_count))),
          "Сумма заказов": _format_metric("Сумма заказов", orders_sum),
          "% выкупа": _format_percent(buyout_percent),
          "Комиссия, ₽": _format_metric("Комиссия, ₽", commission),
          "Эквайринг, ₽": _format_metric("Эквайринг, ₽", acquiring),
          "Реклама, ₽": _format_metric("Реклама, ₽", advertising),
          "% рекламы": _format_percent(ad_pct),
          "Дополнительные расходы, ₽": _format_metric("Дополнительные расходы, ₽", additional_expenses),
          "Предварительная прибыль, ₽": _format_metric("Предварительная прибыль, ₽", preliminary_profit),
          "% маржинальности": _format_percent(margin_pct),
        }
      )
      if not aggregate_by_period:
        to_store.append(
          (
            sku,
            row_date,
            orders_count,
            orders_sum,
            commission,
            acquiring,
            advertising,
            additional_expenses,
            preliminary_profit,
            additional_rate,
          )
        )

    if to_store:
      conn.executemany(
        """
        INSERT INTO preliminary_order_economics(
          sku, date, orders_count, orders_sum, commission_rub, acquiring_rub,
          advertising_rub, additional_expenses_rub, preliminary_profit_rub, additional_rate, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(sku, date) DO UPDATE SET
          orders_count=excluded.orders_count,
          orders_sum=excluded.orders_sum,
          commission_rub=excluded.commission_rub,
          acquiring_rub=excluded.acquiring_rub,
          advertising_rub=excluded.advertising_rub,
          additional_expenses_rub=excluded.additional_expenses_rub,
          preliminary_profit_rub=excluded.preliminary_profit_rub,
          additional_rate=excluded.additional_rate,
          updated_at=CURRENT_TIMESTAMP
        """,
        to_store,
      )
    conn.commit()

  return {
    "rows": payload,
    "additional_rate": f"{additional_rate * 100:.2f}%",
    "buyout_percent": _format_percent(buyout_percent),
    "expense_components": used_additional_cols,
  }

# ── Tasks Kanban page ────────────────────────────────────────────────────────
TASKS_HTML = """\
<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Задачи</title>
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:var(--bg,#f1f5f9);color:var(--ink,#1e293b);min-height:100vh}}
.report-nav{{padding:8px 16px;background:var(--card,#fff);border-bottom:1px solid #e2e8f0;display:flex;gap:8px;flex-wrap:wrap;align-items:center}}
.page-wrap{{max-width:1400px;margin:0 auto;padding:16px}}

/* create form */
.create-card{{background:var(--card,#fff);border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.08);padding:16px 20px;margin-bottom:20px}}
.create-card h2{{font-size:14px;font-weight:600;color:var(--muted,#64748b);text-transform:uppercase;letter-spacing:.05em;margin-bottom:12px}}
.form-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}
.form-grid label{{display:flex;flex-direction:column;gap:4px;font-size:13px;font-weight:500;color:var(--ink,#1e293b)}}
.form-grid input,.form-grid textarea{{border:1px solid #e2e8f0;border-radius:6px;padding:7px 10px;font-size:13px;font-family:inherit;resize:vertical;background:#fff;color:var(--ink,#1e293b);transition:border-color .15s}}
.form-grid input:focus,.form-grid textarea:focus{{outline:none;border-color:#0f766e}}
.form-grid textarea{{min-height:52px}}
.form-actions{{display:flex;justify-content:flex-end;margin-top:10px}}
.btn-add{{background:#0f766e;color:#fff;border:none;border-radius:7px;padding:8px 20px;font-size:13px;font-weight:600;cursor:pointer;transition:background .15s}}
.btn-add:hover{{background:#115e59}}

/* kanban board */
.board{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;align-items:start}}
@media(max-width:900px){{.board{{grid-template-columns:repeat(2,1fr)}}}}
@media(max-width:540px){{.board{{grid-template-columns:1fr}}}}

.column{{background:var(--col-bg,#f8fafc);border-radius:10px;padding:10px;min-height:200px;transition:background .15s}}
.column.drag-over{{background:#e0f2f1}}
.col-header{{display:flex;align-items:center;gap:6px;padding:4px 2px 10px;font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.06em}}
.col-header .dot{{width:10px;height:10px;border-radius:50%;flex-shrink:0}}
.col-count{{margin-left:auto;background:#e2e8f0;color:#475569;font-size:11px;font-weight:700;border-radius:10px;padding:1px 7px}}

/* cards */
.card{{background:var(--card,#fff);border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.08);padding:10px 12px;margin-bottom:8px;cursor:pointer;position:relative;transition:box-shadow .15s,opacity .15s}}
.card:hover{{box-shadow:0 3px 8px rgba(0,0,0,.12)}}
.card.dragging{{opacity:.45}}
.card-title{{font-size:13px;font-weight:600;line-height:1.35;padding-right:24px;color:var(--ink,#1e293b)}}
.card-meta{{display:flex;align-items:center;gap:6px;margin-top:6px;flex-wrap:wrap}}
.due-badge{{font-size:11px;color:var(--muted,#64748b);background:#f1f5f9;border-radius:4px;padding:1px 6px}}
.due-badge.overdue{{background:#fee2e2;color:#b91c1c}}

.card-details{{margin-top:10px;padding-top:10px;border-top:1px solid #f1f5f9;display:none;font-size:12px;color:var(--ink,#334155)}}
.card.expanded .card-details{{display:block}}
.detail-row{{margin-bottom:7px}}
.detail-label{{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.05em;color:var(--muted,#64748b);margin-bottom:2px}}
.detail-val{{white-space:pre-wrap;word-break:break-word}}
.created-at{{font-size:11px;color:#94a3b8;margin-top:8px}}

.card-edit-btn{{position:absolute;top:8px;right:8px;background:none;border:none;cursor:pointer;padding:2px;border-radius:4px;opacity:.45;transition:opacity .15s;line-height:1;display:flex;align-items:center}}
.card:hover .card-edit-btn{{opacity:.9}}
.card-edit-btn svg{{width:14px;height:14px;fill:none;stroke:#475569;stroke-width:1.8}}

/* done column muted */
.col-done .card{{opacity:.7}}
.col-done .card:hover{{opacity:.95}}
.col-done .card-title{{text-decoration:line-through;color:#94a3b8}}

/* modal */
.modal-overlay{{position:fixed;inset:0;background:rgba(0,0,0,.35);z-index:1000;display:flex;align-items:center;justify-content:center;padding:16px}}
.modal{{background:#fff;border-radius:12px;width:100%;max-width:460px;box-shadow:0 8px 32px rgba(0,0,0,.18);padding:24px}}
.modal h3{{font-size:15px;font-weight:700;margin-bottom:16px}}
.modal-form label{{display:flex;flex-direction:column;gap:4px;font-size:13px;font-weight:500;margin-bottom:10px}}
.modal-form input,.modal-form textarea,.modal-form select{{border:1px solid #e2e8f0;border-radius:6px;padding:7px 10px;font-size:13px;font-family:inherit;resize:vertical;width:100%;color:var(--ink,#1e293b)}}
.modal-form input:focus,.modal-form textarea:focus,.modal-form select:focus{{outline:none;border-color:#0f766e}}
.modal-form textarea{{min-height:56px}}
.modal-actions{{display:flex;justify-content:space-between;align-items:center;margin-top:16px}}
.btn-save{{background:#0f766e;color:#fff;border:none;border-radius:7px;padding:8px 18px;font-size:13px;font-weight:600;cursor:pointer}}
.btn-save:hover{{background:#115e59}}
.btn-cancel{{background:#f1f5f9;color:#475569;border:none;border-radius:7px;padding:8px 14px;font-size:13px;font-weight:500;cursor:pointer}}
.btn-cancel:hover{{background:#e2e8f0}}
.btn-delete{{background:none;border:none;cursor:pointer;font-size:12px;color:#ef4444;font-weight:500;padding:4px 0}}
.btn-delete:hover{{text-decoration:underline}}
</style>
</head>
<body>
<div class="report-nav">{report_nav}</div>
<div class="page-wrap">
  <div class="create-card">
    <h2>Новая задача</h2>
    <div class="form-grid">
      <label>Что сделать<input id="f-title" type="text" placeholder="Заголовок задачи" required></label>
      <label>Зачем<textarea id="f-why" placeholder="Цель / мотивация" required></textarea></label>
      <label>Какой результат<textarea id="f-result" placeholder="Ожидаемый результат" required></textarea></label>
      <label>Когда (срок)<input id="f-due" type="date" required></label>
    </div>
    <div class="form-actions">
      <button class="btn-add" onclick="submitCreate()">+ Добавить задачу</button>
    </div>
  </div>

  <div class="board">
    <div class="column" id="col-backlog" data-status="backlog" ondragover="onDragOver(event)" ondrop="onDrop(event)">
      <div class="col-header"><span class="dot" style="background:#94a3b8"></span>Бэклог<span class="col-count" id="cnt-backlog">0</span></div>
      <div class="col-cards" id="cards-backlog"></div>
    </div>
    <div class="column" id="col-in_progress" data-status="in_progress" ondragover="onDragOver(event)" ondrop="onDrop(event)">
      <div class="col-header"><span class="dot" style="background:#f97316"></span>В работе<span class="col-count" id="cnt-in_progress">0</span></div>
      <div class="col-cards" id="cards-in_progress"></div>
    </div>
    <div class="column" id="col-review" data-status="review" ondragover="onDragOver(event)" ondrop="onDrop(event)">
      <div class="col-header"><span class="dot" style="background:#3b82f6"></span>На проверке<span class="col-count" id="cnt-review">0</span></div>
      <div class="col-cards" id="cards-review"></div>
    </div>
    <div class="column col-done" id="col-done" data-status="done" ondragover="onDragOver(event)" ondrop="onDrop(event)">
      <div class="col-header"><span class="dot" style="background:#22c55e"></span>Выполнено<span class="col-count" id="cnt-done">0</span></div>
      <div class="col-cards" id="cards-done"></div>
    </div>
  </div>
</div>

<!-- Edit modal -->
<div class="modal-overlay" id="modal" style="display:none" onclick="onOverlayClick(event)">
  <div class="modal">
    <h3>Редактировать задачу</h3>
    <div class="modal-form">
      <label>Что сделать<input id="m-title" type="text" required></label>
      <label>Зачем<textarea id="m-why" required></textarea></label>
      <label>Какой результат<textarea id="m-result" required></textarea></label>
      <label>Когда<input id="m-due" type="date" required></label>
      <label>Статус
        <select id="m-status">
          <option value="backlog">Бэклог</option>
          <option value="in_progress">В работе</option>
          <option value="review">На проверке</option>
          <option value="done">Выполнено</option>
        </select>
      </label>
    </div>
    <div class="modal-actions">
      <button class="btn-delete" onclick="confirmDelete()">Удалить задачу</button>
      <div style="display:flex;gap:8px">
        <button class="btn-cancel" onclick="closeModal()">Отмена</button>
        <button class="btn-save" onclick="saveModal()">Сохранить</button>
      </div>
    </div>
  </div>
</div>

<script>
let tasks = [];
let editId = null;
let dragId = null;

const today = new Date().toISOString().slice(0,10);

async function loadTasks() {{
  const r = await fetch('/api/tasks');
  const d = await r.json();
  tasks = d.tasks || [];
  renderBoard();
}}

function renderBoard() {{
  const cols = ['backlog','in_progress','review','done'];
  cols.forEach(s => {{
    const list = tasks.filter(t => t.status === s);
    document.getElementById('cnt-' + s).textContent = list.length;
    document.getElementById('cards-' + s).innerHTML = list.map(cardHtml).join('');
  }});
  bindCards();
}}

function fmtDate(iso) {{
  if (!iso) return '';
  const [y,m,d] = iso.split('-');
  return d + '.' + m + '.' + y;
}}

function cardHtml(t) {{
  const over = t.due_date && t.due_date < today;
  const dueCls = over ? 'due-badge overdue' : 'due-badge';
  return `<div class="card" id="card-${{t.id}}" draggable="true" data-id="${{t.id}}">
    <div class="card-title">${{esc(t.title)}}</div>
    <div class="card-meta">
      <span class="${{dueCls}}">${{fmtDate(t.due_date)}}</span>
    </div>
    <div class="card-details">
      <div class="detail-row"><div class="detail-label">Зачем</div><div class="detail-val">${{esc(t.why)}}</div></div>
      <div class="detail-row"><div class="detail-label">Какой результат</div><div class="detail-val">${{esc(t.result)}}</div></div>
      <div class="created-at">Создана: ${{t.created_at ? t.created_at.slice(0,16) : ''}}</div>
    </div>
    <button class="card-edit-btn" onclick="openEdit(event,${{t.id}})" title="Редактировать">
      <svg viewBox="0 0 24 24"><path d="M11 4H4a2 2 0 0 0-2 2v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2v-7"/><path d="M18.5 2.5a2.121 2.121 0 0 1 3 3L12 15l-4 1 1-4 9.5-9.5z"/></svg>
    </button>
  </div>`;
}}

function esc(s) {{
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}}

function bindCards() {{
  document.querySelectorAll('.card').forEach(el => {{
    el.addEventListener('click', onCardClick);
    el.addEventListener('dragstart', onDragStart);
    el.addEventListener('dragend', onDragEnd);
  }});
}}

function onCardClick(e) {{
  if (e.target.closest('.card-edit-btn')) return;
  e.currentTarget.classList.toggle('expanded');
}}

// Drag and drop
function onDragStart(e) {{
  dragId = parseInt(e.currentTarget.dataset.id);
  e.currentTarget.classList.add('dragging');
  e.dataTransfer.effectAllowed = 'move';
}}
function onDragEnd(e) {{
  e.currentTarget.classList.remove('dragging');
}}
function onDragOver(e) {{
  e.preventDefault();
  e.currentTarget.classList.add('drag-over');
}}
document.querySelectorAll('.column').forEach(col => {{
  col.addEventListener('dragleave', () => col.classList.remove('drag-over'));
}});
async function onDrop(e) {{
  e.preventDefault();
  const col = e.currentTarget;
  col.classList.remove('drag-over');
  const newStatus = col.dataset.status;
  if (!dragId) return;
  const task = tasks.find(t => t.id === dragId);
  if (!task || task.status === newStatus) return;
  await fetch('/api/tasks/update', {{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{id:dragId,status:newStatus}})}});
  await loadTasks();
}}

// Create
async function submitCreate() {{
  const title = document.getElementById('f-title').value.trim();
  const why = document.getElementById('f-why').value.trim();
  const result = document.getElementById('f-result').value.trim();
  const due = document.getElementById('f-due').value.trim();
  if (!title || !why || !result || !due) {{
    alert('Все поля обязательны для заполнения');
    return;
  }}
  await fetch('/api/tasks', {{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{title,why,result,due_date:due}})}});
  document.getElementById('f-title').value='';
  document.getElementById('f-why').value='';
  document.getElementById('f-result').value='';
  document.getElementById('f-due').value='';
  await loadTasks();
}}

// Edit modal
function openEdit(e, id) {{
  e.stopPropagation();
  const t = tasks.find(x => x.id === id);
  if (!t) return;
  editId = id;
  document.getElementById('m-title').value = t.title;
  document.getElementById('m-why').value = t.why;
  document.getElementById('m-result').value = t.result;
  document.getElementById('m-due').value = t.due_date;
  document.getElementById('m-status').value = t.status;
  document.getElementById('modal').style.display = 'flex';
}}
function closeModal() {{
  document.getElementById('modal').style.display = 'none';
  editId = null;
}}
function onOverlayClick(e) {{
  if (e.target === document.getElementById('modal')) closeModal();
}}
async function saveModal() {{
  const title = document.getElementById('m-title').value.trim();
  const why = document.getElementById('m-why').value.trim();
  const result = document.getElementById('m-result').value.trim();
  const due_date = document.getElementById('m-due').value.trim();
  const status = document.getElementById('m-status').value;
  if (!title || !why || !result || !due_date) {{
    alert('Все поля обязательны для заполнения');
    return;
  }}
  await fetch('/api/tasks/update', {{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{id:editId,title,why,result,due_date,status}})}});
  closeModal();
  await loadTasks();
}}
async function confirmDelete() {{
  if (!confirm('Удалить задачу?')) return;
  await fetch('/api/tasks/delete', {{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{id:editId}})}});
  closeModal();
  await loadTasks();
}}

loadTasks();
</script>
</body>
</html>
"""

# ── OZON HTML templates (from mipao) ─────────────────────────────────────
OZON_BUYOUT_ORDER_DAY_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{page_title}</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0f766e; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #e0f2fe, transparent 240px), var(--bg); }}
    .wrap {{ width: calc(100% - 32px); max-width: 3000px; margin: 0 auto; padding: 10px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .76rem; margin-bottom: 3px; font-weight: 700; }}
    input, select {{ padding: 7px 8px; border: 1px solid #cbd5e1; border-radius: 7px; font-size: 12px; min-width: 160px; background: #fff; }}
    button {{ padding: 8px 12px; border: 0; border-radius: 7px; background: var(--accent); color: #fff; font-weight: 700; cursor: pointer; font-size: 12px; }}
    .quick-range {{ display: flex; gap: 6px; align-items: center; flex-wrap: wrap; }}
    .quick-range button {{ padding: 7px 10px; background: #e6f4f1; color: var(--accent); }}
    .meta {{ margin-top: 8px; color: var(--muted); font-size: .78rem; }}
    .workspace {{ --articles-width: 332px; display: grid; grid-template-columns: var(--articles-width) minmax(0, 1fr); gap: 12px; align-items: start; margin-top: 10px; }}
    .workspace.articles-collapsed {{ --articles-width: 44px; }}
    .main-pane {{ display: flex; flex-direction: column; gap: 12px; min-width: 0; }}
    .tbl {{ --graph-col-width: 58px; --metric-col-width: 240px; --total-col-width: 120px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow-x: auto; overflow-y: clip; -webkit-overflow-scrolling: touch; }}
    .articles {{ position: sticky; top: 8px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: hidden; max-height: calc(100vh - 18px); display: flex; flex-direction: column; min-width: 0; transition: width .18s ease, box-shadow .18s ease, border-color .18s ease; }}
    .articles-inner {{ display: flex; flex-direction: column; min-height: 0; height: 100%; background: #fff; }}
    .articles-head {{ padding: 7px 8px; border-bottom: 1px solid #eef2f7; display: flex; justify-content: space-between; gap: 8px; align-items: center; }}
    .articles-title {{ font-weight: 800; font-size: .78rem; }}
    .articles-actions {{ display: flex; gap: 8px; }}
    .articles-actions button {{ padding: 4px 6px; border-radius: 6px; font-size: .68rem; background: #e6f4f1; color: var(--accent); }}
    .articles-mini {{ display: none; align-items: center; justify-content: center; gap: 6px; height: 100%; min-height: 260px; padding: 8px 0; writing-mode: vertical-rl; transform: rotate(180deg); color: var(--accent); font-weight: 800; font-size: .72rem; letter-spacing: .08em; background: linear-gradient(180deg, #f0fdfa, #ecfeff); }}
    .articles-mini button {{ writing-mode: horizontal-tb; transform: rotate(180deg); margin: 0; width: 24px; height: 24px; padding: 0; border-radius: 999px; background: #0f766e; color: #fff; font-size: .9rem; line-height: 1; }}
    .subject-filter {{ padding: 6px 8px; border-bottom: 1px solid #eef2f7; }}
    .subject-filter label {{ font-size: .68rem; margin-bottom: 2px; }}
    .subject-filter select {{ width: 100%; min-width: 0; padding: 5px 6px; font-size: .72rem; }}
    .article-list {{ overflow: auto; user-select: none; }}
    .article-table {{ width: 100%; min-width: 0; border-collapse: collapse; table-layout: fixed; font-size: 10px; }}
    .article-table th, .article-table td {{ border-bottom: 1px solid #eef2f7; padding: 3px; line-height: 1.15; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
    .article-table th {{ position: sticky; top: 0; z-index: 1; background: #f8fafc; color: #475569; font-size: .62rem; cursor: pointer; }}
    .article-table th:first-child, .article-table td:first-child {{ text-align: left; width: 44%; }}
    .article-table th:nth-child(2), .article-table td:nth-child(2) {{ text-align: right; width: 12%; }}
    .article-table th:nth-child(3), .article-table td:nth-child(3) {{ text-align: right; width: 12%; }}
    .article-table th:nth-child(4), .article-table td:nth-child(4) {{ text-align: right; width: 12%; }}
    .article-table th:nth-child(5), .article-table td:nth-child(5) {{ text-align: right; width: 20%; }}
    .article-table tr {{ cursor: default; }}
    .article-table tbody tr:hover {{ background: #f0fdfa; }}
    .article-table tbody tr.selected {{ background: #ccfbf1; font-weight: 800; color: #115e59; }}
    .article-name-cell {{ display: flex; align-items: center; min-width: 0; gap: 2px; }}
    .article-name-text {{ display: flex; min-width: 0; flex: 1 1 auto; overflow: hidden; }}
    .article-name-start {{ min-width: 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
    .article-name-end {{ flex: 0 0 auto; white-space: nowrap; }}
    .article-actions {{ display: inline-flex; align-items: center; gap: 2px; flex: 0 0 auto; }}
    .article-copy {{ background: none; border: none; cursor: pointer; padding: 0 1px; margin-left: 0; font-size: .8em; color: #94a3b8; opacity: .75; line-height: 1; vertical-align: middle; }}
    .article-copy:hover {{ color: var(--accent); opacity: 1; }}
    .article-copy.copied {{ color: #16a34a; opacity: 1; }}
    .article-link {{ color: var(--accent); text-decoration: none; font-weight: 800; margin-left: 0; }}
    .article-link:hover {{ text-decoration: underline; }}
    .tbl table {{ width: max-content; min-width: 100%; border-collapse: collapse; font-size: 11px; }}
    .tbl th, .tbl td {{ border-bottom: 1px solid #eef2f7; padding: 6px 7px; text-align: right; white-space: nowrap; }}
    .tbl th:first-child, .tbl td:first-child {{ position: sticky; left: 0; z-index: 1; text-align: center; background: #fff; width: var(--graph-col-width); min-width: var(--graph-col-width); max-width: var(--graph-col-width); padding-left: 0; padding-right: 0; }}
    .tbl td:first-child {{ vertical-align: middle; }}
    .tbl th:nth-child(2), .tbl td:nth-child(2) {{ position: sticky; left: var(--graph-col-width); z-index: 1; text-align: left; background: #fff; width: var(--metric-col-width); min-width: var(--metric-col-width); max-width: var(--metric-col-width); font-weight: 700; }}
    .tbl th:nth-child(3), .tbl td:nth-child(3) {{ position: sticky; left: calc(var(--graph-col-width) + var(--metric-col-width)); z-index: 1; background: #fff; width: var(--total-col-width); min-width: var(--total-col-width); max-width: var(--total-col-width); font-weight: 800; }}
    .tbl th {{ position: sticky; top: 0; z-index: 2; background: #f8fafc; }}
    .tbl th.weekend-date {{ color: #7f1d1d; }}
    .tbl th:first-child, .tbl th:nth-child(2), .tbl th:nth-child(3) {{ z-index: 3; background: #f8fafc; }}
    .tbl tr.group-row td {{ background: #eef6f4; color: #0f766e; font-size: .78rem; font-weight: 800; text-transform: uppercase; letter-spacing: .04em; }}
    .tbl tr.group-row td:first-child {{ background: #eef6f4; }}
    .tbl tr.group-row td:nth-child(2) {{ background: #eef6f4; }}
    .tbl tr.group-row td:nth-child(3) {{ background: #eef6f4; }}
    .group-toggle {{ width: 22px; height: 22px; padding: 0; margin: 0; border-radius: 999px; background: #0f766e; color: #fff; font-size: .92rem; line-height: 1; }}
    .group-label {{ cursor: pointer; }}
    .metric-toggle {{ display: inline-block; width: 16px; height: 16px; min-width: 16px; padding: 0; margin: 0; border: 0; border-radius: 0; background: transparent; box-shadow: none; accent-color: var(--accent); cursor: pointer; vertical-align: middle; appearance: auto; -webkit-appearance: checkbox; }}
    .chart-card {{ background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; }}
    .chart-head {{ display: flex; justify-content: space-between; align-items: center; gap: 10px; margin-bottom: 10px; }}
    .chart-title {{ font-size: .95rem; font-weight: 800; }}
    .chart-subtitle {{ color: var(--muted); font-size: .74rem; }}
    .chart-legend {{ display: flex; flex-wrap: wrap; gap: 8px 14px; margin: 0 0 10px; }}
    .chart-legend-item {{ display: inline-flex; align-items: center; gap: 7px; padding: 5px 9px; border: 1px solid #dbeafe; border-radius: 999px; background: #f8fafc; color: #0f172a; font-size: .78rem; font-weight: 600; }}
    .chart-swatch {{ width: 10px; height: 10px; border-radius: 999px; flex: 0 0 auto; }}
    .chart-empty {{ color: var(--muted); font-size: .78rem; padding: 14px 0 4px; }}
    .chart-svg-wrap {{ position: relative; width: 100%; overflow-x: auto; border: 1px solid #eef2f7; border-radius: 10px; background: linear-gradient(180deg, #fcfffe, #f8fafc); }}
    .chart-svg-wrap[data-fit-width="1"] {{ overflow-x: hidden; }}
    .chart-svg {{ display: block; min-width: var(--chart-min-width, 980px); width: 100%; height: var(--chart-height, 860px); }}
    .chart-svg-wrap[data-fit-width="0"] .chart-svg {{ width: var(--chart-min-width, 980px); height: auto; aspect-ratio: var(--chart-aspect-ratio, 980 / 860); }}
    .chart-svg-wrap[data-fit-width="1"] .chart-svg {{ min-width: 0; height: auto; aspect-ratio: var(--chart-aspect-ratio, 980 / 860); }}
    .chart-axis {{ stroke: #cbd5e1; stroke-width: 1; }}
    .chart-grid {{ stroke: #e5e7eb; stroke-width: 1; stroke-dasharray: 3 4; }}
    .chart-label {{ fill: #64748b; font-size: var(--chart-font-size, 11px); }}
    .chart-zone-label {{ fill: #0f766e; font-size: var(--chart-zone-font-size, 12px); font-weight: 700; }}
    .chart-line {{ fill: none; stroke-width: var(--chart-line-width, 2.5); stroke-linecap: round; stroke-linejoin: round; }}
    .chart-dot {{ stroke: #fff; stroke-width: var(--chart-dot-stroke, 1.5); }}
    .chart-hover-line {{ stroke: #0f766e; stroke-width: 1.5; stroke-dasharray: 4 4; opacity: .55; }}
    .chart-hit {{ fill: transparent; cursor: crosshair; }}
    .chart-tooltip {{ position: absolute; display: none; min-width: 180px; max-width: 280px; padding: 8px 10px; border: 1px solid #99f6e4; border-radius: 10px; background: rgba(255, 255, 255, .96); box-shadow: 0 14px 28px rgba(15, 118, 110, .16); color: #0f172a; font-size: 12px; line-height: 1.35; pointer-events: none; }}
    .chart-tooltip-date {{ font-weight: 800; margin-bottom: 6px; color: #0f766e; }}
    .chart-tooltip-row {{ display: flex; align-items: center; justify-content: space-between; gap: 10px; margin-top: 4px; }}
    .chart-tooltip-name {{ display: inline-flex; align-items: center; gap: 6px; min-width: 0; }}
    .chart-tooltip-name span:last-child {{ overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
    .chart-tooltip-dot {{ width: 8px; height: 8px; border-radius: 999px; flex: 0 0 auto; }}
    .chart-tooltip-value {{ font-weight: 700; text-align: right; white-space: nowrap; }}
    .workspace.articles-collapsed .articles {{ overflow: visible; }}
    .workspace.articles-collapsed .articles .articles-inner {{ display: none; }}
    .workspace.articles-collapsed .articles .articles-mini {{ display: flex; }}
    .workspace.articles-collapsed .articles:hover,
    .workspace.articles-collapsed .articles:focus-within {{ width: 332px; z-index: 20; box-shadow: 0 18px 40px rgba(15, 118, 110, .18); border-color: #99f6e4; }}
    .workspace.articles-collapsed .articles:hover .articles-inner,
    .workspace.articles-collapsed .articles:focus-within .articles-inner {{ display: flex; }}
    .workspace.articles-collapsed .articles:hover .articles-mini,
    .workspace.articles-collapsed .articles:focus-within .articles-mini {{ display: none; }}
    @media (max-width: 1100px) {{
      .workspace {{ grid-template-columns: 1fr; }}
      .workspace.articles-collapsed {{ --articles-width: 1fr; }}
      .articles {{ order: 2; }}
      .articles {{ max-height: 260px; }}
      .workspace.articles-collapsed .articles {{ width: auto; overflow: hidden; }}
      .workspace.articles-collapsed .articles .articles-inner {{ display: flex; }}
      .workspace.articles-collapsed .articles .articles-mini {{ display: none; }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>{page_title}</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="article">Артикул</label>
        <input id="article" type="text" placeholder="например 12-0042-001">
      </div>
      <div>
        <label for="df">Период с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Период по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div class="quick-range">
        <button type="button" id="range-7">7 дней</button>
        <button type="button" id="range-14">14 дней</button>
        <button type="button" id="range-31">Месяц</button>
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="workspace" id="workspace">
      <aside class="articles">
        <div class="articles-inner">
          <div class="articles-head">
            <div class="articles-title">Артикулы</div>
            <div class="articles-actions">
              <button type="button" id="toggle-articles" title="Свернуть список">◂</button>
              <button type="button" id="select-all">Все</button>
              <button type="button" id="clear-all">Сброс</button>
            </div>
          </div>
          <div class="subject-filter">
            <label for="subject">Предмет</label>
            <select id="subject">
              <option value="">Все предметы</option>
            </select>
          </div>
          <div class="subject-filter">
            <label for="strategy">Ярлыки</label>
            <select id="strategy" multiple size="4" style="height:72px;" title="Ctrl/Cmd для множественного выбора"></select>
          </div>
          <div class="subject-filter">
            <label for="revenue-category">Категория по выручке</label>
            <select id="revenue-category">
              <option value="">Все категории</option>
            </select>
          </div>
          <div class="article-list" id="article-list"></div>
        </div>
        <div class="articles-mini">
          <button type="button" id="expand-articles" title="Развернуть список">▸</button>
          <span>Артикулы</span>
        </div>
      </aside>
      <div class="main-pane">
        <div class="tbl" id="tbl"></div>
        <section class="chart-card">
          <div class="chart-head">
            <div>
              <div class="chart-title">График метрик</div>
              <div class="chart-subtitle">Суммы и проценты на одном графике. Проценты вынесены ниже и масштабируются отдельно.</div>
            </div>
          </div>
          <div class="chart-legend" id="chart-legend"></div>
          <div class="chart-svg-wrap" id="chart-wrap"></div>
        </section>
      </div>
    </div>
  </div>
<script>
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const articleList = document.getElementById('article-list');
const subjectEl = document.getElementById('subject');
const strategyEl = document.getElementById('strategy');
const revenueCategoryEl = document.getElementById('revenue-category');
const chartLegend = document.getElementById('chart-legend');
const chartWrap = document.getElementById('chart-wrap');
const workspaceEl = document.getElementById('workspace');
const reportGranularity = '{granularity}';
const filterKey = `ozon.analytics.buyoutOrderDay.v2.${{reportGranularity}}.filters`;
let requestSeq = 0;
let filterTimer = null;
let selectedArticles = new Set();
let articlesLoadedFor = '';
let subjectsLoadedFor = '';
function fillMultiSelect(select, values) {{
  const prevVals = new Set([...select.selectedOptions].map(o=>o.value));
  const optionsHtml = (values || []).map(v => {{
    const sel = prevVals.has(v) ? ' selected' : '';
    return `<option value="${{String(v).replaceAll('"','&quot;')}}"${{sel}}>${{v}}</option>`;
  }}).join('');
  select.innerHTML = select.multiple ? optionsHtml : `<option value="">Все ярлыки</option>${{optionsHtml}}`;
}}
let articleItems = [];
let articleSort = {{ key: 'revenue', dir: 'desc' }};
let lastArticleIndex = -1;
let activeChartMetrics = new Set();
let latestChartPayload = null;
let articlesCollapsed = false;
let collapsedGroups = new Set();
let chartMetricsInitialized = false;
const chartPalette = ['#0f766e', '#f97316', '#0284c7', '#dc2626', '#7c3aed', '#16a34a', '#ca8a04', '#db2777', '#0891b2', '#4f46e5'];

function syncArticlesPanel() {{
  workspaceEl.classList.toggle('articles-collapsed', articlesCollapsed);
  const toggleBtn = document.getElementById('toggle-articles');
  const expandBtn = document.getElementById('expand-articles');
  if (toggleBtn) {{
    toggleBtn.textContent = articlesCollapsed ? '▸' : '◂';
    toggleBtn.title = articlesCollapsed ? 'Развернуть список' : 'Свернуть список';
  }}
  if (expandBtn) expandBtn.title = articlesCollapsed ? 'Развернуть список' : 'Свернуть список';
}}

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.article !== undefined) document.getElementById('article').value = saved.article;
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
    if (saved.subject !== undefined) subjectEl.dataset.pendingValue = saved.subject;
    // labels not restored from storage - always start empty
    if (saved.revenueCategory !== undefined) revenueCategoryEl.dataset.pendingValue = saved.revenueCategory;
    if (Array.isArray(saved.selectedArticles)) selectedArticles = new Set(saved.selectedArticles.map(String).filter(value => /^\d+$/.test(value)));
    if (Array.isArray(saved.chartMetrics)) {{
      activeChartMetrics = new Set(saved.chartMetrics);
      chartMetricsInitialized = true;
    }}
    if (Array.isArray(saved.collapsedGroups)) collapsedGroups = new Set(saved.collapsedGroups.map(String).filter(Boolean));
    articlesCollapsed = Boolean(saved.articlesCollapsed);
  }} catch (err) {{}}
  syncArticlesPanel();
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    article: document.getElementById('article').value,
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value,
    subject: subjectEl.value,
    labels: [...strategyEl.selectedOptions].map(o=>o.value).join(","),
    revenueCategory: revenueCategoryEl.value,
    selectedArticles: [...selectedArticles],
    chartMetrics: [...activeChartMetrics],
    collapsedGroups: [...collapsedGroups],
    articlesCollapsed
  }}));
}}

function isoDate(date) {{
  return date.toISOString().slice(0, 10);
}}

function applyQuickRange(days) {{
  const end = new Date();
  const start = new Date(end);
  start.setDate(start.getDate() - (days - 1));
  document.getElementById('df').value = isoDate(start);
  document.getElementById('dt').value = isoDate(end);
  loadData();
}}

function scheduleLoad() {{
  clearTimeout(filterTimer);
  filterTimer = setTimeout(loadData, 350);
}}

function resetArticleFilters() {{
  subjectEl.value = '';
  [...strategyEl.options].forEach(o=>o.selected=false);
  revenueCategoryEl.value = '';
  delete subjectEl.dataset.pendingValue;
  
  delete revenueCategoryEl.dataset.pendingValue;
}}

function escapeHtml(value) {{
  return String(value)
    .replaceAll('&', '&amp;')
    .replaceAll('<', '&lt;')
    .replaceAll('>', '&gt;')
    .replaceAll('"', '&quot;')
    .replaceAll("'", '&#39;');
}}

function formatCompactNumber(value, fractionDigits = 0) {{
  if (value === null || value === undefined || value === '') return '';
  const numeric = Number(value);
  if (!Number.isFinite(numeric)) return '';
  return numeric.toLocaleString('ru-RU', {{
    minimumFractionDigits: fractionDigits,
    maximumFractionDigits: fractionDigits
  }});
}}

function parseMetricValue(value) {{
  const text = String(value ?? '').trim();
  if (!text) return 0;
  const normalized = text.replaceAll(' ', '').replace('%', '').replace(',', '.');
  const numeric = Number(normalized);
  return Number.isFinite(numeric) ? numeric : 0;
}}

function chartRowsFromPayload(data) {{
  return (data.rows || []).filter(row => String(row.metric || '').trim());
}}

function ensureDefaultChartMetrics(rows) {{
  if (chartMetricsInitialized || activeChartMetrics.size) return;
  const defaults = [
    'Сумма заказов',
    'Сумма выкупов в эту дату',
    'Реклама',
    'ДРР от заказов',
    'Маржинальность от выкупов за дату',
  ];
  const available = new Set(rows.map(row => row.metric));
  for (const metric of defaults) {{
    if (available.has(metric)) activeChartMetrics.add(metric);
  }}
  if (!activeChartMetrics.size) {{
    for (const row of rows.slice(0, 4)) activeChartMetrics.add(row.metric);
  }}
  chartMetricsInitialized = true;
}}

function renderChart(data) {{
  latestChartPayload = data;
  const rows = chartRowsFromPayload(data);
  ensureDefaultChartMetrics(rows);
  const selectedRows = rows.filter(row => activeChartMetrics.has(row.metric));
  const dates = data.dates || [];
  if (!dates.length || !selectedRows.length) {{
    chartLegend.innerHTML = '';
    chartWrap.dataset.fitWidth = '0';
    chartWrap.innerHTML = '<div class="chart-empty">Выберите хотя бы одну метрику с данными.</div>';
    return;
  }}

  const isWeekChart = reportGranularity === 'week';
  const chartConfig = isWeekChart
    ? {{
        minWidth: Math.max(840, dates.length * 150 + 260),
        height: 760,
        left: 74,
        right: 26,
        top: 30,
        bottom: 112,
        gap: 44,
        moneyHeight: 390,
        percentHeight: 184,
        fontSize: 12,
        zoneFontSize: 13,
        lineWidth: 3.2,
        dotStroke: 1.8,
        dotRadius: 4.8,
      }}
    : {{
        minWidth: Math.max(980, dates.length * 84 + 220),
        height: 860,
        left: 64,
        right: 18,
        top: 24,
        bottom: 110,
        gap: 52,
        moneyHeight: 450,
        percentHeight: 220,
        fontSize: 11,
        zoneFontSize: 12,
        lineWidth: 2.5,
        dotStroke: 1.5,
        dotRadius: 3.5,
      }};
  const percentRows = selectedRows.filter(row => String(row.kind || '').startsWith('percent'));
  const moneyRows = selectedRows.filter(row => !String(row.kind || '').startsWith('percent'));
  const moneySeries = moneyRows.map(row => row.values.map(parseMetricValue));
  const percentSeries = percentRows.map(row => row.values.map(parseMetricValue));
  const moneyMax = Math.max(0, ...moneySeries.flat());
  const percentMax = Math.max(0, ...percentSeries.flat());

  const width = chartConfig.minWidth;
  const height = chartConfig.height;
  const preserveAspectRatio = 'xMidYMid meet';
  // Явно задаем размер SVG, чтобы контейнер не растягивал недельный график.
  const svgInlineStyle = isWeekChart
    ? `width:${{width}}px;height:auto;display:block`
    : 'width:100%;height:auto;display:block';
  const left = chartConfig.left;
  const right = chartConfig.right;
  const top = chartConfig.top;
  const bottom = chartConfig.bottom;
  const gap = chartConfig.gap;
  const plotWidth = width - left - right;
  const moneyHeight = chartConfig.moneyHeight;
  const percentHeight = chartConfig.percentHeight;
  const moneyTop = top;
  const percentTop = moneyTop + moneyHeight + gap;
  const moneyBottom = moneyTop + moneyHeight;
  const percentBottom = percentTop + percentHeight;
  const xStep = dates.length > 1 ? plotWidth / (dates.length - 1) : 0;
  const moneyScale = moneyMax > 0 ? moneyHeight / moneyMax : 1;
  const percentScale = percentMax > 0 ? percentHeight / percentMax : 1;
  const formatDate = (iso) => {{
    if (String(iso).includes('..')) {{
      return String(iso).split('..').map(part => {{
        const [yy, mm, dd] = part.split('-');
        return `${{dd}}.${{mm}}`;
      }}).join('-');
    }}
    const [y, m, d] = String(iso).split('-');
    return `${{d}}.${{m}}`;
  }};
  const formatTooltipDate = (iso) => {{
    if (String(iso).includes('..')) {{
      return String(iso).split('..').map(part => {{
        const [y, m, d] = part.split('-');
        return `${{d}}.${{m}}.${{y}}`;
      }}).join(' - ');
    }}
    const [y, m, d] = String(iso).split('-');
    return `${{d}}.${{m}}.${{y}}`;
  }};

  const gridLines = [];
  for (let idx = 0; idx <= 4; idx++) {{
    const moneyY = moneyTop + (moneyHeight / 4) * idx;
    const percentY = percentTop + (percentHeight / 4) * idx;
    const moneyValue = moneyMax * (1 - idx / 4);
    const percentValue = percentMax * (1 - idx / 4);
    gridLines.push(`<line class="chart-grid" x1="${{left}}" y1="${{moneyY}}" x2="${{width - right}}" y2="${{moneyY}}"></line>`);
    gridLines.push(`<text class="chart-label" x="${{left - 8}}" y="${{moneyY + 4}}" text-anchor="end">${{formatCompactNumber(moneyValue, 0)}}</text>`);
    gridLines.push(`<line class="chart-grid" x1="${{left}}" y1="${{percentY}}" x2="${{width - right}}" y2="${{percentY}}"></line>`);
    gridLines.push(`<text class="chart-label" x="${{left - 8}}" y="${{percentY + 4}}" text-anchor="end">${{formatCompactNumber(percentValue, 0)}}%</text>`);
  }}

  const xLabels = dates.map((day, idx) => {{
    const x = left + xStep * idx;
    return `
      <line class="chart-grid" x1="${{x}}" y1="${{moneyTop}}" x2="${{x}}" y2="${{percentBottom}}"></line>
      <text class="chart-label" x="${{x}}" y="${{height - 18}}" text-anchor="end" transform="rotate(-45 ${{x}} ${{height - 18}})">${{formatDate(day)}}</text>
    `;
  }}).join('');
  const allSeries = selectedRows.map((row, idx) => ({{
    row,
    color: chartPalette[idx % chartPalette.length],
  }}));
  const legendHtml = allSeries.map(series => `
    <div class="chart-legend-item">
      <span class="chart-swatch" style="background:${{series.color}}"></span>
      <span>${{escapeHtml(series.row.metric)}}</span>
    </div>
  `).join('');

  const seriesSvg = selectedRows.map((row, idx) => {{
    const isPercent = String(row.kind || '').startsWith('percent');
    const values = row.values.map(parseMetricValue);
    const color = chartPalette[idx % chartPalette.length];
    const zoneTop = isPercent ? percentTop : moneyTop;
    const zoneBottom = isPercent ? percentBottom : moneyBottom;
    const scale = isPercent ? percentScale : moneyScale;
    const points = values.map((value, pointIdx) => {{
      const x = left + xStep * pointIdx;
      const y = zoneBottom - value * scale;
      return `${{x}},${{y}}`;
    }}).join(' ');
    const dots = values.map((value, pointIdx) => {{
      const x = left + xStep * pointIdx;
      const y = zoneBottom - value * scale;
      return `<circle class="chart-dot" cx="${{x}}" cy="${{y}}" r="${{chartConfig.dotRadius}}" fill="${{color}}"><title>${{escapeHtml(row.metric)}}: ${{row.values[pointIdx]}}</title></circle>`;
    }}).join('');
    const labelX = width - right - 8;
    const lastValue = values[values.length - 1] || 0;
    const labelY = zoneBottom - lastValue * scale - 6;
    return `
      <polyline class="chart-line" points="${{points}}" stroke="${{color}}"></polyline>
      ${{dots}}
      <text class="chart-label" x="${{labelX}}" y="${{Math.max(zoneTop + 12, Math.min(zoneBottom - 6, labelY))}}" text-anchor="end" fill="${{color}}">${{escapeHtml(row.metric)}}</text>
    `;
  }}).join('');

  const hoverTargets = dates.map((day, idx) => {{
    const x = left + xStep * idx;
    const rectWidth = dates.length > 1 ? Math.max(24, xStep) : plotWidth;
    const rectX = dates.length > 1 ? x - rectWidth / 2 : left;
    return `<rect class="chart-hit" data-index="${{idx}}" x="${{rectX}}" y="${{moneyTop}}" width="${{rectWidth}}" height="${{percentBottom - moneyTop}}"></rect>`;
  }}).join('');

  chartWrap.style.setProperty('--chart-min-width', `${{width}}px`);
  chartWrap.style.setProperty('--chart-height', `${{height}}px`);
  chartWrap.style.setProperty('--chart-aspect-ratio', `${{width}} / ${{height}}`);
  chartWrap.style.setProperty('--chart-font-size', `${{chartConfig.fontSize}}px`);
  chartWrap.style.setProperty('--chart-zone-font-size', `${{chartConfig.zoneFontSize}}px`);
  chartWrap.style.setProperty('--chart-line-width', `${{chartConfig.lineWidth}}`);
  chartWrap.style.setProperty('--chart-dot-stroke', `${{chartConfig.dotStroke}}`);
  chartWrap.dataset.fitWidth = isWeekChart ? '0' : '1';
  chartLegend.innerHTML = legendHtml;
  chartWrap.innerHTML = `
    <div class="chart-tooltip" id="chart-tooltip"></div>
    <svg class="chart-svg" style="${{svgInlineStyle}}" viewBox="0 0 ${{width}} ${{height}}" preserveAspectRatio="${{preserveAspectRatio}}">
      <text class="chart-zone-label" x="${{left}}" y="${{moneyTop - 6}}">Суммы</text>
      <text class="chart-zone-label" x="${{left}}" y="${{percentTop - 6}}">Проценты</text>
      <line class="chart-axis" x1="${{left}}" y1="${{moneyTop}}" x2="${{left}}" y2="${{moneyBottom}}"></line>
      <line class="chart-axis" x1="${{left}}" y1="${{moneyBottom}}" x2="${{width - right}}" y2="${{moneyBottom}}"></line>
      <line class="chart-axis" x1="${{left}}" y1="${{percentTop}}" x2="${{left}}" y2="${{percentBottom}}"></line>
      <line class="chart-axis" x1="${{left}}" y1="${{percentBottom}}" x2="${{width - right}}" y2="${{percentBottom}}"></line>
      ${{gridLines.join('')}}
      ${{xLabels}}
      ${{seriesSvg}}
      <line class="chart-hover-line" id="chart-hover-line" x1="${{left}}" y1="${{moneyTop}}" x2="${{left}}" y2="${{percentBottom}}" visibility="hidden"></line>
      ${{hoverTargets}}
    </svg>
  `;

  const svgEl = chartWrap.querySelector('.chart-svg');
  const tooltipEl = document.getElementById('chart-tooltip');
  const hoverLineEl = document.getElementById('chart-hover-line');
  const showTooltip = (index, clientX, clientY) => {{
    const x = left + xStep * index;
    hoverLineEl.setAttribute('x1', String(x));
    hoverLineEl.setAttribute('x2', String(x));
    hoverLineEl.setAttribute('visibility', 'visible');
    tooltipEl.innerHTML = `
      <div class="chart-tooltip-date">${{formatTooltipDate(dates[index])}}</div>
      ${{allSeries.map(series => `
        <div class="chart-tooltip-row">
          <div class="chart-tooltip-name">
            <span class="chart-tooltip-dot" style="background:${{series.color}}"></span>
            <span>${{escapeHtml(series.row.metric)}}</span>
          </div>
          <div class="chart-tooltip-value">${{escapeHtml(String((series.row.values || [])[index] || '0'))}}</div>
        </div>
      `).join('')}}
    `;
    tooltipEl.style.display = 'block';
    const wrapRect = chartWrap.getBoundingClientRect();
    const svgRect = svgEl.getBoundingClientRect();
    const tooltipWidth = tooltipEl.offsetWidth;
    const tooltipHeight = tooltipEl.offsetHeight;
    let leftPos = clientX - wrapRect.left + 14;
    let topPos = clientY - wrapRect.top - tooltipHeight - 14;
    if (leftPos + tooltipWidth > wrapRect.width - 8) leftPos = clientX - wrapRect.left - tooltipWidth - 14;
    if (leftPos < 8) leftPos = 8;
    if (topPos < 8) topPos = clientY - wrapRect.top + 14;
    if (topPos + tooltipHeight > wrapRect.height - 8) topPos = Math.max(8, wrapRect.height - tooltipHeight - 8);
    tooltipEl.style.left = `${{leftPos + chartWrap.scrollLeft}}px`;
    tooltipEl.style.top = `${{topPos + chartWrap.scrollTop}}px`;
  }};
  const hideTooltip = () => {{
    hoverLineEl.setAttribute('visibility', 'hidden');
    tooltipEl.style.display = 'none';
  }};
  for (const target of [...chartWrap.querySelectorAll('.chart-hit')]) {{
    target.addEventListener('mouseenter', (event) => showTooltip(Number(target.dataset.index || 0), event.clientX, event.clientY));
    target.addEventListener('mousemove', (event) => showTooltip(Number(target.dataset.index || 0), event.clientX, event.clientY));
    target.addEventListener('mouseleave', hideTooltip);
  }}
}}

function cellColorStyle(kind, prevRaw, currRaw, row) {{
  if (prevRaw === null || prevRaw === undefined || currRaw === null || currRaw === undefined) return '';
  if (kind === 'stock') {{
    const diff = currRaw - prevRaw;
    const absThreshold = Number(row?.color_threshold ?? 20);
    if (diff < absThreshold) return '';
    return 'background:#dcfce7;color:#047857;font-weight:900;box-shadow:inset 0 0 0 1px #86efac';
  }}
  if (String(row?.metric || '') === 'СПП') {{
    const diff = currRaw - prevRaw;
    const threshold = Number(row?.color_threshold ?? 1);
    // Для СПП подсветка должна включаться уже при изменении ровно на порог.
    if (Math.abs(diff) < threshold) return '';
    return diff > 0
      ? 'color:#047857;font-weight:900;text-shadow:0 0 0 #047857'
      : 'color:#b91c1c;font-weight:900;text-shadow:0 0 0 #b91c1c';
  }}
  const isPercent = kind === 'percent_income' || kind === 'percent_expense';
  const threshold = Number(row?.color_threshold ?? (isPercent ? 5 : 10));
  const colorMode = row?.color_mode || (isPercent ? 'absolute' : 'percent');
  let beneficial;
  if (isPercent) {{
    const diff = currRaw - prevRaw;
    if (Math.abs(diff) <= threshold) return '';
    beneficial = kind === 'percent_income' ? diff > 0 : diff < 0;
  }} else {{
    const diff = currRaw - prevRaw;
    if (colorMode === 'absolute') {{
      if (Math.abs(diff) <= threshold) return '';
      beneficial = kind === 'income' ? diff > 0 : diff < 0;
    }} else {{
      if (Math.abs(prevRaw) < 1) return '';
      const pctChange = diff / Math.abs(prevRaw) * 100;
      if (Math.abs(pctChange) <= threshold) return '';
      beneficial = kind === 'income' ? pctChange > 0 : pctChange < 0;
    }}
  }}
  return beneficial ? 'color:#1a8c40;font-weight:600' : 'color:#c0392b;font-weight:600';
}}

function isHeatmapMetric(row) {{
  const metric = String(row?.metric || '');
  const group = String(row?.group || '');
  const kind = String(row?.kind || '');
  if (metric === 'СПП') return true;
  if (group === 'Воронка' && kind === 'percent_income') return true;
  return metric === 'ДРР от выкупов в эту дату' || metric === 'Маржинальность от выкупов за дату';
}}

function heatmapCellStyle(row, raw) {{
  if (!isHeatmapMetric(row) || raw === null || raw === undefined || !Number.isFinite(raw)) return '';
  if (String(row?.metric || '') === 'СПП') {{
    const values = (row?.raw_values || []).filter(value => Number.isFinite(value));
    if (!values.length) return '';
    const min = Math.min(...values);
    const max = Math.max(...values);
    const span = max - min;
    const score = span > 0 ? (raw - min) / span : 1;
    const hue = Math.round(8 + score * 132);
    const lightness = Math.round(96 - score * 14);
    return `background:hsl(${{hue}} 58% ${{lightness}}%);color:#0f172a;font-weight:700`;
  }}
  const values = (row?.raw_values || []).filter(value => Number.isFinite(value));
  if (!values.length) return '';
  const min = Math.min(...values);
  const max = Math.max(...values);
  const span = max - min;
  let score = span > 0 ? (raw - min) / span : 1;
  if (String(row?.metric || '') === 'ДРР от выкупов в эту дату') score = 1 - score;
  const hue = Math.round(8 + score * 132);
  const lightness = Math.round(94 - score * 18);
  return `background:hsl(${{hue}} 72% ${{lightness}}%);color:#0f172a;font-weight:700`;
}}

function metricColumnWidth(rows) {{
  const labels = ['Метрика', ...(rows || []).map(row => String(row.metric || ''))];
  const canvas = document.createElement('canvas');
  const ctx = canvas.getContext('2d');
  if (!ctx) return 240;
  ctx.font = '700 11px "Segoe UI", "Trebuchet MS", sans-serif';
  const textWidth = Math.max(...labels.map(label => ctx.measureText(label).width), 0);
  return Math.max(240, Math.min(560, Math.ceil(textWidth + 28)));
}}

function render(data) {{
  const dates = data.dates || [];
  const rows = data.rows || [];
  const formatDate = (iso) => {{
    if (String(iso).includes('..')) {{
      return String(iso).split('..').map(part => {{
        const [yy, mm, dd] = part.split('-');
        return `${{dd}}.${{mm}}.${{String(yy).slice(2)}}`;
      }}).join('-');
    }}
    const [y, m, d] = String(iso).split('-');
    return `${{d}}.${{m}}.${{String(y).slice(2)}}`;
  }};
  const isWeekendDate = (iso) => {{
    if (String(iso).includes('..')) return false;
    const day = new Date(`${{iso}}T00:00:00`).getDay();
    return day === 0 || day === 6;
  }};
  const head = '<tr><th>Гр.</th><th>Метрика</th><th>Итого / среднее</th>' + dates.map(d => `<th class="${{isWeekendDate(d) ? 'weekend-date' : ''}}">${{formatDate(d)}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.reduce((html, row, idx) => {{
        const prev = idx > 0 ? rows[idx - 1].group : null;
        const group = row.group || '';
        const kind = row.kind || 'none';
        const rawVals = row.raw_values || [];
        const groupCollapsed = group ? collapsedGroups.has(group) : false;
        const groupRow = group && group !== prev
          ? `<tr class="group-row" data-group-header="${{escapeHtml(group)}}">
              <td><button type="button" class="group-toggle" data-group-toggle="${{escapeHtml(group)}}" aria-label="Свернуть группу">${{groupCollapsed ? '+' : '−'}}</button></td>
              <td class="group-label" data-group-toggle="${{escapeHtml(group)}}">${{group}}</td>
              <td></td>
              ${{dates.map(() => '<td></td>').join('')}}
            </tr>`
          : '';
        const checked = activeChartMetrics.has(row.metric) ? ' checked' : '';
        const cells = (row.values || []).map((v, i) => {{
          const heatmapStyle = heatmapCellStyle(row, rawVals[i]);
          const changeStyle = (kind !== 'none' && kind !== 'neutral' && i > 0)
            ? cellColorStyle(kind, rawVals[i - 1], rawVals[i], row)
            : '';
          const style = heatmapStyle && changeStyle && String(row.metric || '') === 'СПП'
            ? `${{heatmapStyle}};${{changeStyle}}`
            : (heatmapStyle || changeStyle);
          return style ? `<td style="${{style}}">${{v}}</td>` : `<td>${{v}}</td>`;
        }}).join('');
        const hiddenAttr = groupCollapsed ? ' hidden' : '';
        return html + groupRow + `<tr data-group="${{escapeHtml(group)}}"${{hiddenAttr}}><td><input class="metric-toggle" type="checkbox" data-metric="${{escapeHtml(row.metric)}}"${{checked}}></td><td>${{row.metric}}</td><td>${{row.total || ''}}</td>${{cells}}</tr>`;
      }}, '')
    : `<tr><td colspan="${{dates.length + 3}}">Нет данных за выбранный период</td></tr>`;
  tbl.style.setProperty('--metric-col-width', `${{metricColumnWidth(rows)}}px`);
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
  requestAnimationFrame(() => setupStickyHeader());
  for (const toggle of [...tbl.querySelectorAll('[data-group-toggle]')]) {{
    toggle.addEventListener('click', () => {{
      const group = toggle.dataset.groupToggle || '';
      if (!group) return;
      if (collapsedGroups.has(group)) collapsedGroups.delete(group);
      else collapsedGroups.add(group);
      saveFilters();
      render(data);
    }});
  }}
  for (const input of [...tbl.querySelectorAll('input[data-metric]')]) {{
    input.addEventListener('change', () => {{
      const metric = input.dataset.metric;
      if (input.checked) activeChartMetrics.add(metric);
      else activeChartMetrics.delete(metric);
      chartMetricsInitialized = true;
      saveFilters();
      renderChart(data);
    }});
  }}
  renderChart(data);
}}

function renderArticles(articles) {{
  const normalized = articles.map(item => typeof item === 'string'
    ? {{ article: item, nmid: '', stock: null, buyouts: 0, revenue: 0, drr: null, turnover: null }}
    : item
  );
  const middleArticleHtml = (value) => {{
    const text = String(value || '');
    if (text.length <= 12) return `<span class="article-name-text"><span class="article-name-start">${{escapeHtml(text)}}</span></span>`;
    const endSize = Math.min(8, Math.max(4, Math.floor(text.length * 0.35)));
    return `<span class="article-name-text"><span class="article-name-start">${{escapeHtml(text.slice(0, -endSize))}}</span><span class="article-name-end">${{escapeHtml(text.slice(-endSize))}}</span></span>`;
  }};
  articleItems = [...normalized].sort((left, right) => {{
    const key = articleSort.key;
    const dir = articleSort.dir === 'asc' ? 1 : -1;
    const leftValue = key === 'article' ? String(left.article || '') : Number(left[key] || 0);
    const rightValue = key === 'article' ? String(right.article || '') : Number(right[key] || 0);
    if (key === 'article') return leftValue.localeCompare(rightValue, 'ru') * dir;
    return (leftValue - rightValue) * dir || String(left.article || '').localeCompare(String(right.article || ''), 'ru');
  }});
  const sortMark = (key) => articleSort.key === key ? (articleSort.dir === 'asc' ? ' ▲' : ' ▼') : '';
  articleList.innerHTML = articleItems.length
    ? `<table class="article-table">
        <thead><tr>
          <th data-sort="article">Артикул${{sortMark('article')}}</th>
          <th data-sort="stock">Ост.${{sortMark('stock')}}</th>
          <th data-sort="drr">ДРР${{sortMark('drr')}}</th>
          <th data-sort="turnover">Обор.${{sortMark('turnover')}}</th>
          <th data-sort="revenue">Выручка${{sortMark('revenue')}}</th>
        </tr></thead>
	        <tbody>${{articleItems.map(item => {{
	          const article = String(item.article || '');
	          const itemKey = String(item.nmid || item.article || '').trim();
	          const selected = selectedArticles.has(itemKey) ? ' selected' : '';
	          const safeArticle = escapeHtml(article);
	          const articleName = middleArticleHtml(article);
	          const nmid = String(item.nmid || '').trim();
          const copyBtn = `<button class="article-copy" data-copy="${{safeArticle}}" data-skip-select="1" title="Скопировать артикул">⎘</button>`;
          const ozonLink = /^\d+$/.test(nmid)
            ? `<a class="article-link" href="https://ozon.ru/product/${{nmid}}" target="_blank" rel="noopener noreferrer" title="Открыть товар на Ozon" data-skip-select="1">↗</a>`
            : '';
          const pct = item.drr != null ? item.drr.toFixed(1) + '%' : '—';
          const turnoverStr = (item.turnover != null ? item.turnover : 999) + 'д';
	          return `<tr class="article-item${{selected}}" data-article="${{safeArticle}}" data-key="${{escapeHtml(itemKey)}}" title="${{safeArticle}}">
	            <td><span class="article-name-cell">${{articleName}}<span class="article-actions">${{copyBtn}}${{ozonLink}}</span></span></td>
            <td>${{formatCompactNumber(item.stock)}}</td>
            <td class="drr-cell">${{pct}}</td>
            <td>${{turnoverStr}}</td>
            <td>${{formatCompactNumber(item.revenue)}}</td>
          </tr>`;
        }}).join('')}}</tbody>
      </table>`
    : '<div class="meta">Нет артикулов</div>';
  for (const th of [...articleList.querySelectorAll('th[data-sort]')]) {{
    th.addEventListener('click', () => {{
      const key = th.dataset.sort;
      articleSort = {{
        key,
        dir: articleSort.key === key && articleSort.dir === 'desc' ? 'asc' : 'desc'
      }};
      renderArticles(normalized);
    }});
  }}
  for (const btn of articleList.querySelectorAll('.article-copy')) {{
    btn.addEventListener('click', (e) => {{
      e.stopPropagation();
      navigator.clipboard.writeText(btn.dataset.copy).then(() => {{
        btn.classList.add('copied'); btn.textContent = '✓';
        setTimeout(() => {{ btn.classList.remove('copied'); btn.textContent = '⎘'; }}, 1500);
      }});
    }});
  }}
  for (const [idx, item] of [...articleList.querySelectorAll('.article-item')].entries()) {{
    item.addEventListener('click', (event) => {{
	      if (event.target.closest('[data-skip-select="1"]')) return;
	      const article = item.dataset.key || item.dataset.article;
	      if (event.shiftKey && lastArticleIndex >= 0) {{
	        const [from, to] = [lastArticleIndex, idx].sort((a, b) => a - b);
	        for (let pos = from; pos <= to; pos++) selectedArticles.add(String(articleItems[pos].nmid || articleItems[pos].article || '').trim());
      }} else if (event.ctrlKey || event.metaKey) {{
        if (selectedArticles.has(article)) {{
          selectedArticles.delete(article);
        }} else {{
          selectedArticles.add(article);
        }}
        lastArticleIndex = idx;
      }} else {{
        if (selectedArticles.has(article) && selectedArticles.size === 1) {{
          selectedArticles.clear();
        }} else {{
          selectedArticles = new Set([article]);
        }}
        lastArticleIndex = idx;
      }}
      saveFilters();
      renderArticles(articles);
      loadData();
    }});
  }}
}}

async function loadSubjects() {{
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  const key = `${{df}}..${{dt}}`;
  if (key === subjectsLoadedFor) return;
  subjectsLoadedFor = key;
  const currentSubject = subjectEl.dataset.pendingValue !== undefined ? subjectEl.dataset.pendingValue : subjectEl.value;
  const currentStrategy = [...strategyEl.selectedOptions].map(o=>o.value).join(',');
  const currentRevenueCategory = revenueCategoryEl.dataset.pendingValue !== undefined ? revenueCategoryEl.dataset.pendingValue : revenueCategoryEl.value;
  delete subjectEl.dataset.pendingValue;
  
  delete revenueCategoryEl.dataset.pendingValue;
  const qs = new URLSearchParams({{ date_from: df, date_to: dt}});
  const data = await fetchJsonChecked(`/api/ozon/analytics/buyout-filter-options?${{qs.toString()}}`);
  const subjects = data.subjects || [];
  const labelsData = data.labels || [];
  const revenueCategories = data.revenue_categories || [];
  subjectEl.innerHTML = '<option value="">Все предметы</option>' + subjects.map(subject => {{
    const selected = subject === currentSubject ? ' selected' : '';
    const safeSubject = escapeHtml(subject);
    return `<option value="${{safeSubject}}"${{selected}}>${{safeSubject}}</option>`;
  }}).join('');
  fillMultiSelect(strategyEl, labelsData);
  revenueCategoryEl.innerHTML = '<option value="">Все категории</option>' + revenueCategories.map(category => {{
    const selected = category === currentRevenueCategory ? ' selected' : '';
    const safeCategory = escapeHtml(category);
    return `<option value="${{safeCategory}}"${{selected}}>${{safeCategory}}</option>`;
  }}).join('');
}}

async function loadArticles() {{
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  await loadSubjects();
  const subject = subjectEl.value;
  const strategy = [...strategyEl.selectedOptions].map(o=>o.value).join(",");
  const revenueCategory = revenueCategoryEl.value;
  const key = `${{df}}..${{dt}}..${{subject}}..${{[...strategyEl.selectedOptions].map(o=>o.value).join(",")}}..${{revenueCategory}}`;
  if (key === articlesLoadedFor) return;
  articlesLoadedFor = key;
  const qs = new URLSearchParams({{ date_from: df, date_to: dt, subject, labels: strategy, revenue_category: revenueCategory}});
  const data = await fetchJsonChecked(`/api/ozon/analytics/buyout-articles?${{qs.toString()}}`);
  const articles = data.articles || [];
  if (selectedArticles.size) {{
    const available = new Set(articles.map(item => String((item && (item.nmid || item.article)) || '').trim()).filter(Boolean));
    selectedArticles = new Set([...selectedArticles].filter(item => available.has(item)));
  }}
  renderArticles(articles);
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const article = document.getElementById('article').value;
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  const subject = subjectEl.value;
  const strategy = [...strategyEl.selectedOptions].map(o=>o.value).join(",");
  const revenueCategory = revenueCategoryEl.value;
  meta.textContent = 'Загружаю...';
  try {{
    await loadArticles();
    const qs = new URLSearchParams({{ article, date_from: df, date_to: dt, subject, labels: strategy, revenue_category: revenueCategory, granularity: reportGranularity}});
    for (const articleName of selectedArticles) qs.append('articles', articleName);
    const data = await fetchJsonChecked(`/api/ozon/analytics/buyout-order-day?${{qs.toString()}}`);
    if (seq !== requestSeq) return;
    render(data);
    const selection = selectedArticles.size ? ` | Выбрано артикулов: ${{selectedArticles.size}}` : '';
    meta.textContent = `Дат: ${{(data.dates || []).length}} | Показан период: ${{data.effective_from || df}}..${{data.effective_to || dt}}${{selection}}`;
  }} catch (err) {{
    if (seq !== requestSeq) return;
    const message = err && err.message ? err.message : 'Не удалось загрузить данные';
    meta.textContent = `Ошибка: ${{message}}`;
  }}
}}

async function fetchJsonChecked(url) {{
  const res = await fetch(url);
  let data = null;
  try {{
    data = await res.json();
  }} catch (_) {{
    if (!res.ok) throw new Error(`HTTP ${{res.status}}`);
    throw new Error('Некорректный ответ сервера');
  }}
  if (!res.ok) throw new Error((data && data.error) ? data.error : `HTTP ${{res.status}}`);
  if (data && data.error) throw new Error(data.error);
  return data;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('range-7').addEventListener('click', () => applyQuickRange(7));
document.getElementById('range-14').addEventListener('click', () => applyQuickRange(14));
document.getElementById('range-31').addEventListener('click', () => applyQuickRange(31));
document.getElementById('article').addEventListener('input', scheduleLoad);
document.getElementById('article').addEventListener('keydown', (e) => {{ if (e.key === 'Enter') loadData(); }});
subjectEl.addEventListener('change', () => {{
  selectedArticles.clear();
  articlesLoadedFor = '';
  saveFilters();
  loadData();
}});
strategyEl.addEventListener('change', () => {{
  selectedArticles.clear();
  articlesLoadedFor = '';
  saveFilters();
  loadData();
}});
revenueCategoryEl.addEventListener('change', () => {{
  selectedArticles.clear();
  articlesLoadedFor = '';
  saveFilters();
  loadData();
}});
document.getElementById('df').addEventListener('change', () => {{ subjectsLoadedFor = ''; articlesLoadedFor = ''; loadData(); }});
document.getElementById('dt').addEventListener('change', () => {{ subjectsLoadedFor = ''; articlesLoadedFor = ''; loadData(); }});
document.getElementById('select-all').addEventListener('click', async () => {{
  selectedArticles.clear();
  resetArticleFilters();
  articlesLoadedFor = '';
  saveFilters();
  await loadArticles();
  loadData();
}});
document.getElementById('clear-all').addEventListener('click', () => {{
  selectedArticles.clear();
  resetArticleFilters();
  articlesLoadedFor = '';
  saveFilters();
  loadArticles();
  loadData();
}});
document.getElementById('toggle-articles').addEventListener('click', () => {{
  articlesCollapsed = !articlesCollapsed;
  syncArticlesPanel();
  saveFilters();
}});
document.getElementById('expand-articles').addEventListener('click', () => {{
  articlesCollapsed = false;
  syncArticlesPanel();
  saveFilters();
}});
let _stickyClone = null;
let _stickyPageSL = null;
let _stickyTblSL = null;
function setupStickyHeader() {{
  if (_stickyClone) {{ _stickyClone.remove(); _stickyClone = null; }}
  if (_stickyPageSL) window.removeEventListener('scroll', _stickyPageSL);
  if (_stickyTblSL) tbl.removeEventListener('scroll', _stickyTblSL);
  const origRow = tbl.querySelector('thead tr');
  if (!origRow) return;
  const wrap = document.createElement('div');
  wrap.style.cssText = 'position:fixed;top:0;z-index:150;overflow:hidden;display:none;pointer-events:none;';
  const cloneTable = document.createElement('table');
  cloneTable.style.cssText = 'border-collapse:collapse;font-size:11px;table-layout:fixed;';
  const cloneThead = document.createElement('thead');
  const cloneRow = origRow.cloneNode(true);
  cloneThead.appendChild(cloneRow);
  cloneTable.appendChild(cloneThead);
  wrap.appendChild(cloneTable);
  document.body.appendChild(wrap);
  _stickyClone = wrap;
  function syncWidths() {{
    const tblRect = tbl.getBoundingClientRect();
    wrap.style.left = tblRect.left + 'px';
    wrap.style.width = tblRect.width + 'px';
    const orig = origRow.querySelectorAll('th');
    const clone = cloneRow.querySelectorAll('th');
    orig.forEach((th, i) => {{
      if (!clone[i]) return;
      const w = th.getBoundingClientRect().width;
      clone[i].style.cssText = th.style.cssText + `;width:${{w}}px;min-width:${{w}}px;max-width:${{w}}px;background:#f8fafc;border-bottom:1px solid #eef2f7;padding:6px 7px;font-size:11px;text-align:${{getComputedStyle(th).textAlign}};white-space:nowrap;`;
    }});
    cloneTable.style.transform = `translateX(-${{tbl.scrollLeft}}px)`;
  }}
  function update() {{
    const r = origRow.getBoundingClientRect();
    if (r.bottom <= 0) {{ syncWidths(); wrap.style.display = 'block'; }}
    else wrap.style.display = 'none';
  }}
  _stickyPageSL = () => update();
  _stickyTblSL = () => {{ if (wrap.style.display !== 'none') cloneTable.style.transform = `translateX(-${{tbl.scrollLeft}}px)`; }};
  window.addEventListener('scroll', _stickyPageSL, {{passive: true}});
  tbl.addEventListener('scroll', _stickyTblSL, {{passive: true}});
  update();
}}
restoreFilters();
loadData();
</script>
</body>
</html>
"""

ANALYTICS_DAY_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Аналитика по дням</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0369a1; --accent-2: #0891b2; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #cffafe, transparent 240px), var(--bg); }}
    .wrap {{ width: min(100% - 28px, 1800px); margin: 0 auto; padding: 20px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .84rem; margin-bottom: 4px; font-weight: 700; }}
    input {{ padding: 9px 10px; border: 1px solid #cbd5e1; border-radius: 8px; font-size: 14px; }}
    button {{ padding: 10px 14px; border: 0; border-radius: 8px; background: linear-gradient(135deg, var(--accent), var(--accent-2)); color: #fff; font-weight: 700; cursor: pointer; }}
    .kpis {{ margin-top: 12px; display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 10px; }}
    .kpi {{ background: #fff; border: 1px solid #e2e8f0; border-radius: 12px; padding: 12px; }}
    .kpi .name {{ color: var(--muted); font-size: .78rem; font-weight: 700; text-transform: uppercase; letter-spacing: .4px; }}
    .kpi .val {{ margin-top: 6px; font-size: 1.15rem; font-weight: 800; }}
    .meta {{ margin-top: 10px; color: var(--muted); font-size: .9rem; }}
    .tbl {{ margin-top: 12px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: auto; max-height: calc(100vh - 320px); }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid #eef2f7; padding: 8px 10px; text-align: left; white-space: nowrap; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1; }}
    th {{ position: sticky; top: 0; background: #f8fafc; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>Аналитика по дням</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="df">Дата с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Дата по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="kpis">
      <div class="kpi"><div class="name">Дней</div><div class="val" id="kpi-days">0</div></div>
      <div class="kpi"><div class="name">Продажи</div><div class="val" id="kpi-sales">0.00</div></div>
      <div class="kpi"><div class="name">Реклама</div><div class="val" id="kpi-ads">0.00</div></div>
      <div class="kpi"><div class="name">Чистая прибыль</div><div class="val" id="kpi-profit">0.00</div></div>
      <div class="kpi"><div class="name">ДРР</div><div class="val" id="kpi-drr">0%</div></div>
      <div class="kpi"><div class="name">% маржи</div><div class="val" id="kpi-margin">0%</div></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="tbl" id="tbl"></div>
  </div>
<script>
const cols = ["Дата", "Продажи", "Реклама", "Чистая прибыль", "ДРР", "% маржи"];
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const filterKey = 'wb.analytics.day.filters';
let requestSeq = 0;

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
  }} catch (err) {{}}
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value
  }}));
}}

function render(rows) {{
  const isNumericLike = (value) => /^-?\d[\d\s.,%]*$/.test(String(value ?? '').trim());
  const head = '<tr>' + cols.map(c => `<th>${{c}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.map(r => '<tr>' + cols.map(c => {{
        const value = r[c] ?? '';
        const cls = isNumericLike(value) ? ' class="num"' : '';
        return `<td${{cls}}>${{value}}</td>`;
      }}).join('') + '</tr>').join('')
    : `<tr><td colspan="${{cols.length}}">Нет данных за выбранный период</td></tr>`;
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
}}

function renderSummary(summary) {{
  document.getElementById('kpi-days').textContent = summary.days || '0';
  document.getElementById('kpi-sales').textContent = summary.sales || '0.00';
  document.getElementById('kpi-ads').textContent = summary.ads || '0.00';
  document.getElementById('kpi-profit').textContent = summary.profit || '0.00';
  document.getElementById('kpi-drr').textContent = summary.drr || '0%';
  document.getElementById('kpi-margin').textContent = summary.margin || '0%';
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  meta.textContent = 'Загружаю...';
  const res = await fetch(`/api/analytics/day?date_from=${{encodeURIComponent(df)}}&date_to=${{encodeURIComponent(dt)}}`);
  const data = await res.json();
  if (seq !== requestSeq) return;
  renderSummary(data.summary || {{}});
  render(data.rows || []);
  meta.textContent = `Строк: ${{(data.rows || []).length}}`;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('df').addEventListener('change', loadData);
document.getElementById('dt').addEventListener('change', loadData);
restoreFilters();
loadData();
</script>
</body>
</html>
"""

ANALYTICS_PERIOD_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Период по артикулам</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0f766e; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #e0f2fe, transparent 240px), var(--bg); }}
    .wrap {{ width: min(100% - 28px, 1800px); margin: 0 auto; padding: 20px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .84rem; margin-bottom: 4px; font-weight: 700; }}
    input {{ padding: 9px 10px; border: 1px solid #cbd5e1; border-radius: 8px; font-size: 14px; }}
    button {{ padding: 10px 14px; border: 0; border-radius: 8px; background: var(--accent); color: #fff; font-weight: 700; cursor: pointer; }}
    .meta {{ margin-top: 10px; color: var(--muted); font-size: .9rem; }}
    .tbl {{ margin-top: 12px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: auto; max-height: calc(100vh - 230px); }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid #eef2f7; padding: 8px 10px; text-align: left; white-space: nowrap; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1; }}
    th {{ position: sticky; top: 0; background: #f8fafc; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>Период по артикулам</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="article">Артикул</label>
        <input id="article" type="text" placeholder="например 12-0040-019">
      </div>
      <div>
        <label for="df">Период с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Период по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="tbl" id="tbl"></div>
  </div>
<script>
const cols = ["Артикул", "Продажи по нашей цене", "Реклама", "Чистая прибыль", "ДРР", "% маржи"];
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const filterKey = 'wb.analytics.period.filters';
let requestSeq = 0;
let filterTimer = null;

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.article !== undefined) document.getElementById('article').value = saved.article;
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
  }} catch (err) {{}}
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    article: document.getElementById('article').value,
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value
  }}));
}}

function scheduleLoad() {{
  clearTimeout(filterTimer);
  filterTimer = setTimeout(loadData, 350);
}}

function render(rows) {{
  const isNumericLike = (value) => /^-?\d[\d\s.,%]*$/.test(String(value ?? '').trim());
  const head = '<tr>' + cols.map(c => `<th>${{c}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.map(r => '<tr>' + cols.map(c => {{
        const value = r[c] ?? '';
        const cls = isNumericLike(value) ? ' class="num"' : '';
        return `<td${{cls}}>${{value}}</td>`;
      }}).join('') + '</tr>').join('')
    : `<tr><td colspan="${{cols.length}}">Нет данных за выбранный период</td></tr>`;
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const article = document.getElementById('article').value;
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  meta.textContent = 'Загружаю...';
  const qs = new URLSearchParams({{ article, date_from: df, date_to: dt}});
  const res = await fetch(`/api/analytics/period?${{qs.toString()}}`);
  const data = await res.json();
  if (seq !== requestSeq) return;
  render(data.rows || []);
  meta.textContent = `Строк: ${{(data.rows || []).length}}`;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('article').addEventListener('input', scheduleLoad);
document.getElementById('article').addEventListener('keydown', (e) => {{ if (e.key === 'Enter') loadData(); }});
document.getElementById('df').addEventListener('change', loadData);
document.getElementById('dt').addEventListener('change', loadData);
restoreFilters();
loadData();
</script>
</body>
</html>
"""

ANALYTICS_ARTICLE_DAY_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Дневная аналитика артикула</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0f766e; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #dcfce7, transparent 240px), var(--bg); }}
    .wrap {{ width: min(100% - 28px, 1800px); margin: 0 auto; padding: 20px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .84rem; margin-bottom: 4px; font-weight: 700; }}
    input {{ padding: 9px 10px; border: 1px solid #cbd5e1; border-radius: 8px; font-size: 14px; min-width: 180px; }}
    button {{ padding: 10px 14px; border: 0; border-radius: 8px; background: var(--accent); color: #fff; font-weight: 700; cursor: pointer; }}
    .meta {{ margin-top: 10px; color: var(--muted); font-size: .9rem; }}
    .tbl {{ margin-top: 12px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: auto; max-height: calc(100vh - 230px); }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid #eef2f7; padding: 8px 10px; text-align: left; white-space: nowrap; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1; }}
    th {{ position: sticky; top: 0; background: #f8fafc; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>Дни по артикулу</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="article">Артикул</label>
        <input id="article" type="text" placeholder="например 12-0040-019">
      </div>
      <div>
        <label for="df">Дата с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Дата по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="tbl" id="tbl"></div>
  </div>
<script>
const cols = [
  "Артикул / SKU",
  "Дата",
  "Продажи по нашей цене",
  "Реклама",
  "Чистая прибыль",
  "ДРР",
  "% маржи"
];
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const filterKey = 'wb.analytics.articleDay.filters';
let requestSeq = 0;
let filterTimer = null;

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.article !== undefined) document.getElementById('article').value = saved.article;
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
  }} catch (err) {{}}
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    article: document.getElementById('article').value,
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value
  }}));
}}

function scheduleLoad() {{
  clearTimeout(filterTimer);
  filterTimer = setTimeout(loadData, 350);
}}

function render(rows) {{
  const isNumericLike = (value) => /^-?\d[\d\s.,%]*$/.test(String(value ?? '').trim());
  const head = '<tr>' + cols.map(c => `<th>${{c}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.map(r => '<tr>' + cols.map(c => {{
        const value = r[c] ?? '';
        const cls = isNumericLike(value) ? ' class="num"' : '';
        return `<td${{cls}}>${{value}}</td>`;
      }}).join('') + '</tr>').join('')
    : `<tr><td colspan="${{cols.length}}">Нет данных по фильтру</td></tr>`;
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const article = document.getElementById('article').value;
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  meta.textContent = 'Загружаю...';
  const qs = new URLSearchParams({{ article, date_from: df, date_to: dt}});
  const res = await fetch(`/api/analytics/article-day?${{qs.toString()}}`);
  const data = await res.json();
  if (seq !== requestSeq) return;
  render(data.rows || []);
  meta.textContent = `Строк: ${{(data.rows || []).length}}`;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('article').addEventListener('input', scheduleLoad);
document.getElementById('article').addEventListener('keydown', (e) => {{ if (e.key === 'Enter') loadData(); }});
document.getElementById('df').addEventListener('change', loadData);
document.getElementById('dt').addEventListener('change', loadData);
restoreFilters();
loadData();
</script>
</body>
</html>
"""

PRELIMINARY_ECONOMICS_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Предварительная экономика по заказам</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0f766e; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #e0f2fe, transparent 240px), var(--bg); }}
    .wrap {{ width: min(100% - 28px, 1800px); margin: 0 auto; padding: 20px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .84rem; margin-bottom: 4px; font-weight: 700; }}
    input {{ padding: 9px 10px; border: 1px solid #cbd5e1; border-radius: 8px; font-size: 14px; }}
    button {{ padding: 10px 14px; border: 0; border-radius: 8px; background: var(--accent); color: #fff; font-weight: 700; cursor: pointer; }}
    .meta {{ margin-top: 10px; color: var(--muted); font-size: .9rem; }}
    .tbl {{ margin-top: 12px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: auto; max-height: calc(100vh - 230px); }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid #eef2f7; padding: 8px 10px; text-align: left; white-space: nowrap; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1; }}
    th {{ position: sticky; top: 0; background: #f8fafc; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>Предварительная экономика по заказам</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="article">Артикул</label>
        <input id="article" type="text" placeholder="например 12-0040-019">
      </div>
      <div>
        <label for="buyout">% выкупа</label>
        <input id="buyout" type="number" min="0" max="100" step="0.1" value="30">
      </div>
      <div>
        <label for="df">Период с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Период по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="tbl" id="tbl"></div>
  </div>
<script>
const cols = [
  "Артикул / SKU",
  "Дата",
  "Количество заказов",
  "Сумма заказов",
  "% выкупа",
  "Комиссия, ₽",
  "Эквайринг, ₽",
  "Реклама, ₽",
  "% рекламы",
  "Дополнительные расходы, ₽",
  "Предварительная прибыль, ₽"
  ,"% маржинальности"
];
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const filterKey = 'wb.analytics.preliminary.filters';
let requestSeq = 0;
let filterTimer = null;

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.article !== undefined) document.getElementById('article').value = saved.article;
    if (saved.buyout !== undefined) document.getElementById('buyout').value = saved.buyout;
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
  }} catch (err) {{}}
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    article: document.getElementById('article').value,
    buyout: document.getElementById('buyout').value,
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value
  }}));
}}

function scheduleLoad() {{
  clearTimeout(filterTimer);
  filterTimer = setTimeout(loadData, 350);
}}

function render(rows) {{
  const isNumericLike = (value) => /^-?\d[\d\s.,%]*$/.test(String(value ?? '').trim());
  const head = '<tr>' + cols.map(c => `<th>${{c}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.map(r => '<tr>' + cols.map(c => {{
        const value = r[c] ?? '';
        const cls = isNumericLike(value) ? ' class="num"' : '';
        return `<td${{cls}}>${{value}}</td>`;
      }}).join('') + '</tr>').join('')
    : `<tr><td colspan="${{cols.length}}">Нет данных за выбранный период</td></tr>`;
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const article = document.getElementById('article').value;
  const buyout = document.getElementById('buyout').value || '30';
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  meta.textContent = 'Загружаю...';
  const qs = new URLSearchParams({{ article, date_from: df, date_to: dt, buyout_percent: buyout}});
  const res = await fetch(`/api/analytics/preliminary-economics?${{qs.toString()}}`);
  const data = await res.json();
  if (seq !== requestSeq) return;
  render(data.rows || []);
  const extra = (data.expense_components || []).length
    ? ` | Статьи допрасходов: ${{data.expense_components.join(', ')}}`
    : '';
  meta.textContent = `Строк: ${{(data.rows || []).length}} | % выкупа: ${{data.buyout_percent || '30.00%'}} | Зафиксированный % допрасходов: ${{data.additional_rate || '0.00%'}}${{extra}}`;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('article').addEventListener('input', scheduleLoad);
document.getElementById('article').addEventListener('keydown', (e) => {{ if (e.key === 'Enter') loadData(); }});
document.getElementById('buyout').addEventListener('input', scheduleLoad);
document.getElementById('df').addEventListener('change', loadData);
document.getElementById('dt').addEventListener('change', loadData);
restoreFilters();
loadData();
</script>
</body>
</html>
"""

PRELIMINARY_ECONOMICS_SUMMARY_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Предварительная экономика по заказам (период)</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0f766e; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #e0f2fe, transparent 240px), var(--bg); }}
    .wrap {{ width: min(100% - 28px, 1800px); margin: 0 auto; padding: 20px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .84rem; margin-bottom: 4px; font-weight: 700; }}
    input {{ padding: 9px 10px; border: 1px solid #cbd5e1; border-radius: 8px; font-size: 14px; }}
    button {{ padding: 10px 14px; border: 0; border-radius: 8px; background: var(--accent); color: #fff; font-weight: 700; cursor: pointer; }}
    .meta {{ margin-top: 10px; color: var(--muted); font-size: .9rem; }}
    .tbl {{ margin-top: 12px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: auto; max-height: calc(100vh - 230px); }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid #eef2f7; padding: 8px 10px; text-align: left; white-space: nowrap; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1; }}
    th {{ position: sticky; top: 0; background: #f8fafc; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>Предварительная экономика по периоду</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="article">Артикул</label>
        <input id="article" type="text" placeholder="например 12-0040-019">
      </div>
      <div>
        <label for="buyout">% выкупа</label>
        <input id="buyout" type="number" min="0" max="100" step="0.1" value="30">
      </div>
      <div>
        <label for="df">Период с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Период по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="tbl" id="tbl"></div>
  </div>
<script>
const cols = [
  "Артикул / SKU",
  "Период",
  "Количество заказов",
  "Сумма заказов",
  "% выкупа",
  "Комиссия, ₽",
  "Эквайринг, ₽",
  "Реклама, ₽",
  "% рекламы",
  "Дополнительные расходы, ₽",
  "Предварительная прибыль, ₽",
  "% маржинальности"
];
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const filterKey = 'wb.analytics.preliminarySummary.filters';
let requestSeq = 0;
let filterTimer = null;

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.article !== undefined) document.getElementById('article').value = saved.article;
    if (saved.buyout !== undefined) document.getElementById('buyout').value = saved.buyout;
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
  }} catch (err) {{}}
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    article: document.getElementById('article').value,
    buyout: document.getElementById('buyout').value,
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value
  }}));
}}

function scheduleLoad() {{
  clearTimeout(filterTimer);
  filterTimer = setTimeout(loadData, 350);
}}

function render(rows) {{
  const isNumericLike = (value) => /^-?\d[\d\s.,%]*$/.test(String(value ?? '').trim());
  const head = '<tr>' + cols.map(c => `<th>${{c}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.map(r => '<tr>' + cols.map(c => {{
        const value = r[c] ?? '';
        const cls = isNumericLike(value) ? ' class="num"' : '';
        return `<td${{cls}}>${{value}}</td>`;
      }}).join('') + '</tr>').join('')
    : `<tr><td colspan="${{cols.length}}">Нет данных за выбранный период</td></tr>`;
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const article = document.getElementById('article').value;
  const buyout = document.getElementById('buyout').value || '30';
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  meta.textContent = 'Загружаю...';
  const qs = new URLSearchParams({{ article, date_from: df, date_to: dt, buyout_percent: buyout, aggregate: '1'}});
  const res = await fetch(`/api/analytics/preliminary-economics?${{qs.toString()}}`);
  const data = await res.json();
  if (seq !== requestSeq) return;
  render(data.rows || []);
  const extra = (data.expense_components || []).length
    ? ` | Статьи допрасходов: ${{data.expense_components.join(', ')}}`
    : '';
  meta.textContent = `Строк: ${{(data.rows || []).length}} | % выкупа: ${{data.buyout_percent || '30.00%'}} | Зафиксированный % допрасходов: ${{data.additional_rate || '0.00%'}}${{extra}}`;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('article').addEventListener('input', scheduleLoad);
document.getElementById('article').addEventListener('keydown', (e) => {{ if (e.key === 'Enter') loadData(); }});
document.getElementById('buyout').addEventListener('input', scheduleLoad);
document.getElementById('df').addEventListener('change', loadData);
document.getElementById('dt').addEventListener('change', loadData);
restoreFilters();
loadData();
</script>
</body>
</html>
"""

PLANNING_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Планирование</title>
  <style>
    :root {{ --bg:#f8fafc; --card:#fff; --ink:#111827; --muted:#6b7280; --accent:#0f766e; --line:#e5e7eb; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:"Segoe UI","Trebuchet MS",sans-serif; color:var(--ink); background:linear-gradient(180deg,#e0f2fe,transparent 240px),var(--bg); }}
    .wrap {{ width:calc(100% - 32px); max-width:3000px; margin:0 auto; padding:10px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .ctrl-panel {{ background:var(--card); border:1px solid var(--line); border-radius:10px; padding:14px 16px; display:flex; flex-direction:column; gap:10px; }}
    .ctrl-row {{ display:flex; flex-wrap:wrap; align-items:flex-end; gap:8px; }}
    .ctrl-field {{ display:flex; flex-direction:column; gap:3px; }}
    .ctrl-field label {{ font-size:11px; color:#64748b; font-weight:600; text-transform:uppercase; letter-spacing:.04em; white-space:nowrap; }}
    .ctrl-field input, .ctrl-field select {{ height:32px; padding:0 8px; border:1px solid #cbd5e1; border-radius:6px; font-size:13px; background:#f8fafc; color:#0f172a; min-width:0; }}
    .ctrl-field select {{ padding-right:24px; }}
    .ctrl-sep {{ width:1px; align-self:stretch; background:var(--line); margin:0 4px; flex-shrink:0; }}
    .ctrl-group {{ display:flex; align-items:flex-end; gap:6px; }}
    .ctrl-group-label {{ font-size:10px; font-weight:700; text-transform:uppercase; letter-spacing:.06em; color:#94a3b8; margin-bottom:4px; }}
    .ctrl-btn {{ height:32px; padding:0 12px; border-radius:6px; border:none; cursor:pointer; font-size:13px; font-weight:500; color:#fff; white-space:nowrap; transition:filter .15s; }}
    .ctrl-btn:hover {{ filter:brightness(.88); }}
    .ctrl-btn:disabled {{ opacity:.5; cursor:default; }}
    .ctrl-btn-primary {{ background:#2563eb; }}
    .ctrl-btn-teal   {{ background:#0f766e; }}
    .ctrl-btn-purple {{ background:#7c3aed; }}
    .ctrl-btn-gray   {{ background:#64748b; }}
    .ctrl-btn-red    {{ background:#b91c1c; }}
    .ctrl-btn-green  {{ background:#065f46; }}
    .ctrl-btn-slate  {{ background:#334155; }}
    .ctrl-scenario-select {{ height:32px; padding:0 6px; border:1px solid #cbd5e1; border-radius:6px; font-size:12px; background:#f8fafc; color:#0f172a; max-width:130px; }}
    .ctrl-scenario-input  {{ height:32px; padding:0 8px; border:1px solid #cbd5e1; border-radius:6px; font-size:12px; background:#f8fafc; color:#0f172a; width:110px; }}
    .ctrl-tax-input {{ height:32px; width:60px; padding:0 8px; border:1px solid #cbd5e1; border-radius:6px; font-size:13px; background:#f8fafc; color:#0f172a; }}
    .panel {{ background:var(--card); border:1px solid var(--line); border-radius:8px; padding:12px; display:flex; flex-wrap:wrap; gap:10px; align-items:end; }}
    label {{ display:block; color:var(--muted); font-size:.72rem; margin-bottom:3px; font-weight:700; }}
    input, select {{ padding:7px 8px; border:1px solid #cbd5e1; border-radius:7px; font-size:12px; min-width:150px; background:#fff; }}
    button {{ padding:8px 12px; border:0; border-radius:7px; background:var(--accent); color:#fff; font-weight:700; cursor:pointer; font-size:12px; }}
    .meta {{ margin:8px 0 0; color:var(--muted); font-size:.78rem; }}
    .summary {{ display:flex; flex-wrap:nowrap; gap:10px; margin-top:10px; overflow-x:auto; }}
    .summary .kpi {{ flex:1 1 0; min-width:0; }}
    .kpi {{ background:#fff; border:1px solid var(--line); border-radius:8px; padding:10px 12px; }}
    .kpi-name {{ color:var(--muted); font-size:.72rem; font-weight:700; text-transform:uppercase; letter-spacing:.3px; }}
    .kpi-rub {{ margin-top:5px; font-size:1.15rem; font-weight:900; color:#0f172a; }}
    .kpi-qty {{ margin-top:2px; font-size:.82rem; font-weight:600; color:#475569; }}
    .tbl {{ margin-top:10px; background:#fff; border:1px solid var(--line); border-radius:8px; overflow:auto; max-height:calc(100vh - 260px); }}
    table {{ width:max-content; min-width:100%; border-collapse:collapse; font-size:11px; }}
    th, td {{ border-bottom:1px solid #eef2f7; padding:5px 6px; text-align:right; }}
    th {{ position:sticky; top:0; z-index:1; background:#f8fafc; color:#334155; font-weight:800; white-space:normal; word-break:normal; vertical-align:bottom; line-height:1.35; }}
    td {{ white-space:nowrap; }}
    .col-ck {{ width:22px; min-width:22px; max-width:22px; padding:4px 2px; text-align:center; position:sticky; left:0; z-index:2; background:#fff; }}
    th.col-ck {{ z-index:3; background:#f8fafc; }}
    .col-sku {{ width:88px; min-width:60px; max-width:100px; text-align:left; position:sticky; left:22px; z-index:2; background:#fff; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    th.col-sku {{ z-index:3; background:#f8fafc; }}
    .col-subject {{ text-align:left; min-width:54px; max-width:110px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    th.col-subject {{ white-space:normal; max-width:80px; }}
    .col-text {{ text-align:left; white-space:nowrap; }}
    tbody tr:hover td {{ background:#f0fdfa !important; }}
    tbody tr:hover td.col-ck, tbody tr:hover td.col-sku {{ background:#f0fdfa !important; }}
    tbody tr.row-selected td {{ background:#d1fae5 !important; }}
    tbody tr.row-selected td.col-ck, tbody tr.row-selected td.col-sku {{ background:#d1fae5 !important; }}
    .empty {{ padding:18px; color:var(--muted); font-size:.9rem; }}
    th.g-sep {{ border-left:2px solid #94a3b8; }}
    td.g-sep {{ border-left:2px solid #e2e8f0; }}
    .price-up {{ color:#16a34a; }}
    .price-down {{ color:#dc2626; }}
    .price-arrow {{ font-size:9px; margin-left:2px; }}
    @media (max-width:1200px) {{ .summary {{ flex-wrap:nowrap; }} }}
    .ck-cell {{ text-align:center; width:26px; min-width:26px; padding:4px 2px; }}
    input[type=checkbox] {{ min-width:auto; width:14px; height:14px; cursor:pointer; padding:0; accent-color:var(--accent); }}
    .pct-cell {{ display:flex; justify-content:flex-end; }}
    .pct-spinner {{ display:inline-flex; border:1px solid #cbd5e1; border-radius:5px; overflow:hidden; height:22px; }}
    .pct-spinner input[type=number] {{ width:42px; min-width:0; border:none; padding:2px 4px; text-align:center; font-size:11px; -moz-appearance:textfield; background:#fff; }}
    .pct-spinner input[type=number]::-webkit-inner-spin-button {{ display:none; }}
    .pct-spinner input[type=number]:focus {{ outline:none; }}
    .pct-arrows {{ display:flex; flex-direction:column; border-left:1px solid #cbd5e1; }}
    .pct-btn, .drr-btn, .conv-btn {{ width:15px; padding:0; font-size:7px; background:#f8fafc; color:#475569; border:none; border-radius:0; flex:1; line-height:1; cursor:pointer; min-width:0; }}
    .pct-btn:first-child, .drr-btn:first-child, .conv-btn:first-child {{ border-bottom:1px solid #e2e8f0; }}
    .pct-btn:hover, .drr-btn:hover, .conv-btn:hover {{ background:#e2e8f0; }}
    .pct-cell {{ gap:3px; align-items:center; }}
    .val-arrow {{ font-size:9px; font-weight:700; min-width:14px; width:14px; text-align:center; flex-shrink:0; }}
    .val-arrow:not(:empty) {{ cursor:pointer; padding:0 2px; }}
    .val-arrow:not(:empty):hover {{ opacity:0.6; }}
    .pct-spinner.no-data {{ opacity:0.35; pointer-events:none; }}
    .pct-spinner input:disabled {{ color:#94a3b8; background:#f8fafc; text-align:center; }}
    .val-arrow.val-up {{ color:#16a34a; }}
    .val-arrow.val-down {{ color:#dc2626; }}
    input.val-changed-up {{ color:#16a34a; font-weight:700; }}
    input.val-changed-down {{ color:#dc2626; font-weight:700; }}
    thead tr:first-child th:not([rowspan]) {{ font-size:9px; text-transform:uppercase; letter-spacing:.04em; padding:3px 8px; }}
    thead tr:last-child th {{ top:22px; z-index:1; }}
    th.col-ck, th.col-sku {{ top:0 !important; z-index:4 !important; }}
    .th-params {{ background:#dbeafe !important; color:#1d4ed8; }}
    .th-result {{ background:#dcfce7 !important; color:#166534; }}
    .th-lost  {{ background:#fef9c3 !important; color:#854d0e; }}
    th.params-col {{ background:#eff6ff; }}
    th.result-col {{ background:#f0fdf4; }}
    th.lost-col   {{ background:#fefce8; }}
    .th-econ {{ background:#fdf4ff !important; color:#7e22ce; }}
    th.econ-col {{ background:#fdf4ff; }}
    td.econ-col {{ background:#fdf4ff44; }}
    th[data-sort-key] {{ cursor:pointer; user-select:none; }}
    th[data-sort-key]:hover {{ filter:brightness(0.93); }}
    th[data-sort-key][data-sort-dir="asc"]::after  {{ content:" ▲"; font-size:.65em; opacity:.7; }}
    th[data-sort-key][data-sort-dir="desc"]::after {{ content:" ▼"; font-size:.65em; opacity:.7; }}
    #bulk-bar {{ display:none; position:fixed; bottom:22px; left:50%; transform:translateX(-50%); background:#0f172a; color:#fff; border-radius:10px; padding:10px 16px; align-items:center; gap:10px; font-size:12px; box-shadow:0 4px 24px rgba(0,0,0,.35); z-index:100; white-space:nowrap; }}
    #bulk-bar input[type=number] {{ width:64px; min-width:0; padding:5px 7px; color:#0f172a; font-size:12px; }}
    #bulk-bar .apply-btn {{ background:#0f766e; }}
    #bulk-bar .reset-btn {{ background:transparent; border:1px solid #475569; color:#94a3b8; }}
    #bulk-bar .close-btn {{ background:transparent; border:none; color:#94a3b8; font-size:14px; padding:2px 6px; cursor:pointer; }}
    .methodology {{ margin-top:18px; background:#fff; border:1px solid var(--line); border-radius:8px; padding:16px 20px; font-size:.8rem; line-height:1.6; color:#334155; }}
    .methodology h2 {{ margin:0 0 12px; font-size:.88rem; font-weight:800; color:#0f172a; }}
    .methodology h3 {{ margin:14px 0 4px; font-size:.78rem; font-weight:800; color:var(--accent); text-transform:uppercase; letter-spacing:.04em; }}
    .methodology p {{ margin:0 0 6px; }}
    .methodology ul {{ margin:2px 0 6px; padding-left:18px; }}
    .methodology li {{ margin-bottom:2px; }}
    .methodology .formula {{ font-family:monospace; background:#f1f5f9; border-radius:4px; padding:2px 6px; font-size:.77rem; }}
  </style>
</head>
<body>
<div class="wrap">
  <div class="top">
    <h1>Планирование</h1>
    <div class="report-nav">{report_nav}</div>
  </div>
  <div class="ctrl-panel">
    <!-- Ряд 1: параметры -->
    <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#94a3b8;margin-bottom:2px">Параметры расчёта</div>
    <div class="ctrl-row">
      <div class="ctrl-field">
        <label for="forecast-to">Прогноз до</label>
        <input id="forecast-to" type="date" value="{forecast_to}" style="width:130px">
      </div>
      <div class="ctrl-field">
        <label for="return-delay">Возврат, дней</label>
        <input id="return-delay" type="number" min="1" max="60" step="1" value="8" style="width:72px">
      </div>
      <div class="ctrl-field">
        <label for="transit-days">Поставка WB, дней</label>
        <input id="transit-days" type="number" min="0" max="30" step="1" value="3" style="width:72px">
      </div>
      <div class="ctrl-field">
        <label for="tax-pct">Налоги, %</label>
        <input id="tax-pct" type="number" min="0" max="50" step="0.1" value="0" style="width:72px">
      </div>
      <div class="ctrl-field">
        <label for="admin-pct">Адм. расходы, %</label>
        <input id="admin-pct" type="number" min="0" max="50" step="0.1" value="0" style="width:72px">
      </div>
      <div class="ctrl-sep"></div>
      <div>
        <div class="ctrl-group-label">Фильтры</div>
        <div style="display:flex;gap:6px;align-items:flex-end">
          <div class="ctrl-field">
            <label for="subject">Предмет</label>
            <select id="subject" style="width:160px"><option value="">Все предметы</option></select>
          </div>
          <div class="ctrl-field">
            <label for="strategy">Ярлыки</label>
            <select id="strategy" multiple size="4" style="width:140px;height:72px;" title="Ctrl/Cmd для множественного выбора"></select>
          </div>
          <div class="ctrl-field">
            <label for="revenue-category">Категория</label>
            <select id="revenue-category" style="width:140px"><option value="">Все категории</option></select>
          </div>
          <div class="ctrl-field">
            <label for="weeks-filter">Недель, макс.</label>
            <input id="weeks-filter" type="number" min="1" max="52" step="1" placeholder="Все" style="width:72px">
          </div>
        </div>
      </div>
      <button id="load" class="ctrl-btn ctrl-btn-primary" style="align-self:flex-end">Посчитать</button>
    </div>
    <!-- Ряд 2: действия -->
    <div class="ctrl-row" style="border-top:1px solid var(--line);padding-top:10px">
      <!-- Оптимизация -->
      <div class="ctrl-group">
        <div>
          <div class="ctrl-group-label">Оптимизация</div>
          <div style="display:flex;gap:6px">
            <button id="no-oos-btn" class="ctrl-btn ctrl-btn-teal" title="Подобрать минимальное повышение цены, чтобы не уйти в OOS до конца периода">Без OOS</button>
            <button id="optimize-btn" class="ctrl-btn ctrl-btn-purple" title="Найти цену и ДРР для максимального оборота при марже не ниже 0%">⚡ Макс. выручку</button>
          </div>
        </div>
      </div>
      <div class="ctrl-sep"></div>
      <!-- Сброс -->
      <div class="ctrl-group">
        <div>
          <div class="ctrl-group-label">Сбросить</div>
          <div style="display:flex;gap:6px">
            <button id="reset-prices" class="ctrl-btn ctrl-btn-gray">Цены</button>
            <button id="reset-drr"    class="ctrl-btn ctrl-btn-gray">ДРР</button>
            <button id="reset-conv"   class="ctrl-btn ctrl-btn-gray">Конверсии</button>
          </div>
        </div>
      </div>
      <div class="ctrl-sep"></div>
      <!-- Сценарии -->
      <div class="ctrl-group">
        <div>
          <div class="ctrl-group-label">Сценарий</div>
          <div style="display:flex;gap:6px;align-items:center">
            <select id="scenario-select" class="ctrl-scenario-select"><option value="">— выбрать —</option></select>
            <button id="scenario-load"   class="ctrl-btn ctrl-btn-teal"  style="padding:0 10px;font-size:12px">Загрузить</button>
            <button id="scenario-delete" class="ctrl-btn ctrl-btn-red"   style="padding:0 10px;font-size:12px">Удалить</button>
            <input type="text" id="scenario-name" class="ctrl-scenario-input" placeholder="Новый сценарий…">
            <button id="scenario-save"   class="ctrl-btn ctrl-btn-primary" style="padding:0 10px;font-size:12px">Сохранить</button>
          </div>
        </div>
      </div>
      <div class="ctrl-sep"></div>
      <!-- Экспорт и прочее -->
      <div class="ctrl-group" style="margin-left:auto">
        <div>
          <div class="ctrl-group-label">Экспорт и вид</div>
          <div style="display:flex;gap:6px;align-items:center">
            <button id="export-xls" class="ctrl-btn ctrl-btn-green">⬇ XLS</button>
            <div style="position:relative">
              <button id="col-settings-btn" class="ctrl-btn ctrl-btn-slate" style="padding:0 10px;font-size:16px" title="Настройка столбцов">⚙</button>
              <div id="col-settings-panel" style="display:none;position:absolute;right:0;top:36px;background:#fff;border:1px solid #cbd5e1;border-radius:8px;padding:12px 16px;z-index:300;box-shadow:0 4px 16px rgba(0,0,0,.18);white-space:nowrap;min-width:160px"></div>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>
  <div class="meta" id="meta">Загрузка...</div>
  <div style="margin-top:18px;margin-bottom:6px;display:flex;align-items:center;gap:10px">
    <span style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#94a3b8">Общие метрики</span>
    <div style="flex:1;height:1px;background:var(--line)"></div>
  </div>
  <div class="summary" id="summary"></div>
  <div class="tbl" id="table"></div>
</div>
<div id="bulk-bar">
  <span>Выбрано: <strong id="bulk-count">0</strong></span>
  <select id="bulk-type" style="padding:4px 6px;font-size:12px;background:#1e293b;color:#fff;border:1px solid #475569;border-radius:4px">
    <option value="price">Цена, %</option>
    <option value="drr">ДРР, %</option>
    <option value="ctr">CTR, %</option>
    <option value="cr1">CR1, %</option>
    <option value="cr2">CR2, %</option>
  </select>
  <input type="number" id="bulk-pct" value="0" step="0.5" min="-90" max="300" placeholder="±δ">
  <button class="apply-btn" id="bulk-apply">+ Применить</button>
  <button class="reset-btn" id="bulk-reset">Сбросить</button>
  <button class="close-btn" id="bulk-close">✕</button>
</div>
<script>
const stateKey = 'wb.analytics.planning.filters';
const fields = {{
  forecastTo: document.getElementById('forecast-to'),
  returnDelay: document.getElementById('return-delay'),
  transitDays: document.getElementById('transit-days'),
  subject: document.getElementById('subject'),
  strategy: document.getElementById('strategy'),
  revenueCategory: document.getElementById('revenue-category'),
  weeksFilter: document.getElementById('weeks-filter'),
}};
const meta = document.getElementById('meta');
const summaryEl = document.getElementById('summary');
const tableEl = document.getElementById('table');
const bulkBar = document.getElementById('bulk-bar');
const bulkCount = document.getElementById('bulk-count');
const bulkPct = document.getElementById('bulk-pct');

let rowDataByNmid = {{}};
const priceChangesKey = 'wb.analytics.planning.priceChanges';
let priceChanges = {{}};
let rowForecastCache = {{}};
const drrChangesKey = 'wb.analytics.planning.drrChanges';
let drrChanges = {{}};  // nmid → delta_drr_pct (float, added to base_drr_pct)
const convChangesKey = 'wb.analytics.planning.convChanges';
let convChanges = {{}};  // nmid → {{ ctr, cr1, cr2 }} (absolute %, null = use base)

function savePriceChanges() {{
  try {{ localStorage.setItem(priceChangesKey, JSON.stringify(priceChanges)); }} catch(e) {{}}
}}
function loadPriceChanges() {{
  try {{ priceChanges = JSON.parse(localStorage.getItem(priceChangesKey) || '{{}}'); }} catch(e) {{ priceChanges = {{}}; }}
}}
function saveDrrChanges() {{
  try {{ localStorage.setItem(drrChangesKey, JSON.stringify(drrChanges)); }} catch(e) {{}}
}}
function loadDrrChanges() {{
  try {{ drrChanges = JSON.parse(localStorage.getItem(drrChangesKey) || '{{}}'); }} catch(e) {{ drrChanges = {{}}; }}
}}
function saveConvChanges() {{
  try {{ localStorage.setItem(convChangesKey, JSON.stringify(convChanges)); }} catch(e) {{}}
}}
function loadConvChanges() {{
  try {{ convChanges = JSON.parse(localStorage.getItem(convChangesKey) || '{{}}'); }} catch(e) {{ convChanges = {{}}; }}
  // чистим нули, сохранённые старой версией кода (|| 0 вместо delete)
  for (const nmid of Object.keys(convChanges)) {{
    for (const k of Object.keys(convChanges[nmid] || {{}})) {{
      if (!(convChanges[nmid][k] > 0)) delete convChanges[nmid][k];
    }}
    if (!Object.keys(convChanges[nmid] || {{}}).length) delete convChanges[nmid];
  }}
}}
loadPriceChanges();
loadDrrChanges();
loadConvChanges();

// ── Сценарии ──────────────────────────────────────────────────────────────
const scenariosKey = 'wb.analytics.planning.scenarios';
function loadScenarioList() {{
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  const sel = document.getElementById('scenario-select');
  const cur = sel.value;
  sel.innerHTML = '<option value="">— выбрать —</option>';
  Object.keys(all).sort().forEach(name => {{
    const opt = document.createElement('option');
    opt.value = name; opt.textContent = name;
    sel.appendChild(opt);
  }});
  if (cur && all[cur]) sel.value = cur;
}}
loadScenarioList();

document.getElementById('scenario-save').addEventListener('click', () => {{
  const name = document.getElementById('scenario-name').value.trim();
  if (!name) return;
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  all[name] = {{
    price: Object.assign({{}}, priceChanges),
    drr:   Object.assign({{}}, drrChanges),
    conv:  JSON.parse(JSON.stringify(convChanges)),
  }};
  localStorage.setItem(scenariosKey, JSON.stringify(all));
  loadScenarioList();
  document.getElementById('scenario-select').value = name;
  document.getElementById('scenario-name').value = '';
}});

document.getElementById('scenario-load').addEventListener('click', () => {{
  const name = document.getElementById('scenario-select').value;
  if (!name) return;
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  const sc = all[name];
  if (!sc) return;
  priceChanges = sc.price || {{}}; savePriceChanges();
  drrChanges   = sc.drr   || {{}}; saveDrrChanges();
  convChanges  = sc.conv  || {{}}; saveConvChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const pInp = tr.querySelector('.pct-input');
    if (pInp) pInp.value = priceChanges[nmid] ?? 0;
    const dInp = tr.querySelector('.drr-input');
    if (dInp) dInp.value = (drrChanges[nmid] != null ? drrChanges[nmid] : (row.base_drr_pct || 0)).toFixed(1);
    ['ctr','cr1','cr2'].forEach(k => {{
      const inp = tr.querySelector(`.conv-input[data-conv="${{k}}"]`);
      if (inp && !inp.disabled) {{
        const v = (convChanges[nmid] || {{}})[k];
        inp.value = (v > 0 ? v : (row[`base_${{k}}`] || 0)).toFixed(2);
      }}
    }});
    updateRowPlanned(nmid);
  }});
}});

document.getElementById('scenario-delete').addEventListener('click', () => {{
  const name = document.getElementById('scenario-select').value;
  if (!name) return;
  if (!confirm(`Удалить сценарий «${{name}}»?`)) return;
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  delete all[name];
  localStorage.setItem(scenariosKey, JSON.stringify(all));
  loadScenarioList();
}});

// ── Экспорт XLS ───────────────────────────────────────────────────────────
document.getElementById('export-xls').addEventListener('click', () => {{
  const taxOnlyPct   = Number(document.getElementById('tax-pct')?.value) || 0;
  const adminOnlyPct = Number(document.getElementById('admin-pct')?.value) || 0;
  const exportRows = Object.values(rowDataByNmid).map(row => {{
    const f = calcRowForecast(row);
    const drrPct    = drrChanges[row.nmid] != null ? drrChanges[row.nmid] : (row.base_drr_pct || 0);
    const revenue   = f.forecastBuyouts * f.plannedClient;
    const cogsTotal = f.forecastBuyouts * (row.cogs || 0);
    const wbComm    = revenue * ((row.wb_commission_pct || 0) / 100);
    const acquiring = revenue * 0.03;
    const drrCost   = revenue * (drrPct / 100);
    const grossProfit = revenue - cogsTotal - wbComm - acquiring - drrCost;
    const marginPct   = revenue > 0 ? grossProfit / revenue * 100 : 0;
    const adminCost   = revenue * (adminOnlyPct / 100);
    const taxCost     = revenue * (taxOnlyPct / 100);
    const netProfit   = grossProfit - adminCost - taxCost;
    const profitPct   = revenue > 0 ? netProfit / revenue * 100 : 0;
    return {{
      article:           row.article,
      subject:           row.subject,
      strategy:          (row.labels||[]).join(", "),
      revenue_cat:       row.revenue_category,
      stock:             row.stock,
      wh_stock:          row.warehouse_stock,
      stock_val:         row.stock * (row.client_price || row.average_check || 0),
      base_price:        row.baseline_seller_price,
      base_spp:          row.baseline_spp,
      base_client:       row.baseline_client_check,
      base_speed:        row.base_speed,
      cur_price:         row.current_price,
      cur_spp:           row.current_spp,
      cur_client:        row.client_price,
      elasticity:        row.elasticity,
      cur_speed:         row.current_speed,
      price_pct:         priceChanges[row.nmid] ?? 0,
      plan_drr:          drrPct,
      ads_elasticity:    row.ads_elasticity,
      base_ctr:          row.base_ctr,
      base_cr1:          row.base_cr1,
      base_cr2:          row.base_cr2,
      planned_price:     f.plannedSeller,
      planned_client:    f.plannedClient,
      planned_speed:     f.plannedSpeed,
      buyout_percent:    row.buyout_percent,
      forecast_orders:   f.forecastOrders,
      forecast_buyouts:  f.forecastBuyouts,
      forecast_revenue:  f.forecastRevenue,
      stockout_date:     f.stockoutDate || '',
      lost_orders:       f.lostOrders,
      lost_revenue:      f.lostRevenue,
      comm_pct:          row.wb_commission_pct || 0,
      comm_rub:          wbComm,
      acq_rub:           acquiring,
      drr_rub:           drrCost,
      cogs_total:        cogsTotal,
      margin_pct:        marginPct,
      margin_rub:        grossProfit,
      admin_rub:         adminCost,
      tax_rub:           taxCost,
      profit_pct:        profitPct,
      profit_rub:        netProfit,
    }};
  }});
  fetch('/api/analytics/planning/export', {{
    method: 'POST',
    headers: {{'Content-Type': 'application/json'}},
    body: JSON.stringify({{ rows: exportRows }}),
  }}).then(r => {{
    if (!r.ok) return r.text().then(t => {{ throw new Error(t); }});
    return r.blob();
  }}).then(blob => {{
    const a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    const today = new Date().toISOString().slice(0, 10);
    a.download = `planning_${{today}}.xlsx`;
    a.click();
  }}).catch(err => alert('Ошибка экспорта: ' + err));
}});

let planningParams = {{ forecast_days: 0, return_delay_days: 8, transit_days: 3, forecast_from: '' }};

function calcConvFactor(row) {{
  const c = convChanges[row.nmid] || {{}};
  let f = 1.0;
  if (c.ctr > 0 && row.base_ctr > 0) f *= c.ctr / row.base_ctr;
  if (c.cr1 > 0 && row.base_cr1 > 0) f *= c.cr1 / row.base_cr1;
  if (c.cr2 > 0 && row.base_cr2 > 0) f *= c.cr2 / row.base_cr2;
  return f;
}}

function simulatePlanning(stock, dailySpeed, buyoutRate, forecastDays, returnDelay, startDateStr, warehouseStock, transitDays) {{
  let available = Math.max(0, stock);
  const demand = Math.max(0, dailySpeed);
  buyoutRate = Math.max(0, Math.min(1, buyoutRate));
  const delay = Math.max(1, Math.round(returnDelay));
  const transit = Math.max(0, Math.round(transitDays || 0));
  const arrivals = {{}};
  if (warehouseStock > 0) arrivals[transit] = Math.max(0, warehouseStock);
  const returnsByDay = {{}};
  let totalOrders = 0;
  let stockoutDate = '';
  const startMs = startDateStr ? new Date(startDateStr).getTime() : 0;
  for (let idx = 0; idx < Math.max(0, forecastDays); idx++) {{
    available += (returnsByDay[idx] || 0) + (arrivals[idx] || 0);
    delete returnsByDay[idx];
    delete arrivals[idx];
    const ordered = Math.min(demand, available);
    if (demand > available && !stockoutDate && startMs) {{
      const d = new Date(startMs + idx * 86400000);
      stockoutDate = d.toISOString().slice(0, 10);
    }}
    available -= ordered;
    totalOrders += ordered;
    const returning = ordered * (1 - buyoutRate);
    const dueIdx = idx + delay;
    if (returning > 0 && dueIdx < forecastDays) {{
      returnsByDay[dueIdx] = (returnsByDay[dueIdx] || 0) + returning;
    }}
  }}
  return {{ totalOrders, stockoutDate }};
}}

function formatNumber(value, digits = 0) {{
  const num = Number(value || 0);
  return num.toLocaleString('ru-RU', {{ maximumFractionDigits: digits, minimumFractionDigits: digits }});
}}
function formatMoney(value) {{ return formatNumber(value, 0); }}

function saveState() {{
  try {{
    localStorage.setItem(stateKey, JSON.stringify({{
      forecastTo: fields.forecastTo.value,
      returnDelay: fields.returnDelay.value,
      transitDays: fields.transitDays.value,
      subject: fields.subject.value,
      labels: [...(fields.strategy?.selectedOptions||[])].map(o=>o.value).join(","),
      revenueCategory: fields.revenueCategory.value,
    }}));
  }} catch (err) {{}}
}}

function restoreState() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(stateKey) || '{{}}');
    if (saved.forecastTo) fields.forecastTo.value = saved.forecastTo;
    if (saved.returnDelay) fields.returnDelay.value = saved.returnDelay;
    if (saved.transitDays != null) fields.transitDays.value = saved.transitDays;
    fields.subject.dataset.pending = saved.subject || '';
    // labels not restored from storage - always start empty
    fields.revenueCategory.dataset.pending = saved.revenueCategory || '';
  }} catch (err) {{}}
}}

function fillSelect(select, values, placeholder) {{
  const current = select.value || select.dataset.pending || '';
  select.innerHTML = `<option value="">${{placeholder}}</option>` + (values || []).map(v => `<option value="${{String(v).replaceAll('"','&quot;')}}">${{v}}</option>`).join('');
  if ([...select.options].some(o => o.value === current)) select.value = current;
  select.dataset.pending = '';
}}
function fillMultiSelect(select, values) {{
  const prevVals = new Set([...select.selectedOptions].map(o=>o.value));
  const optionsHtml = (values || []).map(v => {{
    const sel = prevVals.has(v) ? ' selected' : '';
    return `<option value="${{String(v).replaceAll('"','&quot;')}}"${{sel}}>${{v}}</option>`;
  }}).join('');
  select.innerHTML = select.multiple ? optionsHtml : `<option value="">Все ярлыки</option>${{optionsHtml}}`;
}}

function renderSummary(summary) {{
  const avgMargin = summary.total_econ_revenue > 0
    ? (summary.total_gross_profit / summary.total_econ_revenue * 100).toFixed(1)
    : '—';
  summaryEl.innerHTML = `
    <div class="kpi">
      <div class="kpi-name">Себестоимость остатков</div>
      <div class="kpi-rub">${{formatMoney(summary.stock_cost)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.stock)}} шт на складе</div>
      <div class="kpi-qty" style="color:#64748b">В тек. ценах: ${{formatMoney(summary.stock_value_current || summary.stock_value)}}</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Плановые заказы</div>
      <div class="kpi-rub">${{formatMoney(summary.orders_revenue)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.forecast_orders)}} шт</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Плановые выкупы</div>
      <div class="kpi-rub">${{formatMoney(summary.forecast_revenue)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.forecast_buyouts)}} шт</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Упущенные продажи</div>
      <div class="kpi-rub">${{formatMoney(summary.lost_revenue)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.lost_orders)}} шт</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Расходы на рекламу</div>
      <div class="kpi-rub">${{formatMoney(summary.total_drr_cost)}}</div>
      <div class="kpi-qty">${{summary.total_econ_revenue > 0 ? (summary.total_drr_cost / summary.total_econ_revenue * 100).toFixed(1) : '—'}}% от выкупов</div>
    </div>
    <div class="kpi" style="border-color:#a855f7">
      <div class="kpi-name" style="color:#7e22ce">Маржинальность</div>
      <div class="kpi-rub" style="color:#7e22ce">${{avgMargin}}%</div>
      <div class="kpi-qty">валовая прибыль ${{formatMoney(summary.total_gross_profit)}}</div>
    </div>
    <div class="kpi" style="border-color:#a855f7">
      <div class="kpi-name" style="color:#7e22ce">Чистая прибыль</div>
      <div class="kpi-rub" style="color:${{(summary.total_net_profit||0)>=0?'#166534':'#dc2626'}}">${{formatMoney(summary.total_net_profit)}}</div>
      <div class="kpi-qty">после налогов и адм.</div>
    </div>`;
}}

function priceArrow(current, baseline, invertLogic) {{
  if (!baseline || !current) return '';
  const delta = (current - baseline) / baseline;
  if (Math.abs(delta) < 0.01) return '';
  const up = invertLogic ? delta < 0 : delta > 0;
  return `<span class="price-arrow ${{up ? 'price-up' : 'price-down'}}">${{up ? '↑' : '↓'}}</span>`;
}}

function calcNoOOS(row) {{
  const {{ forecast_days, return_delay_days, transit_days, forecast_from }} = planningParams;
  const buyoutRate = row.buyout_percent / 100;
  const wh = row.warehouse_stock || 0;
  const curPct = priceChanges[row.nmid] ?? 0;
  // check OOS at current planned speed (already includes DRR/conv factors)
  const {{ plannedSpeed: curSpeed }} = calcPlanned(row, curPct);
  const {{ stockoutDate: sd0 }} = simulatePlanning(row.stock, curSpeed, buyoutRate, forecast_days, return_delay_days, forecast_from, wh, transit_days);
  if (!sd0) return 0;  // no OOS at current plan → nothing to do
  if (row.elasticity >= 0) return null;  // can't reduce speed by raising price
  const baseClient = row.baseline_client_check;
  if (!baseClient) return null;
  // binary search starting from current price pct (don't lower existing change)
  let lo = curPct, hi = Math.max(300, curPct + 300);
  for (let i = 0; i < 50; i++) {{
    const mid = (lo + hi) / 2;
    const {{ plannedSpeed }} = calcPlanned(row, mid);
    const {{ stockoutDate }} = simulatePlanning(row.stock, plannedSpeed, buyoutRate, forecast_days, return_delay_days, forecast_from, wh, transit_days);
    if (!stockoutDate) hi = mid; else lo = mid;
    if (hi - lo < 0.05) break;
  }}
  if (hi >= curPct + 299) return null;  // can't avoid OOS even with huge increase
  return Math.ceil(hi);
}}

// Variant of calcPlanned with explicit drrPct (doesn't read global drrChanges)
function calcPlannedEx(row, pricePct, drrPct) {{
  const p = Number(pricePct) || 0;
  const plannedSeller = row.current_price * (1 + p / 100);
  const plannedClient = plannedSeller * (1 - row.current_spp / 100);
  const baseClient = row.baseline_client_check || plannedClient;
  const delta = (baseClient > 0 && plannedClient > 0) ? (plannedClient / baseClient - 1) : 0;
  const priceFactor = Math.max(0, 1 + row.elasticity * delta);
  let adsFactor = 1.0;
  const d = Number(drrPct) || 0;
  if (row.ads_elasticity > 0 && (row.base_drr_pct || 0) > 0) {{
    const adsRatio = Math.max(0.1, d / row.base_drr_pct);
    adsFactor = Math.pow(adsRatio, row.ads_elasticity);
  }}
  return {{ plannedSeller, plannedClient, plannedSpeed: row.base_speed * priceFactor * adsFactor * calcConvFactor(row) }};
}}

function netRevEx(row, pricePct, drrPct) {{
  const {{ plannedSpeed, plannedClient }} = calcPlannedEx(row, pricePct, drrPct);
  const {{ forecast_days, return_delay_days, transit_days, forecast_from }} = planningParams;
  const {{ totalOrders }} = simulatePlanning(
    row.stock, plannedSpeed, row.buyout_percent / 100,
    forecast_days, return_delay_days, forecast_from,
    row.warehouse_stock || 0, transit_days
  );
  const gross = totalOrders * (row.buyout_percent / 100) * plannedClient;
  return gross * (1 - (Number(drrPct) || 0) / 100);
}}

function findOptimalRevenue(row) {{
  if (!row.stock || !planningParams.forecast_days) return {{ pricePct: 0, drrPct: row.base_drr_pct || 0 }};
  const hasAds = row.ads_elasticity > 0 && (row.base_drr_pct || 0) > 0;
  const baseDrr = row.base_drr_pct || 0;

  // Cap upward price search: extrapolating far beyond data range is unreliable.
  // Inelastic products (|e|≈0) should not be pushed to +100% — market won't absorb it.
  const absE = Math.abs(row.elasticity || 0);
  const maxPriceUp = absE === 0 ? 5 : absE < 0.3 ? 10 : absE < 0.7 ? 20 : 30;

  let bestPct = 0, bestDrr = baseDrr, bestRev = -Infinity;

  if (hasAds) {{
    // Phase 1: joint coarse grid — price × DRR
    for (let p = -30; p <= maxPriceUp; p += 3) {{
      for (let d = 0; d <= 25; d++) {{
        const r = netRevEx(row, p, d);
        if (r > bestRev) {{ bestRev = r; bestPct = p; bestDrr = d; }}
      }}
    }}
    // Phase 2: refine around best — price ±3 step 1 × DRR ±1 step 0.1
    const p0 = bestPct, d0 = bestDrr;
    for (let p = p0 - 3; p <= Math.min(p0 + 3, maxPriceUp); p++) {{
      for (let di = -10; di <= 10; di++) {{
        const d = Math.round((d0 + di * 0.1) * 10) / 10;
        if (d < 0) continue;
        const r = netRevEx(row, p, d);
        if (r > bestRev) {{ bestRev = r; bestPct = p; bestDrr = d; }}
      }}
    }}
  }} else {{
    // No ads elasticity — search only price
    for (let p = -30; p <= maxPriceUp; p += 3) {{
      const r = netRevEx(row, p, baseDrr);
      if (r > bestRev) {{ bestRev = r; bestPct = p; }}
    }}
    const p0 = bestPct;
    for (let p = p0 - 3; p <= Math.min(p0 + 3, maxPriceUp); p++) {{
      const r = netRevEx(row, p, baseDrr);
      if (r > bestRev) {{ bestRev = r; bestPct = p; }}
    }}
    bestDrr = baseDrr;
  }}

  return {{ pricePct: bestPct, drrPct: Math.round(bestDrr * 10) / 10 }};
}}

function calcPlanned(row, pct) {{
  const p = Number(pct) || 0;
  const plannedSeller = row.current_price * (1 + p / 100);
  const plannedClient = plannedSeller * (1 - row.current_spp / 100);
  const baseClient = row.baseline_client_check || plannedClient;
  const delta = (baseClient > 0 && plannedClient > 0) ? (plannedClient / baseClient - 1) : 0;
  const priceFactor = Math.max(0, 1 + row.elasticity * delta);
  // ads factor: per-row planned DRR% (absolute), applied via bivariate elasticity
  let adsFactor = 1.0;
  const plannedDrr = drrChanges[row.nmid] != null ? drrChanges[row.nmid] : (row.base_drr_pct || 0);
  if (row.ads_elasticity > 0 && (row.base_drr_pct || 0) > 0) {{
    // ratio of planned to base DRR%; floor at 0.1 models organic traffic baseline
    const adsRatio = Math.max(0.1, plannedDrr / row.base_drr_pct);
    adsFactor = Math.pow(adsRatio, row.ads_elasticity);
  }}
  // conversion factor
  const convFactor = calcConvFactor(row);
  return {{ plannedSeller, plannedClient, plannedSpeed: row.base_speed * priceFactor * adsFactor * convFactor }};
}}

function calcRowForecast(row) {{
  const pct = priceChanges[row.nmid] ?? 0;
  const {{ plannedSeller, plannedClient, plannedSpeed }} = calcPlanned(row, pct);
  const {{ forecast_days, return_delay_days, transit_days, forecast_from }} = planningParams;
  const {{ totalOrders: forecastOrders, stockoutDate }} = simulatePlanning(
    row.stock, plannedSpeed, row.buyout_percent / 100,
    forecast_days, return_delay_days, forecast_from,
    row.warehouse_stock || 0, transit_days
  );
  const buyoutRate = row.buyout_percent / 100;
  const effectiveCheck = plannedClient > 0 ? plannedClient : row.average_check;
  const forecastBuyouts = forecastOrders * buyoutRate;
  const forecastRevenue = forecastBuyouts * effectiveCheck;
  const potentialOrders = plannedSpeed * forecast_days;
  const lostOrders = Math.max(0, potentialOrders - forecastOrders);
  const lostRevenue = lostOrders * buyoutRate * effectiveCheck;
  return {{ plannedSeller, plannedClient, plannedSpeed, effectiveCheck, forecastOrders, forecastBuyouts, forecastRevenue, stockoutDate, lostOrders, lostRevenue }};
}}

function recalcSummary() {{
  const s = {{ stock_cost: 0, stock_value: 0, stock_value_current: 0, forecast_orders: 0, forecast_buyouts: 0, orders_revenue: 0, forecast_revenue: 0, lost_orders: 0, lost_revenue: 0, stock: 0, total_gross_profit: 0, total_net_profit: 0, total_econ_revenue: 0, total_drr_cost: 0 }};
  Object.values(rowDataByNmid).forEach(row => {{
    const c = rowForecastCache[row.nmid] || {{}};
    s.stock += row.stock;
    s.stock_cost += row.stock * (row.cogs || 0);
    s.stock_value += row.stock * (row.client_price || row.average_check || 0);
    s.stock_value_current += row.stock * (row.client_price || row.average_check || 0);
    s.forecast_orders += c.forecastOrders || 0;
    s.forecast_buyouts += c.forecastBuyouts || 0;
    s.orders_revenue += c.ordersRevenue || 0;
    s.forecast_revenue += c.forecastRevenue || 0;
    s.lost_orders += c.lostOrders || 0;
    s.lost_revenue += c.lostRevenue || 0;
    s.total_gross_profit += c.grossProfit || 0;
    s.total_net_profit   += c.netProfit || 0;
    s.total_econ_revenue += c.econRevenue || 0;
    s.total_drr_cost     += c.drrCost || 0;
  }});
  renderSummary(s);
}}

function updateRowPlanned(nmid, skipSummary) {{
  const row = rowDataByNmid[nmid];
  if (!row) return;
  const f = calcRowForecast(row);

  // Economics
  const taxOnlyPct   = Number(document.getElementById('tax-pct')?.value) || 0;
  const adminOnlyPct = Number(document.getElementById('admin-pct')?.value) || 0;
  const drrPct = drrChanges[nmid] != null ? drrChanges[nmid] : (row.base_drr_pct || 0);
  const revenue    = f.forecastBuyouts * f.plannedClient;
  const cogsTotal  = f.forecastBuyouts * (row.cogs || 0);
  const wbComm     = revenue * ((row.wb_commission_pct || 0) / 100);
  const acquiring  = revenue * 0.03;
  const drrCost    = revenue * (drrPct / 100);
  const grossProfit = revenue - cogsTotal - wbComm - acquiring - drrCost;
  const marginPct   = revenue > 0 ? grossProfit / revenue * 100 : 0;
  const adminCost   = revenue * (adminOnlyPct / 100);
  const taxCost     = revenue * (taxOnlyPct / 100);
  const netProfit   = grossProfit - adminCost - taxCost;
  const profitPct   = revenue > 0 ? netProfit / revenue * 100 : 0;

  rowForecastCache[nmid] = {{
    forecastOrders: f.forecastOrders,
    forecastBuyouts: f.forecastBuyouts,
    ordersRevenue: f.forecastOrders * f.effectiveCheck,
    forecastRevenue: f.forecastRevenue,
    lostOrders: f.lostOrders,
    lostRevenue: f.lostRevenue,
    grossProfit,
    netProfit,
    econRevenue: revenue,
    drrCost,
  }};
  const tr = tableEl.querySelector(`tr[data-nmid="${{nmid}}"]`);
  if (!tr) return;
  tr.querySelector('.cell-ps').textContent = formatMoney(f.plannedSeller);
  tr.querySelector('.cell-pc').textContent = formatMoney(f.plannedClient);
  tr.querySelector('.cell-pspd').textContent = formatNumber(f.plannedSpeed, 2);
  tr.querySelector('.cell-fo').textContent = formatNumber(f.forecastOrders);
  tr.querySelector('.cell-fb').textContent = formatNumber(f.forecastBuyouts);
  tr.querySelector('.cell-fr').textContent = formatMoney(f.forecastRevenue);
  tr.querySelector('.cell-oos').textContent = f.stockoutDate || '—';
  tr.querySelector('.cell-lo').textContent = formatNumber(f.lostOrders);
  tr.querySelector('.cell-lr').textContent = formatMoney(f.lostRevenue);

  // Economics cells
  function setCell(sel, text, val, colorVal) {{
    const el = tr.querySelector(sel);
    if (!el) return;
    el.textContent = text;
    if (val !== undefined) el.dataset[Object.keys(el.dataset)[0] || 'v'] = val;
    if (colorVal !== undefined) el.style.color = colorVal >= 0 ? '#166534' : '#dc2626';
  }}
  tr.querySelector('.cell-econ-comm-rub') && (tr.querySelector('.cell-econ-comm-rub').textContent = formatMoney(wbComm));
  tr.querySelector('.cell-econ-acq-rub')  && (tr.querySelector('.cell-econ-acq-rub').textContent  = formatMoney(acquiring));
  tr.querySelector('.cell-econ-drr-rub')  && (tr.querySelector('.cell-econ-drr-rub').textContent  = formatMoney(drrCost));
  tr.querySelector('.cell-econ-cogs')     && (tr.querySelector('.cell-econ-cogs').textContent     = formatMoney(cogsTotal));
  const marginCell = tr.querySelector('.cell-econ-margin');
  if (marginCell) {{
    marginCell.textContent = marginPct.toFixed(1) + '%';
    marginCell.dataset.margin = marginPct.toFixed(1);
    marginCell.style.color = marginPct >= 0 ? '#166534' : '#dc2626';
  }}
  tr.querySelector('.cell-econ-margin-rub') && (tr.querySelector('.cell-econ-margin-rub').textContent = formatMoney(grossProfit));
  tr.querySelector('.cell-econ-admin-rub')  && (tr.querySelector('.cell-econ-admin-rub').textContent  = formatMoney(adminCost));
  tr.querySelector('.cell-econ-tax-rub')    && (tr.querySelector('.cell-econ-tax-rub').textContent    = formatMoney(taxCost));
  const profitPctCell = tr.querySelector('.cell-econ-profit-pct');
  if (profitPctCell) {{
    profitPctCell.textContent = profitPct.toFixed(1) + '%';
    profitPctCell.dataset.profitPct = profitPct.toFixed(1);
    profitPctCell.style.color = profitPct >= 0 ? '#166534' : '#dc2626';
  }}
  const profitCell = tr.querySelector('.cell-econ-profit');
  if (profitCell) {{
    profitCell.textContent = formatMoney(netProfit);
    profitCell.dataset.profit = netProfit.toFixed(0);
    profitCell.style.color = netProfit >= 0 ? '#166534' : '#dc2626';
  }}

  // WB stock warning
  const {{ transit_days }} = planningParams;
  const daysLeft = f.plannedSpeed > 0 ? row.stock / f.plannedSpeed : Infinity;
  const stockCell = tr.querySelector('.cell-wbstock');
  if (stockCell) {{
    const warn = isFinite(daysLeft) && daysLeft < transit_days && (row.warehouse_stock || 0) > 0;
    stockCell.style.color = warn ? '#dc2626' : '';
    stockCell.title = warn ? `Кончится через ~${{daysLeft.toFixed(1)}} дн., срок поставки ${{transit_days}} дн.` : '';
  }}

  // arrow indicators for drr / conv inputs
  function applyValArrow(input, arrowEl, baseVal) {{
    if (!input || !arrowEl) return;
    const step = parseFloat(input.step) || 1;
    const dec = step <= 0.15 ? 2 : 1;
    // Compare as strings at the same precision to avoid float mismatch (e.g. base=36.65 → toFixed(1)="36.6" but Math.round(36.65*10)=367)
    const curStr  = parseFloat(input.value).toFixed(dec);
    const baseStr = (baseVal || 0).toFixed(dec);
    const cur  = parseFloat(curStr);
    const base = parseFloat(baseStr);
    input.classList.remove('val-changed-up', 'val-changed-down');
    arrowEl.classList.remove('val-up', 'val-down');
    arrowEl.textContent = '';
    if (cur !== base) {{
      const up = cur > base;
      input.classList.add(up ? 'val-changed-up' : 'val-changed-down');
      arrowEl.classList.add(up ? 'val-up' : 'val-down');
      arrowEl.textContent = up ? '↑' : '↓';
    }}
  }}
  applyValArrow(tr.querySelector('.drr-input'), tr.querySelector('.drr-arrow'), row.base_drr_pct);
  applyValArrow(tr.querySelector('.conv-input[data-conv="ctr"]'), tr.querySelector('.ctr-arrow'), row.base_ctr);
  applyValArrow(tr.querySelector('.conv-input[data-conv="cr1"]'), tr.querySelector('.cr1-arrow'), row.base_cr1);
  applyValArrow(tr.querySelector('.conv-input[data-conv="cr2"]'), tr.querySelector('.cr2-arrow'), row.base_cr2);

  if (!skipSummary) recalcSummary();
}}

function applyWeeksFilter() {{
  const max = Number(fields.weeksFilter.value);
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const row = rowDataByNmid[tr.dataset.nmid];
    if (!row) return;
    const weeks = row.base_speed > 0 ? row.stock / row.base_speed / 7 : Infinity;
    tr.style.display = (!max || weeks <= max) ? '' : 'none';
  }});
}}

function applyElasticityFilter() {{
  const val = fields.elasticityFilter.value;
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const hasElasticity = row.elasticity !== 0;
    const visible = !val || (val === 'elastic' ? hasElasticity : !hasElasticity);
    tr.style.display = visible ? '' : 'none';
  }});
}}

function updateBulkBar() {{
  const checked = tableEl.querySelectorAll('input.row-ck:checked');
  const n = checked.length;
  bulkCount.textContent = n;
  bulkBar.style.display = n > 0 ? 'flex' : 'none';
}}

function renderTable(rows) {{
  if (!rows.length) {{
    tableEl.innerHTML = '<div class="empty">Нет товаров для выбранных фильтров</div>';
    return;
  }}
  rowDataByNmid = {{}};
  rowForecastCache = {{}};
  const _taxPctInit   = Number(document.getElementById('tax-pct')?.value) || 0;
  const _adminPctInit = Number(document.getElementById('admin-pct')?.value) || 0;
  rows.forEach(r => {{
    rowDataByNmid[r.nmid] = r;
    const _rev = r.forecast_buyouts * (r.planned_price || r.client_price || r.average_check || 0);
    const _cogs = r.forecast_buyouts * (r.cogs || 0);
    const _comm = _rev * ((r.wb_commission_pct || 0) / 100);
    const _acq  = _rev * 0.03;
    const _drr  = _rev * ((r.base_drr_pct || 0) / 100);
    const _gross = _rev - _cogs - _comm - _acq - _drr;
    const _net   = _gross - _rev * (_adminPctInit / 100) - _rev * (_taxPctInit / 100);
    rowForecastCache[r.nmid] = {{
      forecastOrders: r.forecast_orders,
      forecastBuyouts: r.forecast_buyouts,
      ordersRevenue: r.forecast_orders * (r.effective_check || r.average_check),
      forecastRevenue: r.forecast_revenue,
      lostOrders: r.lost_orders,
      lostRevenue: r.lost_revenue,
      grossProfit: _gross,
      netProfit: _net,
      econRevenue: _rev,
    }};
  }});

  const head = `<tr>
    <th class="col-ck" rowspan="2"><input type="checkbox" id="ck-all" title="Выбрать все"></th>
    <th class="col-sku" rowspan="2" data-sort-key="article">SKU</th>
    <th class="col-subject" rowspan="2" data-sort-key="subject">Предмет</th>
    <th class="col-text" rowspan="2" data-sort-key="strategy">Ярлыки</th>
    <th class="col-text" rowspan="2" data-sort-key="revenue_category">Категория</th>
    <th class="g-sep" colspan="3" data-col-group="остатки">Остатки</th>
    <th class="g-sep" colspan="4" data-col-group="базовый">Базовый период 7д</th>
    <th class="g-sep" colspan="5" data-col-group="текущее">Текущее</th>
    <th class="g-sep th-params" colspan="6" data-col-group="параметры">Параметры плана</th>
    <th class="g-sep th-result" colspan="7" data-col-group="результат">Результат</th>
    <th class="g-sep th-lost" colspan="3" data-col-group="упущено">Упущено</th>
    <th class="g-sep th-econ" colspan="10" data-col-group="экономика">Экономика</th>
  </tr><tr>
    <th class="g-sep" data-col-group="остатки" data-col-key="c-stock" data-sort-key="stock">WB, шт</th><th data-col-group="остатки" data-col-key="c-wh-stock" data-sort-key="warehouse_stock">Склад, шт</th><th data-col-group="остатки" data-col-key="c-stock-val" data-sort-key="stock_value">Сумма, ₽</th>
    <th class="g-sep" data-col-group="базовый" data-col-key="c-base-price" data-sort-key="baseline_seller_price">Цена, ₽</th><th data-col-group="базовый" data-col-key="c-base-spp" data-sort-key="baseline_spp">СПП, %</th><th data-col-group="базовый" data-col-key="c-base-client" data-sort-key="baseline_client_check">Кл. цена, ₽</th><th data-col-group="базовый" data-col-key="c-base-speed" data-sort-key="base_speed">Скорость, шт</th>
    <th class="g-sep" data-col-group="текущее" data-col-key="c-cur-price" data-sort-key="current_price">Тек. цена, ₽</th><th data-col-group="текущее" data-col-key="c-cur-spp" data-sort-key="current_spp">СПП, %</th><th data-col-group="текущее" data-col-key="c-cur-client" data-sort-key="client_price">Кл. цена, ₽</th><th data-col-group="текущее" data-col-key="c-cur-elast" data-sort-key="elasticity">Эл. цены</th><th data-col-group="текущее" data-col-key="c-cur-speed" data-sort-key="current_speed">Скорость, шт</th>
    <th class="g-sep params-col" data-col-group="параметры" data-col-key="c-plan-price">Цена, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-drr" data-sort-key="base_drr_pct">ДРР, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-drr-e" data-sort-key="ads_elasticity">Эл.рекл.</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-ctr" data-sort-key="base_ctr">CTR, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-cr1" data-sort-key="base_cr1">CR1, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-cr2" data-sort-key="base_cr2">CR2, %</th>
    <th class="g-sep result-col" data-col-group="результат" data-col-key="c-res-price">Пл. цена, ₽</th><th class="result-col" data-col-group="результат" data-col-key="c-res-client">Пл. кл., ₽</th><th class="result-col" data-col-group="результат" data-col-key="c-res-speed" data-sort-key="plannedSpeed">Скорость, шт</th><th class="result-col" data-col-group="результат" data-col-key="c-res-buyout" data-sort-key="buyout_percent">% выкупа</th><th class="result-col" data-col-group="результат" data-col-key="c-res-orders" data-sort-key="forecast_orders">Заказы, шт</th><th class="result-col" data-col-group="результат" data-col-key="c-res-buyouts" data-sort-key="forecast_buyouts">Выкупы, шт</th><th class="result-col" data-col-group="результат" data-col-key="c-res-revenue" data-sort-key="forecast_revenue">Выручка, ₽</th>
    <th class="g-sep lost-col" data-col-group="упущено" data-col-key="c-res-oos">Дата стопа</th><th class="lost-col" data-col-group="упущено" data-col-key="c-lost-qty" data-sort-key="lost_orders">Упущ., шт</th><th class="lost-col" data-col-group="упущено" data-col-key="c-lost-rub" data-sort-key="lost_revenue">Упущ., ₽</th>
    <th class="g-sep econ-col" data-col-group="экономика" data-col-key="c-econ-comm-pct" data-sort-key="wb_commission_pct">Ком. %</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-comm-rub" data-sort-key="comm_rub">Ком. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-acq-rub" data-sort-key="acq_rub">Эквайр. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-drr-rub" data-sort-key="drr_rub">Реклама ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-cogs" data-sort-key="cogs">Себест. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-margin-pct" data-sort-key="margin_pct">Маржа %</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-margin-rub" data-sort-key="margin_rub">Маржа ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-admin-rub" data-sort-key="admin_rub">Адм. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-tax-rub" data-sort-key="tax_rub">Налоги ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-profit-pct" data-sort-key="profit_pct">Прибыль %</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-profit-rub" data-sort-key="net_profit">Прибыль ₽</th>
  </tr>`;

  const body = rows.map(row => {{
    const pct = priceChanges[row.nmid] ?? 0;
    const drrVal = drrChanges[row.nmid] != null ? drrChanges[row.nmid] : (row.base_drr_pct || 0);
    const ctrVal = (convChanges[row.nmid] || {{}}).ctr != null ? (convChanges[row.nmid] || {{}}).ctr : (row.base_ctr || 0);
    const cr1Val = (convChanges[row.nmid] || {{}}).cr1 != null ? (convChanges[row.nmid] || {{}}).cr1 : (row.base_cr1 || 0);
    const cr2Val = (convChanges[row.nmid] || {{}}).cr2 != null ? (convChanges[row.nmid] || {{}}).cr2 : (row.base_cr2 || 0);
    const {{ plannedSeller, plannedClient, plannedSpeed }} = calcPlanned(row, pct);
    return `<tr data-nmid="${{row.nmid}}">
      <td class="col-ck"><input type="checkbox" class="row-ck"></td>
      <td class="col-sku" title="${{row.article || row.nmid}}">${{row.article || row.nmid}}</td>
      <td class="col-subject" title="${{row.subject || ''}}">${{row.subject || ''}}</td>
      <td class="col-text">${{(row.labels||[]).join(', ')}}</td>
      <td class="col-text">${{row.revenue_category || ''}}</td>
      <td class="g-sep cell-wbstock" data-col-group="остатки" data-col-key="c-stock">${{formatNumber(row.stock)}}</td>
      <td data-col-group="остатки" data-col-key="c-wh-stock">${{formatNumber(row.warehouse_stock)}}</td>
      <td data-col-group="остатки" data-col-key="c-stock-val">${{formatMoney(row.stock * (row.client_price || row.average_check || 0))}}</td>
      <td class="g-sep" data-col-group="базовый" data-col-key="c-base-price">${{formatMoney(row.baseline_seller_price)}}</td>
      <td data-col-group="базовый" data-col-key="c-base-spp">${{formatNumber(row.baseline_spp, 1)}}%</td>
      <td data-col-group="базовый" data-col-key="c-base-client">${{formatMoney(row.baseline_client_check)}}</td>
      <td data-col-group="базовый" data-col-key="c-base-speed">${{formatNumber(row.base_speed, 2)}}</td>
      <td class="g-sep" data-col-group="текущее" data-col-key="c-cur-price">${{formatMoney(row.current_price)}}${{priceArrow(row.current_price, row.baseline_seller_price)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-spp">${{formatNumber(row.current_spp, 1)}}%${{priceArrow(row.current_spp, row.baseline_spp, true)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-client">${{formatMoney(row.client_price)}}${{priceArrow(row.client_price, row.baseline_client_check)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-elast">${{formatNumber(row.elasticity, 2)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-speed">${{formatNumber(row.current_speed, 2)}}</td>
      <td class="g-sep" data-col-group="параметры" data-col-key="c-plan-price">
        <div class="pct-cell">
          <div class="pct-spinner">
            <input type="number" class="pct-input" value="${{pct}}" step="1" min="-90" max="300">
            <div class="pct-arrows">
              <button class="pct-btn" data-d="1">▲</button>
              <button class="pct-btn" data-d="-1">▼</button>
            </div>
          </div>
        </div>
      </td>
      <td data-col-group="параметры" data-col-key="c-plan-drr">
        <div class="pct-cell" title="Базовый ДРР: ${{formatNumber(row.base_drr_pct, 1)}}%">
          <span class="val-arrow drr-arrow"></span>
          <div class="pct-spinner">
            <input type="number" class="drr-input" value="${{(+drrVal || 0).toFixed(1)}}" step="0.5" min="0" max="100">
            <div class="pct-arrows">
              <button class="drr-btn" data-d="0.5">▲</button>
              <button class="drr-btn" data-d="-0.5">▼</button>
            </div>
          </div>
        </div>
      </td>
      <td class="cell-drr-e" data-col-group="параметры" data-col-key="c-plan-drr-e">${{row.ads_elasticity ? formatNumber(row.ads_elasticity, 2) : '—'}}</td>
      <td data-col-group="параметры" data-col-key="c-plan-ctr">
        ${{(function(){{
          const noBase = !row.base_ctr && !(convChanges[row.nmid]||{{}}).ctr;
          const v = noBase ? '' : (+ctrVal||0).toFixed(2);
          return `<div class="pct-cell" title="${{noBase ? 'Нет данных за период' : 'Базовый CTR: '+formatNumber(row.base_ctr,2)+'%'}}">
            <span class="val-arrow ctr-arrow"></span>
            <div class="pct-spinner${{noBase?' no-data':''}}">
              <input type="number" class="conv-input" data-conv="ctr" value="${{v}}" placeholder="${{noBase?'—':''}}" step="0.1" min="0" max="100" ${{noBase?'disabled':''}}>
              <div class="pct-arrows"><button class="conv-btn" data-conv="ctr" data-d="0.1">▲</button><button class="conv-btn" data-conv="ctr" data-d="-0.1">▼</button></div>
            </div></div>`;
        }})()}}
      </td>
      <td data-col-group="параметры" data-col-key="c-plan-cr1">
        ${{(function(){{
          const noBase = !row.base_cr1 && !(convChanges[row.nmid]||{{}}).cr1;
          const v = noBase ? '' : (+cr1Val||0).toFixed(2);
          return `<div class="pct-cell" title="${{noBase ? 'Нет данных за период' : 'Базовый CR1: '+formatNumber(row.base_cr1,2)+'%'}}">
            <span class="val-arrow cr1-arrow"></span>
            <div class="pct-spinner${{noBase?' no-data':''}}">
              <input type="number" class="conv-input" data-conv="cr1" value="${{v}}" placeholder="${{noBase?'—':''}}" step="0.1" min="0" max="100" ${{noBase?'disabled':''}}>
              <div class="pct-arrows"><button class="conv-btn" data-conv="cr1" data-d="0.1">▲</button><button class="conv-btn" data-conv="cr1" data-d="-0.1">▼</button></div>
            </div></div>`;
        }})()}}
      </td>
      <td data-col-group="параметры" data-col-key="c-plan-cr2">
        ${{(function(){{
          const noBase = !row.base_cr2 && !(convChanges[row.nmid]||{{}}).cr2;
          const v = noBase ? '' : (+cr2Val||0).toFixed(2);
          return `<div class="pct-cell" title="${{noBase ? 'Нет данных за период' : 'Базовый CR2: '+formatNumber(row.base_cr2,2)+'%'}}">
            <span class="val-arrow cr2-arrow"></span>
            <div class="pct-spinner${{noBase?' no-data':''}}">
              <input type="number" class="conv-input" data-conv="cr2" value="${{v}}" placeholder="${{noBase?'—':''}}" step="0.1" min="0" max="100" ${{noBase?'disabled':''}}>
              <div class="pct-arrows"><button class="conv-btn" data-conv="cr2" data-d="0.1">▲</button><button class="conv-btn" data-conv="cr2" data-d="-0.1">▼</button></div>
            </div></div>`;
        }})()}}
      </td>
      <td class="g-sep cell-ps" data-col-group="результат" data-col-key="c-res-price">${{formatMoney(plannedSeller)}}</td>
      <td class="cell-pc" data-col-group="результат" data-col-key="c-res-client">${{formatMoney(plannedClient)}}</td>
      <td class="cell-pspd" data-col-group="результат" data-col-key="c-res-speed">${{formatNumber(plannedSpeed, 2)}}</td>
      <td data-col-group="результат" data-col-key="c-res-buyout">${{formatNumber(row.buyout_percent)}}%</td>
      <td class="cell-fo" data-col-group="результат" data-col-key="c-res-orders">${{formatNumber(row.forecast_orders)}}</td>
      <td class="cell-fb" data-col-group="результат" data-col-key="c-res-buyouts">${{formatNumber(row.forecast_buyouts)}}</td>
      <td class="cell-fr" data-col-group="результат" data-col-key="c-res-revenue">${{formatMoney(row.forecast_revenue)}}</td>
      <td class="cell-oos g-sep" data-col-group="упущено" data-col-key="c-res-oos">${{row.stockout_date || '—'}}</td>
      <td class="g-sep cell-lo" data-col-group="упущено" data-col-key="c-lost-qty">${{formatNumber(row.lost_orders)}}</td>
      <td class="cell-lr" data-col-group="упущено" data-col-key="c-lost-rub">${{formatMoney(row.lost_revenue)}}</td>
      <td class="g-sep econ-col cell-wb-comm" data-col-group="экономика" data-col-key="c-econ-comm-pct">${{formatNumber(row.wb_commission_pct || 0, 1)}}%</td>
      <td class="econ-col cell-econ-comm-rub" data-col-group="экономика" data-col-key="c-econ-comm-rub">—</td>
      <td class="econ-col cell-econ-acq-rub" data-col-group="экономика" data-col-key="c-econ-acq-rub">—</td>
      <td class="econ-col cell-econ-drr-rub" data-col-group="экономика" data-col-key="c-econ-drr-rub">—</td>
      <td class="econ-col cell-econ-cogs" data-col-group="экономика" data-col-key="c-econ-cogs">—</td>
      <td class="econ-col cell-econ-margin" data-col-group="экономика" data-col-key="c-econ-margin-pct" data-margin="0">—</td>
      <td class="econ-col cell-econ-margin-rub" data-col-group="экономика" data-col-key="c-econ-margin-rub">—</td>
      <td class="econ-col cell-econ-admin-rub" data-col-group="экономика" data-col-key="c-econ-admin-rub">—</td>
      <td class="econ-col cell-econ-tax-rub" data-col-group="экономика" data-col-key="c-econ-tax-rub">—</td>
      <td class="econ-col cell-econ-profit-pct" data-col-group="экономика" data-col-key="c-econ-profit-pct" data-profit-pct="0">—</td>
      <td class="econ-col cell-econ-profit" data-col-group="экономика" data-col-key="c-econ-profit-rub" data-profit="0">—</td>
    </tr>`;
  }}).join('');

  tableEl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;

  // checkbox: select all
  document.getElementById('ck-all').addEventListener('change', e => {{
    tableEl.querySelectorAll('input.row-ck').forEach(ck => {{
      ck.checked = e.target.checked;
      ck.closest('tr')?.classList.toggle('row-selected', e.target.checked);
    }});
    updateBulkBar();
  }});

  // checkbox: individual
  tableEl.addEventListener('change', e => {{
    if (e.target.classList.contains('row-ck')) {{
      e.target.closest('tr')?.classList.toggle('row-selected', e.target.checked);
      updateBulkBar();
    }}
  }});

  // price pct input & buttons
  tableEl.addEventListener('input', e => {{
    const tr = e.target.closest('tr');
    if (!tr) return;
    const nmid = tr.dataset.nmid;
    if (e.target.classList.contains('pct-input')) {{
      priceChanges[nmid] = Number(e.target.value) || 0;
      savePriceChanges();
      updateRowPlanned(nmid);
    }} else if (e.target.classList.contains('drr-input')) {{
      drrChanges[nmid] = Number(e.target.value) || 0;
      saveDrrChanges();
      updateRowPlanned(nmid);
    }} else if (e.target.classList.contains('conv-input')) {{
      const conv = e.target.dataset.conv;
      const val = parseFloat(e.target.value);
      if (!convChanges[nmid]) convChanges[nmid] = {{}};
      if (val > 0) {{
        convChanges[nmid][conv] = val;
      }} else {{
        delete convChanges[nmid][conv];
      }}
      saveConvChanges();
      updateRowPlanned(nmid);
    }}
  }});
  tableEl.addEventListener('click', e => {{
    // arrow click → reset that value to baseline
    const arrow = e.target.closest('.val-arrow');
    if (arrow && arrow.textContent) {{
      const tr = arrow.closest('tr');
      const nmid = tr?.dataset.nmid;
      const row = rowDataByNmid[nmid];
      if (!row) return;
      if (arrow.classList.contains('drr-arrow')) {{
        delete drrChanges[nmid];
        saveDrrChanges();
        const inp = tr.querySelector('.drr-input');
        if (inp) inp.value = (row.base_drr_pct || 0).toFixed(1);
      }} else {{
        let conv = null;
        if (arrow.classList.contains('ctr-arrow')) conv = 'ctr';
        else if (arrow.classList.contains('cr1-arrow')) conv = 'cr1';
        else if (arrow.classList.contains('cr2-arrow')) conv = 'cr2';
        if (conv) {{
          if (convChanges[nmid]) {{
            delete convChanges[nmid][conv];
            if (!Object.keys(convChanges[nmid]).length) delete convChanges[nmid];
          }}
          saveConvChanges();
          const inp = tr.querySelector(`.conv-input[data-conv="${{conv}}"]`);
          if (inp && !inp.disabled) inp.value = (row[`base_${{conv}}`] || 0).toFixed(2);
        }}
      }}
      updateRowPlanned(nmid);
      return;
    }}
    const btn = e.target.closest('.pct-btn, .drr-btn, .conv-btn');
    if (!btn) return;
    const tr = btn.closest('tr');
    const nmid = tr.dataset.nmid;
    if (btn.classList.contains('pct-btn')) {{
      const input = tr.querySelector('.pct-input');
      const newVal = Math.max(-90, Math.min(300, (Number(input.value) || 0) + Number(btn.dataset.d)));
      input.value = newVal;
      priceChanges[nmid] = newVal;
      savePriceChanges();
    }} else if (btn.classList.contains('drr-btn')) {{
      const input = tr.querySelector('.drr-input');
      const newVal = Math.max(0, Math.min(100, Math.round(((Number(input.value) || 0) + Number(btn.dataset.d)) * 10) / 10));
      input.value = newVal.toFixed(1);
      drrChanges[nmid] = newVal;
      saveDrrChanges();
    }} else if (btn.classList.contains('conv-btn')) {{
      const conv = btn.dataset.conv;
      const input = tr.querySelector(`.conv-input[data-conv="${{conv}}"]`);
      const newVal = Math.max(0, Math.min(100, Math.round(((Number(input.value) || 0) + Number(btn.dataset.d)) * 10) / 10));
      input.value = newVal;
      if (!convChanges[nmid]) convChanges[nmid] = {{}};
      convChanges[nmid][conv] = newVal;
      saveConvChanges();
    }}
    updateRowPlanned(nmid);
  }});
}}

// bulk bar actions
const bulkType = document.getElementById('bulk-type');
document.getElementById('bulk-apply').addEventListener('click', () => {{
  const delta = Number(bulkPct.value) || 0;
  const type = bulkType?.value || 'price';
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const ck = tr.querySelector('input.row-ck');
    if (!ck?.checked) return;
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (type === 'price') {{
      const cur = priceChanges[nmid] ?? 0;
      const nv = Math.max(-90, Math.min(300, Math.round((cur + delta) * 10) / 10));
      priceChanges[nmid] = nv;
      const inp = tr.querySelector('.pct-input');
      if (inp) inp.value = nv;
    }} else if (type === 'drr') {{
      const cur = drrChanges[nmid] != null ? drrChanges[nmid] : (row?.base_drr_pct || 0);
      const nv = Math.max(0, Math.min(100, Math.round((cur + delta) * 10) / 10));
      drrChanges[nmid] = nv;
      const inp = tr.querySelector('.drr-input');
      if (inp) inp.value = nv.toFixed(1);
    }} else {{
      if (!convChanges[nmid]) convChanges[nmid] = {{}};
      const cur = convChanges[nmid][type] > 0 ? convChanges[nmid][type] : (row?.[`base_${{type}}`] || 0);
      const nv = Math.max(0, Math.min(100, Math.round((cur + delta) * 100) / 100));
      convChanges[nmid][type] = nv;
      const inp = tr.querySelector(`.conv-input[data-conv="${{type}}"]`);
      if (inp) inp.value = nv.toFixed(2);
    }}
    updateRowPlanned(nmid);
  }});
  savePriceChanges(); saveDrrChanges(); saveConvChanges();
}});
document.getElementById('bulk-reset').addEventListener('click', () => {{
  const type = bulkType?.value || 'price';
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const ck = tr.querySelector('input.row-ck');
    if (!ck?.checked) return;
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (type === 'price') {{
      priceChanges[nmid] = 0;
      const inp = tr.querySelector('.pct-input'); if (inp) inp.value = 0;
    }} else if (type === 'drr') {{
      delete drrChanges[nmid];
      const inp = tr.querySelector('.drr-input');
      if (inp) inp.value = row?.base_drr_pct || 0;
    }} else {{
      if (convChanges[nmid]) delete convChanges[nmid][type];
      const inp = tr.querySelector(`.conv-input[data-conv="${{type}}"]`);
      if (inp) inp.value = row?.[`base_${{type}}`] || 0;
    }}
    updateRowPlanned(nmid);
  }});
  savePriceChanges(); saveDrrChanges(); saveConvChanges();
}});
document.getElementById('no-oos-btn').addEventListener('click', () => {{
  const selected = new Set([...tableEl.querySelectorAll('input.row-ck:checked')].map(ck => ck.closest('tr')?.dataset.nmid).filter(Boolean));
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    if (selected.size > 0 && !selected.has(nmid)) return;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const pct = calcNoOOS(row);
    if (pct === null || pct === 0) return;
    priceChanges[nmid] = pct;
    const input = tr.querySelector('.pct-input');
    if (input) input.value = pct;
    updateRowPlanned(nmid);
  }});
  savePriceChanges();
}});
document.getElementById('optimize-btn').addEventListener('click', () => {{
  const btn = document.getElementById('optimize-btn');
  btn.disabled = true;
  const selected = new Set([...tableEl.querySelectorAll('input.row-ck:checked')]
    .map(ck => ck.closest('tr')?.dataset.nmid).filter(Boolean));
  const rows = [...tableEl.querySelectorAll('tr[data-nmid]')]
    .map(tr => tr.dataset.nmid)
    .filter(nmid => selected.size === 0 || selected.has(nmid))
    .filter(nmid => rowDataByNmid[nmid]);
  const total = rows.length;
  let idx = 0;
  const CHUNK = 5;
  function processChunk() {{
    const end = Math.min(idx + CHUNK, total);
    for (; idx < end; idx++) {{
      const nmid = rows[idx];
      const row = rowDataByNmid[nmid];
      const {{ pricePct, drrPct }} = findOptimalRevenue(row);
      priceChanges[nmid] = pricePct;
      drrChanges[nmid] = drrPct;
      const tr = tableEl.querySelector(`tr[data-nmid="${{nmid}}"]`);
      if (tr) {{
        const pInp = tr.querySelector('.pct-input');
        if (pInp) pInp.value = pricePct;
        const dInp = tr.querySelector('.drr-input');
        if (dInp) dInp.value = drrPct.toFixed(1);
        updateRowPlanned(nmid);
      }}
    }}
    btn.textContent = `⏳ ${{idx}}/${{total}}…`;
    if (idx < total) {{
      setTimeout(processChunk, 0);
    }} else {{
      savePriceChanges();
      saveDrrChanges();
      const sname = 'Макс. выручка';
      const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
      all[sname] = {{
        price: Object.assign({{}}, priceChanges),
        drr:   Object.assign({{}}, drrChanges),
        conv:  JSON.parse(JSON.stringify(convChanges)),
      }};
      localStorage.setItem(scenariosKey, JSON.stringify(all));
      loadScenarioList();
      document.getElementById('scenario-select').value = sname;
      btn.disabled = false; btn.textContent = '⚡ Макс. выручку';
    }}
  }}
  setTimeout(processChunk, 0);
}});
document.getElementById('reset-prices').addEventListener('click', () => {{
  priceChanges = {{}};
  savePriceChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const input = tr.querySelector('.pct-input');
    if (input) input.value = 0;
    updateRowPlanned(nmid);
  }});
}});
document.getElementById('reset-drr').addEventListener('click', () => {{
  drrChanges = {{}};
  saveDrrChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const drrInp = tr.querySelector('.drr-input');
    if (drrInp) drrInp.value = (row.base_drr_pct || 0).toFixed(1);
    updateRowPlanned(nmid, true);
  }});
  recalcSummary();
}});
document.getElementById('reset-conv').addEventListener('click', () => {{
  convChanges = {{}};
  saveConvChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    ['ctr','cr1','cr2'].forEach(k => {{
      const inp = tr.querySelector(`.conv-input[data-conv="${{k}}"]`);
      if (inp && !inp.disabled) inp.value = (row[`base_${{k}}`] || 0).toFixed(2);
    }});
    updateRowPlanned(nmid, true);
  }});
  recalcSummary();
}});
document.getElementById('bulk-close').addEventListener('click', () => {{
  tableEl.querySelectorAll('input.row-ck').forEach(ck => {{ ck.checked = false; }});
  const ckAll = document.getElementById('ck-all');
  if (ckAll) ckAll.checked = false;
  updateBulkBar();
}});

async function loadData() {{
  saveState();
  meta.textContent = 'Считаю...';
  priceChanges = {{}};
  const qs = new URLSearchParams({{
    forecast_to: fields.forecastTo.value,
    return_delay_days: fields.returnDelay.value || '8',
    transit_days: fields.transitDays.value || '3',
    subject: fields.subject.value,
    labels: [...(fields.strategy?.selectedOptions||[])].map(o=>o.value).join(","),
    revenue_category: fields.revenueCategory.value,
    price_change_pct: '0',
  }});
  const resp = await fetch(`/api/analytics/planning?${{qs.toString()}}`);
  const data = await resp.json();
  if (!resp.ok) {{
    meta.textContent = data.error || `Ошибка ${{resp.status}}`;
    summaryEl.innerHTML = '';
    tableEl.innerHTML = '';
    return;
  }}
  fillSelect(fields.subject, data.options?.subjects || [], 'Все предметы');
  fillMultiSelect(fields.strategy, data.options?.labels || []);
  fillSelect(fields.revenueCategory, data.options?.revenue_categories || [], 'Все категории');
  planningParams = {{
    forecast_days: data.forecast_days || 0,
    return_delay_days: data.return_delay_days || 8,
    transit_days: data.transit_days ?? 3,
    forecast_from: data.forecast_from || '',
  }};
  renderSummary(data.summary || {{}});
  renderTable(data.rows || []);
  // fill economics + WB stock warning for all rows — skip per-row summary recalc, do one pass at the end
  Object.keys(rowDataByNmid).forEach(nmid => updateRowPlanned(nmid, true));
  recalcSummary();
  applyWeeksFilter();
  applyColVisibility();
  meta.textContent = `База скорости: ${{data.baseline_from}}..${{data.baseline_to}} | % выкупа WB за этот период | Прогноз: ${{data.forecast_from}}..${{data.forecast_to}} | Дней: ${{data.forecast_days}} | SKU: ${{(data.rows || []).length}}`;
}}

let timer = null;
function scheduleLoad() {{ clearTimeout(timer); timer = setTimeout(loadData, 250); }}
for (const el of Object.values(fields)) {{
  if (el === fields.weeksFilter) {{
    el.addEventListener('input', applyWeeksFilter);
  }} else {{
    el.addEventListener('change', loadData);
    el.addEventListener('input', scheduleLoad);
  }}
}}
document.getElementById('load').addEventListener('click', loadData);

// tax/admin: recalc all rows on change
['tax-pct', 'admin-pct'].forEach(id => {{
  document.getElementById(id).addEventListener('input', () => {{
    Object.keys(rowDataByNmid).forEach(nmid => updateRowPlanned(nmid, true));
    recalcSummary();
  }});
}});

// ——— Column visibility (gear) ———
const COL_SCHEMA = [
  {{ group:'остатки', label:'Остатки', cols:[
    {{ key:'c-stock',     label:'WB, шт' }},
    {{ key:'c-wh-stock',  label:'Склад, шт' }},
    {{ key:'c-stock-val', label:'Сумма, ₽' }},
  ]}},
  {{ group:'базовый', label:'Базовый период', cols:[
    {{ key:'c-base-price',  label:'Цена, ₽' }},
    {{ key:'c-base-spp',    label:'СПП, %' }},
    {{ key:'c-base-client', label:'Кл. цена, ₽' }},
    {{ key:'c-base-speed',  label:'Скорость, шт' }},
  ]}},
  {{ group:'текущее', label:'Текущее', cols:[
    {{ key:'c-cur-price',  label:'Цена, ₽' }},
    {{ key:'c-cur-spp',    label:'СПП, %' }},
    {{ key:'c-cur-client', label:'Кл. цена, ₽' }},
    {{ key:'c-cur-elast',  label:'Эластичность' }},
    {{ key:'c-cur-speed',  label:'Скорость, шт' }},
  ]}},
  {{ group:'параметры', label:'Параметры', cols:[
    {{ key:'c-plan-price', label:'Цена, %' }},
    {{ key:'c-plan-drr',   label:'ДРР, %' }},
    {{ key:'c-plan-drr-e', label:'Эл. рекл.' }},
    {{ key:'c-plan-ctr',   label:'CTR, %' }},
    {{ key:'c-plan-cr1',   label:'CR корзина, %' }},
    {{ key:'c-plan-cr2',   label:'CR заказ, %' }},
  ]}},
  {{ group:'результат', label:'Результат', cols:[
    {{ key:'c-res-price',   label:'Пл. цена, ₽' }},
    {{ key:'c-res-client',  label:'Пл. кл. цена, ₽' }},
    {{ key:'c-res-speed',   label:'Скорость, шт/д' }},
    {{ key:'c-res-buyout',  label:'% выкупа' }},
    {{ key:'c-res-orders',  label:'Заказы, шт' }},
    {{ key:'c-res-buyouts', label:'Выкупы, шт' }},
    {{ key:'c-res-revenue', label:'Выручка, ₽' }},
  ]}},
  {{ group:'упущено', label:'Упущенные продажи', cols:[
    {{ key:'c-res-oos',  label:'Дата стопа' }},
    {{ key:'c-lost-qty', label:'Упущ., шт' }},
    {{ key:'c-lost-rub', label:'Упущ., ₽' }},
  ]}},
  {{ group:'экономика', label:'Экономика', cols:[
    {{ key:'c-econ-comm-pct',   label:'Комиссия, %' }},
    {{ key:'c-econ-comm-rub',   label:'Комиссия, ₽' }},
    {{ key:'c-econ-acq-rub',    label:'Эквайринг, ₽' }},
    {{ key:'c-econ-drr-rub',    label:'Реклама, ₽' }},
    {{ key:'c-econ-cogs',       label:'Себестоимость, ₽' }},
    {{ key:'c-econ-margin-pct', label:'Маржа, %' }},
    {{ key:'c-econ-margin-rub', label:'Маржа, ₽' }},
    {{ key:'c-econ-admin-rub',  label:'Адм. расходы, ₽' }},
    {{ key:'c-econ-tax-rub',    label:'Налоги, ₽' }},
    {{ key:'c-econ-profit-pct', label:'Прибыль, %' }},
    {{ key:'c-econ-profit-rub', label:'Прибыль, ₽' }},
  ]}},
];

const COL_VIS_KEY = 'wb.planning.col-vis';
let colVisibility = {{}};
try {{ colVisibility = JSON.parse(localStorage.getItem(COL_VIS_KEY) || '{{}}'); }} catch(e) {{}}

const _colVisStyle = document.createElement('style');
_colVisStyle.id = 'col-vis-style';
document.head.appendChild(_colVisStyle);

function applyColVisibility() {{
  const rules = [];
  COL_SCHEMA.forEach(grp => {{
    const allHidden = grp.cols.every(c => colVisibility[c.key] === false);
    if (allHidden) {{
      rules.push(`[data-col-group="${{grp.group}}"] {{ display:none !important; }}`);
    }} else {{
      grp.cols.forEach(c => {{
        if (colVisibility[c.key] === false) {{
          rules.push(`[data-col-key="${{c.key}}"] {{ display:none !important; }}`);
        }}
      }});
    }}
  }});
  document.getElementById('col-vis-style').textContent = rules.join('\\n');
}}

function _isGroupChecked(grp) {{ return grp.cols.some(c => colVisibility[c.key] !== false); }}
function _isGroupIndeterminate(grp) {{
  const vis = grp.cols.filter(c => colVisibility[c.key] !== false).length;
  return vis > 0 && vis < grp.cols.length;
}}

const colSettingsBtn   = document.getElementById('col-settings-btn');
const colSettingsPanel = document.getElementById('col-settings-panel');

function _buildColPanel() {{
  colSettingsPanel.innerHTML = `
    <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#94a3b8;margin-bottom:8px">Столбцы таблицы</div>
    ${{COL_SCHEMA.map(grp => `
      <div style="margin-bottom:6px">
        <label style="display:flex;align-items:center;gap:6px;cursor:pointer;font-size:12px;font-weight:700;color:#334155;padding:2px 0">
          <input type="checkbox" data-grp="${{grp.group}}"
            ${{_isGroupChecked(grp) ? 'checked' : ''}}
            style="accent-color:#7e22ce"> ${{grp.label}}
        </label>
        <div style="padding-left:18px">
          ${{grp.cols.map(c => `
            <label style="display:flex;align-items:center;gap:5px;cursor:pointer;font-size:12px;color:#475569;padding:1px 0">
              <input type="checkbox" data-col="${{c.key}}"
                ${{colVisibility[c.key] !== false ? 'checked' : ''}}
                style="accent-color:#7e22ce"> ${{c.label}}
            </label>`).join('')}}
        </div>
      </div>`).join('')}}`;

  // group toggle
  colSettingsPanel.querySelectorAll('input[data-grp]').forEach(cb => {{
    const grp = COL_SCHEMA.find(g => g.group === cb.dataset.grp);
    if (!grp) return;
    // set indeterminate state
    if (_isGroupIndeterminate(grp)) cb.indeterminate = true;
    cb.addEventListener('change', () => {{
      grp.cols.forEach(c => {{
        colVisibility[c.key] = cb.checked;
        const colCb = colSettingsPanel.querySelector(`input[data-col="${{c.key}}"]`);
        if (colCb) colCb.checked = cb.checked;
      }});
      localStorage.setItem(COL_VIS_KEY, JSON.stringify(colVisibility));
      applyColVisibility();
    }});
  }});

  // individual column toggle
  colSettingsPanel.querySelectorAll('input[data-col]').forEach(cb => {{
    cb.addEventListener('change', () => {{
      colVisibility[cb.dataset.col] = cb.checked;
      localStorage.setItem(COL_VIS_KEY, JSON.stringify(colVisibility));
      applyColVisibility();
      // update parent group checkbox
      const grp = COL_SCHEMA.find(g => g.cols.some(c => c.key === cb.dataset.col));
      if (grp) {{
        const grpCb = colSettingsPanel.querySelector(`input[data-grp="${{grp.group}}"]`);
        if (grpCb) {{
          const checked = _isGroupChecked(grp);
          const indet = _isGroupIndeterminate(grp);
          grpCb.checked = checked;
          grpCb.indeterminate = indet;
        }}
      }}
    }});
  }});
}}

colSettingsBtn.addEventListener('click', e => {{
  e.stopPropagation();
  if (colSettingsPanel.style.display === 'none') {{
    _buildColPanel();
    colSettingsPanel.style.display = 'block';
  }} else {{
    colSettingsPanel.style.display = 'none';
  }}
}});
document.addEventListener('click', () => {{ colSettingsPanel.style.display = 'none'; }});

// ——— Sorting ———
let sortState = {{ key: null, asc: true }};
tableEl.addEventListener('click', e => {{
  const th = e.target.closest('th[data-sort-key]');
  if (!th) return;
  const key = th.dataset.sortKey;
  sortState.asc = sortState.key === key ? !sortState.asc : true;
  sortState.key = key;
  tableEl.querySelectorAll('th[data-sort-key]').forEach(t => delete t.dataset.sortDir);
  th.dataset.sortDir = sortState.asc ? 'asc' : 'desc';
  const tbody = tableEl.querySelector('tbody');
  if (!tbody) return;
  const tRows = [...tbody.querySelectorAll('tr[data-nmid]')];
  const textKeys = new Set(['article','subject','strategy','revenue_category']);
  tRows.sort((a, b) => {{
    const nA = a.dataset.nmid, nB = b.dataset.nmid;
    let va, vb;
    if (key === 'margin_pct') {{
      va = parseFloat(a.querySelector('.cell-econ-margin')?.dataset.margin || 0);
      vb = parseFloat(b.querySelector('.cell-econ-margin')?.dataset.margin || 0);
    }} else if (key === 'profit_pct') {{
      va = parseFloat(a.querySelector('.cell-econ-profit-pct')?.dataset.profitPct || 0);
      vb = parseFloat(b.querySelector('.cell-econ-profit-pct')?.dataset.profitPct || 0);
    }} else if (key === 'net_profit') {{
      va = parseFloat(a.querySelector('.cell-econ-profit')?.dataset.profit || 0);
      vb = parseFloat(b.querySelector('.cell-econ-profit')?.dataset.profit || 0);
    }} else if (key === 'margin_rub') {{
      va = parseFloat(a.querySelector('.cell-econ-margin-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
      vb = parseFloat(b.querySelector('.cell-econ-margin-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
    }} else if (key === 'comm_rub') {{
      va = parseFloat(a.querySelector('.cell-econ-comm-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
      vb = parseFloat(b.querySelector('.cell-econ-comm-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
    }} else if (key === 'drr_rub') {{
      va = parseFloat(a.querySelector('.cell-econ-drr-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
      vb = parseFloat(b.querySelector('.cell-econ-drr-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
    }} else if (key === 'stock_value') {{
      va = (rowDataByNmid[a.dataset.nmid]?.stock || 0) * (rowDataByNmid[a.dataset.nmid]?.client_price || 0);
      vb = (rowDataByNmid[b.dataset.nmid]?.stock || 0) * (rowDataByNmid[b.dataset.nmid]?.client_price || 0);
    }} else if (key === 'plannedSpeed') {{
      va = rowForecastCache[nA]?.forecastOrders / Math.max(1, planningParams.forecast_days) || 0;
      vb = rowForecastCache[nB]?.forecastOrders / Math.max(1, planningParams.forecast_days) || 0;
    }} else if (textKeys.has(key)) {{
      va = (rowDataByNmid[nA]?.[key] || '').toLowerCase();
      vb = (rowDataByNmid[nB]?.[key] || '').toLowerCase();
      return sortState.asc ? va.localeCompare(vb,'ru') : vb.localeCompare(va,'ru');
    }} else {{
      va = rowDataByNmid[nA]?.[key] ?? (rowForecastCache[nA]?.[key] ?? 0);
      vb = rowDataByNmid[nB]?.[key] ?? (rowForecastCache[nB]?.[key] ?? 0);
    }}
    return sortState.asc ? va - vb : vb - va;
  }});
  tRows.forEach(r => tbody.appendChild(r));
}});

restoreState();
loadData();
</script>
<div class="methodology">
  <h2>Методология расчётов</h2>

  <h3>Скорость продаж (базовая)</h3>
  <p>Среднедневное количество заказов за последние 7 дней базового периода (последний день с данными исключается — он может быть неполным).<br>
  <span class="formula">скорость = заказы_за_7_дней / 7</span><br>
  <em>Пример: за 7 дней было 350 заказов → базовая скорость = 50 шт/день.</em></p>

  <h3>Эластичность цены</h3>
  <p>Показывает, насколько меняется спрос при изменении цены. Рассчитывается методом лог-лог МНК по 60 дням истории заказов:<br>
  <span class="formula">ln(заказы) = E × ln(цена) + const</span></p>
  <p>Условия расчёта: ≥ 7 дней с заказами, разброс цен ≥ 5%, два ценовых уровня с разрывом ≥ 3%, R² ≥ 0.30.<br>
  Диапазон значений: от −3.0 до 0.0 (отрицательная — при росте цены спрос падает).</p>
  <p><strong>Как читать в таблице:</strong></p>
  <ul>
    <li><strong>−1.0</strong> (по умолчанию, нет данных) — единичная эластичность: +10% цены → −10% скорости.<br>
    <em>50 шт/день × (1 + (−1.0) × 0.10) = 45 шт/день при цене +10%</em></li>
    <li><strong>−0.5</strong> — слабая реакция: +20% цены → −10% скорости.<br>
    <em>50 × (1 + (−0.5) × 0.20) = 45 шт/день</em></li>
    <li><strong>−2.0</strong> — сильная реакция: +10% цены → −20% скорости.<br>
    <em>50 × (1 + (−2.0) × 0.10) = 40 шт/день</em></li>
    <li><strong>0.0</strong> — данные есть, но цена статистически не влияет на спрос (R² &lt; 0.30 или цена не менялась).</li>
  </ul>
  <p>Если по SKU нет надёжных данных — берётся среднее по предмету → стратегии → категории. Если нигде нет — используется <strong>−1.0</strong> (консервативный дефолт).</p>

  <h3>ДРР и эластичность рекламы</h3>
  <p><strong>Базовый ДРР%</strong> = фактические расходы на рекламу / фактическая выручка с выкупов × 100%, за последние 14 дней (последние 2 дня исключаются — выкупы там ещё не зафиксированы).<br>
  <span class="formula">ДРР = Σ расходы_реклама_14д / Σ buyoutSum_14д × 100%</span><br>
  <em>Пример: реклама 70 000 ₽, выручка с выкупов 500 000 ₽ → ДРР = 14%.</em></p>
  <p><strong>Эластичность рекламы</strong> показывает, как расходы на рекламу влияют на скорость продаж. Рассчитывается двумерным МНК вместе с ценовой эластичностью:<br>
  <span class="formula">ln(заказы) = E_цена × ln(цена) + E_реклама × ln(1 + расход/день) + const</span><br>
  Диапазон E_реклама: от 0.0 до 1.5.</p>
  <p><strong>Как работает в планировании:</strong><br>
  <span class="formula">ads_factor = (план_ДРР% / баз_ДРР%) ^ E_реклама</span><br>
  (минимальный порог ratio = 0.10, чтобы при ДРР = 0 скорость не падала в ноль — органика сохраняется)</p>
  <ul>
    <li><em>ДРР 14% → 28% (×2), E_реклама = 0.5: ads_factor = 2.0^0.5 = 1.41 → +41% к скорости</em></li>
    <li><em>ДРР 14% → 7% (÷2), E_реклама = 0.5: ads_factor = 0.5^0.5 = 0.71 → −29% к скорости</em></li>
    <li><em>ДРР 14% → 0% (выкл.), E_реклама = 0.23: ads_factor = 0.10^0.23 = 0.60 → −40% к скорости</em></li>
  </ul>

  <h3>Конверсии воронки (CTR, CR1, CR2)</h3>
  <p>Данные из ручной выгрузки WB (funnel_impressions_upload) за базовые 7 дней:<br>
  <span class="formula">CTR = переходы / показы × 100%</span> &nbsp;
  <span class="formula">CR1 = корзины / переходы × 100%</span> &nbsp;
  <span class="formula">CR2 = заказы / корзины × 100%</span></p>
  <p>Изменение конверсий — прямой мультипликатор (улучшение карточки, фото, описания):<br>
  <span class="formula">conv_factor = (CTR_план/CTR_баз) × (CR1_план/CR1_баз) × (CR2_план/CR2_баз)</span><br>
  <em>Пример: CTR +20%, CR1 без изм., CR2 +10% → conv_factor = 1.20 × 1.0 × 1.10 = 1.32 → +32% к скорости.</em></p>

  <h3>Итоговая плановая скорость</h3>
  <p>Три рычага перемножаются:<br>
  <span class="formula">план_скорость = базовая_скорость × price_factor × ads_factor × conv_factor</span></p>
  <ul>
    <li><strong>price_factor</strong> = <span class="formula">max(0, 1 + E_цена × Δцена_дол.)</span> — линейная аппроксимация эластичности</li>
    <li><strong>ads_factor</strong> = <span class="formula">(план_ДРР / баз_ДРР) ^ E_реклама</span> — степенной закон</li>
    <li><strong>conv_factor</strong> — прямой мультипликатор конверсий</li>
  </ul>
  <p><em>Пример: скорость 50 шт/день, цена +10% (E=−1.0), ДРР ×2 (E_рекл=0.5), CTR +20%:<br>
  price_factor = 1 + (−1.0) × 0.10 = 0.90 &nbsp;|&nbsp; ads_factor = 2.0^0.5 = 1.41 &nbsp;|&nbsp; conv_factor = 1.20<br>
  план_скорость = 50 × 0.90 × 1.41 × 1.20 ≈ 76 шт/день</em></p>
  <p style="color:#64748b;font-size:.75rem;margin-top:8px">Все изменения цены, ДРР и конверсий хранятся в localStorage браузера — прогноз пересчитывается мгновенно на стороне клиента без запросов к серверу.</p>
</div>
</body>
</html>
"""

OZON_PLANNING_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Планирование</title>
  <style>
    :root {{ --bg:#f8fafc; --card:#fff; --ink:#111827; --muted:#6b7280; --accent:#0f766e; --line:#e5e7eb; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:"Segoe UI","Trebuchet MS",sans-serif; color:var(--ink); background:linear-gradient(180deg,#e0f2fe,transparent 240px),var(--bg); }}
    .wrap {{ width:calc(100% - 32px); max-width:3000px; margin:0 auto; padding:10px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .ctrl-panel {{ background:var(--card); border:1px solid var(--line); border-radius:10px; padding:14px 16px; display:flex; flex-direction:column; gap:10px; }}
    .ctrl-row {{ display:flex; flex-wrap:wrap; align-items:flex-end; gap:8px; }}
    .ctrl-field {{ display:flex; flex-direction:column; gap:3px; }}
    .ctrl-field label {{ font-size:11px; color:#64748b; font-weight:600; text-transform:uppercase; letter-spacing:.04em; white-space:nowrap; }}
    .ctrl-field input, .ctrl-field select {{ height:32px; padding:0 8px; border:1px solid #cbd5e1; border-radius:6px; font-size:13px; background:#f8fafc; color:#0f172a; min-width:0; }}
    .ctrl-field select {{ padding-right:24px; }}
    .ctrl-sep {{ width:1px; align-self:stretch; background:var(--line); margin:0 4px; flex-shrink:0; }}
    .ctrl-group {{ display:flex; align-items:flex-end; gap:6px; }}
    .ctrl-group-label {{ font-size:10px; font-weight:700; text-transform:uppercase; letter-spacing:.06em; color:#94a3b8; margin-bottom:4px; }}
    .ctrl-btn {{ height:32px; padding:0 12px; border-radius:6px; border:none; cursor:pointer; font-size:13px; font-weight:500; color:#fff; white-space:nowrap; transition:filter .15s; }}
    .ctrl-btn:hover {{ filter:brightness(.88); }}
    .ctrl-btn:disabled {{ opacity:.5; cursor:default; }}
    .ctrl-btn-primary {{ background:#2563eb; }}
    .ctrl-btn-teal   {{ background:#0f766e; }}
    .ctrl-btn-purple {{ background:#7c3aed; }}
    .ctrl-btn-gray   {{ background:#64748b; }}
    .ctrl-btn-red    {{ background:#b91c1c; }}
    .ctrl-btn-green  {{ background:#065f46; }}
    .ctrl-btn-slate  {{ background:#334155; }}
    .ctrl-scenario-select {{ height:32px; padding:0 6px; border:1px solid #cbd5e1; border-radius:6px; font-size:12px; background:#f8fafc; color:#0f172a; max-width:130px; }}
    .ctrl-scenario-input  {{ height:32px; padding:0 8px; border:1px solid #cbd5e1; border-radius:6px; font-size:12px; background:#f8fafc; color:#0f172a; width:110px; }}
    .ctrl-tax-input {{ height:32px; width:60px; padding:0 8px; border:1px solid #cbd5e1; border-radius:6px; font-size:13px; background:#f8fafc; color:#0f172a; }}
    .panel {{ background:var(--card); border:1px solid var(--line); border-radius:8px; padding:12px; display:flex; flex-wrap:wrap; gap:10px; align-items:end; }}
    label {{ display:block; color:var(--muted); font-size:.72rem; margin-bottom:3px; font-weight:700; }}
    input, select {{ padding:7px 8px; border:1px solid #cbd5e1; border-radius:7px; font-size:12px; min-width:150px; background:#fff; }}
    button {{ padding:8px 12px; border:0; border-radius:7px; background:var(--accent); color:#fff; font-weight:700; cursor:pointer; font-size:12px; }}
    .meta {{ margin:8px 0 0; color:var(--muted); font-size:.78rem; }}
    .summary {{ display:flex; flex-wrap:nowrap; gap:10px; margin-top:10px; overflow-x:auto; }}
    .summary .kpi {{ flex:1 1 0; min-width:0; }}
    .kpi {{ background:#fff; border:1px solid var(--line); border-radius:8px; padding:10px 12px; }}
    .kpi-name {{ color:var(--muted); font-size:.72rem; font-weight:700; text-transform:uppercase; letter-spacing:.3px; }}
    .kpi-rub {{ margin-top:5px; font-size:1.15rem; font-weight:900; color:#0f172a; }}
    .kpi-qty {{ margin-top:2px; font-size:.82rem; font-weight:600; color:#475569; }}
    .tbl {{ margin-top:10px; background:#fff; border:1px solid var(--line); border-radius:8px; overflow:auto; max-height:calc(100vh - 260px); }}
    table {{ width:max-content; min-width:100%; border-collapse:collapse; font-size:11px; }}
    th, td {{ border-bottom:1px solid #eef2f7; padding:5px 6px; text-align:right; }}
    th {{ position:sticky; top:0; z-index:1; background:#f8fafc; color:#334155; font-weight:800; white-space:normal; word-break:normal; vertical-align:bottom; line-height:1.35; }}
    td {{ white-space:nowrap; }}
    .col-ck {{ width:22px; min-width:22px; max-width:22px; padding:4px 2px; text-align:center; position:sticky; left:0; z-index:2; background:#fff; }}
    th.col-ck {{ z-index:3; background:#f8fafc; }}
    .col-sku {{ width:88px; min-width:60px; max-width:100px; text-align:left; position:sticky; left:22px; z-index:2; background:#fff; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    th.col-sku {{ z-index:3; background:#f8fafc; }}
    .col-subject {{ text-align:left; min-width:54px; max-width:110px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    th.col-subject {{ white-space:normal; max-width:80px; }}
    .col-text {{ text-align:left; white-space:nowrap; }}
    tbody tr:hover td {{ background:#f0fdfa !important; }}
    tbody tr:hover td.col-ck, tbody tr:hover td.col-sku {{ background:#f0fdfa !important; }}
    tbody tr.row-selected td {{ background:#d1fae5 !important; }}
    tbody tr.row-selected td.col-ck, tbody tr.row-selected td.col-sku {{ background:#d1fae5 !important; }}
    .empty {{ padding:18px; color:var(--muted); font-size:.9rem; }}
    th.g-sep {{ border-left:2px solid #94a3b8; }}
    td.g-sep {{ border-left:2px solid #e2e8f0; }}
    .price-up {{ color:#16a34a; }}
    .price-down {{ color:#dc2626; }}
    .price-arrow {{ font-size:9px; margin-left:2px; }}
    @media (max-width:1200px) {{ .summary {{ flex-wrap:nowrap; }} }}
    .ck-cell {{ text-align:center; width:26px; min-width:26px; padding:4px 2px; }}
    input[type=checkbox] {{ min-width:auto; width:14px; height:14px; cursor:pointer; padding:0; accent-color:var(--accent); }}
    .pct-cell {{ display:flex; justify-content:flex-end; }}
    .pct-spinner {{ display:inline-flex; border:1px solid #cbd5e1; border-radius:5px; overflow:hidden; height:22px; }}
    .pct-spinner input[type=number] {{ width:42px; min-width:0; border:none; padding:2px 4px; text-align:center; font-size:11px; -moz-appearance:textfield; background:#fff; }}
    .pct-spinner input[type=number]::-webkit-inner-spin-button {{ display:none; }}
    .pct-spinner input[type=number]:focus {{ outline:none; }}
    .pct-arrows {{ display:flex; flex-direction:column; border-left:1px solid #cbd5e1; }}
    .pct-btn, .drr-btn, .conv-btn {{ width:15px; padding:0; font-size:7px; background:#f8fafc; color:#475569; border:none; border-radius:0; flex:1; line-height:1; cursor:pointer; min-width:0; }}
    .pct-btn:first-child, .drr-btn:first-child, .conv-btn:first-child {{ border-bottom:1px solid #e2e8f0; }}
    .pct-btn:hover, .drr-btn:hover, .conv-btn:hover {{ background:#e2e8f0; }}
    .pct-cell {{ gap:3px; align-items:center; }}
    .val-arrow {{ font-size:9px; font-weight:700; min-width:14px; width:14px; text-align:center; flex-shrink:0; }}
    .val-arrow:not(:empty) {{ cursor:pointer; padding:0 2px; }}
    .val-arrow:not(:empty):hover {{ opacity:0.6; }}
    .pct-spinner.no-data {{ opacity:0.35; pointer-events:none; }}
    .pct-spinner input:disabled {{ color:#94a3b8; background:#f8fafc; text-align:center; }}
    .val-arrow.val-up {{ color:#16a34a; }}
    .val-arrow.val-down {{ color:#dc2626; }}
    input.val-changed-up {{ color:#16a34a; font-weight:700; }}
    input.val-changed-down {{ color:#dc2626; font-weight:700; }}
    thead tr:first-child th:not([rowspan]) {{ font-size:9px; text-transform:uppercase; letter-spacing:.04em; padding:3px 8px; }}
    thead tr:last-child th {{ top:22px; z-index:1; }}
    th.col-ck, th.col-sku {{ top:0 !important; z-index:4 !important; }}
    .th-params {{ background:#dbeafe !important; color:#1d4ed8; }}
    .th-result {{ background:#dcfce7 !important; color:#166534; }}
    .th-lost  {{ background:#fef9c3 !important; color:#854d0e; }}
    th.params-col {{ background:#eff6ff; }}
    th.result-col {{ background:#f0fdf4; }}
    th.lost-col   {{ background:#fefce8; }}
    .th-econ {{ background:#fdf4ff !important; color:#7e22ce; }}
    th.econ-col {{ background:#fdf4ff; }}
    td.econ-col {{ background:#fdf4ff44; }}
    th[data-sort-key] {{ cursor:pointer; user-select:none; }}
    th[data-sort-key]:hover {{ filter:brightness(0.93); }}
    th[data-sort-key][data-sort-dir="asc"]::after  {{ content:" ▲"; font-size:.65em; opacity:.7; }}
    th[data-sort-key][data-sort-dir="desc"]::after {{ content:" ▼"; font-size:.65em; opacity:.7; }}
    #bulk-bar {{ display:none; position:fixed; bottom:22px; left:50%; transform:translateX(-50%); background:#0f172a; color:#fff; border-radius:10px; padding:10px 16px; align-items:center; gap:10px; font-size:12px; box-shadow:0 4px 24px rgba(0,0,0,.35); z-index:100; white-space:nowrap; }}
    #bulk-bar input[type=number] {{ width:64px; min-width:0; padding:5px 7px; color:#0f172a; font-size:12px; }}
    #bulk-bar .apply-btn {{ background:#0f766e; }}
    #bulk-bar .reset-btn {{ background:transparent; border:1px solid #475569; color:#94a3b8; }}
    #bulk-bar .close-btn {{ background:transparent; border:none; color:#94a3b8; font-size:14px; padding:2px 6px; cursor:pointer; }}
    .methodology {{ margin-top:18px; background:#fff; border:1px solid var(--line); border-radius:8px; padding:16px 20px; font-size:.8rem; line-height:1.6; color:#334155; }}
    .methodology h2 {{ margin:0 0 12px; font-size:.88rem; font-weight:800; color:#0f172a; }}
    .methodology h3 {{ margin:14px 0 4px; font-size:.78rem; font-weight:800; color:var(--accent); text-transform:uppercase; letter-spacing:.04em; }}
    .methodology p {{ margin:0 0 6px; }}
    .methodology ul {{ margin:2px 0 6px; padding-left:18px; }}
    .methodology li {{ margin-bottom:2px; }}
    .methodology .formula {{ font-family:monospace; background:#f1f5f9; border-radius:4px; padding:2px 6px; font-size:.77rem; }}
  </style>
</head>
<body>
<div class="wrap">
  <div class="top">
    <h1>Планирование</h1>
    <div class="report-nav">{report_nav}</div>
  </div>
  <div class="ctrl-panel">
    <!-- Ряд 1: параметры -->
    <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#94a3b8;margin-bottom:2px">Параметры расчёта</div>
    <div class="ctrl-row">
      <div class="ctrl-field">
        <label for="forecast-to">Прогноз до</label>
        <input id="forecast-to" type="date" value="{forecast_to}" style="width:130px">
      </div>
      <div class="ctrl-field">
        <label for="return-delay">Возврат, дней</label>
        <input id="return-delay" type="number" min="1" max="60" step="1" value="8" style="width:72px">
      </div>
      <div class="ctrl-field">
        <label for="transit-days">Поставка FBO, дней</label>
        <input id="transit-days" type="number" min="0" max="30" step="1" value="3" style="width:72px">
      </div>
      <div class="ctrl-field">
        <label for="tax-pct">Налоги, %</label>
        <input id="tax-pct" type="number" min="0" max="50" step="0.1" value="0" style="width:72px">
      </div>
      <div class="ctrl-field">
        <label for="admin-pct">Адм. расходы, %</label>
        <input id="admin-pct" type="number" min="0" max="50" step="0.1" value="0" style="width:72px">
      </div>
      <div class="ctrl-sep"></div>
      <div>
        <div class="ctrl-group-label">Фильтры</div>
        <div style="display:flex;gap:6px;align-items:flex-end">
          <div class="ctrl-field">
            <label for="subject">Предмет</label>
            <select id="subject" style="width:160px"><option value="">Все предметы</option></select>
          </div>
          <div class="ctrl-field" style="display:none">
            <label for="strategy">Ярлыки</label>
            <select id="strategy" multiple size="4" style="width:140px;height:72px;" title="Ctrl/Cmd для множественного выбора"></select>
          </div>
          <div class="ctrl-field">
            <label for="revenue-category">Категория</label>
            <select id="revenue-category" style="width:140px"><option value="">Все категории</option></select>
          </div>
          <div class="ctrl-field">
            <label for="weeks-filter">Недель, макс.</label>
            <input id="weeks-filter" type="number" min="1" max="52" step="1" placeholder="Все" style="width:72px">
          </div>
        </div>
      </div>
      <button id="load" class="ctrl-btn ctrl-btn-primary" style="align-self:flex-end">Посчитать</button>
    </div>
    <!-- Ряд 2: действия -->
    <div class="ctrl-row" style="border-top:1px solid var(--line);padding-top:10px">
      <!-- Оптимизация -->
      <div class="ctrl-group">
        <div>
          <div class="ctrl-group-label">Оптимизация</div>
          <div style="display:flex;gap:6px">
            <button id="no-oos-btn" class="ctrl-btn ctrl-btn-teal" title="Подобрать минимальное повышение цены, чтобы не уйти в OOS до конца периода">Без OOS</button>
            <button id="optimize-btn" class="ctrl-btn ctrl-btn-purple" title="Найти цену и ДРР для максимального оборота при марже не ниже 0%">⚡ Макс. выручку</button>
          </div>
        </div>
      </div>
      <div class="ctrl-sep"></div>
      <!-- Сброс -->
      <div class="ctrl-group">
        <div>
          <div class="ctrl-group-label">Сбросить</div>
          <div style="display:flex;gap:6px">
            <button id="reset-prices" class="ctrl-btn ctrl-btn-gray">Цены</button>
            <button id="reset-drr"    class="ctrl-btn ctrl-btn-gray">ДРР</button>
            <button id="reset-conv"   class="ctrl-btn ctrl-btn-gray">Конверсии</button>
          </div>
        </div>
      </div>
      <div class="ctrl-sep"></div>
      <!-- Сценарии -->
      <div class="ctrl-group">
        <div>
          <div class="ctrl-group-label">Сценарий</div>
          <div style="display:flex;gap:6px;align-items:center">
            <select id="scenario-select" class="ctrl-scenario-select"><option value="">— выбрать —</option></select>
            <button id="scenario-load"   class="ctrl-btn ctrl-btn-teal"  style="padding:0 10px;font-size:12px">Загрузить</button>
            <button id="scenario-delete" class="ctrl-btn ctrl-btn-red"   style="padding:0 10px;font-size:12px">Удалить</button>
            <input type="text" id="scenario-name" class="ctrl-scenario-input" placeholder="Новый сценарий…">
            <button id="scenario-save"   class="ctrl-btn ctrl-btn-primary" style="padding:0 10px;font-size:12px">Сохранить</button>
          </div>
        </div>
      </div>
      <div class="ctrl-sep"></div>
      <!-- Экспорт и прочее -->
      <div class="ctrl-group" style="margin-left:auto">
        <div>
          <div class="ctrl-group-label">Экспорт и вид</div>
          <div style="display:flex;gap:6px;align-items:center">
            <button id="export-xls" class="ctrl-btn ctrl-btn-green">⬇ XLS</button>
            <div style="position:relative">
              <button id="col-settings-btn" class="ctrl-btn ctrl-btn-slate" style="padding:0 10px;font-size:16px" title="Настройка столбцов">⚙</button>
              <div id="col-settings-panel" style="display:none;position:absolute;right:0;top:36px;background:#fff;border:1px solid #cbd5e1;border-radius:8px;padding:12px 16px;z-index:300;box-shadow:0 4px 16px rgba(0,0,0,.18);white-space:nowrap;min-width:160px"></div>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>
  <div class="meta" id="meta">Загрузка...</div>
  <div style="margin-top:18px;margin-bottom:6px;display:flex;align-items:center;gap:10px">
    <span style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#94a3b8">Общие метрики</span>
    <div style="flex:1;height:1px;background:var(--line)"></div>
  </div>
  <div class="summary" id="summary"></div>
  <div class="tbl" id="table"></div>
</div>
<div id="bulk-bar">
  <span>Выбрано: <strong id="bulk-count">0</strong></span>
  <select id="bulk-type" style="padding:4px 6px;font-size:12px;background:#1e293b;color:#fff;border:1px solid #475569;border-radius:4px">
    <option value="price">Цена, %</option>
    <option value="drr">ДРР, %</option>
    <option value="ctr">CTR, %</option>
    <option value="cr1">CR1, %</option>
    <option value="cr2">CR2, %</option>
  </select>
  <input type="number" id="bulk-pct" value="0" step="0.5" min="-90" max="300" placeholder="±δ">
  <button class="apply-btn" id="bulk-apply">+ Применить</button>
  <button class="reset-btn" id="bulk-reset">Сбросить</button>
  <button class="close-btn" id="bulk-close">✕</button>
</div>
<script>
const stateKey = 'ozon.planning.filters';
const fields = {{
  forecastTo: document.getElementById('forecast-to'),
  returnDelay: document.getElementById('return-delay'),
  transitDays: document.getElementById('transit-days'),
  subject: document.getElementById('subject'),
  strategy: document.getElementById('strategy'),
  revenueCategory: document.getElementById('revenue-category'),
  weeksFilter: document.getElementById('weeks-filter'),
}};
const meta = document.getElementById('meta');
const summaryEl = document.getElementById('summary');
const tableEl = document.getElementById('table');
const bulkBar = document.getElementById('bulk-bar');
const bulkCount = document.getElementById('bulk-count');
const bulkPct = document.getElementById('bulk-pct');

let rowDataByNmid = {{}};
const priceChangesKey = 'ozon.planning.priceChanges';
let priceChanges = {{}};
let rowForecastCache = {{}};
const drrChangesKey = 'ozon.planning.drrChanges';
let drrChanges = {{}};  // nmid → delta_drr_pct (float, added to base_drr_pct)
const convChangesKey = 'ozon.planning.convChanges';
let convChanges = {{}};  // nmid → {{ ctr, cr1, cr2 }} (absolute %, null = use base)

function savePriceChanges() {{
  try {{ localStorage.setItem(priceChangesKey, JSON.stringify(priceChanges)); }} catch(e) {{}}
}}
function loadPriceChanges() {{
  try {{ priceChanges = JSON.parse(localStorage.getItem(priceChangesKey) || '{{}}'); }} catch(e) {{ priceChanges = {{}}; }}
}}
function saveDrrChanges() {{
  try {{ localStorage.setItem(drrChangesKey, JSON.stringify(drrChanges)); }} catch(e) {{}}
}}
function loadDrrChanges() {{
  try {{ drrChanges = JSON.parse(localStorage.getItem(drrChangesKey) || '{{}}'); }} catch(e) {{ drrChanges = {{}}; }}
}}
function saveConvChanges() {{
  try {{ localStorage.setItem(convChangesKey, JSON.stringify(convChanges)); }} catch(e) {{}}
}}
function loadConvChanges() {{
  try {{ convChanges = JSON.parse(localStorage.getItem(convChangesKey) || '{{}}'); }} catch(e) {{ convChanges = {{}}; }}
  // чистим нули, сохранённые старой версией кода (|| 0 вместо delete)
  for (const nmid of Object.keys(convChanges)) {{
    for (const k of Object.keys(convChanges[nmid] || {{}})) {{
      if (!(convChanges[nmid][k] > 0)) delete convChanges[nmid][k];
    }}
    if (!Object.keys(convChanges[nmid] || {{}}).length) delete convChanges[nmid];
  }}
}}
loadPriceChanges();
loadDrrChanges();
loadConvChanges();

// ── Сценарии ──────────────────────────────────────────────────────────────
const scenariosKey = 'ozon.planning.scenarios';
function loadScenarioList() {{
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  const sel = document.getElementById('scenario-select');
  const cur = sel.value;
  sel.innerHTML = '<option value="">— выбрать —</option>';
  Object.keys(all).sort().forEach(name => {{
    const opt = document.createElement('option');
    opt.value = name; opt.textContent = name;
    sel.appendChild(opt);
  }});
  if (cur && all[cur]) sel.value = cur;
}}
loadScenarioList();

document.getElementById('scenario-save').addEventListener('click', () => {{
  const name = document.getElementById('scenario-name').value.trim();
  if (!name) return;
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  all[name] = {{
    price: Object.assign({{}}, priceChanges),
    drr:   Object.assign({{}}, drrChanges),
    conv:  JSON.parse(JSON.stringify(convChanges)),
  }};
  localStorage.setItem(scenariosKey, JSON.stringify(all));
  loadScenarioList();
  document.getElementById('scenario-select').value = name;
  document.getElementById('scenario-name').value = '';
}});

document.getElementById('scenario-load').addEventListener('click', () => {{
  const name = document.getElementById('scenario-select').value;
  if (!name) return;
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  const sc = all[name];
  if (!sc) return;
  priceChanges = sc.price || {{}}; savePriceChanges();
  drrChanges   = sc.drr   || {{}}; saveDrrChanges();
  convChanges  = sc.conv  || {{}}; saveConvChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const pInp = tr.querySelector('.pct-input');
    if (pInp) pInp.value = priceChanges[nmid] ?? 0;
    const dInp = tr.querySelector('.drr-input');
    if (dInp) dInp.value = (drrChanges[nmid] != null ? drrChanges[nmid] : (row.base_drr_pct || 0)).toFixed(1);
    ['ctr','cr1','cr2'].forEach(k => {{
      const inp = tr.querySelector(`.conv-input[data-conv="${{k}}"]`);
      if (inp && !inp.disabled) {{
        const v = (convChanges[nmid] || {{}})[k];
        inp.value = (v > 0 ? v : (row[`base_${{k}}`] || 0)).toFixed(2);
      }}
    }});
    updateRowPlanned(nmid);
  }});
}});

document.getElementById('scenario-delete').addEventListener('click', () => {{
  const name = document.getElementById('scenario-select').value;
  if (!name) return;
  if (!confirm(`Удалить сценарий «${{name}}»?`)) return;
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  delete all[name];
  localStorage.setItem(scenariosKey, JSON.stringify(all));
  loadScenarioList();
}});

// ── Экспорт XLS ───────────────────────────────────────────────────────────
document.getElementById('export-xls').addEventListener('click', () => {{
  const taxOnlyPct   = Number(document.getElementById('tax-pct')?.value) || 0;
  const adminOnlyPct = Number(document.getElementById('admin-pct')?.value) || 0;
  const exportRows = Object.values(rowDataByNmid).map(row => {{
    const f = calcRowForecast(row);
    const drrPct    = drrChanges[row.nmid] != null ? drrChanges[row.nmid] : (row.base_drr_pct || 0);
    const revenue   = f.forecastBuyouts * f.plannedClient;
    const cogsTotal = f.forecastBuyouts * (row.cogs || 0);
    const wbComm    = revenue * ((row.wb_commission_pct || 0) / 100);
    const acquiring = revenue * 0.03;
    const drrCost   = revenue * (drrPct / 100);
    const grossProfit = revenue - cogsTotal - wbComm - acquiring - drrCost;
    const marginPct   = revenue > 0 ? grossProfit / revenue * 100 : 0;
    const adminCost   = revenue * (adminOnlyPct / 100);
    const taxCost     = revenue * (taxOnlyPct / 100);
    const netProfit   = grossProfit - adminCost - taxCost;
    const profitPct   = revenue > 0 ? netProfit / revenue * 100 : 0;
    return {{
      article:           row.article,
      subject:           row.subject,
      strategy:          (row.labels||[]).join(", "),
      revenue_cat:       row.revenue_category,
      stock:             row.stock,
      wh_stock:          row.warehouse_stock,
      stock_val:         row.stock * (row.client_price || row.average_check || 0),
      base_price:        row.baseline_seller_price,
      base_spp:          row.baseline_spp,
      base_client:       row.baseline_client_check,
      base_speed:        row.base_speed,
      cur_price:         row.current_price,
      cur_spp:           row.current_spp,
      cur_client:        row.client_price,
      elasticity:        row.elasticity,
      cur_speed:         row.current_speed,
      price_pct:         priceChanges[row.nmid] ?? 0,
      plan_drr:          drrPct,
      ads_elasticity:    row.ads_elasticity,
      base_ctr:          row.base_ctr,
      base_cr1:          row.base_cr1,
      base_cr2:          row.base_cr2,
      planned_price:     f.plannedSeller,
      planned_client:    f.plannedClient,
      planned_speed:     f.plannedSpeed,
      buyout_percent:    row.buyout_percent,
      forecast_orders:   f.forecastOrders,
      forecast_buyouts:  f.forecastBuyouts,
      forecast_revenue:  f.forecastRevenue,
      stockout_date:     f.stockoutDate || '',
      lost_orders:       f.lostOrders,
      lost_revenue:      f.lostRevenue,
      comm_pct:          row.wb_commission_pct || 0,
      comm_rub:          wbComm,
      acq_rub:           acquiring,
      drr_rub:           drrCost,
      cogs_total:        cogsTotal,
      margin_pct:        marginPct,
      margin_rub:        grossProfit,
      admin_rub:         adminCost,
      tax_rub:           taxCost,
      profit_pct:        profitPct,
      profit_rub:        netProfit,
    }};
  }});
  fetch('/api/ozon/planning/export', {{
    method: 'POST',
    headers: {{'Content-Type': 'application/json'}},
    body: JSON.stringify({{ rows: exportRows }}),
  }}).then(r => {{
    if (!r.ok) return r.text().then(t => {{ throw new Error(t); }});
    return r.blob();
  }}).then(blob => {{
    const a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    const today = new Date().toISOString().slice(0, 10);
    a.download = `planning_${{today}}.xlsx`;
    a.click();
  }}).catch(err => alert('Ошибка экспорта: ' + err));
}});

let planningParams = {{ forecast_days: 0, return_delay_days: 8, transit_days: 3, forecast_from: '' }};

function calcConvFactor(row) {{
  const c = convChanges[row.nmid] || {{}};
  let f = 1.0;
  if (c.ctr > 0 && row.base_ctr > 0) f *= c.ctr / row.base_ctr;
  if (c.cr1 > 0 && row.base_cr1 > 0) f *= c.cr1 / row.base_cr1;
  if (c.cr2 > 0 && row.base_cr2 > 0) f *= c.cr2 / row.base_cr2;
  return f;
}}

function simulatePlanning(stock, dailySpeed, buyoutRate, forecastDays, returnDelay, startDateStr, warehouseStock, transitDays) {{
  let available = Math.max(0, stock);
  const demand = Math.max(0, dailySpeed);
  buyoutRate = Math.max(0, Math.min(1, buyoutRate));
  const delay = Math.max(1, Math.round(returnDelay));
  const transit = Math.max(0, Math.round(transitDays || 0));
  const arrivals = {{}};
  if (warehouseStock > 0) arrivals[transit] = Math.max(0, warehouseStock);
  const returnsByDay = {{}};
  let totalOrders = 0;
  let stockoutDate = '';
  const startMs = startDateStr ? new Date(startDateStr).getTime() : 0;
  for (let idx = 0; idx < Math.max(0, forecastDays); idx++) {{
    available += (returnsByDay[idx] || 0) + (arrivals[idx] || 0);
    delete returnsByDay[idx];
    delete arrivals[idx];
    const ordered = Math.min(demand, available);
    if (demand > available && !stockoutDate && startMs) {{
      const d = new Date(startMs + idx * 86400000);
      stockoutDate = d.toISOString().slice(0, 10);
    }}
    available -= ordered;
    totalOrders += ordered;
    const returning = ordered * (1 - buyoutRate);
    const dueIdx = idx + delay;
    if (returning > 0 && dueIdx < forecastDays) {{
      returnsByDay[dueIdx] = (returnsByDay[dueIdx] || 0) + returning;
    }}
  }}
  return {{ totalOrders, stockoutDate }};
}}

function formatNumber(value, digits = 0) {{
  const num = Number(value || 0);
  return num.toLocaleString('ru-RU', {{ maximumFractionDigits: digits, minimumFractionDigits: digits }});
}}
function formatMoney(value) {{ return formatNumber(value, 0); }}

function saveState() {{
  try {{
    localStorage.setItem(stateKey, JSON.stringify({{
      forecastTo: fields.forecastTo.value,
      returnDelay: fields.returnDelay.value,
      transitDays: fields.transitDays.value,
      subject: fields.subject.value,
      labels: [...(fields.strategy?.selectedOptions||[])].map(o=>o.value).join(","),
      revenueCategory: fields.revenueCategory.value,
    }}));
  }} catch (err) {{}}
}}

function restoreState() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(stateKey) || '{{}}');
    if (saved.forecastTo) fields.forecastTo.value = saved.forecastTo;
    if (saved.returnDelay) fields.returnDelay.value = saved.returnDelay;
    if (saved.transitDays != null) fields.transitDays.value = saved.transitDays;
    fields.subject.dataset.pending = saved.subject || '';
    // labels not restored from storage - always start empty
    fields.revenueCategory.dataset.pending = saved.revenueCategory || '';
  }} catch (err) {{}}
}}

function fillSelect(select, values, placeholder) {{
  const current = select.value || select.dataset.pending || '';
  select.innerHTML = `<option value="">${{placeholder}}</option>` + (values || []).map(v => `<option value="${{String(v).replaceAll('"','&quot;')}}">${{v}}</option>`).join('');
  if ([...select.options].some(o => o.value === current)) select.value = current;
  select.dataset.pending = '';
}}
function fillMultiSelect(select, values) {{
  const prevVals = new Set([...select.selectedOptions].map(o=>o.value));
  const optionsHtml = (values || []).map(v => {{
    const sel = prevVals.has(v) ? ' selected' : '';
    return `<option value="${{String(v).replaceAll('"','&quot;')}}"${{sel}}>${{v}}</option>`;
  }}).join('');
  select.innerHTML = select.multiple ? optionsHtml : `<option value="">Все ярлыки</option>${{optionsHtml}}`;
}}

function renderSummary(summary) {{
  const avgMargin = summary.total_econ_revenue > 0
    ? (summary.total_gross_profit / summary.total_econ_revenue * 100).toFixed(1)
    : '—';
  summaryEl.innerHTML = `
    <div class="kpi">
      <div class="kpi-name">Себестоимость остатков</div>
      <div class="kpi-rub">${{formatMoney(summary.stock_cost)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.stock)}} шт на складе</div>
      <div class="kpi-qty" style="color:#64748b">В тек. ценах: ${{formatMoney(summary.stock_value_current || summary.stock_value)}}</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Плановые заказы</div>
      <div class="kpi-rub">${{formatMoney(summary.orders_revenue)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.forecast_orders)}} шт</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Плановые выкупы</div>
      <div class="kpi-rub">${{formatMoney(summary.forecast_revenue)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.forecast_buyouts)}} шт</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Упущенные продажи</div>
      <div class="kpi-rub">${{formatMoney(summary.lost_revenue)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.lost_orders)}} шт</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Расходы на рекламу</div>
      <div class="kpi-rub">${{formatMoney(summary.total_drr_cost)}}</div>
      <div class="kpi-qty">${{summary.total_econ_revenue > 0 ? (summary.total_drr_cost / summary.total_econ_revenue * 100).toFixed(1) : '—'}}% от выкупов</div>
    </div>
    <div class="kpi" style="border-color:#a855f7">
      <div class="kpi-name" style="color:#7e22ce">Маржинальность</div>
      <div class="kpi-rub" style="color:#7e22ce">${{avgMargin}}%</div>
      <div class="kpi-qty">валовая прибыль ${{formatMoney(summary.total_gross_profit)}}</div>
    </div>
    <div class="kpi" style="border-color:#a855f7">
      <div class="kpi-name" style="color:#7e22ce">Чистая прибыль</div>
      <div class="kpi-rub" style="color:${{(summary.total_net_profit||0)>=0?'#166534':'#dc2626'}}">${{formatMoney(summary.total_net_profit)}}</div>
      <div class="kpi-qty">после налогов и адм.</div>
    </div>`;
}}

function priceArrow(current, baseline, invertLogic) {{
  if (!baseline || !current) return '';
  const delta = (current - baseline) / baseline;
  if (Math.abs(delta) < 0.01) return '';
  const up = invertLogic ? delta < 0 : delta > 0;
  return `<span class="price-arrow ${{up ? 'price-up' : 'price-down'}}">${{up ? '↑' : '↓'}}</span>`;
}}

function calcNoOOS(row) {{
  const {{ forecast_days, return_delay_days, transit_days, forecast_from }} = planningParams;
  const buyoutRate = row.buyout_percent / 100;
  const wh = row.warehouse_stock || 0;
  const curPct = priceChanges[row.nmid] ?? 0;
  // check OOS at current planned speed (already includes DRR/conv factors)
  const {{ plannedSpeed: curSpeed }} = calcPlanned(row, curPct);
  const {{ stockoutDate: sd0 }} = simulatePlanning(row.stock, curSpeed, buyoutRate, forecast_days, return_delay_days, forecast_from, wh, transit_days);
  if (!sd0) return 0;  // no OOS at current plan → nothing to do
  if (row.elasticity >= 0) return null;  // can't reduce speed by raising price
  const baseClient = row.baseline_client_check;
  if (!baseClient) return null;
  // binary search starting from current price pct (don't lower existing change)
  let lo = curPct, hi = Math.max(300, curPct + 300);
  for (let i = 0; i < 50; i++) {{
    const mid = (lo + hi) / 2;
    const {{ plannedSpeed }} = calcPlanned(row, mid);
    const {{ stockoutDate }} = simulatePlanning(row.stock, plannedSpeed, buyoutRate, forecast_days, return_delay_days, forecast_from, wh, transit_days);
    if (!stockoutDate) hi = mid; else lo = mid;
    if (hi - lo < 0.05) break;
  }}
  if (hi >= curPct + 299) return null;  // can't avoid OOS even with huge increase
  return Math.ceil(hi);
}}

// Variant of calcPlanned with explicit drrPct (doesn't read global drrChanges)
function calcPlannedEx(row, pricePct, drrPct) {{
  const p = Number(pricePct) || 0;
  const plannedSeller = row.current_price * (1 + p / 100);
  const plannedClient = plannedSeller * (1 - row.current_spp / 100);
  const baseClient = row.baseline_client_check || plannedClient;
  const delta = (baseClient > 0 && plannedClient > 0) ? (plannedClient / baseClient - 1) : 0;
  const priceFactor = Math.max(0, 1 + row.elasticity * delta);
  let adsFactor = 1.0;
  const d = Number(drrPct) || 0;
  if (row.ads_elasticity > 0 && (row.base_drr_pct || 0) > 0) {{
    const adsRatio = Math.max(0.1, d / row.base_drr_pct);
    adsFactor = Math.pow(adsRatio, row.ads_elasticity);
  }}
  return {{ plannedSeller, plannedClient, plannedSpeed: row.base_speed * priceFactor * adsFactor * calcConvFactor(row) }};
}}

function netRevEx(row, pricePct, drrPct) {{
  const {{ plannedSpeed, plannedClient }} = calcPlannedEx(row, pricePct, drrPct);
  const {{ forecast_days, return_delay_days, transit_days, forecast_from }} = planningParams;
  const {{ totalOrders }} = simulatePlanning(
    row.stock, plannedSpeed, row.buyout_percent / 100,
    forecast_days, return_delay_days, forecast_from,
    row.warehouse_stock || 0, transit_days
  );
  const gross = totalOrders * (row.buyout_percent / 100) * plannedClient;
  return gross * (1 - (Number(drrPct) || 0) / 100);
}}

function findOptimalRevenue(row) {{
  if (!row.stock || !planningParams.forecast_days) return {{ pricePct: 0, drrPct: row.base_drr_pct || 0 }};
  const hasAds = row.ads_elasticity > 0 && (row.base_drr_pct || 0) > 0;
  const baseDrr = row.base_drr_pct || 0;

  // Cap upward price search: extrapolating far beyond data range is unreliable.
  // Inelastic products (|e|≈0) should not be pushed to +100% — market won't absorb it.
  const absE = Math.abs(row.elasticity || 0);
  const maxPriceUp = absE === 0 ? 5 : absE < 0.3 ? 10 : absE < 0.7 ? 20 : 30;

  let bestPct = 0, bestDrr = baseDrr, bestRev = -Infinity;

  if (hasAds) {{
    // Phase 1: joint coarse grid — price × DRR
    for (let p = -30; p <= maxPriceUp; p += 3) {{
      for (let d = 0; d <= 25; d++) {{
        const r = netRevEx(row, p, d);
        if (r > bestRev) {{ bestRev = r; bestPct = p; bestDrr = d; }}
      }}
    }}
    // Phase 2: refine around best — price ±3 step 1 × DRR ±1 step 0.1
    const p0 = bestPct, d0 = bestDrr;
    for (let p = p0 - 3; p <= Math.min(p0 + 3, maxPriceUp); p++) {{
      for (let di = -10; di <= 10; di++) {{
        const d = Math.round((d0 + di * 0.1) * 10) / 10;
        if (d < 0) continue;
        const r = netRevEx(row, p, d);
        if (r > bestRev) {{ bestRev = r; bestPct = p; bestDrr = d; }}
      }}
    }}
  }} else {{
    // No ads elasticity — search only price
    for (let p = -30; p <= maxPriceUp; p += 3) {{
      const r = netRevEx(row, p, baseDrr);
      if (r > bestRev) {{ bestRev = r; bestPct = p; }}
    }}
    const p0 = bestPct;
    for (let p = p0 - 3; p <= Math.min(p0 + 3, maxPriceUp); p++) {{
      const r = netRevEx(row, p, baseDrr);
      if (r > bestRev) {{ bestRev = r; bestPct = p; }}
    }}
    bestDrr = baseDrr;
  }}

  return {{ pricePct: bestPct, drrPct: Math.round(bestDrr * 10) / 10 }};
}}

function calcPlanned(row, pct) {{
  const p = Number(pct) || 0;
  const plannedSeller = row.current_price * (1 + p / 100);
  const plannedClient = plannedSeller * (1 - row.current_spp / 100);
  const baseClient = row.baseline_client_check || plannedClient;
  const delta = (baseClient > 0 && plannedClient > 0) ? (plannedClient / baseClient - 1) : 0;
  const priceFactor = Math.max(0, 1 + row.elasticity * delta);
  // ads factor: per-row planned DRR% (absolute), applied via bivariate elasticity
  let adsFactor = 1.0;
  const plannedDrr = drrChanges[row.nmid] != null ? drrChanges[row.nmid] : (row.base_drr_pct || 0);
  if (row.ads_elasticity > 0 && (row.base_drr_pct || 0) > 0) {{
    // ratio of planned to base DRR%; floor at 0.1 models organic traffic baseline
    const adsRatio = Math.max(0.1, plannedDrr / row.base_drr_pct);
    adsFactor = Math.pow(adsRatio, row.ads_elasticity);
  }}
  // conversion factor
  const convFactor = calcConvFactor(row);
  return {{ plannedSeller, plannedClient, plannedSpeed: row.base_speed * priceFactor * adsFactor * convFactor }};
}}

function calcRowForecast(row) {{
  const pct = priceChanges[row.nmid] ?? 0;
  const {{ plannedSeller, plannedClient, plannedSpeed }} = calcPlanned(row, pct);
  const {{ forecast_days, return_delay_days, transit_days, forecast_from }} = planningParams;
  const {{ totalOrders: forecastOrders, stockoutDate }} = simulatePlanning(
    row.stock, plannedSpeed, row.buyout_percent / 100,
    forecast_days, return_delay_days, forecast_from,
    row.warehouse_stock || 0, transit_days
  );
  const buyoutRate = row.buyout_percent / 100;
  const effectiveCheck = plannedClient > 0 ? plannedClient : row.average_check;
  const forecastBuyouts = forecastOrders * buyoutRate;
  const forecastRevenue = forecastBuyouts * effectiveCheck;
  const potentialOrders = plannedSpeed * forecast_days;
  const lostOrders = Math.max(0, potentialOrders - forecastOrders);
  const lostRevenue = lostOrders * buyoutRate * effectiveCheck;
  return {{ plannedSeller, plannedClient, plannedSpeed, effectiveCheck, forecastOrders, forecastBuyouts, forecastRevenue, stockoutDate, lostOrders, lostRevenue }};
}}

function recalcSummary() {{
  const s = {{ stock_cost: 0, stock_value: 0, stock_value_current: 0, forecast_orders: 0, forecast_buyouts: 0, orders_revenue: 0, forecast_revenue: 0, lost_orders: 0, lost_revenue: 0, stock: 0, total_gross_profit: 0, total_net_profit: 0, total_econ_revenue: 0, total_drr_cost: 0 }};
  Object.values(rowDataByNmid).forEach(row => {{
    const c = rowForecastCache[row.nmid] || {{}};
    s.stock += row.stock;
    s.stock_cost += row.stock * (row.cogs || 0);
    s.stock_value += row.stock * (row.client_price || row.average_check || 0);
    s.stock_value_current += row.stock * (row.client_price || row.average_check || 0);
    s.forecast_orders += c.forecastOrders || 0;
    s.forecast_buyouts += c.forecastBuyouts || 0;
    s.orders_revenue += c.ordersRevenue || 0;
    s.forecast_revenue += c.forecastRevenue || 0;
    s.lost_orders += c.lostOrders || 0;
    s.lost_revenue += c.lostRevenue || 0;
    s.total_gross_profit += c.grossProfit || 0;
    s.total_net_profit   += c.netProfit || 0;
    s.total_econ_revenue += c.econRevenue || 0;
    s.total_drr_cost     += c.drrCost || 0;
  }});
  renderSummary(s);
}}

function updateRowPlanned(nmid, skipSummary) {{
  const row = rowDataByNmid[nmid];
  if (!row) return;
  const f = calcRowForecast(row);

  // Economics
  const taxOnlyPct   = Number(document.getElementById('tax-pct')?.value) || 0;
  const adminOnlyPct = Number(document.getElementById('admin-pct')?.value) || 0;
  const drrPct = drrChanges[nmid] != null ? drrChanges[nmid] : (row.base_drr_pct || 0);
  const revenue    = f.forecastBuyouts * f.plannedClient;
  const cogsTotal  = f.forecastBuyouts * (row.cogs || 0);
  const wbComm     = revenue * ((row.wb_commission_pct || 0) / 100);
  const acquiring  = revenue * 0.03;
  const drrCost    = revenue * (drrPct / 100);
  const grossProfit = revenue - cogsTotal - wbComm - acquiring - drrCost;
  const marginPct   = revenue > 0 ? grossProfit / revenue * 100 : 0;
  const adminCost   = revenue * (adminOnlyPct / 100);
  const taxCost     = revenue * (taxOnlyPct / 100);
  const netProfit   = grossProfit - adminCost - taxCost;
  const profitPct   = revenue > 0 ? netProfit / revenue * 100 : 0;

  rowForecastCache[nmid] = {{
    forecastOrders: f.forecastOrders,
    forecastBuyouts: f.forecastBuyouts,
    ordersRevenue: f.forecastOrders * f.effectiveCheck,
    forecastRevenue: f.forecastRevenue,
    lostOrders: f.lostOrders,
    lostRevenue: f.lostRevenue,
    grossProfit,
    netProfit,
    econRevenue: revenue,
    drrCost,
  }};
  const tr = tableEl.querySelector(`tr[data-nmid="${{nmid}}"]`);
  if (!tr) return;
  tr.querySelector('.cell-ps').textContent = formatMoney(f.plannedSeller);
  tr.querySelector('.cell-pc').textContent = formatMoney(f.plannedClient);
  tr.querySelector('.cell-pspd').textContent = formatNumber(f.plannedSpeed, 2);
  tr.querySelector('.cell-fo').textContent = formatNumber(f.forecastOrders);
  tr.querySelector('.cell-fb').textContent = formatNumber(f.forecastBuyouts);
  tr.querySelector('.cell-fr').textContent = formatMoney(f.forecastRevenue);
  tr.querySelector('.cell-oos').textContent = f.stockoutDate || '—';
  tr.querySelector('.cell-lo').textContent = formatNumber(f.lostOrders);
  tr.querySelector('.cell-lr').textContent = formatMoney(f.lostRevenue);

  // Economics cells
  function setCell(sel, text, val, colorVal) {{
    const el = tr.querySelector(sel);
    if (!el) return;
    el.textContent = text;
    if (val !== undefined) el.dataset[Object.keys(el.dataset)[0] || 'v'] = val;
    if (colorVal !== undefined) el.style.color = colorVal >= 0 ? '#166534' : '#dc2626';
  }}
  tr.querySelector('.cell-econ-comm-rub') && (tr.querySelector('.cell-econ-comm-rub').textContent = formatMoney(wbComm));
  tr.querySelector('.cell-econ-acq-rub')  && (tr.querySelector('.cell-econ-acq-rub').textContent  = formatMoney(acquiring));
  tr.querySelector('.cell-econ-drr-rub')  && (tr.querySelector('.cell-econ-drr-rub').textContent  = formatMoney(drrCost));
  tr.querySelector('.cell-econ-cogs')     && (tr.querySelector('.cell-econ-cogs').textContent     = formatMoney(cogsTotal));
  const marginCell = tr.querySelector('.cell-econ-margin');
  if (marginCell) {{
    marginCell.textContent = marginPct.toFixed(1) + '%';
    marginCell.dataset.margin = marginPct.toFixed(1);
    marginCell.style.color = marginPct >= 0 ? '#166534' : '#dc2626';
  }}
  tr.querySelector('.cell-econ-margin-rub') && (tr.querySelector('.cell-econ-margin-rub').textContent = formatMoney(grossProfit));
  tr.querySelector('.cell-econ-admin-rub')  && (tr.querySelector('.cell-econ-admin-rub').textContent  = formatMoney(adminCost));
  tr.querySelector('.cell-econ-tax-rub')    && (tr.querySelector('.cell-econ-tax-rub').textContent    = formatMoney(taxCost));
  const profitPctCell = tr.querySelector('.cell-econ-profit-pct');
  if (profitPctCell) {{
    profitPctCell.textContent = profitPct.toFixed(1) + '%';
    profitPctCell.dataset.profitPct = profitPct.toFixed(1);
    profitPctCell.style.color = profitPct >= 0 ? '#166534' : '#dc2626';
  }}
  const profitCell = tr.querySelector('.cell-econ-profit');
  if (profitCell) {{
    profitCell.textContent = formatMoney(netProfit);
    profitCell.dataset.profit = netProfit.toFixed(0);
    profitCell.style.color = netProfit >= 0 ? '#166534' : '#dc2626';
  }}

  // WB stock warning
  const {{ transit_days }} = planningParams;
  const daysLeft = f.plannedSpeed > 0 ? row.stock / f.plannedSpeed : Infinity;
  const stockCell = tr.querySelector('.cell-wbstock');
  if (stockCell) {{
    const warn = isFinite(daysLeft) && daysLeft < transit_days && (row.warehouse_stock || 0) > 0;
    stockCell.style.color = warn ? '#dc2626' : '';
    stockCell.title = warn ? `Кончится через ~${{daysLeft.toFixed(1)}} дн., срок поставки ${{transit_days}} дн.` : '';
  }}

  // arrow indicators for drr / conv inputs
  function applyValArrow(input, arrowEl, baseVal) {{
    if (!input || !arrowEl) return;
    const step = parseFloat(input.step) || 1;
    const dec = step <= 0.15 ? 2 : 1;
    // Compare as strings at the same precision to avoid float mismatch (e.g. base=36.65 → toFixed(1)="36.6" but Math.round(36.65*10)=367)
    const curStr  = parseFloat(input.value).toFixed(dec);
    const baseStr = (baseVal || 0).toFixed(dec);
    const cur  = parseFloat(curStr);
    const base = parseFloat(baseStr);
    input.classList.remove('val-changed-up', 'val-changed-down');
    arrowEl.classList.remove('val-up', 'val-down');
    arrowEl.textContent = '';
    if (cur !== base) {{
      const up = cur > base;
      input.classList.add(up ? 'val-changed-up' : 'val-changed-down');
      arrowEl.classList.add(up ? 'val-up' : 'val-down');
      arrowEl.textContent = up ? '↑' : '↓';
    }}
  }}
  applyValArrow(tr.querySelector('.drr-input'), tr.querySelector('.drr-arrow'), row.base_drr_pct);
  applyValArrow(tr.querySelector('.conv-input[data-conv="ctr"]'), tr.querySelector('.ctr-arrow'), row.base_ctr);
  applyValArrow(tr.querySelector('.conv-input[data-conv="cr1"]'), tr.querySelector('.cr1-arrow'), row.base_cr1);
  applyValArrow(tr.querySelector('.conv-input[data-conv="cr2"]'), tr.querySelector('.cr2-arrow'), row.base_cr2);

  if (!skipSummary) recalcSummary();
}}

function applyWeeksFilter() {{
  const max = Number(fields.weeksFilter.value);
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const row = rowDataByNmid[tr.dataset.nmid];
    if (!row) return;
    const weeks = row.base_speed > 0 ? row.stock / row.base_speed / 7 : Infinity;
    tr.style.display = (!max || weeks <= max) ? '' : 'none';
  }});
}}

function applyElasticityFilter() {{
  const val = fields.elasticityFilter.value;
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const hasElasticity = row.elasticity !== 0;
    const visible = !val || (val === 'elastic' ? hasElasticity : !hasElasticity);
    tr.style.display = visible ? '' : 'none';
  }});
}}

function updateBulkBar() {{
  const checked = tableEl.querySelectorAll('input.row-ck:checked');
  const n = checked.length;
  bulkCount.textContent = n;
  bulkBar.style.display = n > 0 ? 'flex' : 'none';
}}

function renderTable(rows) {{
  if (!rows.length) {{
    tableEl.innerHTML = '<div class="empty">Нет товаров для выбранных фильтров</div>';
    return;
  }}
  rowDataByNmid = {{}};
  rowForecastCache = {{}};
  const _taxPctInit   = Number(document.getElementById('tax-pct')?.value) || 0;
  const _adminPctInit = Number(document.getElementById('admin-pct')?.value) || 0;
  rows.forEach(r => {{
    rowDataByNmid[r.nmid] = r;
    const _rev = r.forecast_buyouts * (r.planned_price || r.client_price || r.average_check || 0);
    const _cogs = r.forecast_buyouts * (r.cogs || 0);
    const _comm = _rev * ((r.wb_commission_pct || 0) / 100);
    const _acq  = _rev * 0.03;
    const _drr  = _rev * ((r.base_drr_pct || 0) / 100);
    const _gross = _rev - _cogs - _comm - _acq - _drr;
    const _net   = _gross - _rev * (_adminPctInit / 100) - _rev * (_taxPctInit / 100);
    rowForecastCache[r.nmid] = {{
      forecastOrders: r.forecast_orders,
      forecastBuyouts: r.forecast_buyouts,
      ordersRevenue: r.forecast_orders * (r.effective_check || r.average_check),
      forecastRevenue: r.forecast_revenue,
      lostOrders: r.lost_orders,
      lostRevenue: r.lost_revenue,
      grossProfit: _gross,
      netProfit: _net,
      econRevenue: _rev,
    }};
  }});

  const head = `<tr>
    <th class="col-ck" rowspan="2"><input type="checkbox" id="ck-all" title="Выбрать все"></th>
    <th class="col-sku" rowspan="2" data-sort-key="article">SKU</th>
    <th class="col-subject" rowspan="2" data-sort-key="subject">Предмет</th>
    <th class="col-text" rowspan="2" data-sort-key="strategy" style="display:none">Ярлыки</th>
    <th class="col-text" rowspan="2" data-sort-key="revenue_category">Категория</th>
    <th class="g-sep" colspan="3" data-col-group="остатки">Остатки</th>
    <th class="g-sep" colspan="4" data-col-group="базовый">Базовый период 7д</th>
    <th class="g-sep" colspan="5" data-col-group="текущее">Текущее</th>
    <th class="g-sep th-params" colspan="6" data-col-group="параметры">Параметры плана</th>
    <th class="g-sep th-result" colspan="7" data-col-group="результат">Результат</th>
    <th class="g-sep th-lost" colspan="3" data-col-group="упущено">Упущено</th>
    <th class="g-sep th-econ" colspan="10" data-col-group="экономика">Экономика</th>
  </tr><tr>
    <th class="g-sep" data-col-group="остатки" data-col-key="c-stock" data-sort-key="stock">WB, шт</th><th data-col-group="остатки" data-col-key="c-wh-stock" data-sort-key="warehouse_stock">Склад, шт</th><th data-col-group="остатки" data-col-key="c-stock-val" data-sort-key="stock_value">Сумма, ₽</th>
    <th class="g-sep" data-col-group="базовый" data-col-key="c-base-price" data-sort-key="baseline_seller_price">Цена, ₽</th><th data-col-group="базовый" data-col-key="c-base-spp" data-sort-key="baseline_spp">СПП, %</th><th data-col-group="базовый" data-col-key="c-base-client" data-sort-key="baseline_client_check">Кл. цена, ₽</th><th data-col-group="базовый" data-col-key="c-base-speed" data-sort-key="base_speed">Скорость, шт</th>
    <th class="g-sep" data-col-group="текущее" data-col-key="c-cur-price" data-sort-key="current_price">Тек. цена, ₽</th><th data-col-group="текущее" data-col-key="c-cur-spp" data-sort-key="current_spp">СПП, %</th><th data-col-group="текущее" data-col-key="c-cur-client" data-sort-key="client_price">Кл. цена, ₽</th><th data-col-group="текущее" data-col-key="c-cur-elast" data-sort-key="elasticity">Эл. цены</th><th data-col-group="текущее" data-col-key="c-cur-speed" data-sort-key="current_speed">Скорость, шт</th>
    <th class="g-sep params-col" data-col-group="параметры" data-col-key="c-plan-price">Цена, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-drr" data-sort-key="base_drr_pct">ДРР, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-drr-e" data-sort-key="ads_elasticity">Эл.рекл.</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-ctr" data-sort-key="base_ctr">CTR, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-cr1" data-sort-key="base_cr1">CR1, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-cr2" data-sort-key="base_cr2">CR2, %</th>
    <th class="g-sep result-col" data-col-group="результат" data-col-key="c-res-price">Пл. цена, ₽</th><th class="result-col" data-col-group="результат" data-col-key="c-res-client">Пл. кл., ₽</th><th class="result-col" data-col-group="результат" data-col-key="c-res-speed" data-sort-key="plannedSpeed">Скорость, шт</th><th class="result-col" data-col-group="результат" data-col-key="c-res-buyout" data-sort-key="buyout_percent">% выкупа</th><th class="result-col" data-col-group="результат" data-col-key="c-res-orders" data-sort-key="forecast_orders">Заказы, шт</th><th class="result-col" data-col-group="результат" data-col-key="c-res-buyouts" data-sort-key="forecast_buyouts">Выкупы, шт</th><th class="result-col" data-col-group="результат" data-col-key="c-res-revenue" data-sort-key="forecast_revenue">Выручка, ₽</th>
    <th class="g-sep lost-col" data-col-group="упущено" data-col-key="c-res-oos">Дата стопа</th><th class="lost-col" data-col-group="упущено" data-col-key="c-lost-qty" data-sort-key="lost_orders">Упущ., шт</th><th class="lost-col" data-col-group="упущено" data-col-key="c-lost-rub" data-sort-key="lost_revenue">Упущ., ₽</th>
    <th class="g-sep econ-col" data-col-group="экономика" data-col-key="c-econ-comm-pct" data-sort-key="wb_commission_pct">Ком. %</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-comm-rub" data-sort-key="comm_rub">Ком. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-acq-rub" data-sort-key="acq_rub">Эквайр. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-drr-rub" data-sort-key="drr_rub">Реклама ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-cogs" data-sort-key="cogs">Себест. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-margin-pct" data-sort-key="margin_pct">Маржа %</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-margin-rub" data-sort-key="margin_rub">Маржа ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-admin-rub" data-sort-key="admin_rub">Адм. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-tax-rub" data-sort-key="tax_rub">Налоги ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-profit-pct" data-sort-key="profit_pct">Прибыль %</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-profit-rub" data-sort-key="net_profit">Прибыль ₽</th>
  </tr>`;

  const body = rows.map(row => {{
    const pct = priceChanges[row.nmid] ?? 0;
    const drrVal = drrChanges[row.nmid] != null ? drrChanges[row.nmid] : (row.base_drr_pct || 0);
    const ctrVal = (convChanges[row.nmid] || {{}}).ctr != null ? (convChanges[row.nmid] || {{}}).ctr : (row.base_ctr || 0);
    const cr1Val = (convChanges[row.nmid] || {{}}).cr1 != null ? (convChanges[row.nmid] || {{}}).cr1 : (row.base_cr1 || 0);
    const cr2Val = (convChanges[row.nmid] || {{}}).cr2 != null ? (convChanges[row.nmid] || {{}}).cr2 : (row.base_cr2 || 0);
    const {{ plannedSeller, plannedClient, plannedSpeed }} = calcPlanned(row, pct);
    return `<tr data-nmid="${{row.nmid}}">
      <td class="col-ck"><input type="checkbox" class="row-ck"></td>
      <td class="col-sku" title="${{row.article || row.nmid}}">${{row.article || row.nmid}}</td>
      <td class="col-subject" title="${{row.subject || ''}}">${{row.subject || ''}}</td>
      <td class="col-text" style="display:none">${{(row.labels||[]).join(', ')}}</td>
      <td class="col-text">${{row.revenue_category || ''}}</td>
      <td class="g-sep cell-wbstock" data-col-group="остатки" data-col-key="c-stock">${{formatNumber(row.stock)}}</td>
      <td data-col-group="остатки" data-col-key="c-wh-stock">${{formatNumber(row.warehouse_stock)}}</td>
      <td data-col-group="остатки" data-col-key="c-stock-val">${{formatMoney(row.stock * (row.client_price || row.average_check || 0))}}</td>
      <td class="g-sep" data-col-group="базовый" data-col-key="c-base-price">${{formatMoney(row.baseline_seller_price)}}</td>
      <td data-col-group="базовый" data-col-key="c-base-spp">${{formatNumber(row.baseline_spp, 1)}}%</td>
      <td data-col-group="базовый" data-col-key="c-base-client">${{formatMoney(row.baseline_client_check)}}</td>
      <td data-col-group="базовый" data-col-key="c-base-speed">${{formatNumber(row.base_speed, 2)}}</td>
      <td class="g-sep" data-col-group="текущее" data-col-key="c-cur-price">${{formatMoney(row.current_price)}}${{priceArrow(row.current_price, row.baseline_seller_price)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-spp">${{formatNumber(row.current_spp, 1)}}%${{priceArrow(row.current_spp, row.baseline_spp, true)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-client">${{formatMoney(row.client_price)}}${{priceArrow(row.client_price, row.baseline_client_check)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-elast">${{formatNumber(row.elasticity, 2)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-speed">${{formatNumber(row.current_speed, 2)}}</td>
      <td class="g-sep" data-col-group="параметры" data-col-key="c-plan-price">
        <div class="pct-cell">
          <div class="pct-spinner">
            <input type="number" class="pct-input" value="${{pct}}" step="1" min="-90" max="300">
            <div class="pct-arrows">
              <button class="pct-btn" data-d="1">▲</button>
              <button class="pct-btn" data-d="-1">▼</button>
            </div>
          </div>
        </div>
      </td>
      <td data-col-group="параметры" data-col-key="c-plan-drr">
        <div class="pct-cell" title="Базовый ДРР: ${{formatNumber(row.base_drr_pct, 1)}}%">
          <span class="val-arrow drr-arrow"></span>
          <div class="pct-spinner">
            <input type="number" class="drr-input" value="${{(+drrVal || 0).toFixed(1)}}" step="0.5" min="0" max="100">
            <div class="pct-arrows">
              <button class="drr-btn" data-d="0.5">▲</button>
              <button class="drr-btn" data-d="-0.5">▼</button>
            </div>
          </div>
        </div>
      </td>
      <td class="cell-drr-e" data-col-group="параметры" data-col-key="c-plan-drr-e">${{row.ads_elasticity ? formatNumber(row.ads_elasticity, 2) : '—'}}</td>
      <td data-col-group="параметры" data-col-key="c-plan-ctr">
        ${{(function(){{
          const noBase = !row.base_ctr && !(convChanges[row.nmid]||{{}}).ctr;
          const v = noBase ? '' : (+ctrVal||0).toFixed(2);
          return `<div class="pct-cell" title="${{noBase ? 'Нет данных за период' : 'Базовый CTR: '+formatNumber(row.base_ctr,2)+'%'}}">
            <span class="val-arrow ctr-arrow"></span>
            <div class="pct-spinner${{noBase?' no-data':''}}">
              <input type="number" class="conv-input" data-conv="ctr" value="${{v}}" placeholder="${{noBase?'—':''}}" step="0.1" min="0" max="100" ${{noBase?'disabled':''}}>
              <div class="pct-arrows"><button class="conv-btn" data-conv="ctr" data-d="0.1">▲</button><button class="conv-btn" data-conv="ctr" data-d="-0.1">▼</button></div>
            </div></div>`;
        }})()}}
      </td>
      <td data-col-group="параметры" data-col-key="c-plan-cr1">
        ${{(function(){{
          const noBase = !row.base_cr1 && !(convChanges[row.nmid]||{{}}).cr1;
          const v = noBase ? '' : (+cr1Val||0).toFixed(2);
          return `<div class="pct-cell" title="${{noBase ? 'Нет данных за период' : 'Базовый CR1: '+formatNumber(row.base_cr1,2)+'%'}}">
            <span class="val-arrow cr1-arrow"></span>
            <div class="pct-spinner${{noBase?' no-data':''}}">
              <input type="number" class="conv-input" data-conv="cr1" value="${{v}}" placeholder="${{noBase?'—':''}}" step="0.1" min="0" max="100" ${{noBase?'disabled':''}}>
              <div class="pct-arrows"><button class="conv-btn" data-conv="cr1" data-d="0.1">▲</button><button class="conv-btn" data-conv="cr1" data-d="-0.1">▼</button></div>
            </div></div>`;
        }})()}}
      </td>
      <td data-col-group="параметры" data-col-key="c-plan-cr2">
        ${{(function(){{
          const noBase = !row.base_cr2 && !(convChanges[row.nmid]||{{}}).cr2;
          const v = noBase ? '' : (+cr2Val||0).toFixed(2);
          return `<div class="pct-cell" title="${{noBase ? 'Нет данных за период' : 'Базовый CR2: '+formatNumber(row.base_cr2,2)+'%'}}">
            <span class="val-arrow cr2-arrow"></span>
            <div class="pct-spinner${{noBase?' no-data':''}}">
              <input type="number" class="conv-input" data-conv="cr2" value="${{v}}" placeholder="${{noBase?'—':''}}" step="0.1" min="0" max="100" ${{noBase?'disabled':''}}>
              <div class="pct-arrows"><button class="conv-btn" data-conv="cr2" data-d="0.1">▲</button><button class="conv-btn" data-conv="cr2" data-d="-0.1">▼</button></div>
            </div></div>`;
        }})()}}
      </td>
      <td class="g-sep cell-ps" data-col-group="результат" data-col-key="c-res-price">${{formatMoney(plannedSeller)}}</td>
      <td class="cell-pc" data-col-group="результат" data-col-key="c-res-client">${{formatMoney(plannedClient)}}</td>
      <td class="cell-pspd" data-col-group="результат" data-col-key="c-res-speed">${{formatNumber(plannedSpeed, 2)}}</td>
      <td data-col-group="результат" data-col-key="c-res-buyout">${{formatNumber(row.buyout_percent)}}%</td>
      <td class="cell-fo" data-col-group="результат" data-col-key="c-res-orders">${{formatNumber(row.forecast_orders)}}</td>
      <td class="cell-fb" data-col-group="результат" data-col-key="c-res-buyouts">${{formatNumber(row.forecast_buyouts)}}</td>
      <td class="cell-fr" data-col-group="результат" data-col-key="c-res-revenue">${{formatMoney(row.forecast_revenue)}}</td>
      <td class="cell-oos g-sep" data-col-group="упущено" data-col-key="c-res-oos">${{row.stockout_date || '—'}}</td>
      <td class="g-sep cell-lo" data-col-group="упущено" data-col-key="c-lost-qty">${{formatNumber(row.lost_orders)}}</td>
      <td class="cell-lr" data-col-group="упущено" data-col-key="c-lost-rub">${{formatMoney(row.lost_revenue)}}</td>
      <td class="g-sep econ-col cell-wb-comm" data-col-group="экономика" data-col-key="c-econ-comm-pct">${{formatNumber(row.wb_commission_pct || 0, 1)}}%</td>
      <td class="econ-col cell-econ-comm-rub" data-col-group="экономика" data-col-key="c-econ-comm-rub">—</td>
      <td class="econ-col cell-econ-acq-rub" data-col-group="экономика" data-col-key="c-econ-acq-rub">—</td>
      <td class="econ-col cell-econ-drr-rub" data-col-group="экономика" data-col-key="c-econ-drr-rub">—</td>
      <td class="econ-col cell-econ-cogs" data-col-group="экономика" data-col-key="c-econ-cogs">—</td>
      <td class="econ-col cell-econ-margin" data-col-group="экономика" data-col-key="c-econ-margin-pct" data-margin="0">—</td>
      <td class="econ-col cell-econ-margin-rub" data-col-group="экономика" data-col-key="c-econ-margin-rub">—</td>
      <td class="econ-col cell-econ-admin-rub" data-col-group="экономика" data-col-key="c-econ-admin-rub">—</td>
      <td class="econ-col cell-econ-tax-rub" data-col-group="экономика" data-col-key="c-econ-tax-rub">—</td>
      <td class="econ-col cell-econ-profit-pct" data-col-group="экономика" data-col-key="c-econ-profit-pct" data-profit-pct="0">—</td>
      <td class="econ-col cell-econ-profit" data-col-group="экономика" data-col-key="c-econ-profit-rub" data-profit="0">—</td>
    </tr>`;
  }}).join('');

  tableEl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;

  // checkbox: select all
  document.getElementById('ck-all').addEventListener('change', e => {{
    tableEl.querySelectorAll('input.row-ck').forEach(ck => {{
      ck.checked = e.target.checked;
      ck.closest('tr')?.classList.toggle('row-selected', e.target.checked);
    }});
    updateBulkBar();
  }});

  // checkbox: individual
  tableEl.addEventListener('change', e => {{
    if (e.target.classList.contains('row-ck')) {{
      e.target.closest('tr')?.classList.toggle('row-selected', e.target.checked);
      updateBulkBar();
    }}
  }});

  // price pct input & buttons
  tableEl.addEventListener('input', e => {{
    const tr = e.target.closest('tr');
    if (!tr) return;
    const nmid = tr.dataset.nmid;
    if (e.target.classList.contains('pct-input')) {{
      priceChanges[nmid] = Number(e.target.value) || 0;
      savePriceChanges();
      updateRowPlanned(nmid);
    }} else if (e.target.classList.contains('drr-input')) {{
      drrChanges[nmid] = Number(e.target.value) || 0;
      saveDrrChanges();
      updateRowPlanned(nmid);
    }} else if (e.target.classList.contains('conv-input')) {{
      const conv = e.target.dataset.conv;
      const val = parseFloat(e.target.value);
      if (!convChanges[nmid]) convChanges[nmid] = {{}};
      if (val > 0) {{
        convChanges[nmid][conv] = val;
      }} else {{
        delete convChanges[nmid][conv];
      }}
      saveConvChanges();
      updateRowPlanned(nmid);
    }}
  }});
  tableEl.addEventListener('click', e => {{
    // arrow click → reset that value to baseline
    const arrow = e.target.closest('.val-arrow');
    if (arrow && arrow.textContent) {{
      const tr = arrow.closest('tr');
      const nmid = tr?.dataset.nmid;
      const row = rowDataByNmid[nmid];
      if (!row) return;
      if (arrow.classList.contains('drr-arrow')) {{
        delete drrChanges[nmid];
        saveDrrChanges();
        const inp = tr.querySelector('.drr-input');
        if (inp) inp.value = (row.base_drr_pct || 0).toFixed(1);
      }} else {{
        let conv = null;
        if (arrow.classList.contains('ctr-arrow')) conv = 'ctr';
        else if (arrow.classList.contains('cr1-arrow')) conv = 'cr1';
        else if (arrow.classList.contains('cr2-arrow')) conv = 'cr2';
        if (conv) {{
          if (convChanges[nmid]) {{
            delete convChanges[nmid][conv];
            if (!Object.keys(convChanges[nmid]).length) delete convChanges[nmid];
          }}
          saveConvChanges();
          const inp = tr.querySelector(`.conv-input[data-conv="${{conv}}"]`);
          if (inp && !inp.disabled) inp.value = (row[`base_${{conv}}`] || 0).toFixed(2);
        }}
      }}
      updateRowPlanned(nmid);
      return;
    }}
    const btn = e.target.closest('.pct-btn, .drr-btn, .conv-btn');
    if (!btn) return;
    const tr = btn.closest('tr');
    const nmid = tr.dataset.nmid;
    if (btn.classList.contains('pct-btn')) {{
      const input = tr.querySelector('.pct-input');
      const newVal = Math.max(-90, Math.min(300, (Number(input.value) || 0) + Number(btn.dataset.d)));
      input.value = newVal;
      priceChanges[nmid] = newVal;
      savePriceChanges();
    }} else if (btn.classList.contains('drr-btn')) {{
      const input = tr.querySelector('.drr-input');
      const newVal = Math.max(0, Math.min(100, Math.round(((Number(input.value) || 0) + Number(btn.dataset.d)) * 10) / 10));
      input.value = newVal.toFixed(1);
      drrChanges[nmid] = newVal;
      saveDrrChanges();
    }} else if (btn.classList.contains('conv-btn')) {{
      const conv = btn.dataset.conv;
      const input = tr.querySelector(`.conv-input[data-conv="${{conv}}"]`);
      const newVal = Math.max(0, Math.min(100, Math.round(((Number(input.value) || 0) + Number(btn.dataset.d)) * 10) / 10));
      input.value = newVal;
      if (!convChanges[nmid]) convChanges[nmid] = {{}};
      convChanges[nmid][conv] = newVal;
      saveConvChanges();
    }}
    updateRowPlanned(nmid);
  }});
}}

// bulk bar actions
const bulkType = document.getElementById('bulk-type');
document.getElementById('bulk-apply').addEventListener('click', () => {{
  const delta = Number(bulkPct.value) || 0;
  const type = bulkType?.value || 'price';
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const ck = tr.querySelector('input.row-ck');
    if (!ck?.checked) return;
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (type === 'price') {{
      const cur = priceChanges[nmid] ?? 0;
      const nv = Math.max(-90, Math.min(300, Math.round((cur + delta) * 10) / 10));
      priceChanges[nmid] = nv;
      const inp = tr.querySelector('.pct-input');
      if (inp) inp.value = nv;
    }} else if (type === 'drr') {{
      const cur = drrChanges[nmid] != null ? drrChanges[nmid] : (row?.base_drr_pct || 0);
      const nv = Math.max(0, Math.min(100, Math.round((cur + delta) * 10) / 10));
      drrChanges[nmid] = nv;
      const inp = tr.querySelector('.drr-input');
      if (inp) inp.value = nv.toFixed(1);
    }} else {{
      if (!convChanges[nmid]) convChanges[nmid] = {{}};
      const cur = convChanges[nmid][type] > 0 ? convChanges[nmid][type] : (row?.[`base_${{type}}`] || 0);
      const nv = Math.max(0, Math.min(100, Math.round((cur + delta) * 100) / 100));
      convChanges[nmid][type] = nv;
      const inp = tr.querySelector(`.conv-input[data-conv="${{type}}"]`);
      if (inp) inp.value = nv.toFixed(2);
    }}
    updateRowPlanned(nmid);
  }});
  savePriceChanges(); saveDrrChanges(); saveConvChanges();
}});
document.getElementById('bulk-reset').addEventListener('click', () => {{
  const type = bulkType?.value || 'price';
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const ck = tr.querySelector('input.row-ck');
    if (!ck?.checked) return;
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (type === 'price') {{
      priceChanges[nmid] = 0;
      const inp = tr.querySelector('.pct-input'); if (inp) inp.value = 0;
    }} else if (type === 'drr') {{
      delete drrChanges[nmid];
      const inp = tr.querySelector('.drr-input');
      if (inp) inp.value = row?.base_drr_pct || 0;
    }} else {{
      if (convChanges[nmid]) delete convChanges[nmid][type];
      const inp = tr.querySelector(`.conv-input[data-conv="${{type}}"]`);
      if (inp) inp.value = row?.[`base_${{type}}`] || 0;
    }}
    updateRowPlanned(nmid);
  }});
  savePriceChanges(); saveDrrChanges(); saveConvChanges();
}});
document.getElementById('no-oos-btn').addEventListener('click', () => {{
  const selected = new Set([...tableEl.querySelectorAll('input.row-ck:checked')].map(ck => ck.closest('tr')?.dataset.nmid).filter(Boolean));
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    if (selected.size > 0 && !selected.has(nmid)) return;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const pct = calcNoOOS(row);
    if (pct === null || pct === 0) return;
    priceChanges[nmid] = pct;
    const input = tr.querySelector('.pct-input');
    if (input) input.value = pct;
    updateRowPlanned(nmid);
  }});
  savePriceChanges();
}});
document.getElementById('optimize-btn').addEventListener('click', () => {{
  const btn = document.getElementById('optimize-btn');
  btn.disabled = true;
  const selected = new Set([...tableEl.querySelectorAll('input.row-ck:checked')]
    .map(ck => ck.closest('tr')?.dataset.nmid).filter(Boolean));
  const rows = [...tableEl.querySelectorAll('tr[data-nmid]')]
    .map(tr => tr.dataset.nmid)
    .filter(nmid => selected.size === 0 || selected.has(nmid))
    .filter(nmid => rowDataByNmid[nmid]);
  const total = rows.length;
  let idx = 0;
  const CHUNK = 5;
  function processChunk() {{
    const end = Math.min(idx + CHUNK, total);
    for (; idx < end; idx++) {{
      const nmid = rows[idx];
      const row = rowDataByNmid[nmid];
      const {{ pricePct, drrPct }} = findOptimalRevenue(row);
      priceChanges[nmid] = pricePct;
      drrChanges[nmid] = drrPct;
      const tr = tableEl.querySelector(`tr[data-nmid="${{nmid}}"]`);
      if (tr) {{
        const pInp = tr.querySelector('.pct-input');
        if (pInp) pInp.value = pricePct;
        const dInp = tr.querySelector('.drr-input');
        if (dInp) dInp.value = drrPct.toFixed(1);
        updateRowPlanned(nmid);
      }}
    }}
    btn.textContent = `⏳ ${{idx}}/${{total}}…`;
    if (idx < total) {{
      setTimeout(processChunk, 0);
    }} else {{
      savePriceChanges();
      saveDrrChanges();
      const sname = 'Макс. выручка';
      const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
      all[sname] = {{
        price: Object.assign({{}}, priceChanges),
        drr:   Object.assign({{}}, drrChanges),
        conv:  JSON.parse(JSON.stringify(convChanges)),
      }};
      localStorage.setItem(scenariosKey, JSON.stringify(all));
      loadScenarioList();
      document.getElementById('scenario-select').value = sname;
      btn.disabled = false; btn.textContent = '⚡ Макс. выручку';
    }}
  }}
  setTimeout(processChunk, 0);
}});
document.getElementById('reset-prices').addEventListener('click', () => {{
  priceChanges = {{}};
  savePriceChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const input = tr.querySelector('.pct-input');
    if (input) input.value = 0;
    updateRowPlanned(nmid);
  }});
}});
document.getElementById('reset-drr').addEventListener('click', () => {{
  drrChanges = {{}};
  saveDrrChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const drrInp = tr.querySelector('.drr-input');
    if (drrInp) drrInp.value = (row.base_drr_pct || 0).toFixed(1);
    updateRowPlanned(nmid, true);
  }});
  recalcSummary();
}});
document.getElementById('reset-conv').addEventListener('click', () => {{
  convChanges = {{}};
  saveConvChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    ['ctr','cr1','cr2'].forEach(k => {{
      const inp = tr.querySelector(`.conv-input[data-conv="${{k}}"]`);
      if (inp && !inp.disabled) inp.value = (row[`base_${{k}}`] || 0).toFixed(2);
    }});
    updateRowPlanned(nmid, true);
  }});
  recalcSummary();
}});
document.getElementById('bulk-close').addEventListener('click', () => {{
  tableEl.querySelectorAll('input.row-ck').forEach(ck => {{ ck.checked = false; }});
  const ckAll = document.getElementById('ck-all');
  if (ckAll) ckAll.checked = false;
  updateBulkBar();
}});

async function loadData() {{
  saveState();
  meta.textContent = 'Считаю...';
  priceChanges = {{}};
  const qs = new URLSearchParams({{
    forecast_to: fields.forecastTo.value,
    return_delay_days: fields.returnDelay.value || '8',
    transit_days: fields.transitDays.value || '3',
    subject: fields.subject.value,
    labels: [...(fields.strategy?.selectedOptions||[])].map(o=>o.value).join(","),
    revenue_category: fields.revenueCategory.value,
    price_change_pct: '0',
  }});
  const resp = await fetch(`/api/ozon/planning?${{qs.toString()}}`);
  const data = await resp.json();
  if (!resp.ok) {{
    meta.textContent = data.error || `Ошибка ${{resp.status}}`;
    summaryEl.innerHTML = '';
    tableEl.innerHTML = '';
    return;
  }}
  fillSelect(fields.subject, data.options?.subjects || [], 'Все предметы');
  fillMultiSelect(fields.strategy, data.options?.labels || []);
  fillSelect(fields.revenueCategory, data.options?.revenue_categories || [], 'Все категории');
  planningParams = {{
    forecast_days: data.forecast_days || 0,
    return_delay_days: data.return_delay_days || 8,
    transit_days: data.transit_days ?? 3,
    forecast_from: data.forecast_from || '',
  }};
  renderSummary(data.summary || {{}});
  renderTable(data.rows || []);
  // fill economics + WB stock warning for all rows — skip per-row summary recalc, do one pass at the end
  Object.keys(rowDataByNmid).forEach(nmid => updateRowPlanned(nmid, true));
  recalcSummary();
  applyWeeksFilter();
  applyColVisibility();
  meta.textContent = `База скорости: ${{data.baseline_from}}..${{data.baseline_to}} | % выкупа WB за этот период | Прогноз: ${{data.forecast_from}}..${{data.forecast_to}} | Дней: ${{data.forecast_days}} | SKU: ${{(data.rows || []).length}}`;
}}

let timer = null;
function scheduleLoad() {{ clearTimeout(timer); timer = setTimeout(loadData, 250); }}
for (const el of Object.values(fields)) {{
  if (el === fields.weeksFilter) {{
    el.addEventListener('input', applyWeeksFilter);
  }} else {{
    el.addEventListener('change', loadData);
    el.addEventListener('input', scheduleLoad);
  }}
}}
document.getElementById('load').addEventListener('click', loadData);

// tax/admin: recalc all rows on change
['tax-pct', 'admin-pct'].forEach(id => {{
  document.getElementById(id).addEventListener('input', () => {{
    Object.keys(rowDataByNmid).forEach(nmid => updateRowPlanned(nmid, true));
    recalcSummary();
  }});
}});

// ——— Column visibility (gear) ———
const COL_SCHEMA = [
  {{ group:'остатки', label:'Остатки', cols:[
    {{ key:'c-stock',     label:'WB, шт' }},
    {{ key:'c-wh-stock',  label:'Склад, шт' }},
    {{ key:'c-stock-val', label:'Сумма, ₽' }},
  ]}},
  {{ group:'базовый', label:'Базовый период', cols:[
    {{ key:'c-base-price',  label:'Цена, ₽' }},
    {{ key:'c-base-spp',    label:'СПП, %' }},
    {{ key:'c-base-client', label:'Кл. цена, ₽' }},
    {{ key:'c-base-speed',  label:'Скорость, шт' }},
  ]}},
  {{ group:'текущее', label:'Текущее', cols:[
    {{ key:'c-cur-price',  label:'Цена, ₽' }},
    {{ key:'c-cur-spp',    label:'СПП, %' }},
    {{ key:'c-cur-client', label:'Кл. цена, ₽' }},
    {{ key:'c-cur-elast',  label:'Эластичность' }},
    {{ key:'c-cur-speed',  label:'Скорость, шт' }},
  ]}},
  {{ group:'параметры', label:'Параметры', cols:[
    {{ key:'c-plan-price', label:'Цена, %' }},
    {{ key:'c-plan-drr',   label:'ДРР, %' }},
    {{ key:'c-plan-drr-e', label:'Эл. рекл.' }},
    {{ key:'c-plan-ctr',   label:'CTR, %' }},
    {{ key:'c-plan-cr1',   label:'CR корзина, %' }},
    {{ key:'c-plan-cr2',   label:'CR заказ, %' }},
  ]}},
  {{ group:'результат', label:'Результат', cols:[
    {{ key:'c-res-price',   label:'Пл. цена, ₽' }},
    {{ key:'c-res-client',  label:'Пл. кл. цена, ₽' }},
    {{ key:'c-res-speed',   label:'Скорость, шт/д' }},
    {{ key:'c-res-buyout',  label:'% выкупа' }},
    {{ key:'c-res-orders',  label:'Заказы, шт' }},
    {{ key:'c-res-buyouts', label:'Выкупы, шт' }},
    {{ key:'c-res-revenue', label:'Выручка, ₽' }},
  ]}},
  {{ group:'упущено', label:'Упущенные продажи', cols:[
    {{ key:'c-res-oos',  label:'Дата стопа' }},
    {{ key:'c-lost-qty', label:'Упущ., шт' }},
    {{ key:'c-lost-rub', label:'Упущ., ₽' }},
  ]}},
  {{ group:'экономика', label:'Экономика', cols:[
    {{ key:'c-econ-comm-pct',   label:'Комиссия, %' }},
    {{ key:'c-econ-comm-rub',   label:'Комиссия, ₽' }},
    {{ key:'c-econ-acq-rub',    label:'Эквайринг, ₽' }},
    {{ key:'c-econ-drr-rub',    label:'Реклама, ₽' }},
    {{ key:'c-econ-cogs',       label:'Себестоимость, ₽' }},
    {{ key:'c-econ-margin-pct', label:'Маржа, %' }},
    {{ key:'c-econ-margin-rub', label:'Маржа, ₽' }},
    {{ key:'c-econ-admin-rub',  label:'Адм. расходы, ₽' }},
    {{ key:'c-econ-tax-rub',    label:'Налоги, ₽' }},
    {{ key:'c-econ-profit-pct', label:'Прибыль, %' }},
    {{ key:'c-econ-profit-rub', label:'Прибыль, ₽' }},
  ]}},
];

const COL_VIS_KEY = 'wb.planning.col-vis';
let colVisibility = {{}};
try {{ colVisibility = JSON.parse(localStorage.getItem(COL_VIS_KEY) || '{{}}'); }} catch(e) {{}}

const _colVisStyle = document.createElement('style');
_colVisStyle.id = 'col-vis-style';
document.head.appendChild(_colVisStyle);

function applyColVisibility() {{
  const rules = [];
  COL_SCHEMA.forEach(grp => {{
    const allHidden = grp.cols.every(c => colVisibility[c.key] === false);
    if (allHidden) {{
      rules.push(`[data-col-group="${{grp.group}}"] {{ display:none !important; }}`);
    }} else {{
      grp.cols.forEach(c => {{
        if (colVisibility[c.key] === false) {{
          rules.push(`[data-col-key="${{c.key}}"] {{ display:none !important; }}`);
        }}
      }});
    }}
  }});
  document.getElementById('col-vis-style').textContent = rules.join('\\n');
}}

function _isGroupChecked(grp) {{ return grp.cols.some(c => colVisibility[c.key] !== false); }}
function _isGroupIndeterminate(grp) {{
  const vis = grp.cols.filter(c => colVisibility[c.key] !== false).length;
  return vis > 0 && vis < grp.cols.length;
}}

const colSettingsBtn   = document.getElementById('col-settings-btn');
const colSettingsPanel = document.getElementById('col-settings-panel');

function _buildColPanel() {{
  colSettingsPanel.innerHTML = `
    <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#94a3b8;margin-bottom:8px">Столбцы таблицы</div>
    ${{COL_SCHEMA.map(grp => `
      <div style="margin-bottom:6px">
        <label style="display:flex;align-items:center;gap:6px;cursor:pointer;font-size:12px;font-weight:700;color:#334155;padding:2px 0">
          <input type="checkbox" data-grp="${{grp.group}}"
            ${{_isGroupChecked(grp) ? 'checked' : ''}}
            style="accent-color:#7e22ce"> ${{grp.label}}
        </label>
        <div style="padding-left:18px">
          ${{grp.cols.map(c => `
            <label style="display:flex;align-items:center;gap:5px;cursor:pointer;font-size:12px;color:#475569;padding:1px 0">
              <input type="checkbox" data-col="${{c.key}}"
                ${{colVisibility[c.key] !== false ? 'checked' : ''}}
                style="accent-color:#7e22ce"> ${{c.label}}
            </label>`).join('')}}
        </div>
      </div>`).join('')}}`;

  // group toggle
  colSettingsPanel.querySelectorAll('input[data-grp]').forEach(cb => {{
    const grp = COL_SCHEMA.find(g => g.group === cb.dataset.grp);
    if (!grp) return;
    // set indeterminate state
    if (_isGroupIndeterminate(grp)) cb.indeterminate = true;
    cb.addEventListener('change', () => {{
      grp.cols.forEach(c => {{
        colVisibility[c.key] = cb.checked;
        const colCb = colSettingsPanel.querySelector(`input[data-col="${{c.key}}"]`);
        if (colCb) colCb.checked = cb.checked;
      }});
      localStorage.setItem(COL_VIS_KEY, JSON.stringify(colVisibility));
      applyColVisibility();
    }});
  }});

  // individual column toggle
  colSettingsPanel.querySelectorAll('input[data-col]').forEach(cb => {{
    cb.addEventListener('change', () => {{
      colVisibility[cb.dataset.col] = cb.checked;
      localStorage.setItem(COL_VIS_KEY, JSON.stringify(colVisibility));
      applyColVisibility();
      // update parent group checkbox
      const grp = COL_SCHEMA.find(g => g.cols.some(c => c.key === cb.dataset.col));
      if (grp) {{
        const grpCb = colSettingsPanel.querySelector(`input[data-grp="${{grp.group}}"]`);
        if (grpCb) {{
          const checked = _isGroupChecked(grp);
          const indet = _isGroupIndeterminate(grp);
          grpCb.checked = checked;
          grpCb.indeterminate = indet;
        }}
      }}
    }});
  }});
}}

colSettingsBtn.addEventListener('click', e => {{
  e.stopPropagation();
  if (colSettingsPanel.style.display === 'none') {{
    _buildColPanel();
    colSettingsPanel.style.display = 'block';
  }} else {{
    colSettingsPanel.style.display = 'none';
  }}
}});
document.addEventListener('click', () => {{ colSettingsPanel.style.display = 'none'; }});

// ——— Sorting ———
let sortState = {{ key: null, asc: true }};
tableEl.addEventListener('click', e => {{
  const th = e.target.closest('th[data-sort-key]');
  if (!th) return;
  const key = th.dataset.sortKey;
  sortState.asc = sortState.key === key ? !sortState.asc : true;
  sortState.key = key;
  tableEl.querySelectorAll('th[data-sort-key]').forEach(t => delete t.dataset.sortDir);
  th.dataset.sortDir = sortState.asc ? 'asc' : 'desc';
  const tbody = tableEl.querySelector('tbody');
  if (!tbody) return;
  const tRows = [...tbody.querySelectorAll('tr[data-nmid]')];
  const textKeys = new Set(['article','subject','strategy','revenue_category']);
  tRows.sort((a, b) => {{
    const nA = a.dataset.nmid, nB = b.dataset.nmid;
    let va, vb;
    if (key === 'margin_pct') {{
      va = parseFloat(a.querySelector('.cell-econ-margin')?.dataset.margin || 0);
      vb = parseFloat(b.querySelector('.cell-econ-margin')?.dataset.margin || 0);
    }} else if (key === 'profit_pct') {{
      va = parseFloat(a.querySelector('.cell-econ-profit-pct')?.dataset.profitPct || 0);
      vb = parseFloat(b.querySelector('.cell-econ-profit-pct')?.dataset.profitPct || 0);
    }} else if (key === 'net_profit') {{
      va = parseFloat(a.querySelector('.cell-econ-profit')?.dataset.profit || 0);
      vb = parseFloat(b.querySelector('.cell-econ-profit')?.dataset.profit || 0);
    }} else if (key === 'margin_rub') {{
      va = parseFloat(a.querySelector('.cell-econ-margin-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
      vb = parseFloat(b.querySelector('.cell-econ-margin-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
    }} else if (key === 'comm_rub') {{
      va = parseFloat(a.querySelector('.cell-econ-comm-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
      vb = parseFloat(b.querySelector('.cell-econ-comm-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
    }} else if (key === 'drr_rub') {{
      va = parseFloat(a.querySelector('.cell-econ-drr-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
      vb = parseFloat(b.querySelector('.cell-econ-drr-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
    }} else if (key === 'stock_value') {{
      va = (rowDataByNmid[a.dataset.nmid]?.stock || 0) * (rowDataByNmid[a.dataset.nmid]?.client_price || 0);
      vb = (rowDataByNmid[b.dataset.nmid]?.stock || 0) * (rowDataByNmid[b.dataset.nmid]?.client_price || 0);
    }} else if (key === 'plannedSpeed') {{
      va = rowForecastCache[nA]?.forecastOrders / Math.max(1, planningParams.forecast_days) || 0;
      vb = rowForecastCache[nB]?.forecastOrders / Math.max(1, planningParams.forecast_days) || 0;
    }} else if (textKeys.has(key)) {{
      va = (rowDataByNmid[nA]?.[key] || '').toLowerCase();
      vb = (rowDataByNmid[nB]?.[key] || '').toLowerCase();
      return sortState.asc ? va.localeCompare(vb,'ru') : vb.localeCompare(va,'ru');
    }} else {{
      va = rowDataByNmid[nA]?.[key] ?? (rowForecastCache[nA]?.[key] ?? 0);
      vb = rowDataByNmid[nB]?.[key] ?? (rowForecastCache[nB]?.[key] ?? 0);
    }}
    return sortState.asc ? va - vb : vb - va;
  }});
  tRows.forEach(r => tbody.appendChild(r));
}});

restoreState();
loadData();
</script>
<div class="methodology">
  <h2>Методология расчётов</h2>

  <h3>Скорость продаж (базовая)</h3>
  <p>Среднедневное количество заказов за последние 7 дней базового периода (последний день с данными исключается — он может быть неполным).<br>
  <span class="formula">скорость = заказы_за_7_дней / 7</span><br>
  <em>Пример: за 7 дней было 350 заказов → базовая скорость = 50 шт/день.</em></p>

  <h3>Эластичность цены</h3>
  <p>Показывает, насколько меняется спрос при изменении цены. Рассчитывается методом лог-лог МНК по 60 дням истории заказов:<br>
  <span class="formula">ln(заказы) = E × ln(цена) + const</span></p>
  <p>Условия расчёта: ≥ 7 дней с заказами, разброс цен ≥ 5%, два ценовых уровня с разрывом ≥ 3%, R² ≥ 0.30.<br>
  Диапазон значений: от −3.0 до 0.0 (отрицательная — при росте цены спрос падает).</p>
  <p><strong>Как читать в таблице:</strong></p>
  <ul>
    <li><strong>−1.0</strong> (по умолчанию, нет данных) — единичная эластичность: +10% цены → −10% скорости.<br>
    <em>50 шт/день × (1 + (−1.0) × 0.10) = 45 шт/день при цене +10%</em></li>
    <li><strong>−0.5</strong> — слабая реакция: +20% цены → −10% скорости.<br>
    <em>50 × (1 + (−0.5) × 0.20) = 45 шт/день</em></li>
    <li><strong>−2.0</strong> — сильная реакция: +10% цены → −20% скорости.<br>
    <em>50 × (1 + (−2.0) × 0.10) = 40 шт/день</em></li>
    <li><strong>0.0</strong> — данные есть, но цена статистически не влияет на спрос (R² &lt; 0.30 или цена не менялась).</li>
  </ul>
  <p>Если по SKU нет надёжных данных — берётся среднее по предмету → стратегии → категории. Если нигде нет — используется <strong>−1.0</strong> (консервативный дефолт).</p>

  <h3>ДРР и эластичность рекламы</h3>
  <p><strong>Базовый ДРР%</strong> = фактические расходы на рекламу / фактическая выручка с выкупов × 100%, за последние 14 дней (последние 2 дня исключаются — выкупы там ещё не зафиксированы).<br>
  <span class="formula">ДРР = Σ расходы_реклама_14д / Σ buyoutSum_14д × 100%</span><br>
  <em>Пример: реклама 70 000 ₽, выручка с выкупов 500 000 ₽ → ДРР = 14%.</em></p>
  <p><strong>Эластичность рекламы</strong> показывает, как расходы на рекламу влияют на скорость продаж. Рассчитывается двумерным МНК вместе с ценовой эластичностью:<br>
  <span class="formula">ln(заказы) = E_цена × ln(цена) + E_реклама × ln(1 + расход/день) + const</span><br>
  Диапазон E_реклама: от 0.0 до 1.5.</p>
  <p><strong>Как работает в планировании:</strong><br>
  <span class="formula">ads_factor = (план_ДРР% / баз_ДРР%) ^ E_реклама</span><br>
  (минимальный порог ratio = 0.10, чтобы при ДРР = 0 скорость не падала в ноль — органика сохраняется)</p>
  <ul>
    <li><em>ДРР 14% → 28% (×2), E_реклама = 0.5: ads_factor = 2.0^0.5 = 1.41 → +41% к скорости</em></li>
    <li><em>ДРР 14% → 7% (÷2), E_реклама = 0.5: ads_factor = 0.5^0.5 = 0.71 → −29% к скорости</em></li>
    <li><em>ДРР 14% → 0% (выкл.), E_реклама = 0.23: ads_factor = 0.10^0.23 = 0.60 → −40% к скорости</em></li>
  </ul>

  <h3>Конверсии воронки (CTR, CR1, CR2)</h3>
  <p>Данные из ручной выгрузки WB (funnel_impressions_upload) за базовые 7 дней:<br>
  <span class="formula">CTR = переходы / показы × 100%</span> &nbsp;
  <span class="formula">CR1 = корзины / переходы × 100%</span> &nbsp;
  <span class="formula">CR2 = заказы / корзины × 100%</span></p>
  <p>Изменение конверсий — прямой мультипликатор (улучшение карточки, фото, описания):<br>
  <span class="formula">conv_factor = (CTR_план/CTR_баз) × (CR1_план/CR1_баз) × (CR2_план/CR2_баз)</span><br>
  <em>Пример: CTR +20%, CR1 без изм., CR2 +10% → conv_factor = 1.20 × 1.0 × 1.10 = 1.32 → +32% к скорости.</em></p>

  <h3>Итоговая плановая скорость</h3>
  <p>Три рычага перемножаются:<br>
  <span class="formula">план_скорость = базовая_скорость × price_factor × ads_factor × conv_factor</span></p>
  <ul>
    <li><strong>price_factor</strong> = <span class="formula">max(0, 1 + E_цена × Δцена_дол.)</span> — линейная аппроксимация эластичности</li>
    <li><strong>ads_factor</strong> = <span class="formula">(план_ДРР / баз_ДРР) ^ E_реклама</span> — степенной закон</li>
    <li><strong>conv_factor</strong> — прямой мультипликатор конверсий</li>
  </ul>
  <p><em>Пример: скорость 50 шт/день, цена +10% (E=−1.0), ДРР ×2 (E_рекл=0.5), CTR +20%:<br>
  price_factor = 1 + (−1.0) × 0.10 = 0.90 &nbsp;|&nbsp; ads_factor = 2.0^0.5 = 1.41 &nbsp;|&nbsp; conv_factor = 1.20<br>
  план_скорость = 50 × 0.90 × 1.41 × 1.20 ≈ 76 шт/день</em></p>
  <p style="color:#64748b;font-size:.75rem;margin-top:8px">Все изменения цены, ДРР и конверсий хранятся в localStorage браузера — прогноз пересчитывается мгновенно на стороне клиента без запросов к серверу.</p>
</div>
</body>
</html>
"""

AD_BIDDER_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>WB Биддер</title>
  <style>
    :root {{ --bg:#f8fafc; --card:#fff; --ink:#111827; --muted:#6b7280; --accent:#0f766e; --wb:#cb11ab; --line:#e5e7eb; --ok:#047857; --bad:#b91c1c; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:"Segoe UI","Trebuchet MS",sans-serif; color:var(--ink); background:linear-gradient(180deg,#e0f2fe,transparent 240px),var(--bg); }}
    .wrap {{ width:calc(100% - 32px); max-width:3000px; margin:0 auto; padding:10px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    label {{ display:block; color:var(--muted); font-size:.68rem; margin-bottom:2px; font-weight:700; }}
    input,select {{ padding:6px 7px; border:1px solid #cbd5e1; border-radius:7px; font-size:12px; background:#fff; min-width:0; }}
    button {{ padding:7px 10px; border:0; border-radius:7px; background:var(--accent); color:#fff; font-weight:800; cursor:pointer; font-size:12px; }}
    button.ghost {{ background:#e6f4f1; color:var(--accent); }}
    button.gray {{ background:#f1f5f9; color:#334155; }}
    button:disabled {{ opacity:.55; cursor:not-allowed; }}
    .workspace {{ --articles-width:332px; display:grid; grid-template-columns:var(--articles-width) minmax(0,1fr); gap:12px; align-items:start; margin-top:10px; }}
    .workspace.articles-collapsed {{ --articles-width:44px; }}
    .main-pane {{ display:flex; flex-direction:column; gap:12px; min-width:0; }}
    .panel {{ background:#fff; border:1px solid var(--line); border-radius:12px; padding:12px; }}
    .toolbar {{ display:flex; align-items:end; gap:10px; flex-wrap:wrap; }}
    .toolbar h2 {{ margin:0 auto 2px 0; font-size:.95rem; }}
    .toolbar input {{ width:116px; }}
    .toolbar select {{ width:92px; }}
    .quick-dates {{ display:flex; align-items:center; gap:4px; flex-wrap:wrap; padding-bottom:1px; }}
    .quick-dates button {{ padding:6px 8px; border-radius:7px; background:#f1f5f9; color:#334155; font-size:.72rem; }}
    .quick-dates button:hover {{ background:#e6f4f1; color:var(--accent); }}
    .status {{ color:var(--muted); font-size:.78rem; min-height:1.2em; }}
    .status.ok {{ color:var(--ok); }}
    .status.err {{ color:var(--bad); }}
    .articles {{ position:sticky; top:8px; background:#fff; border:1px solid var(--line); border-radius:12px; overflow:hidden; max-height:calc(100vh - 18px); display:flex; flex-direction:column; min-width:0; transition:width .18s ease,box-shadow .18s ease,border-color .18s ease; }}
    .articles-inner {{ display:flex; flex-direction:column; min-height:0; height:100%; background:#fff; }}
    .articles-head {{ padding:7px 8px; border-bottom:1px solid #eef2f7; display:flex; justify-content:space-between; gap:8px; align-items:center; }}
    .articles-title {{ font-weight:800; font-size:.78rem; }}
    .articles-actions {{ display:flex; gap:8px; }}
    .articles-actions button {{ padding:4px 6px; border-radius:6px; font-size:.68rem; background:#e6f4f1; color:var(--accent); }}
    .articles-mini {{ display:none; align-items:center; justify-content:center; gap:6px; height:100%; min-height:260px; padding:8px 0; writing-mode:vertical-rl; transform:rotate(180deg); color:var(--accent); font-weight:800; font-size:.72rem; letter-spacing:.08em; background:linear-gradient(180deg,#f0fdfa,#ecfeff); }}
    .articles-mini button {{ writing-mode:horizontal-tb; transform:rotate(180deg); margin:0; width:24px; height:24px; padding:0; border-radius:999px; background:var(--accent); color:#fff; font-size:.9rem; line-height:1; }}
    .subject-filter {{ padding:6px 8px; border-bottom:1px solid #eef2f7; }}
    .subject-filter input,.subject-filter select {{ width:100%; padding:5px 6px; font-size:.72rem; }}
    .article-list {{ overflow:auto; user-select:none; }}
    .article-table {{ width:100%; min-width:0; border-collapse:collapse; table-layout:fixed; font-size:10px; }}
    .article-table th,.article-table td {{ border-bottom:1px solid #eef2f7; padding:4px 3px; line-height:1.15; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    .article-table th {{ position:sticky; top:0; z-index:1; background:#f8fafc; color:#475569; font-size:.62rem; text-align:right; cursor:pointer; }}
    .article-table th:first-child,.article-table td:first-child {{ text-align:left; width:38%; }}
    .article-table th:nth-child(2),.article-table td:nth-child(2) {{ text-align:right; width:12%; }}
    .article-table th:nth-child(3),.article-table td:nth-child(3) {{ text-align:right; width:12%; }}
    .article-table th:nth-child(4),.article-table td:nth-child(4) {{ text-align:right; width:12%; }}
    .article-table th:nth-child(5),.article-table td:nth-child(5) {{ text-align:right; width:26%; }}
    .article-table tr {{ cursor:default; }}
    .article-table tbody tr:hover {{ background:#f0fdfa; }}
    .article-table tbody tr.selected {{ background:#ccfbf1; font-weight:800; color:#115e59; }}
    .article-link,.campaign-link {{ color:var(--accent); text-decoration:none; font-weight:900; margin-left:4px; }}
    .campaigns-list-link {{ color:#7c3aed; text-decoration:none; font-weight:900; margin-left:4px; }}
    .article-link:hover,.campaign-link:hover,.campaigns-list-link:hover {{ text-decoration:underline; }}
    .article-copy {{ background:none; border:none; cursor:pointer; padding:0 1px; margin-left:3px; font-size:.8em; color:#94a3b8; opacity:.75; line-height:1; vertical-align:middle; }}
    .article-copy:hover {{ color:var(--accent); opacity:1; }}
    .article-copy.copied {{ color:#16a34a; opacity:1; }}
    .tbl {{ background:#fff; border:1px solid var(--line); border-radius:12px; overflow:auto; -webkit-overflow-scrolling:touch; }}
    .campaign-table {{ max-height:46vh; min-height:180px; }}
    .tbl table {{ width:max-content; min-width:100%; border-collapse:collapse; font-size:11px; }}
    .tbl th,.tbl td {{ border-bottom:1px solid #eef2f7; padding:6px 7px; text-align:right; white-space:nowrap; vertical-align:middle; }}
    .tbl th {{ position:sticky; top:0; z-index:2; background:#f8fafc; color:#475569; font-weight:800; }}
    .tbl th:first-child,.tbl td:first-child {{ position:sticky; left:0; z-index:1; text-align:left; background:#fff; min-width:260px; max-width:340px; }}
    .tbl th:first-child {{ z-index:3; background:#f8fafc; }}
    .campaign-name {{ font-weight:800; max-width:320px; overflow:hidden; text-overflow:ellipsis; }}
    .meta {{ color:var(--muted); font-size:.76rem; line-height:1.35; }}
    .pill {{ display:inline-flex; align-items:center; border-radius:999px; padding:3px 7px; font-size:.72rem; font-weight:800; background:#f1f5f9; color:#334155; }}
    .pill.on {{ background:#dcfce7; color:#166534; }}
    .pill.off {{ background:#fee2e2; color:#991b1b; }}
    .row-controls {{ display:flex; gap:8px; align-items:center; justify-content:flex-start; }}
    .control-box {{ display:inline-flex; align-items:center; gap:5px; padding:4px 6px; border:1px solid #e2e8f0; border-radius:8px; background:#f8fafc; }}
    .control-box span {{ color:#64748b; font-size:.7rem; font-weight:800; }}
    .control-box.default-budget {{ border-color:#99f6e4; background:#ecfeff; }}
    .calc-chip {{ display:inline-flex; align-items:center; justify-content:center; min-width:28px; padding:2px 5px; border-radius:999px; background:#ccfbf1; color:#0f766e; font-size:.64rem; font-weight:900; }}
    .budget {{ width:72px; text-align:right; padding:5px 6px; }}
    .pause-input {{ width:74px; padding:5px 6px; }}
    .swipe-toggle {{ display:inline-flex; align-items:center; gap:6px; padding:0; border:0; background:transparent; color:#334155; font-size:.72rem; font-weight:800; white-space:nowrap; cursor:grab; touch-action:pan-y; user-select:none; }}
    .swipe-toggle:active {{ cursor:grabbing; }}
    .swipe-toggle:focus-visible {{ outline:2px solid #99f6e4; outline-offset:3px; border-radius:999px; }}
    .swipe-track {{ position:relative; width:34px; height:18px; border-radius:999px; background:#e2e8f0; box-shadow:inset 0 0 0 1px #cbd5e1; transition:background .16s ease,box-shadow .16s ease; flex:0 0 auto; }}
    .swipe-thumb {{ position:absolute; left:2px; top:2px; width:14px; height:14px; border-radius:50%; background:#fff; box-shadow:0 1px 3px rgba(15,23,42,.22); transition:transform .16s ease; }}
    .swipe-toggle.is-on .swipe-track {{ background:#0f766e; box-shadow:inset 0 0 0 1px #0f766e; }}
    .swipe-toggle.is-on .swipe-thumb {{ transform:translateX(16px); }}
    .swipe-toggle.is-saving {{ opacity:.68; }}
    .swipe-label {{ color:#334155; }}
    .type-chip {{ display:inline-flex; align-items:center; padding:2px 6px; border-radius:6px; background:#f1f5f9; color:#334155; font-size:.68rem; font-weight:800; }}
    .nm-count {{ display:inline-flex; align-items:center; padding:2px 6px; border-radius:999px; background:#ecfdf5; color:#047857; font-size:.68rem; font-weight:900; }}
    .nm-count.multi {{ background:#fee2e2; color:#991b1b; }}
    .log-panel {{ padding:0; overflow:hidden; }}
    .log-head {{ display:flex; align-items:center; gap:8px; padding:9px 10px; border-bottom:1px solid #eef2f7; }}
    .log-head h2 {{ margin:0; font-size:.86rem; flex:1; }}
    .log-list {{ max-height:34vh; min-height:150px; overflow:auto; background:#fff; }}
    .log-row {{ display:grid; grid-template-columns:124px 76px minmax(150px,1fr); gap:8px; padding:7px 10px; border-bottom:1px solid #eef2f7; align-items:start; }}
    .log-time {{ color:#64748b; font-size:.72rem; font-variant-numeric:tabular-nums; }}
    .log-kind {{ justify-self:start; border-radius:999px; padding:2px 7px; font-size:.68rem; font-weight:900; background:#f1f5f9; color:#334155; }}
    .log-kind.manual {{ background:#e0f2fe; color:#075985; }}
    .log-kind.poll {{ background:#ecfdf5; color:#047857; }}
    .log-kind.auto {{ background:#fef3c7; color:#92400e; }}
    .log-kind.error {{ background:#fee2e2; color:#991b1b; }}
    .log-title {{ font-size:.78rem; font-weight:800; }}
    .log-details {{ margin-top:2px; color:#64748b; font-size:.72rem; line-height:1.3; white-space:normal; }}
    .num {{ font-variant-numeric:tabular-nums; }}
    .danger {{ color:var(--bad); font-weight:800; }}
    .ok-text {{ color:var(--ok); font-weight:800; }}
    .workspace.articles-collapsed .articles {{ overflow:visible; }}
    .workspace.articles-collapsed .articles .articles-inner {{ display:none; }}
    .workspace.articles-collapsed .articles .articles-mini {{ display:flex; }}
    .workspace.articles-collapsed .articles:hover,.workspace.articles-collapsed .articles:focus-within {{ width:332px; z-index:20; box-shadow:0 18px 40px rgba(15,118,110,.18); border-color:#99f6e4; }}
    .workspace.articles-collapsed .articles:hover .articles-inner,.workspace.articles-collapsed .articles:focus-within .articles-inner {{ display:flex; }}
    .workspace.articles-collapsed .articles:hover .articles-mini,.workspace.articles-collapsed .articles:focus-within .articles-mini {{ display:none; }}
    @media (max-width:1100px) {{
      .workspace {{ grid-template-columns:1fr; }}
      .workspace.articles-collapsed {{ --articles-width:1fr; }}
      .articles {{ order:2; max-height:260px; }}
      .workspace.articles-collapsed .articles {{ width:auto; overflow:hidden; }}
      .workspace.articles-collapsed .articles .articles-inner {{ display:flex; }}
      .workspace.articles-collapsed .articles .articles-mini {{ display:none; }}
      .row-controls {{ flex-wrap:wrap; }}
      .campaign-table {{ max-height:360px; }}
      .log-row {{ grid-template-columns:96px 66px minmax(120px,1fr); gap:6px; }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>WB Биддер</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div class="toolbar">
        <h2 id="campaign-title">Кампании артикула</h2>
        <div>
          <label for="df">Период с</label>
          <input id="df" type="date" value="{date_from}" style="width:130px">
        </div>
        <div>
          <label for="dt">Период по</label>
          <input id="dt" type="date" value="{date_to}" style="width:130px">
        </div>
        <div class="quick-dates" aria-label="Быстрый период">
          <button type="button" data-quick-period="yesterday">Вчера</button>
          <button type="button" data-quick-period="today">Сегодня</button>
          <button type="button" data-quick-period="3">3</button>
          <button type="button" data-quick-period="7">7</button>
          <button type="button" data-quick-period="14">14</button>
        </div>
        <button id="run" class="ghost" type="button">Проверить</button>
        <button id="refresh" class="gray" type="button">Обновить из API</button>
        <div id="status" class="status"></div>
      </div>
    </div>
    <div class="workspace" id="workspace">
      <aside class="articles">
        <div class="articles-inner">
          <div class="articles-head">
            <div class="articles-title">Артикулы</div>
            <div class="articles-actions">
              <button type="button" id="toggle-articles" title="Свернуть список">◂</button>
            </div>
          </div>
          <div class="subject-filter">
            <label for="q">Поиск</label>
            <input id="q" type="text" placeholder="Артикул, название или nmId">
          </div>
          <div class="subject-filter">
            <label for="subject">Предмет</label>
            <select id="subject"><option value="">Все предметы</option></select>
          </div>
          <div class="subject-filter">
            <label for="strategy">Ярлыки</label>
            <select id="strategy" multiple size="4" style="height:72px;" title="Ctrl/Cmd для множественного выбора"></select>
          </div>
          <div class="subject-filter">
            <label for="revenue-category">Категория по выручке</label>
            <select id="revenue-category"><option value="">Все категории</option></select>
          </div>
          <div class="article-list" id="article-list"></div>
        </div>
        <div class="articles-mini">
          <button type="button" id="expand-articles" title="Развернуть список">▸</button>
          <span>Артикулы</span>
        </div>
      </aside>
      <div class="main-pane">
        <div class="tbl campaign-table">
          <table>
            <thead>
              <tr>
                <th>Кампания</th>
                <th>Тип</th>
                <th>Бюджет</th>
                <th>Пауза</th>
                <th>Сегодня</th>
                <th>Статус</th>
                <th>Расход</th>
                <th>Показы</th>
                <th>Клики</th>
                <th>CTR</th>
                <th>CPC</th>
                <th>CPM</th>
                <th>Заказы</th>
                <th>CPO</th>
                <th>ДРР</th>
              </tr>
            </thead>
            <tbody id="campaigns"><tr><td colspan="15" class="meta">Выберите артикул.</td></tr></tbody>
          </table>
        </div>
        <section class="panel log-panel">
          <div class="log-head">
            <h2>Лог рекламного кабинета</h2>
            <button id="reload-log" class="gray" type="button">Обновить</button>
          </div>
          <div class="log-list" id="ad-log"><div class="meta" style="padding:10px;">Загружаю...</div></div>
        </section>
      </div>
    </div>
  </div>
<script>
let selectedArticle = null;
let allArticleItems = [];
let articleItems = [];
let campaigns = [];
let articlesCollapsed = false;
let articleSort = {{ key: 'revenue', dir: 'desc' }};
let articlesLoadedFor = '';
let subjectsLoadedFor = '';
const workspaceEl = document.getElementById('workspace');
const q = document.getElementById('q');
const articleList = document.getElementById('article-list');
const subjectEl = document.getElementById('subject');
const strategyEl = document.getElementById('strategy');
const revenueCategoryEl = document.getElementById('revenue-category');
const campaignsEl = document.getElementById('campaigns');
const statusEl = document.getElementById('status');
const adLogEl = document.getElementById('ad-log');
const SERVER_TODAY = '{today_date}';
function escapeHtml(value) {{
  return String(value ?? '').replaceAll('&','&amp;').replaceAll('<','&lt;').replaceAll('>','&gt;').replaceAll('"','&quot;').replaceAll("'",'&#39;');
}}
function fmt(value, digits=0) {{
  const numeric = Number(value || 0);
  return numeric.toLocaleString('ru-RU', {{ maximumFractionDigits: digits }});
}}
function setStatus(text, cls='') {{
  statusEl.textContent = text;
  statusEl.className = 'status ' + cls;
}}
function isoDateLocal(date) {{
  const year = date.getFullYear();
  const month = String(date.getMonth() + 1).padStart(2, '0');
  const day = String(date.getDate()).padStart(2, '0');
  return `${{year}}-${{month}}-${{day}}`;
}}
function reloadPeriodData() {{
  subjectsLoadedFor = '';
  articlesLoadedFor = '';
  loadArticles(true);
  loadCampaigns();
}}
function applyQuickPeriod(value) {{
  const today = new Date(`${{SERVER_TODAY}}T00:00:00`);
  today.setHours(0, 0, 0, 0);
  let from = new Date(today);
  let to = new Date(today);
  if (value === 'yesterday') {{
    from.setDate(today.getDate() - 1);
    to.setDate(today.getDate() - 1);
  }} else if (value !== 'today') {{
    const days = Number(value || 1);
    to.setDate(today.getDate() - 1);
    from = new Date(to);
    from.setDate(to.getDate() - Math.max(1, days) + 1);
  }}
  document.getElementById('df').value = isoDateLocal(from);
  document.getElementById('dt').value = isoDateLocal(to);
  reloadPeriodData();
}}
function logKindLabel(kind) {{
  return {{manual:'ручное', poll:'опрос', auto:'авто', error:'ошибка'}}[kind] || kind || 'событие';
}}
function renderAdLog(logs) {{
  adLogEl.innerHTML = logs.length ? logs.map(item => {{
    const created = String(item.created_at || '').replace('T', ' ').slice(0, 16);
    const kind = escapeHtml(item.event_type || 'event');
    return `<div class="log-row">
      <div class="log-time">${{escapeHtml(created)}}</div>
      <div><span class="log-kind ${{kind}}">${{escapeHtml(logKindLabel(item.event_type))}}</span></div>
      <div>
        <div class="log-title">${{escapeHtml(item.title || item.advert_id || 'Событие')}}</div>
        <div class="log-details">${{escapeHtml(item.details || '')}}</div>
      </div>
    </div>`;
  }}).join('') : '<div class="meta" style="padding:10px;">Пока нет событий.</div>';
}}
async function loadAdLog() {{
  try {{
    const data = await fetch('/api/ads/logs?limit=200').then(r => r.json());
    renderAdLog(data.logs || []);
  }} catch (err) {{
    adLogEl.innerHTML = '<div class="meta" style="padding:10px;">Не удалось загрузить лог.</div>';
  }}
}}
function syncArticlesPanel() {{
  workspaceEl.classList.toggle('articles-collapsed', articlesCollapsed);
  const toggleBtn = document.getElementById('toggle-articles');
  if (toggleBtn) {{
    toggleBtn.textContent = articlesCollapsed ? '▸' : '◂';
    toggleBtn.title = articlesCollapsed ? 'Развернуть список' : 'Свернуть список';
  }}
}}
function fillMultiSelect(select, values) {{
  const prevVals = new Set([...select.selectedOptions].map(o => o.value));
  select.innerHTML = (values || []).map(value => {{
    const safe = escapeHtml(value);
    const selected = prevVals.has(value) ? ' selected' : '';
    return `<option value="${{safe}}"${{selected}}>${{safe}}</option>`;
  }}).join('');
}}
function renderArticles(items) {{
  const sourceItems = items || allArticleItems;
  const query = q.value.trim().toLowerCase();
  const normalized = sourceItems.filter(item => {{
    if (!query) return true;
    return String(item.article || '').toLowerCase().includes(query) || String(item.nmid || '').includes(query);
  }});
  articleItems = [...normalized].sort((left, right) => {{
    const key = articleSort.key;
    const dir = articleSort.dir === 'asc' ? 1 : -1;
    const leftValue = key === 'article' ? String(left.article || '') : Number(left[key] || 0);
    const rightValue = key === 'article' ? String(right.article || '') : Number(right[key] || 0);
    if (key === 'article') return leftValue.localeCompare(rightValue, 'ru') * dir;
    return (leftValue - rightValue) * dir || String(left.article || '').localeCompare(String(right.article || ''), 'ru');
  }});
  const sortMark = key => articleSort.key === key ? (articleSort.dir === 'asc' ? ' ▲' : ' ▼') : '';
  articleList.innerHTML = articleItems.length
    ? `<table class="article-table"><thead><tr>
        <th data-sort="article">Артикул${{sortMark('article')}}</th>
        <th data-sort="stock">Ост.${{sortMark('stock')}}</th>
        <th data-sort="drr">ДРР${{sortMark('drr')}}</th>
        <th data-sort="turnover">Обор.${{sortMark('turnover')}}</th>
        <th data-sort="revenue">Выручка${{sortMark('revenue')}}</th>
      </tr></thead><tbody>${{articleItems.map(item => {{
        const selected = selectedArticle && String(selectedArticle.nmid) === String(item.nmid) ? ' selected' : '';
        const drr = item.drr != null ? Number(item.drr).toFixed(1) + '%' : '—';
        const turnover = item.turnover != null ? item.turnover + 'д' : '—';
        const article = String(item.article || item.nmid || '');
        const safeArticle = escapeHtml(article);
        const nmid = String(item.nmid || '').trim();
        const copyBtn = `<button class="article-copy" data-copy="${{safeArticle}}" data-skip-select="1" title="Скопировать артикул">⎘</button>`;
        const wbLink = /^\\d+$/.test(nmid)
          ? `<a class="article-link" href="https://www.wildberries.ru/catalog/${{nmid}}/detail.aspx" target="_blank" rel="noopener noreferrer" title="Открыть товар на WB" data-skip-select="1">↗</a>`
          : '';
        const campaignsListLink = /^\\d+$/.test(nmid)
          ? `<a class="campaigns-list-link" href="https://cmp.wildberries.ru/campaigns/list?search=${{nmid}}" target="_blank" rel="noopener noreferrer" title="Рекламные кампании WB" data-skip-select="1">р</a>`
          : '';
        return `<tr class="article-item${{selected}}" data-nmid="${{escapeHtml(item.nmid)}}" data-article="${{escapeHtml(item.article)}}" data-name="${{escapeHtml(item.name || '')}}" title="${{escapeHtml(item.article || item.nmid)}}">
          <td>${{safeArticle}}${{copyBtn}}${{wbLink}}${{campaignsListLink}}</td>
          <td>${{fmt(item.stock)}}</td>
          <td>${{drr}}</td>
          <td>${{turnover}}</td>
          <td>${{fmt(item.revenue)}}</td>
        </tr>`;
      }}).join('')}}</tbody></table>`
    : '<div class="meta" style="padding:8px;">Нет артикулов</div>';
  for (const th of articleList.querySelectorAll('th[data-sort]')) {{
    th.addEventListener('click', () => {{
      const key = th.dataset.sort;
      articleSort = {{ key, dir: articleSort.key === key && articleSort.dir === 'desc' ? 'asc' : 'desc' }};
      renderArticles(sourceItems);
    }});
  }}
  for (const row of articleList.querySelectorAll('.article-item')) {{
    row.addEventListener('click', event => {{
      if (event.target.closest('[data-skip-select="1"]')) return;
      selectedArticle = {{ nmid: row.dataset.nmid, article: row.dataset.article, name: row.dataset.name }};
      document.getElementById('campaign-title').textContent = `Кампании: ${{selectedArticle.article || selectedArticle.nmid}}`;
      renderArticles(articleItems);
      loadCampaigns();
    }});
  }}
  for (const btn of articleList.querySelectorAll('.article-copy')) {{
    btn.addEventListener('click', event => {{
      event.stopPropagation();
      const text = btn.dataset.copy || '';
      navigator.clipboard.writeText(text).then(() => {{
        btn.classList.add('copied');
        btn.textContent = '✓';
        setTimeout(() => {{ btn.classList.remove('copied'); btn.textContent = '⎘'; }}, 1500);
      }});
    }});
  }}
}}
async function loadFilterOptions() {{
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  const key = `${{df}}..${{dt}}`;
  if (key === subjectsLoadedFor) return;
  subjectsLoadedFor = key;
  const currentSubject = subjectEl.value;
  const currentRevenueCategory = revenueCategoryEl.value;
  const data = await fetch(`/api/analytics/buyout-filter-options?${{new URLSearchParams({{date_from:df,date_to:dt}}).toString()}}`).then(r => r.json());
  subjectEl.innerHTML = '<option value="">Все предметы</option>' + (data.subjects || []).map(value => `<option value="${{escapeHtml(value)}}"${{value === currentSubject ? ' selected' : ''}}>${{escapeHtml(value)}}</option>`).join('');
  fillMultiSelect(strategyEl, data.labels || []);
  revenueCategoryEl.innerHTML = '<option value="">Все категории</option>' + (data.revenue_categories || []).map(value => `<option value="${{escapeHtml(value)}}"${{value === currentRevenueCategory ? ' selected' : ''}}>${{escapeHtml(value)}}</option>`).join('');
}}
async function loadArticles(force=false) {{
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  await loadFilterOptions();
  const labels = [...strategyEl.selectedOptions].map(o => o.value).join(',');
  const key = `${{df}}..${{dt}}..${{subjectEl.value}}..${{labels}}..${{revenueCategoryEl.value}}`;
  if (!force && key === articlesLoadedFor) {{
    renderArticles(articleItems);
    return;
  }}
  articlesLoadedFor = key;
  const qs = new URLSearchParams({{date_from:df,date_to:dt,subject:subjectEl.value,labels,revenue_category:revenueCategoryEl.value}});
  const data = await fetch(`/api/analytics/buyout-articles?${{qs.toString()}}`).then(r => r.json());
  allArticleItems = data.articles || [];
  renderArticles(allArticleItems);
}}
function rowState(c) {{
  const s = c.settings || {{}};
  const spend = Number(s.today_spend || 0);
  const budget = Number(s.daily_budget || 0);
  if (s.auto_pause && budget > 0 && spend >= budget) return '<span class="pill off">перерасход</span>';
  return s.schedule_auto ? '<span class="pill on">авто</span>' : '<span class="pill off">ручн.</span>';
}}
function wbState(c) {{
  const status = String(c.status || '');
  if (status === '4') return '<span class="pill" title="WB статус 4">готова</span>';
  if (status === '9') return '<span class="pill on" title="WB статус 9">активна</span>';
  if (status === '11') return '<span class="pill off" title="WB статус 11">пауза</span>';
  if (status === '7') return '<span class="pill off" title="WB статус 7">завершена</span>';
  return `<span class="pill" title="WB статус ${{escapeHtml(status || '—')}}">WB ${{escapeHtml(status || '—')}}</span>`;
}}
function campaignSortValue(c) {{
  const spend = Number((c.metrics || {{}}).spend || 0);
  return [-spend, String(c.name || '')];
}}
function sortCampaigns(items) {{
  return [...items].sort((left, right) => {{
    const a = campaignSortValue(left);
    const b = campaignSortValue(right);
    return (a[0] - b[0]) || String(a[1] || '').localeCompare(String(b[1] || ''), 'ru');
  }});
}}
function budgetPctLabel(budget, avgOrders) {{
  const avg = Number(avgOrders || 0);
  if (avg <= 0) return '—%';
  const pct = Number(budget || 0) / avg * 100;
  if (!Number.isFinite(pct)) return '—%';
  const digits = Math.abs(pct) < 10 ? 1 : 0;
  return pct.toFixed(digits).replace(/\\.0$/, '') + '%';
}}
function budgetPctTitle(budget, avgOrders) {{
  const avg = Number(avgOrders || 0);
  if (avg <= 0) return 'Нет средней суммы заказов за 7 дней';
  return `Бюджет ${{fmt(budget)}} ₽ = ${{budgetPctLabel(budget, avg)}} от средней суммы заказов за 7 дней ${{fmt(avg)}} ₽`;
}}
function syncBudgetPct(row) {{
  const input = row && row.querySelector('.budget');
  const chip = row && row.querySelector('.calc-chip');
  const box = row && row.querySelector('.control-box');
  if (!input || !chip) return;
  const avg = Number(chip.dataset.avgOrders || 0);
  const budget = Number(input.value || 0);
  chip.textContent = budgetPctLabel(budget, avg);
  if (box) box.title = budgetPctTitle(budget, avg);
}}
function swipeToggleHtml(cls, checked, label, title) {{
  const on = Boolean(checked);
  return `<button class="swipe-toggle ${{cls}}${{on ? ' is-on' : ''}}" type="button" role="switch" aria-checked="${{on ? 'true' : 'false'}}" data-checked="${{on ? '1' : '0'}}" title="${{escapeHtml(title || label)}}">
    <span class="swipe-track" aria-hidden="true"><span class="swipe-thumb"></span></span>
    <span class="swipe-label">${{escapeHtml(label)}}</span>
  </button>`;
}}
function swipeToggleValue(row, selector) {{
  const el = row.querySelector(selector);
  return Boolean(el && el.dataset.checked === '1');
}}
function setSwipeToggle(el, checked) {{
  const on = Boolean(checked);
  el.dataset.checked = on ? '1' : '0';
  el.setAttribute('aria-checked', on ? 'true' : 'false');
  el.classList.toggle('is-on', on);
}}
function renderCampaigns(data) {{
  campaigns = sortCampaigns(data.campaigns || []);
  campaignsEl.innerHTML = campaigns.length
    ? campaigns.map(c => {{
        const m = c.metrics || {{}};
        const s = c.settings || {{}};
        const last = c.last_action ? `${{c.last_action.actual_action || ''}} ${{c.last_action.status_code || ''}}` : '';
        const type = c.type || '—';
        const nmCount = Number(c.nm_count || 0);
        const budgetValue = Number(s.daily_budget || 0);
        const avgOrders = Number(s.avg_orders_7d || 0);
        const budgetTitle = budgetPctTitle(budgetValue, avgOrders);
        const budgetPct = budgetPctLabel(budgetValue, avgOrders);
        const nmCountHtml = nmCount ? `<span class="nm-count${{nmCount > 1 ? ' multi' : ''}}" title="Товаров в кампании">${{nmCount}} SKU</span>` : '<span class="nm-count" title="Количество товаров неизвестно">?</span>';
        const campaignUrl = `https://cmp.wildberries.ru/campaigns/edit/${{encodeURIComponent(c.advert_id)}}`;
        return `<tr data-advert-id="${{escapeHtml(c.advert_id)}}">
          <td><div class="campaign-name">${{escapeHtml(c.name || c.advert_id)}} <a class="campaign-link" href="${{campaignUrl}}" target="_blank" rel="noopener noreferrer" title="Открыть РК в WB">↗</a> <span class="type-chip">${{escapeHtml(type)}}</span></div><div class="meta">${{escapeHtml(c.advert_id)}} · ${{escapeHtml(c.source || '')}}${{last ? ' · ' + escapeHtml(last) : ''}}</div></td>
          <td><span class="type-chip">${{escapeHtml(type)}}</span><div class="meta" style="margin-top:3px;">${{nmCountHtml}}</div></td>
          <td>
            <div class="row-controls">
              <span class="control-box${{s.daily_budget_is_default ? ' default-budget' : ''}}" title="${{escapeHtml(budgetTitle)}}"><span>₽/день</span><input class="budget" type="number" min="0" step="100" value="${{budgetValue}}"><span class="calc-chip" data-avg-orders="${{avgOrders}}">${{escapeHtml(budgetPct)}}</span></span>
              ${{swipeToggleHtml('row-autopause', s.auto_pause, 'автовыкл', 'Автовыключение по дневному бюджету')}}
            </div>
          </td>
          <td>
            <div class="row-controls">
              <span class="control-box"><span>Пауза</span><input class="pause-input" value="${{escapeHtml(s.pause_interval || '23-08')}}"></span>
              ${{swipeToggleHtml('row-schedule-auto', s.schedule_auto, 'авто', 'Автоматическая пауза по расписанию')}}
            </div>
          </td>
          <td class="num">${{fmt(s.today_spend, 0)}} / ${{fmt(s.daily_budget, 0)}}<div class="meta">${{s.last_spend_sync_at ? escapeHtml(String(s.last_spend_sync_at).slice(11,16)) : ''}}</div></td>
          <td>${{wbState(c)}} ${{rowState(c)}}<div class="meta">WB ${{escapeHtml(String(c.status || ''))}}</div></td>
          <td class="num">${{fmt(m.spend)}}</td>
          <td class="num">${{fmt(m.views)}}</td>
          <td class="num">${{fmt(m.clicks)}}</td>
          <td class="num">${{fmt(m.ctr,1)}}%</td>
          <td class="num">${{fmt(m.cpc,1)}}</td>
          <td class="num">${{fmt(m.cpm,1)}}</td>
          <td class="num">${{fmt(m.orders)}}</td>
          <td class="num">${{fmt(m.cpo,1)}}</td>
          <td class="num">${{fmt(m.drr,1)}}%</td>
        </tr>`;
      }}).join('')
    : '<tr><td colspan="15" class="meta">Кампании не найдены.</td></tr>';
  if (data.api_error) setStatus('API: ' + data.api_error, 'err');
  else if (data.refresh_started) {{
    setStatus('WB обновляется в фоне, результат будет в логе', 'ok');
    setTimeout(loadAdLog, 1200);
  }}
}}
async function loadCampaigns(refreshDetails=false) {{
  if (!selectedArticle) {{
    campaignsEl.innerHTML = '<tr><td colspan="15" class="meta">Выберите артикул.</td></tr>';
    return;
  }}
  campaignsEl.innerHTML = '<tr><td colspan="15" class="meta">Загружаю...</td></tr>';
  const days = 30;
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  const data = await fetch(`/api/ads/campaigns?nmid=${{encodeURIComponent(selectedArticle.nmid)}}&days=${{days}}&date_from=${{encodeURIComponent(df)}}&date_to=${{encodeURIComponent(dt)}}${{refreshDetails ? '&refresh=1' : ''}}`).then(r => r.json());
  renderCampaigns(data);
}}
async function saveCampaignRow(row) {{
  if (!selectedArticle) return;
  const advertId = row.dataset.advertId;
  const campaign = campaigns.find(item => String(item.advert_id) === String(advertId)) || {{}};
  const payload = {{
    advert_id: advertId,
    enabled: swipeToggleValue(row, '.row-schedule-auto'),
    daily_budget: row.querySelector('.budget').value,
    auto_pause: swipeToggleValue(row, '.row-autopause'),
    pause_interval: row.querySelector('.pause-input').value,
    nmid: selectedArticle.nmid,
    article: selectedArticle.article,
    name: campaign.name || ''
  }};
  const resp = await fetch('/api/ads/campaign-settings', {{ method:'POST', headers:{{'Content-Type':'application/json'}}, body:JSON.stringify(payload) }});
  const data = await resp.json();
  if (!resp.ok) return setStatus(data.error || 'Ошибка', 'err');
  setStatus('Сохранено', 'ok');
  const idx = campaigns.findIndex(item => String(item.advert_id) === String(advertId));
  if (idx >= 0) {{
    const prev = campaigns[idx].settings || {{}};
    campaigns[idx].settings = {{
      ...prev,
      ...data,
      daily_budget: Number(data.daily_budget || payload.daily_budget || 0),
      auto_pause: Boolean(data.auto_pause),
      schedule_auto: Boolean(data.enabled),
      enabled: Boolean(data.enabled),
      pause_interval: `${{String(data.pause_start || '23:00').slice(0,2)}}-${{String(data.pause_end || '08:00').slice(0,2)}}`,
      daily_budget_is_default: false
    }};
    renderCampaigns({{campaigns}});
  }}
  loadAdLog();
}}
campaignsEl.addEventListener('input', event => {{
  const input = event.target.closest('.budget');
  if (!input) return;
  syncBudgetPct(input.closest('tr'));
}});
campaignsEl.addEventListener('change', event => {{
  const input = event.target.closest('.budget,.pause-input');
  if (!input) return;
  saveCampaignRow(input.closest('tr'));
}});
campaignsEl.addEventListener('pointerdown', event => {{
  const toggle = event.target.closest('.swipe-toggle');
  if (!toggle) return;
  event.preventDefault();
  toggle.dataset.dragStartX = String(event.clientX);
  toggle.dataset.dragging = '1';
  try {{ toggle.setPointerCapture(event.pointerId); }} catch (err) {{}}
}});
campaignsEl.addEventListener('pointerup', event => {{
  const toggle = event.target.closest('.swipe-toggle');
  if (!toggle || toggle.dataset.dragging !== '1') return;
  event.preventDefault();
  const startX = Number(toggle.dataset.dragStartX || event.clientX);
  const dx = event.clientX - startX;
  delete toggle.dataset.dragStartX;
  delete toggle.dataset.dragging;
  try {{ toggle.releasePointerCapture(event.pointerId); }} catch (err) {{}}
  if (Math.abs(dx) < 22) return;
  const next = dx > 0;
  if (next === (toggle.dataset.checked === '1')) return;
  setSwipeToggle(toggle, next);
  toggle.classList.add('is-saving');
  saveCampaignRow(toggle.closest('tr')).finally(() => toggle.classList.remove('is-saving'));
}});
campaignsEl.addEventListener('pointercancel', event => {{
  const toggle = event.target.closest('.swipe-toggle');
  if (!toggle) return;
  delete toggle.dataset.dragStartX;
  delete toggle.dataset.dragging;
}});
document.getElementById('run').addEventListener('click', async () => {{
  setStatus('Проверяю...');
  const resp = await fetch('/api/ads/executor/run-once', {{ method:'POST', headers:{{'Content-Type':'application/json'}}, body:JSON.stringify({{force_spend_sync:true}}) }});
  const data = await resp.json();
  if (!resp.ok) return setStatus(data.error || 'Ошибка', 'err');
  setStatus(`Готово, действий: ${{(data.actions || []).length}}, расход: ${{(data.spend_synced || []).length}}`, 'ok');
  loadCampaigns();
  loadAdLog();
}});
document.getElementById('refresh').addEventListener('click', () => {{
  loadCampaigns(true);
}});
document.getElementById('reload-log').addEventListener('click', loadAdLog);
q.addEventListener('input', () => {{ clearTimeout(window.__adSearch); window.__adSearch = setTimeout(() => renderArticles(allArticleItems), 180); }});
for (const id of ['df','dt']) {{
  document.getElementById(id).addEventListener('change', () => {{
    reloadPeriodData();
  }});
}}
for (const btn of document.querySelectorAll('[data-quick-period]')) {{
  btn.addEventListener('click', () => applyQuickPeriod(btn.dataset.quickPeriod || 'today'));
}}
for (const el of [subjectEl, strategyEl, revenueCategoryEl]) {{
  el.addEventListener('change', () => {{
    selectedArticle = null;
    document.getElementById('campaign-title').textContent = 'Кампании артикула';
    campaignsEl.innerHTML = '<tr><td colspan="15" class="meta">Выберите артикул.</td></tr>';
    articlesLoadedFor = '';
    loadArticles(true);
  }});
}}
document.getElementById('toggle-articles').addEventListener('click', () => {{ articlesCollapsed = !articlesCollapsed; syncArticlesPanel(); }});
document.getElementById('expand-articles').addEventListener('click', () => {{ articlesCollapsed = false; syncArticlesPanel(); }});
syncArticlesPanel();
loadArticles(true);
loadAdLog();
</script>
</body>
</html>
"""

_FAVICON_SVG = (
    "<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 32 32'>"
    "<rect width='32' height='32' rx='7' fill='#0a0f1e'/>"
    "<defs><linearGradient id='g' x1='0' y1='0' x2='32' y2='32' gradientUnits='userSpaceOnUse'>"
    "<stop offset='0' stop-color='#38bdf8'/><stop offset='1' stop-color='#14b8a6'/>"
    "</linearGradient></defs>"
    "<ellipse cx='16' cy='22' rx='13' ry='3' fill='#0c4a6e' opacity='.5'/>"
    "<path d='M3 22 A13 13 0 0 1 29 22Z' fill='url(#g)'/>"
    "<path d='M6 14 Q16 9 26 14' stroke='rgba(255,255,255,.18)' stroke-width='.8' fill='none'/>"
    "<line x1='16' y1='9' x2='16' y2='22' stroke='rgba(255,255,255,.18)' stroke-width='.8'/>"
    "<path d='M6 19 Q12 11 18 8' stroke='rgba(255,255,255,.45)' stroke-width='1.2' fill='none' stroke-linecap='round'/>"
    "</svg>"
)
FAVICON_LINK = f'<link rel="icon" type="image/svg+xml" href="data:image/svg+xml;base64,{base64.b64encode(_FAVICON_SVG.encode()).decode()}">'

BRAND_HEADER_HTML = """<div style="background:linear-gradient(135deg,#0a0f1e 0%,#1a1145 50%,#0c3a5e 100%);border-bottom:1px solid rgba(56,189,248,0.12);box-shadow:0 2px 16px rgba(0,0,0,0.38);position:relative;z-index:100;">
  <div style="display:flex;align-items:center;gap:9px;max-width:3000px;margin:0 auto;height:48px;padding:0 20px;">
    <svg width="38" height="26" viewBox="0 0 44 30" fill="none" xmlns="http://www.w3.org/2000/svg" style="flex-shrink:0;">
      <defs>
        <linearGradient id="psG" x1="0" y1="0" x2="44" y2="30" gradientUnits="userSpaceOnUse">
          <stop offset="0%" stop-color="#38bdf8"/>
          <stop offset="60%" stop-color="#14b8a6"/>
          <stop offset="100%" stop-color="#0891b2"/>
        </linearGradient>
        <radialGradient id="psShine" cx="35%" cy="25%" r="55%">
          <stop offset="0%" stop-color="rgba(255,255,255,0.28)"/>
          <stop offset="100%" stop-color="rgba(255,255,255,0)"/>
        </radialGradient>
      </defs>
      <ellipse cx="22" cy="26" rx="19" ry="4.5" fill="#0c4a6e" opacity="0.5"/>
      <path d="M3 26 A19 19 0 0 1 41 26 Z" fill="url(#psG)"/>
      <path d="M3 26 A19 19 0 0 1 41 26 Z" fill="url(#psShine)"/>
      <ellipse cx="22" cy="26" rx="19" ry="4.5" fill="none" stroke="rgba(255,255,255,0.22)" stroke-width="0.8"/>
      <path d="M8 16 Q22 10 36 16" stroke="rgba(255,255,255,0.17)" stroke-width="0.8" fill="none"/>
      <line x1="22" y1="7" x2="22" y2="26" stroke="rgba(255,255,255,0.18)" stroke-width="0.8"/>
      <path d="M13 9 Q13 26 13 26" stroke="rgba(255,255,255,0.13)" stroke-width="0.7" fill="none"/>
      <path d="M31 9 Q31 26 31 26" stroke="rgba(255,255,255,0.13)" stroke-width="0.7" fill="none"/>
      <path d="M8 21 Q16 12 25 9" stroke="rgba(255,255,255,0.38)" stroke-width="1.3" fill="none" stroke-linecap="round"/>
    </svg>
    <div style="display:flex;flex-direction:column;gap:0;line-height:1;">
      <span title="Не СФЕРА. Но тоже стильно." style="font-size:1.16rem;font-weight:900;background:linear-gradient(90deg,#7dd3fc,#2dd4bf);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;letter-spacing:0;cursor:default;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">ПолуСфера</span>
      <span style="font-size:0.56rem;font-weight:600;color:rgba(148,163,184,0.75);letter-spacing:0.08em;text-transform:uppercase;margin-top:1px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">by ИИ Технологии</span>
    </div>
    <!--CABINET_LABEL-->
  </div>
</div>"""

def _cabinet_switcher_html(cabinet: dict | None) -> str:
    if not cabinet:
        return ""
    name = escape(cabinet.get("name", ""))
    mp = cabinet.get("marketplace", "wb")
    if mp == "wb":
        color = "#3b82f6"
    elif mp == "ozon":
        color = "#f97316"
    else:
        color = "#8b5cf6"

    # Embed cabinet list directly — no fetch needed, avoids Basic Auth issue
    try:
        all_cabinets = get_platform().list_cabinets()
    except Exception:
        all_cabinets = []

    cab_buttons = ""
    for c in all_cabinets:
        dot_color = "#3b82f6" if c.get("marketplace") == "wb" else "#f97316" if c.get("marketplace") == "ozon" else "#8b5cf6"
        cab_id = escape(c.get("cabinet_id", ""))
        cab_name = escape(c.get("name", ""))
        active_style = "border-color:#4f46e5;" if c.get("cabinet_id") == cabinet.get("cabinet_id") else ""
        cab_buttons += (
            f'<button data-cab="{cab_id}" '
            f'style="box-sizing:border-box;width:100%;min-width:0;padding:10px 14px;background:#0f172a;color:#e2e8f0;'
            f'border:1.5px solid #334155;border-radius:8px;cursor:pointer;text-align:left;'
            f'font-size:.9rem;font-weight:600;display:flex;align-items:center;gap:8px;{active_style}">'
            f'<span style="width:8px;height:8px;border-radius:50%;background:{dot_color};flex-shrink:0;"></span>'
            f'{cab_name}</button>'
        )

    return (
        f'<div style="margin-left:auto;display:flex;align-items:center;gap:8px;">'
        f'<button id="cab-badge" onclick="document.getElementById(\'cab-modal\').style.display=\'flex\';document.getElementById(\'cab-pin-form\').style.display=\'none\'" '
        f'style="cursor:pointer;background:{color};color:#fff;font-size:.75rem;font-weight:700;'
        f'padding:4px 12px;border-radius:999px;white-space:nowrap;'
        f'border:1.5px solid rgba(255,255,255,.25);user-select:none;">{name} ▾</button></div>'
        f'<div id="cab-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);'
        f'z-index:9999;align-items:center;justify-content:center;padding:12px;overflow:hidden;" '
        f'onclick="if(event.target===this)this.style.display=\'none\'">'
        f'<div style="box-sizing:border-box;background:#1e293b;border-radius:16px;padding:clamp(16px,4vw,24px);'
        f'width:min(420px,calc(100vw - 24px));max-width:100%;'
        f'box-shadow:0 20px 60px rgba(0,0,0,.6);max-height:90vh;overflow-y:auto;overflow-x:hidden;">'
        f'<div style="color:#f1f5f9;font-size:1rem;font-weight:700;margin-bottom:14px;">Сменить кабинет</div>'
        f'<div id="cab-list" style="display:flex;flex-direction:column;gap:6px;">{cab_buttons}</div>'
        f'<div id="cab-pin-form" style="display:none;margin-top:14px;">'
        f'<div id="cab-pin-label" style="color:#94a3b8;font-size:.82rem;margin-bottom:12px;text-align:center;font-weight:600;"></div>'
        f'<div style="display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:clamp(8px,2.5vw,12px);width:100%;margin-bottom:12px;">'
        f'<input id="pin0" type="password" maxlength="1" inputmode="numeric" autocomplete="off" '
        f'style="box-sizing:border-box;width:100%;min-width:0;max-width:86px;aspect-ratio:1/1;justify-self:center;text-align:center;font-size:clamp(1.4rem,7vw,2rem);font-weight:700;'
        f'background:#0f172a;color:#f1f5f9;border:2px solid #334155;border-radius:10px;outline:none;-webkit-text-security:disc;">'
        f'<input id="pin1" type="password" maxlength="1" inputmode="numeric" autocomplete="off" '
        f'style="box-sizing:border-box;width:100%;min-width:0;max-width:86px;aspect-ratio:1/1;justify-self:center;text-align:center;font-size:clamp(1.4rem,7vw,2rem);font-weight:700;'
        f'background:#0f172a;color:#f1f5f9;border:2px solid #334155;border-radius:10px;outline:none;-webkit-text-security:disc;">'
        f'<input id="pin2" type="password" maxlength="1" inputmode="numeric" autocomplete="off" '
        f'style="box-sizing:border-box;width:100%;min-width:0;max-width:86px;aspect-ratio:1/1;justify-self:center;text-align:center;font-size:clamp(1.4rem,7vw,2rem);font-weight:700;'
        f'background:#0f172a;color:#f1f5f9;border:2px solid #334155;border-radius:10px;outline:none;-webkit-text-security:disc;">'
        f'</div>'
        f'<div id="pin-error" style="color:#f87171;font-size:.82rem;text-align:center;min-height:1.2em;margin-bottom:8px;"></div>'
        f'<button onclick="document.getElementById(\'cab-pin-form\').style.display=\'none\';[\'pin0\',\'pin1\',\'pin2\'].forEach(function(x){{var e=document.getElementById(x);if(e)e.value=\'\';}});" '
        f'style="width:100%;padding:9px;background:#334155;color:#94a3b8;'
        f'border:none;border-radius:8px;cursor:pointer;font-size:.85rem;">← Назад</button>'
        f'</div>'
        f'<button onclick="document.getElementById(\'cab-modal\').style.display=\'none\'" '
        f'style="margin-top:8px;width:100%;padding:9px;background:transparent;color:#475569;'
        f'border:1px solid #334155;border-radius:8px;cursor:pointer;font-size:.82rem;">Отмена</button>'
        f'</div></div>'
        f'<script>(function(){{'
        f'var selCabId=null;'
        f'document.getElementById("cab-list").addEventListener("click",function(e){{'
        f'var btn=e.target.closest("button[data-cab]");'
        f'if(!btn)return;'
        f'selCabId=btn.getAttribute("data-cab");'
        f'document.getElementById("cab-pin-label").textContent=btn.textContent.trim();'
        f'document.getElementById("cab-pin-form").style.display="block";'
        f'document.getElementById("pin-error").textContent="";'
        f'["pin0","pin1","pin2"].forEach(function(id){{var el=document.getElementById(id);if(el)el.value="";}});'
        f'var p=document.getElementById("pin0");if(p)p.focus();'
        f'}});'
        f'[0,1,2].forEach(function(i){{'
        f'var el=document.getElementById("pin"+i);'
        f'if(!el)return;'
        f'el.addEventListener("input",function(){{'
        f'this.value=this.value.replace(/[^0-9]/g,"");'
        f'if(this.value.length===1){{'
        f'if(i<2){{var nx=document.getElementById("pin"+(i+1));if(nx)nx.focus();}}'
        f'else{{'
        f'var pin=(document.getElementById("pin0").value||"")+(document.getElementById("pin1").value||"")+(document.getElementById("pin2").value||"");'
        f'fetch("/api/switch-cabinet",{{method:"POST",headers:{{"Content-Type":"application/json"}},'
        f'body:JSON.stringify({{cabinet_id:selCabId,pin:pin}})}}).then(function(r){{return r.json();}}).then(function(d){{'
        f'if(d.ok){{window.location.reload();}}'
        f'else{{document.getElementById("pin-error").textContent=d.error||"Неверный PIN";'
        f'["pin0","pin1","pin2"].forEach(function(id){{var el=document.getElementById(id);if(el)el.value="";}});'
        f'var p=document.getElementById("pin0");if(p)p.focus();}}'
        f'}}).catch(function(){{document.getElementById("pin-error").textContent="Ошибка соединения";}});'
        f'}}'
        f'}}'
        f'}});'
        f'}});'
        f'}})()</script>'
    )

def _inject_brand(html: str, cabinet: dict | None = None) -> str:
    cab_html = _cabinet_switcher_html(cabinet)
    header = BRAND_HEADER_HTML.replace("<!--CABINET_LABEL-->", cab_html)
    return (html
            .replace("</head>", FAVICON_LINK + "</head>", 1)
            .replace("<body>", "<body>" + header, 1))

def _sync_panel_html(cabinet: dict | None, marketplace: str) -> str:
    if marketplace == "wb":
        if not _has_wb_cabinet(cabinet):
            return (
                '<div class="sync-card disabled">'
                '<h2>Wildberries</h2>'
                '<p>Для выбранного кабинета нет активного WB-ключа.</p>'
                '<div class="missing">Нет кабинета WB</div>'
                '</div>'
            )
        return (
            '<div class="sync-card wb">'
            '<h2>Wildberries</h2>'
            '<p>Полная или частичная загрузка WB: продажи, заказы, остатки, реклама, воронка.</p>'
            '<div class="load-grid">'
            '<button type="button" class="primary" data-load-mode="all">Загрузить всё WB</button>'
            '<button type="button" class="secondary" data-load-mode="sales">Продажи</button>'
            '<button type="button" class="secondary" data-load-mode="orders">Заказы</button>'
            '<button type="button" class="secondary" data-load-mode="stocks">Остатки</button>'
            '<button type="button" class="secondary" data-load-mode="ads">Реклама</button>'
            '<button type="button" class="secondary" data-load-mode="funnel">Воронка API</button>'
            '<button type="button" class="secondary" data-load-mode="all" data-skip-ads="1">Все без рекламы</button>'
            '<button type="button" class="secondary" data-load-mode="all" data-skip-funnel="1">Все без воронки</button>'
            '</div>'
            + _funnel_upload_panel_html(cabinet) +
            '</div>'
        )
    if not _has_ozon_cabinet(cabinet):
        return (
            '<div class="sync-card disabled ozon">'
            '<h2>OZON</h2>'
            '<p>Для выбранного кабинета нет активных Ozon ключей.</p>'
            '<div class="missing">Нет кабинета OZON</div>'
            '</div>'
        )
    return (
        '<div class="sync-card ozon">'
        '<h2>OZON</h2>'
        '<p>Полный обмен Ozon, быстрый режим без рекламы или отдельная загрузка рекламных расходов.</p>'
        '<div class="load-grid">'
        '<button type="button" class="primary" data-ozon-sync="full">Загрузить всё OZON</button>'
        '<button type="button" class="secondary" data-ozon-sync="fast">OZON без рекламы</button>'
        '<button type="button" class="secondary" data-ozon-sync="ads">Реклама</button>'
        '</div>'
        '</div>'
    )

def _funnel_upload_panel_html(cabinet: dict | None) -> str:
    if not _has_wb_cabinet(cabinet):
        return ""
    return (
        '<div class="wb-funnel-upload">'
        '<h3>Воронка WB из Excel</h3>'
        '<div class="hint">Файл обновит только даты, которые есть внутри него. Остальные даты останутся без изменений.</div>'
        '<input id="funnel-file" type="file" accept=".xlsx">'
        '<div class="funnel-actions">'
        '<button type="button" id="funnel-upload">Загрузить воронку</button>'
        '<button type="button" id="funnel-force-upload" class="force">Загрузить всё равно</button>'
        '</div>'
        '<div id="funnel-status" class="funnel-status">Поддерживается Excel WB: лист Товары, дата в колонке J, показы в K, переходы в M.</div>'
        '</div>'
    )

def _home_errors_html(cabinet: dict | None) -> str:
    yesterday = (date.today() - timedelta(days=1)).isoformat()
    issues: list[dict[str, str]] = []

    def add_issue(title: str, detail: str) -> None:
        issues.append({"title": title, "detail": detail})

    def _count_for_day(conn: sqlite3.Connection, table: str, date_expr: str, day: str) -> int:
        if not _table_exists(conn, table):
            return 0
        row = conn.execute(f"SELECT COUNT(*) AS n FROM {table} WHERE {date_expr} = ?", (day,)).fetchone()
        return int(row["n"] or 0) if row else 0

    try:
        with _db_connect() as conn:
            if _has_wb_cabinet(cabinet):
                funnel_count = _count_for_day(conn, "funnel_analytics", "date", yesterday)
                if funnel_count <= 0:
                    add_issue(
                        "Воронка WB за вчера не загружена",
                        f"В `funnel_analytics` нет строк за {yesterday}. Загрузите воронку WB из Excel или через API.",
                    )
                else:
                    pass

                raw_orders_count = _count_for_day(conn, "raw_orders", "substr(date, 1, 10)", yesterday)
                buyout_count = _count_for_day(conn, "buyout_order_day", '"Дата"', yesterday)
                if raw_orders_count <= 0 and buyout_count <= 0:
                    add_issue(
                        "Обмена WB за вчера не было",
                        f"Нет строк за {yesterday} ни в `raw_orders`, ни в `buyout_order_day`. Запустите обмен WB за вчера.",
                    )

            if _table_exists(conn, "SKU"):
                columns = _table_columns(conn, "SKU")
                cogs_col = _first_existing(columns, ["себестоимость", "cost_price", "cogs"])
                article_col = _first_existing(columns, ["Артикул поставщика", "Артикул WB", "SKU Ozon", "sku"])
                name_col = _first_existing(columns, ["Название", "Наименование", "name"])
                if cogs_col:
                    cogs_expr = (
                        f"CAST(REPLACE(REPLACE(COALESCE(NULLIF(TRIM({_sql_ident(cogs_col)}), ''), '0'), ' ', ''), ',', '.') AS REAL)"
                    )
                    missing_count = int(conn.execute(
                        f"SELECT COUNT(*) AS n FROM SKU WHERE TRIM(COALESCE({_sql_ident(cogs_col)}, '')) = '' OR {cogs_expr} <= 0"
                    ).fetchone()["n"] or 0)
                    if missing_count > 0:
                        select_parts = []
                        if article_col:
                            select_parts.append(f"COALESCE(NULLIF(TRIM({_sql_ident(article_col)}), ''), '') AS article")
                        else:
                            select_parts.append("'' AS article")
                        if name_col:
                            select_parts.append(f"COALESCE(NULLIF(TRIM({_sql_ident(name_col)}), ''), '') AS name")
                        else:
                            select_parts.append("'' AS name")
                        sample_rows = conn.execute(
                            "SELECT "
                            + ", ".join(select_parts)
                            + f" FROM SKU WHERE TRIM(COALESCE({_sql_ident(cogs_col)}, '')) = '' OR {cogs_expr} <= 0 LIMIT 5"
                        ).fetchall()
                        examples = []
                        for row in sample_rows:
                            article = str(row["article"] or "").strip()
                            name = str(row["name"] or "").strip()
                            examples.append(" — ".join(part for part in [article, name] if part) or "без артикула")
                        suffix = f": {', '.join(examples)}" if examples else ""
                        add_issue(
                            "Есть товары без себестоимости",
                            f"В SKU найдено {missing_count} строк без себестоимости{suffix}.",
                        )
                else:
                    add_issue("Нет колонки себестоимости", "В таблице SKU не найдена колонка `себестоимость` / `cost_price` / `cogs`.")
    except Exception as exc:
        add_issue("Не удалось проверить ошибки", str(exc))

    if issues:
        items = "".join(
            f'<li><div class="err-title">{escape(item["title"])}</div><div class="err-detail">{escape(item["detail"])}</div></li>'
            for item in issues
        )
        status = f"{len(issues)}"
        return (
            '<section class="error-panel has-errors">'
            '<div class="error-head"><h2>Ошибки</h2><span>' + status + '</span></div>'
            '<ul class="error-list">' + items + '</ul>'
            '</section>'
        )

    return ""

HTML = """\
<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Загрузка данных</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
         background: #f0f2f5; display: flex; flex-direction: column;
         align-items: stretch; min-height: 100vh; }}
  .page-nav {{ width: calc(100% - 28px); max-width: 3000px; margin: 0 auto; padding: 6px 0 0; }}
  .top {{ display:flex; align-items:center; gap:7px; margin-bottom:5px; flex-wrap:nowrap; overflow:hidden; }}
  .top h1 {{ font-size:.82rem; font-weight:700; color:#6b7280; white-space:nowrap; margin:0; flex-shrink:0; }}
  .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
  .report-nav::-webkit-scrollbar {{ display:none; }}
  .card-wrap {{ display: flex; justify-content: center; padding: 8px 16px 18px; }}
  .card {{ background: #fff; border-radius: 10px; padding: 22px 28px 20px;
           box-shadow: 0 2px 12px rgba(0,0,0,.08); width: 100%; max-width: 1100px; }}
  h1 {{ font-size: 1.25rem; font-weight: 700; color: #1a1a2e; margin-bottom: 14px; }}
  label {{ display: block; font-size: .85rem; font-weight: 600;
           color: #555; margin-bottom: 4px; margin-top: 8px; }}
  input[type=date] {{ width: 100%; padding: 7px 10px; border: 1.5px solid #d1d5db;
                      border-radius: 7px; font-size: .95rem; outline: none;
                      transition: border-color .2s; }}
  input[type=date]:focus {{ border-color: #4f46e5; }}
  .row {{ display: flex; gap: 16px; }}
  .row > div {{ flex: 1; }}
  .quick {{ display: flex; gap: 8px; margin-top: 9px; flex-wrap: wrap; }}
  .quick button {{ margin-top: 0; width: auto; padding: 6px 10px; font-size: .8rem; background: #e5e7eb; color: #111827; }}
  .quick button:hover:not(:disabled) {{ background: #d1d5db; }}
  .load-grid {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 8px; margin-top: 14px; }}
  .load-grid button {{ margin-top: 0; }}
  .load-grid .primary {{ grid-column: 1 / -1; background: #0f766e; }}
  .load-grid .primary:hover:not(:disabled) {{ background: #115e59; }}
  .load-grid .secondary {{ background: #4f46e5; }}
  .load-grid .secondary:hover:not(:disabled) {{ background: #4338ca; }}
  .sync-card.wb .load-grid {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
  .sync-grid {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:14px; margin-top:16px; align-items:start; }}
  .sync-card {{ border:1px solid #e5e7eb; border-radius:10px; padding:14px; background:#f8fafc; }}
  .sync-card h2 {{ font-size:.96rem; color:#111827; margin:0 0 6px; }}
  .sync-card p {{ color:#4b5563; font-size:.79rem; line-height:1.35; margin:0 0 8px; }}
  .sync-card.ozon h2 {{ color:#c2410c; }}
  .sync-card.ozon .primary {{ background:#f97316; }}
  .sync-card.ozon .primary:hover:not(:disabled) {{ background:#ea580c; }}
  .sync-card.ozon .secondary {{ background:#6b7280; }}
  .sync-card.ozon .secondary:hover:not(:disabled) {{ background:#4b5563; }}
  .sync-card.ozon .load-grid {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
  .sync-card.disabled {{ background:#f3f4f6; border-style:dashed; }}
  .sync-card.disabled h2 {{ color:#6b7280; }}
  .sync-card.disabled .missing {{ margin-top:8px; padding:8px 10px; border-radius:7px; background:#fff7ed; color:#9a3412; font-weight:700; font-size:.82rem; }}
  .error-panel {{ margin: 0 0 14px; border-radius: 10px; border: 1px solid #e5e7eb; padding: 13px 14px; }}
  .error-panel.has-errors {{ background: #fff7ed; border-color: #fed7aa; }}
  .error-panel.ok {{ background: #ecfdf5; border-color: #a7f3d0; }}
  .error-head {{ display:flex; align-items:center; justify-content:space-between; gap:10px; margin-bottom:8px; }}
  .error-head h2 {{ font-size:.98rem; color:#111827; margin:0; }}
  .error-head span {{ min-width: 28px; padding: 3px 8px; border-radius: 999px; text-align:center; font-size:.78rem; font-weight:800; background:#fff; color:#9a3412; border:1px solid #fed7aa; }}
  .error-panel.ok .error-head span {{ color:#047857; border-color:#a7f3d0; }}
  .error-list {{ list-style:none; display:grid; gap:8px; margin:0; padding:0; }}
  .error-list li {{ background:#fff; border:1px solid #fed7aa; border-radius:8px; padding:9px 10px; }}
  .err-title {{ color:#9a3412; font-size:.86rem; font-weight:800; margin-bottom:3px; }}
  .err-detail {{ color:#7c2d12; font-size:.78rem; line-height:1.35; }}
  .error-empty {{ color:#047857; font-size:.82rem; line-height:1.35; }}
  .hint {{ margin-top: 6px; color: #4b5563; font-size: .77rem; line-height:1.25; }}
  .wb-funnel-upload {{ margin-top: 11px; padding-top: 10px; border-top: 1px solid #dbeafe; }}
  .wb-funnel-upload h3 {{ font-size: .84rem; color: #1f2937; margin: 0 0 4px; }}
  .wb-funnel-upload input[type=file] {{ width: 100%; margin-top: 7px; padding: 7px;
                                        border: 1.5px solid #d1d5db; border-radius: 7px; background: #fff; font-size:.78rem; }}
  .funnel-actions {{ display: flex; gap: 8px; margin-top: 8px; }}
  .funnel-actions button {{ margin-top: 0; }}
  .funnel-actions .force {{ display: none; background: #b45309; }}
  .funnel-status {{ margin-top: 7px; color: #4b5563; font-size: .74rem; line-height: 1.25; }}
  .funnel-status.ok {{ color: #047857; }}
  .funnel-status.err {{ color: #b91c1c; }}
  .funnel-status.warn {{ color: #92400e; }}
  button {{ margin-top: 18px; width: 100%; padding: 9px 11px;
            background: #4f46e5; color: #fff; border: none; border-radius: 7px;
            font-size: .9rem; font-weight: 700; cursor: pointer;
            transition: background .2s; }}
  button:hover:not(:disabled) {{ background: #4338ca; }}
  button:disabled {{ background: #a5b4fc; cursor: not-allowed; }}
  #log-wrap {{ display: none; margin-top: 14px; }}
  #log {{ background: #111827; color: #d1fae5; border-radius: 8px; padding: 16px;
          font-size: .8rem; font-family: "Menlo", "Courier New", monospace;
          max-height: 320px; overflow-y: auto; white-space: pre-wrap; word-break: break-all; }}
  .status {{ margin-top: 8px; font-size: .84rem; font-weight: 600; }}
  .status.ok {{ color: #059669; }}
  .status.err {{ color: #dc2626; }}
  @media (max-width: 760px) {{ .sync-grid {{ grid-template-columns:1fr; }} .row {{ flex-direction:column; gap:0; }} }}
</style>
</head>
<body>
<div class="page-nav">
  <div class="top">
    <h1>Загрузка данных</h1>
    <div class="report-nav">{report_nav}</div>
  </div>
</div>
<div class="card-wrap">
<div class="card">
  <h1>Синхронизация кабинета</h1>
  {errors_panel}
  <form id="form">
    <div class="row">
      <div>
        <label for="df">Дата с</label>
        <input type="date" id="df" name="date_from" value="{date_from}" required>
      </div>
      <div>
        <label for="dt">Дата по</label>
        <input type="date" id="dt" name="date_to" value="{date_to}" required>
      </div>
    </div>
    <div class="quick">
      <button type="button" data-range="yesterday">За вчера</button>
      <button type="button" data-range="last2">2 дня</button>
      <button type="button" data-range="today">Сегодня</button>
      <button type="button" data-range="last7">Последние 7 дней</button>
    </div>
    <div class="sync-grid">
      {wb_sync_panel}
      {ozon_sync_panel}
    </div>
    <div class="hint">Синхронизация через кнопки идет по всему выбранному кабинету. Если ключей маркетплейса нет, синк для него не запускается.</div>
  </form>
  <div id="log-wrap">
    <div id="log"></div>
    <div id="status" class="status"></div>
  </div>
</div>
</div>
<script>
const form = document.getElementById('form');
const loadButtons = [...document.querySelectorAll('[data-load-mode]')];
const ozonButtons = [...document.querySelectorAll('[data-ozon-sync]')];
const funnelButtons = [...document.querySelectorAll('#funnel-upload, #funnel-force-upload')];
const syncButtons = [...loadButtons, ...ozonButtons, ...funnelButtons];
const logEl = document.getElementById('log');
const logWrap = document.getElementById('log-wrap');
const statusEl = document.getElementById('status');
const dfEl = document.getElementById('df');
const dtEl = document.getElementById('dt');
const funnelFileEl = document.getElementById('funnel-file');
const funnelUploadBtn = document.getElementById('funnel-upload');
const funnelForceBtn = document.getElementById('funnel-force-upload');
const funnelStatusEl = document.getElementById('funnel-status');
let lastFunnelUpload = null;

function isoDate(d) {{
  return d.toISOString().slice(0, 10);
}}

for (const quickBtn of document.querySelectorAll('.quick button[data-range]')) {{
  quickBtn.addEventListener('click', () => {{
    const today = new Date();
    const mode = quickBtn.dataset.range;
    if (mode === 'today') {{
      const t = isoDate(today);
      dfEl.value = t;
      dtEl.value = t;
      return;
    }}
    if (mode === 'yesterday') {{
      const y = new Date(today);
      y.setDate(y.getDate() - 1);
      const yd = isoDate(y);
      dfEl.value = yd;
      dtEl.value = yd;
      return;
    }}
    if (mode === 'last2') {{
      const from = new Date(today);
      const to = new Date(today);
      from.setDate(from.getDate() - 2);
      to.setDate(to.getDate() - 1);
      dfEl.value = isoDate(from);
      dtEl.value = isoDate(to);
      return;
    }}
    if (mode === 'last7') {{
      const from = new Date(today);
      from.setDate(from.getDate() - 6);
      dfEl.value = isoDate(from);
      dtEl.value = isoDate(today);
    }}
  }});
}}

function appendLog(text) {{
  logEl.textContent += text + '\\n';
  logEl.scrollTop = logEl.scrollHeight;
}}

async function startLoad(buttonOrMode) {{
  const activeButton = typeof buttonOrMode === 'string'
    ? loadButtons.find(button => button.dataset.loadMode === buttonOrMode)
    : buttonOrMode;
  const mode = activeButton ? (activeButton.dataset.loadMode || 'all') : (buttonOrMode || 'all');
  const skipAds = activeButton && activeButton.dataset.skipAds === '1';
  const skipFunnel = activeButton && activeButton.dataset.skipFunnel === '1';
  const df = dfEl.value;
  const dt = dtEl.value;

  syncButtons.forEach(button => button.disabled = true);
  const oldText = activeButton ? activeButton.textContent : '';
  if (activeButton) activeButton.textContent = '⏳ Загружаю...';
  logEl.textContent = '';
  statusEl.textContent = '';
  statusEl.className = 'status';
  logWrap.style.display = 'block';
  appendLog(`WB: запрос отправлен ${{df}}..${{dt}}, режим: ${{mode}}`);
  appendLog('Загрузка идет по всему кабинету WB.');
  statusEl.textContent = 'Выполняется...';

  const url = `/stream?date_from=${{df}}&date_to=${{dt}}&mode=${{encodeURIComponent(mode)}}&skip_ads=${{skipAds ? '1' : '0'}}&skip_funnel=${{skipFunnel ? '1' : '0'}}`;
  const resp = await fetch(url);
  if (!resp.ok || !resp.body) {{
    statusEl.textContent = `❌ Не удалось открыть поток логов (HTTP ${{resp.status}})`;
    statusEl.className = 'status err';
    syncButtons.forEach(button => button.disabled = false);
    if (activeButton) activeButton.textContent = oldText;
    return;
  }}
  const reader = resp.body.getReader();
  const decoder = new TextDecoder();
  let buf = '';

  while (true) {{
    const {{ done, value }} = await reader.read();
    if (done) break;
    buf += decoder.decode(value, {{ stream: true }});
    const lines = buf.split('\\n');
    buf = lines.pop();
    for (const line of lines) {{
      if (line.startsWith('data: ')) {{
        const payload = JSON.parse(line.slice(6));
        if (payload.type === 'log') {{
          appendLog(payload.text);
        }} else if (payload.type === 'heartbeat') {{
          statusEl.textContent = payload.text;
        }} else if (payload.type === 'done') {{
          statusEl.textContent = payload.text;
          statusEl.className = 'status ok';
        }} else if (payload.type === 'error') {{
          statusEl.textContent = payload.text;
          statusEl.className = 'status err';
        }}
      }}
    }}
  }}

  syncButtons.forEach(button => button.disabled = false);
  if (activeButton) activeButton.textContent = oldText;
}}

function funnelAsBase64(file) {{
  return new Promise((resolve, reject) => {{
    const reader = new FileReader();
    reader.onload = () => {{
      const text = String(reader.result || '');
      resolve(text.includes(',') ? text.split(',', 2)[1] : text);
    }};
    reader.onerror = reject;
    reader.readAsDataURL(file);
  }});
}}

async function uploadFunnel(force = false) {{
  if (!funnelFileEl || !funnelUploadBtn || !funnelStatusEl) return;
  const file = funnelFileEl.files && funnelFileEl.files[0];
  if (!file) {{
    funnelStatusEl.textContent = 'Выбери файл .xlsx';
    funnelStatusEl.className = 'funnel-status err';
    return;
  }}
  syncButtons.forEach(button => button.disabled = true);
  if (funnelForceBtn) funnelForceBtn.style.display = 'none';
  funnelStatusEl.textContent = 'Загружаю...';
  funnelStatusEl.className = 'funnel-status';
  try {{
    const content = lastFunnelUpload && lastFunnelUpload.name === file.name
      ? lastFunnelUpload.content
      : await funnelAsBase64(file);
    lastFunnelUpload = {{ name: file.name, content }};
    const resp = await fetch('/api/analytics/funnel-upload', {{
      method: 'POST',
      headers: {{ 'Content-Type': 'application/json' }},
      body: JSON.stringify({{ filename: file.name, content, force }})
    }});
    const data = await resp.json();
    if (data.warning) {{
      const details = (data.details || [])
        .map(item => `${{item.date}}: в файле ${{item.uploaded}}, в базе ${{item.existing}}, меньше на ${{item.missing}}`)
        .join('; ');
      funnelStatusEl.className = 'funnel-status warn';
      funnelStatusEl.textContent = `${{data.message}} ${{details}}`;
      if (funnelForceBtn) funnelForceBtn.style.display = 'inline-block';
      return;
    }}
    if (!resp.ok) throw new Error(data.error || `HTTP ${{resp.status}}`);
    funnelStatusEl.className = 'funnel-status ok';
    funnelStatusEl.textContent = `Готово: строк ${{data.rows}}, товаров ${{data.nmids}}, период ${{data.date_from}}..${{data.date_to}}, показов ${{data.impressions}}. buyout_order_day: ${{data.buyout_order_day_rows}} строк.`;
  }} catch (err) {{
    funnelStatusEl.className = 'funnel-status err';
    funnelStatusEl.textContent = `Ошибка: ${{err.message || err}}`;
  }} finally {{
    syncButtons.forEach(button => button.disabled = false);
  }}
}}

async function startOzonSync(mode) {{
  const df = dfEl.value;
  const dt = dtEl.value;
  const activeButton = ozonButtons.find(button => button.dataset.ozonSync === mode);
  const withAds = mode === 'full';
  const adsOnly = mode === 'ads';
  syncButtons.forEach(button => button.disabled = true);
  const oldText = activeButton ? activeButton.textContent : '';
  if (activeButton) activeButton.textContent = '⏳ Загружаю...';
  logEl.textContent = '';
  statusEl.textContent = '';
  statusEl.className = 'status';
  logWrap.style.display = 'block';
  appendLog(`OZON: запрос отправлен ${{df}}..${{dt}}`);
  appendLog(adsOnly ? 'Режим: только реклама.' : withAds ? 'Режим: полный синк с рекламой.' : 'Режим: быстрый синк без рекламы.');
  statusEl.textContent = 'Выполняется...';

  const url = `/ozon/stream?date_from=${{df}}&date_to=${{dt}}&mode=${{encodeURIComponent(mode)}}&skip_ads=${{withAds || adsOnly ? '0' : '1'}}`;
  const resp = await fetch(url);
  if (!resp.ok || !resp.body) {{
    statusEl.textContent = `❌ Не удалось открыть поток логов (HTTP ${{resp.status}})`;
    statusEl.className = 'status err';
    syncButtons.forEach(button => button.disabled = false);
    if (activeButton) activeButton.textContent = oldText;
    return;
  }}
  const reader = resp.body.getReader();
  const decoder = new TextDecoder();
  let buf = '';
  while (true) {{
    const {{ done, value }} = await reader.read();
    if (done) break;
    buf += decoder.decode(value, {{ stream: true }});
    const lines = buf.split('\\n');
    buf = lines.pop();
    for (const line of lines) {{
      if (!line.startsWith('data: ')) continue;
      const payload = JSON.parse(line.slice(6));
      if (payload.type === 'log') {{
        appendLog(payload.text);
      }} else if (payload.type === 'heartbeat') {{
        statusEl.textContent = payload.text;
      }} else if (payload.type === 'done') {{
        statusEl.textContent = payload.text;
        statusEl.className = 'status ok';
      }} else if (payload.type === 'error') {{
        statusEl.textContent = payload.text;
        statusEl.className = 'status err';
      }}
    }}
  }}
  syncButtons.forEach(button => button.disabled = false);
  if (activeButton) activeButton.textContent = oldText;
}}

form.addEventListener('submit', e => {{
  e.preventDefault();
}});
for (const button of loadButtons) {{
  button.addEventListener('click', () => startLoad(button));
}}
for (const button of ozonButtons) {{
  button.addEventListener('click', () => startOzonSync(button.dataset.ozonSync || 'fast'));
}}
if (funnelUploadBtn) funnelUploadBtn.addEventListener('click', () => uploadFunnel(false));
if (funnelForceBtn) funnelForceBtn.addEventListener('click', () => uploadFunnel(true));

</script>
</body>
</html>
"""

ANALYTICS_HOME_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>WB Analytics</title>
  <style>
    :root {{
      --bg: #f7f6f2;
      --card: #ffffff;
      --ink: #1f2937;
      --muted: #6b7280;
      --accent: #0f766e;
      --accent-2: #0ea5e9;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: radial-gradient(circle at 20% 20%, #d1fae5 0, transparent 42%), radial-gradient(circle at 80% 0%, #bae6fd 0, transparent 35%), var(--bg); }}
    .wrap {{ max-width: 980px; margin: 0 auto; padding: 24px 16px 40px; }}
    .top {{ display: flex; justify-content: space-between; gap: 12px; align-items: center; margin-bottom: 20px; }}
    .top a {{ color: var(--accent); text-decoration: none; font-weight: 700; }}
    h1 {{ margin: 0; font-size: 1.7rem; }}
    .sub {{ color: var(--muted); margin: 8px 0 0; }}
    .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 14px; margin-top: 18px; }}
    .card {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 14px; padding: 18px; box-shadow: 0 8px 24px rgba(15, 118, 110, .08); }}
    .card h2 {{ margin: 0 0 8px; font-size: 1.1rem; }}
    .card p {{ margin: 0 0 14px; color: var(--muted); line-height: 1.4; }}
    .btn {{ display: inline-block; padding: 10px 14px; border-radius: 10px; color: #fff; text-decoration: none; font-weight: 700; background: linear-gradient(135deg, var(--accent), var(--accent-2)); }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <div>
        <h1>Аналитика Wildberries</h1>
        <p class="sub">Пользовательские страницы только для аналитических таблиц.</p>
      </div>
      <a href="/">← К загрузке</a>
    </div>
    <div class="grid">
      <div class="card">
        <h2>Выкупы по дням</h2>
        <p>Pivot-таблица: метрики в строках, даты в колонках, факт выкупов и выкупы заказов выбранной даты.</p>
        <a class="btn" href="/analytics/buyout-order-day">Открыть дни</a>
      </div>
      <div class="card">
        <h2>Выкупы по неделям</h2>
        <p>Та же рабочая таблица, собранная недельными колонками для более спокойного анализа периода.</p>
        <a class="btn" href="/analytics/buyout-order-week">Открыть недели</a>
      </div>
      <div class="card">
        <h2>Планирование</h2>
        <p>Остатки, скорость, прогнозы, экономика и ручные параметры для планирования поставок и цен.</p>
        <a class="btn" href="/analytics/planning">Открыть план</a>
      </div>
      <div class="card">
        <h2>Комментарии</h2>
        <p>Рабочие пометки менеджера по товарам, типам решений и истории комментариев.</p>
        <a class="btn" href="/analytics/comments">Открыть комментарии</a>
      </div>
    </div>
  </div>
</body>
</html>
"""

BUYOUT_ORDER_DAY_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{page_title}</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0f766e; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #e0f2fe, transparent 240px), var(--bg); }}
    .wrap {{ width: calc(100% - 32px); max-width: 3000px; margin: 0 auto; padding: 10px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .76rem; margin-bottom: 3px; font-weight: 700; }}
    input, select {{ padding: 7px 8px; border: 1px solid #cbd5e1; border-radius: 7px; font-size: 12px; min-width: 160px; background: #fff; }}
    button {{ padding: 8px 12px; border: 0; border-radius: 7px; background: var(--accent); color: #fff; font-weight: 700; cursor: pointer; font-size: 12px; }}
    .quick-range {{ display: flex; gap: 6px; align-items: center; flex-wrap: wrap; }}
    .quick-range button {{ padding: 7px 10px; background: #e6f4f1; color: var(--accent); }}
    .meta {{ margin-top: 8px; color: var(--muted); font-size: .78rem; }}
    .workspace {{ --articles-width: 332px; display: grid; grid-template-columns: var(--articles-width) minmax(0, 1fr); gap: 12px; align-items: start; margin-top: 10px; }}
    .workspace.articles-collapsed {{ --articles-width: 44px; }}
    .main-pane {{ display: flex; flex-direction: column; gap: 12px; min-width: 0; }}
    .tbl {{ --graph-col-width: 58px; --metric-col-width: 240px; --total-col-width: 120px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow-x: auto; overflow-y: clip; -webkit-overflow-scrolling: touch; }}
    .articles {{ position: sticky; top: 8px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: hidden; max-height: calc(100vh - 18px); display: flex; flex-direction: column; min-width: 0; transition: width .18s ease, box-shadow .18s ease, border-color .18s ease; }}
    .articles-inner {{ display: flex; flex-direction: column; min-height: 0; height: 100%; background: #fff; }}
    .articles-head {{ padding: 7px 8px; border-bottom: 1px solid #eef2f7; display: flex; justify-content: space-between; gap: 8px; align-items: center; }}
    .articles-title {{ font-weight: 800; font-size: .78rem; }}
    .articles-actions {{ display: flex; gap: 8px; }}
    .articles-actions button {{ padding: 4px 6px; border-radius: 6px; font-size: .68rem; background: #e6f4f1; color: var(--accent); }}
    .articles-mini {{ display: none; align-items: center; justify-content: center; gap: 6px; height: 100%; min-height: 260px; padding: 8px 0; writing-mode: vertical-rl; transform: rotate(180deg); color: var(--accent); font-weight: 800; font-size: .72rem; letter-spacing: .08em; background: linear-gradient(180deg, #f0fdfa, #ecfeff); }}
    .articles-mini button {{ writing-mode: horizontal-tb; transform: rotate(180deg); margin: 0; width: 24px; height: 24px; padding: 0; border-radius: 999px; background: #0f766e; color: #fff; font-size: .9rem; line-height: 1; }}
    .subject-filter {{ padding: 6px 8px; border-bottom: 1px solid #eef2f7; }}
    .subject-filter label {{ font-size: .68rem; margin-bottom: 2px; }}
    .subject-filter select {{ width: 100%; min-width: 0; padding: 5px 6px; font-size: .72rem; }}
    .article-list {{ overflow: auto; user-select: none; }}
    .article-table {{ width: 100%; min-width: 0; border-collapse: collapse; table-layout: fixed; font-size: 10px; }}
    .article-table th, .article-table td {{ border-bottom: 1px solid #eef2f7; padding: 3px; line-height: 1.15; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
    .article-table th {{ position: sticky; top: 0; z-index: 1; background: #f8fafc; color: #475569; font-size: .62rem; cursor: pointer; }}
    .article-table th:first-child, .article-table td:first-child {{ text-align: left; width: 38%; }}
    .article-table th:nth-child(2), .article-table td:nth-child(2) {{ text-align: right; width: 12%; }}
    .article-table th:nth-child(3), .article-table td:nth-child(3) {{ text-align: right; width: 12%; }}
    .article-table th:nth-child(4), .article-table td:nth-child(4) {{ text-align: right; width: 12%; }}
    .article-table th:nth-child(5), .article-table td:nth-child(5) {{ text-align: right; width: 26%; }}
    .article-table tr {{ cursor: default; }}
    .article-table tbody tr:hover {{ background: #f0fdfa; }}
    .article-table tbody tr.selected {{ background: #ccfbf1; font-weight: 800; color: #115e59; }}
    .article-link {{ color: var(--accent); text-decoration: none; font-weight: 800; margin-left: 4px; }}
    .campaigns-list-link {{ color: #7c3aed; text-decoration: none; font-weight: 900; margin-left: 4px; }}
    .article-link:hover, .campaigns-list-link:hover {{ text-decoration: underline; }}
    .article-copy {{ background:none;border:none;cursor:pointer;padding:0 1px;margin-left:3px;
                     font-size:.8em;color:#94a3b8;opacity:.7;line-height:1;vertical-align:middle; }}
    .article-copy:hover {{ color:var(--accent);opacity:1; }}
    .article-copy.copied {{ color:#16a34a;opacity:1; }}
    .tbl table {{ width: max-content; min-width: 100%; border-collapse: collapse; font-size: 11px; }}
    .tbl th, .tbl td {{ border-bottom: 1px solid #eef2f7; padding: 6px 7px; text-align: right; white-space: nowrap; }}
    .tbl th:first-child, .tbl td:first-child {{ position: sticky; left: 0; z-index: 1; text-align: center; background: #fff; width: var(--graph-col-width); min-width: var(--graph-col-width); max-width: var(--graph-col-width); padding-left: 0; padding-right: 0; }}
    .tbl td:first-child {{ vertical-align: middle; }}
    .tbl th:nth-child(2), .tbl td:nth-child(2) {{ position: sticky; left: var(--graph-col-width); z-index: 1; text-align: left; background: #fff; width: var(--metric-col-width); min-width: var(--metric-col-width); max-width: var(--metric-col-width); font-weight: 700; }}
    .tbl th:nth-child(3), .tbl td:nth-child(3) {{ position: sticky; left: calc(var(--graph-col-width) + var(--metric-col-width)); z-index: 1; background: #fff; width: var(--total-col-width); min-width: var(--total-col-width); max-width: var(--total-col-width); font-weight: 800; }}
    .tbl th {{ position: sticky; top: 0; z-index: 2; background: #f8fafc; }}
    .tbl th.weekend-date {{ color: #7f1d1d; }}
    .tbl th:first-child, .tbl th:nth-child(2), .tbl th:nth-child(3) {{ z-index: 3; background: #f8fafc; }}
    .tbl tr.group-row td {{ background: #eef6f4; color: #0f766e; font-size: .78rem; font-weight: 800; text-transform: uppercase; letter-spacing: .04em; }}
    .tbl tr.group-row td:first-child {{ background: #eef6f4; }}
    .tbl tr.group-row td:nth-child(2) {{ background: #eef6f4; }}
    .tbl tr.group-row td:nth-child(3) {{ background: #eef6f4; }}
    .group-toggle {{ width: 22px; height: 22px; padding: 0; margin: 0; border-radius: 999px; background: #0f766e; color: #fff; font-size: .92rem; line-height: 1; }}
    .group-label {{ cursor: pointer; }}
    .metric-toggle {{ display: inline-block; width: 16px; height: 16px; min-width: 16px; padding: 0; margin: 0; border: 0; border-radius: 0; background: transparent; box-shadow: none; accent-color: var(--accent); cursor: pointer; vertical-align: middle; appearance: auto; -webkit-appearance: checkbox; }}
    .chart-card {{ background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; }}
    .chart-head {{ display: flex; justify-content: space-between; align-items: center; gap: 10px; margin-bottom: 10px; }}
    .chart-title {{ font-size: .95rem; font-weight: 800; }}
    .chart-subtitle {{ color: var(--muted); font-size: .74rem; }}
    .chart-legend {{ display: flex; flex-wrap: wrap; gap: 8px 14px; margin: 0 0 10px; }}
    .chart-legend-item {{ display: inline-flex; align-items: center; gap: 7px; padding: 5px 9px; border: 1px solid #dbeafe; border-radius: 999px; background: #f8fafc; color: #0f172a; font-size: .78rem; font-weight: 600; }}
    .chart-swatch {{ width: 10px; height: 10px; border-radius: 999px; flex: 0 0 auto; }}
    .chart-empty {{ color: var(--muted); font-size: .78rem; padding: 14px 0 4px; }}
    .chart-svg-wrap {{ position: relative; width: 100%; overflow-x: auto; border: 1px solid #eef2f7; border-radius: 10px; background: linear-gradient(180deg, #fcfffe, #f8fafc); }}
    .chart-svg-wrap[data-fit-width="1"] {{ overflow-x: hidden; }}
    .chart-svg {{ display: block; min-width: var(--chart-min-width, 980px); width: 100%; height: var(--chart-height, 860px); }}
    .chart-svg-wrap[data-fit-width="0"] .chart-svg {{ width: var(--chart-min-width, 980px); height: auto; aspect-ratio: var(--chart-aspect-ratio, 980 / 860); }}
    .chart-svg-wrap[data-fit-width="1"] .chart-svg {{ min-width: 0; height: auto; aspect-ratio: var(--chart-aspect-ratio, 980 / 860); }}
    .chart-axis {{ stroke: #cbd5e1; stroke-width: 1; }}
    .chart-grid {{ stroke: #e5e7eb; stroke-width: 1; stroke-dasharray: 3 4; }}
    .chart-label {{ fill: #64748b; font-size: var(--chart-font-size, 11px); }}
    .chart-zone-label {{ fill: #0f766e; font-size: var(--chart-zone-font-size, 12px); font-weight: 700; }}
    .chart-line {{ fill: none; stroke-width: var(--chart-line-width, 2.5); stroke-linecap: round; stroke-linejoin: round; }}
    .chart-dot {{ stroke: #fff; stroke-width: var(--chart-dot-stroke, 1.5); }}
    .chart-hover-line {{ stroke: #0f766e; stroke-width: 1.5; stroke-dasharray: 4 4; opacity: .55; }}
    .chart-hit {{ fill: transparent; cursor: crosshair; }}
    .chart-tooltip {{ position: absolute; display: none; min-width: 180px; max-width: 280px; padding: 8px 10px; border: 1px solid #99f6e4; border-radius: 10px; background: rgba(255, 255, 255, .96); box-shadow: 0 14px 28px rgba(15, 118, 110, .16); color: #0f172a; font-size: 12px; line-height: 1.35; pointer-events: none; }}
    .chart-tooltip-date {{ font-weight: 800; margin-bottom: 6px; color: #0f766e; }}
    .chart-tooltip-row {{ display: flex; align-items: center; justify-content: space-between; gap: 10px; margin-top: 4px; }}
    .chart-tooltip-name {{ display: inline-flex; align-items: center; gap: 6px; min-width: 0; }}
    .chart-tooltip-name span:last-child {{ overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
    .chart-tooltip-dot {{ width: 8px; height: 8px; border-radius: 999px; flex: 0 0 auto; }}
    .chart-tooltip-value {{ font-weight: 700; text-align: right; white-space: nowrap; }}
    .workspace.articles-collapsed .articles {{ overflow: visible; }}
    .workspace.articles-collapsed .articles .articles-inner {{ display: none; }}
    .workspace.articles-collapsed .articles .articles-mini {{ display: flex; }}
    .workspace.articles-collapsed .articles:hover,
    .workspace.articles-collapsed .articles:focus-within {{ width: 332px; z-index: 20; box-shadow: 0 18px 40px rgba(15, 118, 110, .18); border-color: #99f6e4; }}
    .workspace.articles-collapsed .articles:hover .articles-inner,
    .workspace.articles-collapsed .articles:focus-within .articles-inner {{ display: flex; }}
    .workspace.articles-collapsed .articles:hover .articles-mini,
    .workspace.articles-collapsed .articles:focus-within .articles-mini {{ display: none; }}
    @media (max-width: 1100px) {{
      .workspace {{ grid-template-columns: 1fr; }}
      .workspace.articles-collapsed {{ --articles-width: 1fr; }}
      .articles {{ order: 2; }}
      .articles {{ max-height: 260px; }}
      .workspace.articles-collapsed .articles {{ width: auto; overflow: hidden; }}
      .workspace.articles-collapsed .articles .articles-inner {{ display: flex; }}
      .workspace.articles-collapsed .articles .articles-mini {{ display: none; }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>{page_title}</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="article">Артикул</label>
        <input id="article" type="text" placeholder="например 12-0042-001">
      </div>
      <div>
        <label for="df">Период с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Период по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div class="quick-range">
        <button type="button" id="range-7">7 дней</button>
        <button type="button" id="range-14">14 дней</button>
        <button type="button" id="range-31">Месяц</button>
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="workspace" id="workspace">
      <aside class="articles">
        <div class="articles-inner">
          <div class="articles-head">
            <div class="articles-title">Артикулы</div>
            <div class="articles-actions">
              <button type="button" id="toggle-articles" title="Свернуть список">◂</button>
              <button type="button" id="select-all">Все</button>
              <button type="button" id="clear-all">Сброс</button>
            </div>
          </div>
          <div class="subject-filter">
            <label for="subject">Предмет</label>
            <select id="subject">
              <option value="">Все предметы</option>
            </select>
          </div>
          <div class="subject-filter">
            <label for="strategy">Ярлыки</label>
            <select id="strategy" multiple size="4" style="height:72px;" title="Ctrl/Cmd для множественного выбора"></select>
          </div>
          <div class="subject-filter">
            <label for="revenue-category">Категория по выручке</label>
            <select id="revenue-category">
              <option value="">Все категории</option>
            </select>
          </div>
          <div class="article-list" id="article-list"></div>
        </div>
        <div class="articles-mini">
          <button type="button" id="expand-articles" title="Развернуть список">▸</button>
          <span>Артикулы</span>
        </div>
      </aside>
      <div class="main-pane">
        <div class="tbl" id="tbl"></div>
        <section class="chart-card">
          <div class="chart-head">
            <div>
              <div class="chart-title">График метрик</div>
              <div class="chart-subtitle">Суммы и проценты на одном графике. Проценты вынесены ниже и масштабируются отдельно.</div>
            </div>
          </div>
          <div class="chart-legend" id="chart-legend"></div>
          <div class="chart-svg-wrap" id="chart-wrap"></div>
        </section>
      </div>
    </div>
  </div>
<script>
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const articleList = document.getElementById('article-list');
const subjectEl = document.getElementById('subject');
const strategyEl = document.getElementById('strategy');
const revenueCategoryEl = document.getElementById('revenue-category');
const chartLegend = document.getElementById('chart-legend');
const chartWrap = document.getElementById('chart-wrap');
const workspaceEl = document.getElementById('workspace');
const reportGranularity = '{granularity}';
const filterKey = `wb.analytics.buyoutOrderDay.${{reportGranularity}}.filters`;
let requestSeq = 0;
let filterTimer = null;
let selectedArticles = new Set();
let articlesLoadedFor = '';
let subjectsLoadedFor = '';
function fillMultiSelect(select, values) {{
  const prevVals = new Set([...select.selectedOptions].map(o=>o.value));
  const optionsHtml = (values || []).map(v => {{
    const sel = prevVals.has(v) ? ' selected' : '';
    return `<option value="${{String(v).replaceAll('"','&quot;')}}"${{sel}}>${{v}}</option>`;
  }}).join('');
  select.innerHTML = select.multiple ? optionsHtml : `<option value="">Все ярлыки</option>${{optionsHtml}}`;
}}
let articleItems = [];
let articleSort = {{ key: 'revenue', dir: 'desc' }};
let lastArticleIndex = -1;
let activeChartMetrics = new Set();
let latestChartPayload = null;
let articlesCollapsed = false;
let collapsedGroups = new Set();
let chartMetricsInitialized = false;
const chartPalette = ['#0f766e', '#f97316', '#0284c7', '#dc2626', '#7c3aed', '#16a34a', '#ca8a04', '#db2777', '#0891b2', '#4f46e5'];

function syncArticlesPanel() {{
  workspaceEl.classList.toggle('articles-collapsed', articlesCollapsed);
  const toggleBtn = document.getElementById('toggle-articles');
  const expandBtn = document.getElementById('expand-articles');
  if (toggleBtn) {{
    toggleBtn.textContent = articlesCollapsed ? '▸' : '◂';
    toggleBtn.title = articlesCollapsed ? 'Развернуть список' : 'Свернуть список';
  }}
  if (expandBtn) expandBtn.title = articlesCollapsed ? 'Развернуть список' : 'Свернуть список';
}}

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.article !== undefined) document.getElementById('article').value = saved.article;
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
    if (saved.subject !== undefined) subjectEl.dataset.pendingValue = saved.subject;
    // labels not restored from storage - always start empty
    if (saved.revenueCategory !== undefined) revenueCategoryEl.dataset.pendingValue = saved.revenueCategory;
    if (Array.isArray(saved.selectedArticles)) selectedArticles = new Set(saved.selectedArticles.map(String).filter(value => /^\d+$/.test(value)));
    if (Array.isArray(saved.chartMetrics)) {{
      activeChartMetrics = new Set(saved.chartMetrics);
      chartMetricsInitialized = true;
    }}
    if (Array.isArray(saved.collapsedGroups)) collapsedGroups = new Set(saved.collapsedGroups.map(String).filter(Boolean));
    articlesCollapsed = Boolean(saved.articlesCollapsed);
  }} catch (err) {{}}
  syncArticlesPanel();
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    article: document.getElementById('article').value,
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value,
    subject: subjectEl.value,
    labels: [...strategyEl.selectedOptions].map(o=>o.value).join(","),
    revenueCategory: revenueCategoryEl.value,
    selectedArticles: [...selectedArticles],
    chartMetrics: [...activeChartMetrics],
    collapsedGroups: [...collapsedGroups],
    articlesCollapsed
  }}));
}}

function isoDate(date) {{
  return date.toISOString().slice(0, 10);
}}

function applyQuickRange(days) {{
  const end = new Date();
  const start = new Date(end);
  start.setDate(start.getDate() - (days - 1));
  document.getElementById('df').value = isoDate(start);
  document.getElementById('dt').value = isoDate(end);
  loadData();
}}

function scheduleLoad() {{
  clearTimeout(filterTimer);
  filterTimer = setTimeout(loadData, 350);
}}

function resetArticleFilters() {{
  subjectEl.value = '';
  [...strategyEl.options].forEach(o=>o.selected=false);
  revenueCategoryEl.value = '';
  delete subjectEl.dataset.pendingValue;
  
  delete revenueCategoryEl.dataset.pendingValue;
}}

function escapeHtml(value) {{
  return String(value)
    .replaceAll('&', '&amp;')
    .replaceAll('<', '&lt;')
    .replaceAll('>', '&gt;')
    .replaceAll('"', '&quot;')
    .replaceAll("'", '&#39;');
}}

function formatCompactNumber(value, fractionDigits = 0) {{
  if (value === null || value === undefined || value === '') return '';
  const numeric = Number(value);
  if (!Number.isFinite(numeric)) return '';
  return numeric.toLocaleString('ru-RU', {{
    minimumFractionDigits: fractionDigits,
    maximumFractionDigits: fractionDigits
  }});
}}

function parseMetricValue(value) {{
  const text = String(value ?? '').trim();
  if (!text) return 0;
  const normalized = text.replaceAll(' ', '').replace('%', '').replace(',', '.');
  const numeric = Number(normalized);
  return Number.isFinite(numeric) ? numeric : 0;
}}

function chartRowsFromPayload(data) {{
  return (data.rows || []).filter(row => String(row.metric || '').trim());
}}

function ensureDefaultChartMetrics(rows) {{
  if (chartMetricsInitialized || activeChartMetrics.size) return;
  const defaults = [
    'Сумма заказов',
    'Сумма выкупов в эту дату',
    'Реклама',
    'ДРР от заказов',
    'Маржинальность от выкупов за дату',
  ];
  const available = new Set(rows.map(row => row.metric));
  for (const metric of defaults) {{
    if (available.has(metric)) activeChartMetrics.add(metric);
  }}
  if (!activeChartMetrics.size) {{
    for (const row of rows.slice(0, 4)) activeChartMetrics.add(row.metric);
  }}
  chartMetricsInitialized = true;
}}

function renderChart(data) {{
  latestChartPayload = data;
  const rows = chartRowsFromPayload(data);
  ensureDefaultChartMetrics(rows);
  const selectedRows = rows.filter(row => activeChartMetrics.has(row.metric));
  const dates = data.dates || [];
  if (!dates.length || !selectedRows.length) {{
    chartLegend.innerHTML = '';
    chartWrap.dataset.fitWidth = '0';
    chartWrap.innerHTML = '<div class="chart-empty">Выберите хотя бы одну метрику с данными.</div>';
    return;
  }}

  const isWeekChart = reportGranularity === 'week';
  const chartConfig = isWeekChart
    ? {{
        minWidth: Math.max(840, dates.length * 150 + 260),
        height: 760,
        left: 74,
        right: 26,
        top: 30,
        bottom: 112,
        gap: 44,
        moneyHeight: 390,
        percentHeight: 184,
        fontSize: 12,
        zoneFontSize: 13,
        lineWidth: 3.2,
        dotStroke: 1.8,
        dotRadius: 4.8,
      }}
    : {{
        minWidth: Math.max(980, dates.length * 84 + 220),
        height: 860,
        left: 64,
        right: 18,
        top: 24,
        bottom: 110,
        gap: 52,
        moneyHeight: 450,
        percentHeight: 220,
        fontSize: 11,
        zoneFontSize: 12,
        lineWidth: 2.5,
        dotStroke: 1.5,
        dotRadius: 3.5,
      }};
  const percentRows = selectedRows.filter(row => String(row.kind || '').startsWith('percent'));
  const moneyRows = selectedRows.filter(row => !String(row.kind || '').startsWith('percent'));
  const moneySeries = moneyRows.map(row => row.values.map(parseMetricValue));
  const percentSeries = percentRows.map(row => row.values.map(parseMetricValue));
  const moneyMax = Math.max(0, ...moneySeries.flat());
  const percentMax = Math.max(0, ...percentSeries.flat());

  const width = chartConfig.minWidth;
  const height = chartConfig.height;
  const preserveAspectRatio = 'xMidYMid meet';
  // Явно задаем размер SVG, чтобы контейнер не растягивал недельный график.
  const svgInlineStyle = isWeekChart
    ? `width:${{width}}px;height:auto;display:block`
    : 'width:100%;height:auto;display:block';
  const left = chartConfig.left;
  const right = chartConfig.right;
  const top = chartConfig.top;
  const bottom = chartConfig.bottom;
  const gap = chartConfig.gap;
  const plotWidth = width - left - right;
  const moneyHeight = chartConfig.moneyHeight;
  const percentHeight = chartConfig.percentHeight;
  const moneyTop = top;
  const percentTop = moneyTop + moneyHeight + gap;
  const moneyBottom = moneyTop + moneyHeight;
  const percentBottom = percentTop + percentHeight;
  const xStep = dates.length > 1 ? plotWidth / (dates.length - 1) : 0;
  const moneyScale = moneyMax > 0 ? moneyHeight / moneyMax : 1;
  const percentScale = percentMax > 0 ? percentHeight / percentMax : 1;
  const formatDate = (iso) => {{
    if (String(iso).includes('..')) {{
      return String(iso).split('..').map(part => {{
        const [yy, mm, dd] = part.split('-');
        return `${{dd}}.${{mm}}`;
      }}).join('-');
    }}
    const [y, m, d] = String(iso).split('-');
    return `${{d}}.${{m}}`;
  }};
  const formatTooltipDate = (iso) => {{
    if (String(iso).includes('..')) {{
      return String(iso).split('..').map(part => {{
        const [y, m, d] = part.split('-');
        return `${{d}}.${{m}}.${{y}}`;
      }}).join(' - ');
    }}
    const [y, m, d] = String(iso).split('-');
    return `${{d}}.${{m}}.${{y}}`;
  }};

  const gridLines = [];
  for (let idx = 0; idx <= 4; idx++) {{
    const moneyY = moneyTop + (moneyHeight / 4) * idx;
    const percentY = percentTop + (percentHeight / 4) * idx;
    const moneyValue = moneyMax * (1 - idx / 4);
    const percentValue = percentMax * (1 - idx / 4);
    gridLines.push(`<line class="chart-grid" x1="${{left}}" y1="${{moneyY}}" x2="${{width - right}}" y2="${{moneyY}}"></line>`);
    gridLines.push(`<text class="chart-label" x="${{left - 8}}" y="${{moneyY + 4}}" text-anchor="end">${{formatCompactNumber(moneyValue, 0)}}</text>`);
    gridLines.push(`<line class="chart-grid" x1="${{left}}" y1="${{percentY}}" x2="${{width - right}}" y2="${{percentY}}"></line>`);
    gridLines.push(`<text class="chart-label" x="${{left - 8}}" y="${{percentY + 4}}" text-anchor="end">${{formatCompactNumber(percentValue, 0)}}%</text>`);
  }}

  const xLabels = dates.map((day, idx) => {{
    const x = left + xStep * idx;
    return `
      <line class="chart-grid" x1="${{x}}" y1="${{moneyTop}}" x2="${{x}}" y2="${{percentBottom}}"></line>
      <text class="chart-label" x="${{x}}" y="${{height - 18}}" text-anchor="end" transform="rotate(-45 ${{x}} ${{height - 18}})">${{formatDate(day)}}</text>
    `;
  }}).join('');
  const allSeries = selectedRows.map((row, idx) => ({{
    row,
    color: chartPalette[idx % chartPalette.length],
  }}));
  const legendHtml = allSeries.map(series => `
    <div class="chart-legend-item">
      <span class="chart-swatch" style="background:${{series.color}}"></span>
      <span>${{escapeHtml(series.row.metric)}}</span>
    </div>
  `).join('');

  const seriesSvg = selectedRows.map((row, idx) => {{
    const isPercent = String(row.kind || '').startsWith('percent');
    const values = row.values.map(parseMetricValue);
    const color = chartPalette[idx % chartPalette.length];
    const zoneTop = isPercent ? percentTop : moneyTop;
    const zoneBottom = isPercent ? percentBottom : moneyBottom;
    const scale = isPercent ? percentScale : moneyScale;
    const points = values.map((value, pointIdx) => {{
      const x = left + xStep * pointIdx;
      const y = zoneBottom - value * scale;
      return `${{x}},${{y}}`;
    }}).join(' ');
    const dots = values.map((value, pointIdx) => {{
      const x = left + xStep * pointIdx;
      const y = zoneBottom - value * scale;
      return `<circle class="chart-dot" cx="${{x}}" cy="${{y}}" r="${{chartConfig.dotRadius}}" fill="${{color}}"><title>${{escapeHtml(row.metric)}}: ${{row.values[pointIdx]}}</title></circle>`;
    }}).join('');
    const labelX = width - right - 8;
    const lastValue = values[values.length - 1] || 0;
    const labelY = zoneBottom - lastValue * scale - 6;
    return `
      <polyline class="chart-line" points="${{points}}" stroke="${{color}}"></polyline>
      ${{dots}}
      <text class="chart-label" x="${{labelX}}" y="${{Math.max(zoneTop + 12, Math.min(zoneBottom - 6, labelY))}}" text-anchor="end" fill="${{color}}">${{escapeHtml(row.metric)}}</text>
    `;
  }}).join('');

  const hoverTargets = dates.map((day, idx) => {{
    const x = left + xStep * idx;
    const rectWidth = dates.length > 1 ? Math.max(24, xStep) : plotWidth;
    const rectX = dates.length > 1 ? x - rectWidth / 2 : left;
    return `<rect class="chart-hit" data-index="${{idx}}" x="${{rectX}}" y="${{moneyTop}}" width="${{rectWidth}}" height="${{percentBottom - moneyTop}}"></rect>`;
  }}).join('');

  chartWrap.style.setProperty('--chart-min-width', `${{width}}px`);
  chartWrap.style.setProperty('--chart-height', `${{height}}px`);
  chartWrap.style.setProperty('--chart-aspect-ratio', `${{width}} / ${{height}}`);
  chartWrap.style.setProperty('--chart-font-size', `${{chartConfig.fontSize}}px`);
  chartWrap.style.setProperty('--chart-zone-font-size', `${{chartConfig.zoneFontSize}}px`);
  chartWrap.style.setProperty('--chart-line-width', `${{chartConfig.lineWidth}}`);
  chartWrap.style.setProperty('--chart-dot-stroke', `${{chartConfig.dotStroke}}`);
  chartWrap.dataset.fitWidth = isWeekChart ? '0' : '1';
  chartLegend.innerHTML = legendHtml;
  chartWrap.innerHTML = `
    <div class="chart-tooltip" id="chart-tooltip"></div>
    <svg class="chart-svg" style="${{svgInlineStyle}}" viewBox="0 0 ${{width}} ${{height}}" preserveAspectRatio="${{preserveAspectRatio}}">
      <text class="chart-zone-label" x="${{left}}" y="${{moneyTop - 6}}">Суммы</text>
      <text class="chart-zone-label" x="${{left}}" y="${{percentTop - 6}}">Проценты</text>
      <line class="chart-axis" x1="${{left}}" y1="${{moneyTop}}" x2="${{left}}" y2="${{moneyBottom}}"></line>
      <line class="chart-axis" x1="${{left}}" y1="${{moneyBottom}}" x2="${{width - right}}" y2="${{moneyBottom}}"></line>
      <line class="chart-axis" x1="${{left}}" y1="${{percentTop}}" x2="${{left}}" y2="${{percentBottom}}"></line>
      <line class="chart-axis" x1="${{left}}" y1="${{percentBottom}}" x2="${{width - right}}" y2="${{percentBottom}}"></line>
      ${{gridLines.join('')}}
      ${{xLabels}}
      ${{seriesSvg}}
      <line class="chart-hover-line" id="chart-hover-line" x1="${{left}}" y1="${{moneyTop}}" x2="${{left}}" y2="${{percentBottom}}" visibility="hidden"></line>
      ${{hoverTargets}}
    </svg>
  `;

  const svgEl = chartWrap.querySelector('.chart-svg');
  const tooltipEl = document.getElementById('chart-tooltip');
  const hoverLineEl = document.getElementById('chart-hover-line');
  const showTooltip = (index, clientX, clientY) => {{
    const x = left + xStep * index;
    hoverLineEl.setAttribute('x1', String(x));
    hoverLineEl.setAttribute('x2', String(x));
    hoverLineEl.setAttribute('visibility', 'visible');
    tooltipEl.innerHTML = `
      <div class="chart-tooltip-date">${{formatTooltipDate(dates[index])}}</div>
      ${{allSeries.map(series => `
        <div class="chart-tooltip-row">
          <div class="chart-tooltip-name">
            <span class="chart-tooltip-dot" style="background:${{series.color}}"></span>
            <span>${{escapeHtml(series.row.metric)}}</span>
          </div>
          <div class="chart-tooltip-value">${{escapeHtml(String((series.row.values || [])[index] || '0'))}}</div>
        </div>
      `).join('')}}
    `;
    tooltipEl.style.display = 'block';
    const wrapRect = chartWrap.getBoundingClientRect();
    const svgRect = svgEl.getBoundingClientRect();
    const tooltipWidth = tooltipEl.offsetWidth;
    const tooltipHeight = tooltipEl.offsetHeight;
    let leftPos = clientX - wrapRect.left + 14;
    let topPos = clientY - wrapRect.top - tooltipHeight - 14;
    if (leftPos + tooltipWidth > wrapRect.width - 8) leftPos = clientX - wrapRect.left - tooltipWidth - 14;
    if (leftPos < 8) leftPos = 8;
    if (topPos < 8) topPos = clientY - wrapRect.top + 14;
    if (topPos + tooltipHeight > wrapRect.height - 8) topPos = Math.max(8, wrapRect.height - tooltipHeight - 8);
    tooltipEl.style.left = `${{leftPos + chartWrap.scrollLeft}}px`;
    tooltipEl.style.top = `${{topPos + chartWrap.scrollTop}}px`;
  }};
  const hideTooltip = () => {{
    hoverLineEl.setAttribute('visibility', 'hidden');
    tooltipEl.style.display = 'none';
  }};
  for (const target of [...chartWrap.querySelectorAll('.chart-hit')]) {{
    target.addEventListener('mouseenter', (event) => showTooltip(Number(target.dataset.index || 0), event.clientX, event.clientY));
    target.addEventListener('mousemove', (event) => showTooltip(Number(target.dataset.index || 0), event.clientX, event.clientY));
    target.addEventListener('mouseleave', hideTooltip);
  }}
}}

function cellColorStyle(kind, prevRaw, currRaw, row) {{
  if (prevRaw === null || prevRaw === undefined || currRaw === null || currRaw === undefined) return '';
  if (kind === 'stock') {{
    const diff = currRaw - prevRaw;
    const absThreshold = Number(row?.color_threshold ?? 20);
    if (diff < absThreshold) return '';
    return 'background:#dcfce7;color:#047857;font-weight:900;box-shadow:inset 0 0 0 1px #86efac';
  }}
  if (String(row?.metric || '') === 'СПП') {{
    const diff = currRaw - prevRaw;
    const threshold = Number(row?.color_threshold ?? 1);
    // Для СПП подсветка должна включаться уже при изменении ровно на порог.
    if (Math.abs(diff) < threshold) return '';
    return diff > 0
      ? 'color:#047857;font-weight:900;text-shadow:0 0 0 #047857'
      : 'color:#b91c1c;font-weight:900;text-shadow:0 0 0 #b91c1c';
  }}
  const isPercent = kind === 'percent_income' || kind === 'percent_expense';
  const threshold = Number(row?.color_threshold ?? (isPercent ? 5 : 10));
  const colorMode = row?.color_mode || (isPercent ? 'absolute' : 'percent');
  let beneficial;
  if (isPercent) {{
    const diff = currRaw - prevRaw;
    if (Math.abs(diff) <= threshold) return '';
    beneficial = kind === 'percent_income' ? diff > 0 : diff < 0;
  }} else {{
    const diff = currRaw - prevRaw;
    if (colorMode === 'absolute') {{
      if (Math.abs(diff) <= threshold) return '';
      beneficial = kind === 'income' ? diff > 0 : diff < 0;
    }} else {{
      if (Math.abs(prevRaw) < 1) return '';
      const pctChange = diff / Math.abs(prevRaw) * 100;
      if (Math.abs(pctChange) <= threshold) return '';
      beneficial = kind === 'income' ? pctChange > 0 : pctChange < 0;
    }}
  }}
  return beneficial ? 'color:#1a8c40;font-weight:600' : 'color:#c0392b;font-weight:600';
}}

function isHeatmapMetric(row) {{
  const metric = String(row?.metric || '');
  const group = String(row?.group || '');
  const kind = String(row?.kind || '');
  if (metric === 'СПП') return true;
  if (group === 'Воронка' && kind === 'percent_income') return true;
  return metric === 'ДРР от выкупов в эту дату' || metric === 'Маржинальность от выкупов за дату';
}}

function heatmapCellStyle(row, raw) {{
  if (!isHeatmapMetric(row) || raw === null || raw === undefined || !Number.isFinite(raw)) return '';
  if (String(row?.metric || '') === 'СПП') {{
    const values = (row?.raw_values || []).filter(value => Number.isFinite(value));
    if (!values.length) return '';
    const min = Math.min(...values);
    const max = Math.max(...values);
    const span = max - min;
    const score = span > 0 ? (raw - min) / span : 1;
    const hue = Math.round(8 + score * 132);
    const lightness = Math.round(96 - score * 14);
    return `background:hsl(${{hue}} 58% ${{lightness}}%);color:#0f172a;font-weight:700`;
  }}
  const values = (row?.raw_values || []).filter(value => Number.isFinite(value));
  if (!values.length) return '';
  const min = Math.min(...values);
  const max = Math.max(...values);
  const span = max - min;
  let score = span > 0 ? (raw - min) / span : 1;
  if (String(row?.metric || '') === 'ДРР от выкупов в эту дату') score = 1 - score;
  const hue = Math.round(8 + score * 132);
  const lightness = Math.round(94 - score * 18);
  return `background:hsl(${{hue}} 72% ${{lightness}}%);color:#0f172a;font-weight:700`;
}}

function metricColumnWidth(rows) {{
  const labels = ['Метрика', ...(rows || []).map(row => String(row.metric || ''))];
  const canvas = document.createElement('canvas');
  const ctx = canvas.getContext('2d');
  if (!ctx) return 240;
  ctx.font = '700 11px "Segoe UI", "Trebuchet MS", sans-serif';
  const textWidth = Math.max(...labels.map(label => ctx.measureText(label).width), 0);
  return Math.max(240, Math.min(560, Math.ceil(textWidth + 28)));
}}

function render(data) {{
  const dates = data.dates || [];
  const rows = data.rows || [];
  const formatDate = (iso) => {{
    if (String(iso).includes('..')) {{
      return String(iso).split('..').map(part => {{
        const [yy, mm, dd] = part.split('-');
        return `${{dd}}.${{mm}}.${{String(yy).slice(2)}}`;
      }}).join('-');
    }}
    const [y, m, d] = String(iso).split('-');
    return `${{d}}.${{m}}.${{String(y).slice(2)}}`;
  }};
  const isWeekendDate = (iso) => {{
    if (String(iso).includes('..')) return false;
    const day = new Date(`${{iso}}T00:00:00`).getDay();
    return day === 0 || day === 6;
  }};
  const head = '<tr><th>Гр.</th><th>Метрика</th><th>Итого / среднее</th>' + dates.map(d => `<th class="${{isWeekendDate(d) ? 'weekend-date' : ''}}">${{formatDate(d)}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.reduce((html, row, idx) => {{
        const prev = idx > 0 ? rows[idx - 1].group : null;
        const group = row.group || '';
        const kind = row.kind || 'none';
        const rawVals = row.raw_values || [];
        const groupCollapsed = group ? collapsedGroups.has(group) : false;
        const groupRow = group && group !== prev
          ? `<tr class="group-row" data-group-header="${{escapeHtml(group)}}">
              <td><button type="button" class="group-toggle" data-group-toggle="${{escapeHtml(group)}}" aria-label="Свернуть группу">${{groupCollapsed ? '+' : '−'}}</button></td>
              <td class="group-label" data-group-toggle="${{escapeHtml(group)}}">${{group}}</td>
              <td></td>
              ${{dates.map(() => '<td></td>').join('')}}
            </tr>`
          : '';
        const checked = activeChartMetrics.has(row.metric) ? ' checked' : '';
        const cells = (row.values || []).map((v, i) => {{
          const heatmapStyle = heatmapCellStyle(row, rawVals[i]);
          const changeStyle = (kind !== 'none' && kind !== 'neutral' && i > 0)
            ? cellColorStyle(kind, rawVals[i - 1], rawVals[i], row)
            : '';
          const style = heatmapStyle && changeStyle && String(row.metric || '') === 'СПП'
            ? `${{heatmapStyle}};${{changeStyle}}`
            : (heatmapStyle || changeStyle);
          return style ? `<td style="${{style}}">${{v}}</td>` : `<td>${{v}}</td>`;
        }}).join('');
        const hiddenAttr = groupCollapsed ? ' hidden' : '';
        return html + groupRow + `<tr data-group="${{escapeHtml(group)}}"${{hiddenAttr}}><td><input class="metric-toggle" type="checkbox" data-metric="${{escapeHtml(row.metric)}}"${{checked}}></td><td>${{row.metric}}</td><td>${{row.total || ''}}</td>${{cells}}</tr>`;
      }}, '')
    : `<tr><td colspan="${{dates.length + 3}}">Нет данных за выбранный период</td></tr>`;
  tbl.style.setProperty('--metric-col-width', `${{metricColumnWidth(rows)}}px`);
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
  requestAnimationFrame(() => setupStickyHeader());
  for (const toggle of [...tbl.querySelectorAll('[data-group-toggle]')]) {{
    toggle.addEventListener('click', () => {{
      const group = toggle.dataset.groupToggle || '';
      if (!group) return;
      if (collapsedGroups.has(group)) collapsedGroups.delete(group);
      else collapsedGroups.add(group);
      saveFilters();
      render(data);
    }});
  }}
  for (const input of [...tbl.querySelectorAll('input[data-metric]')]) {{
    input.addEventListener('change', () => {{
      const metric = input.dataset.metric;
      if (input.checked) activeChartMetrics.add(metric);
      else activeChartMetrics.delete(metric);
      chartMetricsInitialized = true;
      saveFilters();
      renderChart(data);
    }});
  }}
  renderChart(data);
}}

function renderArticles(articles) {{
  const normalized = articles.map(item => typeof item === 'string'
    ? {{ article: item, nmid: '', stock: null, buyouts: 0, revenue: 0, drr: null, turnover: null }}
    : item
  );
  articleItems = [...normalized].sort((left, right) => {{
    const key = articleSort.key;
    const dir = articleSort.dir === 'asc' ? 1 : -1;
    const leftValue = key === 'article' ? String(left.article || '') : Number(left[key] || 0);
    const rightValue = key === 'article' ? String(right.article || '') : Number(right[key] || 0);
    if (key === 'article') return leftValue.localeCompare(rightValue, 'ru') * dir;
    return (leftValue - rightValue) * dir || String(left.article || '').localeCompare(String(right.article || ''), 'ru');
  }});
  const sortMark = (key) => articleSort.key === key ? (articleSort.dir === 'asc' ? ' ▲' : ' ▼') : '';
  articleList.innerHTML = articleItems.length
    ? `<table class="article-table">
        <thead><tr>
          <th data-sort="article">Артикул${{sortMark('article')}}</th>
          <th data-sort="stock">Ост.${{sortMark('stock')}}</th>
          <th data-sort="drr">ДРР${{sortMark('drr')}}</th>
          <th data-sort="turnover">Обор.${{sortMark('turnover')}}</th>
          <th data-sort="revenue">Выручка${{sortMark('revenue')}}</th>
        </tr></thead>
	        <tbody>${{articleItems.map(item => {{
	          const article = String(item.article || '');
	          const itemKey = String(item.nmid || item.article || '').trim();
	          const selected = selectedArticles.has(itemKey) ? ' selected' : '';
	          const safeArticle = escapeHtml(article);
	          const nmid = String(item.nmid || '').trim();
          const copyBtn = `<button class="article-copy" data-copy="${{safeArticle}}" data-skip-select="1" title="Скопировать артикул">⎘</button>`;
          const wbLink = /^\d+$/.test(nmid)
            ? `<a class="article-link" href="https://www.wildberries.ru/catalog/${{nmid}}/detail.aspx" target="_blank" rel="noopener noreferrer" title="Открыть товар на WB" data-skip-select="1">↗</a>`
            : '';
          const campaignsListLink = /^\d+$/.test(nmid)
            ? `<a class="campaigns-list-link" href="https://cmp.wildberries.ru/campaigns/list?search=${{nmid}}" target="_blank" rel="noopener noreferrer" title="Рекламные кампании WB" data-skip-select="1">р</a>`
            : '';
          const pct = item.drr != null ? item.drr.toFixed(1) + '%' : '—';
          const turnoverStr = (item.turnover != null ? item.turnover : 999) + 'д';
	          return `<tr class="article-item${{selected}}" data-article="${{safeArticle}}" data-key="${{escapeHtml(itemKey)}}" title="${{safeArticle}}">
	            <td>${{safeArticle}}${{copyBtn}}${{wbLink}}${{campaignsListLink}}</td>
            <td>${{formatCompactNumber(item.stock)}}</td>
            <td class="drr-cell">${{pct}}</td>
            <td>${{turnoverStr}}</td>
            <td>${{formatCompactNumber(item.revenue)}}</td>
          </tr>`;
        }}).join('')}}</tbody>
      </table>`
    : '<div class="meta">Нет артикулов</div>';
  for (const th of [...articleList.querySelectorAll('th[data-sort]')]) {{
    th.addEventListener('click', () => {{
      const key = th.dataset.sort;
      articleSort = {{
        key,
        dir: articleSort.key === key && articleSort.dir === 'desc' ? 'asc' : 'desc'
      }};
      renderArticles(normalized);
    }});
  }}
  for (const btn of articleList.querySelectorAll('.article-copy')) {{
    btn.addEventListener('click', (e) => {{
      e.stopPropagation();
      navigator.clipboard.writeText(btn.dataset.copy).then(() => {{
        btn.classList.add('copied'); btn.textContent = '✓';
        setTimeout(() => {{ btn.classList.remove('copied'); btn.textContent = '⎘'; }}, 1500);
      }});
    }});
  }}
  for (const [idx, item] of [...articleList.querySelectorAll('.article-item')].entries()) {{
    item.addEventListener('click', (event) => {{
	      if (event.target.closest('[data-skip-select="1"]')) return;
	      const article = item.dataset.key || item.dataset.article;
	      if (event.shiftKey && lastArticleIndex >= 0) {{
	        const [from, to] = [lastArticleIndex, idx].sort((a, b) => a - b);
	        for (let pos = from; pos <= to; pos++) selectedArticles.add(String(articleItems[pos].nmid || articleItems[pos].article || '').trim());
      }} else if (event.ctrlKey || event.metaKey) {{
        if (selectedArticles.has(article)) {{
          selectedArticles.delete(article);
        }} else {{
          selectedArticles.add(article);
        }}
        lastArticleIndex = idx;
      }} else {{
        if (selectedArticles.has(article) && selectedArticles.size === 1) {{
          selectedArticles.clear();
        }} else {{
          selectedArticles = new Set([article]);
        }}
        lastArticleIndex = idx;
      }}
      saveFilters();
      renderArticles(articles);
      loadData();
    }});
  }}
}}

async function loadSubjects() {{
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  const key = `${{df}}..${{dt}}`;
  if (key === subjectsLoadedFor) return;
  subjectsLoadedFor = key;
  const currentSubject = subjectEl.dataset.pendingValue !== undefined ? subjectEl.dataset.pendingValue : subjectEl.value;
  const currentStrategy = [...strategyEl.selectedOptions].map(o=>o.value).join(',');
  const currentRevenueCategory = revenueCategoryEl.dataset.pendingValue !== undefined ? revenueCategoryEl.dataset.pendingValue : revenueCategoryEl.value;
  delete subjectEl.dataset.pendingValue;
  
  delete revenueCategoryEl.dataset.pendingValue;
  const qs = new URLSearchParams({{ date_from: df, date_to: dt}});
  const data = await fetchJsonChecked(`/api/analytics/buyout-filter-options?${{qs.toString()}}`);
  const subjects = data.subjects || [];
  const labelsData = data.labels || [];
  const revenueCategories = data.revenue_categories || [];
  subjectEl.innerHTML = '<option value="">Все предметы</option>' + subjects.map(subject => {{
    const selected = subject === currentSubject ? ' selected' : '';
    const safeSubject = escapeHtml(subject);
    return `<option value="${{safeSubject}}"${{selected}}>${{safeSubject}}</option>`;
  }}).join('');
  fillMultiSelect(strategyEl, labelsData);
  revenueCategoryEl.innerHTML = '<option value="">Все категории</option>' + revenueCategories.map(category => {{
    const selected = category === currentRevenueCategory ? ' selected' : '';
    const safeCategory = escapeHtml(category);
    return `<option value="${{safeCategory}}"${{selected}}>${{safeCategory}}</option>`;
  }}).join('');
}}

async function loadArticles() {{
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  await loadSubjects();
  const subject = subjectEl.value;
  const strategy = [...strategyEl.selectedOptions].map(o=>o.value).join(",");
  const revenueCategory = revenueCategoryEl.value;
  const key = `${{df}}..${{dt}}..${{subject}}..${{[...strategyEl.selectedOptions].map(o=>o.value).join(",")}}..${{revenueCategory}}`;
  if (key === articlesLoadedFor) return;
  articlesLoadedFor = key;
  const qs = new URLSearchParams({{ date_from: df, date_to: dt, subject, labels: strategy, revenue_category: revenueCategory}});
  const data = await fetchJsonChecked(`/api/analytics/buyout-articles?${{qs.toString()}}`);
  const articles = data.articles || [];
  if (selectedArticles.size) {{
    const available = new Set(articles.map(item => String((item && (item.nmid || item.article)) || '').trim()).filter(Boolean));
    selectedArticles = new Set([...selectedArticles].filter(item => available.has(item)));
  }}
  renderArticles(articles);
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const article = document.getElementById('article').value;
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  const subject = subjectEl.value;
  const strategy = [...strategyEl.selectedOptions].map(o=>o.value).join(",");
  const revenueCategory = revenueCategoryEl.value;
  meta.textContent = 'Загружаю...';
  try {{
    const qs = new URLSearchParams({{ article, date_from: df, date_to: dt, subject, labels: strategy, revenue_category: revenueCategory, granularity: reportGranularity}});
    for (const articleName of selectedArticles) qs.append('articles', articleName);
    // Run articles and pivot in parallel — pivot doesn't need articles list first
    const [_, data] = await Promise.all([
      loadArticles(),
      fetchJsonChecked(`/api/analytics/buyout-order-day?${{qs.toString()}}`)
    ]);
    if (seq !== requestSeq) return;
    render(data);
    const selection = selectedArticles.size ? ` | Выбрано артикулов: ${{selectedArticles.size}}` : '';
    meta.textContent = `Дат: ${{(data.dates || []).length}} | Показан период: ${{data.effective_from || df}}..${{data.effective_to || dt}}${{selection}}`;
  }} catch (err) {{
    if (seq !== requestSeq) return;
    const message = err && err.message ? err.message : 'Не удалось загрузить данные';
    meta.textContent = `Ошибка: ${{message}}`;
  }}
}}

async function fetchJsonChecked(url) {{
  const res = await fetch(url);
  let data = null;
  try {{
    data = await res.json();
  }} catch (_) {{
    if (!res.ok) throw new Error(`HTTP ${{res.status}}`);
    throw new Error('Некорректный ответ сервера');
  }}
  if (!res.ok) throw new Error((data && data.error) ? data.error : `HTTP ${{res.status}}`);
  if (data && data.error) throw new Error(data.error);
  return data;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('range-7').addEventListener('click', () => applyQuickRange(7));
document.getElementById('range-14').addEventListener('click', () => applyQuickRange(14));
document.getElementById('range-31').addEventListener('click', () => applyQuickRange(31));
document.getElementById('article').addEventListener('input', scheduleLoad);
document.getElementById('article').addEventListener('keydown', (e) => {{ if (e.key === 'Enter') loadData(); }});
subjectEl.addEventListener('change', () => {{
  selectedArticles.clear();
  articlesLoadedFor = '';
  saveFilters();
  loadData();
}});
strategyEl.addEventListener('change', () => {{
  selectedArticles.clear();
  articlesLoadedFor = '';
  saveFilters();
  loadData();
}});
revenueCategoryEl.addEventListener('change', () => {{
  selectedArticles.clear();
  articlesLoadedFor = '';
  saveFilters();
  loadData();
}});
document.getElementById('df').addEventListener('change', () => {{ subjectsLoadedFor = ''; articlesLoadedFor = ''; loadData(); }});
document.getElementById('dt').addEventListener('change', () => {{ subjectsLoadedFor = ''; articlesLoadedFor = ''; loadData(); }});
document.getElementById('select-all').addEventListener('click', async () => {{
  selectedArticles.clear();
  resetArticleFilters();
  articlesLoadedFor = '';
  saveFilters();
  await loadArticles();
  loadData();
}});
document.getElementById('clear-all').addEventListener('click', () => {{
  selectedArticles.clear();
  resetArticleFilters();
  articlesLoadedFor = '';
  saveFilters();
  loadArticles();
  loadData();
}});
document.getElementById('toggle-articles').addEventListener('click', () => {{
  articlesCollapsed = !articlesCollapsed;
  syncArticlesPanel();
  saveFilters();
}});
document.getElementById('expand-articles').addEventListener('click', () => {{
  articlesCollapsed = false;
  syncArticlesPanel();
  saveFilters();
}});
let _stickyClone = null;
let _stickyPageSL = null;
let _stickyTblSL = null;
function setupStickyHeader() {{
  if (_stickyClone) {{ _stickyClone.remove(); _stickyClone = null; }}
  if (_stickyPageSL) window.removeEventListener('scroll', _stickyPageSL);
  if (_stickyTblSL) tbl.removeEventListener('scroll', _stickyTblSL);
  const origRow = tbl.querySelector('thead tr');
  if (!origRow) return;
  const wrap = document.createElement('div');
  wrap.style.cssText = 'position:fixed;top:0;z-index:150;overflow:hidden;display:none;pointer-events:none;';
  const cloneTable = document.createElement('table');
  cloneTable.style.cssText = 'border-collapse:collapse;font-size:11px;table-layout:fixed;';
  const cloneThead = document.createElement('thead');
  const cloneRow = origRow.cloneNode(true);
  cloneThead.appendChild(cloneRow);
  cloneTable.appendChild(cloneThead);
  wrap.appendChild(cloneTable);
  document.body.appendChild(wrap);
  _stickyClone = wrap;
  function syncWidths() {{
    const tblRect = tbl.getBoundingClientRect();
    wrap.style.left = tblRect.left + 'px';
    wrap.style.width = tblRect.width + 'px';
    const orig = origRow.querySelectorAll('th');
    const clone = cloneRow.querySelectorAll('th');
    orig.forEach((th, i) => {{
      if (!clone[i]) return;
      const w = th.getBoundingClientRect().width;
      clone[i].style.cssText = th.style.cssText + `;width:${{w}}px;min-width:${{w}}px;max-width:${{w}}px;background:#f8fafc;border-bottom:1px solid #eef2f7;padding:6px 7px;font-size:11px;text-align:${{getComputedStyle(th).textAlign}};white-space:nowrap;`;
    }});
    cloneTable.style.transform = `translateX(-${{tbl.scrollLeft}}px)`;
  }}
  function update() {{
    const r = origRow.getBoundingClientRect();
    if (r.bottom <= 0) {{ syncWidths(); wrap.style.display = 'block'; }}
    else wrap.style.display = 'none';
  }}
  _stickyPageSL = () => update();
  _stickyTblSL = () => {{ if (wrap.style.display !== 'none') cloneTable.style.transform = `translateX(-${{tbl.scrollLeft}}px)`; }};
  window.addEventListener('scroll', _stickyPageSL, {{passive: true}});
  tbl.addEventListener('scroll', _stickyTblSL, {{passive: true}});
  update();
}}
restoreFilters();
loadData();
</script>
</body>
</html>
"""

ANALYTICS_DAY_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Аналитика по дням</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0369a1; --accent-2: #0891b2; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #cffafe, transparent 240px), var(--bg); }}
    .wrap {{ width: min(100% - 28px, 1800px); margin: 0 auto; padding: 20px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .84rem; margin-bottom: 4px; font-weight: 700; }}
    input {{ padding: 9px 10px; border: 1px solid #cbd5e1; border-radius: 8px; font-size: 14px; }}
    button {{ padding: 10px 14px; border: 0; border-radius: 8px; background: linear-gradient(135deg, var(--accent), var(--accent-2)); color: #fff; font-weight: 700; cursor: pointer; }}
    .kpis {{ margin-top: 12px; display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 10px; }}
    .kpi {{ background: #fff; border: 1px solid #e2e8f0; border-radius: 12px; padding: 12px; }}
    .kpi .name {{ color: var(--muted); font-size: .78rem; font-weight: 700; text-transform: uppercase; letter-spacing: .4px; }}
    .kpi .val {{ margin-top: 6px; font-size: 1.15rem; font-weight: 800; }}
    .meta {{ margin-top: 10px; color: var(--muted); font-size: .9rem; }}
    .tbl {{ margin-top: 12px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: auto; max-height: calc(100vh - 320px); }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid #eef2f7; padding: 8px 10px; text-align: left; white-space: nowrap; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1; }}
    th {{ position: sticky; top: 0; background: #f8fafc; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>Аналитика по дням</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="df">Дата с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Дата по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="kpis">
      <div class="kpi"><div class="name">Дней</div><div class="val" id="kpi-days">0</div></div>
      <div class="kpi"><div class="name">Продажи</div><div class="val" id="kpi-sales">0.00</div></div>
      <div class="kpi"><div class="name">Реклама</div><div class="val" id="kpi-ads">0.00</div></div>
      <div class="kpi"><div class="name">Чистая прибыль</div><div class="val" id="kpi-profit">0.00</div></div>
      <div class="kpi"><div class="name">ДРР</div><div class="val" id="kpi-drr">0%</div></div>
      <div class="kpi"><div class="name">% маржи</div><div class="val" id="kpi-margin">0%</div></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="tbl" id="tbl"></div>
  </div>
<script>
const cols = ["Дата", "Продажи", "Реклама", "Чистая прибыль", "ДРР", "% маржи"];
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const filterKey = 'wb.analytics.day.filters';
let requestSeq = 0;

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
  }} catch (err) {{}}
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value
  }}));
}}

function render(rows) {{
  const isNumericLike = (value) => /^-?\d[\d\s.,%]*$/.test(String(value ?? '').trim());
  const head = '<tr>' + cols.map(c => `<th>${{c}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.map(r => '<tr>' + cols.map(c => {{
        const value = r[c] ?? '';
        const cls = isNumericLike(value) ? ' class="num"' : '';
        return `<td${{cls}}>${{value}}</td>`;
      }}).join('') + '</tr>').join('')
    : `<tr><td colspan="${{cols.length}}">Нет данных за выбранный период</td></tr>`;
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
}}

function renderSummary(summary) {{
  document.getElementById('kpi-days').textContent = summary.days || '0';
  document.getElementById('kpi-sales').textContent = summary.sales || '0.00';
  document.getElementById('kpi-ads').textContent = summary.ads || '0.00';
  document.getElementById('kpi-profit').textContent = summary.profit || '0.00';
  document.getElementById('kpi-drr').textContent = summary.drr || '0%';
  document.getElementById('kpi-margin').textContent = summary.margin || '0%';
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  meta.textContent = 'Загружаю...';
  const res = await fetch(`/api/analytics/day?date_from=${{encodeURIComponent(df)}}&date_to=${{encodeURIComponent(dt)}}`);
  const data = await res.json();
  if (seq !== requestSeq) return;
  renderSummary(data.summary || {{}});
  render(data.rows || []);
  meta.textContent = `Строк: ${{(data.rows || []).length}}`;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('df').addEventListener('change', loadData);
document.getElementById('dt').addEventListener('change', loadData);
restoreFilters();
loadData();
</script>
</body>
</html>
"""

ANALYTICS_PERIOD_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Период по артикулам</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0f766e; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #e0f2fe, transparent 240px), var(--bg); }}
    .wrap {{ width: min(100% - 28px, 1800px); margin: 0 auto; padding: 20px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .84rem; margin-bottom: 4px; font-weight: 700; }}
    input {{ padding: 9px 10px; border: 1px solid #cbd5e1; border-radius: 8px; font-size: 14px; }}
    button {{ padding: 10px 14px; border: 0; border-radius: 8px; background: var(--accent); color: #fff; font-weight: 700; cursor: pointer; }}
    .meta {{ margin-top: 10px; color: var(--muted); font-size: .9rem; }}
    .tbl {{ margin-top: 12px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: auto; max-height: calc(100vh - 230px); }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid #eef2f7; padding: 8px 10px; text-align: left; white-space: nowrap; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1; }}
    th {{ position: sticky; top: 0; background: #f8fafc; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>Период по артикулам</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="article">Артикул</label>
        <input id="article" type="text" placeholder="например 12-0040-019">
      </div>
      <div>
        <label for="df">Период с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Период по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="tbl" id="tbl"></div>
  </div>
<script>
const cols = ["Артикул", "Продажи по нашей цене", "Реклама", "Чистая прибыль", "ДРР", "% маржи"];
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const filterKey = 'wb.analytics.period.filters';
let requestSeq = 0;
let filterTimer = null;

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.article !== undefined) document.getElementById('article').value = saved.article;
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
  }} catch (err) {{}}
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    article: document.getElementById('article').value,
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value
  }}));
}}

function scheduleLoad() {{
  clearTimeout(filterTimer);
  filterTimer = setTimeout(loadData, 350);
}}

function render(rows) {{
  const isNumericLike = (value) => /^-?\d[\d\s.,%]*$/.test(String(value ?? '').trim());
  const head = '<tr>' + cols.map(c => `<th>${{c}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.map(r => '<tr>' + cols.map(c => {{
        const value = r[c] ?? '';
        const cls = isNumericLike(value) ? ' class="num"' : '';
        return `<td${{cls}}>${{value}}</td>`;
      }}).join('') + '</tr>').join('')
    : `<tr><td colspan="${{cols.length}}">Нет данных за выбранный период</td></tr>`;
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const article = document.getElementById('article').value;
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  meta.textContent = 'Загружаю...';
  const qs = new URLSearchParams({{ article, date_from: df, date_to: dt}});
  const res = await fetch(`/api/analytics/period?${{qs.toString()}}`);
  const data = await res.json();
  if (seq !== requestSeq) return;
  render(data.rows || []);
  meta.textContent = `Строк: ${{(data.rows || []).length}}`;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('article').addEventListener('input', scheduleLoad);
document.getElementById('article').addEventListener('keydown', (e) => {{ if (e.key === 'Enter') loadData(); }});
document.getElementById('df').addEventListener('change', loadData);
document.getElementById('dt').addEventListener('change', loadData);
restoreFilters();
loadData();
</script>
</body>
</html>
"""

ANALYTICS_ARTICLE_DAY_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Дневная аналитика артикула</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0f766e; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #dcfce7, transparent 240px), var(--bg); }}
    .wrap {{ width: min(100% - 28px, 1800px); margin: 0 auto; padding: 20px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .84rem; margin-bottom: 4px; font-weight: 700; }}
    input {{ padding: 9px 10px; border: 1px solid #cbd5e1; border-radius: 8px; font-size: 14px; min-width: 180px; }}
    button {{ padding: 10px 14px; border: 0; border-radius: 8px; background: var(--accent); color: #fff; font-weight: 700; cursor: pointer; }}
    .meta {{ margin-top: 10px; color: var(--muted); font-size: .9rem; }}
    .tbl {{ margin-top: 12px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: auto; max-height: calc(100vh - 230px); }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid #eef2f7; padding: 8px 10px; text-align: left; white-space: nowrap; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1; }}
    th {{ position: sticky; top: 0; background: #f8fafc; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>Дни по артикулу</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="article">Артикул</label>
        <input id="article" type="text" placeholder="например 12-0040-019">
      </div>
      <div>
        <label for="df">Дата с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Дата по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="tbl" id="tbl"></div>
  </div>
<script>
const cols = [
  "Артикул / SKU",
  "Дата",
  "Продажи по нашей цене",
  "Реклама",
  "Чистая прибыль",
  "ДРР",
  "% маржи"
];
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const filterKey = 'wb.analytics.articleDay.filters';
let requestSeq = 0;
let filterTimer = null;

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.article !== undefined) document.getElementById('article').value = saved.article;
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
  }} catch (err) {{}}
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    article: document.getElementById('article').value,
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value
  }}));
}}

function scheduleLoad() {{
  clearTimeout(filterTimer);
  filterTimer = setTimeout(loadData, 350);
}}

function render(rows) {{
  const isNumericLike = (value) => /^-?\d[\d\s.,%]*$/.test(String(value ?? '').trim());
  const head = '<tr>' + cols.map(c => `<th>${{c}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.map(r => '<tr>' + cols.map(c => {{
        const value = r[c] ?? '';
        const cls = isNumericLike(value) ? ' class="num"' : '';
        return `<td${{cls}}>${{value}}</td>`;
      }}).join('') + '</tr>').join('')
    : `<tr><td colspan="${{cols.length}}">Нет данных по фильтру</td></tr>`;
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const article = document.getElementById('article').value;
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  meta.textContent = 'Загружаю...';
  const qs = new URLSearchParams({{ article, date_from: df, date_to: dt}});
  const res = await fetch(`/api/analytics/article-day?${{qs.toString()}}`);
  const data = await res.json();
  if (seq !== requestSeq) return;
  render(data.rows || []);
  meta.textContent = `Строк: ${{(data.rows || []).length}}`;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('article').addEventListener('input', scheduleLoad);
document.getElementById('article').addEventListener('keydown', (e) => {{ if (e.key === 'Enter') loadData(); }});
document.getElementById('df').addEventListener('change', loadData);
document.getElementById('dt').addEventListener('change', loadData);
restoreFilters();
loadData();
</script>
</body>
</html>
"""

PRELIMINARY_ECONOMICS_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Предварительная экономика по заказам</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0f766e; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #e0f2fe, transparent 240px), var(--bg); }}
    .wrap {{ width: min(100% - 28px, 1800px); margin: 0 auto; padding: 20px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .84rem; margin-bottom: 4px; font-weight: 700; }}
    input {{ padding: 9px 10px; border: 1px solid #cbd5e1; border-radius: 8px; font-size: 14px; }}
    button {{ padding: 10px 14px; border: 0; border-radius: 8px; background: var(--accent); color: #fff; font-weight: 700; cursor: pointer; }}
    .meta {{ margin-top: 10px; color: var(--muted); font-size: .9rem; }}
    .tbl {{ margin-top: 12px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: auto; max-height: calc(100vh - 230px); }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid #eef2f7; padding: 8px 10px; text-align: left; white-space: nowrap; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1; }}
    th {{ position: sticky; top: 0; background: #f8fafc; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>Предварительная экономика по заказам</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="article">Артикул</label>
        <input id="article" type="text" placeholder="например 12-0040-019">
      </div>
      <div>
        <label for="buyout">% выкупа</label>
        <input id="buyout" type="number" min="0" max="100" step="0.1" value="30">
      </div>
      <div>
        <label for="df">Период с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Период по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="tbl" id="tbl"></div>
  </div>
<script>
const cols = [
  "Артикул / SKU",
  "Дата",
  "Количество заказов",
  "Сумма заказов",
  "% выкупа",
  "Комиссия, ₽",
  "Эквайринг, ₽",
  "Реклама, ₽",
  "% рекламы",
  "Дополнительные расходы, ₽",
  "Предварительная прибыль, ₽"
  ,"% маржинальности"
];
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const filterKey = 'wb.analytics.preliminary.filters';
let requestSeq = 0;
let filterTimer = null;

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.article !== undefined) document.getElementById('article').value = saved.article;
    if (saved.buyout !== undefined) document.getElementById('buyout').value = saved.buyout;
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
  }} catch (err) {{}}
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    article: document.getElementById('article').value,
    buyout: document.getElementById('buyout').value,
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value
  }}));
}}

function scheduleLoad() {{
  clearTimeout(filterTimer);
  filterTimer = setTimeout(loadData, 350);
}}

function render(rows) {{
  const isNumericLike = (value) => /^-?\d[\d\s.,%]*$/.test(String(value ?? '').trim());
  const head = '<tr>' + cols.map(c => `<th>${{c}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.map(r => '<tr>' + cols.map(c => {{
        const value = r[c] ?? '';
        const cls = isNumericLike(value) ? ' class="num"' : '';
        return `<td${{cls}}>${{value}}</td>`;
      }}).join('') + '</tr>').join('')
    : `<tr><td colspan="${{cols.length}}">Нет данных за выбранный период</td></tr>`;
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const article = document.getElementById('article').value;
  const buyout = document.getElementById('buyout').value || '30';
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  meta.textContent = 'Загружаю...';
  const qs = new URLSearchParams({{ article, date_from: df, date_to: dt, buyout_percent: buyout}});
  const res = await fetch(`/api/analytics/preliminary-economics?${{qs.toString()}}`);
  const data = await res.json();
  if (seq !== requestSeq) return;
  render(data.rows || []);
  const extra = (data.expense_components || []).length
    ? ` | Статьи допрасходов: ${{data.expense_components.join(', ')}}`
    : '';
  meta.textContent = `Строк: ${{(data.rows || []).length}} | % выкупа: ${{data.buyout_percent || '30.00%'}} | Зафиксированный % допрасходов: ${{data.additional_rate || '0.00%'}}${{extra}}`;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('article').addEventListener('input', scheduleLoad);
document.getElementById('article').addEventListener('keydown', (e) => {{ if (e.key === 'Enter') loadData(); }});
document.getElementById('buyout').addEventListener('input', scheduleLoad);
document.getElementById('df').addEventListener('change', loadData);
document.getElementById('dt').addEventListener('change', loadData);
restoreFilters();
loadData();
</script>
</body>
</html>
"""

PRELIMINARY_ECONOMICS_SUMMARY_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Предварительная экономика по заказам (период)</title>
  <style>
    :root {{ --bg: #f8fafc; --card: #fff; --ink: #111827; --muted: #6b7280; --accent: #0f766e; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", "Trebuchet MS", sans-serif; color: var(--ink); background: linear-gradient(180deg, #e0f2fe, transparent 240px), var(--bg); }}
    .wrap {{ width: min(100% - 28px, 1800px); margin: 0 auto; padding: 20px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .panel {{ background: var(--card); border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ display: block; color: var(--muted); font-size: .84rem; margin-bottom: 4px; font-weight: 700; }}
    input {{ padding: 9px 10px; border: 1px solid #cbd5e1; border-radius: 8px; font-size: 14px; }}
    button {{ padding: 10px 14px; border: 0; border-radius: 8px; background: var(--accent); color: #fff; font-weight: 700; cursor: pointer; }}
    .meta {{ margin-top: 10px; color: var(--muted); font-size: .9rem; }}
    .tbl {{ margin-top: 12px; background: #fff; border: 1px solid #e5e7eb; border-radius: 12px; overflow: auto; max-height: calc(100vh - 230px); }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid #eef2f7; padding: 8px 10px; text-align: left; white-space: nowrap; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1; }}
    th {{ position: sticky; top: 0; background: #f8fafc; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>Предварительная экономика по периоду</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <div>
        <label for="article">Артикул</label>
        <input id="article" type="text" placeholder="например 12-0040-019">
      </div>
      <div>
        <label for="buyout">% выкупа</label>
        <input id="buyout" type="number" min="0" max="100" step="0.1" value="30">
      </div>
      <div>
        <label for="df">Период с</label>
        <input id="df" type="date" value="{date_from}">
      </div>
      <div>
        <label for="dt">Период по</label>
        <input id="dt" type="date" value="{date_to}">
      </div>
      <div><button id="load">Показать</button></div>
    </div>
    <div class="meta" id="meta">Загрузка...</div>
    <div class="tbl" id="tbl"></div>
  </div>
<script>
const cols = [
  "Артикул / SKU",
  "Период",
  "Количество заказов",
  "Сумма заказов",
  "% выкупа",
  "Комиссия, ₽",
  "Эквайринг, ₽",
  "Реклама, ₽",
  "% рекламы",
  "Дополнительные расходы, ₽",
  "Предварительная прибыль, ₽",
  "% маржинальности"
];
const meta = document.getElementById('meta');
const tbl = document.getElementById('tbl');
const filterKey = 'wb.analytics.preliminarySummary.filters';
let requestSeq = 0;
let filterTimer = null;

function restoreFilters() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(filterKey) || '{{}}');
    if (saved.article !== undefined) document.getElementById('article').value = saved.article;
    if (saved.buyout !== undefined) document.getElementById('buyout').value = saved.buyout;
    if (saved.df) document.getElementById('df').value = saved.df;
    if (saved.dt) document.getElementById('dt').value = saved.dt;
  }} catch (err) {{}}
}}

function saveFilters() {{
  localStorage.setItem(filterKey, JSON.stringify({{
    article: document.getElementById('article').value,
    buyout: document.getElementById('buyout').value,
    df: document.getElementById('df').value,
    dt: document.getElementById('dt').value
  }}));
}}

function scheduleLoad() {{
  clearTimeout(filterTimer);
  filterTimer = setTimeout(loadData, 350);
}}

function render(rows) {{
  const isNumericLike = (value) => /^-?\d[\d\s.,%]*$/.test(String(value ?? '').trim());
  const head = '<tr>' + cols.map(c => `<th>${{c}}</th>`).join('') + '</tr>';
  const body = rows.length
    ? rows.map(r => '<tr>' + cols.map(c => {{
        const value = r[c] ?? '';
        const cls = isNumericLike(value) ? ' class="num"' : '';
        return `<td${{cls}}>${{value}}</td>`;
      }}).join('') + '</tr>').join('')
    : `<tr><td colspan="${{cols.length}}">Нет данных за выбранный период</td></tr>`;
  tbl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;
}}

async function loadData() {{
  const seq = ++requestSeq;
  saveFilters();
  const article = document.getElementById('article').value;
  const buyout = document.getElementById('buyout').value || '30';
  const df = document.getElementById('df').value;
  const dt = document.getElementById('dt').value;
  meta.textContent = 'Загружаю...';
  const qs = new URLSearchParams({{ article, date_from: df, date_to: dt, buyout_percent: buyout, aggregate: '1'}});
  const res = await fetch(`/api/analytics/preliminary-economics?${{qs.toString()}}`);
  const data = await res.json();
  if (seq !== requestSeq) return;
  render(data.rows || []);
  const extra = (data.expense_components || []).length
    ? ` | Статьи допрасходов: ${{data.expense_components.join(', ')}}`
    : '';
  meta.textContent = `Строк: ${{(data.rows || []).length}} | % выкупа: ${{data.buyout_percent || '30.00%'}} | Зафиксированный % допрасходов: ${{data.additional_rate || '0.00%'}}${{extra}}`;
}}

document.getElementById('load').addEventListener('click', loadData);
document.getElementById('article').addEventListener('input', scheduleLoad);
document.getElementById('article').addEventListener('keydown', (e) => {{ if (e.key === 'Enter') loadData(); }});
document.getElementById('buyout').addEventListener('input', scheduleLoad);
document.getElementById('df').addEventListener('change', loadData);
document.getElementById('dt').addEventListener('change', loadData);
restoreFilters();
loadData();
</script>
</body>
</html>
"""

PLANNING_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Планирование</title>
  <style>
    :root {{ --bg:#f8fafc; --card:#fff; --ink:#111827; --muted:#6b7280; --accent:#0f766e; --line:#e5e7eb; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:"Segoe UI","Trebuchet MS",sans-serif; color:var(--ink); background:linear-gradient(180deg,#e0f2fe,transparent 240px),var(--bg); }}
    .wrap {{ width:calc(100% - 32px); max-width:3000px; margin:0 auto; padding:10px 0 28px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .ctrl-panel {{ background:var(--card); border:1px solid var(--line); border-radius:10px; padding:14px 16px; display:flex; flex-direction:column; gap:10px; }}
    .ctrl-row {{ display:flex; flex-wrap:wrap; align-items:flex-end; gap:8px; }}
    .ctrl-field {{ display:flex; flex-direction:column; gap:3px; }}
    .ctrl-field label {{ font-size:11px; color:#64748b; font-weight:600; text-transform:uppercase; letter-spacing:.04em; white-space:nowrap; }}
    .ctrl-field input, .ctrl-field select {{ height:32px; padding:0 8px; border:1px solid #cbd5e1; border-radius:6px; font-size:13px; background:#f8fafc; color:#0f172a; min-width:0; }}
    .ctrl-field select {{ padding-right:24px; }}
    .ctrl-sep {{ width:1px; align-self:stretch; background:var(--line); margin:0 4px; flex-shrink:0; }}
    .ctrl-group {{ display:flex; align-items:flex-end; gap:6px; }}
    .ctrl-group-label {{ font-size:10px; font-weight:700; text-transform:uppercase; letter-spacing:.06em; color:#94a3b8; margin-bottom:4px; }}
    .ctrl-btn {{ height:32px; padding:0 12px; border-radius:6px; border:none; cursor:pointer; font-size:13px; font-weight:500; color:#fff; white-space:nowrap; transition:filter .15s; }}
    .ctrl-btn:hover {{ filter:brightness(.88); }}
    .ctrl-btn:disabled {{ opacity:.5; cursor:default; }}
    .ctrl-btn-primary {{ background:#2563eb; }}
    .ctrl-btn-teal   {{ background:#0f766e; }}
    .ctrl-btn-purple {{ background:#7c3aed; }}
    .ctrl-btn-gray   {{ background:#64748b; }}
    .ctrl-btn-red    {{ background:#b91c1c; }}
    .ctrl-btn-green  {{ background:#065f46; }}
    .ctrl-btn-slate  {{ background:#334155; }}
    .ctrl-scenario-select {{ height:32px; padding:0 6px; border:1px solid #cbd5e1; border-radius:6px; font-size:12px; background:#f8fafc; color:#0f172a; max-width:130px; }}
    .ctrl-scenario-input  {{ height:32px; padding:0 8px; border:1px solid #cbd5e1; border-radius:6px; font-size:12px; background:#f8fafc; color:#0f172a; width:110px; }}
    .ctrl-tax-input {{ height:32px; width:60px; padding:0 8px; border:1px solid #cbd5e1; border-radius:6px; font-size:13px; background:#f8fafc; color:#0f172a; }}
    .panel {{ background:var(--card); border:1px solid var(--line); border-radius:8px; padding:12px; display:flex; flex-wrap:wrap; gap:10px; align-items:end; }}
    label {{ display:block; color:var(--muted); font-size:.72rem; margin-bottom:3px; font-weight:700; }}
    input, select {{ padding:7px 8px; border:1px solid #cbd5e1; border-radius:7px; font-size:12px; min-width:150px; background:#fff; }}
    button {{ padding:8px 12px; border:0; border-radius:7px; background:var(--accent); color:#fff; font-weight:700; cursor:pointer; font-size:12px; }}
    .meta {{ margin:8px 0 0; color:var(--muted); font-size:.78rem; }}
    .summary {{ display:flex; flex-wrap:nowrap; gap:10px; margin-top:10px; overflow-x:auto; }}
    .summary .kpi {{ flex:1 1 0; min-width:0; }}
    .kpi {{ background:#fff; border:1px solid var(--line); border-radius:8px; padding:10px 12px; }}
    .kpi-name {{ color:var(--muted); font-size:.72rem; font-weight:700; text-transform:uppercase; letter-spacing:.3px; }}
    .kpi-rub {{ margin-top:5px; font-size:1.15rem; font-weight:900; color:#0f172a; }}
    .kpi-qty {{ margin-top:2px; font-size:.82rem; font-weight:600; color:#475569; }}
    .tbl {{ margin-top:10px; background:#fff; border:1px solid var(--line); border-radius:8px; overflow:auto; max-height:calc(100vh - 260px); }}
    table {{ width:max-content; min-width:100%; border-collapse:collapse; font-size:11px; }}
    th, td {{ border-bottom:1px solid #eef2f7; padding:5px 6px; text-align:right; }}
    th {{ position:sticky; top:0; z-index:1; background:#f8fafc; color:#334155; font-weight:800; white-space:normal; word-break:normal; vertical-align:bottom; line-height:1.35; }}
    td {{ white-space:nowrap; }}
    .col-ck {{ width:22px; min-width:22px; max-width:22px; padding:4px 2px; text-align:center; position:sticky; left:0; z-index:2; background:#fff; }}
    th.col-ck {{ z-index:3; background:#f8fafc; }}
    .col-sku {{ width:88px; min-width:60px; max-width:100px; text-align:left; position:sticky; left:22px; z-index:2; background:#fff; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    th.col-sku {{ z-index:3; background:#f8fafc; }}
    .col-subject {{ text-align:left; min-width:54px; max-width:110px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    th.col-subject {{ white-space:normal; max-width:80px; }}
    .col-text {{ text-align:left; white-space:nowrap; }}
    tbody tr:hover td {{ background:#f0fdfa !important; }}
    tbody tr:hover td.col-ck, tbody tr:hover td.col-sku {{ background:#f0fdfa !important; }}
    tbody tr.row-selected td {{ background:#d1fae5 !important; }}
    tbody tr.row-selected td.col-ck, tbody tr.row-selected td.col-sku {{ background:#d1fae5 !important; }}
    .empty {{ padding:18px; color:var(--muted); font-size:.9rem; }}
    th.g-sep {{ border-left:2px solid #94a3b8; }}
    td.g-sep {{ border-left:2px solid #e2e8f0; }}
    .price-up {{ color:#16a34a; }}
    .price-down {{ color:#dc2626; }}
    .price-arrow {{ font-size:9px; margin-left:2px; }}
    @media (max-width:1200px) {{ .summary {{ flex-wrap:nowrap; }} }}
    .ck-cell {{ text-align:center; width:26px; min-width:26px; padding:4px 2px; }}
    input[type=checkbox] {{ min-width:auto; width:14px; height:14px; cursor:pointer; padding:0; accent-color:var(--accent); }}
    .pct-cell {{ display:flex; justify-content:flex-end; }}
    .pct-spinner {{ display:inline-flex; border:1px solid #cbd5e1; border-radius:5px; overflow:hidden; height:22px; }}
    .pct-spinner input[type=number] {{ width:42px; min-width:0; border:none; padding:2px 4px; text-align:center; font-size:11px; -moz-appearance:textfield; background:#fff; }}
    .pct-spinner input[type=number]::-webkit-inner-spin-button {{ display:none; }}
    .pct-spinner input[type=number]:focus {{ outline:none; }}
    .pct-arrows {{ display:flex; flex-direction:column; border-left:1px solid #cbd5e1; }}
    .pct-btn, .drr-btn, .conv-btn {{ width:15px; padding:0; font-size:7px; background:#f8fafc; color:#475569; border:none; border-radius:0; flex:1; line-height:1; cursor:pointer; min-width:0; }}
    .pct-btn:first-child, .drr-btn:first-child, .conv-btn:first-child {{ border-bottom:1px solid #e2e8f0; }}
    .pct-btn:hover, .drr-btn:hover, .conv-btn:hover {{ background:#e2e8f0; }}
    .pct-cell {{ gap:3px; align-items:center; }}
    .val-arrow {{ font-size:9px; font-weight:700; min-width:14px; width:14px; text-align:center; flex-shrink:0; }}
    .val-arrow:not(:empty) {{ cursor:pointer; padding:0 2px; }}
    .val-arrow:not(:empty):hover {{ opacity:0.6; }}
    .pct-spinner.no-data {{ opacity:0.35; pointer-events:none; }}
    .pct-spinner input:disabled {{ color:#94a3b8; background:#f8fafc; text-align:center; }}
    .val-arrow.val-up {{ color:#16a34a; }}
    .val-arrow.val-down {{ color:#dc2626; }}
    input.val-changed-up {{ color:#16a34a; font-weight:700; }}
    input.val-changed-down {{ color:#dc2626; font-weight:700; }}
    thead tr:first-child th:not([rowspan]) {{ font-size:9px; text-transform:uppercase; letter-spacing:.04em; padding:3px 8px; }}
    thead tr:last-child th {{ top:22px; z-index:1; }}
    th.col-ck, th.col-sku {{ top:0 !important; z-index:4 !important; }}
    .th-params {{ background:#dbeafe !important; color:#1d4ed8; }}
    .th-result {{ background:#dcfce7 !important; color:#166534; }}
    .th-lost  {{ background:#fef9c3 !important; color:#854d0e; }}
    th.params-col {{ background:#eff6ff; }}
    th.result-col {{ background:#f0fdf4; }}
    th.lost-col   {{ background:#fefce8; }}
    .th-econ {{ background:#fdf4ff !important; color:#7e22ce; }}
    th.econ-col {{ background:#fdf4ff; }}
    td.econ-col {{ background:#fdf4ff44; }}
    th[data-sort-key] {{ cursor:pointer; user-select:none; }}
    th[data-sort-key]:hover {{ filter:brightness(0.93); }}
    th[data-sort-key][data-sort-dir="asc"]::after  {{ content:" ▲"; font-size:.65em; opacity:.7; }}
    th[data-sort-key][data-sort-dir="desc"]::after {{ content:" ▼"; font-size:.65em; opacity:.7; }}
    #bulk-bar {{ display:none; position:fixed; bottom:22px; left:50%; transform:translateX(-50%); background:#0f172a; color:#fff; border-radius:10px; padding:10px 16px; align-items:center; gap:10px; font-size:12px; box-shadow:0 4px 24px rgba(0,0,0,.35); z-index:100; white-space:nowrap; }}
    #bulk-bar input[type=number] {{ width:64px; min-width:0; padding:5px 7px; color:#0f172a; font-size:12px; }}
    #bulk-bar .apply-btn {{ background:#0f766e; }}
    #bulk-bar .reset-btn {{ background:transparent; border:1px solid #475569; color:#94a3b8; }}
    #bulk-bar .close-btn {{ background:transparent; border:none; color:#94a3b8; font-size:14px; padding:2px 6px; cursor:pointer; }}
    .methodology {{ margin-top:18px; background:#fff; border:1px solid var(--line); border-radius:8px; padding:16px 20px; font-size:.8rem; line-height:1.6; color:#334155; }}
    .methodology h2 {{ margin:0 0 12px; font-size:.88rem; font-weight:800; color:#0f172a; }}
    .methodology h3 {{ margin:14px 0 4px; font-size:.78rem; font-weight:800; color:var(--accent); text-transform:uppercase; letter-spacing:.04em; }}
    .methodology p {{ margin:0 0 6px; }}
    .methodology ul {{ margin:2px 0 6px; padding-left:18px; }}
    .methodology li {{ margin-bottom:2px; }}
    .methodology .formula {{ font-family:monospace; background:#f1f5f9; border-radius:4px; padding:2px 6px; font-size:.77rem; }}
  </style>
</head>
<body>
<div class="wrap">
  <div class="top">
    <h1>Планирование</h1>
    <div class="report-nav">{report_nav}</div>
  </div>
  <div class="ctrl-panel">
    <!-- Ряд 1: параметры -->
    <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#94a3b8;margin-bottom:2px">Параметры расчёта</div>
    <div class="ctrl-row">
      <div class="ctrl-field">
        <label for="forecast-to">Прогноз до</label>
        <input id="forecast-to" type="date" value="{forecast_to}" style="width:130px">
      </div>
      <div class="ctrl-field">
        <label for="return-delay">Возврат, дней</label>
        <input id="return-delay" type="number" min="1" max="60" step="1" value="8" style="width:72px">
      </div>
      <div class="ctrl-field">
        <label for="transit-days">Поставка WB, дней</label>
        <input id="transit-days" type="number" min="0" max="30" step="1" value="3" style="width:72px">
      </div>
      <div class="ctrl-field">
        <label for="tax-pct">Налоги, %</label>
        <input id="tax-pct" type="number" min="0" max="50" step="0.1" value="0" style="width:72px">
      </div>
      <div class="ctrl-field">
        <label for="admin-pct">Адм. расходы, %</label>
        <input id="admin-pct" type="number" min="0" max="50" step="0.1" value="0" style="width:72px">
      </div>
      <div class="ctrl-sep"></div>
      <div>
        <div class="ctrl-group-label">Фильтры</div>
        <div style="display:flex;gap:6px;align-items:flex-end">
          <div class="ctrl-field">
            <label for="subject">Предмет</label>
            <select id="subject" style="width:160px"><option value="">Все предметы</option></select>
          </div>
          <div class="ctrl-field">
            <label for="strategy">Ярлыки</label>
            <select id="strategy" multiple size="4" style="width:140px;height:72px;" title="Ctrl/Cmd для множественного выбора"></select>
          </div>
          <div class="ctrl-field">
            <label for="revenue-category">Категория</label>
            <select id="revenue-category" style="width:140px"><option value="">Все категории</option></select>
          </div>
          <div class="ctrl-field">
            <label for="weeks-filter">Недель, макс.</label>
            <input id="weeks-filter" type="number" min="1" max="52" step="1" placeholder="Все" style="width:72px">
          </div>
        </div>
      </div>
      <button id="load" class="ctrl-btn ctrl-btn-primary" style="align-self:flex-end">Посчитать</button>
    </div>
    <!-- Ряд 2: действия -->
    <div class="ctrl-row" style="border-top:1px solid var(--line);padding-top:10px">
      <!-- Оптимизация -->
      <div class="ctrl-group">
        <div>
          <div class="ctrl-group-label">Оптимизация</div>
          <div style="display:flex;gap:6px">
            <button id="no-oos-btn" class="ctrl-btn ctrl-btn-teal" title="Подобрать минимальное повышение цены, чтобы не уйти в OOS до конца периода">Без OOS</button>
            <button id="optimize-btn" class="ctrl-btn ctrl-btn-purple" title="Найти цену и ДРР для максимального оборота при марже не ниже 0%">⚡ Макс. выручку</button>
          </div>
        </div>
      </div>
      <div class="ctrl-sep"></div>
      <!-- Сброс -->
      <div class="ctrl-group">
        <div>
          <div class="ctrl-group-label">Сбросить</div>
          <div style="display:flex;gap:6px">
            <button id="reset-prices" class="ctrl-btn ctrl-btn-gray">Цены</button>
            <button id="reset-drr"    class="ctrl-btn ctrl-btn-gray">ДРР</button>
            <button id="reset-conv"   class="ctrl-btn ctrl-btn-gray">Конверсии</button>
          </div>
        </div>
      </div>
      <div class="ctrl-sep"></div>
      <!-- Сценарии -->
      <div class="ctrl-group">
        <div>
          <div class="ctrl-group-label">Сценарий</div>
          <div style="display:flex;gap:6px;align-items:center">
            <select id="scenario-select" class="ctrl-scenario-select"><option value="">— выбрать —</option></select>
            <button id="scenario-load"   class="ctrl-btn ctrl-btn-teal"  style="padding:0 10px;font-size:12px">Загрузить</button>
            <button id="scenario-delete" class="ctrl-btn ctrl-btn-red"   style="padding:0 10px;font-size:12px">Удалить</button>
            <input type="text" id="scenario-name" class="ctrl-scenario-input" placeholder="Новый сценарий…">
            <button id="scenario-save"   class="ctrl-btn ctrl-btn-primary" style="padding:0 10px;font-size:12px">Сохранить</button>
          </div>
        </div>
      </div>
      <div class="ctrl-sep"></div>
      <!-- Экспорт и прочее -->
      <div class="ctrl-group" style="margin-left:auto">
        <div>
          <div class="ctrl-group-label">Экспорт и вид</div>
          <div style="display:flex;gap:6px;align-items:center">
            <button id="export-xls" class="ctrl-btn ctrl-btn-green">⬇ XLS</button>
            <div style="position:relative">
              <button id="col-settings-btn" class="ctrl-btn ctrl-btn-slate" style="padding:0 10px;font-size:16px" title="Настройка столбцов">⚙</button>
              <div id="col-settings-panel" style="display:none;position:absolute;right:0;top:36px;background:#fff;border:1px solid #cbd5e1;border-radius:8px;padding:12px 16px;z-index:300;box-shadow:0 4px 16px rgba(0,0,0,.18);white-space:nowrap;min-width:160px"></div>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>
  <div class="meta" id="meta">Загрузка...</div>
  <div style="margin-top:18px;margin-bottom:6px;display:flex;align-items:center;gap:10px">
    <span style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#94a3b8">Общие метрики</span>
    <div style="flex:1;height:1px;background:var(--line)"></div>
  </div>
  <div class="summary" id="summary"></div>
  <div class="tbl" id="table"></div>
</div>
<div id="bulk-bar">
  <span>Выбрано: <strong id="bulk-count">0</strong></span>
  <select id="bulk-type" style="padding:4px 6px;font-size:12px;background:#1e293b;color:#fff;border:1px solid #475569;border-radius:4px">
    <option value="price">Цена, %</option>
    <option value="drr">ДРР, %</option>
    <option value="ctr">CTR, %</option>
    <option value="cr1">CR1, %</option>
    <option value="cr2">CR2, %</option>
  </select>
  <input type="number" id="bulk-pct" value="0" step="0.5" min="-90" max="300" placeholder="±δ">
  <button class="apply-btn" id="bulk-apply">+ Применить</button>
  <button class="reset-btn" id="bulk-reset">Сбросить</button>
  <button class="close-btn" id="bulk-close">✕</button>
</div>
<script>
const stateKey = 'wb.analytics.planning.filters';
const fields = {{
  forecastTo: document.getElementById('forecast-to'),
  returnDelay: document.getElementById('return-delay'),
  transitDays: document.getElementById('transit-days'),
  subject: document.getElementById('subject'),
  strategy: document.getElementById('strategy'),
  revenueCategory: document.getElementById('revenue-category'),
  weeksFilter: document.getElementById('weeks-filter'),
}};
const meta = document.getElementById('meta');
const summaryEl = document.getElementById('summary');
const tableEl = document.getElementById('table');
const bulkBar = document.getElementById('bulk-bar');
const bulkCount = document.getElementById('bulk-count');
const bulkPct = document.getElementById('bulk-pct');

let rowDataByNmid = {{}};
const priceChangesKey = 'wb.analytics.planning.priceChanges';
let priceChanges = {{}};
let rowForecastCache = {{}};
const drrChangesKey = 'wb.analytics.planning.drrChanges';
const MAX_PLAN_DRR = 50;
let drrChanges = {{}};  // nmid → planned_drr_pct (float)
function normalizeDrrValue(value) {{
  const n = Number(value);
  if (!Number.isFinite(n)) return 0;
  return Math.max(0, Math.min(MAX_PLAN_DRR, Math.round(n * 10) / 10));
}}
function sanitizeDrrChanges() {{
  let changed = false;
  Object.keys(drrChanges || {{}}).forEach(nmid => {{
    const normalized = normalizeDrrValue(drrChanges[nmid]);
    if (normalized !== drrChanges[nmid]) {{ drrChanges[nmid] = normalized; changed = true; }}
  }});
  return changed;
}}
const convChangesKey = 'wb.analytics.planning.convChanges';
let convChanges = {{}};  // nmid → {{ ctr, cr1, cr2 }} (absolute %, null = use base)

function savePriceChanges() {{
  try {{ localStorage.setItem(priceChangesKey, JSON.stringify(priceChanges)); }} catch(e) {{}}
}}
function loadPriceChanges() {{
  try {{ priceChanges = JSON.parse(localStorage.getItem(priceChangesKey) || '{{}}'); }} catch(e) {{ priceChanges = {{}}; }}
}}
function saveDrrChanges() {{
  try {{ localStorage.setItem(drrChangesKey, JSON.stringify(drrChanges)); }} catch(e) {{}}
}}
function loadDrrChanges() {{
  try {{ drrChanges = JSON.parse(localStorage.getItem(drrChangesKey) || '{{}}'); }} catch(e) {{ drrChanges = {{}}; }}
  if (sanitizeDrrChanges()) saveDrrChanges();
}}
function saveConvChanges() {{
  try {{ localStorage.setItem(convChangesKey, JSON.stringify(convChanges)); }} catch(e) {{}}
}}
function loadConvChanges() {{
  try {{ convChanges = JSON.parse(localStorage.getItem(convChangesKey) || '{{}}'); }} catch(e) {{ convChanges = {{}}; }}
  // чистим нули, сохранённые старой версией кода (|| 0 вместо delete)
  for (const nmid of Object.keys(convChanges)) {{
    for (const k of Object.keys(convChanges[nmid] || {{}})) {{
      if (!(convChanges[nmid][k] > 0)) delete convChanges[nmid][k];
    }}
    if (!Object.keys(convChanges[nmid] || {{}}).length) delete convChanges[nmid];
  }}
}}
loadPriceChanges();
loadDrrChanges();
loadConvChanges();

// ── Сценарии ──────────────────────────────────────────────────────────────
const scenariosKey = 'wb.analytics.planning.scenarios';
function loadScenarioList() {{
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  const sel = document.getElementById('scenario-select');
  const cur = sel.value;
  sel.innerHTML = '<option value="">— выбрать —</option>';
  Object.keys(all).sort().forEach(name => {{
    const opt = document.createElement('option');
    opt.value = name; opt.textContent = name;
    sel.appendChild(opt);
  }});
  if (cur && all[cur]) sel.value = cur;
}}
loadScenarioList();

document.getElementById('scenario-save').addEventListener('click', () => {{
  const name = document.getElementById('scenario-name').value.trim();
  if (!name) return;
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  all[name] = {{
    price: Object.assign({{}}, priceChanges),
    drr:   Object.assign({{}}, drrChanges),
    conv:  JSON.parse(JSON.stringify(convChanges)),
  }};
  localStorage.setItem(scenariosKey, JSON.stringify(all));
  loadScenarioList();
  document.getElementById('scenario-select').value = name;
  document.getElementById('scenario-name').value = '';
}});

document.getElementById('scenario-load').addEventListener('click', () => {{
  const name = document.getElementById('scenario-select').value;
  if (!name) return;
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  const sc = all[name];
  if (!sc) return;
  priceChanges = sc.price || {{}}; savePriceChanges();
  drrChanges   = sc.drr   || {{}};
  sanitizeDrrChanges();
  saveDrrChanges();
  convChanges  = sc.conv  || {{}}; saveConvChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const pInp = tr.querySelector('.pct-input');
    if (pInp) pInp.value = priceChanges[nmid] ?? 0;
    const dInp = tr.querySelector('.drr-input');
    if (dInp) dInp.value = normalizeDrrValue(drrChanges[nmid] != null ? drrChanges[nmid] : (row.base_drr_pct || 0)).toFixed(1);
    ['ctr','cr1','cr2'].forEach(k => {{
      const inp = tr.querySelector(`.conv-input[data-conv="${{k}}"]`);
      if (inp && !inp.disabled) {{
        const v = (convChanges[nmid] || {{}})[k];
        inp.value = (v > 0 ? v : (row[`base_${{k}}`] || 0)).toFixed(2);
      }}
    }});
    updateRowPlanned(nmid);
  }});
}});

document.getElementById('scenario-delete').addEventListener('click', () => {{
  const name = document.getElementById('scenario-select').value;
  if (!name) return;
  if (!confirm(`Удалить сценарий «${{name}}»?`)) return;
  const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
  delete all[name];
  localStorage.setItem(scenariosKey, JSON.stringify(all));
  loadScenarioList();
}});

// ── Экспорт XLS ───────────────────────────────────────────────────────────
document.getElementById('export-xls').addEventListener('click', () => {{
  const taxOnlyPct   = Number(document.getElementById('tax-pct')?.value) || 0;
  const adminOnlyPct = Number(document.getElementById('admin-pct')?.value) || 0;
  const exportRows = Object.values(rowDataByNmid).map(row => {{
    const f = calcRowForecast(row);
    const drrPct    = normalizeDrrValue(drrChanges[row.nmid] != null ? drrChanges[row.nmid] : (row.base_drr_pct || 0));
    const revenue   = f.forecastBuyouts * f.plannedClient;
    const cogsTotal = f.forecastBuyouts * (row.cogs || 0);
    const wbComm    = revenue * ((row.wb_commission_pct || 0) / 100);
    const acquiring = revenue * 0.03;
    const drrCost   = revenue * (drrPct / 100);
    const grossProfit = revenue - cogsTotal - wbComm - acquiring - drrCost;
    const marginPct   = revenue > 0 ? grossProfit / revenue * 100 : 0;
    const adminCost   = revenue * (adminOnlyPct / 100);
    const taxCost     = revenue * (taxOnlyPct / 100);
    const netProfit   = grossProfit - adminCost - taxCost;
    const profitPct   = revenue > 0 ? netProfit / revenue * 100 : 0;
    return {{
      article:           row.article,
      nmid:              row.nmid,
      subject:           row.subject,
      strategy:          (row.labels||[]).join(", "),
      revenue_cat:       row.revenue_category,
      stock:             row.stock,
      wh_stock:          row.warehouse_stock,
      stock_val:         (row.stock + (row.warehouse_stock || 0)) * (row.client_price || row.average_check || 0),
      base_price:        row.baseline_seller_price,
      base_spp:          row.baseline_spp,
      base_client:       row.baseline_client_check,
      base_speed:        row.base_speed,
      cur_price:         row.current_price,
      cur_spp:           row.current_spp,
      cur_client:        row.client_price,
      elasticity:        row.elasticity,
      cur_speed:         row.current_speed,
      price_pct:         priceChanges[row.nmid] ?? 0,
      plan_drr:          drrPct,
      ads_elasticity:    row.ads_elasticity,
      base_ctr:          row.base_ctr,
      base_cr1:          row.base_cr1,
      base_cr2:          row.base_cr2,
      planned_price:     f.plannedSeller,
      planned_client:    f.plannedClient,
      planned_speed:     f.plannedSpeed,
      buyout_percent:    row.buyout_percent,
      forecast_orders:   f.forecastOrders,
      forecast_buyouts:  f.forecastBuyouts,
      forecast_revenue:  f.forecastRevenue,
      stockout_date:     f.stockoutDate || '',
      lost_orders:       f.lostOrders,
      lost_revenue:      f.lostRevenue,
      comm_pct:          row.wb_commission_pct || 0,
      comm_rub:          wbComm,
      acq_rub:           acquiring,
      drr_rub:           drrCost,
      cogs_total:        cogsTotal,
      margin_pct:        marginPct,
      margin_rub:        grossProfit,
      admin_rub:         adminCost,
      tax_rub:           taxCost,
      profit_pct:        profitPct,
      profit_rub:        netProfit,
    }};
  }});
  fetch('/api/analytics/planning/export', {{
    method: 'POST',
    headers: {{'Content-Type': 'application/json'}},
    body: JSON.stringify({{ rows: exportRows }}),
  }}).then(r => {{
    if (!r.ok) return r.text().then(t => {{ throw new Error(t); }});
    return r.blob();
  }}).then(blob => {{
    const a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    const today = new Date().toISOString().slice(0, 10);
    a.download = `planning_${{today}}.xlsx`;
    a.click();
  }}).catch(err => alert('Ошибка экспорта: ' + err));
}});

let planningParams = {{ forecast_days: 0, return_delay_days: 8, transit_days: 3, forecast_from: '' }};

function calcConvFactor(row) {{
  const c = convChanges[row.nmid] || {{}};
  let f = 1.0;
  if (c.ctr > 0 && row.base_ctr > 0) f *= c.ctr / row.base_ctr;
  if (c.cr1 > 0 && row.base_cr1 > 0) f *= c.cr1 / row.base_cr1;
  if (c.cr2 > 0 && row.base_cr2 > 0) f *= c.cr2 / row.base_cr2;
  return f;
}}

function simulatePlanning(stock, dailySpeed, buyoutRate, forecastDays, returnDelay, startDateStr, warehouseStock, transitDays) {{
  let available = Math.max(0, stock);
  const demand = Math.max(0, dailySpeed);
  buyoutRate = Math.max(0, Math.min(1, buyoutRate));
  const delay = Math.max(1, Math.round(returnDelay));
  const transit = Math.max(0, Math.round(transitDays || 0));
  const arrivals = {{}};
  if (warehouseStock > 0) arrivals[transit] = Math.max(0, warehouseStock);
  const returnsByDay = {{}};
  let totalOrders = 0;
  let stockoutDate = '';
  const startMs = startDateStr ? new Date(startDateStr).getTime() : 0;
  for (let idx = 0; idx < Math.max(0, forecastDays); idx++) {{
    available += (returnsByDay[idx] || 0) + (arrivals[idx] || 0);
    delete returnsByDay[idx];
    delete arrivals[idx];
    const ordered = Math.min(demand, available);
    if (demand > available && !stockoutDate && startMs) {{
      const d = new Date(startMs + idx * 86400000);
      stockoutDate = d.toISOString().slice(0, 10);
    }}
    available -= ordered;
    totalOrders += ordered;
    const returning = ordered * (1 - buyoutRate);
    const dueIdx = idx + delay;
    if (returning > 0 && dueIdx < forecastDays) {{
      returnsByDay[dueIdx] = (returnsByDay[dueIdx] || 0) + returning;
    }}
  }}
  return {{ totalOrders, stockoutDate }};
}}

function formatNumber(value, digits = 0) {{
  const num = Number(value || 0);
  return num.toLocaleString('ru-RU', {{ maximumFractionDigits: digits, minimumFractionDigits: digits }});
}}
function formatMoney(value) {{ return formatNumber(value, 0); }}

function saveState() {{
  try {{
    localStorage.setItem(stateKey, JSON.stringify({{
      forecastTo: fields.forecastTo.value,
      returnDelay: fields.returnDelay.value,
      transitDays: fields.transitDays.value,
      subject: fields.subject.value,
      labels: [...(fields.strategy?.selectedOptions||[])].map(o=>o.value).join(","),
      revenueCategory: fields.revenueCategory.value,
    }}));
  }} catch (err) {{}}
}}

function restoreState() {{
  try {{
    const saved = JSON.parse(localStorage.getItem(stateKey) || '{{}}');
    if (saved.forecastTo) fields.forecastTo.value = saved.forecastTo;
    if (saved.returnDelay) fields.returnDelay.value = saved.returnDelay;
    if (saved.transitDays != null) fields.transitDays.value = saved.transitDays;
    fields.subject.dataset.pending = saved.subject || '';
    // labels not restored from storage - always start empty
    fields.revenueCategory.dataset.pending = saved.revenueCategory || '';
  }} catch (err) {{}}
}}

function fillSelect(select, values, placeholder) {{
  const current = select.value || select.dataset.pending || '';
  select.innerHTML = `<option value="">${{placeholder}}</option>` + (values || []).map(v => `<option value="${{String(v).replaceAll('"','&quot;')}}">${{v}}</option>`).join('');
  if ([...select.options].some(o => o.value === current)) select.value = current;
  select.dataset.pending = '';
}}
function fillMultiSelect(select, values) {{
  const prevVals = new Set([...select.selectedOptions].map(o=>o.value));
  const optionsHtml = (values || []).map(v => {{
    const sel = prevVals.has(v) ? ' selected' : '';
    return `<option value="${{String(v).replaceAll('"','&quot;')}}"${{sel}}>${{v}}</option>`;
  }}).join('');
  select.innerHTML = select.multiple ? optionsHtml : `<option value="">Все ярлыки</option>${{optionsHtml}}`;
}}

function renderSummary(summary) {{
  const avgMargin = summary.total_econ_revenue > 0
    ? (summary.total_gross_profit / summary.total_econ_revenue * 100).toFixed(1)
    : '—';
  summaryEl.innerHTML = `
    <div class="kpi">
      <div class="kpi-name">Себестоимость остатков</div>
      <div class="kpi-rub">${{formatMoney(summary.stock_cost)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.stock)}} шт на складе</div>
      <div class="kpi-qty" style="color:#64748b">В тек. ценах: ${{formatMoney(summary.stock_value_current || summary.stock_value)}}</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Плановые заказы</div>
      <div class="kpi-rub">${{formatMoney(summary.orders_revenue)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.forecast_orders)}} шт</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Плановые выкупы</div>
      <div class="kpi-rub">${{formatMoney(summary.forecast_revenue)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.forecast_buyouts)}} шт</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Упущенные продажи</div>
      <div class="kpi-rub">${{formatMoney(summary.lost_revenue)}}</div>
      <div class="kpi-qty">${{formatNumber(summary.lost_orders)}} шт</div>
    </div>
    <div class="kpi">
      <div class="kpi-name">Расходы на рекламу</div>
      <div class="kpi-rub">${{formatMoney(summary.total_drr_cost)}}</div>
      <div class="kpi-qty">${{summary.total_econ_revenue > 0 ? (summary.total_drr_cost / summary.total_econ_revenue * 100).toFixed(1) : '—'}}% от выкупов</div>
    </div>
    <div class="kpi" style="border-color:#a855f7">
      <div class="kpi-name" style="color:#7e22ce">Маржинальность</div>
      <div class="kpi-rub" style="color:#7e22ce">${{avgMargin}}%</div>
      <div class="kpi-qty">валовая прибыль ${{formatMoney(summary.total_gross_profit)}}</div>
    </div>
    <div class="kpi" style="border-color:#a855f7">
      <div class="kpi-name" style="color:#7e22ce">Чистая прибыль</div>
      <div class="kpi-rub" style="color:${{(summary.total_net_profit||0)>=0?'#166534':'#dc2626'}}">${{formatMoney(summary.total_net_profit)}}</div>
      <div class="kpi-qty">после налогов и адм.</div>
    </div>`;
}}

function priceArrow(current, baseline, invertLogic) {{
  if (!baseline || !current) return '';
  const delta = (current - baseline) / baseline;
  if (Math.abs(delta) < 0.01) return '';
  const up = invertLogic ? delta < 0 : delta > 0;
  return `<span class="price-arrow ${{up ? 'price-up' : 'price-down'}}">${{up ? '↑' : '↓'}}</span>`;
}}

function calcNoOOS(row) {{
  const {{ forecast_days, return_delay_days, transit_days, forecast_from }} = planningParams;
  const buyoutRate = row.buyout_percent / 100;
  const wh = row.warehouse_stock || 0;
  const curPct = priceChanges[row.nmid] ?? 0;
  // check OOS at current planned speed (already includes DRR/conv factors)
  const {{ plannedSpeed: curSpeed }} = calcPlanned(row, curPct);
  const {{ stockoutDate: sd0 }} = simulatePlanning(row.stock, curSpeed, buyoutRate, forecast_days, return_delay_days, forecast_from, wh, transit_days);
  if (!sd0) return 0;  // no OOS at current plan → nothing to do
  if (row.elasticity >= 0) return null;  // can't reduce speed by raising price
  const baseClient = row.baseline_client_check;
  if (!baseClient) return null;
  // binary search starting from current price pct (don't lower existing change)
  let lo = curPct, hi = Math.max(300, curPct + 300);
  for (let i = 0; i < 50; i++) {{
    const mid = (lo + hi) / 2;
    const {{ plannedSpeed }} = calcPlanned(row, mid);
    const {{ stockoutDate }} = simulatePlanning(row.stock, plannedSpeed, buyoutRate, forecast_days, return_delay_days, forecast_from, wh, transit_days);
    if (!stockoutDate) hi = mid; else lo = mid;
    if (hi - lo < 0.05) break;
  }}
  if (hi >= curPct + 299) return null;  // can't avoid OOS even with huge increase
  return Math.ceil(hi);
}}

function adsResponseFactor(row, drrPct) {{
  const e = Math.max(0, Number(row.ads_elasticity) || 0);
  if (e <= 0) return 1.0;
  const anchor = Math.max(0.1, Number(row.drr_anchor_pct) || Math.min(Math.max(Number(row.base_drr_pct) || 0, 6), 12));
  const organic = Math.max(0.15, Math.min(0.75, Number(row.organic_floor) || 0.45));
  const baseDrr = Math.max(0, Number(row.base_drr_pct) || 0);
  const plannedDrr = Math.max(0, Number(drrPct) || 0);
  const raw = d => organic + (1 - organic) * Math.pow(Math.max(0, d) / anchor, e);
  const baseRaw = Math.max(0.05, raw(baseDrr));
  const factor = raw(plannedDrr) / baseRaw;
  return Math.max(0.25, Math.min(3.0, factor));
}}

// Variant of calcPlanned with explicit drrPct (doesn't read global drrChanges)
function calcPlannedEx(row, pricePct, drrPct) {{
  const p = Number(pricePct) || 0;
  const plannedSeller = row.current_price * (1 + p / 100);
  const plannedClient = plannedSeller * (1 - row.current_spp / 100);
  const speedAnchor = row.current_speed || row.base_speed || 0;
  const clientAnchor = row.client_price || row.baseline_client_check || plannedClient;
  const delta = (clientAnchor > 0 && plannedClient > 0) ? (plannedClient / clientAnchor - 1) : 0;
  const priceFactor = Math.max(0, 1 + row.elasticity * delta);
  const adsFactor = adsResponseFactor(row, drrPct);
  return {{ plannedSeller, plannedClient, plannedSpeed: speedAnchor * priceFactor * adsFactor * calcConvFactor(row) }};
}}

function netRevEx(row, pricePct, drrPct) {{
  const {{ plannedSpeed, plannedClient }} = calcPlannedEx(row, pricePct, drrPct);
  const {{ forecast_days, return_delay_days, transit_days, forecast_from }} = planningParams;
  const {{ totalOrders, stockoutDate }} = simulatePlanning(
    row.stock, plannedSpeed, row.buyout_percent / 100,
    forecast_days, return_delay_days, forecast_from,
    row.warehouse_stock || 0, transit_days
  );
  const buyoutRate = row.buyout_percent / 100;
  const gross = totalOrders * buyoutRate * plannedClient;
  const forecastBuyouts = totalOrders * buyoutRate;
  const cogsTotal = forecastBuyouts * (row.cogs || 0);
  const wbComm = gross * ((row.wb_commission_pct || 0) / 100);
  const acquiring = gross * 0.03;
  const drrCost = gross * (normalizeDrrValue(drrPct) / 100);
  const grossProfit = gross - cogsTotal - wbComm - acquiring - drrCost;
  const marginPct = gross > 0 ? grossProfit / gross * 100 : 0;
  return {{ gross, grossProfit, marginPct, totalOrders, stockoutDate, plannedSpeed, plannedClient }};
}}

function getRevenueDrrRange(row) {{
  const currentDrr = normalizeDrrValue(drrChanges[row.nmid] != null ? drrChanges[row.nmid] : (row.base_drr_pct || 0));
  const cur = netRevEx(row, priceChanges[row.nmid] ?? 0, currentDrr);
  const stockBase = Math.max(1, (row.stock || 0) + (row.warehouse_stock || 0));
  const sellThrough = cur.totalOrders / stockBase;
  if (cur.stockoutDate || sellThrough >= 0.85) {{
    return {{ min: 0, max: 6, mode: 'ending' }};
  }}
  if (sellThrough < 0.35 || (row.base_speed || 0) < 0.25) {{
    return {{ min: 12, max: 50, mode: 'launch' }};
  }}
  return {{ min: 6, max: 12, mode: 'normal' }};
}}

function isBetterRevenueCandidate(candidate, best, drr, bestDrr, pricePct, bestPct, mode) {{
  if (!best) return true;
  const eps = Math.max(25, best.gross * 0.002);
  if (candidate.gross > best.gross + eps) return true;
  if (candidate.gross < best.gross - eps) return false;
  if (Math.abs(drr - bestDrr) > 0.05) return drr < bestDrr;
  if (mode === 'ending' && pricePct > bestPct) return true;
  return Math.abs(pricePct) < Math.abs(bestPct);
}}

function findOptimalRevenue(row) {{
  if (!row.stock || !planningParams.forecast_days) return {{ pricePct: 0, drrPct: row.base_drr_pct || 0 }};
  const drrRange = getRevenueDrrRange(row);

  // Cap upward price search: extrapolating far beyond data range is unreliable.
  // Ending products can take a higher price because excess demand cannot be fulfilled anyway.
  const absE = Math.abs(row.elasticity || 0);
  const maxPriceUp = drrRange.mode === 'ending'
    ? (absE < 0.3 ? 25 : absE < 0.7 ? 35 : 45)
    : (absE === 0 ? 5 : absE < 0.3 ? 10 : absE < 0.7 ? 20 : 30);

  let bestPct = 0, bestDrr = drrRange.min, best = null;

  // Objective is gross turnover with non-negative margin after cogs, WB commission, acquiring and ads.
  for (let p = -30; p <= maxPriceUp; p += 3) {{
    for (let d = drrRange.min; d <= drrRange.max; d += 1) {{
      const r = netRevEx(row, p, d);
      if (r.grossProfit >= 0 && isBetterRevenueCandidate(r, best, d, bestDrr, p, bestPct, drrRange.mode)) {{
        best = r; bestPct = p; bestDrr = d;
      }}
    }}
  }}
  // If normal price limits cannot reach zero margin, expand price search.
  // This keeps the "max revenue" action from choosing a loss-making plan.
  if (!best) {{
    for (let p = maxPriceUp + 5; p <= 300; p += 5) {{
      for (let d = drrRange.min; d <= drrRange.max; d += 1) {{
        const r = netRevEx(row, p, d);
        if (r.grossProfit >= 0 && isBetterRevenueCandidate(r, best, d, bestDrr, p, bestPct, drrRange.mode)) {{
          best = r; bestPct = p; bestDrr = d;
        }}
      }}
    }}
  }}
  const p0 = bestPct, d0 = bestDrr;
  const refineMaxPrice = best ? Math.max(maxPriceUp, p0 + 3) : maxPriceUp;
  for (let p = p0 - 3; p <= Math.min(p0 + 3, refineMaxPrice); p++) {{
    for (let di = -2; di <= 2; di++) {{
      const d = Math.round((d0 + di * 0.5) * 10) / 10;
      if (d < drrRange.min || d > drrRange.max) continue;
      const r = netRevEx(row, p, d);
      if (r.grossProfit >= 0 && isBetterRevenueCandidate(r, best, d, bestDrr, p, bestPct, drrRange.mode)) {{
        best = r; bestPct = p; bestDrr = d;
      }}
    }}
  }}
  if (!best) {{
    return {{ pricePct: 0, drrPct: normalizeDrrValue(row.base_drr_pct || 0) }};
  }}

  return {{ pricePct: bestPct, drrPct: Math.round(bestDrr * 10) / 10 }};
}}

function calcPlanned(row, pct) {{
  const p = Number(pct) || 0;
  const plannedSeller = row.current_price * (1 + p / 100);
  const plannedClient = plannedSeller * (1 - row.current_spp / 100);
  const speedAnchor = row.current_speed || row.base_speed || 0;
  const clientAnchor = row.client_price || row.baseline_client_check || plannedClient;
  const delta = (clientAnchor > 0 && plannedClient > 0) ? (plannedClient / clientAnchor - 1) : 0;
  const priceFactor = Math.max(0, 1 + row.elasticity * delta);
  // ads factor: per-row planned DRR% (absolute), normalized to observed baseline
  const plannedDrr = normalizeDrrValue(drrChanges[row.nmid] != null ? drrChanges[row.nmid] : (row.base_drr_pct || 0));
  const adsFactor = adsResponseFactor(row, plannedDrr);
  // conversion factor
  const convFactor = calcConvFactor(row);
  return {{ plannedSeller, plannedClient, plannedSpeed: speedAnchor * priceFactor * adsFactor * convFactor }};
}}

function calcRowForecast(row) {{
  const pct = priceChanges[row.nmid] ?? 0;
  const {{ plannedSeller, plannedClient, plannedSpeed }} = calcPlanned(row, pct);
  const {{ forecast_days, return_delay_days, transit_days, forecast_from }} = planningParams;
  const {{ totalOrders: forecastOrders, stockoutDate }} = simulatePlanning(
    row.stock, plannedSpeed, row.buyout_percent / 100,
    forecast_days, return_delay_days, forecast_from,
    row.warehouse_stock || 0, transit_days
  );
  const buyoutRate = row.buyout_percent / 100;
  const effectiveCheck = plannedClient > 0 ? plannedClient : row.average_check;
  const forecastBuyouts = forecastOrders * buyoutRate;
  const forecastRevenue = forecastBuyouts * effectiveCheck;
  const potentialOrders = plannedSpeed * forecast_days;
  const lostOrders = Math.max(0, potentialOrders - forecastOrders);
  const lostRevenue = lostOrders * buyoutRate * effectiveCheck;
  return {{ plannedSeller, plannedClient, plannedSpeed, effectiveCheck, forecastOrders, forecastBuyouts, forecastRevenue, stockoutDate, lostOrders, lostRevenue }};
}}

function recalcSummary() {{
  const s = {{ stock_cost: 0, stock_value: 0, stock_value_current: 0, forecast_orders: 0, forecast_buyouts: 0, orders_revenue: 0, forecast_revenue: 0, lost_orders: 0, lost_revenue: 0, stock: 0, total_gross_profit: 0, total_net_profit: 0, total_econ_revenue: 0, total_drr_cost: 0 }};
  Object.values(rowDataByNmid).forEach(row => {{
    const c = rowForecastCache[row.nmid] || {{}};
    s.stock += row.stock + (row.warehouse_stock || 0);
    s.stock_cost += (row.stock + (row.warehouse_stock || 0)) * (row.cogs || 0);
    s.stock_value += (row.stock + (row.warehouse_stock || 0)) * (row.client_price || row.average_check || 0);
    s.stock_value_current += (row.stock + (row.warehouse_stock || 0)) * (row.client_price || row.average_check || 0);
    s.forecast_orders += c.forecastOrders || 0;
    s.forecast_buyouts += c.forecastBuyouts || 0;
    s.orders_revenue += c.ordersRevenue || 0;
    s.forecast_revenue += c.forecastRevenue || 0;
    s.lost_orders += c.lostOrders || 0;
    s.lost_revenue += c.lostRevenue || 0;
    s.total_gross_profit += c.grossProfit || 0;
    s.total_net_profit   += c.netProfit || 0;
    s.total_econ_revenue += c.econRevenue || 0;
    s.total_drr_cost     += c.drrCost || 0;
  }});
  renderSummary(s);
}}

function updateRowPlanned(nmid, skipSummary) {{
  const row = rowDataByNmid[nmid];
  if (!row) return;
  const f = calcRowForecast(row);

  // Economics
  const taxOnlyPct   = Number(document.getElementById('tax-pct')?.value) || 0;
  const adminOnlyPct = Number(document.getElementById('admin-pct')?.value) || 0;
  const drrPct = normalizeDrrValue(drrChanges[nmid] != null ? drrChanges[nmid] : (row.base_drr_pct || 0));
  const revenue    = f.forecastBuyouts * f.plannedClient;
  const cogsTotal  = f.forecastBuyouts * (row.cogs || 0);
  const wbComm     = revenue * ((row.wb_commission_pct || 0) / 100);
  const acquiring  = revenue * 0.03;
  const drrCost    = revenue * (drrPct / 100);
  const grossProfit = revenue - cogsTotal - wbComm - acquiring - drrCost;
  const marginPct   = revenue > 0 ? grossProfit / revenue * 100 : 0;
  const adminCost   = revenue * (adminOnlyPct / 100);
  const taxCost     = revenue * (taxOnlyPct / 100);
  const netProfit   = grossProfit - adminCost - taxCost;
  const profitPct   = revenue > 0 ? netProfit / revenue * 100 : 0;

  rowForecastCache[nmid] = {{
    forecastOrders: f.forecastOrders,
    forecastBuyouts: f.forecastBuyouts,
    ordersRevenue: f.forecastOrders * f.effectiveCheck,
    forecastRevenue: f.forecastRevenue,
    lostOrders: f.lostOrders,
    lostRevenue: f.lostRevenue,
    grossProfit,
    netProfit,
    econRevenue: revenue,
    drrCost,
  }};
  const tr = tableEl.querySelector(`tr[data-nmid="${{nmid}}"]`);
  if (!tr) return;
  tr.querySelector('.cell-ps').textContent = formatMoney(f.plannedSeller);
  tr.querySelector('.cell-pc').textContent = formatMoney(f.plannedClient);
  tr.querySelector('.cell-pspd').textContent = formatNumber(f.plannedSpeed, 2);
  tr.querySelector('.cell-fo').textContent = formatNumber(f.forecastOrders);
  tr.querySelector('.cell-fb').textContent = formatNumber(f.forecastBuyouts);
  tr.querySelector('.cell-fr').textContent = formatMoney(f.forecastRevenue);
  tr.querySelector('.cell-oos').textContent = f.stockoutDate || '—';
  tr.querySelector('.cell-lo').textContent = formatNumber(f.lostOrders);
  tr.querySelector('.cell-lr').textContent = formatMoney(f.lostRevenue);

  // Economics cells
  function setCell(sel, text, val, colorVal) {{
    const el = tr.querySelector(sel);
    if (!el) return;
    el.textContent = text;
    if (val !== undefined) el.dataset[Object.keys(el.dataset)[0] || 'v'] = val;
    if (colorVal !== undefined) el.style.color = colorVal >= 0 ? '#166534' : '#dc2626';
  }}
  tr.querySelector('.cell-econ-comm-rub') && (tr.querySelector('.cell-econ-comm-rub').textContent = formatMoney(wbComm));
  tr.querySelector('.cell-econ-acq-rub')  && (tr.querySelector('.cell-econ-acq-rub').textContent  = formatMoney(acquiring));
  tr.querySelector('.cell-econ-drr-rub')  && (tr.querySelector('.cell-econ-drr-rub').textContent  = formatMoney(drrCost));
  tr.querySelector('.cell-econ-cogs')     && (tr.querySelector('.cell-econ-cogs').textContent     = formatMoney(cogsTotal));
  const marginCell = tr.querySelector('.cell-econ-margin');
  if (marginCell) {{
    marginCell.textContent = marginPct.toFixed(1) + '%';
    marginCell.dataset.margin = marginPct.toFixed(1);
    marginCell.style.color = marginPct >= 0 ? '#166534' : '#dc2626';
  }}
  tr.querySelector('.cell-econ-margin-rub') && (tr.querySelector('.cell-econ-margin-rub').textContent = formatMoney(grossProfit));
  tr.querySelector('.cell-econ-admin-rub')  && (tr.querySelector('.cell-econ-admin-rub').textContent  = formatMoney(adminCost));
  tr.querySelector('.cell-econ-tax-rub')    && (tr.querySelector('.cell-econ-tax-rub').textContent    = formatMoney(taxCost));
  const profitPctCell = tr.querySelector('.cell-econ-profit-pct');
  if (profitPctCell) {{
    profitPctCell.textContent = profitPct.toFixed(1) + '%';
    profitPctCell.dataset.profitPct = profitPct.toFixed(1);
    profitPctCell.style.color = profitPct >= 0 ? '#166534' : '#dc2626';
  }}
  const profitCell = tr.querySelector('.cell-econ-profit');
  if (profitCell) {{
    profitCell.textContent = formatMoney(netProfit);
    profitCell.dataset.profit = netProfit.toFixed(0);
    profitCell.style.color = netProfit >= 0 ? '#166534' : '#dc2626';
  }}

  // WB stock warning
  const {{ transit_days }} = planningParams;
  const daysLeft = f.plannedSpeed > 0 ? row.stock / f.plannedSpeed : Infinity;
  const stockCell = tr.querySelector('.cell-wbstock');
  if (stockCell) {{
    const warn = isFinite(daysLeft) && daysLeft < transit_days && (row.warehouse_stock || 0) > 0;
    stockCell.style.color = warn ? '#dc2626' : '';
    stockCell.title = warn ? `Кончится через ~${{daysLeft.toFixed(1)}} дн., срок поставки ${{transit_days}} дн.` : '';
  }}

  // arrow indicators for drr / conv inputs
  function applyValArrow(input, arrowEl, baseVal) {{
    if (!input || !arrowEl) return;
    const step = parseFloat(input.step) || 1;
    const dec = step <= 0.15 ? 2 : 1;
    // Compare as strings at the same precision to avoid float mismatch (e.g. base=36.65 → toFixed(1)="36.6" but Math.round(36.65*10)=367)
    const curStr  = parseFloat(input.value).toFixed(dec);
    const baseStr = (baseVal || 0).toFixed(dec);
    const cur  = parseFloat(curStr);
    const base = parseFloat(baseStr);
    input.classList.remove('val-changed-up', 'val-changed-down');
    arrowEl.classList.remove('val-up', 'val-down');
    arrowEl.textContent = '';
    if (cur !== base) {{
      const up = cur > base;
      input.classList.add(up ? 'val-changed-up' : 'val-changed-down');
      arrowEl.classList.add(up ? 'val-up' : 'val-down');
      arrowEl.textContent = up ? '↑' : '↓';
    }}
  }}
  applyValArrow(tr.querySelector('.drr-input'), tr.querySelector('.drr-arrow'), row.base_drr_pct);
  applyValArrow(tr.querySelector('.conv-input[data-conv="ctr"]'), tr.querySelector('.ctr-arrow'), row.base_ctr);
  applyValArrow(tr.querySelector('.conv-input[data-conv="cr1"]'), tr.querySelector('.cr1-arrow'), row.base_cr1);
  applyValArrow(tr.querySelector('.conv-input[data-conv="cr2"]'), tr.querySelector('.cr2-arrow'), row.base_cr2);

  if (!skipSummary) recalcSummary();
}}

function applyWeeksFilter() {{
  const max = Number(fields.weeksFilter.value);
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const row = rowDataByNmid[tr.dataset.nmid];
    if (!row) return;
    const weeks = row.base_speed > 0 ? row.stock / row.base_speed / 7 : Infinity;
    tr.style.display = (!max || weeks <= max) ? '' : 'none';
  }});
}}

function applyElasticityFilter() {{
  const val = fields.elasticityFilter.value;
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const hasElasticity = row.elasticity !== 0;
    const visible = !val || (val === 'elastic' ? hasElasticity : !hasElasticity);
    tr.style.display = visible ? '' : 'none';
  }});
}}

function updateBulkBar() {{
  const checked = tableEl.querySelectorAll('input.row-ck:checked');
  const n = checked.length;
  bulkCount.textContent = n;
  bulkBar.style.display = n > 0 ? 'flex' : 'none';
}}

function renderTable(rows) {{
  if (!rows.length) {{
    tableEl.innerHTML = '<div class="empty">Нет товаров для выбранных фильтров</div>';
    return;
  }}
  rowDataByNmid = {{}};
  rowForecastCache = {{}};
  const _taxPctInit   = Number(document.getElementById('tax-pct')?.value) || 0;
  const _adminPctInit = Number(document.getElementById('admin-pct')?.value) || 0;
  rows.forEach(r => {{
    rowDataByNmid[r.nmid] = r;
    const _rev = r.forecast_buyouts * (r.planned_price || r.client_price || r.average_check || 0);
    const _cogs = r.forecast_buyouts * (r.cogs || 0);
    const _comm = _rev * ((r.wb_commission_pct || 0) / 100);
    const _acq  = _rev * 0.03;
    const _drr  = _rev * ((r.base_drr_pct || 0) / 100);
    const _gross = _rev - _cogs - _comm - _acq - _drr;
    const _net   = _gross - _rev * (_adminPctInit / 100) - _rev * (_taxPctInit / 100);
    rowForecastCache[r.nmid] = {{
      forecastOrders: r.forecast_orders,
      forecastBuyouts: r.forecast_buyouts,
      ordersRevenue: r.forecast_orders * (r.effective_check || r.average_check),
      forecastRevenue: r.forecast_revenue,
      lostOrders: r.lost_orders,
      lostRevenue: r.lost_revenue,
      grossProfit: _gross,
      netProfit: _net,
      econRevenue: _rev,
    }};
  }});

  const head = `<tr>
    <th class="col-ck" rowspan="2"><input type="checkbox" id="ck-all" title="Выбрать все"></th>
    <th class="col-sku" rowspan="2" data-sort-key="article">SKU</th>
    <th class="col-subject" rowspan="2" data-sort-key="subject">Предмет</th>
    <th class="col-text" rowspan="2" data-sort-key="strategy">Ярлыки</th>
    <th class="col-text" rowspan="2" data-sort-key="revenue_category">Категория</th>
    <th class="g-sep" colspan="3" data-col-group="остатки">Остатки</th>
    <th class="g-sep" colspan="4" data-col-group="базовый">Базовый период 7д</th>
    <th class="g-sep" colspan="5" data-col-group="текущее">Текущее</th>
    <th class="g-sep th-params" colspan="6" data-col-group="параметры">Параметры плана</th>
    <th class="g-sep th-result" colspan="7" data-col-group="результат">Результат</th>
    <th class="g-sep th-lost" colspan="3" data-col-group="упущено">Упущено</th>
    <th class="g-sep th-econ" colspan="10" data-col-group="экономика">Экономика</th>
  </tr><tr>
    <th class="g-sep" data-col-group="остатки" data-col-key="c-stock" data-sort-key="stock">WB, шт</th><th data-col-group="остатки" data-col-key="c-wh-stock" data-sort-key="warehouse_stock">Склад, шт</th><th data-col-group="остатки" data-col-key="c-stock-val" data-sort-key="stock_value">Сумма, ₽</th>
    <th class="g-sep" data-col-group="базовый" data-col-key="c-base-price" data-sort-key="baseline_seller_price">Цена, ₽</th><th data-col-group="базовый" data-col-key="c-base-spp" data-sort-key="baseline_spp">СПП, %</th><th data-col-group="базовый" data-col-key="c-base-client" data-sort-key="baseline_client_check">Кл. цена, ₽</th><th data-col-group="базовый" data-col-key="c-base-speed" data-sort-key="base_speed">Скорость, шт</th>
    <th class="g-sep" data-col-group="текущее" data-col-key="c-cur-price" data-sort-key="current_price">Тек. цена, ₽</th><th data-col-group="текущее" data-col-key="c-cur-spp" data-sort-key="current_spp">СПП, %</th><th data-col-group="текущее" data-col-key="c-cur-client" data-sort-key="client_price">Кл. цена, ₽</th><th data-col-group="текущее" data-col-key="c-cur-elast" data-sort-key="elasticity">Эл. цены</th><th data-col-group="текущее" data-col-key="c-cur-speed" data-sort-key="current_speed">Скорость, шт</th>
    <th class="g-sep params-col" data-col-group="параметры" data-col-key="c-plan-price">Цена, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-drr" data-sort-key="base_drr_pct">ДРР, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-drr-e" data-sort-key="ads_elasticity">Эл.рекл.</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-ctr" data-sort-key="base_ctr">CTR, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-cr1" data-sort-key="base_cr1">CR1, %</th><th class="params-col" data-col-group="параметры" data-col-key="c-plan-cr2" data-sort-key="base_cr2">CR2, %</th>
    <th class="g-sep result-col" data-col-group="результат" data-col-key="c-res-price">Пл. цена, ₽</th><th class="result-col" data-col-group="результат" data-col-key="c-res-client">Пл. кл., ₽</th><th class="result-col" data-col-group="результат" data-col-key="c-res-speed" data-sort-key="plannedSpeed">Скорость, шт</th><th class="result-col" data-col-group="результат" data-col-key="c-res-buyout" data-sort-key="buyout_percent">% выкупа</th><th class="result-col" data-col-group="результат" data-col-key="c-res-orders" data-sort-key="forecast_orders">Заказы, шт</th><th class="result-col" data-col-group="результат" data-col-key="c-res-buyouts" data-sort-key="forecast_buyouts">Выкупы, шт</th><th class="result-col" data-col-group="результат" data-col-key="c-res-revenue" data-sort-key="forecast_revenue">Выручка, ₽</th>
    <th class="g-sep lost-col" data-col-group="упущено" data-col-key="c-res-oos">Дата стопа</th><th class="lost-col" data-col-group="упущено" data-col-key="c-lost-qty" data-sort-key="lost_orders">Упущ., шт</th><th class="lost-col" data-col-group="упущено" data-col-key="c-lost-rub" data-sort-key="lost_revenue">Упущ., ₽</th>
    <th class="g-sep econ-col" data-col-group="экономика" data-col-key="c-econ-comm-pct" data-sort-key="wb_commission_pct">Ком. %</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-comm-rub" data-sort-key="comm_rub">Ком. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-acq-rub" data-sort-key="acq_rub">Эквайр. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-drr-rub" data-sort-key="drr_rub">Реклама ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-cogs" data-sort-key="cogs">Себест. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-margin-pct" data-sort-key="margin_pct">Маржа %</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-margin-rub" data-sort-key="margin_rub">Маржа ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-admin-rub" data-sort-key="admin_rub">Адм. ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-tax-rub" data-sort-key="tax_rub">Налоги ₽</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-profit-pct" data-sort-key="profit_pct">Прибыль %</th><th class="econ-col" data-col-group="экономика" data-col-key="c-econ-profit-rub" data-sort-key="net_profit">Прибыль ₽</th>
  </tr>`;

  const body = rows.map(row => {{
    const pct = priceChanges[row.nmid] ?? 0;
    const drrVal = normalizeDrrValue(drrChanges[row.nmid] != null ? drrChanges[row.nmid] : (row.base_drr_pct || 0));
    const ctrVal = (convChanges[row.nmid] || {{}}).ctr != null ? (convChanges[row.nmid] || {{}}).ctr : (row.base_ctr || 0);
    const cr1Val = (convChanges[row.nmid] || {{}}).cr1 != null ? (convChanges[row.nmid] || {{}}).cr1 : (row.base_cr1 || 0);
    const cr2Val = (convChanges[row.nmid] || {{}}).cr2 != null ? (convChanges[row.nmid] || {{}}).cr2 : (row.base_cr2 || 0);
    const {{ plannedSeller, plannedClient, plannedSpeed }} = calcPlanned(row, pct);
    return `<tr data-nmid="${{row.nmid}}">
      <td class="col-ck"><input type="checkbox" class="row-ck"></td>
      <td class="col-sku" title="${{row.article || row.nmid}}">${{row.article || row.nmid}}</td>
      <td class="col-subject" title="${{row.subject || ''}}">${{row.subject || ''}}</td>
      <td class="col-text">${{(row.labels||[]).join(', ')}}</td>
      <td class="col-text">${{row.revenue_category || ''}}</td>
      <td class="g-sep cell-wbstock" data-col-group="остатки" data-col-key="c-stock">${{formatNumber(row.stock)}}</td>
      <td data-col-group="остатки" data-col-key="c-wh-stock">${{formatNumber(row.warehouse_stock)}}</td>
      <td data-col-group="остатки" data-col-key="c-stock-val">${{formatMoney((row.stock + (row.warehouse_stock || 0)) * (row.client_price || row.average_check || 0))}}</td>
      <td class="g-sep" data-col-group="базовый" data-col-key="c-base-price">${{formatMoney(row.baseline_seller_price)}}</td>
      <td data-col-group="базовый" data-col-key="c-base-spp">${{formatNumber(row.baseline_spp, 1)}}%</td>
      <td data-col-group="базовый" data-col-key="c-base-client">${{formatMoney(row.baseline_client_check)}}</td>
      <td data-col-group="базовый" data-col-key="c-base-speed">${{formatNumber(row.base_speed, 2)}}</td>
      <td class="g-sep" data-col-group="текущее" data-col-key="c-cur-price">${{formatMoney(row.current_price)}}${{priceArrow(row.current_price, row.baseline_seller_price)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-spp">${{formatNumber(row.current_spp, 1)}}%${{priceArrow(row.current_spp, row.baseline_spp, true)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-client">${{formatMoney(row.client_price)}}${{priceArrow(row.client_price, row.baseline_client_check)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-elast">${{formatNumber(row.elasticity, 2)}}</td>
      <td data-col-group="текущее" data-col-key="c-cur-speed">${{formatNumber(row.current_speed, 2)}}</td>
      <td class="g-sep" data-col-group="параметры" data-col-key="c-plan-price">
        <div class="pct-cell">
          <div class="pct-spinner">
            <input type="number" class="pct-input" value="${{pct}}" step="1" min="-90" max="300">
            <div class="pct-arrows">
              <button class="pct-btn" data-d="1">▲</button>
              <button class="pct-btn" data-d="-1">▼</button>
            </div>
          </div>
        </div>
      </td>
      <td data-col-group="параметры" data-col-key="c-plan-drr">
        <div class="pct-cell" title="Базовый ДРР: ${{formatNumber(row.base_drr_pct, 1)}}%">
          <span class="val-arrow drr-arrow"></span>
          <div class="pct-spinner">
            <input type="number" class="drr-input" value="${{(+drrVal || 0).toFixed(1)}}" step="0.5" min="0" max="50">
            <div class="pct-arrows">
              <button class="drr-btn" data-d="0.5">▲</button>
              <button class="drr-btn" data-d="-0.5">▼</button>
            </div>
          </div>
        </div>
      </td>
      <td class="cell-drr-e" data-col-group="параметры" data-col-key="c-plan-drr-e">${{row.ads_elasticity ? formatNumber(row.ads_elasticity, 2) : '—'}}</td>
      <td data-col-group="параметры" data-col-key="c-plan-ctr">
        ${{(function(){{
          const noBase = !row.base_ctr && !(convChanges[row.nmid]||{{}}).ctr;
          const v = noBase ? '' : (+ctrVal||0).toFixed(2);
          return `<div class="pct-cell" title="${{noBase ? 'Нет данных за период' : 'Базовый CTR: '+formatNumber(row.base_ctr,2)+'%'}}">
            <span class="val-arrow ctr-arrow"></span>
            <div class="pct-spinner${{noBase?' no-data':''}}">
              <input type="number" class="conv-input" data-conv="ctr" value="${{v}}" placeholder="${{noBase?'—':''}}" step="0.1" min="0" max="100" ${{noBase?'disabled':''}}>
              <div class="pct-arrows"><button class="conv-btn" data-conv="ctr" data-d="0.1">▲</button><button class="conv-btn" data-conv="ctr" data-d="-0.1">▼</button></div>
            </div></div>`;
        }})()}}
      </td>
      <td data-col-group="параметры" data-col-key="c-plan-cr1">
        ${{(function(){{
          const noBase = !row.base_cr1 && !(convChanges[row.nmid]||{{}}).cr1;
          const v = noBase ? '' : (+cr1Val||0).toFixed(2);
          return `<div class="pct-cell" title="${{noBase ? 'Нет данных за период' : 'Базовый CR1: '+formatNumber(row.base_cr1,2)+'%'}}">
            <span class="val-arrow cr1-arrow"></span>
            <div class="pct-spinner${{noBase?' no-data':''}}">
              <input type="number" class="conv-input" data-conv="cr1" value="${{v}}" placeholder="${{noBase?'—':''}}" step="0.1" min="0" max="100" ${{noBase?'disabled':''}}>
              <div class="pct-arrows"><button class="conv-btn" data-conv="cr1" data-d="0.1">▲</button><button class="conv-btn" data-conv="cr1" data-d="-0.1">▼</button></div>
            </div></div>`;
        }})()}}
      </td>
      <td data-col-group="параметры" data-col-key="c-plan-cr2">
        ${{(function(){{
          const noBase = !row.base_cr2 && !(convChanges[row.nmid]||{{}}).cr2;
          const v = noBase ? '' : (+cr2Val||0).toFixed(2);
          return `<div class="pct-cell" title="${{noBase ? 'Нет данных за период' : 'Базовый CR2: '+formatNumber(row.base_cr2,2)+'%'}}">
            <span class="val-arrow cr2-arrow"></span>
            <div class="pct-spinner${{noBase?' no-data':''}}">
              <input type="number" class="conv-input" data-conv="cr2" value="${{v}}" placeholder="${{noBase?'—':''}}" step="0.1" min="0" max="100" ${{noBase?'disabled':''}}>
              <div class="pct-arrows"><button class="conv-btn" data-conv="cr2" data-d="0.1">▲</button><button class="conv-btn" data-conv="cr2" data-d="-0.1">▼</button></div>
            </div></div>`;
        }})()}}
      </td>
      <td class="g-sep cell-ps" data-col-group="результат" data-col-key="c-res-price">${{formatMoney(plannedSeller)}}</td>
      <td class="cell-pc" data-col-group="результат" data-col-key="c-res-client">${{formatMoney(plannedClient)}}</td>
      <td class="cell-pspd" data-col-group="результат" data-col-key="c-res-speed">${{formatNumber(plannedSpeed, 2)}}</td>
      <td data-col-group="результат" data-col-key="c-res-buyout">${{formatNumber(row.buyout_percent)}}%</td>
      <td class="cell-fo" data-col-group="результат" data-col-key="c-res-orders">${{formatNumber(row.forecast_orders)}}</td>
      <td class="cell-fb" data-col-group="результат" data-col-key="c-res-buyouts">${{formatNumber(row.forecast_buyouts)}}</td>
      <td class="cell-fr" data-col-group="результат" data-col-key="c-res-revenue">${{formatMoney(row.forecast_revenue)}}</td>
      <td class="cell-oos g-sep" data-col-group="упущено" data-col-key="c-res-oos">${{row.stockout_date || '—'}}</td>
      <td class="g-sep cell-lo" data-col-group="упущено" data-col-key="c-lost-qty">${{formatNumber(row.lost_orders)}}</td>
      <td class="cell-lr" data-col-group="упущено" data-col-key="c-lost-rub">${{formatMoney(row.lost_revenue)}}</td>
      <td class="g-sep econ-col cell-wb-comm" data-col-group="экономика" data-col-key="c-econ-comm-pct">${{formatNumber(row.wb_commission_pct || 0, 1)}}%</td>
      <td class="econ-col cell-econ-comm-rub" data-col-group="экономика" data-col-key="c-econ-comm-rub">—</td>
      <td class="econ-col cell-econ-acq-rub" data-col-group="экономика" data-col-key="c-econ-acq-rub">—</td>
      <td class="econ-col cell-econ-drr-rub" data-col-group="экономика" data-col-key="c-econ-drr-rub">—</td>
      <td class="econ-col cell-econ-cogs" data-col-group="экономика" data-col-key="c-econ-cogs">—</td>
      <td class="econ-col cell-econ-margin" data-col-group="экономика" data-col-key="c-econ-margin-pct" data-margin="0">—</td>
      <td class="econ-col cell-econ-margin-rub" data-col-group="экономика" data-col-key="c-econ-margin-rub">—</td>
      <td class="econ-col cell-econ-admin-rub" data-col-group="экономика" data-col-key="c-econ-admin-rub">—</td>
      <td class="econ-col cell-econ-tax-rub" data-col-group="экономика" data-col-key="c-econ-tax-rub">—</td>
      <td class="econ-col cell-econ-profit-pct" data-col-group="экономика" data-col-key="c-econ-profit-pct" data-profit-pct="0">—</td>
      <td class="econ-col cell-econ-profit" data-col-group="экономика" data-col-key="c-econ-profit-rub" data-profit="0">—</td>
    </tr>`;
  }}).join('');

  tableEl.innerHTML = `<table><thead>${{head}}</thead><tbody>${{body}}</tbody></table>`;

  // checkbox: select all
  document.getElementById('ck-all').addEventListener('change', e => {{
    tableEl.querySelectorAll('input.row-ck').forEach(ck => {{
      ck.checked = e.target.checked;
      ck.closest('tr')?.classList.toggle('row-selected', e.target.checked);
    }});
    updateBulkBar();
  }});

  // checkbox: individual
  tableEl.addEventListener('change', e => {{
    if (e.target.classList.contains('row-ck')) {{
      e.target.closest('tr')?.classList.toggle('row-selected', e.target.checked);
      updateBulkBar();
    }}
  }});

  // price pct input & buttons
  tableEl.addEventListener('input', e => {{
    const tr = e.target.closest('tr');
    if (!tr) return;
    const nmid = tr.dataset.nmid;
    if (e.target.classList.contains('pct-input')) {{
      priceChanges[nmid] = Number(e.target.value) || 0;
      savePriceChanges();
      updateRowPlanned(nmid);
    }} else if (e.target.classList.contains('drr-input')) {{
      const nv = normalizeDrrValue(e.target.value);
      e.target.value = nv.toFixed(1);
      drrChanges[nmid] = nv;
      saveDrrChanges();
      updateRowPlanned(nmid);
    }} else if (e.target.classList.contains('conv-input')) {{
      const conv = e.target.dataset.conv;
      const val = parseFloat(e.target.value);
      if (!convChanges[nmid]) convChanges[nmid] = {{}};
      if (val > 0) {{
        convChanges[nmid][conv] = val;
      }} else {{
        delete convChanges[nmid][conv];
      }}
      saveConvChanges();
      updateRowPlanned(nmid);
    }}
  }});
  tableEl.addEventListener('click', e => {{
    // arrow click → reset that value to baseline
    const arrow = e.target.closest('.val-arrow');
    if (arrow && arrow.textContent) {{
      const tr = arrow.closest('tr');
      const nmid = tr?.dataset.nmid;
      const row = rowDataByNmid[nmid];
      if (!row) return;
      if (arrow.classList.contains('drr-arrow')) {{
        delete drrChanges[nmid];
        saveDrrChanges();
        const inp = tr.querySelector('.drr-input');
        if (inp) inp.value = normalizeDrrValue(row.base_drr_pct || 0).toFixed(1);
      }} else {{
        let conv = null;
        if (arrow.classList.contains('ctr-arrow')) conv = 'ctr';
        else if (arrow.classList.contains('cr1-arrow')) conv = 'cr1';
        else if (arrow.classList.contains('cr2-arrow')) conv = 'cr2';
        if (conv) {{
          if (convChanges[nmid]) {{
            delete convChanges[nmid][conv];
            if (!Object.keys(convChanges[nmid]).length) delete convChanges[nmid];
          }}
          saveConvChanges();
          const inp = tr.querySelector(`.conv-input[data-conv="${{conv}}"]`);
          if (inp && !inp.disabled) inp.value = (row[`base_${{conv}}`] || 0).toFixed(2);
        }}
      }}
      updateRowPlanned(nmid);
      return;
    }}
    const btn = e.target.closest('.pct-btn, .drr-btn, .conv-btn');
    if (!btn) return;
    const tr = btn.closest('tr');
    const nmid = tr.dataset.nmid;
    if (btn.classList.contains('pct-btn')) {{
      const input = tr.querySelector('.pct-input');
      const newVal = Math.max(-90, Math.min(300, (Number(input.value) || 0) + Number(btn.dataset.d)));
      input.value = newVal;
      priceChanges[nmid] = newVal;
      savePriceChanges();
    }} else if (btn.classList.contains('drr-btn')) {{
      const input = tr.querySelector('.drr-input');
      const newVal = normalizeDrrValue((Number(input.value) || 0) + Number(btn.dataset.d));
      input.value = newVal.toFixed(1);
      drrChanges[nmid] = newVal;
      saveDrrChanges();
    }} else if (btn.classList.contains('conv-btn')) {{
      const conv = btn.dataset.conv;
      const input = tr.querySelector(`.conv-input[data-conv="${{conv}}"]`);
      const newVal = Math.max(0, Math.min(100, Math.round(((Number(input.value) || 0) + Number(btn.dataset.d)) * 10) / 10));
      input.value = newVal;
      if (!convChanges[nmid]) convChanges[nmid] = {{}};
      convChanges[nmid][conv] = newVal;
      saveConvChanges();
    }}
    updateRowPlanned(nmid);
  }});
}}

// bulk bar actions
const bulkType = document.getElementById('bulk-type');
document.getElementById('bulk-apply').addEventListener('click', () => {{
  const delta = Number(bulkPct.value) || 0;
  const type = bulkType?.value || 'price';
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const ck = tr.querySelector('input.row-ck');
    if (!ck?.checked) return;
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (type === 'price') {{
      const cur = priceChanges[nmid] ?? 0;
      const nv = Math.max(-90, Math.min(300, Math.round((cur + delta) * 10) / 10));
      priceChanges[nmid] = nv;
      const inp = tr.querySelector('.pct-input');
      if (inp) inp.value = nv;
    }} else if (type === 'drr') {{
      const cur = normalizeDrrValue(drrChanges[nmid] != null ? drrChanges[nmid] : (row?.base_drr_pct || 0));
      const nv = normalizeDrrValue(cur + delta);
      drrChanges[nmid] = nv;
      const inp = tr.querySelector('.drr-input');
      if (inp) inp.value = nv.toFixed(1);
    }} else {{
      if (!convChanges[nmid]) convChanges[nmid] = {{}};
      const cur = convChanges[nmid][type] > 0 ? convChanges[nmid][type] : (row?.[`base_${{type}}`] || 0);
      const nv = Math.max(0, Math.min(100, Math.round((cur + delta) * 100) / 100));
      convChanges[nmid][type] = nv;
      const inp = tr.querySelector(`.conv-input[data-conv="${{type}}"]`);
      if (inp) inp.value = nv.toFixed(2);
    }}
    updateRowPlanned(nmid);
  }});
  savePriceChanges(); saveDrrChanges(); saveConvChanges();
}});
document.getElementById('bulk-reset').addEventListener('click', () => {{
  const type = bulkType?.value || 'price';
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const ck = tr.querySelector('input.row-ck');
    if (!ck?.checked) return;
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (type === 'price') {{
      priceChanges[nmid] = 0;
      const inp = tr.querySelector('.pct-input'); if (inp) inp.value = 0;
    }} else if (type === 'drr') {{
      delete drrChanges[nmid];
      const inp = tr.querySelector('.drr-input');
      if (inp) inp.value = normalizeDrrValue(row?.base_drr_pct || 0).toFixed(1);
    }} else {{
      if (convChanges[nmid]) delete convChanges[nmid][type];
      const inp = tr.querySelector(`.conv-input[data-conv="${{type}}"]`);
      if (inp) inp.value = row?.[`base_${{type}}`] || 0;
    }}
    updateRowPlanned(nmid);
  }});
  savePriceChanges(); saveDrrChanges(); saveConvChanges();
}});
document.getElementById('no-oos-btn').addEventListener('click', () => {{
  const selected = new Set([...tableEl.querySelectorAll('input.row-ck:checked')].map(ck => ck.closest('tr')?.dataset.nmid).filter(Boolean));
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    if (selected.size > 0 && !selected.has(nmid)) return;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const pct = calcNoOOS(row);
    if (pct === null || pct === 0) return;
    priceChanges[nmid] = pct;
    const input = tr.querySelector('.pct-input');
    if (input) input.value = pct;
    updateRowPlanned(nmid);
  }});
  savePriceChanges();
}});
document.getElementById('optimize-btn').addEventListener('click', () => {{
  const btn = document.getElementById('optimize-btn');
  btn.disabled = true;
  const selected = new Set([...tableEl.querySelectorAll('input.row-ck:checked')]
    .map(ck => ck.closest('tr')?.dataset.nmid).filter(Boolean));
  const rows = [...tableEl.querySelectorAll('tr[data-nmid]')]
    .map(tr => tr.dataset.nmid)
    .filter(nmid => selected.size === 0 || selected.has(nmid))
    .filter(nmid => rowDataByNmid[nmid]);
  const total = rows.length;
  let idx = 0;
  const CHUNK = 5;
  function processChunk() {{
    const end = Math.min(idx + CHUNK, total);
    for (; idx < end; idx++) {{
      const nmid = rows[idx];
      const row = rowDataByNmid[nmid];
      const {{ pricePct, drrPct }} = findOptimalRevenue(row);
      priceChanges[nmid] = pricePct;
      drrChanges[nmid] = drrPct;
      const tr = tableEl.querySelector(`tr[data-nmid="${{nmid}}"]`);
      if (tr) {{
        const pInp = tr.querySelector('.pct-input');
        if (pInp) pInp.value = pricePct;
        const dInp = tr.querySelector('.drr-input');
        if (dInp) dInp.value = drrPct.toFixed(1);
        updateRowPlanned(nmid);
      }}
    }}
    btn.textContent = `⏳ ${{idx}}/${{total}}…`;
    if (idx < total) {{
      setTimeout(processChunk, 0);
    }} else {{
      savePriceChanges();
      saveDrrChanges();
      const sname = 'Макс. выручка';
      const all = JSON.parse(localStorage.getItem(scenariosKey) || '{{}}');
      all[sname] = {{
        price: Object.assign({{}}, priceChanges),
        drr:   Object.assign({{}}, drrChanges),
        conv:  JSON.parse(JSON.stringify(convChanges)),
      }};
      localStorage.setItem(scenariosKey, JSON.stringify(all));
      loadScenarioList();
      document.getElementById('scenario-select').value = sname;
      btn.disabled = false; btn.textContent = '⚡ Макс. выручку';
    }}
  }}
  setTimeout(processChunk, 0);
}});
document.getElementById('reset-prices').addEventListener('click', () => {{
  priceChanges = {{}};
  savePriceChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const input = tr.querySelector('.pct-input');
    if (input) input.value = 0;
    updateRowPlanned(nmid);
  }});
}});
document.getElementById('reset-drr').addEventListener('click', () => {{
  drrChanges = {{}};
  saveDrrChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    const drrInp = tr.querySelector('.drr-input');
    if (drrInp) drrInp.value = normalizeDrrValue(row.base_drr_pct || 0).toFixed(1);
    updateRowPlanned(nmid, true);
  }});
  recalcSummary();
}});
document.getElementById('reset-conv').addEventListener('click', () => {{
  convChanges = {{}};
  saveConvChanges();
  tableEl.querySelectorAll('tr[data-nmid]').forEach(tr => {{
    const nmid = tr.dataset.nmid;
    const row = rowDataByNmid[nmid];
    if (!row) return;
    ['ctr','cr1','cr2'].forEach(k => {{
      const inp = tr.querySelector(`.conv-input[data-conv="${{k}}"]`);
      if (inp && !inp.disabled) inp.value = (row[`base_${{k}}`] || 0).toFixed(2);
    }});
    updateRowPlanned(nmid, true);
  }});
  recalcSummary();
}});
document.getElementById('bulk-close').addEventListener('click', () => {{
  tableEl.querySelectorAll('input.row-ck').forEach(ck => {{ ck.checked = false; }});
  const ckAll = document.getElementById('ck-all');
  if (ckAll) ckAll.checked = false;
  updateBulkBar();
}});

async function loadData() {{
  saveState();
  meta.textContent = 'Считаю...';
  priceChanges = {{}};
  const qs = new URLSearchParams({{
    forecast_to: fields.forecastTo.value,
    return_delay_days: fields.returnDelay.value || '8',
    transit_days: fields.transitDays.value || '3',
    subject: fields.subject.value,
    labels: [...(fields.strategy?.selectedOptions||[])].map(o=>o.value).join(","),
    revenue_category: fields.revenueCategory.value,
    price_change_pct: '0',
  }});
  const resp = await fetch(`/api/analytics/planning?${{qs.toString()}}`);
  const data = await resp.json();
  if (!resp.ok) {{
    meta.textContent = data.error || `Ошибка ${{resp.status}}`;
    summaryEl.innerHTML = '';
    tableEl.innerHTML = '';
    return;
  }}
  fillSelect(fields.subject, data.options?.subjects || [], 'Все предметы');
  fillMultiSelect(fields.strategy, data.options?.labels || []);
  fillSelect(fields.revenueCategory, data.options?.revenue_categories || [], 'Все категории');
  planningParams = {{
    forecast_days: data.forecast_days || 0,
    return_delay_days: data.return_delay_days || 8,
    transit_days: data.transit_days ?? 3,
    forecast_from: data.forecast_from || '',
  }};
  renderSummary(data.summary || {{}});
  renderTable(data.rows || []);
  // fill economics + WB stock warning for all rows — skip per-row summary recalc, do one pass at the end
  Object.keys(rowDataByNmid).forEach(nmid => updateRowPlanned(nmid, true));
  recalcSummary();
  applyWeeksFilter();
  applyColVisibility();
  meta.textContent = `База скорости: ${{data.baseline_from}}..${{data.baseline_to}} | % выкупа WB за этот период | Прогноз: ${{data.forecast_from}}..${{data.forecast_to}} | Дней: ${{data.forecast_days}} | SKU: ${{(data.rows || []).length}}`;
}}

let timer = null;
function scheduleLoad() {{ clearTimeout(timer); timer = setTimeout(loadData, 250); }}
for (const el of Object.values(fields)) {{
  if (el === fields.weeksFilter) {{
    el.addEventListener('input', applyWeeksFilter);
  }} else {{
    el.addEventListener('change', loadData);
    el.addEventListener('input', scheduleLoad);
  }}
}}
document.getElementById('load').addEventListener('click', loadData);

// tax/admin: recalc all rows on change
['tax-pct', 'admin-pct'].forEach(id => {{
  document.getElementById(id).addEventListener('input', () => {{
    Object.keys(rowDataByNmid).forEach(nmid => updateRowPlanned(nmid, true));
    recalcSummary();
  }});
}});

// ——— Column visibility (gear) ———
const COL_SCHEMA = [
  {{ group:'остатки', label:'Остатки', cols:[
    {{ key:'c-stock',     label:'WB, шт' }},
    {{ key:'c-wh-stock',  label:'Склад, шт' }},
    {{ key:'c-stock-val', label:'Сумма, ₽' }},
  ]}},
  {{ group:'базовый', label:'Базовый период', cols:[
    {{ key:'c-base-price',  label:'Цена, ₽' }},
    {{ key:'c-base-spp',    label:'СПП, %' }},
    {{ key:'c-base-client', label:'Кл. цена, ₽' }},
    {{ key:'c-base-speed',  label:'Скорость, шт' }},
  ]}},
  {{ group:'текущее', label:'Текущее', cols:[
    {{ key:'c-cur-price',  label:'Цена, ₽' }},
    {{ key:'c-cur-spp',    label:'СПП, %' }},
    {{ key:'c-cur-client', label:'Кл. цена, ₽' }},
    {{ key:'c-cur-elast',  label:'Эластичность' }},
    {{ key:'c-cur-speed',  label:'Скорость, шт' }},
  ]}},
  {{ group:'параметры', label:'Параметры', cols:[
    {{ key:'c-plan-price', label:'Цена, %' }},
    {{ key:'c-plan-drr',   label:'ДРР, %' }},
    {{ key:'c-plan-drr-e', label:'Эл. рекл.' }},
    {{ key:'c-plan-ctr',   label:'CTR, %' }},
    {{ key:'c-plan-cr1',   label:'CR корзина, %' }},
    {{ key:'c-plan-cr2',   label:'CR заказ, %' }},
  ]}},
  {{ group:'результат', label:'Результат', cols:[
    {{ key:'c-res-price',   label:'Пл. цена, ₽' }},
    {{ key:'c-res-client',  label:'Пл. кл. цена, ₽' }},
    {{ key:'c-res-speed',   label:'Скорость, шт/д' }},
    {{ key:'c-res-buyout',  label:'% выкупа' }},
    {{ key:'c-res-orders',  label:'Заказы, шт' }},
    {{ key:'c-res-buyouts', label:'Выкупы, шт' }},
    {{ key:'c-res-revenue', label:'Выручка, ₽' }},
  ]}},
  {{ group:'упущено', label:'Упущенные продажи', cols:[
    {{ key:'c-res-oos',  label:'Дата стопа' }},
    {{ key:'c-lost-qty', label:'Упущ., шт' }},
    {{ key:'c-lost-rub', label:'Упущ., ₽' }},
  ]}},
  {{ group:'экономика', label:'Экономика', cols:[
    {{ key:'c-econ-comm-pct',   label:'Комиссия, %' }},
    {{ key:'c-econ-comm-rub',   label:'Комиссия, ₽' }},
    {{ key:'c-econ-acq-rub',    label:'Эквайринг, ₽' }},
    {{ key:'c-econ-drr-rub',    label:'Реклама, ₽' }},
    {{ key:'c-econ-cogs',       label:'Себестоимость, ₽' }},
    {{ key:'c-econ-margin-pct', label:'Маржа, %' }},
    {{ key:'c-econ-margin-rub', label:'Маржа, ₽' }},
    {{ key:'c-econ-admin-rub',  label:'Адм. расходы, ₽' }},
    {{ key:'c-econ-tax-rub',    label:'Налоги, ₽' }},
    {{ key:'c-econ-profit-pct', label:'Прибыль, %' }},
    {{ key:'c-econ-profit-rub', label:'Прибыль, ₽' }},
  ]}},
];

const COL_VIS_KEY = 'wb.planning.col-vis';
let colVisibility = {{}};
try {{ colVisibility = JSON.parse(localStorage.getItem(COL_VIS_KEY) || '{{}}'); }} catch(e) {{}}

const _colVisStyle = document.createElement('style');
_colVisStyle.id = 'col-vis-style';
document.head.appendChild(_colVisStyle);

function applyColVisibility() {{
  const rules = [];
  COL_SCHEMA.forEach(grp => {{
    const allHidden = grp.cols.every(c => colVisibility[c.key] === false);
    if (allHidden) {{
      rules.push(`[data-col-group="${{grp.group}}"] {{ display:none !important; }}`);
    }} else {{
      grp.cols.forEach(c => {{
        if (colVisibility[c.key] === false) {{
          rules.push(`[data-col-key="${{c.key}}"] {{ display:none !important; }}`);
        }}
      }});
    }}
  }});
  document.getElementById('col-vis-style').textContent = rules.join('\\n');
}}

function _isGroupChecked(grp) {{ return grp.cols.some(c => colVisibility[c.key] !== false); }}
function _isGroupIndeterminate(grp) {{
  const vis = grp.cols.filter(c => colVisibility[c.key] !== false).length;
  return vis > 0 && vis < grp.cols.length;
}}

const colSettingsBtn   = document.getElementById('col-settings-btn');
const colSettingsPanel = document.getElementById('col-settings-panel');

function _buildColPanel() {{
  colSettingsPanel.innerHTML = `
    <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#94a3b8;margin-bottom:8px">Столбцы таблицы</div>
    ${{COL_SCHEMA.map(grp => `
      <div style="margin-bottom:6px">
        <label style="display:flex;align-items:center;gap:6px;cursor:pointer;font-size:12px;font-weight:700;color:#334155;padding:2px 0">
          <input type="checkbox" data-grp="${{grp.group}}"
            ${{_isGroupChecked(grp) ? 'checked' : ''}}
            style="accent-color:#7e22ce"> ${{grp.label}}
        </label>
        <div style="padding-left:18px">
          ${{grp.cols.map(c => `
            <label style="display:flex;align-items:center;gap:5px;cursor:pointer;font-size:12px;color:#475569;padding:1px 0">
              <input type="checkbox" data-col="${{c.key}}"
                ${{colVisibility[c.key] !== false ? 'checked' : ''}}
                style="accent-color:#7e22ce"> ${{c.label}}
            </label>`).join('')}}
        </div>
      </div>`).join('')}}`;

  // group toggle
  colSettingsPanel.querySelectorAll('input[data-grp]').forEach(cb => {{
    const grp = COL_SCHEMA.find(g => g.group === cb.dataset.grp);
    if (!grp) return;
    // set indeterminate state
    if (_isGroupIndeterminate(grp)) cb.indeterminate = true;
    cb.addEventListener('change', () => {{
      grp.cols.forEach(c => {{
        colVisibility[c.key] = cb.checked;
        const colCb = colSettingsPanel.querySelector(`input[data-col="${{c.key}}"]`);
        if (colCb) colCb.checked = cb.checked;
      }});
      localStorage.setItem(COL_VIS_KEY, JSON.stringify(colVisibility));
      applyColVisibility();
    }});
  }});

  // individual column toggle
  colSettingsPanel.querySelectorAll('input[data-col]').forEach(cb => {{
    cb.addEventListener('change', () => {{
      colVisibility[cb.dataset.col] = cb.checked;
      localStorage.setItem(COL_VIS_KEY, JSON.stringify(colVisibility));
      applyColVisibility();
      // update parent group checkbox
      const grp = COL_SCHEMA.find(g => g.cols.some(c => c.key === cb.dataset.col));
      if (grp) {{
        const grpCb = colSettingsPanel.querySelector(`input[data-grp="${{grp.group}}"]`);
        if (grpCb) {{
          const checked = _isGroupChecked(grp);
          const indet = _isGroupIndeterminate(grp);
          grpCb.checked = checked;
          grpCb.indeterminate = indet;
        }}
      }}
    }});
  }});
}}

colSettingsBtn.addEventListener('click', e => {{
  e.stopPropagation();
  if (colSettingsPanel.style.display === 'none') {{
    _buildColPanel();
    colSettingsPanel.style.display = 'block';
  }} else {{
    colSettingsPanel.style.display = 'none';
  }}
}});
document.addEventListener('click', () => {{ colSettingsPanel.style.display = 'none'; }});

// ——— Sorting ———
let sortState = {{ key: null, asc: true }};
tableEl.addEventListener('click', e => {{
  const th = e.target.closest('th[data-sort-key]');
  if (!th) return;
  const key = th.dataset.sortKey;
  sortState.asc = sortState.key === key ? !sortState.asc : true;
  sortState.key = key;
  tableEl.querySelectorAll('th[data-sort-key]').forEach(t => delete t.dataset.sortDir);
  th.dataset.sortDir = sortState.asc ? 'asc' : 'desc';
  const tbody = tableEl.querySelector('tbody');
  if (!tbody) return;
  const tRows = [...tbody.querySelectorAll('tr[data-nmid]')];
  const textKeys = new Set(['article','subject','strategy','revenue_category']);
  tRows.sort((a, b) => {{
    const nA = a.dataset.nmid, nB = b.dataset.nmid;
    let va, vb;
    if (key === 'margin_pct') {{
      va = parseFloat(a.querySelector('.cell-econ-margin')?.dataset.margin || 0);
      vb = parseFloat(b.querySelector('.cell-econ-margin')?.dataset.margin || 0);
    }} else if (key === 'profit_pct') {{
      va = parseFloat(a.querySelector('.cell-econ-profit-pct')?.dataset.profitPct || 0);
      vb = parseFloat(b.querySelector('.cell-econ-profit-pct')?.dataset.profitPct || 0);
    }} else if (key === 'net_profit') {{
      va = parseFloat(a.querySelector('.cell-econ-profit')?.dataset.profit || 0);
      vb = parseFloat(b.querySelector('.cell-econ-profit')?.dataset.profit || 0);
    }} else if (key === 'margin_rub') {{
      va = parseFloat(a.querySelector('.cell-econ-margin-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
      vb = parseFloat(b.querySelector('.cell-econ-margin-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
    }} else if (key === 'comm_rub') {{
      va = parseFloat(a.querySelector('.cell-econ-comm-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
      vb = parseFloat(b.querySelector('.cell-econ-comm-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
    }} else if (key === 'drr_rub') {{
      va = parseFloat(a.querySelector('.cell-econ-drr-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
      vb = parseFloat(b.querySelector('.cell-econ-drr-rub')?.textContent?.replace(/[^\d.-]/g,'') || 0);
    }} else if (key === 'stock_value') {{
      va = ((rowDataByNmid[a.dataset.nmid]?.stock || 0) + (rowDataByNmid[a.dataset.nmid]?.warehouse_stock || 0)) * (rowDataByNmid[a.dataset.nmid]?.client_price || 0);
      vb = ((rowDataByNmid[b.dataset.nmid]?.stock || 0) + (rowDataByNmid[b.dataset.nmid]?.warehouse_stock || 0)) * (rowDataByNmid[b.dataset.nmid]?.client_price || 0);
    }} else if (key === 'plannedSpeed') {{
      va = rowForecastCache[nA]?.forecastOrders / Math.max(1, planningParams.forecast_days) || 0;
      vb = rowForecastCache[nB]?.forecastOrders / Math.max(1, planningParams.forecast_days) || 0;
    }} else if (textKeys.has(key)) {{
      va = (rowDataByNmid[nA]?.[key] || '').toLowerCase();
      vb = (rowDataByNmid[nB]?.[key] || '').toLowerCase();
      return sortState.asc ? va.localeCompare(vb,'ru') : vb.localeCompare(va,'ru');
    }} else {{
      va = rowDataByNmid[nA]?.[key] ?? (rowForecastCache[nA]?.[key] ?? 0);
      vb = rowDataByNmid[nB]?.[key] ?? (rowForecastCache[nB]?.[key] ?? 0);
    }}
    return sortState.asc ? va - vb : vb - va;
  }});
  tRows.forEach(r => tbody.appendChild(r));
}});

restoreState();
loadData();
</script>
<div class="methodology">
  <h2>Методология расчётов</h2>

  <h3>Скорость продаж (базовая)</h3>
  <p>Среднедневное количество заказов за последние 7 дней базового периода (последний день с данными исключается — он может быть неполным).<br>
  <span class="formula">скорость = заказы_за_7_дней / 7</span><br>
  <em>Пример: за 7 дней было 350 заказов → базовая скорость = 50 шт/день.</em></p>

  <h3>Эластичность цены</h3>
  <p>Показывает, насколько меняется спрос при изменении цены. Рассчитывается методом лог-лог МНК по 60 дням истории заказов:<br>
  <span class="formula">ln(заказы) = E × ln(цена) + const</span></p>
  <p>Условия расчёта: ≥ 7 дней с заказами, разброс цен ≥ 5%, два ценовых уровня с разрывом ≥ 3%, R² ≥ 0.30.<br>
  Диапазон значений: от −3.0 до 0.0 (отрицательная — при росте цены спрос падает).</p>
  <p><strong>Как читать в таблице:</strong></p>
  <ul>
    <li><strong>−1.0</strong> (по умолчанию, нет данных) — единичная эластичность: +10% цены → −10% скорости.<br>
    <em>50 шт/день × (1 + (−1.0) × 0.10) = 45 шт/день при цене +10%</em></li>
    <li><strong>−0.5</strong> — слабая реакция: +20% цены → −10% скорости.<br>
    <em>50 × (1 + (−0.5) × 0.20) = 45 шт/день</em></li>
    <li><strong>−2.0</strong> — сильная реакция: +10% цены → −20% скорости.<br>
    <em>50 × (1 + (−2.0) × 0.10) = 40 шт/день</em></li>
    <li><strong>0.0</strong> — данные есть, но цена статистически не влияет на спрос (R² &lt; 0.30 или цена не менялась).</li>
  </ul>
  <p>Если по SKU нет надёжных данных — берётся среднее по предмету → стратегии → категории. Если нигде нет — используется <strong>−1.0</strong> (консервативный дефолт).</p>

  <h3>ДРР и эластичность рекламы</h3>
  <p><strong>Базовый ДРР%</strong> = фактические расходы на рекламу / фактическая выручка с выкупов × 100%, за последние 14 дней (последние 2 дня исключаются — выкупы там ещё не зафиксированы).<br>
  <span class="formula">ДРР = Σ расходы_реклама_14д / Σ buyoutSum_14д × 100%</span><br>
  <em>Пример: реклама 70 000 ₽, выручка с выкупов 500 000 ₽ → ДРР = 14%.</em></p>
  <p><strong>Эластичность рекламы</strong> показывает, как расходы на рекламу влияют на скорость продаж. Сначала считается по 60 дням истории через расход/день, затем смешивается с WB-prior по рекламным дням, показам, рекламным заказам и разбросу расходов.<br>
  <span class="formula">E_реклама_эфф = confidence × E_своя + (1 − confidence) × prior</span><br>
  Практический диапазон E_реклама: от 0.08 до 0.85.</p>
  <p><strong>Как работает в планировании:</strong><br>
  <span class="formula">raw_ads = organic_floor + (1 − organic_floor) × (ДРР / anchor_ДРР) ^ E_реклама</span><br>
  <span class="formula">ads_factor = raw_ads(план_ДРР) / raw_ads(факт_ДРР)</span><br>
  anchor_ДРР держится в рабочем коридоре 6–12%, чтобы разовый слив 35–50% не становился нормой модели.</p>
  <ul>
    <li><em>0–6% — товары с близким OOS: не покупаем лишний спрос.</em></li>
    <li><em>6–12% — нормальный рабочий коридор стабильного товара.</em></li>
    <li><em>12–50% — запуск, раскачка, гипотезы, когда остатка достаточно и нужен спрос.</em></li>
  </ul>

  <h3>Конверсии воронки (CTR, CR1, CR2)</h3>
  <p>Данные из ручной выгрузки WB (funnel_impressions_upload) за базовые 7 дней:<br>
  <span class="formula">CTR = переходы / показы × 100%</span> &nbsp;
  <span class="formula">CR1 = корзины / переходы × 100%</span> &nbsp;
  <span class="formula">CR2 = заказы / корзины × 100%</span></p>
  <p>Изменение конверсий — прямой мультипликатор (улучшение карточки, фото, описания):<br>
  <span class="formula">conv_factor = (CTR_план/CTR_баз) × (CR1_план/CR1_баз) × (CR2_план/CR2_баз)</span><br>
  <em>Пример: CTR +20%, CR1 без изм., CR2 +10% → conv_factor = 1.20 × 1.0 × 1.10 = 1.32 → +32% к скорости.</em></p>

  <h3>Итоговая плановая скорость</h3>
  <p>Три рычага перемножаются:<br>
  <span class="formula">план_скорость = базовая_скорость × price_factor × ads_factor × conv_factor</span></p>
  <ul>
    <li><strong>price_factor</strong> = <span class="formula">max(0, 1 + E_цена × Δцена_дол.)</span> — линейная аппроксимация эластичности</li>
    <li><strong>ads_factor</strong> — нормализованный рекламный отклик с органическим полом и anchor_ДРР 6–12%</li>
    <li><strong>conv_factor</strong> — прямой мультипликатор конверсий</li>
  </ul>
  <p><em>Пример: скорость 50 шт/день, цена +10% (E=−1.0), ДРР ×2 (E_рекл=0.5), CTR +20%:<br>
  price_factor = 1 + (−1.0) × 0.10 = 0.90 &nbsp;|&nbsp; ads_factor считается от планового ДРР и органического пола &nbsp;|&nbsp; conv_factor = 1.20<br>
  план_скорость = 50 × 0.90 × 1.41 × 1.20 ≈ 76 шт/день</em></p>
  <p style="color:#64748b;font-size:.75rem;margin-top:8px">Все изменения цены, ДРР и конверсий хранятся в localStorage браузера — прогноз пересчитывается мгновенно на стороне клиента без запросов к серверу.</p>
</div>
</body>
</html>
"""

FUNNEL_UPLOAD_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Загрузка воронки</title>
  <style>
    :root {{ --bg:#f8fafc; --card:#fff; --ink:#111827; --muted:#6b7280; --accent:#0f766e; }}
    * {{ box-sizing: border-box; }}
    body {{ margin:0; font-family:"Segoe UI","Trebuchet MS",sans-serif; color:var(--ink); background:linear-gradient(180deg,#e0f2fe,transparent 240px),var(--bg); }}
    .wrap {{ max-width: 920px; margin:0 auto; padding:22px 16px 40px; }}
    .top {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:var(--muted); white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:2px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    h1 {{ margin:0; font-size:1.35rem; }}
    .panel {{ background:var(--card); border:1px solid #e5e7eb; border-radius:12px; padding:16px; }}
    label {{ display:block; color:var(--muted); font-size:.82rem; font-weight:700; margin-bottom:6px; }}
    input {{ display:block; width:100%; border:1px solid #cbd5e1; border-radius:8px; padding:10px; background:#fff; }}
    button {{ margin-top:12px; padding:10px 14px; border:0; border-radius:8px; background:var(--accent); color:#fff; font-weight:800; cursor:pointer; }}
    button:disabled {{ opacity:.6; cursor:wait; }}
    .meta {{ margin-top:12px; color:var(--muted); font-size:.9rem; line-height:1.45; }}
    .ok {{ color:#047857; }}
    .err {{ color:#b91c1c; }}
    .warn {{ color:#92400e; }}
    .force {{ display:none; margin-top:10px; background:#b45309; }}
    code {{ background:#eef2ff; padding:2px 5px; border-radius:5px; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <div>
        <h1>Загрузка показов и переходов</h1>
        <div class="meta">Поддерживается Excel как выгрузка WB: лист <code>Товары</code>, дата в колонке J, показы в K, переходы в M.</div>
      </div>
      <div class="report-nav">{report_nav}</div>
    </div>
    <div class="panel">
      <label for="file">Файл Excel</label>
      <input id="file" type="file" accept=".xlsx">
      <button id="upload">Загрузить</button>
      <button id="force-upload" class="force">Загрузить всё равно</button>
      <div class="meta" id="status">Файл обновит только даты, которые есть внутри него. Остальные даты останутся без изменений.</div>
    </div>
  </div>
<script>
const fileEl = document.getElementById('file');
const btn = document.getElementById('upload');
const forceBtn = document.getElementById('force-upload');
const statusEl = document.getElementById('status');
let lastUpload = null;

function asBase64(file) {{
  return new Promise((resolve, reject) => {{
    const reader = new FileReader();
    reader.onload = () => {{
      const text = String(reader.result || '');
      resolve(text.includes(',') ? text.split(',', 2)[1] : text);
    }};
    reader.onerror = reject;
    reader.readAsDataURL(file);
  }});
}}

async function uploadFunnel(force = false) {{
  const file = fileEl.files && fileEl.files[0];
  if (!file) {{
    statusEl.textContent = 'Выбери файл .xlsx';
    statusEl.className = 'meta err';
    return;
  }}
  btn.disabled = true;
  forceBtn.disabled = true;
  forceBtn.style.display = 'none';
  statusEl.textContent = 'Загружаю...';
  statusEl.className = 'meta';
  try {{
    const content = lastUpload && lastUpload.name === file.name ? lastUpload.content : await asBase64(file);
    lastUpload = {{ name: file.name, content }};
    const resp = await fetch('/api/analytics/funnel-upload', {{
      method: 'POST',
      headers: {{ 'Content-Type': 'application/json' }},
      body: JSON.stringify({{ filename: file.name, content, force }})
    }});
    const data = await resp.json();
    if (data.warning) {{
      const details = (data.details || [])
        .map(item => `${{item.date}}: в файле ${{item.uploaded}}, в базе ${{item.existing}}, меньше на ${{item.missing}}`)
        .join('; ');
      statusEl.className = 'meta warn';
      statusEl.textContent = `${{data.message}} ${{details}}`;
      forceBtn.style.display = 'inline-block';
      return;
    }}
    if (!resp.ok) throw new Error(data.error || `HTTP ${{resp.status}}`);
    statusEl.className = 'meta ok';
    statusEl.textContent = `Готово: строк ${{data.rows}}, товаров ${{data.nmids}}, период ${{data.date_from}}..${{data.date_to}}, показов ${{data.impressions}}. Обновлены только даты из файла.`;
  }} catch (err) {{
    statusEl.className = 'meta err';
    statusEl.textContent = `Ошибка: ${{err.message || err}}`;
  }} finally {{
    btn.disabled = false;
    forceBtn.disabled = false;
  }}
}}

btn.addEventListener('click', () => uploadFunnel(false));
forceBtn.addEventListener('click', () => uploadFunnel(true));
</script>
</body>
</html>
"""

def _decimal_text(value: object) -> str:
  if value is None:
    return "0"
  try:
    return str(Decimal(str(value).replace(" ", "").replace(",", ".")))
  except InvalidOperation:
    return "0"

def _cell_text(value: object) -> str:
  if value is None:
    return ""
  if isinstance(value, float) and value.is_integer():
    return str(int(value))
  return str(value).strip()

class FunnelUploadWarning(Exception):
  def __init__(self, details: list[dict[str, object]]):
    super().__init__("Файл выглядит неполным")
    self.details = details

def _import_funnel_excel(path: str, source_name: str, force: bool = False) -> dict[str, object]:
  from openpyxl import load_workbook

  workbook = load_workbook(path, read_only=False, data_only=True)
  if "Товары" not in workbook.sheetnames:
    raise ValueError("В файле нет листа 'Товары'")
  sheet = workbook["Товары"]

  header_row = None
  headers: list[str] = []
  for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    values = [str(value).strip() if value is not None else "" for value in row]
    if "Артикул WB" in values and "Дата" in values and "Показы" in values and "Переходы в карточку" in values:
      header_row = row_idx
      headers = values
      break
  if header_row is None:
    raise ValueError("Не нашёл заголовки: нужны 'Артикул WB', 'Дата', 'Показы', 'Переходы в карточку'")

  idx = {header: pos for pos, header in enumerate(headers) if header}
  def first_col(*names: str) -> int | None:
    for name in names:
      if name in idx:
        return idx[name]
    return None

  order_count_idx = first_col("Заказали, шт", "Заказали товаров, шт")
  rows: list[tuple[str, str, str, str, str, str, str, str, str]] = []
  funnel_rows: list[dict[str, str]] = []
  dates: set[str] = set()
  nmids: set[str] = set()
  nmids_by_date: dict[str, set[str]] = {}
  impressions_total = Decimal("0")
  transitions_total = Decimal("0")

  for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
    row_date = _cell_text(row[idx["Дата"]])
    nm_id = _cell_text(row[idx["Артикул WB"]])
    if not row_date or not nm_id:
      continue
    impressions = _decimal_text(row[idx["Показы"]])
    transitions = _decimal_text(row[idx["Переходы в карточку"]])
    cart_count = _decimal_text(row[idx["Положили в корзину"]]) if "Положили в корзину" in idx else "0"
    order_count = _decimal_text(row[order_count_idx]) if order_count_idx is not None else "0"
    buyout_count = _decimal_text(row[idx["Выкупили, шт"]]) if "Выкупили, шт" in idx else "0"
    supplier_article = _cell_text(row[idx["Артикул продавца"]]) if "Артикул продавца" in idx else ""
    product_name = _cell_text(row[idx["Название"]]) if "Название" in idx else ""
    brand = _cell_text(row[idx["Бренд"]]) if "Бренд" in idx else ""
    subject = _cell_text(row[idx["Предмет"]]) if "Предмет" in idx else ""
    dates.add(row_date)
    nmids.add(nm_id)
    nmids_by_date.setdefault(row_date, set()).add(nm_id)
    impressions_total += Decimal(impressions)
    transitions_total += Decimal(transitions)
    rows.append((row_date, nm_id, supplier_article, impressions, transitions, cart_count, order_count, buyout_count, source_name))
    funnel_rows.append({
      "nmId": nm_id,
      "supplierArticle": supplier_article,
      "productName": product_name,
      "brand": brand,
      "currency": "RUB",
      "date": row_date,
      "product_subjectId": "",
      "product_subjectName": subject,
      "addToCartConversion": _decimal_text(row[idx["Конверсия в корзину, %"]]) if "Конверсия в корзину, %" in idx else "0",
      "addToWishlistCount": _decimal_text(row[idx["Добавили в отложенные"]]) if "Добавили в отложенные" in idx else "0",
      "buyoutCount": buyout_count,
      "buyoutPercent": _decimal_text(row[idx["Процент выкупа"]]) if "Процент выкупа" in idx else "0",
      "buyoutSum": _decimal_text(row[idx["Выкупили на сумму, ₽"]]) if "Выкупили на сумму, ₽" in idx else "0",
      "cartCount": cart_count,
      "cartToOrderConversion": _decimal_text(row[idx["Конверсия в заказ, %"]]) if "Конверсия в заказ, %" in idx else "0",
      "openCount": transitions,
      "orderCount": order_count,
      "orderSum": _decimal_text(row[idx["Заказали на сумму, ₽"]]) if "Заказали на сумму, ₽" in idx else "0",
      "cancelCount": _decimal_text(row[idx["Отменили, шт"]]) if "Отменили, шт" in idx else "0",
      "cancelSum": _decimal_text(row[idx["Отменили на сумму, ₽"]]) if "Отменили на сумму, ₽" in idx else "0",
      "rating": _decimal_text(row[idx["Рейтинг карточки"]]) if "Рейтинг карточки" in idx else "0",
      "reviewRating": _decimal_text(row[idx["Рейтинг по отзывам"]]) if "Рейтинг по отзывам" in idx else "0",
      "isDeleted": _cell_text(row[idx["Удаленный товар"]]) if "Удаленный товар" in idx else "",
      "wbClubOrderCount": _decimal_text(row[idx["Заказали ВБ клуб, шт"]]) if "Заказали ВБ клуб, шт" in idx else "0",
      "wbClubOrderSum": _decimal_text(row[idx["Заказали на сумму ВБ клуб, ₽"]]) if "Заказали на сумму ВБ клуб, ₽" in idx else "0",
      "wbClubBuyoutCount": _decimal_text(row[idx["Выкупили ВБ клуб, шт"]]) if "Выкупили ВБ клуб, шт" in idx else "0",
      "wbClubBuyoutSum": _decimal_text(row[idx["Выкупили на сумму ВБ клуб, ₽"]]) if "Выкупили на сумму ВБ клуб, ₽" in idx else "0",
      "wbClubCancelCount": _decimal_text(row[idx["Отменили ВБ клуб, шт"]]) if "Отменили ВБ клуб, шт" in idx else "0",
      "wbClubCancelSum": _decimal_text(row[idx["Отменили на сумму ВБ клуб, ₽"]]) if "Отменили на сумму ВБ клуб, ₽" in idx else "0",
    })

  if not rows:
    raise ValueError("В файле нет строк с датой и Артикул WB")

  with sqlite3.connect(get_current_db_path()) as conn:
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute(
      """
      CREATE TABLE IF NOT EXISTS funnel_impressions_upload (
        date TEXT NOT NULL,
        nmId TEXT NOT NULL,
        supplierArticle TEXT,
        impressions TEXT,
        openCount TEXT,
        cartCount TEXT,
        orderCount TEXT,
        buyoutCount TEXT,
        source_file TEXT,
        PRIMARY KEY (date, nmId)
      )
      """
    )
    funnel_columns = [
      "nmId", "supplierArticle", "productName", "brand", "currency", "date",
      "product_subjectId", "product_subjectName", "addToCartConversion", "addToWishlistCount",
      "buyoutCount", "buyoutPercent", "buyoutSum", "cartCount", "cartToOrderConversion",
      "openCount", "orderCount", "orderSum", "cancelCount", "cancelSum", "rating",
      "reviewRating", "isDeleted", "wbClubOrderCount", "wbClubOrderSum",
      "wbClubBuyoutCount", "wbClubBuyoutSum", "wbClubCancelCount", "wbClubCancelSum",
    ]
    conn.execute(
      "CREATE TABLE IF NOT EXISTS funnel_analytics ("
      + ", ".join(f'"{column}" TEXT' for column in funnel_columns)
      + ")"
    )
    if not force:
      warnings: list[dict[str, object]] = []
      for row_date in sorted(nmids_by_date):
        uploaded_count = len(nmids_by_date[row_date])
        existing_count = conn.execute(
          "SELECT COUNT(DISTINCT nmId) FROM funnel_analytics WHERE date = ?",
          (row_date,),
        ).fetchone()[0] or 0
        missing_count = int(existing_count) - uploaded_count
        # Если в файле заметно меньше товаров, сначала просим явное подтверждение.
        if missing_count >= 10:
          warnings.append({
            "date": row_date,
            "uploaded": uploaded_count,
            "existing": int(existing_count),
            "missing": missing_count,
          })
      if warnings:
        raise FunnelUploadWarning(warnings)
    conn.executemany("DELETE FROM funnel_impressions_upload WHERE date = ?", [(row_date,) for row_date in dates])
    conn.executemany("DELETE FROM funnel_analytics WHERE date = ?", [(row_date,) for row_date in dates])
    conn.executemany(
      """
      INSERT INTO funnel_impressions_upload (
        date, nmId, supplierArticle, impressions, openCount, cartCount, orderCount, buyoutCount, source_file
      ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
      """,
      rows,
    )
    placeholders = ", ".join("?" for _ in funnel_columns)
    quoted_columns = ", ".join(f'"{column}"' for column in funnel_columns)
    conn.executemany(
      f"INSERT INTO funnel_analytics ({quoted_columns}) VALUES ({placeholders})",
      [[row.get(column, "") for column in funnel_columns] for row in funnel_rows],
    )
    conn.commit()

  return {
    "rows": len(rows),
    "nmids": len(nmids),
    "date_from": min(dates),
    "date_to": max(dates),
    "dates": len(dates),
    "impressions": f"{int(impressions_total):,}".replace(",", " "),
    "transitions": f"{int(transitions_total):,}".replace(",", " "),
  }

def _rebuild_buyout_order_day_from_store() -> int:
  src_path = str(ROOT / "src")
  if src_path not in sys.path:
    sys.path.insert(0, src_path)

  from wb_gsheets.sqlite_store import SQLiteStore
  from wb_gsheets.transform import (
    build_buyout_order_day_rows,
    build_nm_mapping,
    sheet_values_to_dicts,
  )

  def _table_name(env_name: str, default: str) -> str:
    return os.getenv(env_name, default).strip() or default

  cabinet = _get_request_cabinet() or {}
  article_filter_type = str(cabinet.get("article_filter_type") or os.getenv("ARTICLE_FILTER_TYPE", "nmId")).strip() or "nmId"
  if article_filter_type not in {"nmId", "vendorCode"}:
    article_filter_type = "nmId"

  sku_table = _table_name("SKU_TABLE", "SKU")
  raw_sales_table = _table_name("RAW_SALES_TABLE", "raw_sales")
  raw_orders_table = _table_name("RAW_ORDERS_TABLE", "raw_orders")
  raw_ads_table = _table_name("RAW_ADS_TABLE", "raw_ads")
  funnel_analytics_table = _table_name("FUNNEL_ANALYTICS_TABLE", "funnel_analytics")

  store = SQLiteStore(get_current_db_path())
  sku_values = store.get_values(sku_table)
  nm_mapping = build_nm_mapping(sku_values, article_filter_type=article_filter_type)
  sales_rows = sheet_values_to_dicts(store.get_values(raw_sales_table))
  orders_rows = sheet_values_to_dicts(store.get_values(raw_orders_table))

  funnel_by_nm: dict[str, dict[str, object]] = {}
  for row in sheet_values_to_dicts(store.get_values(funnel_analytics_table)):
    nm_id = str(row.get("nmId", "")).strip()
    row_date = str(row.get("date", "")).strip()
    if not nm_id or not row_date:
      continue
    entry = funnel_by_nm.setdefault(nm_id, {
      "currency": str(row.get("currency", "") or "RUB"),
      "product": {
        "nmId": nm_id,
        "vendorCode": str(row.get("supplierArticle", "") or ""),
        "title": str(row.get("productName", "") or ""),
        "brandName": str(row.get("brand", "") or ""),
      },
      "history": [],
    })
    entry["history"].append({
      "date": row_date,
      "orderSum": row.get("orderSum") or 0,
      "orderCount": row.get("orderCount") or 0,
      "buyoutCount": row.get("buyoutCount") or 0,
      "buyoutSum": row.get("buyoutSum") or 0,
    })

  rows = build_buyout_order_day_rows(
    orders_rows=orders_rows,
    sales_rows=sales_rows,
    ads_rows=sheet_values_to_dicts(store.get_values(raw_ads_table)),
    nm_mapping=nm_mapping,
    funnel_data=list(funnel_by_nm.values()),
  )
  store.replace_table("buyout_order_day", rows)
  return max(0, len(rows) - 1)

def _render_db_page(selected_table: str | None, page: int, page_size: int, cabinet: dict | None = None) -> bytes:
    store = db_store()
    existing_tables = store.list_tables()
    tables = list(dict.fromkeys(CORE_TABLES + existing_tables))

    current_table = selected_table if selected_table in tables else (tables[0] if tables else None)
    headers: list[str] = []
    rows: list[list[str]] = []
    total = 0
    table_exists = current_table in set(existing_tables) if current_table else False

    if current_table and table_exists:
        offset = max(page - 1, 0) * page_size
        with _db_connect() as conn:
            safe_table = (current_table or "").replace('"', '""')
            table_info = conn.execute(f'PRAGMA table_info("{safe_table}")').fetchall()
            columns = [str(row["name"]) for row in table_info]
            scope_where, scope_params = _db_scope_filter(conn, current_table)
            if columns:
                headers = columns
                total = int(
                    conn.execute(
                        f"SELECT COUNT(*) AS cnt FROM {_sql_ident(current_table)} {scope_where}",
                        scope_params,
                    ).fetchone()["cnt"]
                )
                select_cols = ", ".join(_sql_ident(col) for col in headers)
                rows = [
                    ["" if row[col] is None else str(row[col]) for col in headers]
                    for row in conn.execute(
                        f"SELECT {select_cols} FROM {_sql_ident(current_table)} {scope_where} LIMIT ? OFFSET ?",
                        [*scope_params, page_size, offset],
                    ).fetchall()
                ]

    total_pages = max(1, ceil(total / page_size)) if current_table else 1
    page = min(max(page, 1), total_pages)

    nav_items = []
    existing_set = set(existing_tables)
    for table in tables:
        cls = "active" if table == current_table else ""
        if table not in existing_set:
            cls = (cls + " pending").strip()
        table_q = quote_plus(table)
        suffix = "" if table in existing_set else " (пусто)"
        nav_items.append(
            f'<a class="tbl {cls}" href="{escape(f"/db?table={table_q}&page=1")}">{escape(table)}{escape(suffix)}</a>'
        )
    nav_html = "".join(nav_items)

    split_col_idx = -1
    if current_table == "finance_article_day_detail" and "Реклама" in headers:
        split_col_idx = headers.index("Реклама")

    if headers:
        head_cells = []
        for idx, col in enumerate(headers):
            cls = ' class="sep-left"' if idx == split_col_idx else ""
            head_cells.append(f"<th{cls}>{escape(col)}</th>")
        head = "".join(head_cells)

        body_rows = []
        for row in rows:
            cells = []
            for idx, cell in enumerate(row):
                cls = ' class="sep-left"' if idx == split_col_idx else ""
                cells.append(f"<td{cls}>{escape(cell)}</td>")
            body_rows.append("<tr>" + "".join(cells) + "</tr>")

        body = "".join(body_rows) if body_rows else f'<tr><td colspan="{len(headers)}">Нет данных на этой странице</td></tr>'
        table_html = (
            "<div class=\"meta\">"
            f"<strong>{escape(current_table or '')}</strong>"
            f"<span>Строк: {total}</span>"
            f"<span>Страница {page}/{total_pages}</span>"
            "</div>"
            "<div class=\"tbl-wrap\"><table><thead><tr>"
            f"{head}"
            "</tr></thead><tbody>"
            f"{body}"
            "</tbody></table></div>"
        )
    else:
        if current_table and not table_exists:
            table_html = (
                "<div class=\"empty\">"
                f"Таблица {escape(current_table)} еще не создана. Запусти синхронизацию на главной странице, и она появится автоматически."
                "</div>"
            )
        else:
            table_html = "<div class=\"empty\">Таблица пока пустая.</div>"

    prev_page = max(1, page - 1)
    next_page = min(total_pages, page + 1)
    selected_q = quote_plus(current_table or "")
    pager = ""
    if current_table and table_exists:
        pager = (
            '<div class="pager">'
            f'<a href="{escape(f"/db?table={selected_q}&page={prev_page}")}">← Назад</a>'
            f'<a href="{escape(f"/db?table={selected_q}&page={next_page}")}">Вперед →</a>'
            "</div>"
        )

    report_nav = _report_nav_html_for_cabinet(cabinet)
    html = f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>WB SQLite Browser</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; background: #f3f4f6; color: #111827; }}
    .page-nav {{ width: calc(100% - 32px); max-width: 3000px; margin: 0 auto; padding: 10px 0 8px; }}
    .top {{ display:flex; align-items:center; gap:8px; flex-wrap:nowrap; overflow:hidden; }}
    .top h1 {{ font-size:.82rem; font-weight:700; color:#6b7280; white-space:nowrap; margin:0; flex-shrink:0; }}
    .report-nav {{ display:flex; align-items:center; flex-wrap:nowrap; overflow-x:auto; gap:4px; flex:1; scrollbar-width:none; min-width:0; }}
    .report-nav::-webkit-scrollbar {{ display:none; }}
    .db-meta {{ width: calc(100% - 32px); max-width: 3000px; margin: 0 auto 8px; padding: 6px 10px; border: 1px solid #e5e7eb; border-radius: 8px; background: #fff; color: #64748b; font-size: .76rem; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
    .layout {{ display: grid; grid-template-columns: 250px 1fr; min-height: calc(100vh - 104px); }}
    .side {{ border-right: 1px solid #d1d5db; background: #fff; padding: 12px; overflow-y: auto; }}
    .content {{ width: min(100%, 1800px); margin: 0 auto; padding: 16px; overflow-x: auto; }}
    .tbl {{ display: block; padding: 9px 10px; border-radius: 8px; color: #111827; text-decoration: none; margin-bottom: 6px; }}
    .tbl:hover {{ background: #eef2ff; }}
    .tbl.active {{ background: #4338ca; color: #fff; }}
    .tbl.pending {{ color: #6b7280; background: #f9fafb; }}
    .meta {{ margin-bottom: 10px; display: flex; gap: 14px; align-items: center; }}
    .tbl-wrap {{ background: #fff; border: 1px solid #d1d5db; border-radius: 10px; overflow: auto; max-height: calc(100vh - 180px); }}
    table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid #eef2f7; border-right: 1px solid #f3f4f6; padding: 8px 10px; text-align: left; white-space: nowrap; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1; }}
    th {{ position: sticky; top: 0; background: #f8fafc; }}
    .sep-left {{ border-left: 3px solid #94a3b8; }}
    .pager {{ margin-top: 12px; display: flex; gap: 12px; }}
    .pager a {{ color: #1d4ed8; text-decoration: none; font-weight: 600; }}
    .empty {{ color: #6b7280; padding: 8px; }}
    @media (max-width: 900px) {{
      .layout {{ grid-template-columns: 1fr; }}
      .side {{ border-right: 0; border-bottom: 1px solid #d1d5db; }}
    }}
  </style>
</head>
<body>
  <div class="page-nav">
    <div class="top">
      <h1>SQLite</h1>
      <div class="report-nav">{report_nav}</div>
    </div>
  </div>
  <div class="db-meta">База: {escape(get_current_db_path())}</div>
  <div class="layout">
    <aside class="side">{nav_html}</aside>
    <main class="content">{table_html}{pager}</main>
  </div>
</body>
</html>"""
    return _inject_brand(html, cabinet=cabinet).encode("utf-8")


# ── Цены конкурентов — функции данных ─────────────────────────────────────

def _fetch_competitor_prices_summary(cabinet_id: str) -> dict:
    """Список наших артикулов с агрегатом цен конкурентов за последние 7 дней."""
    import os as _os
    if not _os.path.exists(PRICES_DB_PATH):
        return {"rows": []}
    try:
        pconn = sqlite3.connect(PRICES_DB_PATH)
        pconn.row_factory = sqlite3.Row
        rows = pconn.execute("""
            SELECT
                cp.our_nm_id,
                cp.our_name,
                ph.price        AS our_price,
                ph.stock        AS our_stock,
                COUNT(DISTINCT cp.comp_nm_id)          AS competitors_count,
                CAST(AVG(cprice.price) AS INTEGER)     AS median_price,
                MIN(cprice.price)                      AS min_price,
                MAX(cprice.price)                      AS max_price
            FROM competitor_products cp
            LEFT JOIN competitor_prices cprice
                ON cprice.cabinet_id = cp.cabinet_id
               AND cprice.our_nm_id  = cp.our_nm_id
               AND cprice.comp_nm_id = cp.comp_nm_id
               AND cprice.date >= date('now', '-7 days')
               AND cprice.price > 0
            LEFT JOIN (
                SELECT nm_id, price, stock
                FROM price_history
                WHERE cabinet_id = ? AND date = (
                    SELECT MAX(date) FROM price_history WHERE cabinet_id = ?
                )
            ) ph ON ph.nm_id = cp.our_nm_id
            WHERE cp.cabinet_id = ?
            GROUP BY cp.our_nm_id
            ORDER BY cp.our_nm_id
        """, (cabinet_id, cabinet_id, cabinet_id)).fetchall()
        pconn.close()
        return {"rows": [dict(r) for r in rows]}
    except Exception as e:
        return {"rows": [], "error": str(e)}


def _fetch_competitor_costs(cabinet_id: str) -> dict:
    """Возвращает % расходов WB и себестоимость по каждому артикулу + дефолты по кабинету."""
    import os as _os
    db_path = _os.path.join("data", "cabs", f"{cabinet_id}.db")
    if not _os.path.exists(db_path):
        return {"articles": {}, "defaults": {"commission_pct": 15.0, "logistics_pct": 8.0, "acquiring_pct": 2.0}}
    try:
        import sqlite3 as _sq
        conn = _sq.connect(db_path)

        # Дефолты по кабинету
        drow = conn.execute('''
            SELECT AVG("% комиссии WB"), AVG("% логистики"), AVG("% эквайринга")
            FROM finance_article_day_detail
            WHERE "Дата" >= date("now","-30 days") AND "% комиссии WB" > 0
        ''').fetchone()
        defaults = {
            "commission_pct": round(drow[0] or 15.0, 2),
            "logistics_pct":  round(drow[1] or 0.0,  2),
            "acquiring_pct":  round(drow[2] or 2.0,  2),
        }

        # По артикулу через raw_orders (vendor_code -> nm_id)
        rows = conn.execute('''
            SELECT
                CAST(ro.nmId AS INTEGER)    AS nm_id,
                AVG(f."% комиссии WB")      AS commission_pct,
                AVG(f."% логистики")        AS logistics_pct,
                AVG(f."% эквайринга")       AS acquiring_pct,
                s.себестоимость             AS cost_price
            FROM finance_article_day_detail f
            JOIN (SELECT DISTINCT nmId, supplierArticle FROM raw_orders WHERE nmId IS NOT NULL) ro
                ON ro.supplierArticle = f."Артикул"
            LEFT JOIN SKU s ON s."Артикул поставщика" = f."Артикул"
            WHERE f."Дата" >= date("now","-30 days")
              AND f."% комиссии WB" > 0
            GROUP BY ro.nmId
        ''').fetchall()
        conn.close()

        articles = {}
        for nm_id, comm, log, acq, cost in rows:
            if not nm_id:
                continue
            articles[str(nm_id)] = {
                "commission_pct": round(comm or defaults["commission_pct"], 2),
                "logistics_pct":  round(log  or defaults["logistics_pct"],  2),
                "acquiring_pct":  round(acq  or defaults["acquiring_pct"],  2),
                "cost_price":     float(cost) if cost else None,
            }
        return {"articles": articles, "defaults": defaults}
    except Exception as e:
        return {"articles": {}, "defaults": {"commission_pct": 15.0, "logistics_pct": 8.0, "acquiring_pct": 2.0}, "error": str(e)}


def _fetch_live_prices(nm_ids: list) -> dict:
    """Запрашивает живые цены у агента на ноуте."""
    import urllib.request as _ur
    import urllib.error as _ue
    agent_url = "http://100.65.13.99:8100"
    nm_str = ";".join(str(n) for n in nm_ids)
    try:
        raw = _ur.urlopen(f"{agent_url}/prices?nm={nm_str}", timeout=30).read()
        import json as _json
        data = _json.loads(raw)
        return {str(p["id"]): p for p in data.get("products", [])}
    except _ue.URLError:
        return {"_error": "agent_unavailable"}
    except Exception as e:
        return {"_error": str(e)}


def _fetch_competitor_prices_detail(cabinet_id: str, nm_id: int) -> dict:
    """Список конкурентов с последними ценами для выбранного артикула."""
    import os as _os
    if not _os.path.exists(PRICES_DB_PATH):
        return {"rows": []}
    try:
        pconn = sqlite3.connect(PRICES_DB_PATH)
        pconn.row_factory = sqlite3.Row
        rows = pconn.execute("""
            SELECT
                cp.comp_nm_id,
                cp.comp_name,
                cp.comp_brand,
                cp.comp_seller,
                cp.url,
                cp.subject_name,
                cprice.price,
                cprice.price_basic,
                cprice.stock,
                cprice.rating,
                cprice.feedbacks,
                cprice.date AS last_date
            FROM competitor_products cp
            LEFT JOIN competitor_prices cprice
                ON cprice.cabinet_id = cp.cabinet_id
               AND cprice.our_nm_id  = cp.our_nm_id
               AND cprice.comp_nm_id = cp.comp_nm_id
               AND cprice.date = (
                    SELECT MAX(date) FROM competitor_prices
                    WHERE cabinet_id = cp.cabinet_id
                      AND our_nm_id  = cp.our_nm_id
                      AND comp_nm_id = cp.comp_nm_id
               )
            WHERE cp.cabinet_id = ? AND cp.our_nm_id = ?
            ORDER BY COALESCE(cprice.price, 999999)
        """, (cabinet_id, nm_id)).fetchall()
        pconn.close()
        return {"rows": [dict(r) for r in rows]}
    except Exception as e:
        return {"rows": [], "error": str(e)}


COMPETITOR_PRICES_HTML = """\
<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>Цены конкурентов</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;font-size:13px;background:#f5f6fa;color:#1a1d23}}
.top-bar{{background:#1a1d23;padding:10px 16px;display:flex;align-items:center;gap:16px}}
.top-bar a{{color:#aaa;text-decoration:none;font-size:12px}}
.top-bar a:hover,.top-bar a.active{{color:#fff}}
.page{{padding:16px}}
h2{{font-size:15px;font-weight:600;margin-bottom:12px;color:#1a1d23}}
.section{{background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:16px;overflow:hidden}}
.section-header{{padding:10px 16px;border-bottom:1px solid #eef;font-weight:600;font-size:13px;background:#fafbff;display:flex;align-items:center;gap:8px}}
table{{width:100%;border-collapse:collapse}}
th{{padding:8px 10px;text-align:left;font-size:11px;font-weight:600;color:#888;background:#fafbff;border-bottom:1px solid #eef;position:sticky;top:0;z-index:1}}
td{{padding:7px 10px;border-bottom:1px solid #f2f3f7;vertical-align:middle}}
tr.selected{{background:#eef3ff}}
tr:hover{{background:#f7f9ff;cursor:pointer}}
.price{{font-weight:600;color:#1a1d23}}
.price-low{{color:#1a9e4a}}
.price-high{{color:#d9312e}}
.no-data{{color:#bbb;font-style:italic}}
.badge{{display:inline-block;padding:2px 7px;border-radius:10px;font-size:11px;font-weight:600}}
.badge-count{{background:#eef0ff;color:#5865f2}}
.link-btn{{color:#5865f2;text-decoration:none;font-size:11px}}
.link-btn:hover{{text-decoration:underline}}
.spinner{{text-align:center;padding:32px;color:#aaa}}
#detail-panel{{display:none}}
#detail-panel.visible{{display:block}}
.nm-label{{font-size:11px;color:#888;display:block}}
.costs-block{{background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:10px 16px;margin-bottom:8px;display:flex;align-items:center;gap:20px;flex-wrap:wrap}}
.costs-block b{{font-size:12px;color:#6b7280;margin-right:4px}}
.costs-block label{{display:flex;align-items:center;gap:5px;font-size:12px;color:#374151}}
.costs-block input{{width:62px;height:26px;padding:0 6px;border:1px solid #cbd5e1;border-radius:5px;font-size:12px;text-align:right}}
.new-price-wrap{{display:flex;align-items:center;gap:2px}}
.new-price-wrap input{{width:76px;height:24px;padding:0 5px;border:1px solid #cbd5e1;border-radius:4px;font-size:12px;text-align:right}}
.new-price-wrap button{{width:22px;height:24px;border:1px solid #cbd5e1;background:#f8fafc;border-radius:4px;cursor:pointer;font-size:13px;line-height:1;padding:0}}
.new-price-wrap button:hover{{background:#e0f2fe}}
.our-margin,.new-margin{{white-space:nowrap}}
.margin-warn{{color:#f59e0b;font-size:11px}}
.econ-tbl{{border-collapse:collapse;font-size:11px;width:100%}}
.econ-tbl th{{background:#f1f5f9;padding:3px 8px;font-weight:600;color:#64748b;white-space:nowrap;border:1px solid #e2e8f0;text-align:center}}
.econ-tbl td{{padding:3px 8px;border:1px solid #e2e8f0;white-space:nowrap;text-align:right}}
.econ-tbl .econ-lbl{{text-align:left;font-weight:600;color:#374151;background:#f8fafc}}
.econ-tbl tr#econ-new td{{color:#166534}}
</style>
</head>
<body>
<div class="top-bar">{report_nav}</div>
<div class="page">
  <h2>Цены конкурентов — hld</h2>

  <div class="costs-block" id="costs-block">
    <b>Расходы по умолчанию:</b>
    <label>Комиссия&nbsp;<input id="c-comm" type="number" step="0.1" min="0" max="100" value="15">%</label>
    <label>Эквайринг&nbsp;<input id="c-acq" type="number" step="0.1" min="0" max="100" value="2">%</label>
    <label>Логистика&nbsp;<input id="c-log" type="number" step="0.1" min="0" max="100" value="0">%</label>
    <label>ДРР&nbsp;<input id="c-drr" type="number" step="0.1" min="0" max="100" value="0">%</label>
    <label>Прочие&nbsp;<input id="c-other" type="number" step="0.1" min="0" max="100" value="0">%</label>
    <span id="costs-status" style="font-size:11px;color:#9ca3af;margin-left:8px">загрузка…</span>
  </div>

  <div class="section">
    <div class="section-header">
      Наши артикулы
      <span id="our-count" style="font-weight:400;color:#888;font-size:12px"></span>
    </div>
    <div style="overflow-x:auto">
      <table id="our-table">
        <thead>
          <tr>
            <th>nmId</th>
            <th>Название</th>
            <th>Тек цена</th>
            <th>Маржа</th>
            <th>Новая цена</th>
            <th>Новая маржа</th>
            <th>Остаток</th>
            <th>Конкурентов</th>
            <th>Медиана</th>
            <th>Мин</th>
            <th>Макс</th>
            <th>Позиция</th>
          </tr>
        </thead>
        <tbody id="our-tbody">
          <tr><td colspan="9" class="spinner">Загрузка…</td></tr>
        </tbody>
      </table>
    </div>
  </div>

  <div class="section" id="detail-panel">
    <div class="section-header">
      Конкуренты:&nbsp;
      <span id="detail-title" style="font-weight:400"></span>
      <span id="detail-count" style="font-weight:400;color:#888;font-size:12px;margin-left:8px"></span>
    </div>
    <div id="econ-block" style="overflow-x:auto;margin-bottom:6px">
      <table class="econ-tbl">
        <thead><tr>
          <th></th><th>Цена</th><th>Комиссия</th><th>Логистика</th><th>Эквайринг</th><th>ДРР</th><th>Прочие</th><th>Себест.</th><th>Маржа</th>
        </tr></thead>
        <tbody>
          <tr id="econ-cur"><td class="econ-lbl">Текущая</td><td colspan="8" style="color:#aaa">—</td></tr>
          <tr id="econ-new" style="background:#f0fdf4"><td class="econ-lbl">Новая</td><td colspan="8" style="color:#aaa">—</td></tr>
        </tbody>
      </table>
    </div>
    <div style="overflow-x:auto">
      <table>
        <thead>
          <tr>
            <th>nmId</th>
            <th>Название</th>
            <th>Бренд</th>
            <th>Продавец</th>
            <th>Цена (БД)</th>
            <th>До скидки</th>
            <th>Остаток</th>
            <th style="background:#e8f5e9;color:#2e7d32">Живая цена</th>
            <th style="background:#e8f5e9;color:#2e7d32">Живой склад</th>
            <th>★</th>
            <th>Отзывы</th>
            <th>Дата</th>
            <th></th>
          </tr>
        </thead>
        <tbody id="detail-tbody"></tbody>
      </table>
    </div>
  </div>
</div>
<script>
const CAB = '{cabinet_id}';
function fmt(n){{if(!n&&n!==0)return'<span class="no-data">—</span>';return Number(n).toLocaleString('ru-RU')+' ₽';}}
function fmtN(n){{if(!n&&n!==0)return'<span class="no-data">—</span>';return Number(n).toLocaleString('ru-RU');}}
function posLabel(our,med){{
  if(!our||!med)return'<span class="no-data">—</span>';
  const pct=((our-med)/med*100).toFixed(1);
  const cls=our<=med?'price-low':'price-high';
  return`<span class="${{cls}}">${{pct>0?'+':''}}${{pct}}%</span>`;
}}

// ── Расходы и маржа ──────────────────────────────────────────────────────
let costsData = {{articles:{{}}, defaults:{{commission_pct:15,logistics_pct:0,acquiring_pct:2}}}};

function getInputs(){{
  return {{
    comm:  parseFloat(document.getElementById('c-comm').value)  || 0,
    acq:   parseFloat(document.getElementById('c-acq').value)   || 0,
    log:   parseFloat(document.getElementById('c-log').value)   || 0,
    drr:   parseFloat(document.getElementById('c-drr').value)   || 0,
    other: parseFloat(document.getElementById('c-other').value) || 0,
  }};
}}

function calcMargin(price, costPrice, hasCost, {{comm,acq,log,drr,other}}){{
  if(!price) return null;
  const feesRub = price * (comm + acq + log + drr + other) / 100;
  const profit  = price - feesRub - (hasCost ? (costPrice||0) : 0);
  return profit / price * 100;
}}

function fmtMargin(m, hasCost){{
  if(m==null) return '<span class="no-data">—</span>';
  const warn = hasCost ? '' : '<span class="margin-warn" title="нет себестоимости">⚠</span>&nbsp;';
  const cls  = m>=20 ? 'price-low' : m>=10 ? '' : 'price-high';
  return warn+'<span class="'+cls+'">'+m.toFixed(1)+'%</span>';
}}

function recalcRow(tr, inputs){{
  const nm       = tr.dataset.nm;
  const price    = +tr.dataset.price || 0;
  const art      = costsData.articles[nm] || {{}};
  const rowInputs = {{
    comm:  art.commission_pct ?? inputs.comm,
    acq:   art.acquiring_pct  ?? inputs.acq,
    log:   art.logistics_pct  ?? inputs.log,
    drr:   inputs.drr,
    other: inputs.other,
  }};
  const costPrice = art.cost_price ?? null;
  const hasCost   = costPrice != null && costPrice > 0;
  const mCell = tr.querySelector('.our-margin');
  if(mCell) mCell.innerHTML = fmtMargin(calcMargin(price,costPrice,hasCost,rowInputs), hasCost);
  const inp = tr.querySelector('.new-price-input');
  const nmCell = tr.querySelector('.new-margin');
  if(inp && nmCell) nmCell.innerHTML = fmtMargin(calcMargin(+inp.value,costPrice,hasCost,rowInputs), hasCost);
}}

function recalcAllMargins(){{
  const inputs = getInputs();
  document.querySelectorAll('#our-tbody tr[data-nm]').forEach(tr=>recalcRow(tr,inputs));
}}

async function loadCosts(){{
  try{{
    const r = await fetch(`/api/competitor-prices/costs?cabinet_id=${{CAB}}`);
    costsData = await r.json();
    const d = costsData.defaults||{{}};
    document.getElementById('c-comm').value  = (d.commission_pct||15).toFixed(1);
    document.getElementById('c-acq').value   = (d.acquiring_pct ||2).toFixed(1);
    document.getElementById('c-log').value   = (d.logistics_pct ||0).toFixed(1);
    document.getElementById('costs-status').textContent = 'данные по кабинету загружены';
    recalcAllMargins();
  }}catch(e){{
    document.getElementById('costs-status').textContent = 'ошибка загрузки расходов';
  }}
}}

['c-comm','c-acq','c-log','c-drr','c-other'].forEach(id=>{{
  document.getElementById(id).addEventListener('input', recalcAllMargins);
}});

async function loadSummary(){{
  const resp=await fetch(`/api/competitor-prices/summary?cabinet_id=${{CAB}}`);
  const data=await resp.json();
  const rows=data.rows||[];
  document.getElementById('our-count').textContent=`(${{rows.length}} арт.)`;
  const tbody=document.getElementById('our-tbody');
  if(!rows.length){{
    tbody.innerHTML='<tr><td colspan="13" style="padding:24px;text-align:center;color:#aaa">Нет данных. Запустите wb_competitor_finder.py и wb_price_monitor.py</td></tr>';
    return;
  }}
  tbody.innerHTML=rows.map(r=>`
    <tr data-nm="${{r.our_nm_id}}" data-price="${{r.our_price||0}}" data-name="${{(r.our_name||'').replace(/"/g,'&quot;').replace(/'/g,'&#39;')}}" onclick="if(!event.target.closest('.new-price-wrap'))selectRow(this)">
      <td><b>${{r.our_nm_id}}</b></td>
      <td style="max-width:240px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">${{r.our_name||'—'}}</td>
      <td class="price">${{fmt(r.our_price)}}</td>
      <td class="our-margin" style="white-space:nowrap"><span class="no-data">…</span></td>
      <td>
        <div class="new-price-wrap" onclick="event.stopPropagation()">
          <button onclick="adjustPrice(this.closest('tr'),-100)">↓</button>
          <input class="new-price-input" type="number" value="${{r.our_price||0}}" min="0" step="100" oninput="onNewPriceInput(this.closest('tr'))">
          <button onclick="adjustPrice(this.closest('tr'),100)">↑</button>
        </div>
      </td>
      <td class="new-margin" style="white-space:nowrap"><span class="no-data">…</span></td>
      <td>${{fmtN(r.our_stock)}}</td>
      <td><span class="badge badge-count">${{r.competitors_count||0}}</span></td>
      <td>${{fmt(r.median_price)}}</td>
      <td class="price-low">${{fmt(r.min_price)}}</td>
      <td class="price-high">${{fmt(r.max_price)}}</td>
      <td>${{posLabel(r.our_price,r.median_price)}}</td>
    </tr>`).join('');
  loadCosts();
}}

function adjustPrice(tr, delta){{
  const inp = tr.querySelector('.new-price-input');
  if(!inp) return;
  const newVal = Math.max(0, (+inp.value||0) + delta);
  inp.value = newVal;
  onNewPriceInput(tr);
}}

function onNewPriceInput(tr){{
  const inputs = getInputs();
  const nm       = tr.dataset.nm;
  const art      = costsData.articles[nm] || {{}};
  const rowInputs = {{
    comm:  art.commission_pct ?? inputs.comm,
    acq:   art.acquiring_pct  ?? inputs.acq,
    log:   art.logistics_pct  ?? inputs.log,
    drr:   inputs.drr,
    other: inputs.other,
  }};
  const costPrice = art.cost_price ?? null;
  const hasCost   = costPrice != null && costPrice > 0;
  const inp = tr.querySelector('.new-price-input');
  const cell = tr.querySelector('.new-margin');
  if(cell) cell.innerHTML = fmtMargin(calcMargin(+inp.value,costPrice,hasCost,rowInputs), hasCost);
  // Обновляем блок экономики если эта строка выбрана
  if(tr.classList.contains('selected')) fillEconRow('econ-new', +inp.value, art, rowInputs);
}}

function econCell(price, pct, isRub){{
  if(!price) return '<td>—</td>';
  if(isRub) return `<td>${{Math.round(price).toLocaleString('ru-RU')}} ₽</td>`;
  const rub = Math.round(price * pct / 100);
  return `<td>${{rub.toLocaleString('ru-RU')}} ₽<br><span style="color:#9ca3af">${{pct.toFixed(1)}}%</span></td>`;
}}

function fillEconRow(rowId, price, art, rowInputs){{
  const {{comm,acq,log,drr,other}} = rowInputs;
  const costPrice = art.cost_price ?? null;
  const hasCost   = costPrice != null && costPrice > 0;
  const m = calcMargin(price, costPrice, hasCost, rowInputs);
  const mStr = m==null ? '—' : (hasCost?'':'<span style="color:#f59e0b">⚠</span> ')+'<b>'+(m>=0?'':'')+ m.toFixed(1)+'%</b>';
  const mStyle = m==null?'' : m>=20?'color:#16a34a' : m>=10?'' : 'color:#dc2626';
  const row = document.getElementById(rowId);
  if(!row) return;
  row.innerHTML = `
    <td class="econ-lbl">${{row.id==='econ-cur'?'Текущая':'Новая'}}</td>
    <td><b>${{price.toLocaleString('ru-RU')}} ₽</b></td>
    ${{econCell(price,comm)}}
    ${{econCell(price,log)}}
    ${{econCell(price,acq)}}
    ${{econCell(price,drr)}}
    ${{econCell(price,other)}}
    <td>${{hasCost ? Math.round(costPrice).toLocaleString('ru-RU')+' ₽' : '<span style="color:#f59e0b">нет</span>'}}</td>
    <td style="${{mStyle}}">${{mStr}}</td>
  `;
}}

function updateEconBlock(nmId, selectedTr){{
  const art      = costsData.articles[nmId] || {{}};
  const inputs   = getInputs();
  const rowInputs = {{
    comm:  art.commission_pct ?? inputs.comm,
    acq:   art.acquiring_pct  ?? inputs.acq,
    log:   art.logistics_pct  ?? inputs.log,
    drr:   inputs.drr,
    other: inputs.other,
  }};
  const curPrice = +selectedTr.dataset.price || 0;
  const inp = selectedTr.querySelector('.new-price-input');
  const newPrice = inp ? +inp.value : curPrice;
  fillEconRow('econ-cur', curPrice, art, rowInputs);
  fillEconRow('econ-new', newPrice, art, rowInputs);
}}
async function selectRow(el){{
  const nmId=el.dataset.nm;
  const name=el.dataset.name||'';
  document.querySelectorAll('#our-tbody tr.selected').forEach(r=>r.classList.remove('selected'));
  el.classList.add('selected');
  document.getElementById('detail-title').textContent=name;
  const panel=document.getElementById('detail-panel');
  panel.classList.add('visible');
  updateEconBlock(nmId, el);
  document.getElementById('detail-tbody').innerHTML='<tr><td colspan="11" class="spinner">Загрузка…</td></tr>';
  const resp=await fetch(`/api/competitor-prices/detail?cabinet_id=${{CAB}}&nm_id=${{nmId}}`);
  const data=await resp.json();
  const rows=data.rows||[];
  document.getElementById('detail-count').textContent=`(${{rows.length}} конкурентов)`;
  if(!rows.length){{
    document.getElementById('detail-tbody').innerHTML='<tr><td colspan="11" style="padding:16px;color:#aaa">Нет данных</td></tr>';
    return;
  }}
  const uniqueNms=[...new Set(rows.map(r=>r.comp_nm_id))];
  document.getElementById('detail-tbody').innerHTML=rows.map(r=>`
    <tr data-comp-nm="${{r.comp_nm_id}}">
      <td><small>${{r.comp_nm_id}}</small></td>
      <td style="max-width:260px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis" title="${{(r.comp_name||'').replace(/"/g,'&quot;')}}">${{r.comp_name||'—'}}</td>
      <td>${{r.comp_brand||'—'}}</td>
      <td style="max-width:160px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">${{r.comp_seller||'—'}}</td>
      <td class="price">${{fmt(r.price)}}</td>
      <td style="color:#888">${{fmt(r.price_basic)}}</td>
      <td>${{fmtN(r.stock)}}</td>
      <td class="live-price" style="background:#f1f8e9;color:#aaa;font-size:11px">…</td>
      <td class="live-stock" style="background:#f1f8e9;color:#aaa;font-size:11px">…</td>
      <td>${{r.rating||'—'}}</td>
      <td>${{fmtN(r.feedbacks)}}</td>
      <td style="color:#aaa;font-size:11px">${{r.last_date||'—'}}</td>
      <td><a class="link-btn" href="${{r.url}}" target="_blank">WB ↗</a></td>
    </tr>`).join('');
  // Подгружаем живые цены батчами по 20
  loadLivePrices(uniqueNms);
}}
async function loadLivePrices(nmIds){{
  const BATCH=20;
  for(let i=0;i<nmIds.length;i+=BATCH){{
    const batch=nmIds.slice(i,i+BATCH);
    try{{
      const r=await fetch(`/api/competitor-prices/live-prices?nm=${{batch.join(';')}}`);
      const live=await r.json();
      if(live._error){{
        if(i===0) document.querySelectorAll('#detail-tbody .live-price').forEach(td=>{{td.textContent='нет агента';td.style.color='#e57373';}});
        break;
      }}
      batch.forEach(nm=>{{
        const p=live[String(nm)];
        document.querySelectorAll(`#detail-tbody tr[data-comp-nm="${{nm}}"]`).forEach(tr=>{{
          const lp=tr.querySelector('.live-price');
          const ls=tr.querySelector('.live-stock');
          if(p){{
            lp.innerHTML=`<b style="color:#2e7d32">${{p.price?p.price.toLocaleString('ru-RU')+' ₽':'—'}}</b>`;
            ls.innerHTML=`<span style="color:#1565c0">${{p.stock??'—'}}</span>`;
          }} else {{
            lp.textContent='—'; ls.textContent='—';
          }}
          lp.style.fontSize=''; ls.style.fontSize='';
        }});
      }});
    }}catch(e){{
      console.warn('live-prices error',e);
      break;
    }}
  }}
}}
loadSummary();
</script>
</body>
</html>
"""


COMMENTS_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Комментарии WB</title>
  <style>
    :root{{--bg:#f8fafc;--ink:#111827;--muted:#6b7280;--accent:#0f766e;--line:#e5e7eb;}}
    *{{box-sizing:border-box;}}
    body{{margin:0;font-family:"Segoe UI","Trebuchet MS",sans-serif;color:var(--ink);
         background:linear-gradient(180deg,#e0f2fe,transparent 240px),var(--bg);}}
    /* WB comments styles */
    .wrap{{width:calc(100% - 24px);max-width:3000px;margin:0 auto;padding:8px 0 28px;}}
    .top{{display:flex;align-items:center;gap:8px;margin-bottom:6px;}}
    .top h1{{font-size:.82rem;font-weight:700;color:var(--muted);white-space:nowrap;margin:0;flex-shrink:0;}}
    .report-nav{{display:flex;align-items:center;flex-wrap:nowrap;overflow-x:auto;gap:2px;flex:1;
                 scrollbar-width:none;min-width:0;}}
    .report-nav::-webkit-scrollbar{{display:none;}}
    /* Filters */
    .filters{{background:#fff;border:1px solid var(--line);border-radius:8px;
              padding:8px 12px;margin-bottom:6px;display:flex;flex-wrap:wrap;gap:6px;align-items:flex-end;}}
    .fgroup{{display:flex;flex-direction:column;gap:2px;}}
    .fgroup label{{font-size:10px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.04em;}}
    .fgroup select,.fgroup input{{height:28px;padding:0 7px;border:1px solid #cbd5e1;border-radius:6px;
                                   font-size:11px;background:#f8fafc;color:var(--ink);min-width:0;}}
    .fgroup select:focus,.fgroup input:focus{{outline:none;border-color:var(--accent);background:#fff;}}
    .fgroup-num{{display:flex;flex-direction:column;gap:2px;}}
    .fgroup-num label{{font-size:10px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.04em;}}
    .fgroup-num .num-row{{display:flex;gap:2px;}}
    .fgroup-num select{{height:28px;width:44px;padding:0 2px;border:1px solid #cbd5e1;border-radius:6px 0 0 6px;
                        font-size:12px;background:#f8fafc;border-right:0;}}
    .fgroup-num input{{height:28px;width:72px;padding:0 5px;border:1px solid #cbd5e1;border-radius:0 6px 6px 0;
                       font-size:11px;background:#f8fafc;}}
    .fgroup-num select:focus,.fgroup-num input:focus{{outline:none;border-color:var(--accent);background:#fff;z-index:1;position:relative;}}
    .btn-reset{{height:28px;padding:0 10px;border:1px solid #cbd5e1;border-radius:6px;background:#fff;
                font-size:11px;cursor:pointer;color:#64748b;align-self:flex-end;margin-left:2px;}}
    .btn-reset:hover{{background:#f1f5f9;}}
    #comment-panel{{background:#fff;border:1px solid #a7f3d0;border-radius:8px;padding:10px 14px;margin-bottom:6px;}}
    #comment-panel h3{{margin:0 0 6px;font-size:.78rem;font-weight:800;color:#0f172a;}}
    #cp-hint{{font-size:11px;color:#94a3b8;margin-bottom:6px;}}
    .comment-fields{{display:flex;flex-wrap:wrap;gap:6px;}}
    .comment-field{{display:flex;align-items:flex-start;gap:5px;flex:1;min-width:160px;}}
    .comment-field label{{font-size:10px;font-weight:700;color:#64748b;text-transform:uppercase;
                          letter-spacing:.04em;width:52px;flex-shrink:0;padding-top:6px;}}
    .comment-field textarea{{flex:1;border:1px solid #cbd5e1;border-radius:6px;font-size:11px;
                              padding:5px 7px;resize:none;height:30px;background:#f8fafc;color:#0f172a;font-family:inherit;}}
    .comment-field textarea:focus{{outline:none;border-color:var(--accent);background:#fff;}}
    .comment-field textarea:disabled{{opacity:.45;cursor:not-allowed;}}
    .panel-bottom{{display:flex;align-items:center;margin-top:8px;gap:8px;}}
    #save-btn{{padding:5px 16px;border:none;border-radius:6px;background:var(--accent);
               color:#fff;font-weight:700;font-size:12px;cursor:pointer;}}
    #save-btn:disabled{{opacity:.5;cursor:default;}}
    #save-status{{font-size:11px;color:var(--muted);}}
    .meta{{font-size:.75rem;color:var(--muted);margin-bottom:4px;}}
    .tbl{{background:#fff;border:1px solid var(--line);border-radius:8px;overflow:auto;max-height:480px;}}
    table{{width:100%;border-collapse:collapse;font-size:11px;table-layout:fixed;min-width:830px;}}
    th,td{{border-bottom:1px solid #eef2f7;padding:4px 6px;text-align:right;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
    th{{position:sticky;top:0;z-index:1;background:#f8fafc;color:#334155;font-weight:800;cursor:pointer;user-select:none;position:relative;}}
    th:hover{{background:#f1f5f9;}}
    th .sort-ic{{margin-left:3px;color:#94a3b8;font-size:.7em;}}
    .col-ck{{width:22px;min-width:22px;text-align:center;position:sticky;left:0;z-index:2;background:#fff;cursor:default;}}
    th.col-ck{{z-index:3;background:#f8fafc;cursor:default;}}
    .col-art{{text-align:left;position:sticky;left:22px;z-index:2;background:#fff;overflow:hidden;text-overflow:ellipsis;}}
    th.col-art{{z-index:3;background:#f8fafc;}}
    .col-name{{text-align:left;overflow:hidden;text-overflow:ellipsis;}}
    .col-comment{{text-align:left;overflow:hidden;text-overflow:ellipsis;}}
    .resize-handle{{position:absolute;right:0;top:0;height:100%;width:5px;cursor:col-resize;z-index:10;}}
    .resize-handle:hover,.resize-handle:active{{background:rgba(15,118,110,.3);}}
    input[type=checkbox]{{width:14px;height:14px;cursor:pointer;accent-color:var(--accent);}}
    .art-link{{margin-left:3px;font-size:.75em;color:#94a3b8;text-decoration:none;opacity:.7;}}
    .art-link:hover{{color:var(--accent);opacity:1;}}
    .art-copy{{margin-left:3px;font-size:.75em;color:#94a3b8;background:none;border:none;cursor:pointer;padding:0 1px;opacity:.7;line-height:1;}}
    .art-copy:hover{{color:var(--accent);opacity:1;}}
    .art-copy.copied{{color:#16a34a;opacity:1;}}
    tbody tr.data-row{{cursor:pointer;}}
    tbody tr:hover td{{background:#f0fdfa!important;}}
    tbody tr.row-selected td{{background:#d1fae5!important;}}
    tbody tr.row-selected td.col-ck,tbody tr.row-selected td.col-art{{background:#d1fae5!important;}}
    .empty{{padding:18px;color:var(--muted);font-size:.9rem;}}
    .sel-history{{margin-top:10px;border:2px solid #a7f3d0;border-radius:8px;background:#fff;padding:10px 14px;}}
    .sel-history h3{{margin:0 0 8px;font-size:.78rem;font-weight:800;color:#065f46;}}
    .sel-hint{{font-size:11px;color:#94a3b8;}}
    .sh-article{{margin-bottom:14px;}}
    .sh-art-title{{font-size:.78rem;font-weight:800;color:#0f172a;margin-bottom:6px;display:flex;align-items:center;gap:6px;}}
    .sh-art-title .art-tag{{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:6px;padding:1px 8px;font-size:11px;}}
    .sh-art-title .art-name{{font-size:11px;color:#64748b;font-weight:400;}}
    .sh-cols{{display:flex;gap:6px;overflow-x:auto;}}
    .sh-col{{flex:1;min-width:140px;}}
    .sh-col-hdr{{font-size:10px;font-weight:800;color:#334155;text-transform:uppercase;letter-spacing:.04em;
                 padding:3px 6px;background:#f8fafc;border:1px solid #e5e7eb;border-bottom:2px solid var(--accent);
                 border-radius:4px 4px 0 0;white-space:nowrap;}}
    .sh-entry{{padding:3px 6px;border-left:1px solid #e5e7eb;border-right:1px solid #e5e7eb;border-bottom:1px solid #f1f5f9;}}
    .sh-entry:last-child{{border-bottom:1px solid #e5e7eb;border-radius:0 0 4px 4px;}}
    .sh-entry-date{{font-size:10px;color:#94a3b8;font-weight:700;}}
    .sh-entry.age-1{{background:#d1fae5;}}.sh-entry.age-1 .sh-entry-date{{color:#065f46;font-weight:900;}}
    .sh-entry.age-2{{background:#fef3c7;}}.sh-entry.age-2 .sh-entry-date{{color:#92400e;font-weight:900;}}
    .sh-entry.age-3{{background:#ffedd5;}}.sh-entry.age-3 .sh-entry-date{{color:#9a3412;font-weight:900;}}
    .sh-entry-text{{font-size:11px;color:#0f172a;line-height:1.4;word-break:break-word;}}
    .sh-col-empty{{padding:3px 6px;font-size:11px;color:#94a3b8;border-left:1px solid #e5e7eb;
                   border-right:1px solid #e5e7eb;border-bottom:1px solid #e5e7eb;border-radius:0 0 4px 4px;}}
    .month-comments{{margin-top:10px;background:#fff;border:1px solid var(--line);border-radius:8px;padding:10px 14px;
                     display:grid;grid-template-columns:300px minmax(0,1fr);gap:12px;align-items:start;}}
    .month-cal-head{{display:flex;align-items:center;justify-content:space-between;margin-bottom:8px;gap:8px;}}
    .month-cal-title{{font-size:.78rem;font-weight:800;color:#0f172a;text-transform:capitalize;}}
    .month-nav{{display:flex;gap:4px;}}
    .month-nav button{{width:28px;height:26px;border:1px solid #cbd5e1;border-radius:6px;background:#f8fafc;
                       color:#334155;font-size:15px;font-weight:800;cursor:pointer;line-height:1;}}
    .month-nav button:hover{{background:#f1f5f9;border-color:#94a3b8;}}
    .month-weekdays,.month-grid{{display:grid;grid-template-columns:repeat(7,1fr);gap:4px;}}
    .month-weekdays span{{font-size:10px;font-weight:800;color:#94a3b8;text-align:center;text-transform:uppercase;}}
    .month-day{{height:42px;border:1px solid #e5e7eb;border-radius:6px;background:#fff;color:#334155;
                display:flex;flex-direction:column;align-items:flex-start;justify-content:space-between;
                padding:4px 5px;font-size:11px;cursor:pointer;}}
    .month-day:hover{{border-color:var(--accent);background:#f0fdfa;}}
    .month-day.is-empty{{visibility:hidden;pointer-events:none;}}
    .month-day.is-today{{border-color:#99f6e4;box-shadow:inset 0 0 0 1px #99f6e4;}}
    .month-day.is-selected{{background:#0f766e;border-color:#0f766e;color:#fff;}}
    .month-day-num{{font-weight:800;line-height:1;}}
    .month-day-count{{font-size:10px;font-weight:800;color:#0f766e;background:#ccfbf1;border-radius:999px;
                      padding:1px 5px;align-self:flex-end;min-width:18px;text-align:center;}}
    .month-day.is-selected .month-day-count{{color:#0f766e;background:#fff;}}
    .month-list-head{{display:flex;align-items:baseline;justify-content:space-between;gap:8px;margin-bottom:8px;}}
    .month-list-head h3{{margin:0;font-size:.78rem;font-weight:800;color:#0f172a;}}
    .month-list-count{{font-size:11px;color:#94a3b8;white-space:nowrap;}}
    .month-list{{max-height:360px;overflow:auto;border:1px solid #eef2f7;border-radius:8px;background:#fff;}}
    .month-comment-item{{display:flex;align-items:center;gap:6px;padding:3px 7px;border-bottom:1px solid #eef2f7;
                         min-height:24px;font-size:11px;line-height:1.25;}}
    .month-comment-item:last-child{{border-bottom:0;}}
    .month-comment-article{{font-size:11px;font-weight:900;color:#0f172a;white-space:nowrap;}}
    .month-comment-sep{{color:#94a3b8;}}
    .month-comment-text{{min-width:0;flex:1;color:#0f172a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
    .month-comment-item .type-badge{{flex:0 0 auto;padding:0 5px;font-size:9px;line-height:16px;}}
    @media (max-width:760px){{.month-comments{{grid-template-columns:1fr;}}}}
    .log-section{{margin-top:10px;background:#fff;border:1px solid var(--line);border-radius:8px;padding:10px 14px;}}
    .log-section h3{{margin:0 0 8px;font-size:.78rem;font-weight:800;}}
    .log-filter{{display:flex;gap:6px;align-items:center;margin-bottom:8px;flex-wrap:wrap;}}
    .log-filter input,.log-filter select{{height:28px;padding:0 7px;border:1px solid #cbd5e1;border-radius:6px;font-size:11px;}}
    .log-filter input{{width:180px;}}
    .log-table{{width:100%;border-collapse:collapse;font-size:11px;}}
    .log-table th,.log-table td{{border-bottom:1px solid #eef2f7;padding:4px 6px;text-align:left;}}
    .log-table th{{background:#f8fafc;font-weight:800;color:#334155;white-space:nowrap;}}
    .type-badge{{display:inline-block;padding:1px 6px;border-radius:999px;font-size:10px;font-weight:700;}}
    .type-general{{background:#dbeafe;color:#1d4ed8;}}
    .type-cpc{{background:#fce7f3;color:#9d174d;}}
    .type-shelves{{background:#d1fae5;color:#065f46;}}
    .type-unified{{background:#fef9c3;color:#854d0e;}}
    .type-search{{background:#ede9fe;color:#5b21b6;}}
  </style>
</head>
<body>
<div class="wrap">
  <div class="top">
    <h1>Комментарии WB</h1>
    <div class="report-nav">{report_nav}</div>
  </div>

  <div class="filters" id="filters">
    <div class="fgroup">
      <label>Поиск</label>
      <input id="f-text" type="text" placeholder="Артикул или название" style="width:160px">
    </div>
    <div class="fgroup">
      <label>Предмет</label>
      <select id="f-subject"><option value="">Все предметы</option></select>
    </div>
    <div class="fgroup">
      <label>Ярлыки</label>
      <select id="f-strategy" multiple size="4" style="min-width:120px;height:70px;" title="Ctrl/Cmd для множественного выбора"></select>
    </div>
    <div class="fgroup">
      <label>Категория</label>
      <select id="f-revcat"><option value="">Все категории</option></select>
    </div>
    <div class="fgroup-num">
      <label>ДРР% 7д</label>
      <div class="num-row">
        <select id="f-drr-op"><option value="">=</option><option value="gt">&gt;</option><option value="lt">&lt;</option></select>
        <input id="f-drr-val" type="number" step="0.1" placeholder="—">
      </div>
    </div>
    <div class="fgroup-num">
      <label>Ск. 7д</label>
      <div class="num-row">
        <select id="f-spd-op"><option value="">=</option><option value="gt">&gt;</option><option value="lt">&lt;</option></select>
        <input id="f-spd-val" type="number" step="0.1" placeholder="—">
      </div>
    </div>
    <div class="fgroup-num">
      <label>Оборачив.</label>
      <div class="num-row">
        <select id="f-trn-op"><option value="">=</option><option value="gt">&gt;</option><option value="lt">&lt;</option></select>
        <input id="f-trn-val" type="number" placeholder="—">
      </div>
    </div>
    <button class="btn-reset" id="btn-reset">Сброс</button>
  </div>

  <div id="comment-panel">
    <h3>Добавить комментарий &mdash; выбрано: <span id="sel-count">0</span> артикулов</h3>
    <div id="cp-hint">Выберите строки в таблице (клик, Shift+клик, Ctrl+клик)</div>
    <div class="comment-fields">
      <div class="comment-field"><label>Общий</label>
        <textarea id="txt-general" placeholder="Изменили фото, вошли в акцию..." disabled></textarea></div>
      <div class="comment-field"><label>CPC</label>
        <textarea id="txt-cpc" placeholder="Ставка CPC..." disabled></textarea></div>
      <div class="comment-field"><label>Полки</label>
        <textarea id="txt-shelves" placeholder="Полочная реклама..." disabled></textarea></div>
      <div class="comment-field"><label>Единая</label>
        <textarea id="txt-unified" placeholder="Единая ставка..." disabled></textarea></div>
      <div class="comment-field"><label>Поиск</label>
        <textarea id="txt-search" placeholder="Поисковая реклама..." disabled></textarea></div>
    </div>
    <div class="panel-bottom">
      <button id="save-btn" disabled>Сохранить</button>
      <span id="save-status"></span>
    </div>
  </div>

  <div class="meta" id="meta">Загрузка...</div>
  <div class="tbl" id="tbl-wrap"></div>

  <div class="sel-history" id="sel-history">
    <h3>История выбранных артикулов</h3>
    <div id="sel-history-body"><div class="sel-hint">Выберите артикулы в таблице чтобы увидеть историю.</div></div>
  </div>

  <div class="month-comments">
    <div class="month-cal">
      <div class="month-cal-head">
        <div class="month-cal-title" id="comment-month-title"></div>
        <div class="month-nav">
          <button id="comment-month-prev" type="button" title="Предыдущий месяц">&lsaquo;</button>
          <button id="comment-month-next" type="button" title="Следующий месяц">&rsaquo;</button>
        </div>
      </div>
      <div class="month-weekdays">
        <span>Пн</span><span>Вт</span><span>Ср</span><span>Чт</span><span>Пт</span><span>Сб</span><span>Вс</span>
      </div>
      <div class="month-grid" id="comment-calendar-grid"></div>
    </div>
    <div class="month-list-wrap">
      <div class="month-list-head">
        <h3 id="comment-day-title">Комментарии за день</h3>
        <span class="month-list-count" id="comment-day-count"></span>
      </div>
      <div class="month-list" id="comment-day-list"></div>
    </div>
  </div>

  <div class="log-section">
    <h3>История всех комментариев</h3>
    <div class="log-filter">
      <input id="log-search" type="text" placeholder="Артикул или текст...">
      <select id="log-type">
        <option value="">Все типы</option>
        <option value="general">Общий</option>
        <option value="cpc">CPC</option>
        <option value="shelves">Полки</option>
        <option value="unified">Единая</option>
        <option value="search">Поиск</option>
      </select>
    </div>
    <table class="log-table">
      <thead><tr><th>Дата</th><th>Артикул</th><th>Тип</th><th>Комментарий</th></tr></thead>
      <tbody id="log-body"></tbody>
    </table>
  </div>
</div>
<script>
const TYPE_LABELS={{general:'Общий',cpc:'CPC',shelves:'Полки',unified:'Единая',search:'Поиск'}};
const TYPE_CSS={{general:'type-general',cpc:'type-cpc',shelves:'type-shelves',unified:'type-unified',search:'type-search'}};
const COMMENT_TYPES=['general','cpc','shelves','unified','search'];
const FIELDS=[
  {{id:'txt-general',type:'general'}},{{id:'txt-cpc',type:'cpc'}},
  {{id:'txt-shelves',type:'shelves'}},{{id:'txt-unified',type:'unified'}},{{id:'txt-search',type:'search'}},
];
const COLS=[
  {{k:'article',     label:'Артикул',    num:false, w:160, resize:true}},
  {{k:'name',        label:'Название',   num:false, stretch:true}},
  {{k:'drr_7d',      label:'ДРР% 7д',    num:true,  w:50}},
  {{k:'drr_yday2',   label:'ДРР% поза',  num:true,  w:50}},
  {{k:'drr_yday',    label:'ДРР% вчера', num:true,  w:50}},
  {{k:'speed_7d',    label:'Ск. 7д',     num:true,  w:44}},
  {{k:'speed_yday2', label:'Ск. поза',   num:true,  w:40}},
  {{k:'speed_yday',  label:'Ск. вчера',  num:true,  w:40}},
  {{k:'turnover',    label:'Оборачив.',  num:true,  w:52}},
  {{k:'stock',       label:'Остатки',    num:true,  w:52}},
  {{k:'last_general',label:'Общий',      num:false, w:180, resize:true}},
  {{k:'last_cpc',    label:'CPC',        num:false, w:150, resize:true}},
  {{k:'last_shelves',label:'Полки',      num:false, w:150, resize:true}},
  {{k:'last_unified',label:'Единая',     num:false, w:150, resize:true}},
  {{k:'last_search', label:'Поиск',      num:false, w:150, resize:true}},
];
const _CW_KEY='wb.comments.col_widths';
const colWidths=new Map(Object.entries(JSON.parse(localStorage.getItem(_CW_KEY)||'{{}}')).map(([k,v])=>[k,+v]));
let _rsz=null;
function _doResize(e){{
  if(!_rsz)return;
  const nw=Math.max(50,_rsz.w+(e.clientX-_rsz.x));
  colWidths.set(_rsz.col,nw);
  const c=document.querySelector('col[data-col="'+_rsz.col+'"]');
  if(c)c.style.width=nw+'px';
}}
function _stopResize(){{
  _rsz=null;
  document.removeEventListener('mousemove',_doResize);
  document.removeEventListener('mouseup',_stopResize);
  localStorage.setItem(_CW_KEY,JSON.stringify(Object.fromEntries(colWidths)));
}}
let allRows=[],logRows=[],selected=new Set();
let sortKey='article',sortDir=1,lastClickedNm=null;
let calendarMonthDate=new Date();
calendarMonthDate.setDate(1);
let selectedCalendarDate='';
let calendarInitialized=false;

function esc(s){{return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}}
function numVal(s){{const v=parseFloat(s);return isNaN(v)?null:v;}}
function cmpNum(a,b){{if(a==null&&b==null)return 0;if(a==null)return 1;if(b==null)return -1;return a-b;}}
function pad2(n){{return String(n).padStart(2,'0');}}
function dateKey(d){{return d.getFullYear()+'-'+pad2(d.getMonth()+1)+'-'+pad2(d.getDate());}}
function parseDateKey(key){{
  const p=(key||'').slice(0,10).split('-').map(Number);
  if(p.length!==3||p.some(Number.isNaN))return null;
  return new Date(p[0],p[1]-1,p[2]);
}}
function fmtDateRu(key){{
  const d=parseDateKey(key);
  if(!d)return key||'';
  return pad2(d.getDate())+'.'+pad2(d.getMonth()+1)+'.'+d.getFullYear();
}}

function numMatch(val,opId,inpId){{
  const op=document.getElementById(opId).value;
  const thr=numVal(document.getElementById(inpId).value);
  if(thr===null)return true;if(val==null)return false;
  if(op===''||op==='eq')return Math.abs(val-thr)<0.001;
  if(op==='gt')return val>thr;if(op==='lt')return val<thr;
  return true;
}}

function getFilteredRows(){{
  const txt=(document.getElementById('f-text').value||'').toLowerCase().trim();
  const subj=document.getElementById('f-subject').value;
  const selectedLabels=[...document.getElementById('f-strategy').selectedOptions].map(o=>o.value);
  const revc=document.getElementById('f-revcat').value;
  return allRows.filter(r=>{{
    if(txt&&!(r.article+' '+r.name).toLowerCase().includes(txt))return false;
    if(subj&&r.subject!==subj)return false;
    if(selectedLabels.length>0&&!selectedLabels.some(l=>(r.labels||[]).map(x=>x.toLowerCase()).includes(l.toLowerCase())))return false;
    if(revc&&r.revenue_category!==revc)return false;
    if(!numMatch(r.drr_7d,  'f-drr-op','f-drr-val'))return false;
    if(!numMatch(r.speed_7d,'f-spd-op','f-spd-val'))return false;
    if(!numMatch(r.turnover,'f-trn-op','f-trn-val'))return false;
    return true;
  }});
}}

function getSortedRows(rows){{
  return [...rows].sort((a,b)=>{{
    const col=COLS.find(c=>c.k===sortKey);if(!col)return 0;
    let av=a[sortKey],bv=b[sortKey];
    if(sortKey.startsWith('drr_')||sortKey.startsWith('speed_')){{av=av??0;bv=bv??0;}}
    if(col.num)return sortDir*cmpNum(av,bv);
    return sortDir*((av||'').localeCompare(bv||'','ru'));
  }});
}}

function populateDropdowns(){{
  const sets={{subject:new Set(),strategy:new Set(),revenue_category:new Set()}};
  allRows.forEach(r=>{{
    if(r.subject)sets.subject.add(r.subject);
    (r.labels||[]).forEach(l=>{{if(l)sets.strategy.add(l);}});
    if(r.revenue_category)sets.revenue_category.add(r.revenue_category);
  }});
  function fill(id,set){{
    const sel=document.getElementById(id);const cur=sel.value;
    while(sel.options.length>1)sel.remove(1);
    [...set].sort((a,b)=>a.localeCompare(b,'ru')).forEach(v=>{{
      const o=document.createElement('option');o.value=v;o.textContent=v;sel.appendChild(o);
    }});
    if(cur)sel.value=cur;
  }}
  fill('f-subject',sets.subject);fill('f-strategy',sets.strategy);fill('f-revcat',sets.revenue_category);
}}

async function loadData(){{
  document.getElementById('meta').textContent='Загрузка...';
  try{{
    const r=await fetch('/api/analytics/comments-data');
    const d=await r.json();allRows=d.rows||[];populateDropdowns();renderTable();
  }}catch(e){{document.getElementById('meta').textContent='Ошибка: '+e;}}
}}
async function loadLog(){{
  try{{
    const r=await fetch('/api/analytics/comment-log?limit=2000');
    const d=await r.json();logRows=d.rows||[];renderLog();renderSelHistory();renderCommentCalendar();
  }}catch(e){{}}
}}

function renderTable(){{
  const rows=getSortedRows(getFilteredRows());
  const el=document.getElementById('tbl-wrap');
  document.getElementById('meta').textContent='Артикулов: '+rows.length+' из '+allRows.length;
  if(!rows.length){{el.innerHTML='<div class="empty">Нет данных.</div>';return;}}

  const cgHtml='<colgroup><col style="width:22px;min-width:22px">'+COLS.map(c=>{{
    const hdrMin=Math.ceil(c.label.length*7+20);
    if(c.stretch){{
      const sw=colWidths.has(c.k)?colWidths.get(c.k):null;
      const mw=Math.max(80,hdrMin);
      return sw?'<col data-col="'+c.k+'" style="width:'+sw+'px;min-width:'+mw+'px">'
               :'<col data-col="'+c.k+'" style="min-width:'+mw+'px">';
    }}
    if(c.resize){{
      const w=colWidths.has(c.k)?colWidths.get(c.k):c.w;
      const mw=Math.max(50,hdrMin);
      return '<col data-col="'+c.k+'" style="width:'+w+'px;min-width:'+mw+'px">';
    }}
    const mw=Math.max(c.w,hdrMin);
    return '<col data-col="'+c.k+'" style="width:'+mw+'px;min-width:'+mw+'px">';
  }}).join('')+'</colgroup>';
  const thHtml=COLS.map(c=>{{
    const active=c.k===sortKey;const ic=active?(sortDir>0?'↑':'↓'):'↕';
    const cls=c.k==='article'?'col-art':c.k==='name'?'col-name':c.k.startsWith('last_')?'col-comment':'';
    const rz=(c.resize||c.stretch)?'<span class="resize-handle" data-col="'+c.k+'"></span>':'';
    return `<th data-key="${{c.k}}"${{cls?' class="'+cls+'"':''}}>${{c.label}}${{rz}}<span class="sort-ic">${{ic}}</span></th>`;
  }}).join('');

  function fmtPct(v){{return v!=null?(v===0?'0%':v.toFixed(1)+'%'):'—';}}
  function fmtF(v){{return v?v.toFixed(1):'—';}}
  function fmtI(v){{return v!=null&&v>0?Math.round(v):(v===0?'0':'—');}}

  const wbBase='https://seller.wildberries.ru/new-goods/all-goods?analyticsNmId=';
  const trs=rows.map(r=>{{
    const sel=selected.has(r.nmId)?' row-selected':'';
    const links=`<button class="art-copy" data-copy="${{esc(r.article)}}" data-skip-select="1" title="Копировать">⎘</button>`+
      `<a class="art-link" href="${{wbBase}}${{r.nmId}}" target="_blank" rel="noopener" data-skip-select="1" title="WB кабинет">🔗</a>`;
    return `<tr class="data-row${{sel}}" data-nmid="${{r.nmId}}">
      <td class="col-ck"><input type="checkbox" class="row-ck" data-nmid="${{r.nmId}}" ${{selected.has(r.nmId)?'checked':''}}></td>
      <td class="col-art" title="${{esc(r.article)}}">${{esc(r.article)}}${{links}}</td>
      <td class="col-name" title="${{esc(r.name)}}">${{esc(r.name)}}<a class="art-link" href="https://www.wildberries.ru/catalog/${{r.nmId}}/detail.aspx" target="_blank" rel="noopener" data-skip-select="1" title="WB товар">↗</a></td>
      <td>${{fmtPct(r.drr_7d)}}</td><td>${{fmtPct(r.drr_yday2)}}</td><td>${{fmtPct(r.drr_yday)}}</td>
      <td>${{fmtF(r.speed_7d)}}</td><td>${{fmtI(r.speed_yday2)}}</td><td>${{fmtI(r.speed_yday)}}</td>
      <td>${{r.turnover!=null?Math.round(r.turnover)+'д':'—'}}</td>
      <td>${{r.stock||'—'}}</td>
      <td class="col-comment" title="${{esc(r.last_general)}}">${{esc(r.last_general)}}</td>
      <td class="col-comment" title="${{esc(r.last_cpc)}}">${{esc(r.last_cpc)}}</td>
      <td class="col-comment" title="${{esc(r.last_shelves)}}">${{esc(r.last_shelves)}}</td>
      <td class="col-comment" title="${{esc(r.last_unified)}}">${{esc(r.last_unified)}}</td>
      <td class="col-comment" title="${{esc(r.last_search)}}">${{esc(r.last_search)}}</td>
    </tr>`;
  }}).join('');

  const allVis=rows.length>0&&rows.every(r=>selected.has(r.nmId));
  const ckAll=`<th class="col-ck"><input type="checkbox" id="ck-all" ${{allVis?'checked':''}}></th>`;
  el.innerHTML=`<table>${{cgHtml}}<thead><tr>${{ckAll}}${{thHtml}}</tr></thead><tbody>${{trs}}</tbody></table>`;

  el.querySelectorAll('th[data-key]').forEach(th=>{{
    th.addEventListener('click',(e)=>{{
      if(e.target.classList.contains('resize-handle'))return;
      const k=th.dataset.key;
      if(sortKey===k)sortDir*=-1;else{{sortKey=k;sortDir=1;}}
      renderTable();
    }});
  }});
  el.querySelectorAll('.resize-handle').forEach(h=>{{
    h.addEventListener('mousedown',e=>{{
      const th=h.closest('th');
      _rsz={{col:h.dataset.col,x:e.clientX,w:th.offsetWidth}};
      document.addEventListener('mousemove',_doResize);
      document.addEventListener('mouseup',_stopResize);
      e.preventDefault();e.stopPropagation();
    }});
  }});
  document.getElementById('ck-all').addEventListener('change',e=>{{
    rows.forEach(r=>e.target.checked?selected.add(r.nmId):selected.delete(r.nmId));
    renderTable();updatePanel();
  }});
  el.querySelectorAll('.row-ck').forEach(ck=>{{
    ck.addEventListener('change',e=>{{
      e.stopPropagation();
      const nm=e.target.dataset.nmid;
      e.target.checked?selected.add(nm):selected.delete(nm);
      e.target.closest('tr').classList.toggle('row-selected',e.target.checked);
      lastClickedNm=nm;updatePanel();
    }});
  }});
  el.querySelectorAll('.art-copy').forEach(btn=>{{
    btn.addEventListener('click',e=>{{
      e.stopPropagation();
      navigator.clipboard.writeText(btn.dataset.copy).then(()=>{{
        btn.classList.add('copied');btn.textContent='✓';
        setTimeout(()=>{{btn.classList.remove('copied');btn.textContent='⎘';}},1500);
      }});
    }});
  }});
  el.querySelectorAll('tr.data-row').forEach(tr=>{{
    tr.addEventListener('click',e=>{{
      if(e.target.dataset.skipSelect==='1'||e.target.closest('[data-skip-select="1"]'))return;
      if(e.target.classList.contains('row-ck'))return;
      const nm=tr.dataset.nmid;
      const sortedRows=getSortedRows(getFilteredRows());
      if(e.shiftKey&&lastClickedNm){{
        const li=sortedRows.findIndex(r=>r.nmId===lastClickedNm);
        const ci=sortedRows.findIndex(r=>r.nmId===nm);
        if(li>=0&&ci>=0){{
          const [f,t]=[Math.min(li,ci),Math.max(li,ci)];
          if(!e.ctrlKey&&!e.metaKey)selected.clear();
          sortedRows.slice(f,t+1).forEach(r=>selected.add(r.nmId));
        }}
      }}else if(e.ctrlKey||e.metaKey){{
        selected.has(nm)?selected.delete(nm):selected.add(nm);lastClickedNm=nm;
      }}else{{
        selected.clear();selected.add(nm);lastClickedNm=nm;
      }}
      renderTable();updatePanel();
    }});
  }});
}}

function updatePanel(){{
  const cnt=selected.size;
  document.getElementById('sel-count').textContent=cnt;
  const hint=document.getElementById('cp-hint');
  hint.style.display=cnt>0?'none':'block';
  FIELDS.forEach(f=>{{document.getElementById(f.id).disabled=cnt===0;}});
  document.getElementById('save-btn').disabled=cnt===0;
  renderSelHistory();
}}

function renderSelHistory(){{
  const body=document.getElementById('sel-history-body');
  if(!selected.size){{
    body.innerHTML='<div class="sel-hint">Выберите артикулы в таблице чтобы увидеть историю.</div>';
    return;
  }}
  const _today=new Date();_today.setHours(0,0,0,0);
  function _ageCls(created_at){{
    const d=new Date((created_at||'').slice(0,10));d.setHours(0,0,0,0);
    const age=Math.round((_today-d)/86400000);
    return age===1?'age-1':age===2?'age-2':age===3?'age-3':'';
  }}
  const byArticle=new Map();
  allRows.filter(r=>selected.has(r.nmId)).forEach(r=>{{
    if(!byArticle.has(r.nmId)){{
      byArticle.set(r.nmId,{{article:r.article,name:r.name,types:{{}}}});
      COMMENT_TYPES.forEach(t=>{{byArticle.get(r.nmId).types[t]=[];}});
    }}
  }});
  logRows.filter(r=>selected.has(r.nmId)).forEach(r=>{{
    if(!byArticle.has(r.nmId)){{
      byArticle.set(r.nmId,{{article:r.article,name:'',types:{{}}}});
      COMMENT_TYPES.forEach(t=>{{byArticle.get(r.nmId).types[t]=[];}});
    }}
    const t=r.comment_type;
    if(byArticle.get(r.nmId).types[t])byArticle.get(r.nmId).types[t].push(r);
  }});
  body.innerHTML=[...byArticle.values()].map(g=>{{
    const colsHtml=COMMENT_TYPES.map(t=>{{
      const entries=(g.types[t]||[]).slice().sort((a,b)=>(b.created_at||'').localeCompare(a.created_at||''));
      const hdr=`<div class="sh-col-hdr">${{TYPE_LABELS[t]}}</div>`;
      const content=entries.length
        ?entries.map(e=>{{const _d=(e.created_at||'').slice(0,10).split('-');const _df=_d.length===3?_d[2]+'.'+_d[1]:(e.created_at||'').slice(0,10);const _ac=_ageCls(e.created_at);return`<div class="sh-entry${{_ac?' '+_ac:''}}"><div class="sh-entry-date">${{esc(_df)}}</div><div class="sh-entry-text">${{esc(e.comment_text)}}</div></div>`;}}).join('')
        :`<div class="sh-col-empty">—</div>`;
      return `<div class="sh-col">${{hdr}}${{content}}</div>`;
    }}).join('');
    return `<div class="sh-article"><div class="sh-art-title"><span class="art-tag">${{esc(g.article)}}</span><span class="art-name">${{esc(g.name)}}</span></div><div class="sh-cols">${{colsHtml}}</div></div>`;
  }}).join('');
}}

function ensureCalendarSelection(){{
  if(calendarInitialized)return;
  const first=logRows.find(r=>(r.created_at||'').slice(0,10));
  const d=parseDateKey(first?(first.created_at||'').slice(0,10):dateKey(new Date()));
  if(d){{
    selectedCalendarDate=dateKey(d);
    calendarMonthDate=new Date(d.getFullYear(),d.getMonth(),1);
  }}
  calendarInitialized=true;
}}

function renderCommentCalendar(){{
  ensureCalendarSelection();
  const grid=document.getElementById('comment-calendar-grid');
  const title=document.getElementById('comment-month-title');
  const list=document.getElementById('comment-day-list');
  if(!grid||!title||!list)return;
  const monthStart=new Date(calendarMonthDate.getFullYear(),calendarMonthDate.getMonth(),1);
  const year=monthStart.getFullYear();
  const month=monthStart.getMonth();
  title.textContent=new Intl.DateTimeFormat('ru-RU',{{month:'long',year:'numeric'}}).format(monthStart);
  const counts=new Map();
  logRows.forEach(r=>{{
    const key=(r.created_at||'').slice(0,10);
    if(key)counts.set(key,(counts.get(key)||0)+1);
  }});
  const todayKey=dateKey(new Date());
  const firstOffset=(monthStart.getDay()+6)%7;
  const daysInMonth=new Date(year,month+1,0).getDate();
  const cells=[];
  for(let i=0;i<firstOffset;i++)cells.push('<button class="month-day is-empty" type="button" tabindex="-1"></button>');
  for(let day=1;day<=daysInMonth;day++){{
    const key=dateKey(new Date(year,month,day));
    const cnt=counts.get(key)||0;
    const cls=['month-day'];
    if(key===todayKey)cls.push('is-today');
    if(key===selectedCalendarDate)cls.push('is-selected');
    cells.push(`<button class="${{cls.join(' ')}}" type="button" data-date="${{key}}">
      <span class="month-day-num">${{day}}</span>
      ${{cnt?`<span class="month-day-count">${{cnt}}</span>`:''}}
    </button>`);
  }}
  grid.innerHTML=cells.join('');
  grid.querySelectorAll('.month-day[data-date]').forEach(btn=>{{
    btn.addEventListener('click',()=>{{
      selectedCalendarDate=btn.dataset.date;
      renderCommentCalendar();
    }});
  }});
  renderCommentDayList();
}}

function renderCommentDayList(){{
  const title=document.getElementById('comment-day-title');
  const count=document.getElementById('comment-day-count');
  const list=document.getElementById('comment-day-list');
  if(!title||!count||!list)return;
  const rows=logRows.filter(r=>(r.created_at||'').slice(0,10)===selectedCalendarDate);
  title.textContent='Комментарии за '+fmtDateRu(selectedCalendarDate);
  count.textContent=rows.length?rows.length+' шт.':'нет записей';
  list.innerHTML=rows.map(r=>{{
    const badge=`<span class="type-badge ${{TYPE_CSS[r.comment_type]||''}}">${{TYPE_LABELS[r.comment_type]||r.comment_type}}</span>`;
    return `<div class="month-comment-item" title="${{esc(r.article+' - '+r.comment_text)}}">
      <span class="month-comment-article">${{esc(r.article)}}</span>
      ${{badge}}
      <span class="month-comment-sep">-</span>
      <span class="month-comment-text">${{esc(r.comment_text)}}</span>
    </div>`;
  }}).join('')||'<div class="empty">В этот день комментариев нет.</div>';
}}

function renderLog(){{
  const search=(document.getElementById('log-search').value||'').toLowerCase();
  const tf=document.getElementById('log-type').value;
  const filtered=logRows.filter(r=>{{
    if(tf&&r.comment_type!==tf)return false;
    if(search&&!(r.article+' '+r.comment_text).toLowerCase().includes(search))return false;
    return true;
  }});
  const tbody=document.getElementById('log-body');
  tbody.innerHTML=filtered.map(r=>{{
    const dt=(r.created_at||'').slice(0,16);
    const badge=`<span class="type-badge ${{TYPE_CSS[r.comment_type]||''}}">${{TYPE_LABELS[r.comment_type]||r.comment_type}}</span>`;
    return `<tr><td style="white-space:nowrap">${{esc(dt)}}</td><td>${{esc(r.article)}}</td><td>${{badge}}</td><td>${{esc(r.comment_text)}}</td></tr>`;
  }}).join('')||'<tr><td colspan="4" class="empty">Нет записей.</td></tr>';
}}

document.getElementById('save-btn').addEventListener('click',async()=>{{
  if(!selected.size)return;
  const btn=document.getElementById('save-btn');const status=document.getElementById('save-status');
  btn.disabled=true;status.textContent='Сохранение...';
  const nm_ids=[...selected].map(Number);
  const saves=FIELDS.filter(f=>!document.getElementById(f.id).disabled&&document.getElementById(f.id).value.trim())
    .map(f=>{{return{{type:f.type,text:document.getElementById(f.id).value.trim()}};}} );
  if(!saves.length){{status.textContent='Введите хотя бы один комментарий.';btn.disabled=false;return;}}
  try{{
    for(const s of saves){{
      const resp=await fetch('/api/analytics/comments',{{
        method:'POST',headers:{{'Content-Type':'application/json'}},
        body:JSON.stringify({{nm_ids,comment_type:s.type,comment_text:s.text}})
      }});
      if(!resp.ok)throw new Error(await resp.text());
    }}
    FIELDS.forEach(f=>{{document.getElementById(f.id).value='';}});
    selected.clear();lastClickedNm=null;
    status.textContent='Сохранено!';
    await Promise.all([loadData(),loadLog()]);updatePanel();
    setTimeout(()=>{{status.textContent='';}},3000);
  }}catch(e){{status.textContent='Ошибка: '+e;}}
  btn.disabled=false;
}});

['f-text','f-subject','f-strategy','f-revcat','f-drr-op','f-drr-val','f-spd-op','f-spd-val','f-trn-op','f-trn-val'].forEach(id=>{{
  const el=document.getElementById(id);if(el)el.addEventListener(el.tagName==='SELECT'?'change':'input',renderTable);
}});
document.getElementById('btn-reset').addEventListener('click',()=>{{
  ['f-text'].forEach(id=>document.getElementById(id).value='');
  ['f-subject','f-revcat','f-drr-op','f-spd-op','f-trn-op'].forEach(id=>document.getElementById(id).value=''); document.querySelectorAll('#f-strategy option').forEach(o=>{{o.selected=false;}});
  ['f-drr-val','f-spd-val','f-trn-val'].forEach(id=>document.getElementById(id).value='');
  renderTable();
}});
document.getElementById('log-search').addEventListener('input',renderLog);
document.getElementById('log-type').addEventListener('change',renderLog);
document.getElementById('comment-month-prev').addEventListener('click',()=>{{
  calendarMonthDate=new Date(calendarMonthDate.getFullYear(),calendarMonthDate.getMonth()-1,1);
  renderCommentCalendar();
}});
document.getElementById('comment-month-next').addEventListener('click',()=>{{
  calendarMonthDate=new Date(calendarMonthDate.getFullYear(),calendarMonthDate.getMonth()+1,1);
  renderCommentCalendar();
}});

loadData();
loadLog();
</script>
</body>
</html>"""

OZON_COMMENTS_HTML = """\
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Комментарии OZON</title>
  <style>
    :root{{--bg:#f8fafc;--ink:#111827;--muted:#6b7280;--accent:#ea580c;--line:#e5e7eb;}}
    *{{box-sizing:border-box;}}
    body{{margin:0;font-family:"Segoe UI","Trebuchet MS",sans-serif;color:var(--ink);
         background:linear-gradient(180deg,#fff7ed,transparent 240px),var(--bg);}}
    .wrap{{width:calc(100% - 24px);max-width:3000px;margin:0 auto;padding:8px 0 28px;}}
    .top{{display:flex;align-items:center;gap:8px;margin-bottom:6px;}}
    .top h1{{font-size:.82rem;font-weight:700;color:var(--muted);white-space:nowrap;margin:0;flex-shrink:0;}}
    .report-nav{{display:flex;align-items:center;flex-wrap:nowrap;overflow-x:auto;gap:2px;flex:1;
                 scrollbar-width:none;min-width:0;}}
    .report-nav::-webkit-scrollbar{{display:none;}}
    .filters{{background:#fff;border:1px solid var(--line);border-radius:8px;
              padding:8px 12px;margin-bottom:6px;display:flex;flex-wrap:wrap;gap:6px;align-items:flex-end;}}
    .fgroup{{display:flex;flex-direction:column;gap:2px;}}
    .fgroup label{{font-size:10px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.04em;}}
    .fgroup select,.fgroup input{{height:28px;padding:0 7px;border:1px solid #cbd5e1;border-radius:6px;
                                   font-size:11px;background:#f8fafc;color:var(--ink);min-width:0;}}
    .fgroup select:focus,.fgroup input:focus{{outline:none;border-color:var(--accent);background:#fff;}}
    .fgroup-num{{display:flex;flex-direction:column;gap:2px;}}
    .fgroup-num label{{font-size:10px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.04em;}}
    .fgroup-num .num-row{{display:flex;gap:2px;}}
    .fgroup-num select{{height:28px;width:44px;padding:0 2px;border:1px solid #cbd5e1;border-radius:6px 0 0 6px;
                        font-size:12px;background:#f8fafc;border-right:0;}}
    .fgroup-num input{{height:28px;width:72px;padding:0 5px;border:1px solid #cbd5e1;border-radius:0 6px 6px 0;
                       font-size:11px;background:#f8fafc;}}
    .fgroup-num select:focus,.fgroup-num input:focus{{outline:none;border-color:var(--accent);background:#fff;z-index:1;position:relative;}}
    .btn-reset{{height:28px;padding:0 10px;border:1px solid #cbd5e1;border-radius:6px;background:#fff;
                font-size:11px;cursor:pointer;color:#64748b;align-self:flex-end;margin-left:2px;}}
    .btn-reset:hover{{background:#f1f5f9;}}
    #comment-panel{{display:none;background:#fff;border:1px solid #fed7aa;border-radius:8px;
                    padding:10px 14px;margin-bottom:6px;}}
    #comment-panel h3{{margin:0 0 8px;font-size:.78rem;font-weight:800;color:#0f172a;}}
    .comment-fields{{display:flex;flex-wrap:wrap;gap:6px;}}
    .comment-field{{display:flex;align-items:flex-start;gap:5px;flex:1;min-width:160px;}}
    .comment-field label{{font-size:10px;font-weight:700;color:#64748b;text-transform:uppercase;
                          letter-spacing:.04em;width:52px;flex-shrink:0;padding-top:6px;}}
    .comment-field textarea{{flex:1;border:1px solid #cbd5e1;border-radius:6px;font-size:11px;
                              padding:5px 7px;resize:none;height:30px;background:#f8fafc;
                              color:#0f172a;font-family:inherit;}}
    .comment-field textarea:focus{{outline:none;border-color:var(--accent);background:#fff;}}
    .panel-bottom{{display:flex;align-items:center;margin-top:8px;gap:8px;}}
    #save-btn{{padding:5px 16px;border:none;border-radius:6px;background:var(--accent);
               color:#fff;font-weight:700;font-size:12px;cursor:pointer;}}
    #save-btn:disabled{{opacity:.5;cursor:default;}}
    #save-status{{font-size:11px;color:var(--muted);}}
    .meta{{font-size:.75rem;color:var(--muted);margin-bottom:4px;}}
    .tbl{{background:#fff;border:1px solid var(--line);border-radius:8px;overflow:auto;max-height:480px;}}
    table{{width:max-content;min-width:100%;border-collapse:collapse;font-size:11px;}}
    th,td{{border-bottom:1px solid #eef2f7;padding:4px 6px;text-align:right;white-space:nowrap;}}
    th{{position:sticky;top:0;z-index:1;background:#f8fafc;color:#334155;font-weight:800;cursor:pointer;user-select:none;}}
    th:hover{{background:#f1f5f9;}}
    th .sort-ic{{margin-left:3px;color:#94a3b8;font-size:.7em;}}
    .col-ck{{width:22px;min-width:22px;text-align:center;position:sticky;left:0;z-index:2;background:#fff;cursor:default;}}
    th.col-ck{{z-index:3;background:#f8fafc;cursor:default;}}
    .col-art{{text-align:left;position:sticky;left:22px;z-index:2;background:#fff;
              min-width:80px;max-width:130px;overflow:hidden;text-overflow:ellipsis;}}
    th.col-art{{z-index:3;background:#f8fafc;}}
    .col-name{{text-align:left;min-width:120px;max-width:200px;overflow:hidden;text-overflow:ellipsis;}}
    .col-comment{{text-align:left;min-width:130px;max-width:200px;overflow:hidden;text-overflow:ellipsis;}}
    input[type=checkbox]{{width:14px;height:14px;cursor:pointer;accent-color:var(--accent);}}
    .art-copy{{margin-left:3px;font-size:.75em;color:#94a3b8;background:none;border:none;
               cursor:pointer;padding:0 1px;opacity:.7;line-height:1;}}
    .art-copy:hover{{color:var(--accent);opacity:1;}}
    .art-copy.copied{{color:#16a34a;opacity:1;}}
    tbody tr:hover td{{background:#fff7ed!important;}}
    tbody tr.row-selected td{{background:#fed7aa!important;}}
    tbody tr.row-selected td.col-ck,tbody tr.row-selected td.col-art{{background:#fed7aa!important;}}
    .empty{{padding:18px;color:var(--muted);font-size:.9rem;}}
    .log-section{{margin-top:10px;background:#fff;border:1px solid var(--line);border-radius:8px;padding:10px 14px;}}
    .log-section h3{{margin:0 0 8px;font-size:.78rem;font-weight:800;}}
    .log-filter{{display:flex;gap:6px;align-items:center;margin-bottom:8px;flex-wrap:wrap;}}
    .log-filter input,.log-filter select{{height:28px;padding:0 7px;border:1px solid #cbd5e1;
                                          border-radius:6px;font-size:11px;}}
    .log-filter input{{width:180px;}}
    .log-table{{width:100%;border-collapse:collapse;font-size:11px;}}
    .log-table th,.log-table td{{border-bottom:1px solid #eef2f7;padding:4px 6px;text-align:left;}}
    .log-table th{{background:#f8fafc;font-weight:800;color:#334155;white-space:nowrap;}}
    .type-badge{{display:inline-block;padding:1px 6px;border-radius:999px;font-size:10px;font-weight:700;}}
    .type-general{{background:#dbeafe;color:#1d4ed8;}}
    .type-cpc{{background:#fce7f3;color:#9d174d;}}
    .type-shelves{{background:#d1fae5;color:#065f46;}}
    .type-unified{{background:#fef9c3;color:#854d0e;}}
    .type-search{{background:#ede9fe;color:#5b21b6;}}
    .sel-history{{margin-top:10px;border:2px solid #fed7aa;border-radius:8px;background:#fff;padding:10px 14px;}}
    .sel-history h3{{margin:0 0 10px;font-size:.78rem;font-weight:800;color:#9a3412;}}
    .sh-article{{margin-bottom:12px;}}
    .sh-art-title{{font-size:.78rem;font-weight:800;color:#0f172a;margin-bottom:4px;
                   display:flex;align-items:center;gap:6px;}}
    .sh-art-title span{{background:#fff7ed;border:1px solid #fed7aa;border-radius:6px;
                        padding:1px 8px;font-size:11px;}}
    .sh-dates{{display:flex;flex-direction:column;gap:2px;padding-left:8px;}}
    .sh-date-row{{display:flex;gap:6px;align-items:flex-start;padding:3px 0;
                  border-bottom:1px solid #f1f5f9;}}
    .sh-date{{font-size:10px;font-weight:700;color:#94a3b8;white-space:nowrap;width:70px;flex-shrink:0;padding-top:1px;}}
    .sh-entries{{display:flex;flex-wrap:wrap;gap:4px;flex:1;}}
    .sh-entry{{display:flex;align-items:flex-start;gap:4px;background:#f8fafc;
               border-radius:5px;padding:2px 6px;max-width:400px;}}
    .sh-entry .type-badge{{flex-shrink:0;margin-top:1px;}}
    .sh-entry-text{{font-size:11px;color:#0f172a;line-height:1.4;word-break:break-word;}}
  </style>
</head>
<body>
<div class="wrap">
  <div class="top">
    <h1>Комментарии OZON</h1>
    <div class="report-nav">{report_nav}</div>
  </div>

  <div class="filters" id="filters">
    <div class="fgroup">
      <label>Поиск</label>
      <input id="f-text" type="text" placeholder="Артикул или название" style="width:160px">
    </div>
    <div class="fgroup-num">
      <label>ДРР %</label>
      <div class="num-row">
        <select id="f-drr-op"><option value="">=</option><option value="gt">&gt;</option><option value="lt">&lt;</option></select>
        <input id="f-drr-val" type="number" step="0.1" placeholder="—">
      </div>
    </div>
    <div class="fgroup-num">
      <label>ДРР руб</label>
      <div class="num-row">
        <select id="f-drrr-op"><option value="">=</option><option value="gt">&gt;</option><option value="lt">&lt;</option></select>
        <input id="f-drrr-val" type="number" placeholder="—">
      </div>
    </div>
    <div class="fgroup-num">
      <label>Скорость</label>
      <div class="num-row">
        <select id="f-spd-op"><option value="">=</option><option value="gt">&gt;</option><option value="lt">&lt;</option></select>
        <input id="f-spd-val" type="number" step="0.1" placeholder="—">
      </div>
    </div>
    <div class="fgroup-num">
      <label>Оборачив.</label>
      <div class="num-row">
        <select id="f-trn-op"><option value="">=</option><option value="gt">&gt;</option><option value="lt">&lt;</option></select>
        <input id="f-trn-val" type="number" placeholder="—">
      </div>
    </div>
    <button class="btn-reset" id="btn-reset">Сброс</button>
  </div>

  <div id="comment-panel">
    <h3>Добавить комментарий &mdash; выбрано: <span id="sel-count">0</span> SKU</h3>
    <div class="comment-fields">
      <div class="comment-field"><label>Общий</label>
        <textarea id="txt-general" placeholder="Изменили фото, вошли в акцию..."></textarea></div>
      <div class="comment-field"><label>CPC</label>
        <textarea id="txt-cpc" placeholder="Ставка CPC..."></textarea></div>
      <div class="comment-field"><label>Полки</label>
        <textarea id="txt-shelves" placeholder="Полочная реклама..."></textarea></div>
      <div class="comment-field"><label>Единая</label>
        <textarea id="txt-unified" placeholder="Единая ставка..."></textarea></div>
      <div class="comment-field"><label>Поиск</label>
        <textarea id="txt-search" placeholder="Поисковая реклама..."></textarea></div>
    </div>
    <div class="panel-bottom">
      <button id="save-btn">Сохранить</button>
      <span id="save-status"></span>
    </div>
  </div>

  <div class="meta" id="meta">Загрузка...</div>
  <div class="tbl" id="tbl-wrap"></div>

  <div class="sel-history" id="sel-history" style="display:none">
    <h3>История выбранных SKU</h3>
    <div id="sel-history-body"></div>
  </div>

  <div class="log-section">
    <h3>История комментариев OZON</h3>
    <div class="log-filter">
      <input id="log-search" type="text" placeholder="Артикул или текст...">
      <select id="log-type">
        <option value="">Все типы</option>
        <option value="general">Общий</option>
        <option value="cpc">CPC</option>
        <option value="shelves">Полки</option>
        <option value="unified">Единая</option>
        <option value="search">Поиск</option>
      </select>
    </div>
    <table class="log-table">
      <thead><tr><th>Дата</th><th>Артикул</th><th>Тип</th><th>Комментарий</th></tr></thead>
      <tbody id="log-body"></tbody>
    </table>
  </div>
</div>
<script>
const TYPE_LABELS = {{general:'Общий',cpc:'CPC',shelves:'Полки',unified:'Единая',search:'Поиск'}};
const TYPE_CSS    = {{general:'type-general',cpc:'type-cpc',shelves:'type-shelves',
                     unified:'type-unified',search:'type-search'}};
const FIELDS = [
  {{id:'txt-general',type:'general'}},
  {{id:'txt-cpc',    type:'cpc'}},
  {{id:'txt-shelves',type:'shelves'}},
  {{id:'txt-unified',type:'unified'}},
  {{id:'txt-search', type:'search'}},
];
const COLS = [
  {{k:'offer_id',    label:'Артикул',        num:false}},
  {{k:'item_name',   label:'Название',       num:false}},
  {{k:'drr_pct',     label:'ДРР% (2д)',      num:true}},
  {{k:'drr_rub',     label:'ДРР руб (2д)',   num:true}},
  {{k:'speed',       label:'Скорость зак/д', num:true}},
  {{k:'turnover',    label:'Оборачив. дн',   num:true}},
  {{k:'last_general',label:'Послед. Общий',  num:false}},
  {{k:'last_cpc',    label:'Послед. CPC',    num:false}},
  {{k:'last_shelves',label:'Послед. Полки',  num:false}},
  {{k:'last_unified',label:'Послед. Единая', num:false}},
  {{k:'last_search', label:'Послед. Поиск',  num:false}},
];
let allRows = [], logRows = [], selected = new Set();
let sortKey = 'offer_id', sortDir = 1;

function esc(s){{return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}}
function numVal(s){{const v=parseFloat(s);return isNaN(v)?null:v;}}
function cmpNum(a,b){{if(a==null&&b==null)return 0;if(a==null)return 1;if(b==null)return -1;return a-b;}}

function numMatch(val, opId, inpId){{
  const op = document.getElementById(opId).value;
  const threshold = numVal(document.getElementById(inpId).value);
  if(threshold===null) return true;
  if(val==null) return false;
  if(op===''||op==='eq') return Math.abs(val-threshold)<0.001;
  if(op==='gt') return val>threshold;
  if(op==='lt') return val<threshold;
  return true;
}}

function getFilteredRows(){{
  const txt = (document.getElementById('f-text').value||'').toLowerCase().trim();
  return allRows.filter(r=>{{
    if(txt && !(r.offer_id+' '+r.item_name+' '+r.ozon_sku).toLowerCase().includes(txt)) return false;
    if(!numMatch(r.drr_pct,  'f-drr-op',  'f-drr-val'))  return false;
    if(!numMatch(r.drr_rub,  'f-drrr-op', 'f-drrr-val')) return false;
    if(!numMatch(r.speed,    'f-spd-op',  'f-spd-val'))  return false;
    if(!numMatch(r.turnover, 'f-trn-op',  'f-trn-val'))  return false;
    return true;
  }});
}}

function getSortedRows(rows){{
  return [...rows].sort((a,b)=>{{
    const col = COLS.find(c=>c.k===sortKey);
    if(!col) return 0;
    let av=a[sortKey], bv=b[sortKey];
    if(sortKey==='drr_pct'||sortKey==='drr_rub'){{ av=av??0; bv=bv??0; }}
    if(col.num) return sortDir*cmpNum(av,bv);
    return sortDir*((av||'').localeCompare(bv||'','ru'));
  }});
}}

async function loadData(){{
  document.getElementById('meta').textContent='Загрузка...';
  try{{
    const r=await fetch('/api/ozon/comments-data');
    const d=await r.json();
    allRows=d.rows||[];
    renderTable();
  }}catch(e){{document.getElementById('meta').textContent='Ошибка: '+e;}}
}}
async function loadLog(){{
  try{{
    const r=await fetch('/api/ozon/comment-log?limit=1000');
    const d=await r.json();
    logRows=d.rows||[];
    renderLog();
    renderSelHistory();
  }}catch(e){{}}
}}

function renderTable(){{
  const rows=getSortedRows(getFilteredRows());
  const el=document.getElementById('tbl-wrap');
  document.getElementById('meta').textContent='SKU: '+rows.length+' из '+allRows.length;
  if(!rows.length){{el.innerHTML='<div class="empty">Нет данных.</div>';return;}}

  const thHtml = COLS.map(c=>{{
    const active = c.k===sortKey;
    const ic = active ? (sortDir>0?'↑':'↓') : '↕';
    const cls = c.k==='offer_id'?' class="col-art"' : c.k==='item_name'?' class="col-name"' :
                c.k.startsWith('last_')?' class="col-comment"':'';
    return `<th data-key="${{c.k}}"${{cls}}>${{c.label}}<span class="sort-ic">${{ic}}</span></th>`;
  }}).join('');

  const trs=rows.map(r=>{{
    const sel=selected.has(r.ozon_sku)?' row-selected':'';
    const drr  = r.drr_pct!=null ? r.drr_pct.toFixed(1)+'%' : '—';
    const rub  = r.drr_rub>0 ? Math.round(r.drr_rub).toLocaleString('ru') : (r.drr_rub===0?'0':'—');
    const spd  = r.speed ? r.speed.toFixed(1) : '—';
    const trn  = r.turnover!=null ? Math.round(r.turnover)+'д' : '—';
    const artLinks = `<button class="art-copy" title="Скопировать" data-copy="${{esc(r.offer_id||r.ozon_sku)}}" data-skip-select="1">⎘</button>`;
    const displayArt = esc(r.offer_id||r.ozon_sku);
    return `<tr class="data-row${{sel}}" data-sku="${{r.ozon_sku}}">
      <td class="col-ck"><input type="checkbox" class="row-ck" data-sku="${{r.ozon_sku}}" ${{selected.has(r.ozon_sku)?'checked':''}}></td>
      <td class="col-art" title="${{displayArt}}">${{displayArt}}${{artLinks}}</td>
      <td class="col-name" title="${{esc(r.item_name)}}">${{esc(r.item_name)}}</td>
      <td>${{drr}}</td><td style="text-align:right">${{rub}}</td>
      <td>${{spd}}</td><td>${{trn}}</td>
      <td class="col-comment" title="${{esc(r.last_general)}}">${{esc(r.last_general)}}</td>
      <td class="col-comment" title="${{esc(r.last_cpc)}}">${{esc(r.last_cpc)}}</td>
      <td class="col-comment" title="${{esc(r.last_shelves)}}">${{esc(r.last_shelves)}}</td>
      <td class="col-comment" title="${{esc(r.last_unified)}}">${{esc(r.last_unified)}}</td>
      <td class="col-comment" title="${{esc(r.last_search)}}">${{esc(r.last_search)}}</td>
    </tr>`;
  }}).join('');

  const allVisibleSelected = rows.length>0 && rows.every(r=>selected.has(r.ozon_sku));
  const ckAll=`<th class="col-ck"><input type="checkbox" id="ck-all" ${{allVisibleSelected?'checked':''}}></th>`;
  el.innerHTML=`<table><thead><tr>${{ckAll}}${{thHtml}}</tr></thead><tbody>${{trs}}</tbody></table>`;

  el.querySelectorAll('th[data-key]').forEach(th=>{{
    th.addEventListener('click',()=>{{
      const k=th.dataset.key;
      if(sortKey===k) sortDir*=-1; else {{sortKey=k;sortDir=1;}}
      renderTable();
    }});
  }});
  document.getElementById('ck-all').addEventListener('change',e=>{{
    rows.forEach(r=> e.target.checked?selected.add(r.ozon_sku):selected.delete(r.ozon_sku));
    renderTable();updatePanel();
  }});
  el.querySelectorAll('.row-ck').forEach(ck=>{{
    ck.addEventListener('change',e=>{{
      const sk=e.target.dataset.sku;
      e.target.checked?selected.add(sk):selected.delete(sk);
      e.target.closest('tr').classList.toggle('row-selected',e.target.checked);
      updatePanel();
    }});
  }});
  el.querySelectorAll('.art-copy').forEach(btn=>{{
    btn.addEventListener('click',e=>{{
      e.stopPropagation();
      navigator.clipboard.writeText(btn.dataset.copy).then(()=>{{
        btn.classList.add('copied');btn.textContent='✓';
        setTimeout(()=>{{btn.classList.remove('copied');btn.textContent='⎘';}},1500);
      }});
    }});
  }});
}}

function updatePanel(){{
  const p=document.getElementById('comment-panel');
  document.getElementById('sel-count').textContent=selected.size;
  p.style.display=selected.size>0?'block':'none';
  renderSelHistory();
}}

function renderSelHistory(){{
  const el=document.getElementById('sel-history');
  const body=document.getElementById('sel-history-body');
  if(!selected.size){{el.style.display='none';return;}}
  const selRows=logRows.filter(r=>selected.has(r.ozon_sku));
  if(!selRows.length){{el.style.display='none';return;}}
  el.style.display='block';
  const byArticle=new Map();
  [...allRows.filter(r=>selected.has(r.ozon_sku))].forEach(r=>{{
    if(!byArticle.has(r.ozon_sku)) byArticle.set(r.ozon_sku,{{offer_id:r.offer_id||r.ozon_sku,dates:new Map()}});
  }});
  selRows.forEach(r=>{{
    if(!byArticle.has(r.ozon_sku)) byArticle.set(r.ozon_sku,{{offer_id:r.offer_id||r.ozon_sku,dates:new Map()}});
    const dateKey=(r.created_at||'').slice(0,10);
    const g=byArticle.get(r.ozon_sku);
    if(!g.dates.has(dateKey)) g.dates.set(dateKey,[]);
    g.dates.get(dateKey).push(r);
  }});
  body.innerHTML=[...byArticle.values()].map(g=>{{
    if(!g.dates.size) return '';
    const sortedDates=[...g.dates.keys()].sort((a,b)=>b.localeCompare(a));
    const dateRows=sortedDates.map(d=>{{
      const entries=g.dates.get(d);
      const entryHtml=entries.map(e=>
        `<div class="sh-entry">
          <span class="type-badge ${{TYPE_CSS[e.comment_type]||''}}">${{TYPE_LABELS[e.comment_type]||e.comment_type}}</span>
          <span class="sh-entry-text">${{esc(e.comment_text)}}</span>
        </div>`
      ).join('');
      return `<div class="sh-date-row"><div class="sh-date">${{d}}</div><div class="sh-entries">${{entryHtml}}</div></div>`;
    }}).join('');
    return `<div class="sh-article"><div class="sh-art-title"><span>${{esc(g.offer_id)}}</span></div><div class="sh-dates">${{dateRows}}</div></div>`;
  }}).join('');
}}

function renderLog(){{
  const search=(document.getElementById('log-search').value||'').toLowerCase();
  const tf=document.getElementById('log-type').value;
  const filtered=logRows.filter(r=>{{
    if(tf&&r.comment_type!==tf)return false;
    if(search&&!(r.offer_id+' '+r.comment_text).toLowerCase().includes(search))return false;
    return true;
  }});
  const tbody=document.getElementById('log-body');
  tbody.innerHTML=filtered.map(r=>{{
    const dt=(r.created_at||'').slice(0,16);
    const badge=`<span class="type-badge ${{TYPE_CSS[r.comment_type]||''}}">${{TYPE_LABELS[r.comment_type]||r.comment_type}}</span>`;
    return `<tr><td style="white-space:nowrap">${{esc(dt)}}</td><td>${{esc(r.offer_id)}}</td><td>${{badge}}</td><td>${{esc(r.comment_text)}}</td></tr>`;
  }}).join('')||'<tr><td colspan="4" class="empty">Нет записей.</td></tr>';
}}

document.getElementById('save-btn').addEventListener('click',async()=>{{
  if(!selected.size)return;
  const btn=document.getElementById('save-btn');
  const status=document.getElementById('save-status');
  btn.disabled=true;status.textContent='Сохранение...';
  const skus=[...selected];
  const saves=FIELDS.filter(f=>document.getElementById(f.id).value.trim())
    .map(f=>{{return{{type:f.type,text:document.getElementById(f.id).value.trim()}};}} );
  if(!saves.length){{status.textContent='Введите хотя бы один комментарий.';btn.disabled=false;return;}}
  try{{
    for(const s of saves){{
      const resp=await fetch('/api/ozon/comments',{{
        method:'POST',
        headers:{{'Content-Type':'application/json'}},
        body:JSON.stringify({{skus,comment_type:s.type,comment_text:s.text}})
      }});
      if(!resp.ok)throw new Error(await resp.text());
    }}
    FIELDS.forEach(f=>{{document.getElementById(f.id).value='';}});
    selected.clear();
    status.textContent='Сохранено!';
    await Promise.all([loadData(),loadLog()]);
    updatePanel();
    setTimeout(()=>{{status.textContent='';}},3000);
  }}catch(e){{status.textContent='Ошибка: '+e;}}
  btn.disabled=false;
}});

['f-text','f-drr-op','f-drr-val','f-drrr-op','f-drrr-val',
 'f-spd-op','f-spd-val','f-trn-op','f-trn-val'].forEach(id=>{{
  const el=document.getElementById(id);
  if(el) el.addEventListener(el.tagName==='SELECT'?'change':'input',renderTable);
}});
document.getElementById('btn-reset').addEventListener('click',()=>{{
  document.getElementById('f-text').value='';
  ['f-drr-op','f-drrr-op','f-spd-op','f-trn-op'].forEach(id=>document.getElementById(id).value='');
  ['f-drr-val','f-drrr-val','f-spd-val','f-trn-val'].forEach(id=>document.getElementById(id).value='');
  renderTable();
}});
document.getElementById('log-search').addEventListener('input',renderLog);
document.getElementById('log-type').addEventListener('change',renderLog);

loadData();
loadLog();
</script>
</body>
</html>"""

class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):  # silence default access log
        pass

    def _get_cookie(self, name: str) -> str:
        raw = self.headers.get("Cookie", "")
        for part in raw.split(";"):
            part = part.strip()
            if part.startswith(name + "="):
                return part[len(name) + 1:]
        return ""

    def _set_cabinet_cookie(self, token: str) -> None:
        self.send_header(
            "Set-Cookie",
            f"wb_session={token}; Path=/; HttpOnly; SameSite=Lax; Max-Age=86400",
        )

    def _clear_cabinet_cookie(self) -> None:
        self.send_header(
            "Set-Cookie",
            "wb_session=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0",
        )

    def _resolve_cabinet(self) -> dict | None:
        token = self._get_cookie("wb_session")
        cabinet = get_platform().get_session_cabinet(token)
        _set_request_cabinet(cabinet)
        return cabinet

    def _send_html(self, body: bytes) -> None:
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_json(self, payload: object, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_redirect(self, location: str) -> None:
        self.send_response(302)
        self.send_header("Location", location)
        self.send_header("Content-Length", "0")
        self.end_headers()

    def _send_unauthorized(self) -> None:
        body = "Требуется пароль".encode("utf-8")
        self.send_response(401)
        self.send_header("WWW-Authenticate", 'Basic realm="WB Analytics"')
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _is_authorized(self) -> bool:
        return True

    # ── Cabinet / admin page handlers ───────────────────────────────────────

    def _handle_select_cabinet(self) -> None:
        cabinets = get_platform().list_cabinets()
        mp_color = {"wb": "#3b82f6", "ozon": "#f97316", "both": "#8b5cf6"}
        items_html = ""
        for c in cabinets:
            color = mp_color.get(c.get("marketplace", "wb"), "#3b82f6")
            items_html += (
                f'<button onclick="selectCab(\'{escape(c["cabinet_id"])}\')" '
                f'style="width:100%;padding:14px 18px;margin-bottom:8px;background:#1e293b;'
                f'color:#e2e8f0;border:1.5px solid #334155;border-radius:10px;cursor:pointer;'
                f'text-align:left;font-size:1rem;font-weight:600;display:flex;align-items:center;gap:10px;">'
                f'<span style="width:10px;height:10px;border-radius:50%;background:{color};flex-shrink:0;"></span>'
                f'{escape(c["name"])}</button>'
            )
        body = f"""<!DOCTYPE html>
<html lang="ru">
<head><meta charset="utf-8"><title>Выбор кабинета</title>
<style>*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
background:linear-gradient(135deg,#0a0f1e,#1a1145,#0c3a5e);min-height:100vh;
display:flex;align-items:center;justify-content:center;padding:20px;}}
.card{{background:#0f172a;border:1px solid #1e293b;border-radius:16px;padding:32px;
width:100%;max-width:380px;box-shadow:0 20px 60px rgba(0,0,0,.5);}}
h1{{color:#f1f5f9;font-size:1.2rem;font-weight:700;margin-bottom:20px;}}
.pin-wrap{{display:flex;gap:8px;justify-content:center;margin:12px 0;}}
.pin-wrap input{{width:52px;height:60px;text-align:center;font-size:1.6rem;font-weight:700;
background:#1e293b;color:#f1f5f9;border:2px solid #334155;border-radius:8px;outline:none;}}
.err{{color:#f87171;font-size:.82rem;text-align:center;min-height:1.2em;margin-top:4px;}}
</style></head>
<body>
<div class="card">
  <h1>Выберите кабинет</h1>
  <div id="cab-list">{items_html}</div>
  <div id="pin-form" style="display:none;">
    <div style="color:#94a3b8;font-size:.85rem;margin-bottom:8px;">Введите PIN</div>
    <div class="pin-wrap">
      <input id="p0" type="password" maxlength="1" inputmode="numeric">
      <input id="p1" type="password" maxlength="1" inputmode="numeric">
      <input id="p2" type="password" maxlength="1" inputmode="numeric">
    </div>
    <div id="err" class="err"></div>
    <button onclick="document.getElementById('pin-form').style.display='none';selId=null"
      style="width:100%;padding:8px;background:#334155;color:#94a3b8;border:none;border-radius:8px;cursor:pointer;margin-top:8px;">Назад</button>
  </div>
</div>
<script>
var selId=null;
function selectCab(id){{selId=id;document.getElementById('pin-form').style.display='block';
['p0','p1','p2'].forEach(function(x){{document.getElementById(x).value='';}});
document.getElementById('p0').focus();document.getElementById('err').textContent='';}}
['p0','p1','p2'].forEach(function(id,i){{
  document.getElementById(id).addEventListener('input',function(){{
    this.value=this.value.replace(/\\D/g,'');
    if(this.value.length===1){{
      if(i<2)document.getElementById('p'+(i+1)).focus();
      else{{var pin=document.getElementById('p0').value+document.getElementById('p1').value+document.getElementById('p2').value;
        fetch('/api/switch-cabinet',{{method:'POST',headers:{{'Content-Type':'application/json'}},
          body:JSON.stringify({{cabinet_id:selId,pin:pin}})}}).then(r=>r.json()).then(function(d){{
            if(d.ok){{window.location.href='/';}}
            else{{document.getElementById('err').textContent=d.error||'Неверный PIN';
              ['p0','p1','p2'].forEach(function(x){{document.getElementById(x).value='';}});
              document.getElementById('p0').focus();}}
          }});
      }}
    }}
  }});
}});
</script>
</body></html>"""
        self._send_html(body.encode("utf-8"))

    def _handle_admin_get(self, params: dict) -> None:
        # Step 1: PIN gate
        submitted_pin = (params.get("pin") or [""])[0].strip()
        admin_pin = ADMIN_PIN
        if not admin_pin:
            self._send_html(b"<h1>ADMIN_PIN not configured</h1>")
            return
        pin_ok = submitted_pin == admin_pin
        if not pin_ok:
            body = (
                '<!DOCTYPE html><html lang="ru"><head><meta charset="utf-8"><title>Admin</title>'
                '<style>body{font-family:sans-serif;background:#0f172a;color:#e2e8f0;display:flex;align-items:center;justify-content:center;min-height:100vh;}'
                '.card{background:#1e293b;padding:32px;border-radius:12px;width:300px;}'
                'h2{margin-bottom:16px;}input{width:100%;padding:10px;background:#0f172a;color:#e2e8f0;border:1.5px solid #334155;border-radius:6px;font-size:1rem;}'
                'button{width:100%;margin-top:12px;padding:10px;background:#0f766e;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:1rem;}'
                '</style></head><body><div class="card"><h2>Admin PIN</h2>'
                '<form method="GET" action="/_admin"><input type="password" name="pin" placeholder="PIN" autofocus>'
                '<button type="submit">Войти</button></form></div></body></html>'
            )
            self._send_html(body.encode("utf-8"))
            return

        cabinets = get_platform().list_cabinets()
        rows = ""
        for c in cabinets:
            mp = escape(c.get("marketplace", "wb"))
            rows += (
                f'<tr><td>{escape(c["cabinet_id"])}</td><td>{escape(c["name"])}</td>'
                f'<td>{mp}</td>'
                f'<td><a href="/_admin/edit/{escape(c["cabinet_id"])}?pin={escape(submitted_pin)}">Изменить</a>'
                f' | <form method="POST" action="/_admin/delete/{escape(c["cabinet_id"])}" style="display:inline">'
                f'<input type="hidden" name="admin_pin" value="{escape(submitted_pin)}">'
                f'<button type="submit" onclick="return confirm(\'Удалить?\')">Удалить</button></form></td></tr>'
            )
        body = f"""<!DOCTYPE html><html lang="ru"><head><meta charset="utf-8"><title>Admin — Кабинеты</title>
<style>body{{font-family:sans-serif;background:#0f172a;color:#e2e8f0;padding:24px;}}
table{{border-collapse:collapse;width:100%;margin-bottom:24px;}}
th,td{{padding:8px 12px;border:1px solid #334155;text-align:left;}}
th{{background:#1e293b;}}a{{color:#38bdf8;}}
input,select{{padding:6px 8px;background:#0f172a;color:#e2e8f0;border:1.5px solid #334155;border-radius:4px;width:100%;margin-top:4px;}}
label{{display:block;font-size:.82rem;color:#94a3b8;margin-top:10px;}}
button{{padding:8px 16px;background:#0f766e;color:#fff;border:none;border-radius:6px;cursor:pointer;margin-top:16px;}}
fieldset{{border:1px solid #334155;border-radius:8px;padding:16px;margin-bottom:24px;}}
legend{{color:#94a3b8;padding:0 8px;}}
</style></head><body>
<h1>Кабинеты</h1>
<table><tr><th>ID</th><th>Название</th><th>Маркетплейс</th><th>Действия</th></tr>{rows}</table>
<h2>Создать кабинет</h2>
<form method="POST" action="/_admin/create">
<input type="hidden" name="admin_pin" value="{escape(submitted_pin)}">
<fieldset><legend>Основное</legend>
<label>ID (латиница, без пробелов) <input name="cabinet_id" required></label>
<label>Название <input name="name" required></label>
<label>Маркетплейс
  <select name="marketplace"><option value="wb">WB</option><option value="ozon">Ozon</option><option value="both">WB + Ozon</option></select>
</label>
<label>PIN (3 цифры) <input name="pin" maxlength="3" value="321"></label>
<label>Тип фильтра артикула
  <select name="article_filter_type"><option value="nmId">nmId</option><option value="vendorCode">vendorCode</option></select>
</label>
<label><input type="checkbox" name="disable_scope_filter" value="1"> Отключить фильтр Наши/Не наши</label>
</fieldset>
<fieldset><legend>WB токены</legend>
<label>WB_API_TOKEN <input name="wb_api_token" type="password"></label>
<label>WB_FINANCE_TOKEN (если отдельный) <input name="wb_finance_token" type="password"></label>
<label>WB_ADV_TOKEN (если отдельный) <input name="wb_adv_token" type="password"></label>
</fieldset>
<fieldset><legend>Ozon токены</legend>
<label>OZON_CLIENT_ID <input name="ozon_client_id"></label>
<label>OZON_API_KEY <input name="ozon_api_key" type="password"></label>
<label>OZON_PERFORMANCE_CLIENT_ID <input name="ozon_performance_client_id"></label>
<label>OZON_PERFORMANCE_CLIENT_SECRET <input name="ozon_performance_client_secret" type="password"></label>
</fieldset>
<button type="submit">Создать</button>
</form>
</body></html>"""
        self._send_html(body.encode("utf-8"))

    def _handle_admin_create(self) -> None:
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length).decode("utf-8", errors="replace")
        from urllib.parse import parse_qs as _pqs
        form = {k: v[0] for k, v in _pqs(raw).items() if v}
        admin_pin = form.get("admin_pin", "")
        if not ADMIN_PIN or admin_pin != ADMIN_PIN:
            self._send_json({"error": "unauthorized"}, status=403)
            return
        cabinet_id = form.get("cabinet_id", "").strip()
        if not cabinet_id:
            self._send_html(b"<h1>cabinet_id required</h1>")
            return
        config = {
            "cabinet_id": cabinet_id,
            "name": form.get("name", cabinet_id),
            "marketplace": form.get("marketplace", "wb"),
            "pin": form.get("pin", "321"),
            "article_filter_type": form.get("article_filter_type", "nmId"),
            "disable_scope_filter": bool(form.get("disable_scope_filter")),
            "sku_file": "",
            "wb_api_token": form.get("wb_api_token", ""),
            "wb_finance_token": form.get("wb_finance_token", ""),
            "wb_adv_token": form.get("wb_adv_token", ""),
            "ozon_client_id": form.get("ozon_client_id", ""),
            "ozon_api_key": form.get("ozon_api_key", ""),
            "ozon_performance_client_id": form.get("ozon_performance_client_id", ""),
            "ozon_performance_client_secret": form.get("ozon_performance_client_secret", ""),
        }
        get_platform().create_cabinet(config)
        get_platform().initialize_cabinet_db(cabinet_id, ROOT / "data")
        self._send_redirect(f"/_admin?pin={quote_plus(admin_pin)}")

    def _handle_admin_edit(self, cabinet_id: str) -> None:
        params = parse_qs(urlparse(self.path).query)
        submitted_pin = (params.get("pin") or [""])[0]
        if not ADMIN_PIN or submitted_pin != ADMIN_PIN:
            self._send_redirect("/_admin")
            return
        length = int(self.headers.get("Content-Length", 0))
        if length > 0:
            raw = self.rfile.read(length).decode("utf-8", errors="replace")
            from urllib.parse import parse_qs as _pqs
            form = {k: v[0] for k, v in _pqs(raw).items() if v}
            if form.get("admin_pin", "") != ADMIN_PIN:
                self._send_json({"error": "unauthorized"}, status=403)
                return
            config = {}
            for field in ["name", "marketplace", "pin", "article_filter_type",
                          "wb_api_token", "wb_finance_token", "wb_adv_token",
                          "ozon_client_id", "ozon_api_key",
                          "ozon_performance_client_id", "ozon_performance_client_secret"]:
                if field in form and form[field]:
                    config[field] = form[field]
            if "disable_scope_filter" in form:
                config["disable_scope_filter"] = 1 if form["disable_scope_filter"] == "1" else 0
            get_platform().update_cabinet(cabinet_id, config)
            self._send_redirect(f"/_admin?pin={quote_plus(ADMIN_PIN)}")
            return
        # GET: render edit form
        cab = get_platform().get_cabinet(cabinet_id)
        if not cab:
            self._send_html(b"<h1>Cabinet not found</h1>")
            return
        body = f"""<!DOCTYPE html><html lang="ru"><head><meta charset="utf-8"><title>Edit cabinet</title>
<style>body{{font-family:sans-serif;background:#0f172a;color:#e2e8f0;padding:24px;}}
input,select{{padding:6px 8px;background:#0f172a;color:#e2e8f0;border:1.5px solid #334155;border-radius:4px;width:100%;margin-top:4px;}}
label{{display:block;font-size:.82rem;color:#94a3b8;margin-top:10px;}}
button{{padding:8px 16px;background:#0f766e;color:#fff;border:none;border-radius:6px;cursor:pointer;margin-top:16px;}}
fieldset{{border:1px solid #334155;border-radius:8px;padding:16px;margin-bottom:16px;}}
legend{{color:#94a3b8;padding:0 8px;}}</style></head>
<body><h1>Изменить: {escape(cab["name"])}</h1>
<form method="POST" action="/_admin/edit/{escape(cabinet_id)}?pin={escape(submitted_pin)}">
<input type="hidden" name="admin_pin" value="{escape(submitted_pin)}">
<fieldset><legend>Основное</legend>
<label>Название <input name="name" value="{escape(cab.get('name',''))}"></label>
<label>Маркетплейс
  <select name="marketplace">
    <option value="wb"{' selected' if cab.get('marketplace')=='wb' else ''}>WB</option>
    <option value="ozon"{' selected' if cab.get('marketplace')=='ozon' else ''}>Ozon</option>
    <option value="both"{' selected' if cab.get('marketplace')=='both' else ''}>WB + Ozon</option>
  </select></label>
<label>Новый PIN (3 цифры, оставь пустым — не меняется) <input name="pin" maxlength="3" placeholder="не менять"></label>
</fieldset>
<fieldset><legend>WB токены</legend>
<label>WB_API_TOKEN <input name="wb_api_token" type="password" placeholder="оставь пустым — не меняется"></label>
<label>WB_FINANCE_TOKEN <input name="wb_finance_token" type="password" placeholder="оставь пустым — не меняется"></label>
<label>WB_ADV_TOKEN <input name="wb_adv_token" type="password" placeholder="оставь пустым — не меняется"></label>
</fieldset>
<fieldset><legend>Ozon токены</legend>
<label>OZON_CLIENT_ID <input name="ozon_client_id" value="{escape(cab.get('ozon_client_id',''))}"></label>
<label>OZON_API_KEY <input name="ozon_api_key" type="password" placeholder="оставь пустым — не меняется"></label>
<label>OZON_PERFORMANCE_CLIENT_ID <input name="ozon_performance_client_id" value="{escape(cab.get('ozon_performance_client_id',''))}"></label>
<label>OZON_PERFORMANCE_CLIENT_SECRET <input name="ozon_performance_client_secret" type="password" placeholder="оставь пустым — не меняется"></label>
</fieldset>
<button type="submit">Сохранить</button>
<a href="/_admin?pin={escape(submitted_pin)}" style="margin-left:12px;color:#94a3b8;">← Назад</a>
</form></body></html>"""
        self._send_html(body.encode("utf-8"))

    def _handle_admin_delete(self, cabinet_id: str) -> None:
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length).decode("utf-8", errors="replace")
        from urllib.parse import parse_qs as _pqs
        form = {k: v[0] for k, v in _pqs(raw).items() if v}
        if not ADMIN_PIN or form.get("admin_pin", "") != ADMIN_PIN:
            self._send_json({"error": "unauthorized"}, status=403)
            return
        get_platform().delete_cabinet(cabinet_id)
        self._send_redirect(f"/_admin?pin={quote_plus(ADMIN_PIN)}")

    def do_GET(self):
        if not self._is_authorized():
            self._send_unauthorized()
            return

        parsed = urlparse(self.path)
        params = parse_qs(parsed.query)

        # ── Cabinet resolution ──────────────────────────────────────────────
        cabinet = self._resolve_cabinet()
        _NO_CABINET_PATHS = {"/select-cabinet", "/api/cabinets", "/_admin", "/favicon.ico"}
        if cabinet is None and parsed.path not in _NO_CABINET_PATHS:
            cabinets = get_platform().list_cabinets()
            if len(cabinets) == 1:
                # Auto-select when exactly one cabinet exists
                tok = get_platform().create_session(cabinets[0]["cabinet_id"])
                cabinet = get_platform().get_cabinet(cabinets[0]["cabinet_id"])
                _set_request_cabinet(cabinet)
                self.send_response(302)
                self._set_cabinet_cookie(tok)
                self.send_header("Location", self.path)
                self.send_header("Content-Length", "0")
                self.end_headers()
                return
            self._send_redirect("/select-cabinet")
            return

        if parsed.path == "/select-cabinet":
            self._handle_select_cabinet()
            return

        if parsed.path == "/_admin":
            self._handle_admin_get(params)
            return

        if parsed.path.startswith("/_admin/edit/"):
            cab_id = parsed.path[len("/_admin/edit/"):]
            self._handle_admin_edit(cab_id)
            return

        if parsed.path == "/api/cabinets":
            self._send_json({"cabinets": get_platform().list_cabinets()})
            return

        if parsed.path in DISABLED_PAGE_PATHS:
            self._send_redirect("/")
            return

        if parsed.path in DISABLED_API_PATHS:
            self._send_json({"error": "endpoint disabled"}, status=410)
            return

        if parsed.path == "/":
            today = date.today()
            week_ago = today - timedelta(days=6)
            body = _inject_brand(HTML.format(
                date_from=week_ago.isoformat(),
                date_to=today.isoformat(),
                report_nav=_report_nav_html_for_cabinet(cabinet),
                errors_panel=_home_errors_html(cabinet),
                wb_sync_panel=_sync_panel_html(cabinet, "wb"),
                ozon_sync_panel=_sync_panel_html(cabinet, "ozon"),
            ), cabinet=cabinet).encode()
            self._send_html(body)

        elif parsed.path == "/analytics":
            self._send_redirect("/")

        elif parsed.path == "/analytics/period":
            today = date.today()
            week_ago = today - timedelta(days=6)
            body = _inject_brand(ANALYTICS_PERIOD_HTML.format(
                date_from=week_ago.isoformat(),
                date_to=today.isoformat(),
                report_nav=_report_nav_html_for_cabinet(cabinet),
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/analytics/day":
            today = date.today()
            week_ago = today - timedelta(days=6)
            body = _inject_brand(ANALYTICS_DAY_HTML.format(
                date_from=week_ago.isoformat(),
                date_to=today.isoformat(),
                report_nav=_report_nav_html_for_cabinet(cabinet),
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/analytics/article-day":
            today = date.today()
            week_ago = today - timedelta(days=6)
            body = _inject_brand(ANALYTICS_ARTICLE_DAY_HTML.format(
                date_from=week_ago.isoformat(),
                date_to=today.isoformat(),
                report_nav=_report_nav_html_for_cabinet(cabinet),
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/analytics/buyout-order-day":
            today = date.today()
            month_ago = today - timedelta(days=29)
            body = _inject_brand(BUYOUT_ORDER_DAY_HTML.format(
                date_from=month_ago.isoformat(),
                date_to=today.isoformat(),
                report_nav=_report_nav_html_for_cabinet(cabinet),
                page_title="Выкупы по датам заказов",
                granularity="day",
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/analytics/buyout-order-week":
            today = date.today()
            month_ago = today - timedelta(days=29)
            body = _inject_brand(BUYOUT_ORDER_DAY_HTML.format(
                date_from=month_ago.isoformat(),
                date_to=today.isoformat(),
                report_nav=_report_nav_html_for_cabinet(cabinet),
                page_title="Выкупы по неделям",
                granularity="week",
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/analytics/planning":
            body = _inject_brand(PLANNING_HTML.format(
                forecast_to="2026-06-30",
                report_nav=_report_nav_html_for_cabinet(cabinet),
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/analytics/comments":
            body = _inject_brand(COMMENTS_HTML.format(
                report_nav=_report_nav_html_for_cabinet(cabinet),
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/analytics/competitor-prices":
            cab_id = (cabinet or {}).get("cabinet_id", "hld")
            body = COMPETITOR_PRICES_HTML.format(
                report_nav=_report_nav_html_for_cabinet(cabinet),
                cabinet_id=cab_id,
            ).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/ads/bidder":
            if not _has_wb_cabinet(cabinet):
                self._send_redirect("/")
                return
            today = datetime.now(AD_BIDDER_TZ).date()
            month_ago = today - timedelta(days=29)
            body = _inject_brand(AD_BIDDER_HTML.format(
                report_nav=_report_nav_html_for_cabinet(cabinet),
                date_from=month_ago.isoformat(),
                date_to=today.isoformat(),
                today_date=today.isoformat(),
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/tasks":
            body = _inject_brand(TASKS_HTML.format(
                report_nav=_report_nav_html_for_cabinet(cabinet),
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/api/tasks":
            self._send_json({"tasks": _list_tasks()})

        elif parsed.path == "/analytics/funnel-upload":
            if not _has_wb_cabinet(cabinet):
                self._send_redirect("/")
                return
            body = _inject_brand(FUNNEL_UPLOAD_HTML.format(
                report_nav=_report_nav_html_for_cabinet(cabinet),
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/analytics/preliminary-economics":
          today = date.today()
          week_ago = today - timedelta(days=6)
          body = _inject_brand(PRELIMINARY_ECONOMICS_HTML.format(
            date_from=week_ago.isoformat(),
            date_to=today.isoformat(),
            report_nav=_report_nav_html_for_cabinet(cabinet),
          ), cabinet=cabinet).encode("utf-8")
          self._send_html(body)

        elif parsed.path == "/analytics/preliminary-economics-summary":
          today = date.today()
          week_ago = today - timedelta(days=6)
          body = _inject_brand(PRELIMINARY_ECONOMICS_SUMMARY_HTML.format(
            date_from=week_ago.isoformat(),
            date_to=today.isoformat(),
            report_nav=_report_nav_html_for_cabinet(cabinet),
          ), cabinet=cabinet).encode("utf-8")
          self._send_html(body)

        elif parsed.path == "/api/analytics/period":
            article = (params.get("article") or [""])[0]
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            if not date_from or not date_to:
                self._send_json({"error": "date_from and date_to are required", "rows": []}, status=400)
                return
            rows = _fetch_period_analytics(date_from, date_to, article)
            self._send_json({"rows": rows})

        elif parsed.path == "/api/analytics/day":
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            if not date_from or not date_to:
                self._send_json({"error": "date_from and date_to are required", "rows": [], "summary": {}}, status=400)
                return
            data = _fetch_day_analytics(date_from, date_to)
            self._send_json(data)

        elif parsed.path == "/api/analytics/article-day":
            article = (params.get("article") or [""])[0]
            date_from = (params.get("date_from") or [""])[0] or None
            date_to = (params.get("date_to") or [""])[0] or None
            rows = _fetch_article_day_analytics(article, date_from, date_to)
            self._send_json({"rows": rows})

        elif parsed.path == "/api/analytics/buyout-order-day":
            article = (params.get("article") or [""])[0]
            articles = params.get("articles") or []
            subject = (params.get("subject") or [""])[0]
            labels = [l.strip() for l in (params.get("labels") or [""])[0].split(",") if l.strip()]
            revenue_category = (params.get("revenue_category") or [""])[0]
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            granularity = (params.get("granularity") or ["day"])[0]
            if granularity not in {"day", "week"}:
                granularity = "day"
            if not date_from or not date_to:
                self._send_json({"error": "date_from and date_to are required", "dates": [], "rows": []}, status=400)
                return
            dates, effective_from, effective_to = _date_range_limited(
              date_from,
              date_to,
              max_days=93 if granularity == "week" else 45,
            )
            with _db_connect() as conn:
              if not _table_exists(conn, "buyout_order_day"):
                self._send_json({
                  "dates": dates,
                  "rows": [],
                  "effective_from": effective_from,
                  "effective_to": effective_to,
                  "max_days": 93 if granularity == "week" else 45,
                  "granularity": granularity,
                })
                return
            try:
              payload = _fetch_buyout_order_day_pivot(
                date_from,
                date_to,
                article,
                articles=articles,
                subject=subject,
                labels=labels,
                revenue_category=revenue_category,
                granularity=granularity,
              )
            except ValueError:
              self._send_json({"error": "invalid date format", "dates": [], "rows": []}, status=400)
              return
            except Exception as exc:
              self.log_error("buyout-order-day failed: %s", exc)
              self._send_json({"error": "internal error", "dates": [], "rows": []}, status=500)
              return
            self._send_json(payload)

        elif parsed.path == "/api/analytics/planning":
            forecast_to = (params.get("forecast_to") or ["2026-06-30"])[0]
            subject = (params.get("subject") or [""])[0]
            labels = [l.strip() for l in (params.get("labels") or [""])[0].split(",") if l.strip()]
            revenue_category = (params.get("revenue_category") or [""])[0]
            try:
                return_delay_days = int(float((params.get("return_delay_days") or ["8"])[0]))
            except ValueError:
                return_delay_days = 8
            try:
                transit_days = int(float((params.get("transit_days") or ["3"])[0]))
            except ValueError:
                transit_days = 3
            try:
                price_change_pct = float((params.get("price_change_pct") or ["0"])[0])
            except ValueError:
                price_change_pct = 0.0
            try:
                payload = _fetch_planning(
                    forecast_to,
                    return_delay_days=return_delay_days,
                    transit_days=transit_days,
                    subject=subject,
                    labels=labels,
                    revenue_category=revenue_category,
                    price_change_pct=price_change_pct,
                )
            except ValueError:
                self._send_json({"error": "invalid planning parameters", "summary": {}, "rows": []}, status=400)
                return
            self._send_json(payload)

        elif parsed.path == "/api/competitor-prices/summary":
            cab_id = (params.get("cabinet_id") or ["hld"])[0]
            self._send_json(_fetch_competitor_prices_summary(cab_id))

        elif parsed.path == "/api/competitor-prices/detail":
            cab_id = (params.get("cabinet_id") or ["hld"])[0]
            try:
                nm_id = int((params.get("nm_id") or ["0"])[0])
            except ValueError:
                nm_id = 0
            self._send_json(_fetch_competitor_prices_detail(cab_id, nm_id))

        elif parsed.path == "/api/competitor-prices/live-prices":
            nm_raw = (params.get("nm") or [""])[0]
            nm_ids = [int(x) for x in nm_raw.replace(",", ";").split(";") if x.strip().isdigit()]
            self._send_json(_fetch_live_prices(nm_ids))

        elif parsed.path == "/api/competitor-prices/costs":
            cab_id = (params.get("cabinet_id") or ["hld"])[0]
            self._send_json(_fetch_competitor_costs(cab_id))

        elif parsed.path == "/api/analytics/comments-data":
            self._send_json(_fetch_comments_data())

        elif parsed.path == "/api/analytics/comment-log":
            nm_id = (params.get("nmId") or params.get("nm_id") or [""])[0]
            try:
                limit = min(2000, int((params.get("limit") or ["300"])[0]))
            except ValueError:
                limit = 300
            self._send_json(_fetch_comment_log(nm_id=nm_id, limit=limit))

        elif parsed.path == "/api/ads/articles":
            if not _has_wb_cabinet(cabinet):
                self._send_json({"error": "WB cabinet is required", "articles": []}, status=403)
                return
            query = (params.get("q") or [""])[0]
            self._send_json(_fetch_ad_articles(query))

        elif parsed.path == "/api/ads/campaigns":
            if not _has_wb_cabinet(cabinet):
                self._send_json({"error": "WB cabinet is required", "campaigns": []}, status=403)
                return
            nmid = (params.get("nmid") or [""])[0]
            try:
                days = int((params.get("days") or ["7"])[0])
            except ValueError:
                days = 7
            refresh_details = (params.get("refresh") or [""])[0] in {"1", "true", "yes"}
            date_from = (params.get("date_from") or [""])[0] or None
            date_to = (params.get("date_to") or [""])[0] or None
            show_all = (params.get("show_all") or [""])[0] in {"1", "true", "yes"}
            try:
                self._send_json(_fetch_ad_campaigns(nmid, days, refresh_details=refresh_details, date_from=date_from, date_to=date_to, show_all=show_all))
            except Exception as exc:
                self._send_json({"error": str(exc), "campaigns": []}, status=500)

        elif parsed.path == "/api/ads/logs":
            if not _has_wb_cabinet(cabinet):
                self._send_json({"error": "WB cabinet is required", "logs": []}, status=403)
                return
            try:
                limit = int((params.get("limit") or ["200"])[0])
            except ValueError:
                limit = 200
            self._send_json(_fetch_ad_logs(limit))

        elif parsed.path == "/api/ads/rules":
            if not _has_wb_cabinet(cabinet):
                self._send_json({"error": "WB cabinet is required", "rules": []}, status=403)
                return
            self._send_json(_fetch_ad_rules())

        elif parsed.path == "/api/ads/settings":
            if not _has_wb_cabinet(cabinet):
                self._send_json({"error": "WB cabinet is required"}, status=403)
                return
            self._send_json(_fetch_ad_global_settings())

        elif parsed.path == "/api/analytics/planning/export":
            try:
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length))
                rows = body.get("rows", [])
            except Exception:
                self._send_json({"error": "invalid body"}, status=400)
                return
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                self._send_json({"error": "openpyxl not installed"}, status=500)
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Планирование"
            headers = [
                "Артикул", "Предмет", "Ярлыки", "Категория",
                "Остаток", "WH остаток",
                "Баз. цена", "СПП%", "Цена клиента",
                "Эл. цены", "Эл. рекл.", "Баз. ДРР%", "CTR%", "CR1%", "CR2%",
                "Пл. цена", "Пл. скорость",
                "Заказы", "Выкупы", "Выручка",
                "Дата стокаута", "Упущ. заказы", "Упущ. выручка",
            ]
            ws.append(headers)
            hdr_fill = PatternFill("solid", fgColor="1E293B")
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_align = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = hdr_align
            fields_order = [
                "article", "subject", "strategy", "revenue_cat",
                "stock", "wh_stock",
                "base_price", "spp", "client_price",
                "elasticity", "ads_elasticity", "base_drr", "base_ctr", "base_cr1", "base_cr2",
                "planned_price", "planned_speed",
                "forecast_orders", "forecast_buyouts", "forecast_revenue",
                "stockout_date", "lost_orders", "lost_revenue",
            ]
            for r in rows:
                ws.append([r.get(f, "") for f in fields_order])
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
            buf = io.BytesIO()
            wb.save(buf)
            data = buf.getvalue()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="planning.xlsx"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return

        elif parsed.path == "/api/analytics/buyout-filter-options":
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            if not date_from or not date_to:
                self._send_json({"error": "date_from and date_to are required", "subjects": [], "labels": [], "revenue_categories": []}, status=400)
                return
            with _db_connect() as conn:
                if not _table_exists(conn, "buyout_order_day"):
                    self._send_json({"subjects": [], "labels": [], "revenue_categories": []})
                    return
            try:
              options = _fetch_buyout_filter_options(date_from, date_to)
            except ValueError:
              self._send_json({"error": "invalid date format", "subjects": [], "labels": [], "revenue_categories": []}, status=400)
              return
            except Exception as exc:
              self.log_error("buyout-filter-options failed: %s", exc)
              self._send_json({"error": "internal error", "subjects": [], "labels": [], "revenue_categories": []}, status=500)
              return
            self._send_json(options)

        elif parsed.path == "/api/analytics/buyout-subjects":
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            if not date_from or not date_to:
                self._send_json({"error": "date_from and date_to are required", "subjects": []}, status=400)
                return
            try:
                subjects = _fetch_buyout_subjects(date_from, date_to)
            except ValueError:
                self._send_json({"error": "invalid date format", "subjects": []}, status=400)
                return
            self._send_json({"subjects": subjects})

        elif parsed.path == "/api/analytics/buyout-articles":
            subject = (params.get("subject") or [""])[0]
            labels = [l.strip() for l in (params.get("labels") or [""])[0].split(",") if l.strip()]
            revenue_category = (params.get("revenue_category") or [""])[0]
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            if not date_from or not date_to:
                self._send_json({"error": "date_from and date_to are required", "articles": []}, status=400)
                return
            with _db_connect() as conn:
              if not _table_exists(conn, "buyout_order_day"):
                self._send_json({"articles": []})
                return
            try:
              articles = _fetch_buyout_articles(
                date_from,
                date_to,
                subject=subject,
                labels=labels,
                revenue_category=revenue_category,
              )
            except ValueError:
              self._send_json({"error": "invalid date format", "articles": []}, status=400)
              return
            except Exception as exc:
              self.log_error("buyout-articles failed: %s", exc)
              self._send_json({"error": "internal error", "articles": []}, status=500)
              return
            self._send_json({"articles": articles})

        elif parsed.path == "/api/analytics/preliminary-economics":
          article = (params.get("article") or [""])[0]
          date_from = (params.get("date_from") or [""])[0]
          date_to = (params.get("date_to") or [""])[0]
          buyout_raw = (params.get("buyout_percent") or ["30"])[0]
          aggregate = (params.get("aggregate") or ["0"])[0] == "1"
          if not date_from or not date_to:
            self._send_json(
              {
                "error": "date_from and date_to are required",
                "rows": [],
                "additional_rate": "0.00%",
                "expense_components": [],
              },
              status=400,
            )
            return
          try:
            buyout_percent = float(buyout_raw)
          except ValueError:
            buyout_percent = 30.0
          payload = _fetch_preliminary_economics(
            date_from,
            date_to,
            article,
            buyout_percent=buyout_percent,
            aggregate_by_period=aggregate,
          )
          self._send_json(payload)

        elif parsed.path == "/db":
            selected_table = (params.get("table") or [""])[0] or None
            try:
                page = int((params.get("page") or ["1"])[0])
            except ValueError:
                page = 1
            body = _render_db_page(selected_table=selected_table, page=max(page, 1), page_size=200, cabinet=cabinet)
            self._send_html(body)

        elif parsed.path == "/stream":
            params = parse_qs(parsed.query)
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            skip_ads = (params.get("skip_ads") or ["0"])[0] == "1"
            skip_funnel = (params.get("skip_funnel") or ["0"])[0] == "1"
            only = (params.get("mode") or params.get("only") or ["all"])[0]
            if only not in {"all", "sales", "orders", "stocks", "ads", "funnel"}:
                only = "all"

            if not date_from or not date_to:
                self.send_response(400)
                self.end_headers()
                return
            if not _has_wb_cabinet(cabinet):
                _send_sse_once(self, "error", "Нет кабинета WB")
                return

            self.send_response(200)
            self.send_header("Content-Type", "text/event-stream; charset=utf-8")
            self.send_header("Cache-Control", "no-cache")
            self.send_header("X-Accel-Buffering", "no")
            self.end_headers()

            log_q: queue.Queue = queue.Queue()
            thread = threading.Thread(
                target=run_sync,
                args=(date_from, date_to, skip_ads, skip_funnel, log_q, only, cabinet),
                daemon=True,
            )
            thread.start()

            finished = False
            start_ts = time.time()
            last_heartbeat = 0.0
            while not finished:
                try:
                    kind, text = log_q.get(timeout=1)
                except queue.Empty:
                    if thread.is_alive():
                        now = time.time()
                        if now - last_heartbeat >= 5:
                            elapsed = int(now - start_ts)
                            minutes, seconds = divmod(elapsed, 60)
                            kind = "heartbeat"
                            text = f"Выполняется... {minutes:02d}:{seconds:02d}"
                            last_heartbeat = now
                        else:
                            continue
                    else:
                        kind = "error"
                        text = "❌ Процесс завершился без финального статуса. Проверь лог ниже."
                payload = json.dumps({"type": kind, "text": text})
                try:
                    self.wfile.write(f"data: {payload}\n\n".encode())
                    self.wfile.flush()
                except BrokenPipeError:
                    break
                if kind in ("done", "error"):
                    finished = True

            thread.join(timeout=5)

        elif parsed.path == "/ozon/stream":
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            mode = (params.get("mode") or [""])[0]
            ads_only = mode == "ads" or (params.get("ads_only") or ["0"])[0] == "1"
            skip_ads = (params.get("skip_ads") or ["1"])[0] == "1"

            if not date_from or not date_to:
                self.send_response(400)
                self.end_headers()
                return
            if not _has_ozon_cabinet(cabinet):
                _send_sse_once(self, "error", "Нет кабинета OZON")
                return

            self.send_response(200)
            self.send_header("Content-Type", "text/event-stream; charset=utf-8")
            self.send_header("Cache-Control", "no-cache")
            self.send_header("X-Accel-Buffering", "no")
            self.end_headers()

            log_q_ozon: queue.Queue = queue.Queue()
            ozon_thread = threading.Thread(
                target=run_ozon_sync,
                args=(date_from, date_to, log_q_ozon, skip_ads, cabinet, ads_only),
                daemon=True,
            )
            ozon_thread.start()

            finished = False
            start_ts = time.time()
            last_heartbeat = 0.0
            while not finished:
                try:
                    kind, text = log_q_ozon.get(timeout=1)
                except queue.Empty:
                    if ozon_thread.is_alive():
                        now = time.time()
                        if now - last_heartbeat >= 5:
                            elapsed = int(now - start_ts)
                            minutes, seconds = divmod(elapsed, 60)
                            kind = "heartbeat"
                            text = f"Выполняется... {minutes:02d}:{seconds:02d}"
                            last_heartbeat = now
                        else:
                            continue
                    else:
                        kind = "error"
                        text = "❌ Процесс завершился без финального статуса."
                payload = json.dumps({"type": kind, "text": text})
                try:
                    self.wfile.write(f"data: {payload}\n\n".encode())
                    self.wfile.flush()
                except BrokenPipeError:
                    break
                if kind in ("done", "error"):
                    finished = True

            ozon_thread.join(timeout=5)

        elif parsed.path == "/ozon/analytics/day":
            # Basic OZON daily summary table
            mp = (cabinet or {}).get("marketplace", "wb")
            if mp not in ("ozon", "both"):
                self._send_redirect("/")
                return
            today = date.today()
            week_ago = today - timedelta(days=6)
            ozon_day_page = f"""<!DOCTYPE html>
<html lang="ru">
<head><meta charset="utf-8"><title>OZON по дням</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:#f0f2f5;}}
.wrap{{width:calc(100%-24px);max-width:3000px;margin:0 auto;padding:8px 0 28px;}}
.top{{display:flex;align-items:center;gap:8px;margin-bottom:8px;}}
.top h1{{font-size:.82rem;font-weight:700;color:#6b7280;margin:0;}}
.report-nav{{display:flex;align-items:center;flex-wrap:nowrap;overflow-x:auto;gap:2px;flex:1;scrollbar-width:none;}}
.report-nav a{{color:#f97316;text-decoration:none;font-size:.72rem;font-weight:700;white-space:nowrap;padding:3px 8px;border-radius:999px;background:rgba(249,115,22,.07);}}
.ctrl{{display:flex;gap:8px;margin-bottom:10px;align-items:center;flex-wrap:wrap;}}
.ctrl input[type=date]{{padding:5px 8px;border:1.5px solid #d1d5db;border-radius:6px;font-size:.85rem;}}
.ctrl button{{padding:5px 14px;background:#f97316;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:.85rem;}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.07);}}
th,td{{padding:7px 10px;text-align:right;font-size:.8rem;border-bottom:1px solid #f3f4f6;}}
th{{background:#f9fafb;font-weight:700;text-align:center;}}
td:first-child,th:first-child{{text-align:left;}}
</style></head>
<body>
<div class="wrap">
<div class="top"><h1>OZON</h1><div class="report-nav">{_report_nav_html_for_cabinet(cabinet)}</div></div>
<div class="ctrl">
  <input type="date" id="df" value="{week_ago.isoformat()}">
  <input type="date" id="dt" value="{today.isoformat()}">
  <button onclick="load()">Загрузить</button>
</div>
<div id="tbl"></div>
</div>
<script>
function load(){{
  var df=document.getElementById('df').value,dt=document.getElementById('dt').value;
  fetch('/api/ozon/analytics/day?date_from='+df+'&date_to='+dt).then(r=>r.json()).then(function(d){{
    var rows=d.rows||[];
    if(!rows.length){{document.getElementById('tbl').innerHTML='<p style="padding:12px;color:#6b7280;">Нет данных</p>';return;}}
    var h='<table><tr><th>День</th><th>Выручка</th><th>Заказов</th><th>Выкупов</th><th>Возвратов</th><th>Реклама</th><th>Комиссия</th><th>Логистика</th><th>К выплате</th><th>СПП%</th></tr>';
    rows.forEach(function(r){{
      h+='<tr><td>'+r.day+'</td><td>'+fmt(r.orders_revenue)+'</td><td>'+r.orders_qty+'</td><td>'+r.delivered_qty+'</td><td>'+r.returns_qty+'</td><td>'+fmt(r.ad_spend)+'</td><td>'+fmt(r.sale_commission)+'</td><td>'+fmt(r.delivery_charge)+'</td><td>'+fmt(r.for_pay)+'</td><td>'+(r.avg_spp!=null?r.avg_spp.toFixed(1):'—')+'</td></tr>';
    }});
    h+='</table>';
    document.getElementById('tbl').innerHTML=h;
  }});
}}
function fmt(v){{return Number(v||0).toLocaleString('ru-RU',{{maximumFractionDigits:0}});}}
load();
</script>
</body></html>"""
            body = _inject_brand(ozon_day_page, cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/api/ozon/analytics/day":
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            if not date_from or not date_to:
                self._send_json({"rows": []})
                return
            cab_id = (cabinet or {}).get("cabinet_id", "")
            try:
                with _db_connect() as conn:
                    rows = conn.execute(
                        "SELECT day, orders_revenue, orders_qty, delivered_qty, returns_qty, "
                        "cancellations_qty, ad_spend, sale_commission, delivery_charge, "
                        "return_delivery, for_pay, avg_spp "
                        "FROM ozon_daily_summary "
                        "WHERE cabinet_prefix=? AND day BETWEEN ? AND ? ORDER BY day DESC",
                        (cab_id, date_from, date_to),
                    ).fetchall()
                self._send_json({"rows": [dict(r) for r in rows]})
            except Exception:
                self._send_json({"rows": []})

        elif parsed.path in {"/ozon/analytics/buyout-order-day", "/ozon/analytics/buyout-order-week"}:
            mp = (cabinet or {}).get("marketplace", "wb")
            if mp not in ("ozon", "both"):
                self._send_redirect("/")
                return
            today = date.today()
            month_ago = today - timedelta(days=29)
            granularity = "week" if "week" in parsed.path else "day"
            page_title = "OZON — выкупы по неделям" if granularity == "week" else "OZON — выкупы по датам"
            body = _inject_brand(OZON_BUYOUT_ORDER_DAY_HTML.format(
                date_from=month_ago.isoformat(),
                date_to=today.isoformat(),
                report_nav=_report_nav_html_for_cabinet(cabinet),
                page_title=page_title,
                granularity=granularity,
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/ozon/planning":
            mp = (cabinet or {}).get("marketplace", "wb")
            if mp not in ("ozon", "both"):
                self._send_redirect("/")
                return
            today = date.today()
            forecast_to = (today + timedelta(days=30)).isoformat()
            body = _inject_brand(OZON_PLANNING_HTML.format(
                forecast_to=forecast_to,
                report_nav=_report_nav_html_for_cabinet(cabinet),
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/ozon/comments":
            mp = (cabinet or {}).get("marketplace", "wb")
            if mp not in ("ozon", "both"):
                self._send_redirect("/")
                return
            body = _inject_brand(OZON_COMMENTS_HTML.format(
                report_nav=_report_nav_html_for_cabinet(cabinet),
            ), cabinet=cabinet).encode("utf-8")
            self._send_html(body)

        elif parsed.path == "/api/ozon/comments-data":
            self._send_json(_fetch_ozon_comments_data())

        elif parsed.path == "/api/ozon/comment-log":
            ozon_sku = (params.get("ozon_sku") or [""])[0]
            try:
                limit = min(500, int((params.get("limit") or ["300"])[0]))
            except ValueError:
                limit = 300
            self._send_json(_fetch_ozon_comment_log(ozon_sku=ozon_sku, limit=limit))

        elif parsed.path == "/api/ozon/analytics/buyout-filter-options":
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            if not date_from or not date_to:
                self._send_json({"subjects": [], "labels": [], "revenue_categories": []})
                return
            with _db_connect() as conn:
                if not _table_exists(conn, "ozon_sku_day_analytics"):
                    self._send_json({"subjects": [], "labels": [], "revenue_categories": []})
                    return
            try:
              payload = _fetch_ozon_buyout_filter_options(date_from, date_to)
            except ValueError:
              self._send_json({"error": "invalid date format", "subjects": [], "labels": [], "revenue_categories": []}, status=400)
              return
            except Exception as exc:
              self.log_error("ozon buyout-filter-options failed: %s", exc)
              self._send_json({"error": "internal error", "subjects": [], "labels": [], "revenue_categories": []}, status=500)
              return
            self._send_json(payload)

        elif parsed.path == "/api/ozon/analytics/buyout-articles":
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            subject = (params.get("subject") or [""])[0]
            labels = [value.strip() for value in ((params.get("labels") or [""])[0]).split(",") if value.strip()]
            revenue_category = (params.get("revenue_category") or [""])[0]
            if not date_from or not date_to:
                self._send_json({"articles": []})
                return
            with _db_connect() as conn:
              if not _table_exists(conn, "ozon_sku_day_analytics"):
                self._send_json({"articles": []})
                return
            try:
              articles_payload = _fetch_ozon_buyout_articles(
                date_from,
                date_to,
                subject=subject,
                labels=labels,
                revenue_category=revenue_category,
              )
            except ValueError:
              self._send_json({"error": "invalid date format", "articles": []}, status=400)
              return
            except Exception as exc:
              self.log_error("ozon buyout-articles failed: %s", exc)
              self._send_json({"error": "internal error", "articles": []}, status=500)
              return
            self._send_json({"articles": articles_payload})

        elif parsed.path == "/api/ozon/analytics/buyout-order-day":
            article = (params.get("article") or [""])[0]
            articles = params.get("articles") or []
            subject = (params.get("subject") or [""])[0]
            labels = [value.strip() for value in ((params.get("labels") or [""])[0]).split(",") if value.strip()]
            revenue_category = (params.get("revenue_category") or [""])[0]
            date_from = (params.get("date_from") or [""])[0]
            date_to = (params.get("date_to") or [""])[0]
            granularity = (params.get("granularity") or ["day"])[0]
            if granularity not in {"day", "week"}:
                granularity = "day"
            if not date_from or not date_to:
                self._send_json({"error": "date_from and date_to are required", "dates": [], "rows": []}, status=400)
                return
            dates, effective_from, effective_to = _date_range_limited(
              date_from,
              date_to,
              max_days=93 if granularity == "week" else 45,
            )
            with _db_connect() as conn:
              if not _table_exists(conn, "ozon_sku_day_analytics"):
                self._send_json({
                  "dates": dates,
                  "rows": [],
                  "effective_from": effective_from,
                  "effective_to": effective_to,
                  "max_days": 93 if granularity == "week" else 45,
                  "granularity": granularity,
                })
                return
            try:
              payload = _fetch_ozon_buyout_order_day_pivot(
                date_from, date_to,
                article_query=article,
                articles=articles,
                subject=subject,
                labels=labels,
                revenue_category=revenue_category,
                granularity=granularity,
              )
            except ValueError:
              self._send_json({"error": "invalid date format", "dates": [], "rows": []}, status=400)
              return
            except Exception as exc:
              self.log_error("ozon buyout-order-day failed: %s", exc)
              self._send_json({"error": "internal error", "dates": [], "rows": []}, status=500)
              return
            self._send_json(payload)

        elif parsed.path == "/api/ozon/planning":
            default_forecast = (date.today() + timedelta(days=30)).isoformat()
            forecast_to = (params.get("forecast_to") or [default_forecast])[0]
            subject = (params.get("subject") or [""])[0]
            revenue_category = (params.get("revenue_category") or [""])[0]
            try:
                return_delay_days = int(float((params.get("return_delay_days") or ["8"])[0]))
            except ValueError:
                return_delay_days = 8
            try:
                transit_days = int(float((params.get("transit_days") or ["3"])[0]))
            except ValueError:
                transit_days = 3
            try:
                payload = _fetch_ozon_planning(
                    forecast_to,
                    return_delay_days=return_delay_days,
                    transit_days=transit_days,
                    subject=subject,
                    revenue_category=revenue_category,
                )
            except ValueError:
                self._send_json({"error": "invalid date format", "rows": []}, status=400)
                return
            self._send_json(payload)

        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        if not self._is_authorized():
            self._send_unauthorized()
            return

        parsed = urlparse(self.path)

        # Cabinet-less endpoints (available before selecting a cabinet)
        if parsed.path == "/api/switch-cabinet":
            try:
                length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(length))
                cabinet_id = str(body.get("cabinet_id", "")).strip()
                pin = str(body.get("pin", "")).strip()
            except Exception:
                self._send_json({"error": "invalid body"}, status=400)
                return
            if not cabinet_id or not pin:
                self._send_json({"error": "cabinet_id and pin required"}, status=400)
                return
            if not get_platform().verify_pin(cabinet_id, pin):
                self._send_json({"error": "Неверный PIN"}, status=401)
                return
            token = get_platform().create_session(cabinet_id)
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self._set_cabinet_cookie(token)
            body_bytes = b'{"ok":true}'
            self.send_header("Content-Length", str(len(body_bytes)))
            self.end_headers()
            self.wfile.write(body_bytes)
            return

        if parsed.path == "/api/logout-cabinet":
            token = self._get_cookie("wb_session")
            if token:
                get_platform().delete_session(token)
            self.send_response(200)
            self._clear_cabinet_cookie()
            self.send_header("Content-Type", "application/json; charset=utf-8")
            body_bytes = b'{"ok":true}'
            self.send_header("Content-Length", str(len(body_bytes)))
            self.end_headers()
            self.wfile.write(body_bytes)
            return

        if parsed.path == "/_admin/create":
            self._handle_admin_create()
            return

        if parsed.path.startswith("/_admin/edit/"):
            cab_id = parsed.path[len("/_admin/edit/"):]
            self._handle_admin_edit(cab_id)
            return

        if parsed.path.startswith("/_admin/delete/"):
            cab_id = parsed.path[len("/_admin/delete/"):]
            self._handle_admin_delete(cab_id)
            return

        # All remaining POST routes require a cabinet
        cabinet = self._resolve_cabinet()
        if cabinet is None:
            self._send_json({"error": "no cabinet selected"}, status=403)
            return

        if parsed.path in DISABLED_API_PATHS:
            self._send_json({"error": "endpoint disabled"}, status=410)
            return

        if parsed.path == "/api/ads/settings":
            if not _has_wb_cabinet(cabinet):
                self._send_json({"error": "WB cabinet is required"}, status=403)
                return
            try:
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length)) if content_length else {}
                result = _save_ad_global_schedule(body)
            except Exception as exc:
                self._send_json({"error": str(exc)}, status=400)
                return
            self._send_json(result)
            return

        if parsed.path == "/api/ads/campaign-settings":
            if not _has_wb_cabinet(cabinet):
                self._send_json({"error": "WB cabinet is required"}, status=403)
                return
            try:
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length)) if content_length else {}
                result = _save_ad_campaign_setting(body)
            except Exception as exc:
                self._send_json({"error": str(exc)}, status=400)
                return
            self._send_json(result)
            return

        if parsed.path == "/api/ads/rules":
            if not _has_wb_cabinet(cabinet):
                self._send_json({"error": "WB cabinet is required"}, status=403)
                return
            try:
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length))
                result = _save_ad_rule(body)
            except Exception as exc:
                self._send_json({"error": str(exc)}, status=400)
                return
            self._send_json(result)
            return

        if parsed.path.startswith("/api/ads/rules/") and parsed.path.endswith("/toggle"):
            if not _has_wb_cabinet(cabinet):
                self._send_json({"error": "WB cabinet is required"}, status=403)
                return
            try:
                rule_id = int(parsed.path[len("/api/ads/rules/"):-len("/toggle")])
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length)) if content_length else {}
                result = _toggle_ad_rule(rule_id, bool(body.get("enabled")))
            except Exception as exc:
                self._send_json({"error": str(exc)}, status=400)
                return
            self._send_json(result)
            return

        if parsed.path == "/api/ads/executor/run-once":
            if not _has_wb_cabinet(cabinet):
                self._send_json({"error": "WB cabinet is required"}, status=403)
                return
            try:
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length)) if content_length else {}
                result = _run_ad_bidding_once_for_cabinet(cabinet, force_spend_sync=bool(body.get("force_spend_sync")))
            except Exception as exc:
                self._send_json({"error": str(exc)}, status=500)
                return
            self._send_json(result)
            return

        if parsed.path == "/api/tasks":
            try:
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length))
                title = str(body.get("title") or "").strip()
                why = str(body.get("why") or "").strip()
                result_text = str(body.get("result") or "").strip()
                due_date = str(body.get("due_date") or "").strip()
            except Exception as exc:
                self._send_json({"error": str(exc)}, status=400)
                return
            if not title or not why or not result_text or not due_date:
                self._send_json({"error": "all fields required"}, status=400)
                return
            self._send_json(_create_task(title, why, result_text, due_date))
            return

        elif parsed.path == "/api/tasks/update":
            try:
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length))
                task_id = int(body.get("id") or 0)
            except Exception as exc:
                self._send_json({"error": str(exc)}, status=400)
                return
            if not task_id:
                self._send_json({"error": "id required"}, status=400)
                return
            fields = {k: str(v).strip() for k, v in body.items() if k in ("title","why","result","due_date","status")}
            self._send_json(_update_task(task_id, **fields))
            return

        elif parsed.path == "/api/tasks/delete":
            try:
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length))
                task_id = int(body.get("id") or 0)
            except Exception as exc:
                self._send_json({"error": str(exc)}, status=400)
                return
            if not task_id:
                self._send_json({"error": "id required"}, status=400)
                return
            self._send_json(_delete_task(task_id))
            return

        if parsed.path == "/api/analytics/comments":
            try:
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length))
                nm_ids = [int(x) for x in (body.get("nm_ids") or [])]
                comment_type = str(body.get("comment_type") or "").strip()
                comment_text = str(body.get("comment_text") or "").strip()
            except Exception as exc:
                self._send_json({"error": str(exc)}, status=400)
                return
            result = _save_manager_comments(nm_ids, comment_type, comment_text)
            if "error" in result:
                self._send_json(result, status=400)
                return
            self._send_json(result)
            return

        elif parsed.path == "/api/ozon/comments":
            try:
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length))
                skus = [str(x) for x in (body.get("skus") or [])]
                comment_type = str(body.get("comment_type") or "").strip()
                comment_text = str(body.get("comment_text") or "").strip()
            except Exception as exc:
                self._send_json({"error": str(exc)}, status=400)
                return
            result = _save_ozon_manager_comments(skus, comment_type, comment_text)
            if "error" in result:
                self._send_json(result, status=400)
                return
            self._send_json(result)
            return

        elif parsed.path == "/api/analytics/planning/export":
            try:
                content_length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(content_length))
                rows = body.get("rows", [])
            except Exception:
                self._send_json({"error": "invalid body"}, status=400)
                return
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                self._send_json({"error": "openpyxl not installed"}, status=500)
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Планирование"
            # (header label, field key)
            col_defs = [
                ("Артикул",        "article"),
                ("nmid",           "nmid"),
                ("Предмет",        "subject"),
                ("Ярлыки",         "strategy"),
                ("Категория",      "revenue_cat"),
                ("Остаток WB, шт", "stock"),
                ("Склад, шт",      "wh_stock"),
                ("Сумма ост., ₽",  "stock_val"),
                ("Баз. цена, ₽",   "base_price"),
                ("Баз. СПП, %",    "base_spp"),
                ("Баз. кл. цена",  "base_client"),
                ("Баз. скорость",  "base_speed"),
                ("Тек. цена, ₽",   "cur_price"),
                ("Тек. СПП, %",    "cur_spp"),
                ("Тек. кл. цена",  "cur_client"),
                ("Эл. цены",       "elasticity"),
                ("Тек. скорость",  "cur_speed"),
                ("Цена, %",        "price_pct"),
                ("ДРР, %",         "plan_drr"),
                ("Эл. рекл.",      "ads_elasticity"),
                ("CTR, %",         "base_ctr"),
                ("CR1, %",         "base_cr1"),
                ("CR2, %",         "base_cr2"),
                ("Пл. цена, ₽",    "planned_price"),
                ("Пл. кл., ₽",     "planned_client"),
                ("Пл. скорость",   "planned_speed"),
                ("% выкупа",       "buyout_percent"),
                ("Заказы, шт",     "forecast_orders"),
                ("Выкупы, шт",     "forecast_buyouts"),
                ("Выручка, ₽",     "forecast_revenue"),
                ("Дата стопа",     "stockout_date"),
                ("Упущ., шт",      "lost_orders"),
                ("Упущ., ₽",       "lost_revenue"),
                ("Ком. %",         "comm_pct"),
                ("Ком. ₽",         "comm_rub"),
                ("Эквайр. ₽",      "acq_rub"),
                ("Реклама, ₽",     "drr_rub"),
                ("Себест., ₽",     "cogs_total"),
                ("Маржа, %",       "margin_pct"),
                ("Маржа, ₽",       "margin_rub"),
                ("Адм., ₽",        "admin_rub"),
                ("Налоги, ₽",      "tax_rub"),
                ("Прибыль, %",     "profit_pct"),
                ("Прибыль, ₽",     "profit_rub"),
            ]
            ws.append([h for h, _ in col_defs])
            hdr_fill = PatternFill("solid", fgColor="1E293B")
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_align = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = hdr_align
            nmid_col_idx = next((i for i, (_, f) in enumerate(col_defs) if f == "nmid"), None)
            link_font = Font(color="0563C1", underline="single")
            for r in rows:
                ws.append([r.get(f, "") for _, f in col_defs])
                if nmid_col_idx is not None:
                    nmid_val = r.get("nmid", "")
                    if nmid_val:
                        cell = ws.cell(row=ws.max_row, column=nmid_col_idx + 1)
                        cell.hyperlink = f"https://www.wildberries.ru/catalog/{nmid_val}/detail.aspx"
                        cell.font = link_font
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
            buf = io.BytesIO()
            wb.save(buf)
            data = buf.getvalue()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="planning.xlsx"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if parsed.path != "/api/analytics/funnel-upload":
            self._send_json({"error": "not found"}, status=404)
            return

        if not _has_wb_cabinet(cabinet):
            self._send_json({"error": "WB cabinet is required for funnel upload"}, status=403)
            return

        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError:
            length = 0
        if length <= 0:
            self._send_json({"error": "empty body"}, status=400)
            return

        tmp_path = ""
        try:
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            filename = str(payload.get("filename") or "upload.xlsx")
            content = str(payload.get("content") or "")
            force = bool(payload.get("force"))
            raw = base64.b64decode(content, validate=True)
            if not raw:
                raise ValueError("empty file")
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(raw)
                tmp_path = tmp.name
            result = _import_funnel_excel(tmp_path, filename, force=force)
            result["buyout_order_day_rows"] = _rebuild_buyout_order_day_from_store()
        except FunnelUploadWarning as exc:
            self._send_json({
                "warning": True,
                "message": "В файле по некоторым датам товаров меньше, чем уже есть в базе. Данные не загружены.",
                "details": exc.details,
            }, status=409)
            return
        except (json.JSONDecodeError, UnicodeDecodeError, binascii.Error, ValueError) as exc:
            self._send_json({"error": str(exc)}, status=400)
            return
        except Exception as exc:
            self._send_json({"error": str(exc)}, status=400)
            return
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

        self._send_json(result)

def _ensure_yarlyky_column() -> None:
    import sqlite3 as _sqlite3
    cabs_dir = Path(os.getenv("DATA_DIR", str(Path(__file__).parent / "data" / "cabs")))
    if not cabs_dir.exists():
        return
    for db_path in cabs_dir.glob("*.db"):
        try:
            with _sqlite3.connect(db_path) as conn:
                tables = {r[0] for r in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                ).fetchall()}
                for tbl in tables:
                    if tbl.upper() == "SKU":
                        cols = [r[1] for r in conn.execute(
                            f'PRAGMA table_info("{tbl}")'
                        ).fetchall()]
                        if "Ярлыки" not in cols:
                            conn.execute(
                                f'ALTER TABLE "{tbl}" ADD COLUMN "Ярлыки" TEXT'
                            )
                            conn.commit()
                            print(f"[INFO] {db_path.name}: добавлена колонка Ярлыки в {tbl}")
        except Exception as exc:
            print(f"[WARN] _ensure_yarlyky_column {db_path.name}: {exc}")


def main():
    _ensure_yarlyky_column()
    _start_ad_executor()
    host = os.getenv("WEB_APP_HOST", "127.0.0.1").strip() or "127.0.0.1"
    port = int(os.getenv("WEB_APP_PORT", "8765"))
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"Открой в браузере: http://{host}:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nОстановлен.")

if __name__ == "__main__":
    main()
